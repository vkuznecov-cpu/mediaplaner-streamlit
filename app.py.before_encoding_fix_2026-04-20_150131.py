import streamlit as st
import numpy as np
import pandas as pd
import io
import json
import os
import hashlib
import base64
import zipfile
import re
from dataclasses import dataclass
from io import BytesIO
from openpyxl import load_workbook
import datetime as dt
import plotly.express as px
import streamlit.components.v1 as components

try:
    import xlsxwriter  # noqa: F401
    HAS_XLSXWRITER = True
except Exception:
    HAS_XLSXWRITER = False

try:
    from statsmodels.tsa.holtwinters import ExponentialSmoothing
    HAS_STATSMODELS = True
except Exception:
    ExponentialSmoothing = None
    HAS_STATSMODELS = False


# ---------- Р’РќР•РЁРќР™ Р’Р” (CSS) ----------

st.set_page_config(page_title="РњРµРґРёР°РїР»Р°РЅРµСЂ 12 РјРµСЃСЏС†РµРІ (РєР°РјРїР°РЅРёРё)", layout="wide")

THEME_BORDER = "#1D2A44"
THEME_PLOT_TEXT = "#EAF0FF"
THEME_LEGEND_TEXT = "#D4DDF2"
THEME_CARD_BG = "#111A2E"
VAT_RATE = 0.22
USE_EXCEL_ROUNDDOWN = True

METRIC_PRESETS = {
    "ecom": {
        "label": "E-com",
        "description": "Р‘Р°Р·РѕРІС‹Р№ РїСЂРµСЃРµС‚ РґР»СЏ e-commerce: РїРѕРєР°Р·С‹, CTR, CPC, CR, AOV Рё РїСЂРѕРёР·РІРѕРґРЅС‹Рµ РјРµС‚СЂРёРєРё.",
    },
    "diy": {
        "label": "DIY",
        "description": "РџСЂРµСЃРµС‚ РґР»СЏ DIY: РїСЂРѕРґР°Р¶Рё (РѕС‚ РєР»РёРєРѕРІ), CR РІ РїСЂРѕРґР°Р¶Сѓ, SOV Рё РґРѕР»СЏ РЅРѕРІС‹С… РєР»РёРµРЅС‚РѕРІ.",
    },
    "real_estate": {
        "label": "РќРµРґРІРёР¶РёРјРѕСЃС‚СЊ",
        "description": "РџСЂРµСЃРµС‚ РґР»СЏ РЅРµРґРІРёР¶РёРјРѕСЃС‚Рё: Р»РёРґС‹, С†РµР»РµРІС‹Рµ РѕР±СЂР°С‰РµРЅРёСЏ, CPL/CPQL Р±РµР· РІС‹СЂСѓС‡РєРё Рё AOV.",
    },
}

REAL_ESTATE_FUNNEL_OPTIONS = {
    "simple": "РЈРїСЂРѕС‰С‘РЅРЅР°СЏ РІРѕСЂРѕРЅРєР°",
    "full": "РџРѕР»РЅР°СЏ РІРѕСЂРѕРЅРєР°",
}

st.markdown(
    """
    <style>
    .main {
        background-color: #0B1220;
        color: #EAF0FF;
    }
    .stApp {
        background-color: #0B1220;
        color: #EAF0FF;
    }
    [data-testid="block-container"] {
        padding-top: 1.05rem;
        padding-bottom: 0.9rem;
    }
    section[data-testid="stSidebar"] {
        background-color: #111A2E;
        border-right: 1px solid #1D2A44;
    }
    h1, h2, h3, h4 {
        font-weight: 700;
        letter-spacing: 0.02em;
        color: #EAF0FF;
    }
    p, li, label {
        color: #D4DDF2;
    }
    .stDataFrame {
        border-radius: 12px;
        border: 1px solid #1D2A44;
    }
    .stDownloadButton button {
        background: #0066E0;
        color: #FFFFFF;
        font-weight: 700;
        border-radius: 999px;
        border: 1px solid #2B7EE8;
    }
    .stDownloadButton button:hover {
        background: #0A74F2;
        border-color: #0A74F2;
    }
    .stButton > button, .stFormSubmitButton > button {
        border-radius: 10px;
        border: 1px solid #27406F;
        background: #111A2E;
        color: #EAF0FF;
        font-weight: 600;
    }
    .stButton > button:hover, .stFormSubmitButton > button:hover {
        border-color: #0066E0;
        background: #14203A;
    }
    button[kind="primary"],
    button[kind="primaryFormSubmit"],
    .stButton > button[kind="primary"],
    .stFormSubmitButton > button[kind="primary"],
    .stFormSubmitButton > button[kind="primaryFormSubmit"],
    button[data-testid="baseButton-primary"],
    button[data-testid="stBaseButton-primary"] {
        background: linear-gradient(180deg, #0A74F2 0%, #0066E0 100%) !important;
        border: 1px solid #3D8EF0 !important;
        color: #FFFFFF !important;
        box-shadow: 0 0 0 1px rgba(61, 142, 240, 0.25), 0 6px 16px rgba(0, 102, 224, 0.35);
    }
    button[kind="primary"]:hover,
    button[kind="primaryFormSubmit"]:hover,
    .stButton > button[kind="primary"]:hover,
    .stFormSubmitButton > button[kind="primary"]:hover,
    .stFormSubmitButton > button[kind="primaryFormSubmit"]:hover,
    button[data-testid="baseButton-primary"]:hover,
    button[data-testid="stBaseButton-primary"]:hover {
        background: linear-gradient(180deg, #1C82F6 0%, #0A74F2 100%) !important;
        border-color: #74AEF6 !important;
        color: #FFFFFF !important;
    }
    [data-baseweb="tag"] {
        background-color: #1A2A47 !important;
        border: 1px solid #2C4D82 !important;
        color: #9EC5FF !important;
    }
    button[data-baseweb="tab"] {
        color: #9FB0D1 !important;
    }
    button[data-baseweb="tab"][aria-selected="true"] {
        color: #0066E0 !important;
    }
    [data-baseweb="tab-highlight"] {
        background-color: #0066E0 !important;
    }
    .bottom-tab-nav {
        display: flex;
        gap: 1rem;
        align-items: end;
        border-bottom: 1px solid #1D2A44;
        padding-bottom: 6px;
        margin-top: 8px;
    }
    .bottom-tab-btn {
        background: transparent;
        border: 0;
        border-bottom: 2px solid transparent;
        color: #9FB0D1;
        font-weight: 650;
        font-size: 1rem;
        padding: 0 0 8px 0;
        cursor: pointer;
    }
    .bottom-tab-btn.is-active {
        color: #EAF0FF;
        border-bottom-color: #0066E0;
    }
    .ui-section-title {
        margin: 0.22rem 0 0.42rem 0;
        font-size: 1.35rem;
        font-weight: 700;
        letter-spacing: 0.01em;
        line-height: 1.2;
    }
    .tab-intro {
        margin: 0.2rem 0 0.7rem 0;
        padding: 10px 12px;
        border: 1px solid #1D2A44;
        border-radius: 10px;
        background: rgba(17, 26, 46, 0.55);
        animation: introIn 220ms ease both;
    }
    .tab-intro p {
        margin: 0.12rem 0;
        color: #D4DDF2;
        line-height: 1.45;
    }
    [data-testid="stExpander"] {
        margin-bottom: 0.3rem;
    }
    [data-testid="stExpander"] summary p {
        font-weight: 650;
        letter-spacing: 0.01em;
    }
    /* MOTION START */
    .stButton > button,
    .stFormSubmitButton > button,
    .stDownloadButton > button {
        transition: transform 140ms ease, box-shadow 180ms ease, background-color 180ms ease, border-color 180ms ease;
    }
    .stButton > button:hover,
    .stFormSubmitButton > button:hover,
    .stDownloadButton > button:hover {
        transform: translateY(-1px);
        box-shadow: 0 6px 14px rgba(0, 0, 0, 0.22);
    }
    .stButton > button:active,
    .stFormSubmitButton > button:active,
    .stDownloadButton > button:active {
        transform: translateY(0);
        box-shadow: 0 2px 6px rgba(0, 0, 0, 0.16);
    }
    [data-testid="stExpander"] {
        transition: border-color 180ms ease, box-shadow 180ms ease, transform 160ms ease;
    }
    [data-testid="stExpander"]:hover {
        border-color: #2C4D82;
        box-shadow: 0 6px 16px rgba(0, 0, 0, 0.18);
        transform: translateY(-1px);
    }
    [data-testid="stDataFrame"] {
        transition: border-color 180ms ease, box-shadow 180ms ease;
    }
    [data-testid="stDataFrame"]:hover {
        border-color: #2C4D82;
        box-shadow: 0 8px 20px rgba(0, 0, 0, 0.2);
    }
    [data-baseweb="tab"] {
        transition: color 160ms ease;
    }
    [data-baseweb="tab-highlight"] {
        transition: transform 180ms ease, width 180ms ease, left 180ms ease;
    }
    .ui-section-title {
        animation: fadeInUp 260ms ease both;
    }
    @keyframes fadeInUp {
        from { opacity: 0; transform: translateY(4px); }
        to { opacity: 1; transform: translateY(0); }
    }
    @keyframes introIn {
        from { opacity: 0; transform: translateY(6px); }
        to { opacity: 1; transform: translateY(0); }
    }
    /* MOTION END */
    </style>
    """,
    unsafe_allow_html=True,
)

# ---------- РњРћР”Р•Р›Р¬  Р¤РЈРќРљР¦ Р РђРЎР§РЃРўРђ ----------

@dataclass
class PlanInput:
    impressions: float  # РџРѕРєР°Р·С‹
    ctr: float          # CTR РєР°Рє РґРѕР»СЏ (0.01 = 1%)
    cpc: float          # в‚Ѕ
    cr: float = 0.0     # CR РєР°Рє РґРѕР»СЏ (0.02 = 2%) / CR1 РґР»СЏ РЅРµРґРІРёР¶РёРјРѕСЃС‚Рё
    aov: float = 0.0    # в‚Ѕ
    cr2: float = 0.0    # CR2 РєР°Рє РґРѕР»СЏ (РґР»СЏ РїРѕР»РЅРѕР№ РІРѕСЂРѕРЅРєРё РЅРµРґРІРёР¶РёРјРѕСЃС‚Рё)
    reach: float = 0.0  # РћС…РІР°С‚ (Р°РєС‚СѓР°Р»СЊРЅРѕ РґР»СЏ DIY)
    preset_key: str = "ecom"
    funnel_mode: str = "simple"


def calculate_plan_month(inp: PlanInput) -> dict:
    raw_clicks = inp.impressions * inp.ctr
    clicks = np.floor(raw_clicks) if USE_EXCEL_ROUNDDOWN else raw_clicks
    # Р’ С€Р°Р±Р»РѕРЅРµ Excel СЃС‚РѕРёРјРѕСЃС‚СЊ СЃС‡РёС‚Р°РµС‚СЃСЏ РѕС‚ S*CTR (РґРѕ РѕРєСЂСѓРіР»РµРЅРёСЏ РєР»РёРєРѕРІ).
    cost = raw_clicks * inp.cpc
    leads = np.nan
    target_leads = np.nan
    cr1 = np.nan
    cr2 = np.nan

    if inp.preset_key == "real_estate":
        if inp.funnel_mode == "full":
            leads = np.floor(clicks * inp.cr) if USE_EXCEL_ROUNDDOWN else (clicks * inp.cr)
            target_leads = np.floor(leads * inp.cr2) if USE_EXCEL_ROUNDDOWN else (leads * inp.cr2)
            conv = target_leads
            revenue = 0.0
            cr1 = inp.cr
            cr2 = inp.cr2
            cr_total = (target_leads / clicks) if clicks > 0 else 0.0
        else:
            target_leads = np.floor(clicks * inp.cr) if USE_EXCEL_ROUNDDOWN else (clicks * inp.cr)
            conv = target_leads
            revenue = 0.0
            cr_total = inp.cr
        cpm = cost / (inp.impressions / 1000) if inp.impressions > 0 else 0
        cpa = cost / conv if conv > 0 else 0
        roas = 0.0
        drr = 0.0
    elif inp.preset_key == "diy":
        conv = np.floor(clicks * inp.cr) if USE_EXCEL_ROUNDDOWN else (clicks * inp.cr)
        target_leads = np.floor(conv * inp.cr2) if USE_EXCEL_ROUNDDOWN else (conv * inp.cr2)
        revenue = conv * inp.aov
        cpm = cost / (inp.impressions / 1000) if inp.impressions > 0 else 0
        cpa = cost / conv if conv > 0 else 0
        roas = revenue / cost if cost > 0 else 0
        drr = cost / revenue if revenue > 0 else 0
        cr_total = inp.cr
        cr2 = inp.cr2
    else:
        conv = np.floor(clicks * inp.cr) if USE_EXCEL_ROUNDDOWN else (clicks * inp.cr)
        revenue = conv * inp.aov
        cpm = cost / (inp.impressions / 1000) if inp.impressions > 0 else 0
        cpa = cost / conv if conv > 0 else 0
        roas = revenue / cost if cost > 0 else 0
        drr = cost / revenue if revenue > 0 else 0
        cr_total = inp.cr

    reach = max(0.0, float(inp.reach or 0.0))
    frequency = (inp.impressions / reach) if reach > 0 else 0.0

    return {
        "impressions": inp.impressions,
        "reach": reach,
        "frequency": frequency,
        "ctr": inp.ctr,
        "cpc": inp.cpc,
        "cr": cr_total,
        "cr1": cr1,
        "cr2": cr2,
        "aov": inp.aov,
        "clicks": clicks,
        "conversions": conv,
        "leads": leads,
        "target_leads": target_leads if not pd.isna(target_leads) else conv,
        "cost": cost,
        "revenue": revenue,
        "cpm": cpm,
        "cpa": cpa,
        "roas": roas,
        "drr": drr,
    }


def calc_month_for_all_campaigns(base_campaigns: pd.DataFrame,
                                 coeffs_month: pd.DataFrame,
                                 month_num: int,
                                 month_name: str) -> pd.DataFrame:
    rows = []
    for _, base_row in base_campaigns.iterrows():
        campaign_type = str(base_row["campaign_type"])
        k_row = coeffs_month[coeffs_month["campaign_type"] == campaign_type]
        if k_row.empty:
            k_imp = k_ctr = k_cpc = k_cr = k_aov = 1.0
        else:
            k_row = k_row.iloc[0]
            k_imp = k_row["k_imp"]
            k_ctr = k_row["k_ctr"]
            k_cpc = k_row["k_cpc"]
            k_cr = k_row["k_cr"]
            k_aov = k_row["k_aov"]

        base = PlanInput(
            impressions=base_row["impressions_avg"],
            ctr=base_row["ctr_avg_percent"] / 100.0,
            cpc=base_row["cpc_avg"],
            cr=base_row["cr_avg_percent"] / 100.0,
            aov=base_row["aov_avg"],
        )

        month_inp = PlanInput(
            impressions=base.impressions * k_imp,
            ctr=base.ctr * k_ctr,
            cpc=base.cpc * k_cpc,
            cr=base.cr * k_cr,
            aov=base.aov * k_aov,
        )

        out = calculate_plan_month(month_inp)
        out["month_num"] = month_num
        out["month_name"] = month_name
        out["campaign_type"] = campaign_type
        out["system"] = base_row["system"]
        out["format"] = base_row["format"]

        rows.append(out)

    return pd.DataFrame(rows)


def normalize_coeff_set_type(raw_type: str) -> str:
    """
    Normalize legacy/custom spelling variants of coefficient set types.
    This keeps older session_state values compatible after UI text changes.
    """
    val = str(raw_type or "").strip()
    compact = val.lower().replace(" ", "")
    if compact in {"СЃРїСЂРѕСЃ(РїРѕР·Р°РїСЂРѕСЃР°Рј)", "СЃРїСЂРѕСЃРїРѕР·Р°РїСЂРѕСЃР°Рј", "demand"}:
        return "РЎРїСЂРѕСЃ (РїРѕ Р·Р°РїСЂРѕСЃР°Рј)"
    if compact in {"aov(СЃСЂРµРґРЅРёР№С‡РµРє)", "aovСЃСЂРµРґРЅРёР№С‡РµРє", "aov"}:
        return "AOV (СЃСЂРµРґРЅРёР№ С‡РµРє)"
    if compact in {"РєР°СЃС‚РѕРјРЅС‹Р№РЅР°Р±РѕСЂ", "custom", "customset"}:
        return "РљР°СЃС‚РѕРјРЅС‹Р№ РЅР°Р±РѕСЂ"
    if compact in {"РјРµРґРёР№РЅС‹РµС…РІРѕСЃС‚С‹", "РјРµРґРёР№РЅС‹Р№С…РІРѕСЃС‚", "mediatails", "media_tail", "media tails"}:
        return "РњРµРґРёР№РЅС‹Рµ С…РІРѕСЃС‚С‹"
    return val


def get_coeff_set_id_value(coeff_set: dict) -> str:
    try:
        return str(int(coeff_set.get("id", 0) or 0))
    except Exception:
        return ""


def get_coeff_set_label(coeff_set: dict) -> str:
    set_id = get_coeff_set_id_value(coeff_set)
    set_name = str(coeff_set.get("name", "")).strip()
    if not set_id:
        return set_name
    return f"[{set_id}] {set_name}" if set_name else f"[{set_id}]"


def normalize_coeff_link_value(raw_value, coeff_sets: list[dict], allowed_ids: set[str] | None = None) -> str:
    raw_text = str(raw_value or "").strip()
    if raw_text == "":
        return ""

    coeff_sets_by_id = {
        get_coeff_set_id_value(cs): cs
        for cs in coeff_sets
        if isinstance(cs, dict) and get_coeff_set_id_value(cs)
    }

    match = re.match(r"^\[(\d+)\]", raw_text)
    parsed_id = match.group(1) if match else raw_text if raw_text.isdigit() else ""
    if parsed_id and parsed_id in coeff_sets_by_id:
        if allowed_ids is None or parsed_id in allowed_ids:
            return parsed_id

    for set_id, coeff_set in coeff_sets_by_id.items():
        if allowed_ids is not None and set_id not in allowed_ids:
            continue
        set_name = str(coeff_set.get("name", "")).strip()
        if raw_text == set_name or raw_text == get_coeff_set_label(coeff_set):
            return set_id

    return ""


def ui_section_title(text: str) -> None:
    st.markdown(f"<div class='ui-section-title'>{text}</div>", unsafe_allow_html=True)


def parse_float_loose(value, default: float = 0.0) -> float:
    """Parse numeric values robustly (supports spaces, commas and percent sign)."""
    try:
        if value is None:
            return float(default)
        if isinstance(value, str):
            cleaned = value.strip().replace(" ", "").replace("%", "").replace(",", ".")
            if cleaned == "":
                return float(default)
            return float(cleaned)
        if pd.isna(value):
            return float(default)
        return float(value)
    except Exception:
        return float(default)


def image_file_to_data_uri(image_path: str) -> str | None:
    if not image_path or not os.path.exists(image_path):
        return None
    ext = os.path.splitext(image_path)[1].lower()
    mime = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".webp": "image/webp",
        ".svg": "image/svg+xml",
    }.get(ext)
    if not mime:
        return None
    try:
        raw = open(image_path, "rb").read()
    except Exception:
        return None
    encoded = base64.b64encode(raw).decode("ascii")
    return f"data:{mime};base64,{encoded}"


def _df_to_payload(df: pd.DataFrame | None) -> dict:
    if df is None or not isinstance(df, pd.DataFrame):
        return {"columns": [], "rows": []}
    safe_df = df.replace({np.nan: None})
    return {
        "columns": list(safe_df.columns),
        "rows": safe_df.to_dict(orient="records"),
    }


def _df_from_payload(payload: dict | None) -> pd.DataFrame:
    if not payload or not isinstance(payload, dict):
        return pd.DataFrame()
    cols = payload.get("columns", [])
    rows = payload.get("rows", [])
    if not isinstance(cols, list) or not isinstance(rows, list):
        return pd.DataFrame()
    return pd.DataFrame(rows, columns=cols if cols else None)


def _to_csv_bytes(df: pd.DataFrame) -> bytes:
    if df is None or not isinstance(df, pd.DataFrame):
        return b""
    return df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


def safe_select_columns(df: pd.DataFrame, columns: list[str], fill_value=np.nan) -> pd.DataFrame:
    work = df.copy()
    for col in columns:
        if col not in work.columns:
            work[col] = fill_value
    return work[columns].copy()


def get_real_estate_funnel_mode() -> str:
    mode = str(st.session_state.get("real_estate_funnel_mode", "simple")).strip().lower()
    return mode if mode in REAL_ESTATE_FUNNEL_OPTIONS else "simple"


def get_metric_mode(preset_key: str | None = None, funnel_mode: str | None = None) -> dict:
    preset = str(preset_key or st.session_state.get("metric_preset_key", "ecom")).strip().lower()
    if preset not in METRIC_PRESETS:
        preset = "ecom"
    re_mode = str(funnel_mode or get_real_estate_funnel_mode()).strip().lower()
    if re_mode not in REAL_ESTATE_FUNNEL_OPTIONS:
        re_mode = "simple"
    return {
        "preset_key": preset,
        "is_diy": preset == "diy",
        "is_real_estate": preset == "real_estate",
        "is_real_estate_simple": preset == "real_estate" and re_mode == "simple",
        "is_real_estate_full": preset == "real_estate" and re_mode == "full",
        "real_estate_funnel_mode": re_mode,
        "needs_aov": preset in {"ecom", "diy"},
        "needs_capacity": preset == "diy",
    }


def get_campaign_required_cols(metric_mode: dict) -> list[str]:
    required = ["impressions_avg", "ctr_avg_percent", "cpc_avg", "cr_avg_percent"]
    if metric_mode.get("needs_aov"):
        required.append("aov_avg")
    if metric_mode.get("is_diy"):
        required.append("reach_avg")
        required.append("cr2_avg_percent")
    elif metric_mode.get("is_real_estate_full"):
        required.append("cr2_avg_percent")
    return required


def get_real_estate_display_metric_specs(metric_mode: dict) -> list[tuple[str, str]]:
    common = [
        ("РџРѕРєР°Р·С‹", "impressions"),
        ("РљР»РёРєРё", "clicks"),
        ("CTR", "ctr"),
        ("CPC", "cpc"),
        ("Р‘СЋРґР¶РµС‚", "cost"),
        ("Р‘СЋРґР¶РµС‚ СЃ РќР”РЎ", "cost_with_vat"),
        ("Р‘СЋРґР¶РµС‚ СЃ РќР”РЎ Рё РђРљ", "cost_with_vat_ak"),
        ("CPM", "cpm"),
    ]
    if metric_mode.get("is_real_estate_full"):
        return common + [
            ("CR1 РІ Р›РёРґ", "cr1_pct"),
            ("Р›РёРґС‹", "leads"),
            ("CPL", "cpl"),
            ("CR2 РІ Р¦Рћ", "cr2_pct"),
            ("Р¦Рћ", "target_leads"),
            ("CPQL", "cpql"),
        ]
    return common + [
        ("CR РІ Р¦Рћ", "cr_pct"),
        ("Р¦Рћ", "target_leads"),
        ("CPQL", "cpql"),
    ]


def get_real_estate_table_cols(metric_mode: dict) -> list[str]:
    if metric_mode.get("is_real_estate_full"):
        return [
            "impressions", "clicks", "ctr_pct", "cpc", "cost", "cost_with_vat", "cost_with_vat_ak",
            "cr1_pct", "leads", "cpl", "cr2_pct", "target_leads", "cpql", "cpm",
        ]
    return [
        "impressions", "clicks", "ctr_pct", "cpc", "cost", "cost_with_vat", "cost_with_vat_ak",
        "cr_pct", "target_leads", "cpql", "cpm",
    ]


def compute_real_estate_rates(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "target_leads" not in out.columns and "conversions" in out.columns:
        out["target_leads"] = out["conversions"]
    if "leads" not in out.columns:
        out["leads"] = np.nan
    out["cr1_pct"] = np.where(out["clicks"] > 0, out["leads"] / out["clicks"] * 100.0, 0.0)
    out["cr2_pct"] = np.where(out["leads"] > 0, out["target_leads"] / out["leads"] * 100.0, 0.0)
    return out


def _bootstrap_reference_from_campaigns(campaigns_df: pd.DataFrame | None) -> None:
    """
    Prepare quick-reference sidebar data right after import.
    This keeps "Р‘С‹СЃС‚СЂР°СЏ СЃРІРµСЂРєР°" usable even before full recalculation blocks run.
    """
    if campaigns_df is None or not isinstance(campaigns_df, pd.DataFrame) or campaigns_df.empty:
        return

    metric_mode = get_metric_mode()
    required_cols = {"campaign_type", *get_campaign_required_cols(metric_mode)}
    if not required_cols.issubset(set(campaigns_df.columns)):
        return

    work = campaigns_df.copy()
    work["campaign_type"] = work["campaign_type"].astype(str).str.strip()
    work = work[work["campaign_type"] != ""]
    if work.empty:
        return

    use_vat = bool(st.session_state.get("use_vat_budget_metrics", True))
    base_by_campaign = {}
    total_imp = total_clicks = total_conv = total_cost = total_rev = 0.0

    for _, row in work.iterrows():
        inp = PlanInput(
            impressions=max(0.0, parse_float_loose(row.get("impressions_avg"), 0.0)),
            ctr=max(0.0, parse_float_loose(row.get("ctr_avg_percent"), 0.0) / 100.0),
            cpc=max(0.0, parse_float_loose(row.get("cpc_avg"), 0.0)),
            cr=max(0.0, parse_float_loose(row.get("cr_avg_percent"), 0.0) / 100.0),
            aov=max(0.0, parse_float_loose(row.get("aov_avg"), 0.0)),
            cr2=max(0.0, parse_float_loose(row.get("cr2_avg_percent"), 0.0) / 100.0),
            preset_key=metric_mode["preset_key"],
            funnel_mode=metric_mode["real_estate_funnel_mode"],
        )
        out = calculate_plan_month(inp)
        camp = str(row.get("campaign_type", "")).strip()
        if not camp:
            continue

        imp = float(out.get("impressions", 0.0))
        clicks = float(out.get("clicks", 0.0))
        conv = float(out.get("conversions", 0.0))
        leads = float(out.get("leads", 0.0) or 0.0)
        target_leads = float(out.get("target_leads", conv) or 0.0)
        cost = float(out.get("cost", 0.0))
        rev = float(out.get("revenue", 0.0))
        cost_with_vat = cost * (1.0 + VAT_RATE)
        budget_basis = cost_with_vat if use_vat else cost
        cpo = (budget_basis / conv) if conv > 0 else 0.0
        cpl = (budget_basis / leads) if leads > 0 else 0.0
        cpql = (budget_basis / target_leads) if target_leads > 0 else 0.0
        roas = (rev / budget_basis * 100.0) if budget_basis > 0 else 0.0
        drr = (budget_basis / rev * 100.0) if rev > 0 else 0.0

        base_by_campaign[camp] = {
            "impressions": imp,
            "clicks": clicks,
            "conversions": conv,
            "leads": leads,
            "target_leads": target_leads,
            "cost": cost,
            "cost_with_vat": cost_with_vat,
            "cost_with_vat_ak": cost_with_vat,
            "revenue": rev,
            "ctr": (clicks / imp * 100.0) if imp > 0 else 0.0,
            "cpc": (cost / clicks) if clicks > 0 else 0.0,
            "cr": (conv / clicks * 100.0) if clicks > 0 else 0.0,
            "cpo": cpo,
            "cpl": cpl,
            "cpql": cpql,
            "roas": roas,
            "drr": drr,
        }

        total_imp += imp
        total_clicks += clicks
        total_conv += conv
        total_cost += cost
        total_rev += rev

    if not base_by_campaign:
        return

    total_cost_with_vat = total_cost * (1.0 + VAT_RATE)
    total_budget_basis = total_cost_with_vat if use_vat else total_cost
    st.session_state["mp_ref_base_by_campaign"] = base_by_campaign
    st.session_state["mp_ref_base"] = {
        "impressions": float(total_imp),
        "clicks": float(total_clicks),
        "conversions": float(total_conv),
        "cost": float(total_cost),
        "cost_with_vat": float(total_cost_with_vat),
        "cost_with_vat_ak": float(total_cost_with_vat),
        "revenue": float(total_rev),
        "ctr": float((total_clicks / total_imp * 100.0) if total_imp > 0 else 0.0),
        "cpc": float((total_cost / total_clicks) if total_clicks > 0 else 0.0),
        "cr": float((total_conv / total_clicks * 100.0) if total_clicks > 0 else 0.0),
        "cpo": float((total_budget_basis / total_conv) if total_conv > 0 else 0.0),
        "roas": float((total_rev / total_budget_basis * 100.0) if total_budget_basis > 0 else 0.0),
        "drr": float((total_budget_basis / total_rev * 100.0) if total_rev > 0 else 0.0),
    }


def export_project_state_payload() -> dict:
    coeff_sets_src = st.session_state.get("coeff_sets", [])
    coeff_sets_payload = []
    for cs in coeff_sets_src:
        if not isinstance(cs, dict):
            continue
        coeff_sets_payload.append(
            {
                "id": int(cs.get("id", 0) or 0),
                "name": str(cs.get("name", "")),
                "type": str(cs.get("type", "")),
                "start_month": int(cs.get("start_month", 1) or 1),
                "start_year": int(cs.get("start_year", dt.date.today().year) or dt.date.today().year),
                "period_months": int(cs.get("period_months", 24) or 24),
                "queries": [str(q) for q in cs.get("queries", []) if str(q).strip()],
                "df_data": _df_to_payload(cs.get("df_data")),
                "result": _df_to_payload(cs.get("result")),
            }
        )

    payload = {
        "schema_version": 1,
        "exported_at": dt.datetime.now().isoformat(timespec="seconds"),
        "app": "media_planner",
        "state": {
            "campaigns_df": _df_to_payload(st.session_state.get("campaigns_df")),
            "coeff_sets": coeff_sets_payload,
            "coeff_sets_links_new": _df_to_payload(st.session_state.get("coeff_sets_links_new")),
            "elasticity_df": _df_to_payload(st.session_state.get("elasticity_df")),
            "ak_rules_df": _df_to_payload(st.session_state.get("ak_rules_df")),
            "use_vat_budget_metrics": bool(st.session_state.get("use_vat_budget_metrics", True)),
            "use_ak_budget_metrics": bool(st.session_state.get("use_ak_budget_metrics", False)),
            "ak_mode": str(st.session_state.get("ak_mode", "percent")),
            "ak_fixed_month_wo_vat": float(st.session_state.get("ak_fixed_month_wo_vat", 200000.0) or 0.0),
            "ak_fixed_percent": float(st.session_state.get("ak_fixed_percent", 2.0) or 0.0),
            "planning_months_multiselect": list(st.session_state.get("planning_months_multiselect", [])),
            "metric_preset_key": str(st.session_state.get("metric_preset_key", "ecom")),
            "real_estate_funnel_mode": str(st.session_state.get("real_estate_funnel_mode", "simple")),
        },
    }
    return payload


def import_project_state_payload(payload: dict) -> tuple[bool, str]:
    if not isinstance(payload, dict):
        return False, "РќРµРІРµСЂРЅС‹Р№ С„РѕСЂРјР°С‚ С„Р°Р№Р»Р° РёРјРїРѕСЂС‚Р°."
    if payload.get("schema_version") != 1:
        return False, "РќРµРїРѕРґРґРµСЂР¶РёРІР°РµРјР°СЏ РІРµСЂСЃРёСЏ С„Р°Р№Р»Р° РёРјРїРѕСЂС‚Р°."

    state = payload.get("state")
    if not isinstance(state, dict):
        return False, "Р’ С„Р°Р№Р»Рµ РЅРµС‚ Р±Р»РѕРєР° state."

    st.session_state["campaigns_df"] = _df_from_payload(state.get("campaigns_df"))
    st.session_state["coeff_sets_links_new"] = _df_from_payload(state.get("coeff_sets_links_new"))
    st.session_state["elasticity_df"] = _df_from_payload(state.get("elasticity_df"))
    st.session_state["ak_rules_df"] = _df_from_payload(state.get("ak_rules_df"))

    imported_sets = []
    for idx, cs in enumerate(state.get("coeff_sets", []), start=1):
        if not isinstance(cs, dict):
            continue
        imported_sets.append(
            {
                "id": int(cs.get("id", idx) or idx),
                "name": str(cs.get("name", f"РќР°Р±РѕСЂ {idx}")),
                "type": str(cs.get("type", "РЎРїСЂРѕСЃ (РїРѕ Р·Р°РїСЂРѕСЃР°Рј)")),
                "start_month": int(cs.get("start_month", 1) or 1),
                "start_year": int(cs.get("start_year", dt.date.today().year) or dt.date.today().year),
                "period_months": int(cs.get("period_months", 24) or 24),
                "queries": [str(q) for q in cs.get("queries", []) if str(q).strip()],
                "df_data": _df_from_payload(cs.get("df_data")),
                "result": _df_from_payload(cs.get("result")),
            }
        )
    st.session_state["coeff_sets"] = imported_sets

    st.session_state["use_vat_budget_metrics"] = bool(state.get("use_vat_budget_metrics", True))
    st.session_state["use_ak_budget_metrics"] = bool(state.get("use_ak_budget_metrics", False))
    st.session_state["ak_mode"] = str(state.get("ak_mode", "percent"))
    st.session_state["ak_fixed_month_wo_vat"] = float(state.get("ak_fixed_month_wo_vat", 200000.0) or 0.0)
    st.session_state["ak_fixed_percent"] = float(state.get("ak_fixed_percent", 2.0) or 0.0)
    st.session_state["planning_months_multiselect"] = list(state.get("planning_months_multiselect", []))
    preset_key = str(state.get("metric_preset_key", "ecom"))
    st.session_state["metric_preset_key"] = preset_key if preset_key in METRIC_PRESETS else "ecom"
    st.session_state["real_estate_funnel_mode"] = str(state.get("real_estate_funnel_mode", "simple"))
    _bootstrap_reference_from_campaigns(st.session_state.get("campaigns_df"))

    return True, "Р”Р°РЅРЅС‹Рµ РїСЂРѕРµРєС‚Р° РёРјРїРѕСЂС‚РёСЂРѕРІР°РЅС‹."


if "_pending_project_import_payload" in st.session_state:
    _pending_payload = st.session_state.pop("_pending_project_import_payload")
    _ok, _msg = import_project_state_payload(_pending_payload)
    st.session_state["_pending_project_import_result"] = {"ok": bool(_ok), "msg": str(_msg)}


def queue_project_import_from_upload(uploaded_file, source_key: str) -> None:
    """
    Queue project import from a Streamlit UploadedFile only once per file content.
    Prevents rerun loops when uploader keeps the selected file between reruns.
    """
    sig_key = f"_project_upload_sig_{source_key}"
    if uploaded_file is None:
        st.session_state.pop(sig_key, None)
        return

    raw_bytes = uploaded_file.getvalue()
    payload_sig = hashlib.sha1(raw_bytes).hexdigest()
    if st.session_state.get(sig_key) == payload_sig:
        return

    st.session_state[sig_key] = payload_sig
    try:
        imported_payload = json.loads(raw_bytes.decode("utf-8"))
        st.session_state["_pending_project_import_payload"] = imported_payload
        st.rerun()
    except Exception as e:
        st.error(f"РќРµ СѓРґР°Р»РѕСЃСЊ РїСЂРѕС‡РёС‚Р°С‚СЊ С„Р°Р№Р» РїСЂРѕРµРєС‚Р°: {e}")


def make_ak_rules_signature(ak_rules_df: pd.DataFrame | None) -> tuple:
    if ak_rules_df is None or not isinstance(ak_rules_df, pd.DataFrame) or ak_rules_df.empty:
        return tuple()
    work = ak_rules_df.copy()
    if "min_budget_wo_vat" not in work.columns:
        work["min_budget_wo_vat"] = 0.0
    if "ak_percent" not in work.columns:
        work["ak_percent"] = 0.0
    work["min_budget_wo_vat"] = work["min_budget_wo_vat"].map(lambda v: round(parse_float_loose(v, 0.0), 4))
    work["ak_percent"] = work["ak_percent"].map(lambda v: round(parse_float_loose(v, 0.0), 4))
    work = work[["min_budget_wo_vat", "ak_percent"]].sort_values("min_budget_wo_vat", kind="stable")
    return tuple((float(r["min_budget_wo_vat"]), float(r["ak_percent"])) for _, r in work.iterrows())


def make_vat_ak_settings_signature(
    use_vat_budget_metrics: bool,
    use_ak_budget_metrics: bool,
    ak_mode: str,
    ak_fixed_month_wo_vat: float,
    ak_fixed_percent: float,
    ak_rules_df: pd.DataFrame | None,
) -> tuple:
    return (
        bool(use_vat_budget_metrics),
        bool(use_ak_budget_metrics),
        str(ak_mode),
        round(float(ak_fixed_month_wo_vat), 4) if str(ak_mode) == "fixed" else None,
        round(float(ak_fixed_percent), 4) if str(ak_mode) == "fixed_percent" else None,
        make_ak_rules_signature(ak_rules_df) if str(ak_mode) == "percent" else tuple(),
    )


def normalize_ak_rules_df(ak_rules_df: pd.DataFrame | None) -> pd.DataFrame:
    if ak_rules_df is None or not isinstance(ak_rules_df, pd.DataFrame):
        return pd.DataFrame([{"min_budget_wo_vat": 0.0, "ak_percent": 0.0}])
    work = ak_rules_df.copy()
    for col in ["min_budget_wo_vat", "ak_percent"]:
        if col not in work.columns:
            work[col] = 0.0
    work["min_budget_wo_vat"] = work["min_budget_wo_vat"].map(lambda v: parse_float_loose(v, 0.0))
    work["ak_percent"] = work["ak_percent"].map(lambda v: parse_float_loose(v, 0.0))
    work = work[["min_budget_wo_vat", "ak_percent"]].sort_values("min_budget_wo_vat", kind="stable").reset_index(drop=True)
    if work.empty:
        work = pd.DataFrame([{"min_budget_wo_vat": 0.0, "ak_percent": 0.0}])
    return work


def get_vat_ak_ui_config(metric_mode: dict, is_real_estate_preset: bool) -> dict:
    ak_mode_options = [
        ("percent", "Процент по шкале"),
        ("fixed", "Фиксированная сумма в месяц"),
        ("fixed_percent", "Фиксированный процент"),
    ]
    warning_text = (
        "<b>Учитывать НДС 22% в расчетах</b>: настройка влияет на CPC, CPM, CPL и CPQL."
        if is_real_estate_preset and metric_mode["is_real_estate_full"]
        else "<b>Учитывать НДС 22% в расчетах</b>: настройка влияет на CPC, CPM и CPQL."
        if is_real_estate_preset
        else "<b>Учитывать НДС 22% в расчетах</b>: настройка влияет на CPC, CPM, CPO, ROAS и ДРР."
    )
    mode_label_by_value = {value: label for value, label in ak_mode_options}
    mode_value_by_label = {label: value for value, label in ak_mode_options}
    return {
        "warning_text": warning_text,
        "ak_mode_labels": [label for _, label in ak_mode_options],
        "mode_label_by_value": mode_label_by_value,
        "mode_value_by_label": mode_value_by_label,
    }


def get_vat_ak_apply_state(
    use_vat_budget_metrics: bool,
    use_ak_budget_metrics: bool,
    ak_mode: str,
    ak_fixed_month_wo_vat: float,
    ak_fixed_percent: float,
    ak_rules_df: pd.DataFrame | None,
    last_applied_sig: tuple | None,
) -> dict:
    current_sig = make_vat_ak_settings_signature(
        use_vat_budget_metrics=use_vat_budget_metrics,
        use_ak_budget_metrics=use_ak_budget_metrics,
        ak_mode=ak_mode,
        ak_fixed_month_wo_vat=ak_fixed_month_wo_vat,
        ak_fixed_percent=ak_fixed_percent,
        ak_rules_df=ak_rules_df,
    )
    return {
        "current_sig": current_sig,
        "last_applied_sig": last_applied_sig,
        "dirty": current_sig != last_applied_sig,
    }


def resolve_ak_rate(total_cost_wo_vat: float, ak_rules_df: pd.DataFrame) -> float:
    if ak_rules_df is None or ak_rules_df.empty:
        return 0.0
    rules = ak_rules_df.copy()
    rules["min_budget_wo_vat"] = rules["min_budget_wo_vat"].map(lambda v: parse_float_loose(v, 0.0))
    rules["ak_percent"] = rules["ak_percent"].map(lambda v: parse_float_loose(v, 0.0))
    rules = rules.sort_values("min_budget_wo_vat")
    matched = rules[rules["min_budget_wo_vat"] <= float(total_cost_wo_vat)]
    if matched.empty:
        return 0.0
    return max(0.0, float(matched.iloc[-1]["ak_percent"]) / 100.0)


def apply_budget_basis_metrics(
    df: pd.DataFrame,
    use_vat: bool,
    use_ak: bool = False,
    ak_mode: str = "percent",
    ak_rate_by_month: dict | None = None,
    default_ak_rate: float = 0.0,
    ak_fixed_by_month: dict | None = None,
    default_ak_fixed_wo_vat: float = 0.0,
    default_ak_fixed_rate: float = 0.0,
    vat_rate: float = VAT_RATE,
) -> pd.DataFrame:
    out = df.copy()
    out["cost_with_vat"] = out["cost"].astype(float) * (1.0 + vat_rate)
    out["ak_rate"] = 0.0
    out["ak_fixed_month_wo_vat"] = 0.0
    out["ak_cost_wo_vat"] = 0.0

    if use_ak:
        if ak_mode == "fixed":
            if "month_num" in out.columns:
                if ak_fixed_by_month:
                    out["ak_fixed_month_wo_vat"] = out["month_num"].map(
                        lambda m: float(ak_fixed_by_month.get(int(m), default_ak_fixed_wo_vat))
                    )
                else:
                    out["ak_fixed_month_wo_vat"] = float(default_ak_fixed_wo_vat)
                month_sum = out.groupby("month_num")["cost"].transform("sum")
                share = np.where(month_sum > 0, out["cost"] / month_sum, 0.0)
                out["ak_cost_wo_vat"] = out["ak_fixed_month_wo_vat"] * share
            else:
                total_cost = float(out["cost"].sum())
                fixed_total = float(default_ak_fixed_wo_vat)
                out["ak_fixed_month_wo_vat"] = fixed_total
                share = np.where(total_cost > 0, out["cost"] / total_cost, 0.0)
                out["ak_cost_wo_vat"] = fixed_total * share
            out["ak_rate"] = np.where(out["cost"] > 0, out["ak_cost_wo_vat"] / out["cost"], 0.0)
        elif ak_mode == "fixed_percent":
            out["ak_rate"] = float(default_ak_fixed_rate)
            out["ak_rate"] = pd.to_numeric(out["ak_rate"], errors="coerce").fillna(float(default_ak_fixed_rate))
            out["ak_cost_wo_vat"] = out["cost"] * out["ak_rate"]
        else:
            if "month_num" in out.columns and ak_rate_by_month:
                out["ak_rate"] = out["month_num"].map(
                    lambda m: float(ak_rate_by_month.get(int(m), default_ak_rate))
                )
            else:
                out["ak_rate"] = float(default_ak_rate)
            out["ak_rate"] = pd.to_numeric(out["ak_rate"], errors="coerce").fillna(float(default_ak_rate))
            out["ak_cost_wo_vat"] = out["cost"] * out["ak_rate"]

    out["ak_cost_with_vat"] = out["ak_cost_wo_vat"] * (1.0 + vat_rate)
    out["total_budget_wo_vat_ak"] = out["cost"] + out["ak_cost_wo_vat"]
    out["cost_with_vat_ak"] = (out["cost"] + out["ak_cost_wo_vat"]) * (1.0 + vat_rate)
    out["vat_amount"] = out["cost_with_vat_ak"] - out["total_budget_wo_vat_ak"]
    if use_ak:
        budget_basis = out["cost_with_vat_ak"] if use_vat else (out["cost"] + out["ak_cost_wo_vat"])
    else:
        budget_basis = out["cost_with_vat"] if use_vat else out["cost"]
    out["cpm"] = np.where(
        out["impressions"] > 0,
        budget_basis / (out["impressions"] / 1000.0),
        0.0,
    )
    out["cpa"] = np.where(
        out["conversions"] > 0,
        budget_basis / out["conversions"],
        0.0,
    )
    if "leads" in out.columns:
        out["cpl"] = np.where(out["leads"] > 0, budget_basis / out["leads"], 0.0)
    if "target_leads" in out.columns:
        out["cpql"] = np.where(out["target_leads"] > 0, budget_basis / out["target_leads"], 0.0)
    out["roas"] = np.where(
        budget_basis > 0,
        out["revenue"] / budget_basis,
        0.0,
    )
    out["drr"] = np.where(
        out["revenue"] > 0,
        budget_basis / out["revenue"],
        0.0,
    )
    out["budget_basis"] = budget_basis
    return out


def render_bottom_tab_switcher(current_tab: str, key_suffix: str) -> None:
    st.markdown("---")
    nav_id = f"bottom-tab-nav-{key_suffix}"
    st.markdown(
        f"""
        <div id="{nav_id}" class="bottom-tab-nav">
            <button type="button" class="bottom-tab-btn" data-tab-idx="0">РљРѕСЌС„С„РёС†РёРµРЅС‚С‹</button>
            <button type="button" class="bottom-tab-btn" data-tab-idx="1">РњРµРґРёР°РїР»Р°РЅ</button>
            <button type="button" class="bottom-tab-btn" data-tab-idx="2">Р”РёР°РіСЂР°РјРјС‹</button>
            <button type="button" class="bottom-tab-btn" data-tab-idx="3">Export/Import</button>
        </div>
        """,
        unsafe_allow_html=True,
    )
    components.html(
        f"""
        <script>
        (function() {{
            const tablists = Array.from(window.parent.document.querySelectorAll('[role="tablist"]'));
            if (!tablists.length) return;
            const topmostTablist = tablists
                .map(tl => ({{ tl, top: tl.getBoundingClientRect().top }}))
                .sort((a, b) => a.top - b.top)[0]?.tl;
            const nav = window.parent.document.getElementById("{nav_id}");
            if (!topmostTablist || !nav) return;

            const topTabs = topmostTablist.querySelectorAll('button[role="tab"]');
            const navBtns = nav.querySelectorAll('.bottom-tab-btn');
            if (!topTabs || topTabs.length < 4 || !navBtns.length) return;

            const syncActive = () => {{
                let activeIdx = 0;
                for (let i = 0; i < topTabs.length; i++) {{
                    if (topTabs[i].getAttribute('aria-selected') === 'true') {{
                        activeIdx = i;
                        break;
                    }}
                }}
                navBtns.forEach((btn, i) => {{
                    if (i === activeIdx) btn.classList.add('is-active');
                    else btn.classList.remove('is-active');
                }});
            }};

            navBtns.forEach((btn) => {{
                btn.onclick = () => {{
                    const idx = Number(btn.getAttribute('data-tab-idx') || 0);
                    if (topTabs[idx]) {{
                        topTabs[idx].click();
                        setTimeout(syncActive, 30);
                    }}
                }};
            }});

            syncActive();
            setTimeout(syncActive, 40);
            const obs = new MutationObserver(() => syncActive());
            obs.observe(topmostTablist, {{
                subtree: true,
                attributes: true,
                attributeFilter: ['aria-selected', 'class']
            }});
        }})();
        </script>
        """,
        height=0,
        width=0,
    )


DISPLAY_COL_RENAME = {
    "segment": "РЎРµРіРјРµРЅС‚",
    "campaign_type": "РќР°Р·РІР°РЅРёРµ РєР°РјРїР°РЅРёРё",
    "system": "Р РµРєР»Р°РјРЅР°СЏ СЃРёСЃС‚РµРјР°",
    "format": "Р¤РѕСЂРјР°С‚/С‚Р°СЂРіРµС‚РёРЅРіРё",
    "geo": "Р“Р•Рћ",
    "month_name": "РњРµСЃСЏС†",
    "impressions": "РџРѕРєР°Р·С‹",
    "reach": "РћС…РІР°С‚",
    "frequency": "Р§Р°СЃС‚РѕС‚Р°",
    "ctr_pct": "CTR",
    "ctr": "CTR",
    "cpc": "CPC",
    "cr_pct": "CR",
    "cr": "CR",
    "cr1_pct": "CR1 РІ Р›РёРґ",
    "cr2_pct": "CR2 РІ Р¦Рћ",
    "aov": "AOV",
    "clicks": "РљР»РёРєРё",
    "conversions": "РљРѕРЅРІРµСЂСЃРёРё",
    "leads": "Р›РёРґС‹",
    "target_leads": "Р¦Рћ",
    "cost": "Р‘СЋРґР¶РµС‚",
    "cost_with_vat": "Р‘СЋРґР¶РµС‚ СЃ РќР”РЎ",
    "cost_with_vat_ak": "Р‘СЋРґР¶РµС‚ СЃ РќР”РЎ Рё РђРљ",
    "ak_cost_wo_vat": "РђРљ Р±РµР· РќР”РЎ",
    "total_budget_wo_vat_ak": "РўРѕС‚Р°Р» Р±СЋРґР¶РµС‚ Р±РµР· РќР”РЎ СЃ СѓС‡РµС‚РѕРј РђРљ",
    "vat_amount": "РќР”РЎ, в‚Ѕ",
    "ak_rate_pct": "РђРљ, %",
    "revenue": "Р”РѕС…РѕРґ",
    "cpm": "CPM",
    "cpa": "CPO",
    "cpl": "CPL",
    "cpql": "CPQL",
    "drr_pct": "Р”Р Р ",
    "drr": "Р”Р Р ",
    "ROAS": "ROAS",
    "sov_pct": "SOV, %",
    "available_capacity": "Р”РѕСЃС‚СѓРїРЅР°СЏ РµРјРєРѕСЃС‚СЊ",
    "client_count": "РљРѕР»РёС‡РµСЃС‚РІРѕ РєР»РёРµРЅС‚РѕРІ",
    "absolute_new_clients": "РљРѕР»-РІРѕ Р°Р±СЃРѕР»СЋС‚РЅРѕ РЅРѕРІС‹С… РєР»РёРµРЅС‚РѕРІ",
    "returned_clients": "РљРѕР»-РІРѕ РІРµСЂРЅСѓРІС€РёС…СЃСЏ РєР»РёРµРЅС‚РѕРІ",
    "new_clients": "РљРѕР»-РІРѕ РЅРѕРІС‹С… РєР»РёРµРЅС‚РѕРІ",
    "cac": "CAC",
    "order_frequency": "Р§Р°СЃС‚РѕС‚Р° Р·Р°РєР°Р·Р°",
    "new_clients_share_pct": "Р”РѕР»СЏ РЅРѕРІС‹С… РєР»РёРµРЅС‚РѕРІ, %",
    "budget_share_segment_pct": "Р”РѕР»СЏ СЂРµРєР»Р°РјРЅРѕРіРѕ Р±СЋРґР¶РµС‚Р°, %",
    "order_share_segment_pct": "Р”РѕР»СЏ Р·Р°РєР°Р·РѕРІ, %",
    "revenue_share_segment_pct": "Р”РѕР»СЏ РґРѕС…РѕРґР°, %",
    "shipped_orders": "РљРѕР»-РІРѕ РѕС‚РіСЂСѓР¶РµРЅРЅС‹С… Р·Р°РєР°Р·РѕРІ",
    "shipped_cps": "РЎС‚РѕРёРјРѕСЃС‚СЊ РѕС‚РіСЂ. Р·Р°РєР°Р·Р°, в‚Ѕ СЃ РќР”РЎ",
    "shipped_order_share_segment_pct": "Р”РѕР»СЏ РѕС‚РіСЂ. Р·Р°РєР°Р·РѕРІ, %",
    "shipped_revenue": "Р’С‹СЂСѓС‡РєР° (РѕС‚РіСЂСѓР¶РµРЅРЅС‹Р№ РґРѕС…РѕРґ), в‚Ѕ СЃ РќР”РЎ",
    "shipped_roas": "ROAS РѕС‚РіСЂ.",
    "shipped_drr_pct": "Р”Р Р  РѕС‚РіСЂ., %",
    "shipped_aov": "РЎСЂРµРґРЅРёР№ С‡РµРє РїРѕ РІС‹СЂСѓС‡РєРµ, СЃ РќР”РЎ",
    "shipped_revenue_share_segment_pct": "Р”РѕР»СЏ РІС‹СЂСѓС‡РєРё, %",
}

METRIC_HELP = {
    "impressions": "РџРѕРєР°Р·С‹: РєРѕР»РёС‡РµСЃС‚РІРѕ СЂРµРєР»Р°РјРЅС‹С… РїРѕРєР°Р·РѕРІ.",
    "reach": "РћС…РІР°С‚: С‡РёСЃР»Рѕ СѓРЅРёРєР°Р»СЊРЅС‹С… РїРѕР»СЊР·РѕРІР°С‚РµР»РµР№, СѓРІРёРґРµРІС€РёС… СЂРµРєР»Р°РјСѓ.",
    "frequency": "Р§Р°СЃС‚РѕС‚Р° = РџРѕРєР°Р·С‹ / РћС…РІР°С‚.",
    "clicks": "РљР»РёРєРё: РџРѕРєР°Р·С‹ Г— CTR.",
    "ctr": "CTR, % = РљР»РёРєРё / РџРѕРєР°Р·С‹ Г— 100%.",
    "cpc": "CPC = Р‘СЋРґР¶РµС‚ / РљР»РёРєРё.",
    "cr": "CR, % = РљРѕРЅРІРµСЂСЃРёРё / РљР»РёРєРё Г— 100%.",
    "aov": "AOV = Р”РѕС…РѕРґ / РљРѕРЅРІРµСЂСЃРёРё.",
    "conversions": "РљРѕРЅРІРµСЂСЃРёРё = РљР»РёРєРё Г— CR.",
    "cost": "Р‘СЋРґР¶РµС‚ = РљР»РёРєРё Г— CPC.",
    "cost_with_vat": "Р‘СЋРґР¶РµС‚ СЃ РќР”РЎ = Р‘СЋРґР¶РµС‚ Г— 1.22 (РїСЂРё РІРєР»СЋС‡РµРЅРЅРѕРј СѓС‡РµС‚Рµ РќР”РЎ).",
    "cost_with_vat_ak": "Р‘СЋРґР¶РµС‚ СЃ РќР”РЎ Рё РђРљ = (Р‘СЋРґР¶РµС‚ + РђРљ Р±РµР· РќР”РЎ) Г— 1.22.",
    "total_budget_wo_vat_ak": "РўРѕС‚Р°Р» Р±СЋРґР¶РµС‚ Р±РµР· РќР”РЎ СЃ СѓС‡РµС‚РѕРј РђРљ = Р‘СЋРґР¶РµС‚ + РђРљ Р±РµР· РќР”РЎ.",
    "vat_amount": "РќР”РЎ, в‚Ѕ = РўРѕС‚Р°Р» Р±СЋРґР¶РµС‚ СЃ РќР”РЎ Рё РђРљ - РўРѕС‚Р°Р» Р±СЋРґР¶РµС‚ Р±РµР· РќР”РЎ Рё РђРљ.",
    "revenue": "Р”РѕС…РѕРґ = РљРѕРЅРІРµСЂСЃРёРё Г— AOV.",
    "client_count": "РљРѕР»РёС‡РµСЃС‚РІРѕ РєР»РёРµРЅС‚РѕРІ: Р±Р°Р·РѕРІРѕРµ Р·РЅР°С‡РµРЅРёРµ РїРѕ С‚РёРїСѓ Р Рљ, Рє РєРѕС‚РѕСЂРѕРјСѓ РїСЂРёРјРµРЅСЏРµС‚СЃСЏ РјРµСЃСЏС‡РЅС‹Р№ РєРѕСЌС„С„РёС†РёРµРЅС‚.",
    "absolute_new_clients": "РђР±СЃРѕР»СЋС‚РЅРѕ РЅРѕРІС‹Рµ РєР»РёРµРЅС‚С‹: Р±Р°Р·РѕРІРѕРµ Р·РЅР°С‡РµРЅРёРµ РїРѕ С‚РёРїСѓ Р Рљ, Рє РєРѕС‚РѕСЂРѕРјСѓ РїСЂРёРјРµРЅСЏРµС‚СЃСЏ РјРµСЃСЏС‡РЅС‹Р№ РєРѕСЌС„С„РёС†РёРµРЅС‚.",
    "returned_clients": "Р’РµСЂРЅСѓРІС€РёРµСЃСЏ РєР»РёРµРЅС‚С‹: Р±Р°Р·РѕРІРѕРµ Р·РЅР°С‡РµРЅРёРµ РїРѕ С‚РёРїСѓ Р Рљ, Рє РєРѕС‚РѕСЂРѕРјСѓ РїСЂРёРјРµРЅСЏРµС‚СЃСЏ РјРµСЃСЏС‡РЅС‹Р№ РєРѕСЌС„С„РёС†РёРµРЅС‚.",
    "new_clients": "РќРѕРІС‹Рµ РєР»РёРµРЅС‚С‹ = РђР±СЃРѕР»СЋС‚РЅРѕ РЅРѕРІС‹Рµ РєР»РёРµРЅС‚С‹ + Р’РµСЂРЅСѓРІС€РёРµСЃСЏ РєР»РёРµРЅС‚С‹.",
    "cac": "CAC = Р‘СЋРґР¶РµС‚ СЃ РќР”РЎ Рё РђРљ / РљРѕР»-РІРѕ РЅРѕРІС‹С… РєР»РёРµРЅС‚РѕРІ.",
    "order_frequency": "Р§Р°СЃС‚РѕС‚Р° Р·Р°РєР°Р·Р°: Р±Р°Р·РѕРІРѕРµ Р·РЅР°С‡РµРЅРёРµ РїРѕ С‚РёРїСѓ Р Рљ, Рє РєРѕС‚РѕСЂРѕРјСѓ РїСЂРёРјРµРЅСЏРµС‚СЃСЏ РјРµСЃСЏС‡РЅС‹Р№ РєРѕСЌС„С„РёС†РёРµРЅС‚.",
    "shipped_orders": "РћС‚РіСЂСѓР¶РµРЅРЅС‹Рµ Р·Р°РєР°Р·С‹ = РћС„РѕСЂРјР»РµРЅРЅС‹Рµ Р·Р°РєР°Р·С‹ Г— CR2.",
    "shipped_cps": "РЎС‚РѕРёРјРѕСЃС‚СЊ РѕС‚РіСЂ. Р·Р°РєР°Р·Р°, в‚Ѕ СЃ РќР”РЎ = Р‘СЋРґР¶РµС‚ СЃ РќР”РЎ / РљРѕР»-РІРѕ РѕС‚РіСЂСѓР¶РµРЅРЅС‹С… Р·Р°РєР°Р·РѕРІ.",
    "shipped_revenue": "РћС‚РіСЂСѓР¶РµРЅРЅС‹Р№ РґРѕС…РѕРґ = РљРѕР»-РІРѕ РѕС‚РіСЂСѓР¶РµРЅРЅС‹С… Р·Р°РєР°Р·РѕРІ Г— РЎСЂРµРґРЅРёР№ С‡РµРє.",
    "shipped_roas": "ROAS РѕС‚РіСЂ. = РћС‚РіСЂСѓР¶РµРЅРЅС‹Р№ РґРѕС…РѕРґ / Р‘СЋРґР¶РµС‚РЅР°СЏ Р±Р°Р·Р°.",
    "shipped_drr_pct": "Р”Р Р  РѕС‚РіСЂ., % = Р‘СЋРґР¶РµС‚РЅР°СЏ Р±Р°Р·Р° / РћС‚РіСЂСѓР¶РµРЅРЅС‹Р№ РґРѕС…РѕРґ Г— 100%.",
    "shipped_aov": "РЎСЂРµРґРЅРёР№ С‡РµРє РїРѕ РІС‹СЂСѓС‡РєРµ, СЃ РќР”РЎ = РћС‚РіСЂСѓР¶РµРЅРЅС‹Р№ РґРѕС…РѕРґ / РљРѕР»-РІРѕ РѕС‚РіСЂСѓР¶РµРЅРЅС‹С… Р·Р°РєР°Р·РѕРІ.",
    "cpm": "CPM = Р‘СЋРґР¶РµС‚РЅР°СЏ Р±Р°Р·Р° / (РџРѕРєР°Р·С‹ / 1000). Р‘Р°Р·Р° Р·Р°РІРёСЃРёС‚ РѕС‚ СЂРµР¶РёРјР° РќР”РЎ.",
    "cpo": "CPO = Р‘СЋРґР¶РµС‚РЅР°СЏ Р±Р°Р·Р° / РљРѕРЅРІРµСЂСЃРёРё. Р‘Р°Р·Р° Р·Р°РІРёСЃРёС‚ РѕС‚ СЂРµР¶РёРјР° РќР”РЎ.",
    "roas": "ROAS, % = Р”РѕС…РѕРґ / Р‘СЋРґР¶РµС‚РЅР°СЏ Р±Р°Р·Р° Г— 100%. Р‘Р°Р·Р° Р·Р°РІРёСЃРёС‚ РѕС‚ СЂРµР¶РёРјР° РќР”РЎ.",
    "drr": "Р”Р Р , % = Р‘СЋРґР¶РµС‚РЅР°СЏ Р±Р°Р·Р° / Р”РѕС…РѕРґ Г— 100%. Р‘Р°Р·Р° Р·Р°РІРёСЃРёС‚ РѕС‚ СЂРµР¶РёРјР° РќР”РЎ.",
    "k_imp": "k_imp: РєРѕСЌС„С„РёС†РёРµРЅС‚ РїРѕРєР°Р·РѕРІ. РџСЂРёРјРµРЅСЏРµС‚СЃСЏ Рє РџРѕРєР°Р·С‹.",
    "k_reach": "k_reach: РєРѕСЌС„С„РёС†РёРµРЅС‚ РѕС…РІР°С‚Р°. РџСЂРёРјРµРЅСЏРµС‚СЃСЏ Рє РћС…РІР°С‚.",
    "k_ctr": "k_ctr: РєРѕСЌС„С„РёС†РёРµРЅС‚ CTR. РџСЂРёРјРµРЅСЏРµС‚СЃСЏ Рє CTR.",
    "k_cpc": "k_cpc: РєРѕСЌС„С„РёС†РёРµРЅС‚ CPC. РџСЂРёРјРµРЅСЏРµС‚СЃСЏ Рє CPC.",
    "k_cr": "k_cr: РєРѕСЌС„С„РёС†РёРµРЅС‚ CR. РџСЂРёРјРµРЅСЏРµС‚СЃСЏ Рє CR.",
    "k_aov": "k_aov: РєРѕСЌС„С„РёС†РёРµРЅС‚ AOV. РџСЂРёРјРµРЅСЏРµС‚СЃСЏ Рє AOV.",
    "k_client_count": "k_client_count: РєРѕСЌС„С„РёС†РёРµРЅС‚ РєРѕР»РёС‡РµСЃС‚РІР° РєР»РёРµРЅС‚РѕРІ. РџСЂРёРјРµРЅСЏРµС‚СЃСЏ Рє РјРµС‚СЂРёРєРµ РљРѕР»РёС‡РµСЃС‚РІРѕ РєР»РёРµРЅС‚РѕРІ.",
    "k_absolute_new_clients": "k_absolute_new_clients: РєРѕСЌС„С„РёС†РёРµРЅС‚ РґР»СЏ Р°Р±СЃРѕР»СЋС‚РЅРѕ РЅРѕРІС‹С… РєР»РёРµРЅС‚РѕРІ.",
    "k_returned_clients": "k_returned_clients: РєРѕСЌС„С„РёС†РёРµРЅС‚ РґР»СЏ РІРµСЂРЅСѓРІС€РёС…СЃСЏ РєР»РёРµРЅС‚РѕРІ.",
    "k_order_frequency": "k_order_frequency: РєРѕСЌС„С„РёС†РёРµРЅС‚ С‡Р°СЃС‚РѕС‚С‹ Р·Р°РєР°Р·Р°.",
    "cpc_div": "Р”РµР»РёС‚РµР»СЊ РІР»РёСЏРЅРёСЏ РЅР° CPC: k_cpc = 1 + (k_demand - 1) / div.",
    "ctr_div": "Р”РµР»РёС‚РµР»СЊ РІР»РёСЏРЅРёСЏ РЅР° CTR: k_ctr = 1 - (k_demand - 1) / div.",
    "cr_div": "Р”РµР»РёС‚РµР»СЊ РІР»РёСЏРЅРёСЏ РЅР° CR: k_cr = 1 - (k_demand - 1) / div.",
}


def mhelp(key: str, fallback: str = "") -> str:
    return METRIC_HELP.get(key, fallback)


def reorder_rows_with_segment_subtotals(
    df_show: pd.DataFrame,
    campaign_col: str,
    segment_col: str,
) -> pd.DataFrame:
    """Order rows as: campaigns by segment -> segment subtotal -> ... -> grand total."""
    if df_show is None or df_show.empty:
        return df_show
    if campaign_col not in df_show.columns or segment_col not in df_show.columns:
        return df_show

    work = df_show.copy()
    campaign_vals = work[campaign_col].astype(str)
    is_total_any = campaign_vals.str.startswith("РС‚РѕРіРѕ")
    is_total_grand = campaign_vals.eq("РС‚РѕРіРѕ")
    is_total_segment = is_total_any & ~is_total_grand
    is_campaign = ~is_total_any

    seg_order = ["B2C", "B2B"]
    segments_present = work.loc[is_campaign, segment_col].astype(str).unique().tolist()
    for seg in segments_present:
        if seg not in seg_order:
            seg_order.append(seg)

    ordered_parts = []
    used_idx: set[int] = set()
    for seg in seg_order:
        seg_campaigns = work[is_campaign & (work[segment_col].astype(str) == seg)]
        if not seg_campaigns.empty:
            ordered_parts.append(seg_campaigns)
            used_idx.update(seg_campaigns.index.tolist())
            seg_total = work[is_total_segment & (campaign_vals == f"РС‚РѕРіРѕ {seg}")]
            if not seg_total.empty:
                ordered_parts.append(seg_total)
                used_idx.update(seg_total.index.tolist())

    # Fallback for any rows that didn't get included.
    if ordered_parts:
        ordered = pd.concat(ordered_parts, ignore_index=True)
        campaign_rows_left = work[is_campaign & ~work.index.isin(list(used_idx))]
        seg_totals_left = work[is_total_segment & ~work.index.isin(list(used_idx))]
        if not campaign_rows_left.empty:
            ordered = pd.concat([ordered, campaign_rows_left], ignore_index=True)
            used_idx.update(campaign_rows_left.index.tolist())
        if not seg_totals_left.empty:
            ordered = pd.concat([ordered, seg_totals_left], ignore_index=True)
            used_idx.update(seg_totals_left.index.tolist())
    else:
        ordered = work[is_campaign].copy()
        used_idx.update(ordered.index.tolist())

    grand_total = work[is_total_grand & ~work.index.isin(list(used_idx))]
    if not grand_total.empty:
        ordered = pd.concat([ordered, grand_total], ignore_index=True)
    return ordered


def add_segment_budget_share(
    df: pd.DataFrame | None,
    segment_col: str = "segment",
    budget_col: str = "cost_with_vat_ak",
    out_col: str = "budget_share_segment_pct",
) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df
    out = df.copy()
    if segment_col not in out.columns or budget_col not in out.columns:
        out[out_col] = 0.0
        return out
    seg_total = out.groupby(segment_col)[budget_col].transform("sum")
    out[out_col] = np.where(seg_total > 0, out[budget_col] / seg_total * 100.0, 0.0)
    return out


def add_segment_value_share(
    df: pd.DataFrame | None,
    segment_col: str = "segment",
    value_col: str = "conversions",
    out_col: str = "order_share_segment_pct",
) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df
    out = df.copy()
    if segment_col not in out.columns or value_col not in out.columns:
        out[out_col] = 0.0
        return out
    values = pd.to_numeric(out[value_col], errors="coerce").fillna(0.0)
    seg_total = values.groupby(out[segment_col]).transform("sum")
    out[out_col] = np.where(seg_total > 0, values / seg_total * 100.0, 0.0)
    return out


def forecast_ets_like(
    series: list[float],
    periods: int = 12,
    season_length: int = 12,
    alpha: float = 0.35,
    beta: float = 0.10,
    gamma: float = 0.25,
) -> list[float]:
    """
    ETS forecast with statsmodels when available; otherwise fallback to a
    lightweight additive Holt-Winters implementation.
    """
    vals = [float(x) for x in series if pd.notna(x)]
    n = len(vals)
    if n == 0:
        return [0.0] * periods

    # Primary path: statsmodels ETS (closest to Excel FORECAST.ETS behavior).
    if HAS_STATSMODELS and n >= season_length * 2:
        try:
            # Keep smoothing params inside valid open interval for stable fitting.
            a = min(max(float(alpha), 1e-4), 0.9999)
            b = min(max(float(beta), 1e-4), 0.9999)
            g = min(max(float(gamma), 1e-4), 0.9999)

            model = ExponentialSmoothing(
                vals,
                trend="add",
                seasonal="add",
                seasonal_periods=season_length,
                initialization_method="estimated",
            )
            # Use user-selected alpha/beta/gamma (instead of auto-optimized fit).
            fitted = model.fit(
                optimized=False,
                smoothing_level=a,
                smoothing_trend=b,
                smoothing_seasonal=g,
            )
            fc = fitted.forecast(periods)
            return [max(0.0, float(v)) for v in fc]
        except Exception:
            # fallback below
            pass

    if n < season_length * 2:
        # For short history, repeat last available seasonal pattern.
        tail = vals[-season_length:] if n >= season_length else vals
        rep = []
        while len(rep) < periods:
            rep.extend(tail)
        return [max(0.0, float(v)) for v in rep[:periods]]

    n_seasons = n // season_length
    season_avgs = []
    for j in range(n_seasons):
        chunk = vals[j * season_length:(j + 1) * season_length]
        season_avgs.append(float(np.mean(chunk)))

    seasonals = [0.0] * season_length
    for i in range(season_length):
        acc = 0.0
        cnt = 0
        for j in range(n_seasons):
            idx = j * season_length + i
            if idx < n:
                acc += vals[idx] - season_avgs[j]
                cnt += 1
        seasonals[i] = acc / max(cnt, 1)

    level = vals[0] - seasonals[0]
    trend = (
        ((vals[season_length] - seasonals[0]) - (vals[0] - seasonals[0])) / season_length
        if n > season_length else 0.0
    )

    for t in range(n):
        s = seasonals[t % season_length]
        prev_level = level
        level = alpha * (vals[t] - s) + (1 - alpha) * (level + trend)
        trend = beta * (level - prev_level) + (1 - beta) * trend
        seasonals[t % season_length] = gamma * (vals[t] - level) + (1 - gamma) * s

    fc = []
    for m in range(1, periods + 1):
        idx = (n + m - 1) % season_length
        yhat = level + m * trend + seasonals[idx]
        fc.append(max(0.0, float(yhat)))
    return fc


# ---------- Р­РљРЎРџРћР Рў Р’ РЁРђР‘Р›РћРќ EXCEL ----------

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATHS_ECOM = [
    os.path.join(BASE_DIR, "templates", "РЁР°Р±Р»РѕРЅС‹ РњРџ.xlsx"),
    os.path.join(BASE_DIR, "РЁР°Р±Р»РѕРЅС‹ РњРџ.xlsx"),
    os.path.join(BASE_DIR, "Shablony-MP.xlsx"),
]
TEMPLATE_PATHS_DIY = [
    os.path.join(BASE_DIR, "templates", "РЁР°Р±Р»РѕРЅС‹ РњРџ DIY.xlsx"),
    os.path.join(BASE_DIR, "РЁР°Р±Р»РѕРЅС‹ РњРџ DIY.xlsx"),
    os.path.join(BASE_DIR, "templates", "Shablony-MP-DIY.xlsx"),
]
TEMPLATE_PATHS_REAL_ESTATE_SIMPLE = [
    os.path.join(BASE_DIR, "templates", "РЁР°Р±Р»РѕРЅС‹ РњРџ РќРµРґРІРёР¶РєР° СѓРїСЂ. РІРѕСЂРѕРЅРєР°.xlsx"),
    os.path.join(BASE_DIR, "РЁР°Р±Р»РѕРЅС‹ РњРџ РќРµРґРІРёР¶РєР° СѓРїСЂ. РІРѕСЂРѕРЅРєР°.xlsx"),
]
TEMPLATE_PATHS_REAL_ESTATE_FULL = [
    os.path.join(BASE_DIR, "templates", "РЁР°Р±Р»РѕРЅС‹ РњРџ РќРµРґРІРёР¶РєР° РїРѕР»РЅ. РІРѕСЂРѕРЅРєР°.xlsx"),
    os.path.join(BASE_DIR, "РЁР°Р±Р»РѕРЅС‹ РњРџ РќРµРґРІРёР¶РєР° РїРѕР»РЅ. РІРѕСЂРѕРЅРєР°.xlsx"),
]
TEMPLATE_PATHS_BY_KIND = {
    "ecom": TEMPLATE_PATHS_ECOM,
    "diy": TEMPLATE_PATHS_DIY,
    "real_estate_simple": TEMPLATE_PATHS_REAL_ESTATE_SIMPLE,
    "real_estate_full": TEMPLATE_PATHS_REAL_ESTATE_FULL,
}


def build_excel_from_template(df_all: pd.DataFrame,
                              campaigns: pd.DataFrame,
                              selected_month_nums: list[int],
                              template_kind: str = "ecom",
                              compact_empty_rows: bool = True) -> BytesIO:

    template_paths = TEMPLATE_PATHS_BY_KIND.get(template_kind, TEMPLATE_PATHS_ECOM)
    existing_templates = [p for p in template_paths if os.path.exists(p)]
    template_path = (
        max(existing_templates, key=os.path.getmtime)
        if existing_templates
        else template_paths[0]
    )

    wb = load_workbook(template_path)

    periods_candidates = [s for s in wb.sheetnames if str(s).endswith("_Periods")]
    total_candidates = [s for s in wb.sheetnames if str(s).endswith("_Total")]

    if periods_candidates:
        ws = wb[periods_candidates[0]]
    elif len(wb.worksheets) >= 2:
        ws = wb.worksheets[1]
    else:
        ws = wb.worksheets[0]

    if total_candidates:
        ws_total = wb[total_candidates[0]]
    else:
        ws_total = wb.worksheets[0]

    DEFAULT_START_ROW_JAN = 3
    DEFAULT_BLOCK_STEP = 13
    DEFAULT_ROWS_PER_MONTH = 10
    DEFAULT_START_ROW_TOTAL = 3

    COL_SYSTEM = "B"
    COL_FORMAT = "C"
    COL_TARGETING = "D"
    COL_PERIOD = "F"
    COL_MODEL = "G"

    COL_CPC = "H"
    is_real_estate_template = template_kind in {"real_estate_simple", "real_estate_full"}
    is_real_estate_full_template = template_kind == "real_estate_full"
    COL_GEO = "K" if is_real_estate_template else None
    COL_DEMAND_COEFF = "L" if is_real_estate_template else None
    COL_OTHER = "M" if is_real_estate_template else None
    COL_TOTAL_GEO = "X" if is_real_estate_template else None
    if is_real_estate_template:
        COL_AK = "Q"
        COL_IMPRESSIONS = "V"
        COL_CTR = "AD"
        COL_CR = "AH"
        COL_CR2 = "AL" if is_real_estate_full_template else None
    else:
        COL_AK = "N"
        COL_IMPRESSIONS = "S"
        COL_CTR = "AA"
        COL_CR = "AE"
        COL_CR2 = None
    # In DIY template AN is a formula column ("share of revenue"), while AOV is in AL.
    if template_kind == "diy":
        COL_AOV = "AL"
    elif is_real_estate_template:
        COL_AOV = None
    else:
        COL_AOV = "AN"
    COL_NEW_CLIENTS_SHARE = "W" if template_kind == "diy" else None
    COL_AVAILABLE_CAPACITY = "X" if template_kind == "diy" else None
    COL_REACH = "U" if template_kind == "diy" else None
    COL_FREQUENCY = "T" if template_kind == "diy" else None

    def _safe_text(v) -> str:
        if v is None:
            return ""
        return str(v).strip().lower()

    def _norm_segment(v) -> str:
        t = _safe_text(v).replace(" ", "")
        if "b2b" in t:
            return "B2B"
        if "b2c" in t:
            return "B2C"
        return ""

    def _is_header_row(sheet, r: int) -> bool:
        b = sheet[f"{COL_SYSTEM}{r}"].value
        c = sheet[f"{COL_FORMAT}{r}"].value
        d = sheet[f"{COL_TARGETING}{r}"].value
        f = sheet[f"{COL_PERIOD}{r}"].value
        s = sheet[f"{COL_IMPRESSIONS}{r}"].value
        if not all(v is not None and str(v).strip() != "" for v in (b, c, d, f, s)):
            return False
        # Exclude subtotal/total rows where template keeps dashes in C/D.
        if str(c).strip() == "-" and str(d).strip() == "-":
            return False
        return True

    def _detect_periods_layout() -> tuple[int, int, int]:
        header_rows = []
        for r in range(1, min(ws.max_row, 500) + 1):
            if _is_header_row(ws, r):
                header_rows.append(r)
                if len(header_rows) >= 3:
                    break

        if not header_rows:
            return DEFAULT_START_ROW_JAN, DEFAULT_BLOCK_STEP, DEFAULT_ROWS_PER_MONTH

        header_row = header_rows[0]
        next_header_row = header_rows[1] if len(header_rows) > 1 else None
        start_row = header_row + 1

        if next_header_row:
            rows_per_month = max(1, next_header_row - start_row - 2)
            block_step = next_header_row - header_row
        else:
            search_end = min(ws.max_row, start_row + 50)
            total_row = None
            for r in range(start_row, search_end + 1):
                c = _safe_text(ws[f"{COL_FORMAT}{r}"].value)
                d = _safe_text(ws[f"{COL_TARGETING}{r}"].value)
                b = _safe_text(ws[f"{COL_SYSTEM}{r}"].value)
                if b and c == "-" and d == "-":
                    total_row = r
                    break
            rows_per_month = max(1, total_row - start_row) if total_row else DEFAULT_ROWS_PER_MONTH
            block_step = max(DEFAULT_BLOCK_STEP, rows_per_month + 2)

        return start_row, block_step, rows_per_month

    def _detect_total_layout(default_rows_per_month: int) -> tuple[int, int]:
        header_row = None
        for r in range(1, min(ws_total.max_row, 200) + 1):
            if _is_header_row(ws_total, r):
                header_row = r
                break

        if header_row is None:
            return DEFAULT_START_ROW_TOTAL, default_rows_per_month

        start_row = header_row + 1
        total_like_row = None
        for r in range(start_row, min(ws_total.max_row, start_row + 100) + 1):
            c = _safe_text(ws_total[f"{COL_FORMAT}{r}"].value)
            d = _safe_text(ws_total[f"{COL_TARGETING}{r}"].value)
            b = _safe_text(ws_total[f"{COL_SYSTEM}{r}"].value)
            if b and c == "-" and d == "-":
                total_like_row = r
                break

        rows_per_month = max(1, total_like_row - start_row) if total_like_row else default_rows_per_month
        return start_row, rows_per_month

    START_ROW_JAN, BLOCK_STEP, ROWS_PER_MONTH = _detect_periods_layout()
    START_ROW_TOTAL, ROWS_PER_MONTH_TOTAL = _detect_total_layout(ROWS_PER_MONTH)

    def _write_period_row(row_excel: int, camp_row: pd.Series | None, row_data: pd.Series | None, period_str: str):
        ws.row_dimensions[row_excel].hidden = False
        if camp_row is None or row_data is None:
            ws[f"{COL_SYSTEM}{row_excel}"] = None
            ws[f"{COL_FORMAT}{row_excel}"] = None
            ws[f"{COL_TARGETING}{row_excel}"] = None
            ws[f"{COL_PERIOD}{row_excel}"] = None
            ws[f"{COL_MODEL}{row_excel}"] = None
            if COL_GEO:
                ws[f"{COL_GEO}{row_excel}"] = None
            if COL_DEMAND_COEFF:
                ws[f"{COL_DEMAND_COEFF}{row_excel}"] = None
            if COL_OTHER:
                ws[f"{COL_OTHER}{row_excel}"] = None
            ws[f"{COL_IMPRESSIONS}{row_excel}"] = None
            ws[f"{COL_CTR}{row_excel}"] = None
            ws[f"{COL_CPC}{row_excel}"] = None
            ws[f"{COL_AK}{row_excel}"] = None
            ws[f"{COL_CR}{row_excel}"] = None
            if COL_CR2:
                ws[f"{COL_CR2}{row_excel}"] = None
            if COL_AOV:
                ws[f"{COL_AOV}{row_excel}"] = None
            if COL_NEW_CLIENTS_SHARE:
                ws[f"{COL_NEW_CLIENTS_SHARE}{row_excel}"] = None
            if COL_AVAILABLE_CAPACITY:
                ws[f"{COL_AVAILABLE_CAPACITY}{row_excel}"] = None
            if COL_REACH:
                ws[f"{COL_REACH}{row_excel}"] = None
            if COL_FREQUENCY:
                ws[f"{COL_FREQUENCY}{row_excel}"] = None
            return

        impressions = float(row_data["impressions"])
        ctr = float(row_data["ctr"])
        cpc = float(row_data["cpc"])
        cr = float(row_data["cr1"] if is_real_estate_full_template else row_data["cr"])
        cr2 = float(row_data.get("cr2", 0.0) or 0.0)
        aov = float(row_data.get("aov", 0.0) or 0.0)
        ak_rate = float(row_data.get("ak_rate", 0.0))
        if pd.isna(ak_rate):
            ak_rate = 0.0
        if pd.isna(cr2):
            cr2 = 0.0
        new_clients_share_pct = float(row_data.get("new_clients_share_pct", 0.0))
        if pd.isna(new_clients_share_pct):
            new_clients_share_pct = 0.0
        # App stores this metric in percents (e.g. 25), while Excel percent cells
        # typically expect a fraction (0.25) to display 25%.
        if new_clients_share_pct > 1:
            new_clients_share_pct = new_clients_share_pct / 100.0
        available_capacity = float(row_data.get("available_capacity", 0.0))
        if pd.isna(available_capacity):
            available_capacity = 0.0
        reach = float(row_data.get("reach", 0.0) or 0.0)
        if pd.isna(reach):
            reach = 0.0
        frequency = float(row_data.get("frequency", 0.0) or 0.0)
        if pd.isna(frequency):
            frequency = (impressions / reach) if reach > 0 else 0.0
        demand_coeff = float(row_data.get("k_demand_applied", 1.0) or 1.0)
        if pd.isna(demand_coeff):
            demand_coeff = 1.0

        ws[f"{COL_SYSTEM}{row_excel}"] = camp_row.get("system", "")
        ws[f"{COL_FORMAT}{row_excel}"] = camp_row.get("campaign_type", "")
        ws[f"{COL_TARGETING}{row_excel}"] = camp_row.get("format", "")
        ws[f"{COL_PERIOD}{row_excel}"] = period_str
        ws[f"{COL_MODEL}{row_excel}"] = "CPC"
        if COL_GEO:
            ws[f"{COL_GEO}{row_excel}"] = camp_row.get("geo", "")
        if COL_DEMAND_COEFF:
            ws[f"{COL_DEMAND_COEFF}{row_excel}"] = demand_coeff
        if COL_OTHER:
            ws[f"{COL_OTHER}{row_excel}"] = None
        ws[f"{COL_IMPRESSIONS}{row_excel}"] = impressions
        ws[f"{COL_CTR}{row_excel}"] = ctr
        ws[f"{COL_CPC}{row_excel}"] = cpc
        ws[f"{COL_AK}{row_excel}"] = ak_rate
        ws[f"{COL_CR}{row_excel}"] = cr
        if COL_CR2:
            ws[f"{COL_CR2}{row_excel}"] = cr2
        if COL_AOV:
            ws[f"{COL_AOV}{row_excel}"] = aov
        if COL_NEW_CLIENTS_SHARE:
            ws[f"{COL_NEW_CLIENTS_SHARE}{row_excel}"] = new_clients_share_pct
        if COL_AVAILABLE_CAPACITY:
            ws[f"{COL_AVAILABLE_CAPACITY}{row_excel}"] = available_capacity
        if COL_REACH:
            ws[f"{COL_REACH}{row_excel}"] = reach
        if COL_FREQUENCY:
            ws[f"{COL_FREQUENCY}{row_excel}"] = frequency

    def _write_total_row(row_excel: int, camp_row: pd.Series | None):
        ws_total.row_dimensions[row_excel].hidden = False
        if camp_row is None:
            ws_total[f"{COL_SYSTEM}{row_excel}"] = None
            ws_total[f"{COL_FORMAT}{row_excel}"] = None
            ws_total[f"{COL_TARGETING}{row_excel}"] = None
            if COL_TOTAL_GEO:
                ws_total[f"{COL_TOTAL_GEO}{row_excel}"] = None
            return
        ws_total[f"{COL_SYSTEM}{row_excel}"] = camp_row.get("system", "")
        ws_total[f"{COL_FORMAT}{row_excel}"] = camp_row.get("campaign_type", "")
        ws_total[f"{COL_TARGETING}{row_excel}"] = camp_row.get("format", "")
        if COL_TOTAL_GEO:
            ws_total[f"{COL_TOTAL_GEO}{row_excel}"] = camp_row.get("geo", "")

    def _collect_diy_period_rows(block_start_row: int) -> tuple[list[int], list[int]] | None:
        scan_from = block_start_row
        scan_to = min(ws.max_row, block_start_row + BLOCK_STEP + 8)
        row_b2c_total = None
        row_b2b_total = None
        for r in range(scan_from, scan_to + 1):
            t = _safe_text(ws[f"{COL_SYSTEM}{r}"].value)
            c = _safe_text(ws[f"{COL_FORMAT}{r}"].value)
            d = _safe_text(ws[f"{COL_TARGETING}{r}"].value)
            if c != "-" or d != "-":
                continue
            if "b2c" in t and row_b2c_total is None:
                row_b2c_total = r
            elif "b2b" in t and row_b2b_total is None:
                row_b2b_total = r

        if row_b2c_total is None or row_b2b_total is None or row_b2b_total <= row_b2c_total:
            return None

        b2c_rows = list(range(block_start_row, row_b2c_total))
        b2b_rows = list(range(row_b2c_total + 1, row_b2b_total))
        return b2c_rows, b2b_rows

    def _collect_diy_total_rows() -> tuple[list[int], list[int]] | None:
        scan_from = START_ROW_TOTAL
        scan_to = min(ws_total.max_row, START_ROW_TOTAL + ROWS_PER_MONTH_TOTAL + 40)
        row_b2c_total = None
        row_b2b_total = None
        for r in range(scan_from, scan_to + 1):
            t = _safe_text(ws_total[f"{COL_SYSTEM}{r}"].value)
            c = _safe_text(ws_total[f"{COL_FORMAT}{r}"].value)
            d = _safe_text(ws_total[f"{COL_TARGETING}{r}"].value)
            if c != "-" or d != "-":
                continue
            if "b2c" in t and row_b2c_total is None:
                row_b2c_total = r
            elif "b2b" in t and row_b2b_total is None:
                row_b2b_total = r

        if row_b2c_total is None or row_b2b_total is None or row_b2b_total <= row_b2c_total:
            return None

        b2c_rows = list(range(START_ROW_TOTAL, row_b2c_total))
        b2b_rows = list(range(row_b2c_total + 1, row_b2b_total))
        return b2c_rows, b2b_rows

    year = dt.date.today().year

    for block_index, month_num in enumerate(selected_month_nums):
        block_start_row = START_ROW_JAN + block_index * BLOCK_STEP
        start = dt.date(year, month_num, 1)
        end = dt.date(year, 12, 31) if month_num == 12 else dt.date(year, month_num + 1, 1) - dt.timedelta(days=1)
        period_str = f"{start.strftime('%d.%m.%Y')} - {end.strftime('%d.%m.%Y')}"
        df_month = df_all[df_all["month_num"] == month_num]

        diy_period_rows = _collect_diy_period_rows(block_start_row) if (template_kind == "diy" and "segment" in campaigns.columns) else None

        if diy_period_rows is not None:
            b2c_rows, b2b_rows = diy_period_rows
            camps = campaigns.copy()
            camps["_seg"] = camps["segment"].map(_norm_segment)
            camps_b2c = camps[camps["_seg"] == "B2C"]
            camps_b2b = camps[camps["_seg"] == "B2B"]

            for row_excel, (_, camp) in zip(b2c_rows, camps_b2c.iterrows()):
                row_data = df_month[df_month["campaign_type"] == camp["campaign_type"]]
                _write_period_row(row_excel, camp, (None if row_data.empty else row_data.iloc[0]), period_str)
            for row_excel in b2c_rows[len(camps_b2c):]:
                _write_period_row(row_excel, None, None, period_str)

            for row_excel, (_, camp) in zip(b2b_rows, camps_b2b.iterrows()):
                row_data = df_month[df_month["campaign_type"] == camp["campaign_type"]]
                _write_period_row(row_excel, camp, (None if row_data.empty else row_data.iloc[0]), period_str)
            for row_excel in b2b_rows[len(camps_b2b):]:
                _write_period_row(row_excel, None, None, period_str)
        else:
            for i, (_, camp) in enumerate(campaigns.iterrows()):
                if i >= ROWS_PER_MONTH:
                    break
                row_excel = block_start_row + i
                row_data = df_month[df_month["campaign_type"] == camp["campaign_type"]]
                _write_period_row(row_excel, camp, (None if row_data.empty else row_data.iloc[0]), period_str)
            for i in range(min(len(campaigns), ROWS_PER_MONTH), ROWS_PER_MONTH):
                _write_period_row(block_start_row + i, None, None, period_str)

    if compact_empty_rows:
        rows_to_hide_periods = []
        for block_index, _ in enumerate(selected_month_nums):
            block_start_row = START_ROW_JAN + block_index * BLOCK_STEP
            for i in range(ROWS_PER_MONTH):
                row_excel = block_start_row + i
                is_empty_main = (
                    ws[f"{COL_SYSTEM}{row_excel}"].value in (None, "")
                    and ws[f"{COL_FORMAT}{row_excel}"].value in (None, "")
                    and ws[f"{COL_TARGETING}{row_excel}"].value in (None, "")
                )
                is_empty_metrics = (
                    ws[f"{COL_IMPRESSIONS}{row_excel}"].value in (None, "")
                    and ws[f"{COL_CTR}{row_excel}"].value in (None, "")
                    and ws[f"{COL_CPC}{row_excel}"].value in (None, "")
                    and ws[f"{COL_AK}{row_excel}"].value in (None, "")
                    and ws[f"{COL_CR}{row_excel}"].value in (None, "")
                    and (not COL_AOV or ws[f"{COL_AOV}{row_excel}"].value in (None, ""))
                    and (not COL_NEW_CLIENTS_SHARE or ws[f"{COL_NEW_CLIENTS_SHARE}{row_excel}"].value in (None, ""))
                    and (not COL_AVAILABLE_CAPACITY or ws[f"{COL_AVAILABLE_CAPACITY}{row_excel}"].value in (None, ""))
                    and (not COL_REACH or ws[f"{COL_REACH}{row_excel}"].value in (None, ""))
                    and (not COL_FREQUENCY or ws[f"{COL_FREQUENCY}{row_excel}"].value in (None, ""))
                )
                if is_empty_main and is_empty_metrics:
                    rows_to_hide_periods.append(row_excel)
        for row_idx in sorted(set(rows_to_hide_periods)):
            ws.row_dimensions[row_idx].hidden = True

    diy_total_rows = _collect_diy_total_rows() if (template_kind == "diy" and "segment" in campaigns.columns) else None

    if diy_total_rows is not None:
        b2c_rows_t, b2b_rows_t = diy_total_rows
        total_rows_for_hide = list(b2c_rows_t) + list(b2b_rows_t)
        camps = campaigns.copy()
        camps["_seg"] = camps["segment"].map(_norm_segment)
        camps_b2c = camps[camps["_seg"] == "B2C"]
        camps_b2b = camps[camps["_seg"] == "B2B"]

        for row_excel, (_, camp) in zip(b2c_rows_t, camps_b2c.iterrows()):
            _write_total_row(row_excel, camp)
        for row_excel in b2c_rows_t[len(camps_b2c):]:
            _write_total_row(row_excel, None)

        for row_excel, (_, camp) in zip(b2b_rows_t, camps_b2b.iterrows()):
            _write_total_row(row_excel, camp)
        for row_excel in b2b_rows_t[len(camps_b2b):]:
            _write_total_row(row_excel, None)
    else:
        total_rows_for_hide = list(range(START_ROW_TOTAL, START_ROW_TOTAL + ROWS_PER_MONTH_TOTAL))
        for i in range(ROWS_PER_MONTH_TOTAL):
            row_excel = START_ROW_TOTAL + i
            if i < len(campaigns):
                _write_total_row(row_excel, campaigns.iloc[i])
            else:
                _write_total_row(row_excel, None)

    if compact_empty_rows:
        rows_to_hide_total = []
        for row_excel in total_rows_for_hide:
            if (
                ws_total[f"{COL_SYSTEM}{row_excel}"].value in (None, "")
                and ws_total[f"{COL_FORMAT}{row_excel}"].value in (None, "")
                and ws_total[f"{COL_TARGETING}{row_excel}"].value in (None, "")
            ):
                rows_to_hide_total.append(row_excel)
        for row_idx in sorted(set(rows_to_hide_total)):
            ws_total.row_dimensions[row_idx].hidden = True

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
def resolve_template_path(template_kind: str = "ecom") -> str | None:
    template_paths = TEMPLATE_PATHS_BY_KIND.get(template_kind, TEMPLATE_PATHS_ECOM)
    existing_templates = [p for p in template_paths if os.path.exists(p)]
    if not existing_templates:
        return None
    return max(existing_templates, key=os.path.getmtime)


# ---------- Р—РђР“РћР›РћР’РћРљ  РўРђР‘Р« ----------

coeff_header_uri = image_file_to_data_uri(os.path.join(BASE_DIR, "assets", "coeff_header_mascot.png"))
plan_header_uri = image_file_to_data_uri(os.path.join(BASE_DIR, "assets", "plan_header_mascot.png"))
charts_header_uri = image_file_to_data_uri(os.path.join(BASE_DIR, "assets", "charts_header_mascot.png"))
export_header_uri = image_file_to_data_uri(os.path.join(BASE_DIR, "assets", "export_header_mascot.png"))
top_header_uri = coeff_header_uri or plan_header_uri or charts_header_uri or export_header_uri
if top_header_uri:
    top_header_html = """
        <style>
        .app-top-hero {
            position: relative;
            border: 1px solid #1D2A44;
            border-radius: 14px;
            overflow: hidden;
            margin: 0.15rem 0 0.85rem 0;
            min-height: 250px;
            background: #0B1220;
        }
        .app-top-hero-bg {
            position: absolute;
            inset: 0;
            background-image: url("__TOP_HEADER_URI__");
            background-size: contain;
            background-repeat: no-repeat;
            background-position: calc(100% - 20px) center;
            opacity: 0.0;
            transform: scale(1.0);
            filter: none;
            transition: opacity 220ms ease;
        }
        .app-top-hero-fade {
            position: absolute;
            inset: 0;
            background: linear-gradient(90deg, rgba(8, 16, 30, 0.90) 0%, rgba(8, 16, 30, 0.48) 42%, rgba(8, 16, 30, 0.02) 72%, rgba(8, 16, 30, 0.00) 100%);
        }
        .app-top-hero-content {
            position: relative;
            z-index: 2;
            padding: 14px 16px 14px 16px;
        }
        .app-top-hero-content .intro-card {
            margin: 0.2rem 0 0.2rem 0;
            padding: 12px 14px;
            width: fit-content;
            max-width: calc(100% - 360px);
            border-radius: 12px;
            border: 1px solid #3D8EF0;
            border-left: 4px solid #00CDC5;
            background: linear-gradient(180deg, rgba(0, 102, 224, 0.20) 0%, rgba(17, 26, 46, 0.84) 100%);
            box-shadow: 0 8px 18px rgba(0, 0, 0, 0.20);
        }
        @media (max-width: 1150px) {
            .app-top-hero-content .intro-card {
                width: 100%;
            }
            .app-top-hero-bg {
                background-position: calc(100% - 12px) center;
            }
        }
        .app-top-hero-content .intro-card p {
            margin: 0 0 6px 0;
            color: #EAF0FF;
            line-height: 1.38;
            font-size: clamp(0.93rem, 0.92vw, 1.06rem);
            overflow-wrap: anywhere;
        }
        .app-top-hero-content .intro-card p.one-line {
            white-space: normal;
        }
        .app-top-hero-content .intro-card p:last-child {
            margin-bottom: 0;
        }
        @media (max-width: 1150px) {
            .app-top-hero-content .intro-card p.one-line {
                white-space: normal;
            }
        }
        </style>
        <div class="app-top-hero">
            <div id="app-top-hero-bg" class="app-top-hero-bg"></div>
            <div class="app-top-hero-fade"></div>
            <div class="app-top-hero-content">
                <h1 style="font-weight: 700; letter-spacing: 0.02em; margin: 0 0 6px 0; max-width: 70%;">
                    РњРµРґРёР°РїР»Р°РЅРµСЂ <span style="color: #00CDC5; font-size: 1.25em;">E-promo</span>
                </h1>
                <div class="intro-card">
                    <p><span style="font-weight: 800; color: #9EC5FF;">Р§С‚Рѕ СЌС‚Рѕ:</span>
                    РёРЅСЃС‚СЂСѓРјРµРЅС‚ РґР»СЏ СЂР°СЃС‡РµС‚Р° РјРµРґРёР°РїР»Р°РЅР° РЅР° РІС‹Р±СЂР°РЅРЅС‹Р№ РїРµСЂРёРѕРґ (РѕС‚ 1 РґРѕ 12 РјРµСЃСЏС†РµРІ) РїРѕ С‚РёРїР°Рј СЂРµРєР»Р°РјРЅС‹С… РєР°РјРїР°РЅРёР№.</p>
                    <p class="one-line"><span style="font-weight: 800; color: #9EC5FF;">Р—Р°С‡РµРј РЅСѓР¶РµРЅ:</span>
                    С‡С‚РѕР±С‹ СѓРїСЂРѕСЃС‚РёС‚СЊ Р°Р»РіРѕСЂРёС‚Рј СЂР°СЃС‡РµС‚Р° РјРµРґРёР°РїР»Р°РЅР°, СЃРѕРєСЂР°С‚РёС‚СЊ РІСЂРµРјСЏ РЅР° РїРµСЂРІРёС‡РЅС‹Р№ СЂР°СЃС‡РµС‚ Рё СѓСЃРєРѕСЂРёС‚СЊ РІРЅРµСЃРµРЅРёРµ РїРѕСЃР»РµРґСѓСЋС‰РёС… РїСЂР°РІРѕРє.</p>
                </div>
            </div>
        </div>
        """
    st.markdown(
        top_header_html.replace("__TOP_HEADER_URI__", top_header_uri),
        unsafe_allow_html=True,
    )
else:
    st.markdown(
        """
        <h1 style="font-weight: 700; letter-spacing: 0.02em;">
            РњРµРґРёР°РїР»Р°РЅРµСЂ <span style="color: #00CDC5; font-size: 1.25em;">E-promo</span>
        </h1>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        """
        <div style="
            margin: 0.2rem 0 0.8rem 0;
            padding: 12px 14px;
            border-radius: 12px;
            border: 1px solid #3D8EF0;
            border-left: 4px solid #00CDC5;
            background: linear-gradient(180deg, rgba(0, 102, 224, 0.18) 0%, rgba(17, 26, 46, 0.82) 100%);
            box-shadow: 0 8px 18px rgba(0, 0, 0, 0.20);
        ">
            <p style="margin: 0 0 6px 0; color: #EAF0FF; line-height: 1.45;">
                <span style="font-weight: 800; color: #9EC5FF;">РљР°Рє СЌС‚Рѕ СЂР°Р±РѕС‚Р°РµС‚:</span>
                Р”РѕР±Р°РІСЊС‚Рµ РѕРґРёРЅ РёР»Рё РЅРµСЃРєРѕР»СЊРєРѕ РєРѕСЌС„С„РёС†РёРµРЅС‚РѕРІ РїРѕ РІС‹Р±СЂР°РЅРЅС‹Рј РјРµСЃСЏС†Р°Рј (РѕС‚ 1 РґРѕ 12 РјРµСЃСЏС†РµРІ) РЅР° РѕСЃРЅРѕРІРµ РґРѕСЃС‚СѓРїРЅС‹С… СЃС†РµРЅР°СЂРёРµРІ.
            </p>
            <p style="margin: 0; color: #EAF0FF; line-height: 1.45;">
                <span style="font-weight: 800; color: #9EC5FF;">Р—Р°С‡РµРј РЅСѓР¶РЅРѕ:</span>
                Р§С‚РѕР±С‹ РіРёР±РєРѕ СѓС‡РёС‚С‹РІР°С‚СЊ СЃРµР·РѕРЅРЅС‹Рµ РєРѕР»РµР±Р°РЅРёСЏ, РёР·РјРµРЅРµРЅРёСЏ СЃРїСЂРѕСЃР° РїРѕ РІС‹Р±СЂР°РЅРЅС‹Рј РјРµСЃСЏС†Р°Рј Рё РІР»РёСЏРЅРёРµ РјРµРґРёР№РЅС‹С… С…РІРѕСЃС‚РѕРІ.
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )

ui_compact_tables = True
ui_editor_row_height = 28
ui_table_row_px = 29
ui_table_font_px = "12px"

st.markdown(
    f"""
    <style>
    [data-testid="stDataFrame"] [role="gridcell"],
    [data-testid="stDataFrame"] [role="columnheader"] {{
        font-size: {ui_table_font_px};
        line-height: 1.0 !important;
        padding-top: 2px !important;
        padding-bottom: 2px !important;
        min-height: {ui_table_row_px}px !important;
        height: {ui_table_row_px}px !important;
        max-height: {ui_table_row_px}px !important;
        box-sizing: border-box !important;
    }}
    </style>
    """,
    unsafe_allow_html=True,
)

tab_coeffs, tab_plan, tab_charts, tab_export = st.tabs(
    ["РљРѕСЌС„С„РёС†РёРµРЅС‚С‹", "РњРµРґРёР°РїР»Р°РЅ", "Р”РёР°РіСЂР°РјРјС‹", "Export/Import"]
)
if top_header_uri:
    header_tab_bg_map_json = json.dumps(
        {
            0: coeff_header_uri or "",
            1: plan_header_uri or "",
            2: charts_header_uri or "",
            3: export_header_uri or "",
        },
        ensure_ascii=False,
    )
    components.html(
        """
        <script>
        (function() {
            const TAB_BG_MAP = __TAB_BG_MAP__;
            const VISIBLE_OPACITY = "1.00";
            const COEFF_TAB_INDEX = 0;
            const PLAN_TAB_INDEX = 1;
            const CHARTS_TAB_INDEX = 2;
            const EXPORT_TAB_INDEX = 3;
            const BASE_BG_POSITION = "calc(100% - 20px) center";
            const COEFF_BG_POSITION = "calc(100% + 95px) calc(50% + 77px)";
            const PLAN_BG_POSITION = "calc(100% + 120px) center";
            const CHARTS_BG_SIZE = "47% auto";
            const CHARTS_BG_POSITION = "calc(100% + 120px) calc(50% + 30px)";
            const EXPORT_BG_SIZE = "24.8% auto";
            const EXPORT_BG_POSITION = "calc(100% - 20px) calc(50% + 30px)";

            const getTopTablist = (doc) => {
                const tablists = Array.from(doc.querySelectorAll('[role="tablist"]'));
                if (!tablists.length) return null;
                return tablists
                    .map(tl => ({ tl, top: tl.getBoundingClientRect().top }))
                    .sort((a, b) => a.top - b.top)[0]?.tl || null;
            };

            const applyHeaderImage = () => {
                const doc = window.parent.document;
                const bg = doc.getElementById("app-top-hero-bg");
                if (!bg) return false;
                const topTablist = getTopTablist(doc);
                if (!topTablist) return false;
                const tabs = topTablist.querySelectorAll('button[role="tab"]');
                if (!tabs || !tabs.length) return false;

                let activeIdx = 0;
                for (let i = 0; i < tabs.length; i++) {
                    if (tabs[i].getAttribute("aria-selected") === "true") {
                        activeIdx = i;
                        break;
                    }
                }
                const key = String(activeIdx);
                const imgUrl = TAB_BG_MAP[key] || "";
                bg.style.backgroundImage = imgUrl ? `url(${imgUrl})` : "";
                if (activeIdx === COEFF_TAB_INDEX) {
                    bg.style.backgroundSize = "52% auto";
                    bg.style.backgroundPosition = COEFF_BG_POSITION;
                } else if (activeIdx === PLAN_TAB_INDEX) {
                    bg.style.backgroundSize = "58% auto";
                    bg.style.backgroundPosition = PLAN_BG_POSITION;
                } else if (activeIdx === CHARTS_TAB_INDEX) {
                    bg.style.backgroundSize = CHARTS_BG_SIZE;
                    bg.style.backgroundPosition = CHARTS_BG_POSITION;
                } else if (activeIdx === EXPORT_TAB_INDEX) {
                    bg.style.backgroundSize = EXPORT_BG_SIZE;
                    bg.style.backgroundPosition = EXPORT_BG_POSITION;
                } else {
                    bg.style.backgroundSize = "58% auto";
                    bg.style.backgroundPosition = BASE_BG_POSITION;
                }
                bg.style.opacity = imgUrl ? VISIBLE_OPACITY : "0.0";
                return true;
            };

            let attempts = 0;
            const init = () => {
                applyHeaderImage();
                attempts += 1;
                if (attempts < 30) setTimeout(init, 120);
            };
            init();

            const topTablist = getTopTablist(window.parent.document);
            if (topTablist) {
                const obs = new MutationObserver(() => applyHeaderImage());
                obs.observe(topTablist, {
                    subtree: true,
                    attributes: true,
                    attributeFilter: ["aria-selected", "class"],
                });
            }
        })();
        </script>
        """.replace("__TAB_BG_MAP__", header_tab_bg_map_json),
        height=0,
        width=0,
    )

# Р‘С‹СЃС‚СЂС‹Р№ РёРјРїРѕСЂС‚ РґРѕСЃС‚СѓРїРµРЅ РІСЃРµРіРґР° (РґР°Р¶Рµ РґРѕ РїРµСЂРІРѕРіРѕ СЂР°СЃС‡РµС‚Р° РЅР° РІРєР»Р°РґРєРµ "РњРµРґРёР°РїР»Р°РЅ").
with st.sidebar:
    st.markdown("### Р‘С‹СЃС‚СЂС‹Р№ РёРјРїРѕСЂС‚ РїСЂРѕРµРєС‚Р°")
    st.caption("Р—Р°РіСЂСѓР·РёС‚Рµ JSON РїСЂРѕРµРєС‚Р° СЃСЂР°Р·Сѓ РїРѕСЃР»Рµ Р·Р°РїСѓСЃРєР° РїСЂРёР»РѕР¶РµРЅРёСЏ.")
    uploaded_project_quick = st.file_uploader(
        "РРјРїРѕСЂС‚ РїСЂРѕРµРєС‚Р° (JSON)",
        type=["json"],
        key="upload_project_json_quick",
    )
    queue_project_import_from_upload(uploaded_project_quick, "quick_sidebar")
# ====================================================================
#                        РўРђР‘ В«РљРћР­Р¤Р¤РР¦РР•РќРўР«В»
# ====================================================================

with tab_coeffs:
    st.markdown(
        """
        <div class="tab-intro">
            <p>1) РЎРѕР·РґР°Р№С‚Рµ РЅР°Р±РѕСЂ РєРѕСЌС„С„РёС†РёРµРЅС‚РѕРІ Рё РІС‹Р±РµСЂРёС‚Рµ С‚РёРї: <b>РЎРїСЂРѕСЃ</b>, <b>AOV</b>, <b>РљР°СЃС‚РѕРјРЅС‹Р№ РЅР°Р±РѕСЂ</b> РёР»Рё <b>РњРµРґРёР№РЅС‹Рµ С…РІРѕСЃС‚С‹</b>.</p>
            <p>2) РЈРєР°Р¶РёС‚Рµ СЃС‚Р°СЂС‚РѕРІС‹Р№ РјРµСЃСЏС† Рё РіРѕРґ, Р·Р°С‚РµРј Р·Р°РїРѕР»РЅРёС‚Рµ Р·РЅР°С‡РµРЅРёСЏ РїРѕ РїРµСЂРёРѕРґР°Рј РІ С‚Р°Р±Р»РёС†Рµ РїРѕ РґР°РЅРЅС‹Рј РёР· <a href="https://wordstat.yandex.ru/" target="_blank">Р’РѕСЂРґСЃС‚Р°С‚</a>.</p>
            <p>3) РќР°Р¶РјРёС‚Рµ РєРЅРѕРїРєСѓ РїСЂРёРјРµРЅРµРЅРёСЏ/СЂР°СЃС‡РµС‚Р°, С‡С‚РѕР±С‹ СЃРѕС…СЂР°РЅРёС‚СЊ РЅР°Р±РѕСЂ Рё РїРѕР»СѓС‡РёС‚СЊ РёС‚РѕРіРѕРІС‹Рµ РєРѕСЌС„С„РёС†РёРµРЅС‚С‹.</p>
            <p>4) РџРѕРІС‚РѕСЂРёС‚Рµ РґР»СЏ РІСЃРµС… РЅР°Р±РѕСЂРѕРІ, РєРѕС‚РѕСЂС‹Рµ РїР»Р°РЅРёСЂСѓРµС‚Рµ РёСЃРїРѕР»СЊР·РѕРІР°С‚СЊ РІ В«РњРµРґРёР°РїР»Р°РЅРµВ».</p>
            <p>5) РџСЂРѕРІРµСЂСЊС‚Рµ, С‡С‚Рѕ Сѓ РєР°Р¶РґРѕРіРѕ РЅР°Р±РѕСЂР° Р·Р°РїРѕР»РЅРµРЅС‹ РґР°РЅРЅС‹Рµ, РёРЅР°С‡Рµ РѕРЅ РЅРµ Р±СѓРґРµС‚ РІР»РёСЏС‚СЊ РЅР° СЂР°СЃС‡РµС‚.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    ui_section_title("РљРѕСЌС„С„РёС†РёРµРЅС‚С‹ СЃРµР·РѕРЅРЅРѕСЃС‚Рё")
    st.caption(
        "РЎРѕР·РґР°РІР°Р№С‚Рµ РѕРґРёРЅ РёР»Рё РЅРµСЃРєРѕР»СЊРєРѕ РЅР°Р±РѕСЂРѕРІ РєРѕСЌС„С„РёС†РёРµРЅС‚РѕРІ СЃРµР·РѕРЅРЅРѕСЃС‚Рё: "
        "РїРѕ СЃРїСЂРѕСЃСѓ (Р·Р°РїСЂРѕСЃС‹) РёР»Рё РїРѕ AOV (СЃСЂРµРґРЅРёР№ С‡РµРє). Р”Р»СЏ РєР°Р¶РґРѕРіРѕ РЅР°Р±РѕСЂР° СЃС‡РёС‚Р°Р№С‚Рµ РёРЅРґРµРєСЃС‹ РїРѕ РјРµСЃСЏС†Р°Рј."
    )
    st.markdown(
        """
        <div style="
            margin: 8px 0 14px 0;
            display: inline-block;
            width: fit-content;
            max-width: 100%;
            white-space: nowrap;
            padding: 10px 12px;
            border-radius: 10px;
            border: 1px solid #FF8A66;
            background: rgba(255, 99, 51, 0.14);
            color: #FFD9CC;
            font-weight: 600;
        ">
            Р’Р°Р¶РЅРѕ: РґР»СЏ РєРѕСЂСЂРµРєС‚РЅРѕРіРѕ СЂР°СЃС‡РµС‚Р° РјРµРґРёР°РїР»Р°РЅР° СЃРЅР°С‡Р°Р»Р° РѕР±СЏР·Р°С‚РµР»СЊРЅРѕ СЂР°СЃСЃС‡РёС‚Р°Р№С‚Рµ РјРёРЅРёРјСѓРј РґРІР° РЅР°Р±РѕСЂР°:
            В«РЎРµР·РѕРЅРЅРѕСЃС‚СЊВ» (СЃРїСЂРѕСЃ) Рё В«AOVВ» (СЃСЂРµРґРЅРёР№ С‡РµРє).
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.caption(
        "РСЃС‚РѕС‡РЅРёРєРё РґР°РЅРЅС‹С…: РґР»СЏ В«РЎРїСЂРѕСЃВ» вЂ” Wordstat; РґР»СЏ В«AOVВ» вЂ” С„Р°РєС‚РёС‡РµСЃРєРёР№ СЃСЂРµРґРЅРёР№ С‡РµРє РёР· Р°РЅР°Р»РёС‚РёРєРё; "
        "РґР»СЏ В«РњРµРґРёР№РЅС‹С… С…РІРѕСЃС‚РѕРІВ» вЂ” РѕР±СЂР°С‚РёС‚РµСЃСЊ Рє СЃРїРµС†РёР°Р»РёСЃС‚Р°Рј РїРѕ РјРµРґРёР№РЅРѕР№ СЂРµРєР»Р°РјРµ РґР»СЏ СЂР°СЃС‡РµС‚Р° РІР»РёСЏРЅРёСЏ."
    )

    # ---------------- Р’РЎРџРћРњРћР“РђРўР•Р›Р¬РќР«Р• Р¤РЈРќРљР¦ ----------------

    def generate_months_list(start_month: int, start_year: int, count: int = 24):
        month_names = [
            "РЇРЅРІР°СЂСЊ", "Р¤РµРІСЂР°Р»СЊ", "РњР°СЂС‚", "РђРїСЂРµР»СЊ", "РњР°Р№", "РСЋРЅСЊ",
            "РСЋР»СЊ", "РђРІРіСѓСЃС‚", "РЎРµРЅС‚СЏР±СЂСЊ", "РћРєС‚СЏР±СЂСЊ", "РќРѕСЏР±СЂСЊ", "Р”РµРєР°Р±СЂСЊ"
        ]
        months = []
        current_month = start_month
        current_year = start_year
        for _ in range(count):
            months.append(
                {
                    "period": f"{month_names[current_month - 1]} {current_year}",
                    "month_num": current_month,
                    "year": current_year,
                }
            )
            current_month += 1
            if current_month > 12:
                current_month = 1
                current_year += 1
        return months

    def calculate_seasonality_coeffs_demand(df_data: pd.DataFrame, query_cols: list[str]) -> pd.DataFrame:
        """
        РЎРїСЂРѕСЃ (РїРѕ Р·Р°РїСЂРѕСЃР°Рј), Р»РѕРіРёРєР° РєР°Рє РІ Excel-С„Р°Р№Р»Рµ:
        1) Р”Р»СЏ РєР°Р¶РґРѕРіРѕ РєР°Р»РµРЅРґР°СЂРЅРѕРіРѕ РјРµСЃСЏС†Р° СЃСѓРјРјРёСЂСѓРµРј РїРѕ РєР°Р¶РґРѕРјСѓ Р·Р°РїСЂРѕСЃСѓ (РїРѕ РІСЃРµРј РіРѕРґР°Рј).
        2) Р’РЅСѓС‚СЂРё РјРµСЃСЏС†Р° СЃС‡РёС‚Р°РµРј РІРµСЃР° Р·Р°РїСЂРѕСЃРѕРІ РєР°Рє РґРѕР»Рё (СЃСѓРјРјР° РїРѕ Р·Р°РїСЂРѕСЃСѓ / СЃСѓРјРјР° РїРѕ РІСЃРµРј Р·Р°РїСЂРѕСЃР°Рј РјРµСЃСЏС†Р°).
        3) РЎСЂРµРґ. РІР·РІ. РјРµСЃСЏС†Р° = СЃСѓРјРјР°(СЃСѓРјРјР°_РїРѕ_Р·Р°РїСЂРѕСЃСѓ_РјРµСЃСЏС†Р° * РІРµСЃ_Р·Р°РїСЂРѕСЃР°_РІ_РјРµСЃСЏС†Рµ).
        4) РЎСЂРµРґРЅРµРµ РїРѕ 12 РјРµСЃСЏС†Р°Рј = РЎР Р—РќРђР§(РЎСЂРµРґ. РІР·РІ. РїРѕ РјРµСЃСЏС†Р°Рј).
        5) РљРѕСЌС„. РјРµСЃСЏС†Р° = РЎСЂРµРґ. РІР·РІ. РјРµСЃСЏС†Р° / РЎСЂРµРґРЅРµРµ РїРѕ 12 РјРµСЃСЏС†Р°Рј.
        """
        month_names_map = {
            1: "РЇРЅРІР°СЂСЊ", 2: "Р¤РµРІСЂР°Р»СЊ", 3: "РњР°СЂС‚", 4: "РђРїСЂРµР»СЊ",
            5: "РњР°Р№", 6: "РСЋРЅСЊ", 7: "РСЋР»СЊ", 8: "РђРІРіСѓСЃС‚",
            9: "РЎРµРЅС‚СЏР±СЂСЊ", 10: "РћРєС‚СЏР±СЂСЊ", 11: "РќРѕСЏР±СЂСЊ", 12: "Р”РµРєР°Р±СЂСЊ",
        }

        df = df_data.copy()

        avg_weighted_by_month: dict[int, float] = {}

        for m in range(1, 13):
            df_m = df[df["month_num"] == m]
            if df_m.empty:
                avg_weighted_by_month[m] = 0.0
                continue

            sums_per_query = df_m[query_cols].sum(axis=0)
            total_sum_month = float(sums_per_query.sum())
            if total_sum_month <= 0:
                avg_weighted_by_month[m] = 0.0
                continue

            weights = sums_per_query / total_sum_month
            avg_weighted = float((sums_per_query * weights).sum())
            avg_weighted_by_month[m] = avg_weighted

        values = list(avg_weighted_by_month.values())
        avg_all_months = float(np.mean(values)) if values else 1.0
        if avg_all_months == 0:
            avg_all_months = 1.0

        rows = []
        for m in range(1, 13):
            avg_w = avg_weighted_by_month.get(m, 0.0)
            coeff = avg_w / avg_all_months if avg_all_months > 0 else 1.0
            rows.append(
                {
                    "РќРѕРјРµСЂ РјРµСЃСЏС†Р°": m,
                    "РњРµСЃСЏС†": month_names_map[m],
                    "РЎСЂРµРґ. РІР·РІ.": round(avg_w, 2),
                    "РљРѕСЌС„.": round(coeff, 2),
                }
            )

        return pd.DataFrame(rows)

    def calculate_seasonality_coeffs_aov(df_data: pd.DataFrame) -> pd.DataFrame:
        """
        AOV (СЃСЂРµРґРЅРёР№ С‡РµРє) РїРѕ РјРµСЃСЏС†Р°Рј:
        1) Р”Р»СЏ РєР°Р¶РґРѕРіРѕ РєР°Р»РµРЅРґР°СЂРЅРѕРіРѕ РјРµСЃСЏС†Р° СЃС‡РёС‚Р°РµРј СЃСЂРµРґРЅРёР№ AOV РїРѕ СЃС‚СЂРѕРєР°Рј (РїРѕ РіРѕРґР°Рј).
        2) РЎС‡РёС‚Р°РµРј СЃСЂРµРґРЅРёР№ AOV РїРѕ 12 РјРµСЃСЏС†Р°Рј.
        3) РљРѕСЌС„. AOV РјРµСЃСЏС†Р° = РЎСЂРµРґ. AOV РјРµСЃСЏС†Р° / СЃСЂРµРґРЅРёР№ AOV РїРѕ 12 РјРµСЃСЏС†Р°Рј.
        """
        month_names_map = {
            1: "РЇРЅРІР°СЂСЊ", 2: "Р¤РµРІСЂР°Р»СЊ", 3: "РњР°СЂС‚", 4: "РђРїСЂРµР»СЊ",
            5: "РњР°Р№", 6: "РСЋРЅСЊ", 7: "РСЋР»СЊ", 8: "РђРІРіСѓСЃС‚",
            9: "РЎРµРЅС‚СЏР±СЂСЊ", 10: "РћРєС‚СЏР±СЂСЊ", 11: "РќРѕСЏР±СЂСЊ", 12: "Р”РµРєР°Р±СЂСЊ",
        }

        df = df_data.copy()

        # СЃСЂРµРґРЅРёР№ AOV РїРѕ РєР°Р¶РґРѕРјСѓ РєР°Р»РµРЅРґР°СЂРЅРѕРјСѓ РјРµСЃСЏС†Сѓ
        if "AOV" not in df.columns:
            df["AOV"] = 0.0
        avg_aov_by_month = df.groupby("month_num")["AOV"].mean()

        avg_all = float(avg_aov_by_month.mean()) if not avg_aov_by_month.empty else 1.0
        if avg_all == 0:
            avg_all = 1.0

        rows = []
        for m in range(1, 13):
            avg_m = float(avg_aov_by_month.get(m, 0.0))
            coeff = avg_m / avg_all if avg_all > 0 else 1.0
            rows.append(
                {
                    "РќРѕРјРµСЂ РјРµСЃСЏС†Р°": m,
                    "РњРµСЃСЏС†": month_names_map[m],
                    "РЎСЂРµРґ. AOV": round(avg_m, 2),
                    "РљРѕСЌС„. AOV": round(coeff, 2),
                }
            )

        return pd.DataFrame(rows)

    # ---------------- РќР¦РђР›Р—РђР¦РЇ РЎРћРЎРўРћРЇРќРЇ ----------------

    if "coeff_sets" not in st.session_state:
        st.session_state["coeff_sets"] = []
    if "coeff_active_set_id" not in st.session_state:
        st.session_state["coeff_active_set_id"] = None

    def _set_active_coeff_set(set_id: int) -> None:
        st.session_state["coeff_active_set_id"] = int(set_id)


    # ---------------- Р”РћР‘РђР’Р›Р•РќРР• РќРћР’РћР“Рћ РќРђР‘РћР Рђ ----------------

    col_add_set, col_help = st.columns([2, 3])
    with col_add_set:
        if st.button("Р”РѕР±Р°РІРёС‚СЊ РЅРѕРІС‹Р№ РЅР°Р±РѕСЂ РєРѕСЌС„С„РёС†РёРµРЅС‚РѕРІ", key="add_coeff_set"):
            new_id = len(st.session_state["coeff_sets"]) + 1
            st.session_state["coeff_sets"].append(
                {
                    "id": new_id,
                    "name": f"РќР°Р±РѕСЂ {new_id}",
                    "type": "РЎРїСЂРѕСЃ (РїРѕ Р·Р°РїСЂРѕСЃР°Рј)",  # РёР»Рё "AOV (СЃСЂРµРґРЅРёР№ С‡РµРє)"
                    "start_month": 1,
                    "start_year": 2024,
                    "period_months": 24,
                    "queries": [],
                    "df_data": None,
                    "result": None,
                }
            )
            _set_active_coeff_set(new_id)
            st.rerun()
    with col_help:
        st.empty()

    # ---------------- РЎРџРРЎРћРљ РќРђР‘РћР РћР’ ----------------

    if not st.session_state["coeff_sets"]:
        st.session_state["coeff_active_set_id"] = None
        st.info("РџРѕРєР° РЅРµС‚ РЅРё РѕРґРЅРѕРіРѕ РЅР°Р±РѕСЂР° РєРѕСЌС„С„РёС†РёРµРЅС‚РѕРІ. РќР°Р¶РјРёС‚Рµ В«Р”РѕР±Р°РІРёС‚СЊ РЅРѕРІС‹Р№ РЅР°Р±РѕСЂ РєРѕСЌС„С„РёС†РёРµРЅС‚РѕРІВ».")
    else:
        existing_coeff_set_ids = [int(cs["id"]) for cs in st.session_state["coeff_sets"]]
        if st.session_state.get("coeff_active_set_id") not in existing_coeff_set_ids:
            st.session_state["coeff_active_set_id"] = existing_coeff_set_ids[-1]
        for idx, cs in enumerate(st.session_state["coeff_sets"]):
            set_id = cs["id"]
            with st.expander(f"РќР°Р±РѕСЂ {set_id}: {cs['name']}", expanded=st.session_state.get("coeff_active_set_id") == set_id):

                # ---- 1-2. РќР°Р·РІР°РЅРёРµ Рё С‚РёРї РЅР°Р±РѕСЂР° (РєРѕРјРїР°РєС‚РЅС‹Р№ layout) ----
                meta_c1, meta_c2, _meta_spacer = st.columns([2.2, 2.2, 1.6], vertical_alignment="bottom")
                with meta_c1:
                    cs["name"] = st.text_input(
                        "РќР°Р·РІР°РЅРёРµ РЅР°Р±РѕСЂР°",
                        value=cs["name"],
                        key=f"cs_name_{set_id}",
                        on_change=_set_active_coeff_set,
                        args=(set_id,),
                    )
                with meta_c2:
                    cs["type"] = normalize_coeff_set_type(cs.get("type"))
                    cs["type"] = st.selectbox(
                        "РўРёРї РЅР°Р±РѕСЂР°:",
                        options=["РЎРїСЂРѕСЃ (РїРѕ Р·Р°РїСЂРѕСЃР°Рј)", "AOV (СЃСЂРµРґРЅРёР№ С‡РµРє)", "РљР°СЃС‚РѕРјРЅС‹Р№ РЅР°Р±РѕСЂ", "РњРµРґРёР№РЅС‹Рµ С…РІРѕСЃС‚С‹"],
                        help=(
                            "вЂў РЎРїСЂРѕСЃ (РїРѕ Р·Р°РїСЂРѕСЃР°Рј)\n"
                            "  РђРІС‚РѕРјР°С‚РёС‡РµСЃРєРёРµ РєРѕСЌС„С„РёС†РёРµРЅС‚С‹ РЅР° РѕСЃРЅРѕРІРµ Wordstat.\n\n"
                            "вЂў AOV (СЃСЂРµРґРЅРёР№ С‡РµРє)\n"
                            "  РђРІС‚РѕРјР°С‚РёС‡РµСЃРєРёРµ РєРѕСЌС„С„РёС†РёРµРЅС‚С‹ СЃСЂРµРґРЅРµРіРѕ С‡РµРєР°.\n\n"
                            "вЂў РљР°СЃС‚РѕРјРЅС‹Р№ РЅР°Р±РѕСЂ\n"
                            "  Р СѓС‡РЅРѕР№ РІРІРѕРґ РєРѕСЌС„С„РёС†РёРµРЅС‚РѕРІ РЅР° 12 РјРµСЃСЏС†РµРІ.\n\n"
                            "вЂў РњРµРґРёР№РЅС‹Рµ С…РІРѕСЃС‚С‹\n"
                            "  РљРѕСЌС„С„РёС†РёРµРЅС‚С‹ РїРѕРєР°Р·РѕРІ РІ РјРµСЃСЏС†Р°С…, РЅР°РєР»Р°РґС‹РІР°СЋС‚СЃСЏ РїРѕРІРµСЂС… РєРѕСЌС„С„РёС†РёРµРЅС‚Р° СЃРїСЂРѕСЃР°."
                        ),
                        index=(
                            0 if cs.get("type") == "РЎРїСЂРѕСЃ (РїРѕ Р·Р°РїСЂРѕСЃР°Рј)"
                            else 1 if cs.get("type") == "AOV (СЃСЂРµРґРЅРёР№ С‡РµРє)"
                            else 2 if cs.get("type") == "РљР°СЃС‚РѕРјРЅС‹Р№ РЅР°Р±РѕСЂ"
                            else 3
                        ),
                        key=f"cs_type_{set_id}",
                        on_change=_set_active_coeff_set,
                        args=(set_id,),
                    )

                # ---- 3. РџРµСЂРёРѕРґ 24 РјРµСЃСЏС†Р° (РєРѕРјРїР°РєС‚РЅС‹Р№ layout) ----
                col_m, col_y, _period_spacer = st.columns([1.8, 1.8, 2.4], vertical_alignment="bottom")
                with col_m:
                    cs["start_month"] = st.selectbox(
                        "РќР°С‡Р°Р»СЊРЅС‹Р№ РјРµСЃСЏС† РїРµСЂРёРѕРґР°:",
                        options=list(range(1, 13)),
                        format_func=lambda x: [
                            "РЇРЅРІР°СЂСЊ", "Р¤РµРІСЂР°Р»СЊ", "РњР°СЂС‚", "РђРїСЂРµР»СЊ", "РњР°Р№", "РСЋРЅСЊ",
                            "РСЋР»СЊ", "РђРІРіСѓСЃС‚", "РЎРµРЅС‚СЏР±СЂСЊ", "РћРєС‚СЏР±СЂСЊ", "РќРѕСЏР±СЂСЊ", "Р”РµРєР°Р±СЂСЊ",
                        ][x - 1],
                        index=cs["start_month"] - 1,
                        key=f"cs_month_{set_id}",
                        on_change=_set_active_coeff_set,
                        args=(set_id,),
                    )
                with col_y:
                    cs["start_year"] = st.number_input(
                        "РќР°С‡Р°Р»СЊРЅС‹Р№ РіРѕРґ РїРµСЂРёРѕРґР°:",
                        min_value=2020,
                        max_value=2030,
                        value=cs["start_year"],
                        step=1,
                        key=f"cs_year_{set_id}",
                        on_change=_set_active_coeff_set,
                        args=(set_id,),
                    )
                if "period_months" not in cs:
                    cs["period_months"] = 24
                cs["period_months"] = int(cs.get("period_months", 24) or 24)
                period_len_col, _period_len_spacer = st.columns([1.8, 4.2], vertical_alignment="bottom")
                with period_len_col:
                    cs["period_months"] = int(
                        st.number_input(
                            "Р”Р»РёС‚РµР»СЊРЅРѕСЃС‚СЊ РїРµСЂРёРѕРґР°, РјРµСЃ.",
                            min_value=24,
                            max_value=240,
                            value=int(cs["period_months"]),
                            step=12,
                            key=f"cs_period_months_{set_id}",
                            on_change=_set_active_coeff_set,
                            args=(set_id,),
                            help="РџРѕ СѓРјРѕР»С‡Р°РЅРёСЋ 24 РјРµСЃСЏС†Р°. РњРѕР¶РЅРѕ СѓРІРµР»РёС‡РёС‚СЊ РїРµСЂРёРѕРґ РґРѕ 36, 60, 84 РјРµСЃСЏС†РµРІ Рё Р±РѕР»СЊС€Рµ.",
                        )
                    )
                period_years = cs["period_months"] / 12.0
                st.caption(
                    f"РўРµРєСѓС‰РёР№ РїРµСЂРёРѕРґ: {int(cs['period_months'])} РјРµСЃ. ({period_years:.0f} Рі.)"
                    if float(period_years).is_integer()
                    else f"РўРµРєСѓС‰РёР№ РїРµСЂРёРѕРґ: {int(cs['period_months'])} РјРµСЃ. ({period_years:.1f} Рі.)"
                )

                # ---- 4. РЅС‚РµСЂС„РµР№СЃ РІ Р·Р°РІРёСЃРёРјРѕСЃС‚Рё РѕС‚ С‚РёРїР° РЅР°Р±РѕСЂР° ----

                if cs["type"] == "РЎРїСЂРѕСЃ (РїРѕ Р·Р°РїСЂРѕСЃР°Рј)":
                    # ===== Р Р•Р–Рњ РЎРџР РћРЎ (РџРћ Р—РђРџР РћРЎРђРњ) =====
                    st.markdown("**РџРѕРёСЃРєРѕРІС‹Рµ Р·Р°РїСЂРѕСЃС‹:**")

                    if not cs.get("queries"):
                        cs["queries"] = ["Р—Р°РїСЂРѕСЃ 1"]

                    to_remove = []
                    for q_idx, q in enumerate(cs["queries"]):
                        query_row, _query_spacer = st.columns([1.6, 4.4])
                        with query_row:
                            col_q, col_del = st.columns([5, 1])
                            with col_q:
                                new_q = st.text_input(
                                    f"Р—Р°РїСЂРѕСЃ {q_idx + 1}:",
                                    value=q,
                                    key=f"cs_q_{set_id}_{q_idx}",
                                    label_visibility="collapsed",
                                    on_change=_set_active_coeff_set,
                                    args=(set_id,),
                                )
                                cs["queries"][q_idx] = new_q
                            with col_del:
                                if st.button("рџ—‘пёЏ", key=f"cs_del_{set_id}_{q_idx}"):
                                    _set_active_coeff_set(set_id)
                                    to_remove.append(q_idx)

                    for r in sorted(to_remove, reverse=True):
                        cs["queries"].pop(r)
                        st.rerun()

                    if st.button("вћ• Р”РѕР±Р°РІРёС‚СЊ Р·Р°РїСЂРѕСЃ", key=f"cs_add_q_{set_id}"):
                        _set_active_coeff_set(set_id)
                        cs["queries"].append("")
                        st.rerun()

                    # Р“РµРЅРµСЂР°С†РёСЏ С‚Р°Р±Р»РёС†С‹ РЅР° РІС‹Р±СЂР°РЅРЅС‹Р№ РїРµСЂРёРѕРґ
                    if st.button("РџСЂРёРјРµРЅРёС‚СЊ РЅР°СЃС‚СЂРѕР№РєРё РЅР°Р±РѕСЂР°", key=f"cs_apply_{set_id}", type="primary"):
                        _set_active_coeff_set(set_id)
                        queries_clean = [q.strip() for q in cs["queries"] if q.strip()]
                        if not queries_clean:
                            st.error("Р”РѕР±Р°РІСЊС‚Рµ С…РѕС‚СЏ Р±С‹ РѕРґРёРЅ РїРѕРёСЃРєРѕРІС‹Р№ Р·Р°РїСЂРѕСЃ.")
                        else:
                            period_months = int(cs.get("period_months", 24) or 24)
                            months_period = generate_months_list(cs["start_month"], cs["start_year"], period_months)
                            df_cs = pd.DataFrame(months_period)
                            for q in queries_clean:
                                df_cs[q] = 0
                            cs["df_data"] = df_cs
                            st.success("РќР°СЃС‚СЂРѕРµРЅ РїРµСЂРёРѕРґ Рё Р·Р°РїСЂРѕСЃС‹, Р·Р°РїРѕР»РЅРёС‚Рµ С‚Р°Р±Р»РёС†Сѓ РЅРёР¶Рµ.")
                            st.rerun()

                    # Р РµРґР°РєС‚РѕСЂ РґР°РЅРЅС‹С… Рё СЂР°СЃС‡С‘С‚
                    if cs.get("df_data") is not None:
                        demand_left, demand_right = st.columns([1.65, 1.0], vertical_alignment="top")
                        with demand_left:
                            st.markdown("**Р—Р°РїРѕР»РЅРёС‚Рµ РґР°РЅРЅС‹Рµ (РјРѕР¶РЅРѕ РєРѕРїРёСЂРѕРІР°С‚СЊ РёР· Excel С‡РµСЂРµР· Ctrl+C в†’ Ctrl+V):**")
                            df_edit = st.data_editor(
                                cs["df_data"],
                                use_container_width=True,
                                row_height=ui_editor_row_height,
                                num_rows="fixed",
                                column_config={
                                    "period": st.column_config.TextColumn("РџРµСЂРёРѕРґ", disabled=True),
                                    "month_num": st.column_config.NumberColumn("РњРµСЃСЏС† в„–", disabled=True),
                                    "year": st.column_config.NumberColumn("Р“РѕРґ", disabled=True),
                                },
                                key=f"cs_editor_{set_id}",
                            )
                            cs["df_data"] = df_edit

                            if st.button("Р Р°СЃСЃС‡РёС‚Р°С‚СЊ РєРѕСЌС„С„РёС†РёРµРЅС‚С‹ РґР»СЏ СЌС‚РѕРіРѕ РЅР°Р±РѕСЂР°", key=f"cs_calc_{set_id}", type="primary"):
                                _set_active_coeff_set(set_id)
                                queries_clean = [q.strip() for q in cs["queries"] if q.strip()]
                                if not queries_clean:
                                    st.error("Р”РѕР±Р°РІСЊС‚Рµ С…РѕС‚СЏ Р±С‹ РѕРґРёРЅ РїРѕРёСЃРєРѕРІС‹Р№ Р·Р°РїСЂРѕСЃ.")
                                else:
                                    df_coeffs = calculate_seasonality_coeffs_demand(df_edit, queries_clean)
                                    cs["result"] = df_coeffs
                                    st.success("РљРѕСЌС„С„РёС†РёРµРЅС‚С‹ СЂР°СЃСЃС‡РёС‚Р°РЅС‹!")
                        with demand_right:
                            st.markdown("**РС‚РѕРіРѕРІС‹Рµ РєРѕСЌС„С„РёС†РёРµРЅС‚С‹ СЃРїСЂРѕСЃР°**")
                            if cs.get("result") is not None:
                                st.dataframe(cs["result"], use_container_width=True, height=420)
                            else:
                                st.info("РџРѕСЃР»Рµ СЂР°СЃС‡С‘С‚Р° Р·РґРµСЃСЊ РїРѕСЏРІРёС‚СЃСЏ РёС‚РѕРіРѕРІР°СЏ С‚Р°Р±Р»РёС†Р° СЃРѕ СЃСЂРµРґРЅРµРІР·РІРµС€РµРЅРЅС‹РјРё Р·РЅР°С‡РµРЅРёСЏРјРё Рё РєРѕСЌС„С„РёС†РёРµРЅС‚Р°РјРё РїРѕ РјРµСЃСЏС†Р°Рј.")

                elif cs["type"] == "AOV (СЃСЂРµРґРЅРёР№ С‡РµРє)":
                    # ===== Р Р•Р–Рњ AOV (РЎР Р•Р”РќР™ Р§Р•Рљ) =====
                    st.markdown("**Р”Р°РЅРЅС‹Рµ РїРѕ СЃСЂРµРґРЅРµРјСѓ С‡РµРєСѓ (AOV):**")
                    st.caption(
                        f"Р”Р»СЏ РєР°Р¶РґРѕРіРѕ РјРµСЃСЏС†Р° РІС‹Р±СЂР°РЅРЅРѕРіРѕ РїРµСЂРёРѕРґР° СѓРєР°Р¶РёС‚Рµ СЃСЂРµРґРЅРёР№ С‡РµРє (AOV). РЎРµР№С‡Р°СЃ: {int(cs.get('period_months', 24) or 24)} РјРµСЃ."
                    )

                    # Р“РµРЅРµСЂР°С†РёСЏ С‚Р°Р±Р»РёС†С‹ РЅР° РІС‹Р±СЂР°РЅРЅС‹Р№ РїРµСЂРёРѕРґ (С‚РѕР»СЊРєРѕ AOV)
                    if st.button("РџСЂРёРјРµРЅРёС‚СЊ РЅР°СЃС‚СЂРѕР№РєРё РЅР°Р±РѕСЂР° (AOV)", key=f"cs_apply_aov_{set_id}", type="primary"):
                        _set_active_coeff_set(set_id)
                        period_months = int(cs.get("period_months", 24) or 24)
                        months_period = generate_months_list(cs["start_month"], cs["start_year"], period_months)
                        df_cs = pd.DataFrame(months_period)
                        df_cs["AOV"] = 0.0
                        cs["df_data"] = df_cs
                        st.success("РќР°СЃС‚СЂРѕРµРЅ РїРµСЂРёРѕРґ. Р—Р°РїРѕР»РЅРёС‚Рµ AOV РїРѕ РјРµСЃСЏС†Р°Рј РЅРёР¶Рµ.")
                        st.rerun()

                    # Р РµРґР°РєС‚РѕСЂ РґР°РЅРЅС‹С… Рё СЂР°СЃС‡С‘С‚ РґР»СЏ AOV
                    if cs.get("df_data") is not None:
                        aov_left, aov_right = st.columns([1.65, 1.0], vertical_alignment="top")
                        with aov_left:
                            st.markdown("**Р—Р°РїРѕР»РЅРёС‚Рµ AOV (СЃСЂРµРґРЅРёР№ С‡РµРє) РїРѕ РјРµСЃСЏС†Р°Рј:**")
                            df_edit = st.data_editor(
                                cs["df_data"],
                                use_container_width=True,
                                row_height=ui_editor_row_height,
                                num_rows="fixed",
                                column_config={
                                    "period": st.column_config.TextColumn("РџРµСЂРёРѕРґ", disabled=True),
                                    "month_num": st.column_config.NumberColumn("РњРµСЃСЏС† в„–", disabled=True),
                                    "year": st.column_config.NumberColumn("Р“РѕРґ", disabled=True),
                                    "AOV": st.column_config.NumberColumn(
                                        "AOV (СЃСЂРµРґРЅРёР№ С‡РµРє)", format="%.2f", help=mhelp("aov")
                                    ),
                                },
                                key=f"cs_editor_aov_{set_id}",
                            )
                            cs["df_data"] = df_edit

                        if st.button("Р Р°СЃСЃС‡РёС‚Р°С‚СЊ РєРѕСЌС„С„РёС†РёРµРЅС‚С‹ AOV РґР»СЏ СЌС‚РѕРіРѕ РЅР°Р±РѕСЂР°", key=f"cs_calc_aov_{set_id}", type="primary"):
                            _set_active_coeff_set(set_id)
                            df_coeffs = calculate_seasonality_coeffs_aov(df_edit)
                            cs["result"] = df_coeffs
                            st.success("РљРѕСЌС„С„РёС†РёРµРЅС‚С‹ AOV СЂР°СЃСЃС‡РёС‚Р°РЅС‹!")
                        with aov_right:
                            st.markdown("**РС‚РѕРіРѕРІС‹Рµ РєРѕСЌС„С„РёС†РёРµРЅС‚С‹ AOV**")
                            if cs.get("result") is not None:
                                st.dataframe(cs["result"], use_container_width=True, height=420)
                            else:
                                st.info("РџРѕСЃР»Рµ СЂР°СЃС‡РµС‚Р° Р·РґРµСЃСЊ РїРѕСЏРІРёС‚СЃСЏ РёС‚РѕРіРѕРІР°СЏ С‚Р°Р±Р»РёС†Р° РєРѕСЌС„С„РёС†РёРµРЅС‚РѕРІ.")
                elif cs["type"] == "РљР°СЃС‚РѕРјРЅС‹Р№ РЅР°Р±РѕСЂ":
                    # ===== Р Р•Р–Рњ РљРђРЎРўРћРњРќР«Р™ РќРђР‘РћР  =====
                    st.markdown("**РљР°СЃС‚РѕРјРЅС‹Рµ РєРѕСЌС„С„РёС†РёРµРЅС‚С‹ РЅР° 12 РјРµСЃСЏС†РµРІ:**")
                    st.caption(
                        "Р—Р°РїРѕР»РЅРёС‚Рµ РєРѕСЌС„С„РёС†РёРµРЅС‚С‹ РІСЂСѓС‡РЅСѓСЋ. Р­С‚РѕС‚ РЅР°Р±РѕСЂ РјРѕР¶РЅРѕ РёСЃРїРѕР»СЊР·РѕРІР°С‚СЊ Рё РґР»СЏ СЃРїСЂРѕСЃР°, Рё РґР»СЏ AOV."
                    )

                    month_names_map = {
                        1: "РЇРЅРІР°СЂСЊ", 2: "Р¤РµРІСЂР°Р»СЊ", 3: "РњР°СЂС‚", 4: "РђРїСЂРµР»СЊ",
                        5: "РњР°Р№", 6: "РСЋРЅСЊ", 7: "РСЋР»СЊ", 8: "РђРІРіСѓСЃС‚",
                        9: "РЎРµРЅС‚СЏР±СЂСЊ", 10: "РћРєС‚СЏР±СЂСЊ", 11: "РќРѕСЏР±СЂСЊ", 12: "Р”РµРєР°Р±СЂСЊ",
                    }

                    if (
                        cs.get("df_data") is None
                        or "month_num" not in cs["df_data"].columns
                        or "РљРѕСЌС„." not in cs["df_data"].columns
                    ):
                        cs["df_data"] = pd.DataFrame(
                            {
                                "month_num": list(range(1, 13)),
                                "РњРµСЃСЏС†": [month_names_map[m] for m in range(1, 13)],
                                "РљРѕСЌС„.": [1.0] * 12,
                            }
                        )
                    elif "РљРѕСЌС„. AOV" in cs["df_data"].columns:
                        custom_df = cs["df_data"].copy()
                        custom_df = custom_df.drop(columns=["РљРѕСЌС„. AOV"], errors="ignore")
                        cs["df_data"] = custom_df

                    custom_left, custom_right = st.columns([1.65, 1.0], vertical_alignment="top")
                    with custom_left:
                        df_edit = st.data_editor(
                            cs["df_data"],
                            use_container_width=True,
                            row_height=ui_editor_row_height,
                            num_rows="fixed",
                            column_config={
                                "month_num": st.column_config.NumberColumn("РњРµСЃСЏС† в„–", disabled=True),
                                "РњРµСЃСЏС†": st.column_config.TextColumn("РњРµСЃСЏС†", disabled=True),
                                "РљРѕСЌС„.": st.column_config.NumberColumn(
                                    "РљРѕСЌС„С„РёС†РёРµРЅС‚", format="%.2f",
                                    help="РљРѕСЌС„С„РёС†РёРµРЅС‚ РјРµСЃСЏС†Р°: 1.00 = Р±РµР· РёР·РјРµРЅРµРЅРёР№, 1.20 = +20%, 0.80 = -20%."
                                ),
                            },
                            key=f"cs_editor_custom_{set_id}",
                        )
                        cs["df_data"] = df_edit
                    with custom_right:
                        st.markdown("**Р”РµР№СЃС‚РІРёСЏ**")
                        if st.button("РџСЂРёРјРµРЅРёС‚СЊ РєР°СЃС‚РѕРјРЅС‹Р№ РЅР°Р±РѕСЂ", key=f"cs_apply_custom_{set_id}", type="primary"):
                            _set_active_coeff_set(set_id)
                            df_coeffs = df_edit.copy()
                            if "РљРѕСЌС„." not in df_coeffs.columns:
                                df_coeffs["РљРѕСЌС„."] = 1.0
                            df_coeffs["РќРѕРјРµСЂ РјРµСЃСЏС†Р°"] = df_coeffs["month_num"].astype(int)
                            df_coeffs = df_coeffs[["РќРѕРјРµСЂ РјРµСЃСЏС†Р°", "РњРµСЃСЏС†", "РљРѕСЌС„."]]
                            cs["result"] = df_coeffs
                            st.success("РљР°СЃС‚РѕРјРЅС‹Рµ РєРѕСЌС„С„РёС†РёРµРЅС‚С‹ СЃРѕС…СЂР°РЅРµРЅС‹!")
                        else:
                            st.caption("Р’РЅРѕСЃРёС‚Рµ Р·РЅР°С‡РµРЅРёСЏ СЃР»РµРІР° Рё РїСЂРёРјРµРЅСЏР№С‚Рµ РЅР°Р±РѕСЂ РєРЅРѕРїРєРѕР№ РІС‹С€Рµ.")
                else:
                    # ===== Р Р•Р–РРњ РњР•Р”РР™РќР«Р• РҐР’РћРЎРўР« =====
                    st.markdown("**РњРµРґРёР№РЅС‹Рµ С…РІРѕСЃС‚С‹ (РєРѕСЌС„С„РёС†РёРµРЅС‚С‹ РїРѕРєР°Р·РѕРІ РЅР° 12 РјРµСЃСЏС†РµРІ):**")
                    st.caption(
                        "Р­С‚РѕС‚ РЅР°Р±РѕСЂ РЅР°РєР»Р°РґС‹РІР°РµС‚СЃСЏ РЅР° СЃРїСЂРѕСЃ Рё РІР»РёСЏРµС‚ С‚РѕР»СЊРєРѕ РЅР° РїРѕРєР°Р·С‹ "
                        "(k_imp = k_demand Г— k_media_tail)."
                    )

                    month_names_map = {
                        1: "РЇРЅРІР°СЂСЊ", 2: "Р¤РµРІСЂР°Р»СЊ", 3: "РњР°СЂС‚", 4: "РђРїСЂРµР»СЊ",
                        5: "РњР°Р№", 6: "РСЋРЅСЊ", 7: "РСЋР»СЊ", 8: "РђРІРіСѓСЃС‚",
                        9: "РЎРµРЅС‚СЏР±СЂСЊ", 10: "РћРєС‚СЏР±СЂСЊ", 11: "РќРѕСЏР±СЂСЊ", 12: "Р”РµРєР°Р±СЂСЊ",
                    }

                    if (
                        cs.get("df_data") is None
                        or "month_num" not in cs["df_data"].columns
                        or "РљРѕСЌС„." not in cs["df_data"].columns
                    ):
                        cs["df_data"] = pd.DataFrame(
                            {
                                "month_num": list(range(1, 13)),
                                "РњРµСЃСЏС†": [month_names_map[m] for m in range(1, 13)],
                                "РљРѕСЌС„.": [1.0] * 12,
                            }
                        )

                    media_left, media_right = st.columns([1.65, 1.0], vertical_alignment="top")
                    with media_left:
                        df_edit = st.data_editor(
                            cs["df_data"],
                            use_container_width=True,
                            row_height=ui_editor_row_height,
                            num_rows="fixed",
                            column_config={
                                "month_num": st.column_config.NumberColumn("РњРµСЃСЏС† в„–", disabled=True),
                                "РњРµСЃСЏС†": st.column_config.TextColumn("РњРµСЃСЏС†", disabled=True),
                                "РљРѕСЌС„.": st.column_config.NumberColumn(
                                    "РљРѕСЌС„С„РёС†РёРµРЅС‚ С…РІРѕСЃС‚РѕРІ", format="%.2f",
                                    help="РњРµРґРёР№РЅС‹Р№ С…РІРѕСЃС‚ РґР»СЏ РїРѕРєР°Р·РѕРІ: k_imp = k_demand Г— k_media_tail."
                                ),
                            },
                            key=f"cs_editor_media_tail_{set_id}",
                        )
                        cs["df_data"] = df_edit
                    with media_right:
                        st.markdown("**Р”РµР№СЃС‚РІРёСЏ**")
                        if st.button("РџСЂРёРјРµРЅРёС‚СЊ РЅР°Р±РѕСЂ РјРµРґРёР№РЅС‹С… С…РІРѕСЃС‚РѕРІ", key=f"cs_apply_media_tail_{set_id}", type="primary"):
                            _set_active_coeff_set(set_id)
                            df_coeffs = df_edit.copy()
                            if "РљРѕСЌС„." not in df_coeffs.columns:
                                df_coeffs["РљРѕСЌС„."] = 1.0
                            df_coeffs["РќРѕРјРµСЂ РјРµСЃСЏС†Р°"] = df_coeffs["month_num"].astype(int)
                            df_coeffs = df_coeffs[["РќРѕРјРµСЂ РјРµСЃСЏС†Р°", "РњРµСЃСЏС†", "РљРѕСЌС„."]]
                            cs["result"] = df_coeffs
                            st.success("РќР°Р±РѕСЂ РјРµРґРёР№РЅС‹С… С…РІРѕСЃС‚РѕРІ СЃРѕС…СЂР°РЅРµРЅ!")
                        else:
                            st.caption("РќР°Р±РѕСЂ РІР»РёСЏРµС‚ С‚РѕР»СЊРєРѕ РЅР° РїРѕРєР°Р·С‹ Рё РЅР°РєР»Р°РґС‹РІР°РµС‚СЃСЏ РїРѕРІРµСЂС… СЃРїСЂРѕСЃР°.")
                    cs_type_norm = normalize_coeff_set_type(cs.get("type"))
                    if cs_type_norm == "РЎРїСЂРѕСЃ (РїРѕ Р·Р°РїСЂРѕСЃР°Рј)":
                        st.markdown("### РС‚РѕРіРѕРІС‹Рµ РєРѕСЌС„С„РёС†РёРµРЅС‚С‹")
                        st.dataframe(cs["result"], use_container_width=True)

                    buf = io.BytesIO()
                    sheet_name = (cs["name"] or "РљРѕСЌС„С„РёС†РёРµРЅС‚С‹")[:31]
                    excel_engine = "xlsxwriter" if HAS_XLSXWRITER else "openpyxl"
                    with pd.ExcelWriter(buf, engine=excel_engine) as writer:
                        cs["result"].to_excel(writer, sheet_name=sheet_name, index=False)
                    buf.seek(0)

                    st.download_button(
                        label="рџ“Ґ РЎРєР°С‡Р°С‚СЊ РєРѕСЌС„С„РёС†РёРµРЅС‚С‹ СЌС‚РѕРіРѕ РЅР°Р±РѕСЂР°",
                        data=buf,
                        file_name=f"coeffs_set_{set_id}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"cs_dl_{set_id}",
                    )
    render_bottom_tab_switcher("РљРѕСЌС„С„РёС†РёРµРЅС‚С‹", "coeffs")
# ====================================================================
#                           РўРђР‘ В«РњР•Р”РђРџР›РђРќВ»
# ====================================================================

with tab_plan:
    st.markdown(
        """
        <div class="tab-intro">
            <p>1) Р’С‹Р±РµСЂРёС‚Рµ РјРµСЃСЏС†С‹, РґР»СЏ РєРѕС‚РѕСЂС‹С… РЅСѓР¶РЅРѕ СЂР°СЃСЃС‡РёС‚Р°С‚СЊ РјРµРґРёР°РїР»Р°РЅ.</p>
            <p>2) Р’С‹Р±РµСЂРёС‚Рµ РїСЂРµСЃРµС‚ РјРµС‚СЂРёРє <b>E-com</b>, <b>DIY</b> РёР»Рё <b>РќРµРґРІРёР¶РёРјРѕСЃС‚СЊ</b>. РџСЂРµСЃРµС‚ РІР»РёСЏРµС‚ РЅР° СЃРѕСЃС‚Р°РІ РјРµС‚СЂРёРє Рё Р»РѕРіРёРєСѓ РѕС‚РѕР±СЂР°Р¶РµРЅРёСЏ РѕС‚РґРµР»СЊРЅС‹С… РїРѕРєР°Р·Р°С‚РµР»РµР№ РІ СЂР°СЃС‡РµС‚Рµ.</p>
            <p>3) Р—Р°РїРѕР»РЅРёС‚Рµ РґР°РЅРЅС‹Рµ РїРѕ С‚РёРїР°Рј СЂРµРєР»Р°РјРЅС‹С… РєР°РјРїР°РЅРёР№ Рё Р±Р°Р·РѕРІС‹Рј РјРµС‚СЂРёРєР°Рј РґР»СЏ СЂР°СЃС‡РµС‚Р° СЃСЂРµРґРЅРµРіРѕ РјРµСЃСЏС†Р°. Р—РЅР°С‡РµРЅРёСЏ РІ Р±Р»РѕРє В«РЎСЂРµРґРЅРёР№ РјРµСЃСЏС†В» РІРЅРѕСЃСЏС‚СЃСЏ <b><span style="color:#9EC5FF;">Р±РµР· РќР”РЎ</span></b>. РџРѕРјРЅРёС‚Рµ: СЃРµР·РѕРЅРЅРѕСЃС‚СЊ СЃСЂРµРґРЅРµРіРѕ РјРµСЃСЏС†Р° СЂР°РІРЅР° 1, Рё РѕС‚ РЅРµРµ СЂР°СЃСЃС‡РёС‚С‹РІР°СЋС‚СЃСЏ РІСЃРµ РІС‹Р±СЂР°РЅРЅС‹Рµ РјРµСЃСЏС†С‹ С‡РµСЂРµР· РєРѕСЌС„С„РёС†РёРµРЅС‚С‹.</p>
            <p>4) РџСЂРё РЅРµРѕР±С…РѕРґРёРјРѕСЃС‚Рё РЅР°СЃС‚СЂРѕР№С‚Рµ СѓС‡РµС‚ <b>РќР”РЎ</b> Рё <b>РђРљ</b>: РјРѕР¶РЅРѕ РІРєР»СЋС‡РёС‚СЊ РќР”РЎ, Р·Р°РґР°С‚СЊ С„РёРєСЃРёСЂРѕРІР°РЅРЅСѓСЋ РђРљ РЅР° РјРµСЃСЏС† РёР»Рё РёСЃРїРѕР»СЊР·РѕРІР°С‚СЊ С€РєР°Р»Сѓ РђРљ РѕС‚ TOTAL Р±СЋРґР¶РµС‚Р° РјРµСЃСЏС†Р° Р±РµР· РќР”РЎ.</p>
            <p>5) РќР°Р·РЅР°С‡СЊС‚Рµ РґР»СЏ РєР°Р¶РґРѕРіРѕ С‚РёРїР° Р Рљ РЅР°Р±РѕСЂС‹ РєРѕСЌС„С„РёС†РёРµРЅС‚РѕРІ: <b>РЎРїСЂРѕСЃ</b>, <b>AOV</b>, <b>РљР°СЃС‚РѕРјРЅС‹Р№ РЅР°Р±РѕСЂ</b> Рё РїСЂРё РЅРµРѕР±С…РѕРґРёРјРѕСЃС‚Рё <b>РњРµРґРёР№РЅС‹Рµ С…РІРѕСЃС‚С‹</b>. Р”Р»СЏ РїСЂРµСЃРµС‚Р° <b>РќРµРґРІРёР¶РёРјРѕСЃС‚СЊ</b> РЅР°Р±РѕСЂ <b>AOV</b> РЅРµ РёСЃРїРѕР»СЊР·СѓРµС‚СЃСЏ.</p>
            <p>6) РќР°СЃС‚СЂРѕР№С‚Рµ СЌР»Р°СЃС‚РёС‡РЅРѕСЃС‚СЊ РјРµС‚СЂРёРє Рє СЃРµР·РѕРЅРЅРѕСЃС‚Рё СЃРїСЂРѕСЃР°: РјРѕР¶РЅРѕ РІС‹Р±СЂР°С‚СЊ РїСЂРµСЃРµС‚ РЅР°СЃС‚СЂРѕРµРє РёР»Рё Р·Р°РґР°С‚СЊ Р·РЅР°С‡РµРЅРёСЏ РІСЂСѓС‡РЅСѓСЋ РґР»СЏ <b>CPC</b>, <b>CTR</b> Рё <b>CR</b>.</p>
            <p>7) РџСЂРѕРІРµСЂСЊС‚Рµ СЂР°СЃС‡РµС‚С‹ РїРѕ РєР°Р¶РґРѕРјСѓ РјРµСЃСЏС†Сѓ: СЃР»РµРІР° РїРѕРєР°Р·Р°РЅС‹ СЂРµР·СѓР»СЊС‚Р°С‚С‹, СЃРїСЂР°РІР° РїСЂРёРјРµРЅРµРЅРЅС‹Рµ РєРѕСЌС„С„РёС†РёРµРЅС‚С‹. РљРѕСЌС„С„РёС†РёРµРЅС‚С‹ СЃРїСЂР°РІР° СЂРµРґР°РєС‚РёСЂСѓРµРјС‹Рµ, РїРѕСЌС‚РѕРјСѓ РёР·РјРµРЅРµРЅРёСЏ РјРѕР¶РЅРѕ РІРЅРѕСЃРёС‚СЊ РїСЂСЏРјРѕ РІ СЏС‡РµР№РєРё Рё СЃСЂР°Р·Сѓ СЃРјРѕС‚СЂРµС‚СЊ РЅР° СЂРµР·СѓР»СЊС‚Р°С‚.</p>
            <p>8) РћРїС†РёРѕРЅР°Р»СЊРЅРѕ РёСЃРїРѕР»СЊР·СѓР№С‚Рµ Р»РµРІРѕРµ РјРµРЅСЋ В«Р‘С‹СЃС‚СЂР°СЏ СЃРІРµСЂРєР°В», С‡С‚РѕР±С‹ Р·Р°РєСЂРµРїРёС‚СЊ РїРµСЂРµРґ РіР»Р°Р·Р°РјРё СЃСЂРµРґРЅРёР№ РјРµСЃСЏС† РёР»Рё TOTAL РІС‹Р±СЂР°РЅРЅС‹С… РјРµСЃСЏС†РµРІ Рё СѓРґРѕР±РЅРµРµ СЃСЂР°РІРЅРёРІР°С‚СЊ РґР°РЅРЅС‹Рµ РїСЂРё СЂРµРґР°РєС‚РёСЂРѕРІР°РЅРёРё.</p>
            <p>9) РЎРІРµСЂСЊС‚Рµ РёС‚РѕРіРё РІ Р±Р»РѕРєРµ TOTAL РїРѕ РІС‹Р±СЂР°РЅРЅС‹Рј РјРµСЃСЏС†Р°Рј Рё РїСЂРё РЅРµРѕР±С…РѕРґРёРјРѕСЃС‚Рё СЃРєРѕСЂСЂРµРєС‚РёСЂСѓР№С‚Рµ РЅР°СЃС‚СЂРѕР№РєРё.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    if "use_vat_budget_metrics" not in st.session_state:
        st.session_state["use_vat_budget_metrics"] = True
    if "use_ak_budget_metrics" not in st.session_state:
        st.session_state["use_ak_budget_metrics"] = False
    if "ak_mode" not in st.session_state:
        st.session_state["ak_mode"] = "percent"
    if "ak_fixed_month_wo_vat" not in st.session_state:
        st.session_state["ak_fixed_month_wo_vat"] = 200000.0
    if "ak_fixed_percent" not in st.session_state:
        st.session_state["ak_fixed_percent"] = 2.0
    if "ak_rules_editor_nonce" not in st.session_state:
        st.session_state["ak_rules_editor_nonce"] = 0
    if "elasticity_editor_nonce" not in st.session_state:
        st.session_state["elasticity_editor_nonce"] = 0
    default_ak_rules = pd.DataFrame(
        [
            {"min_budget_wo_vat": 0.0, "ak_percent": 0.0},
            {"min_budget_wo_vat": 1500000.0, "ak_percent": 8.0},
            {"min_budget_wo_vat": 3000000.0, "ak_percent": 4.0},
            {"min_budget_wo_vat": 6000000.0, "ak_percent": 2.0},
            {"min_budget_wo_vat": 10000000.0, "ak_percent": 0.0},
        ]
    )
    if "ak_rules_df" not in st.session_state:
        st.session_state["ak_rules_df"] = default_ak_rules.copy()
    use_vat_budget_metrics = bool(st.session_state.get("use_vat_budget_metrics", True))
    use_ak_budget_metrics = bool(st.session_state.get("use_ak_budget_metrics", False))
    ak_mode = str(st.session_state.get("ak_mode", "fixed"))
    ak_fixed_month_wo_vat = float(st.session_state.get("ak_fixed_month_wo_vat", 200000.0))
    ak_fixed_percent = float(st.session_state.get("ak_fixed_percent", 2.0))
    ak_rules_df = st.session_state.get("ak_rules_df", pd.DataFrame())
    ak_rules_invalid = (
        not isinstance(ak_rules_df, pd.DataFrame)
        or ak_rules_df.empty
        or "min_budget_wo_vat" not in ak_rules_df.columns
        or "ak_percent" not in ak_rules_df.columns
    )
    if not ak_rules_invalid:
        _min_col = pd.to_numeric(ak_rules_df["min_budget_wo_vat"], errors="coerce")
        _pct_col = pd.to_numeric(ak_rules_df["ak_percent"], errors="coerce")
        ak_rules_invalid = _min_col.isna().all() or _pct_col.isna().all()
    if ak_rules_invalid:
        st.session_state["ak_rules_df"] = default_ak_rules.copy()
        ak_rules_df = st.session_state["ak_rules_df"]
        st.session_state["ak_rules_editor_nonce"] = int(st.session_state.get("ak_rules_editor_nonce", 0)) + 1

    month_names_full = {
        1: "РЇРЅРІР°СЂСЊ",
        2: "Р¤РµРІСЂР°Р»СЊ",
        3: "РњР°СЂС‚",
        4: "РђРїСЂРµР»СЊ",
        5: "РњР°Р№",
        6: "РСЋРЅСЊ",
        7: "РСЋР»СЊ",
        8: "РђРІРіСѓСЃС‚",
        9: "РЎРµРЅС‚СЏР±СЂСЊ",
        10: "РћРєС‚СЏР±СЂСЊ",
        11: "РќРѕСЏР±СЂСЊ",
        12: "Р”РµРєР°Р±СЂСЊ",
    }

    def _fmt_ref_int(v: float) -> str:
        return f"{int(round(v)):,}".replace(",", " ")

    def _fmt_ref_rub(v: float) -> str:
        return f"{int(round(v)):,} в‚Ѕ".replace(",", " ")

    def _fmt_ref_rub2(v: float) -> str:
        return f"{float(v):.2f} в‚Ѕ"

    def _fmt_ref_pct(v: float) -> str:
        return f"{float(v):.2f}%"

    def _fmt_ref_roas(v: float) -> str:
        return f"{float(v):.2f}%"

    def _build_ref_df(ref: dict) -> pd.DataFrame:
        if not ref:
            return pd.DataFrame(columns=["РџРѕРєР°Р·Р°С‚РµР»СЊ", "Р—РЅР°С‡РµРЅРёРµ"])
        if is_real_estate_preset:
            rows = [
                {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "РџРѕРєР°Р·С‹", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_int(ref.get("impressions", 0.0))},
                {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "РљР»РёРєРё", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_int(ref.get("clicks", 0.0))},
                {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "Р‘СЋРґР¶РµС‚", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_rub(ref.get("cost", 0.0))},
                {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "Р‘СЋРґР¶РµС‚ СЃ РќР”РЎ", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_rub(ref.get("cost_with_vat", 0.0))},
                {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "Р‘СЋРґР¶РµС‚ СЃ РќР”РЎ Рё РђРљ", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_rub(ref.get("cost_with_vat_ak", 0.0))},
                {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "CTR", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_pct(ref.get("ctr", 0.0))},
                {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "CPC", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_rub2(ref.get("cpc", 0.0))},
            ]
            if metric_mode["is_real_estate_full"]:
                rows += [
                    {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "CR1 РІ Р›РёРґ", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_pct(ref.get("cr1", 0.0))},
                    {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "Р›РёРґС‹", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_int(ref.get("leads", 0.0))},
                    {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "CPL", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_rub(ref.get("cpl", 0.0))},
                    {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "CR2 РІ Р¦Рћ", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_pct(ref.get("cr2", 0.0))},
                    {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "Р¦Рћ", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_int(ref.get("target_leads", ref.get("conversions", 0.0)))},
                    {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "CPQL", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_rub(ref.get("cpql", 0.0))},
                ]
            else:
                rows += [
                    {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "CR РІ Р¦Рћ", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_pct(ref.get("cr", 0.0))},
                    {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "Р¦Рћ", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_int(ref.get("target_leads", ref.get("conversions", 0.0)))},
                    {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "CPQL", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_rub(ref.get("cpql", ref.get("cpo", 0.0)))},
                ]
        else:
            rows = [
                {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "РџРѕРєР°Р·С‹", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_int(ref.get("impressions", 0.0))},
                {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "РљР»РёРєРё", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_int(ref.get("clicks", 0.0))},
                {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "РљРѕРЅРІРµСЂСЃРёРё", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_int(ref.get("conversions", 0.0))},
                {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "Р‘СЋРґР¶РµС‚", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_rub(ref.get("cost", 0.0))},
                {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "Р‘СЋРґР¶РµС‚ СЃ РќР”РЎ", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_rub(ref.get("cost_with_vat", 0.0))},
                {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "Р‘СЋРґР¶РµС‚ СЃ РќР”РЎ Рё РђРљ", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_rub(ref.get("cost_with_vat_ak", 0.0))},
                {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "Р”РѕС…РѕРґ", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_rub(ref.get("revenue", 0.0))},
                {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "CTR", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_pct(ref.get("ctr", 0.0))},
                {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "CPC", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_rub2(ref.get("cpc", 0.0))},
                {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "CR", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_pct(ref.get("cr", 0.0))},
                {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "CPO", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_rub(ref.get("cpo", 0.0))},
                {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "ROAS", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_roas(ref.get("roas", 0.0))},
                {"РџРѕРєР°Р·Р°С‚РµР»СЊ": "Р”Р Р ", "Р—РЅР°С‡РµРЅРёРµ": _fmt_ref_pct(ref.get("drr", 0.0))},
            ]
        return pd.DataFrame(rows)

    # ---------- 0. Р“РѕСЂРёР·РѕРЅС‚ РїР»Р°РЅРёСЂРѕРІР°РЅРёСЏ ----------

    ui_section_title("0. Р“РѕСЂРёР·РѕРЅС‚ РїР»Р°РЅРёСЂРѕРІР°РЅРёСЏ")
    st.caption(
        "Р’С‹Р±СЂР°РЅРЅС‹Рµ РјРµСЃСЏС†С‹ РЅР°РїСЂСЏРјСѓСЋ РІР»РёСЏСЋС‚ РЅР° РїРѕРјРµСЃСЏС‡РЅС‹Р№ СЂР°СЃС‡РµС‚, РёС‚РѕРіРё РІ Р±Р»РѕРєРµ TOTAL "
        "Рё РґР°РЅРЅС‹Рµ РЅР° РІРєР»Р°РґРєРµ В«Р”РёР°РіСЂР°РјРјС‹В»."
    )

    all_month_nums = list(month_names_full.keys())
    all_month_labels = [f"{i}. {month_names_full[i]}" for i in all_month_nums]
    if "planning_months_multiselect" not in st.session_state:
        st.session_state["planning_months_multiselect"] = all_month_labels.copy()

    selected_month_labels = st.multiselect(
        "Р’С‹Р±РµСЂРёС‚Рµ РјРµСЃСЏС†С‹, РґР»СЏ РєРѕС‚РѕСЂС‹С… СЃС‡РёС‚Р°С‚СЊ РјРµРґРёР°РїР»Р°РЅ:",
        options=all_month_labels,
        key="planning_months_multiselect",
    )

    plan_calc_blockers = []
    if not selected_month_labels:
        st.warning("Р’С‹Р±РµСЂРёС‚Рµ С…РѕС‚СЏ Р±С‹ РѕРґРёРЅ РјРµСЃСЏС† РґР»СЏ СЂР°СЃС‡С‘С‚Р° РјРµРґРёР°РїР»Р°РЅР°.")
        plan_calc_blockers.append("РЅРµ РІС‹Р±СЂР°РЅС‹ РјРµСЃСЏС†С‹ РїР»Р°РЅРёСЂРѕРІР°РЅРёСЏ")

    selected_month_nums = [int(label.split(".")[0]) for label in selected_month_labels] if selected_month_labels else []
    months_count = len(selected_month_nums)

    ui_section_title("0.1 РџСЂРµСЃРµС‚ РјРµС‚СЂРёРє")
    st.caption(
        "РџРѕСЃР»Рµ РІС‹Р±РѕСЂР° РіРѕСЂРёР·РѕРЅС‚Р° РїР»Р°РЅРёСЂРѕРІР°РЅРёСЏ РІС‹Р±РµСЂРёС‚Рµ РїСЂРµСЃРµС‚ РјРµС‚СЂРёРє. "
        "Р”РѕСЃС‚СѓРїРЅС‹ РїСЂРµСЃРµС‚С‹: E-com, DIY Рё РќРµРґРІРёР¶РёРјРѕСЃС‚СЊ."
    )
    preset_keys = list(METRIC_PRESETS.keys())
    if "metric_preset_key" not in st.session_state or st.session_state["metric_preset_key"] not in preset_keys:
        st.session_state["metric_preset_key"] = "ecom"
    preset_col, _preset_spacer = st.columns([1, 3])
    with preset_col:
        active_preset_key = st.selectbox(
            "Р’С‹Р±РµСЂРёС‚Рµ РїСЂРµСЃРµС‚ РјРµС‚СЂРёРє:",
            options=preset_keys,
            index=preset_keys.index(st.session_state["metric_preset_key"]),
            format_func=lambda k: METRIC_PRESETS[k]["label"],
            key="metric_preset_key",
        )
    st.caption(f"РђРєС‚РёРІРЅС‹Р№ РїСЂРµСЃРµС‚: {METRIC_PRESETS[active_preset_key]['label']}")
    if "real_estate_funnel_mode" not in st.session_state:
        st.session_state["real_estate_funnel_mode"] = "simple"
    metric_mode = get_metric_mode(active_preset_key, st.session_state.get("real_estate_funnel_mode", "simple"))
    is_diy_preset = metric_mode["is_diy"]
    is_real_estate_preset = metric_mode["is_real_estate"]
    if is_real_estate_preset:
        re_col, _re_spacer = st.columns([1.2, 2.8])
        with re_col:
            selected_re_mode = st.selectbox(
                "Р РµР¶РёРј РІРѕСЂРѕРЅРєРё:",
                options=list(REAL_ESTATE_FUNNEL_OPTIONS.keys()),
                index=list(REAL_ESTATE_FUNNEL_OPTIONS.keys()).index(metric_mode["real_estate_funnel_mode"]),
                format_func=lambda k: REAL_ESTATE_FUNNEL_OPTIONS[k],
                key="real_estate_funnel_mode",
            )
        metric_mode = get_metric_mode(active_preset_key, selected_re_mode)
        if is_diy_preset:
            DISPLAY_COL_RENAME["conversions"] = "РљРѕР»-РІРѕ РѕС„РѕСЂРјР»РµРЅРЅС‹С… Р·Р°РєР°Р·РѕРІ"
            DISPLAY_COL_RENAME["cr"] = "CR1"
            DISPLAY_COL_RENAME["cr_pct"] = "CR1"
            DISPLAY_COL_RENAME["revenue"] = "РћС„РѕСЂРјР»РµРЅРЅС‹Р№ РґРѕС…РѕРґ, в‚Ѕ СЃ РќР”РЎ"
            DISPLAY_COL_RENAME["aov"] = "РЎСЂРµРґРЅРёР№ С‡РµРє РїРѕ РѕС„РѕСЂРјР». РґРѕС…РѕРґСѓ, СЃ РќР”РЎ"
            DISPLAY_COL_RENAME["cpa"] = "РЎС‚РѕРёРјРѕСЃС‚СЊ РѕС„РѕСЂРјР». Р·Р°РєР°Р·Р°, в‚Ѕ СЃ РќР”РЎ"
            DISPLAY_COL_RENAME["ROAS"] = "ROAS РѕС„РѕСЂРјР»."
            DISPLAY_COL_RENAME["drr"] = "Р”Р Р  РѕС‚ РѕС„РѕСЂРјР». РґРѕС…РѕРґР°, %"
            DISPLAY_COL_RENAME["drr_pct"] = "Р”Р Р  РѕС‚ РѕС„РѕСЂРјР». РґРѕС…РѕРґР°, %"
            DISPLAY_COL_RENAME["cr2_pct"] = "CR2"
            DISPLAY_COL_RENAME["target_leads"] = "РљРѕР»-РІРѕ РѕС‚РіСЂСѓР¶РµРЅРЅС‹С… Р·Р°РєР°Р·РѕРІ"
            DISPLAY_COL_RENAME["cpql"] = "РЎС‚РѕРёРјРѕСЃС‚СЊ РѕС‚РіСЂ. Р·Р°РєР°Р·Р°, в‚Ѕ СЃ РќР”РЎ"
    elif is_real_estate_preset:
        DISPLAY_COL_RENAME["conversions"] = "Р¦Рћ"
        DISPLAY_COL_RENAME["target_leads"] = "Р¦Рћ"
        DISPLAY_COL_RENAME["revenue"] = "Р’С‹СЂСѓС‡РєР°"
        if metric_mode["is_real_estate_full"]:
            DISPLAY_COL_RENAME["cr"] = "CR РІ Р¦Рћ"
            DISPLAY_COL_RENAME["cr_pct"] = "CR РІ Р¦Рћ"
            DISPLAY_COL_RENAME["cpa"] = "CPQL"
        else:
            DISPLAY_COL_RENAME["cr"] = "CR РІ Р¦Рћ"
            DISPLAY_COL_RENAME["cr_pct"] = "CR РІ Р¦Рћ"
            DISPLAY_COL_RENAME["cpa"] = "CPQL"
    else:
        DISPLAY_COL_RENAME["conversions"] = "РљРѕРЅРІРµСЂСЃРёРё"
        DISPLAY_COL_RENAME["cr"] = "CR"
        DISPLAY_COL_RENAME["cr_pct"] = "CR"
        DISPLAY_COL_RENAME["revenue"] = "Р”РѕС…РѕРґ"
        DISPLAY_COL_RENAME["cpa"] = "CPO"

    def _df_for_compare(df: pd.DataFrame) -> pd.DataFrame:
        if df is None:
            return pd.DataFrame()
        tmp = df.copy().reset_index(drop=True)
        tmp = tmp.reindex(sorted(tmp.columns), axis=1)
        return tmp.fillna("")

    def _has_unsaved_changes(draft_df: pd.DataFrame, saved_df: pd.DataFrame) -> bool:
        return not _df_for_compare(draft_df).equals(_df_for_compare(saved_df))

    def _table_height_for_rows(rows_count: int, min_height: int = 120, max_height: int = 560) -> int:
        header_px = 34 if ui_compact_tables else 38
        row_px = ui_table_row_px
        h = header_px + max(rows_count, 1) * row_px + 6
        return max(min_height, min(h, max_height))

    # ---------- 1. РўРёРїС‹ Р Рљ (СЃСЂРµРґРЅРёР№ РјРµСЃСЏС†) ----------

    ui_section_title("1. РўРёРїС‹ СЂРµРєР»Р°РјРЅС‹С… РєР°РјРїР°РЅРёР№ РІ СЃСЂРµРґРЅРµРј РјРµСЃСЏС†Рµ")
    with st.expander("РўРёРїС‹ СЂРµРєР»Р°РјРЅС‹С… РєР°РјРїР°РЅРёР№ РІ СЃСЂРµРґРЅРµРј РјРµСЃСЏС†Рµ", expanded=True):

        default_campaigns = pd.DataFrame(
            [
                ["РџРѕРёСЃРє Р±СЂРµРЅРґ", "B2C", "РЇРЅРґРµРєСЃ", "РџРѕРёСЃРє", "",        500_000, 5.0, 15.0, 5.0, 50.0, 5000.0, 220_000, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
                ["Р РЎРЇ Р±Р°РЅРЅРµСЂС‹", "B2C", "РЇРЅРґРµРєСЃ", "Р РЎРЇ Р±Р°РЅРЅРµСЂС‹", "", 1_000_000, 1.0, 10.0, 2.0, 40.0, 3000.0, 450_000, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
                ["Р’РёРґРµРѕ YouTube", "B2B", "YouTube", "Р’РёРґРµРѕ", "",     300_000, 0.7, 20.0, 1.5, 35.0, 4000.0, 180_000, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
            ],
            columns=[
                "campaign_type", "segment", "system", "format", "geo",
                "impressions_avg", "ctr_avg_percent", "cpc_avg", "cr_avg_percent", "cr2_avg_percent", "aov_avg", "reach_avg",
                "available_capacity_avg", "client_count_avg", "absolute_new_clients_avg", "returned_clients_avg", "order_frequency_avg", "new_clients_share_avg_percent",
            ],
        )

        if "campaigns_df" not in st.session_state:
            st.session_state["campaigns_df"] = default_campaigns.copy()
        if "segment" not in st.session_state["campaigns_df"].columns:
            st.session_state["campaigns_df"]["segment"] = "B2C"
        if "geo" not in st.session_state["campaigns_df"].columns:
            st.session_state["campaigns_df"]["geo"] = ""
        if "available_capacity_avg" not in st.session_state["campaigns_df"].columns:
            st.session_state["campaigns_df"]["available_capacity_avg"] = 0.0
        if "client_count_avg" not in st.session_state["campaigns_df"].columns:
            st.session_state["campaigns_df"]["client_count_avg"] = 0.0
        if "absolute_new_clients_avg" not in st.session_state["campaigns_df"].columns:
            st.session_state["campaigns_df"]["absolute_new_clients_avg"] = 0.0
        if "returned_clients_avg" not in st.session_state["campaigns_df"].columns:
            st.session_state["campaigns_df"]["returned_clients_avg"] = 0.0
        if "order_frequency_avg" not in st.session_state["campaigns_df"].columns:
            st.session_state["campaigns_df"]["order_frequency_avg"] = 0.0
        if "new_clients_share_avg_percent" not in st.session_state["campaigns_df"].columns:
            st.session_state["campaigns_df"]["new_clients_share_avg_percent"] = 0.0
        if "reach_avg" not in st.session_state["campaigns_df"].columns:
            st.session_state["campaigns_df"]["reach_avg"] = 0.0
        if "cr2_avg_percent" not in st.session_state["campaigns_df"].columns:
            st.session_state["campaigns_df"]["cr2_avg_percent"] = 0.0

        campaign_cols_order = [
            "campaign_type",
            "system",
            "format",
            "geo",
            "segment",
            "impressions_avg",
            "ctr_avg_percent",
            "cpc_avg",
            "cr_avg_percent",
            "cr2_avg_percent",
            "aov_avg",
            "reach_avg",
            "available_capacity_avg",
            "client_count_avg",
            "absolute_new_clients_avg",
            "returned_clients_avg",
            "order_frequency_avg",
            "new_clients_share_avg_percent",
        ]
        st.session_state["campaigns_df"] = st.session_state["campaigns_df"].reindex(
            columns=[c for c in campaign_cols_order if c in st.session_state["campaigns_df"].columns]
        )

        campaigns_column_config = {
            "campaign_type": st.column_config.TextColumn("РќР°Р·РІР°РЅРёРµ РєР°РјРїР°РЅРёРё"),
            "system": st.column_config.TextColumn("Р РµРєР»Р°РјРЅР°СЏ СЃРёСЃС‚РµРјР°"),
            "format": st.column_config.TextColumn("Р¤РѕСЂРјР°С‚/С‚Р°СЂРіРµС‚РёРЅРіРё"),
            "geo": st.column_config.TextColumn("Р“Р•Рћ"),
            "impressions_avg": st.column_config.NumberColumn(
                "РџРѕРєР°Р·С‹ (СЃСЂРµРґРЅРёР№ РјРµСЃСЏС†)", format="%.0f", help=mhelp("impressions")
            ),
            "ctr_avg_percent": st.column_config.NumberColumn(
                "CTR, % (СЃСЂРµРґРЅРёР№ РјРµСЃСЏС†)", format="%.2f", help=mhelp("ctr")
            ),
            "cpc_avg": st.column_config.NumberColumn(
                "CPC, в‚Ѕ (СЃСЂРµРґРЅРёР№ РјРµСЃСЏС†)", format="%.2f", help=mhelp("cpc")
            ),
            "cr_avg_percent": st.column_config.NumberColumn("CR1, % (средний месяц)" if is_diy_preset else "CR в ЦО, % (средний месяц)" if metric_mode["is_real_estate_simple"] else "CR1 в Лид, % (средний месяц)" if metric_mode["is_real_estate_full"] else "CR, % (средний месяц)", format="%.2f", help=mhelp("cr")),
            "cr2_avg_percent": st.column_config.NumberColumn("CR2, % (средний месяц)", format="%.2f", help="Отгруженные заказы = оформленные заказы × CR2."),
            "reach_avg": st.column_config.NumberColumn(
                "РћС…РІР°С‚ (СЃСЂРµРґРЅРёР№ РјРµСЃСЏС†)", format="%.0f", help=mhelp("reach")
            ),
        }
        campaigns_editor_columns = [
            "campaign_type",
            "system",
            "format",
            "impressions_avg",
            "ctr_avg_percent",
            "cpc_avg",
            "cr_avg_percent",
        ]
        if is_real_estate_preset or is_diy_preset:
            campaigns_editor_columns.insert(3, "geo")
        if metric_mode["is_real_estate_full"] or is_diy_preset:
            campaigns_editor_columns.append("cr2_avg_percent")
        elif metric_mode["needs_aov"]:
            campaigns_editor_columns.append("aov_avg")
        if is_diy_preset:
            campaigns_column_config["segment"] = st.column_config.SelectboxColumn("РЎРµРіРјРµРЅС‚", options=["B2B", "B2C"])
            campaigns_editor_columns.insert(4, "segment")
            campaigns_editor_columns.append("reach_avg")
            campaigns_column_config["available_capacity_avg"] = st.column_config.NumberColumn(
                "Р”РѕСЃС‚СѓРїРЅР°СЏ РµРјРєРѕСЃС‚СЊ (Р±Р°Р·Р°)", format="%.0f"
            )
            campaigns_column_config["client_count_avg"] = st.column_config.NumberColumn(
                "РљРѕР»РёС‡РµСЃС‚РІРѕ РєР»РёРµРЅС‚РѕРІ (Р±Р°Р·Р°)", format="%.0f", help=mhelp("client_count")
            )
            campaigns_column_config["absolute_new_clients_avg"] = st.column_config.NumberColumn(
                "РљРѕР»-РІРѕ Р°Р±СЃРѕР»СЋС‚РЅРѕ РЅРѕРІС‹С… РєР»РёРµРЅС‚РѕРІ (Р±Р°Р·Р°)", format="%.0f", help=mhelp("absolute_new_clients")
            )
            campaigns_column_config["returned_clients_avg"] = st.column_config.NumberColumn(
                "РљРѕР»-РІРѕ РІРµСЂРЅСѓРІС€РёС…СЃСЏ РєР»РёРµРЅС‚РѕРІ (Р±Р°Р·Р°)", format="%.0f", help=mhelp("returned_clients")
            )
            campaigns_column_config["order_frequency_avg"] = st.column_config.NumberColumn(
                "Р§Р°СЃС‚РѕС‚Р° Р·Р°РєР°Р·Р° (Р±Р°Р·Р°)", format="%.2f", help=mhelp("order_frequency")
            )
            campaigns_column_config["new_clients_share_avg_percent"] = st.column_config.NumberColumn(
                "Р”РѕР»СЏ РЅРѕРІС‹С… РєР»РёРµРЅС‚РѕРІ, % (Р±Р°Р·Р°)", format="%.2f"
            )
            campaigns_editor_columns += ["available_capacity_avg", "client_count_avg", "absolute_new_clients_avg", "returned_clients_avg", "order_frequency_avg", "new_clients_share_avg_percent"]

        with st.form("campaigns_editor_form"):
            campaigns_draft = st.data_editor(
                st.session_state["campaigns_df"],
                num_rows="dynamic",
                use_container_width=True,
                row_height=ui_editor_row_height,
                column_order=campaigns_editor_columns,
                column_config=campaigns_column_config,
                key="campaigns_editor",
            )
            if _has_unsaved_changes(campaigns_draft, st.session_state["campaigns_df"]):
                st.caption("Р•СЃС‚СЊ РЅРµСЃРѕС…СЂР°РЅРµРЅРЅС‹Рµ РёР·РјРµРЅРµРЅРёСЏ РІ С‚Р°Р±Р»РёС†Рµ Р Рљ.")
            camp_btn_c1, camp_btn_c2 = st.columns([1.2, 2.8])
            with camp_btn_c1:
                campaigns_saved = st.form_submit_button("РЎРѕС…СЂР°РЅРёС‚СЊ С‚Р°Р±Р»РёС†Сѓ Р Рљ", type="primary")
            with camp_btn_c2:
                campaign_row_added = st.form_submit_button("вћ• Р”РѕР±Р°РІРёС‚СЊ С‚РёРї Р Рљ")

        if campaigns_saved:
            st.session_state["campaigns_df"] = campaigns_draft.copy()

        if campaign_row_added:
            new_row = {
                "campaign_type": "",
                "segment": "B2C",
                "system": "",
                "format": "",
                "geo": "",
                "impressions_avg": 0,
                "ctr_avg_percent": 0.0,
                "cpc_avg": 0.0,
                "cr_avg_percent": 0.0,
                "cr2_avg_percent": 0.0,
                "aov_avg": 0.0,
                "reach_avg": 0.0,
                "available_capacity_avg": 0.0,
                "client_count_avg": 0.0,
                "absolute_new_clients_avg": 0.0,
                "returned_clients_avg": 0.0,
                "order_frequency_avg": 0.0,
                "new_clients_share_avg_percent": 0.0,
            }
            st.session_state["campaigns_df"] = pd.concat(
                [campaigns_draft, pd.DataFrame([new_row])],
                ignore_index=True,
            )
            st.rerun()

        campaigns = st.session_state["campaigns_df"].copy()
        campaigns = campaigns.dropna(subset=["campaign_type"])
        campaigns = campaigns[campaigns["campaign_type"].astype(str).str.strip() != ""]

        if campaigns.empty:
            st.warning("Р”РѕР±Р°РІСЊС‚Рµ С…РѕС‚СЏ Р±С‹ РѕРґРёРЅ С‚РёРї Р Рљ.")
            plan_calc_blockers.append("РЅРµ Р·Р°РїРѕР»РЅРµРЅС‹ С‚РёРїС‹ Р Рљ")

        required_cols = get_campaign_required_cols(metric_mode)

        def row_has_missing(row):
            for col in required_cols:
                val = row.get(col)
                if val is None:
                    return True
                if isinstance(val, str) and val.strip() == "":
                    return True
                try:
                    if pd.isna(val):
                        return True
                except Exception:
                    pass
            return False

        missing_mask = campaigns.apply(row_has_missing, axis=1)

        if missing_mask.any():
            required_labels = ["РџРѕРєР°Р·С‹", "CTR", "CPC"]
            if metric_mode["is_real_estate_full"]:
                required_labels += ["CR1 РІ Р›РёРґ", "CR2 РІ Р¦Рћ"]
            elif metric_mode["is_real_estate_simple"]:
                required_labels += ["CR РІ Р¦Рћ"]
            elif is_diy_preset:
                required_labels += ["CR", "AOV", "РћС…РІР°С‚"]
            else:
                required_labels += ["CR", "AOV"]
            st.error(
                "Р—Р°РїРѕР»РЅРёС‚Рµ, РїРѕР¶Р°Р»СѓР№СЃС‚Р°, РІСЃРµ РїРѕР»СЏ РІ СЃСЂРµРґРЅРµРј РјРµСЃСЏС†Рµ ("
                + ", ".join(required_labels)
                + ") "
                "РґР»СЏ РєР°Р¶РґРѕРіРѕ С‚РёРїР° Р Рљ. РџСЂРѕРІРµСЂСЊ СЃС‚СЂРѕРєРё, РіРґРµ РµСЃС‚СЊ РїСѓСЃС‚С‹Рµ СЏС‡РµР№РєРё."
            )
            plan_calc_blockers.append("РЅРµ Р·Р°РїРѕР»РЅРµРЅС‹ РѕР±СЏР·Р°С‚РµР»СЊРЅС‹Рµ РїРѕР»СЏ РІ С‚Р°Р±Р»РёС†Рµ Р Рљ")

        segment_filter_options = ["Р’СЃРµ", "B2B", "B2C"]
        if "plan_segment_filter" not in st.session_state:
            st.session_state["plan_segment_filter"] = "Р’СЃРµ"
        if "plan_segment_filter_top" not in st.session_state:
            st.session_state["plan_segment_filter_top"] = st.session_state["plan_segment_filter"]
        if "plan_segment_filter_sidebar" not in st.session_state:
            st.session_state["plan_segment_filter_sidebar"] = st.session_state["plan_segment_filter"]

        def _sync_segment_from_top():
            val = str(st.session_state.get("plan_segment_filter_top", "Р’СЃРµ"))
            if val not in segment_filter_options:
                val = "Р’СЃРµ"
            st.session_state["plan_segment_filter"] = val
            st.session_state["plan_segment_filter_sidebar"] = val

        def _sync_segment_from_sidebar():
            val = str(st.session_state.get("plan_segment_filter_sidebar", "Р’СЃРµ"))
            if val not in segment_filter_options:
                val = "Р’СЃРµ"
            st.session_state["plan_segment_filter"] = val
            st.session_state["plan_segment_filter_top"] = val

        if is_diy_preset:
            st.radio(
                "РџРѕРєР°Р·С‹РІР°С‚СЊ РєР°РјРїР°РЅРёРё СЃРµРіРјРµРЅС‚Р°:",
                options=segment_filter_options,
                horizontal=True,
                key="plan_segment_filter_top",
                on_change=_sync_segment_from_top,
            )
            segment_filter = str(st.session_state.get("plan_segment_filter", "Р’СЃРµ"))
            if segment_filter != "Р’СЃРµ":
                campaigns = campaigns[campaigns["segment"].astype(str).str.upper() == segment_filter]
                if campaigns.empty:
                    st.warning(f"Р”Р»СЏ СЃРµРіРјРµРЅС‚Р° {segment_filter} РЅРµС‚ РєР°РјРїР°РЅРёР№.")
                    plan_calc_blockers.append(f"РґР»СЏ СЃРµРіРјРµРЅС‚Р° {segment_filter} РЅРµС‚ РєР°РјРїР°РЅРёР№ РґР»СЏ СЂР°СЃС‡С‘С‚Р°")
            visible_segments = sorted(campaigns["segment"].astype(str).str.upper().unique().tolist())
            show_segment_subtotals = len(visible_segments) > 1
        else:
            st.session_state["plan_segment_filter"] = "Р’СЃРµ"
            st.session_state["plan_segment_filter_top"] = "Р’СЃРµ"
            st.session_state["plan_segment_filter_sidebar"] = "Р’СЃРµ"
            show_segment_subtotals = False

        existing_ctypes = campaigns["campaign_type"].astype(str).unique().tolist()

    with st.expander("1.2 РќР”РЎ Рё РђРљ", expanded=True):
        vat_warning_text = (
            "<b>РЈС‡РёС‚С‹РІР°С‚СЊ РќР”РЎ 22% РІ СЂР°СЃС‡РµС‚Р°С…</b>: РЅР°СЃС‚СЂРѕР№РєР° РІР»РёСЏРµС‚ РЅР° CPC, CPM, CPL Рё CPQL."
            if is_real_estate_preset and metric_mode["is_real_estate_full"]
            else "<b>РЈС‡РёС‚С‹РІР°С‚СЊ РќР”РЎ 22% РІ СЂР°СЃС‡РµС‚Р°С…</b>: РЅР°СЃС‚СЂРѕР№РєР° РІР»РёСЏРµС‚ РЅР° CPC, CPM Рё CPQL."
            if is_real_estate_preset
            else "<b>РЈС‡РёС‚С‹РІР°С‚СЊ РќР”РЎ 22% РІ СЂР°СЃС‡РµС‚Р°С…</b>: РЅР°СЃС‚СЂРѕР№РєР° РІР»РёСЏРµС‚ РЅР° CPC, CPM, CPO, ROAS Рё Р”Р Р ."
        )
        st.markdown(
            f"""
            <div style="
                margin: 10px 0 8px 0;
                display: inline-block;
                width: fit-content;
                max-width: 100%;
                padding: 10px 12px;
                border-radius: 10px;
                border: 1px solid #FF8A66;
                background: rgba(255, 99, 51, 0.14);
                color: #FFD9CC;
                font-weight: 600;
            ">
                {vat_warning_text}
            </div>
            """,
            unsafe_allow_html=True,
        )
        use_vat_budget_metrics = st.checkbox(
                "РЈС‡РёС‚С‹РІР°С‚СЊ РќР”РЎ 22% РІ СЂР°СЃС‡РµС‚Р°С…",
            key="use_vat_budget_metrics",
        )
        use_ak_budget_metrics = st.checkbox(
            "РЈС‡РёС‚С‹РІР°С‚СЊ РђРљ РІ СЂР°СЃС‡РµС‚Р°С…",
            key="use_ak_budget_metrics",
        )
        vat_ak_ui_config = get_vat_ak_ui_config(metric_mode, is_real_estate_preset)
        vat_warning_text = vat_ak_ui_config["warning_text"]
        ak_mode_labels = vat_ak_ui_config["ak_mode_labels"]
        mode_label_by_value = vat_ak_ui_config["mode_label_by_value"]
        mode_value_by_label = vat_ak_ui_config["mode_value_by_label"]
        ak_mode_label = st.selectbox(
            "РЎРїРѕСЃРѕР± СЂР°СЃС‡РµС‚Р° РђРљ",
            options=ak_mode_labels,
            index=ak_mode_labels.index(mode_label_by_value.get(ak_mode, ak_mode_labels[0])),
            key="ak_mode_select",
        )
        ak_mode = mode_value_by_label.get(ak_mode_label, "percent")
        st.session_state["ak_mode"] = ak_mode
        if ak_mode == "fixed":
            ak_fixed_month_wo_vat = st.number_input(
                "РђРљ РІ РјРµСЃСЏС† (Р±РµР· РќР”РЎ), в‚Ѕ",
                min_value=0.0,
                step=10000.0,
                format="%.0f",
                key="ak_fixed_month_wo_vat",
            )
            st.caption("Р¤РёРєСЃРёСЂРѕРІР°РЅРЅР°СЏ РђРљ РјРµСЃСЏС†Р° СЂР°СЃРїСЂРµРґРµР»СЏРµС‚СЃСЏ РјРµР¶РґСѓ С‚РёРїР°РјРё Р Рљ РїСЂРѕРїРѕСЂС†РёРѕРЅР°Р»СЊРЅРѕ РёС… Р±СЋРґР¶РµС‚Сѓ Р±РµР· РќР”РЎ.")
        elif ak_mode == "fixed_percent":
            ak_fixed_percent = st.number_input(
                "РђРљ, % РґР»СЏ РІСЃРµС… С‚РёРїРѕРІ Р Рљ",
                min_value=0.0,
                step=0.5,
                format="%.2f",
                key="ak_fixed_percent",
            )
            st.caption("РћРґРёРЅ Рё С‚РѕС‚ Р¶Рµ РїСЂРѕС†РµРЅС‚ РђРљ РїСЂРёРјРµРЅСЏРµС‚СЃСЏ РєРѕ РІСЃРµРј С‚РёРїР°Рј Р Рљ Рё РјРµСЃСЏС†Р°Рј РІ СЂР°СЃС‡С‘С‚Рµ.")
        if ak_mode == "percent":
            with st.expander("РќР°СЃС‚СЂРѕР№РєР° РђРљ РїРѕ С€РєР°Р»Рµ (TOTAL Р±СЋРґР¶РµС‚Р° РјРµСЃСЏС†Р° Р±РµР· РќР”РЎ)", expanded=use_ak_budget_metrics):
                st.caption(
                    "Р”Р»СЏ СЂРµР¶РёРјР° В«РџСЂРѕС†РµРЅС‚ РїРѕ С€РєР°Р»РµВ»: РђРљ РІС‹Р±РёСЂР°РµС‚СЃСЏ РїРѕ РЅР°РёР±РѕР»СЊС€РµРјСѓ РїРѕРґС…РѕРґСЏС‰РµРјСѓ РїРѕСЂРѕРіСѓ TOTAL Р±СЋРґР¶РµС‚Р° РјРµСЃСЏС†Р° Р±РµР· РќР”РЎ."
                )
                ak_editor_key = f"ak_rules_editor_{int(st.session_state.get('ak_rules_editor_nonce', 0))}"
                scale_col, _spacer_col = st.columns([1, 2], vertical_alignment="top")
                with scale_col:
                    ak_rules_draft = st.data_editor(
                        st.session_state["ak_rules_df"],
                        num_rows="dynamic",
                        use_container_width=True,
                        row_height=ui_editor_row_height,
                        column_config={
                            "min_budget_wo_vat": st.column_config.NumberColumn(
                                "Р‘СЋРґР¶РµС‚ РѕС‚ (Р±РµР· РќР”РЎ), в‚Ѕ",
                                min_value=0.0,
                                step=10000.0,
                                format="%.0f в‚Ѕ",
                            ),
                            "ak_percent": st.column_config.NumberColumn(
                                "РђРљ, %",
                                min_value=0.0,
                                step=0.5,
                                format="%.2f%%",
                            ),
                        },
                        key=ak_editor_key,
                    )
                if isinstance(ak_rules_draft, pd.DataFrame):
                    ak_rules_clean = normalize_ak_rules_df(ak_rules_draft)
                    st.session_state["ak_rules_df"] = ak_rules_clean
                    ak_rules_df = ak_rules_clean
        vat_ak_apply_state = get_vat_ak_apply_state(
            use_vat_budget_metrics=use_vat_budget_metrics,
            use_ak_budget_metrics=use_ak_budget_metrics,
            ak_mode=ak_mode,
            ak_fixed_month_wo_vat=ak_fixed_month_wo_vat,
            ak_fixed_percent=ak_fixed_percent,
            ak_rules_df=st.session_state.get("ak_rules_df"),
            last_applied_sig=st.session_state.get("vat_ak_last_applied_sig"),
        )
        current_vat_ak_sig = vat_ak_apply_state["current_sig"]
        last_applied_vat_ak_sig = vat_ak_apply_state["last_applied_sig"]
        vat_ak_dirty = vat_ak_apply_state["dirty"]
        apply_vat_ak_clicked = st.button(
            "РџСЂРёРјРµРЅРёС‚СЊ РЅР°СЃС‚СЂРѕР№РєРё РќР”РЎ Рё РђРљ",
            type="primary" if vat_ak_dirty else "secondary",
            key="apply_vat_ak_settings_btn",
        )
        if apply_vat_ak_clicked:
            st.session_state["vat_ak_last_applied_sig"] = current_vat_ak_sig
            if vat_ak_dirty:
                st.success("РќР°СЃС‚СЂРѕР№РєРё РќР”РЎ Рё РђРљ РїСЂРёРјРµРЅРµРЅС‹.")
            else:
                st.info("РР·РјРµРЅРµРЅРёР№ РЅРµС‚: С‚РµРєСѓС‰РёРµ РЅР°СЃС‚СЂРѕР№РєРё СѓР¶Рµ РїСЂРёРјРµРЅРµРЅС‹.")
        st.caption(
            "CPC РІСЃРµРіРґР° СЃС‡РёС‚Р°РµС‚СЃСЏ РѕС‚ Р±СЋРґР¶РµС‚Р° Р±РµР· РќР”РЎ. РџСЂРё РІС‹РєР»СЋС‡РµРЅРЅРѕРј СЂРµР¶РёРјРµ РќР”РЎ СЂР°СЃС‡РµС‚С‹ РІРѕР·РІСЂР°С‰Р°СЋС‚СЃСЏ Рє СЃС‚Р°СЂРѕР№ Р»РѕРіРёРєРµ."
        )

    # ---------- 1.3. РќР°Р±РѕСЂС‹ РєРѕСЌС„С„РёС†РёРµРЅС‚РѕРІ РґР»СЏ С‚РёРїРѕРІ Р Рљ ----------

    with st.expander("1.3. РќР°Р±РѕСЂС‹ РєРѕСЌС„С„РёС†РёРµРЅС‚РѕРІ РґР»СЏ С‚РёРїРѕРІ Р Рљ", expanded=True):
        st.markdown(
            """
            <div style="
                margin: 8px 0 14px 0;
                display: inline-block;
                width: fit-content;
                max-width: 100%;
                white-space: nowrap;
                padding: 10px 12px;
                border-radius: 10px;
                border: 1px solid #FF8A66;
                background: rgba(255, 99, 51, 0.14);
                color: #FFD9CC;
                font-weight: 600;
            ">
                Р’Р°Р¶РЅРѕ: Р±РµР· РЅР°Р·РЅР°С‡РµРЅРЅС‹С… РЅР°Р±РѕСЂРѕРІ СЂР°СЃС‡РµС‚ РЅРµ РїРѕР№РґРµС‚.
            </div>
            """,
            unsafe_allow_html=True,
        )

        coeff_sets = st.session_state.get("coeff_sets", [])
        coeff_sets_by_id = {
            get_coeff_set_id_value(cs): cs
            for cs in coeff_sets
            if isinstance(cs, dict) and get_coeff_set_id_value(cs)
        }

        demand_set_ids = [
            get_coeff_set_id_value(cs) for cs in coeff_sets
            if normalize_coeff_set_type(cs.get("type")) in ["РЎРїСЂРѕСЃ (РїРѕ Р·Р°РїСЂРѕСЃР°Рј)", "РљР°СЃС‚РѕРјРЅС‹Р№ РЅР°Р±РѕСЂ"]
            and get_coeff_set_id_value(cs)
        ]
        aov_set_ids = [
            get_coeff_set_id_value(cs) for cs in coeff_sets
            if normalize_coeff_set_type(cs.get("type")) in ["AOV (СЃСЂРµРґРЅРёР№ С‡РµРє)", "РљР°СЃС‚РѕРјРЅС‹Р№ РЅР°Р±РѕСЂ"]
            and get_coeff_set_id_value(cs)
        ]
        media_tail_set_ids = [
            get_coeff_set_id_value(cs) for cs in coeff_sets
            if normalize_coeff_set_type(cs.get("type")) == "РњРµРґРёР№РЅС‹Рµ С…РІРѕСЃС‚С‹"
            and get_coeff_set_id_value(cs)
        ]
        capacity_set_ids = [get_coeff_set_id_value(cs) for cs in coeff_sets if get_coeff_set_id_value(cs)]
        valid_demand_set_ids = set(demand_set_ids)
        valid_aov_set_ids = set(aov_set_ids)
        valid_media_tail_set_ids = set(media_tail_set_ids)
        valid_capacity_set_ids = set(capacity_set_ids)
        demand_set_labels = [get_coeff_set_label(coeff_sets_by_id[set_id]) for set_id in demand_set_ids if set_id in coeff_sets_by_id]
        aov_set_labels = [get_coeff_set_label(coeff_sets_by_id[set_id]) for set_id in aov_set_ids if set_id in coeff_sets_by_id]
        media_tail_set_labels = [get_coeff_set_label(coeff_sets_by_id[set_id]) for set_id in media_tail_set_ids if set_id in coeff_sets_by_id]
        capacity_set_labels = [get_coeff_set_label(coeff_sets_by_id[set_id]) for set_id in capacity_set_ids if set_id in coeff_sets_by_id]

        if "coeff_sets_links_new" not in st.session_state:
            st.session_state["coeff_sets_links_new"] = pd.DataFrame(
                columns=["campaign_type", "demand_set", "aov_set", "media_tail_set", "capacity_set", "client_count_set", "absolute_new_clients_set", "returned_clients_set", "order_frequency_set"]
            )

        coeff_links_prev = st.session_state["coeff_sets_links_new"].copy()
        prev_demand_map = {}
        prev_aov_map = {}
        prev_media_tail_map = {}
        prev_capacity_map = {}
        prev_client_count_map = {}
        prev_absolute_new_clients_map = {}
        prev_returned_clients_map = {}
        prev_order_frequency_map = {}
        for _, r in coeff_links_prev.iterrows():
            ct = str(r.get("campaign_type", "")).strip()
            if not ct:
                continue
            if ct not in prev_demand_map:
                prev_demand_map[ct] = normalize_coeff_link_value(r.get("demand_set", ""), coeff_sets, valid_demand_set_ids)
            if ct not in prev_aov_map:
                prev_aov_map[ct] = normalize_coeff_link_value(r.get("aov_set", ""), coeff_sets, valid_aov_set_ids)
            if ct not in prev_media_tail_map:
                prev_media_tail_map[ct] = normalize_coeff_link_value(r.get("media_tail_set", ""), coeff_sets, valid_media_tail_set_ids)
            if ct not in prev_capacity_map:
                prev_capacity_map[ct] = normalize_coeff_link_value(r.get("capacity_set", ""), coeff_sets, valid_capacity_set_ids)
            if ct not in prev_client_count_map:
                prev_client_count_map[ct] = normalize_coeff_link_value(r.get("client_count_set", ""), coeff_sets, valid_capacity_set_ids)
            if ct not in prev_absolute_new_clients_map:
                prev_absolute_new_clients_map[ct] = normalize_coeff_link_value(r.get("absolute_new_clients_set", ""), coeff_sets, valid_capacity_set_ids)
            if ct not in prev_returned_clients_map:
                prev_returned_clients_map[ct] = normalize_coeff_link_value(r.get("returned_clients_set", ""), coeff_sets, valid_capacity_set_ids)
            if ct not in prev_order_frequency_map:
                prev_order_frequency_map[ct] = normalize_coeff_link_value(r.get("order_frequency_set", ""), coeff_sets, valid_capacity_set_ids)

        coeff_links_new = pd.DataFrame(
            {
                "campaign_type": existing_ctypes,
                "demand_set": [prev_demand_map.get(ct, "") for ct in existing_ctypes],
                "aov_set": [prev_aov_map.get(ct, "") for ct in existing_ctypes],
                "media_tail_set": [prev_media_tail_map.get(ct, "") for ct in existing_ctypes],
                "capacity_set": [prev_capacity_map.get(ct, prev_demand_map.get(ct, "")) for ct in existing_ctypes],
                "client_count_set": [prev_client_count_map.get(ct, prev_capacity_map.get(ct, prev_demand_map.get(ct, ""))) for ct in existing_ctypes],
                "absolute_new_clients_set": [prev_absolute_new_clients_map.get(ct, prev_client_count_map.get(ct, prev_capacity_map.get(ct, prev_demand_map.get(ct, "")))) for ct in existing_ctypes],
                "returned_clients_set": [prev_returned_clients_map.get(ct, prev_client_count_map.get(ct, prev_capacity_map.get(ct, prev_demand_map.get(ct, "")))) for ct in existing_ctypes],
                "order_frequency_set": [prev_order_frequency_map.get(ct, prev_client_count_map.get(ct, prev_capacity_map.get(ct, prev_demand_map.get(ct, "")))) for ct in existing_ctypes],
            }
        )
        if not is_diy_preset:
            coeff_links_new = coeff_links_new.drop(columns=["capacity_set", "client_count_set", "absolute_new_clients_set", "returned_clients_set", "order_frequency_set"], errors="ignore")
        if not metric_mode["needs_aov"]:
            coeff_links_new = coeff_links_new.drop(columns=["aov_set"], errors="ignore")

        dangling_links_detected = False
        if "demand_set" in coeff_links_new.columns:
            invalid_mask = coeff_links_new["demand_set"].astype(str).str.strip().ne("") & ~coeff_links_new["demand_set"].isin(valid_demand_set_ids)
            if invalid_mask.any():
                coeff_links_new.loc[invalid_mask, "demand_set"] = ""
                dangling_links_detected = True
        if "aov_set" in coeff_links_new.columns:
            invalid_mask = coeff_links_new["aov_set"].astype(str).str.strip().ne("") & ~coeff_links_new["aov_set"].isin(valid_aov_set_ids)
            if invalid_mask.any():
                coeff_links_new.loc[invalid_mask, "aov_set"] = ""
                dangling_links_detected = True
        if "media_tail_set" in coeff_links_new.columns:
            invalid_mask = coeff_links_new["media_tail_set"].astype(str).str.strip().ne("") & ~coeff_links_new["media_tail_set"].isin(valid_media_tail_set_ids)
            if invalid_mask.any():
                coeff_links_new.loc[invalid_mask, "media_tail_set"] = ""
                dangling_links_detected = True
        if "capacity_set" in coeff_links_new.columns:
            invalid_mask = coeff_links_new["capacity_set"].astype(str).str.strip().ne("") & ~coeff_links_new["capacity_set"].isin(valid_capacity_set_ids)
            if invalid_mask.any():
                coeff_links_new.loc[invalid_mask, "capacity_set"] = ""
                dangling_links_detected = True
        if "client_count_set" in coeff_links_new.columns:
            invalid_mask = coeff_links_new["client_count_set"].astype(str).str.strip().ne("") & ~coeff_links_new["client_count_set"].isin(valid_capacity_set_ids)
            if invalid_mask.any():
                coeff_links_new.loc[invalid_mask, "client_count_set"] = ""
                dangling_links_detected = True
        if "absolute_new_clients_set" in coeff_links_new.columns:
            invalid_mask = coeff_links_new["absolute_new_clients_set"].astype(str).str.strip().ne("") & ~coeff_links_new["absolute_new_clients_set"].isin(valid_capacity_set_ids)
            if invalid_mask.any():
                coeff_links_new.loc[invalid_mask, "absolute_new_clients_set"] = ""
                dangling_links_detected = True
        if "returned_clients_set" in coeff_links_new.columns:
            invalid_mask = coeff_links_new["returned_clients_set"].astype(str).str.strip().ne("") & ~coeff_links_new["returned_clients_set"].isin(valid_capacity_set_ids)
            if invalid_mask.any():
                coeff_links_new.loc[invalid_mask, "returned_clients_set"] = ""
                dangling_links_detected = True
        if "order_frequency_set" in coeff_links_new.columns:
            invalid_mask = coeff_links_new["order_frequency_set"].astype(str).str.strip().ne("") & ~coeff_links_new["order_frequency_set"].isin(valid_capacity_set_ids)
            if invalid_mask.any():
                coeff_links_new.loc[invalid_mask, "order_frequency_set"] = ""
                dangling_links_detected = True
        if dangling_links_detected:
            st.session_state["coeff_sets_links_new"] = coeff_links_new.copy()

        coeff_links_editor = coeff_links_new.copy()
        coeff_link_allowed_ids = {
            "demand_set": valid_demand_set_ids,
            "aov_set": valid_aov_set_ids,
            "media_tail_set": valid_media_tail_set_ids,
            "capacity_set": valid_capacity_set_ids,
            "client_count_set": valid_capacity_set_ids,
            "absolute_new_clients_set": valid_capacity_set_ids,
            "returned_clients_set": valid_capacity_set_ids,
            "order_frequency_set": valid_capacity_set_ids,
        }
        for col_name, allowed_ids in coeff_link_allowed_ids.items():
            if col_name not in coeff_links_editor.columns:
                continue
            coeff_links_editor[col_name] = coeff_links_editor[col_name].map(
                lambda raw_value: get_coeff_set_label(coeff_sets_by_id[raw_value])
                if str(raw_value).strip() in allowed_ids and str(raw_value).strip() in coeff_sets_by_id
                else ""
            )

        if metric_mode["needs_aov"]:
            st.markdown(
                "Р”Р»СЏ РєР°Р¶РґРѕРіРѕ **С‚РёРїР° Р Рљ** РІС‹Р±РµСЂРёС‚Рµ РЅР°Р±РѕСЂ СЃРµР·РѕРЅРЅРѕСЃС‚Рё **СЃРїСЂРѕСЃР°** Рё РЅР°Р±РѕСЂ СЃРµР·РѕРЅРЅРѕСЃС‚Рё **AOV (СЃСЂРµРґРЅРёР№ С‡РµРє)**."
            )
        else:
            st.markdown(
                "Р”Р»СЏ РєР°Р¶РґРѕРіРѕ **С‚РёРїР° Р Рљ** РІС‹Р±РµСЂРёС‚Рµ РЅР°Р±РѕСЂ СЃРµР·РѕРЅРЅРѕСЃС‚Рё **СЃРїСЂРѕСЃР°**. Р”Р»СЏ РїСЂРµСЃРµС‚Р° В«РќРµРґРІРёР¶РёРјРѕСЃС‚СЊВ» РЅР°Р±РѕСЂ AOV РЅРµ РёСЃРїРѕР»СЊР·СѓРµС‚СЃСЏ."
            )

        with st.form("coeff_sets_links_form"):
            bulk_cols = [2, 2, 1.2] if metric_mode["needs_aov"] else [2, 1.2]
            bulk_columns = st.columns(bulk_cols, vertical_alignment="bottom")
            bulk_c1 = bulk_columns[0]
            bulk_c2 = bulk_columns[1] if metric_mode["needs_aov"] else None
            bulk_c3 = bulk_columns[-1]
            with bulk_c1:
                bulk_demand_set = st.selectbox(
                    "РќР°Р±РѕСЂ СЃРїСЂРѕСЃР° РґР»СЏ РІСЃРµС… Р Рљ",
                    options=[""] + demand_set_labels,
                    help="Р‘С‹СЃС‚СЂРѕ РЅР°Р·РЅР°С‡Р°РµС‚ РѕРґРёРЅ РЅР°Р±РѕСЂ СЃРµР·РѕРЅРЅРѕСЃС‚Рё СЃРїСЂРѕСЃР° РІСЃРµРј С‚РёРїР°Рј Р Рљ РІ С‚Р°Р±Р»РёС†Рµ РЅРёР¶Рµ.",
                    key="bulk_demand_set_for_all",
                )
            bulk_aov_set = ""
            if metric_mode["needs_aov"] and bulk_c2 is not None:
                with bulk_c2:
                    bulk_aov_set = st.selectbox(
                        "РќР°Р±РѕСЂ AOV РґР»СЏ РІСЃРµС… Р Рљ",
                        options=[""] + aov_set_labels,
                        help="Р‘С‹СЃС‚СЂРѕ РЅР°Р·РЅР°С‡Р°РµС‚ РѕРґРёРЅ РЅР°Р±РѕСЂ СЃРµР·РѕРЅРЅРѕСЃС‚Рё AOV РІСЃРµРј С‚РёРїР°Рј Р Рљ РІ С‚Р°Р±Р»РёС†Рµ РЅРёР¶Рµ.",
                        key="bulk_aov_set_for_all",
                    )
            with bulk_c3:
                apply_all_clicked = st.form_submit_button("РџСЂРёРјРµРЅРёС‚СЊ РєРѕ РІСЃРµРј", type="primary")

            links_column_config = {
                "campaign_type": st.column_config.TextColumn(
                    "РўРёРї Р Рљ",
                    disabled=True,
                ),
                "demand_set": st.column_config.SelectboxColumn(
                    "РќР°Р±РѕСЂ СЃРµР·РѕРЅРЅРѕСЃС‚Рё СЃРїСЂРѕСЃР°",
                    options=demand_set_labels,
                    help="Р’Р»РёСЏРµС‚ РЅР° РїРѕРєР°Р·С‹ (k_imp), Р° С‚Р°РєР¶Рµ С‡РµСЂРµР· СЌР»Р°СЃС‚РёС‡РЅРѕСЃС‚СЊ РЅР° CTR/CPC/CR.",
                ),
                "media_tail_set": st.column_config.SelectboxColumn(
                    "РќР°Р±РѕСЂ РјРµРґРёР№РЅС‹С… С…РІРѕСЃС‚РѕРІ (РїРѕРєР°Р·С‹)",
                    options=[""] + media_tail_set_labels,
                    help="РњРЅРѕР¶РёС‚РµР»СЊ РїРѕРєР°Р·РѕРІ РїРѕРІРµСЂС… СЃРїСЂРѕСЃР°: k_imp = k_demand * k_media_tail.",
                ),
            }
            if metric_mode["needs_aov"]:
                links_column_config["aov_set"] = st.column_config.SelectboxColumn(
                    "РќР°Р±РѕСЂ СЃРµР·РѕРЅРЅРѕСЃС‚Рё AOV",
                    options=aov_set_labels,
                    help="Р’Р»РёСЏРµС‚ РЅР° СЃСЂРµРґРЅРёР№ С‡РµРє: k_aov.",
                )
            if is_diy_preset:
                links_column_config["capacity_set"] = st.column_config.SelectboxColumn(
                    "РќР°Р±РѕСЂ РґР»СЏ РµРјРєРѕСЃС‚Рё",
                    options=capacity_set_labels,
                    help="РљРѕСЌС„С„РёС†РёРµРЅС‚С‹ СЌС‚РѕРіРѕ РЅР°Р±РѕСЂР° РїСЂРёРјРµРЅСЏСЋС‚СЃСЏ Рє РґРѕСЃС‚СѓРїРЅРѕР№ РµРјРєРѕСЃС‚Рё (k_capacity).",
                )
                links_column_config["client_count_set"] = st.column_config.SelectboxColumn(
                    "РќР°Р±РѕСЂ РґР»СЏ РєР»РёРµРЅС‚РѕРІ",
                    options=capacity_set_labels,
                    help="РљРѕСЌС„С„РёС†РёРµРЅС‚С‹ СЌС‚РѕРіРѕ РЅР°Р±РѕСЂР° РїСЂРёРјРµРЅСЏСЋС‚СЃСЏ Рє РєРѕР»РёС‡РµСЃС‚РІСѓ РєР»РёРµРЅС‚РѕРІ (k_client_count).",
                )
                links_column_config["absolute_new_clients_set"] = st.column_config.SelectboxColumn(
                    "РќР°Р±РѕСЂ РґР»СЏ РђРќРљ",
                    options=capacity_set_labels,
                    help="РљРѕСЌС„С„РёС†РёРµРЅС‚С‹ СЌС‚РѕРіРѕ РЅР°Р±РѕСЂР° РїСЂРёРјРµРЅСЏСЋС‚СЃСЏ Рє Р°Р±СЃРѕР»СЋС‚РЅРѕ РЅРѕРІС‹Рј РєР»РёРµРЅС‚Р°Рј.",
                )
                links_column_config["returned_clients_set"] = st.column_config.SelectboxColumn(
                    "РќР°Р±РѕСЂ РґР»СЏ РІРµСЂРЅСѓРІС€РёС…СЃСЏ",
                    options=capacity_set_labels,
                    help="РљРѕСЌС„С„РёС†РёРµРЅС‚С‹ СЌС‚РѕРіРѕ РЅР°Р±РѕСЂР° РїСЂРёРјРµРЅСЏСЋС‚СЃСЏ Рє РІРµСЂРЅСѓРІС€РёРјСЃСЏ РєР»РёРµРЅС‚Р°Рј.",
                )
                links_column_config["order_frequency_set"] = st.column_config.SelectboxColumn(
                    "РќР°Р±РѕСЂ РґР»СЏ С‡Р°СЃС‚РѕС‚С‹ Р·Р°РєР°Р·Р°",
                    options=capacity_set_labels,
                    help="РљРѕСЌС„С„РёС†РёРµРЅС‚С‹ СЌС‚РѕРіРѕ РЅР°Р±РѕСЂР° РїСЂРёРјРµРЅСЏСЋС‚СЃСЏ Рє С‡Р°СЃС‚РѕС‚Рµ Р·Р°РєР°Р·Р°.",
                )

            coeff_links_draft = st.data_editor(
                coeff_links_editor,
                num_rows="fixed",
                use_container_width=True,
                row_height=ui_editor_row_height,
                column_config=links_column_config,
                key="coeff_sets_links_new_editor",
            )
            if _has_unsaved_changes(coeff_links_draft, coeff_links_new):
                st.caption("Р•СЃС‚СЊ РЅРµСЃРѕС…СЂР°РЅРµРЅРЅС‹Рµ РёР·РјРµРЅРµРЅРёСЏ РІ РІС‹Р±РѕСЂРµ РЅР°Р±РѕСЂРѕРІ.")
            links_saved = st.form_submit_button("РЎРѕС…СЂР°РЅРёС‚СЊ РІС‹Р±РѕСЂ РЅР°Р±РѕСЂРѕРІ", type="primary")

        if apply_all_clicked or links_saved:
            coeff_links_new = coeff_links_draft.copy()
            for col_name, allowed_ids in coeff_link_allowed_ids.items():
                if col_name in coeff_links_new.columns:
                    coeff_links_new[col_name] = coeff_links_new[col_name].map(
                        lambda raw_value: normalize_coeff_link_value(raw_value, coeff_sets, allowed_ids)
                    )
            if apply_all_clicked:
                bulk_demand_set_id = normalize_coeff_link_value(bulk_demand_set, coeff_sets, valid_demand_set_ids)
                bulk_aov_set_id = normalize_coeff_link_value(bulk_aov_set, coeff_sets, valid_aov_set_ids)
                if bulk_demand_set_id:
                    coeff_links_new["demand_set"] = bulk_demand_set_id
                if bulk_aov_set_id and "aov_set" in coeff_links_new.columns:
                    coeff_links_new["aov_set"] = bulk_aov_set_id
            st.session_state["coeff_sets_links_new"] = coeff_links_new
            if apply_all_clicked:
                # РџРѕСЃР»Рµ РјР°СЃСЃРѕРІРѕРіРѕ РїСЂРёРјРµРЅРµРЅРёСЏ РїРµСЂРµСЂРёСЃРѕРІС‹РІР°РµРј С„РѕСЂРјСѓ,
                # С‡С‚РѕР±С‹ data_editor РѕС‚РѕР±СЂР°Р·РёР» РЅРѕРІС‹Рµ Р·РЅР°С‡РµРЅРёСЏ РІ СЃС‚СЂРѕРєР°С….
                st.session_state.pop("coeff_sets_links_new_editor", None)
                st.rerun()

        aov_link_map = {}
        demand_link_map = {}
        media_tail_link_map = {}
        capacity_link_map = {}
        client_count_link_map = {}
        absolute_new_clients_link_map = {}
        returned_clients_link_map = {}
        order_frequency_link_map = {}

        for _, r in coeff_links_new.iterrows():
            camp = str(r.get("campaign_type", "")).strip()
            if not camp:
                continue
            demand_link_map[camp] = str(r.get("demand_set", "")).strip()
            aov_link_map[camp] = str(r.get("aov_set", "")).strip()
            media_tail_link_map[camp] = str(r.get("media_tail_set", "")).strip()
            capacity_link_map[camp] = str(r.get("capacity_set", "")).strip()
            client_count_link_map[camp] = str(r.get("client_count_set", "")).strip()
            absolute_new_clients_link_map[camp] = str(r.get("absolute_new_clients_set", "")).strip()
            returned_clients_link_map[camp] = str(r.get("returned_clients_set", "")).strip()
            order_frequency_link_map[camp] = str(r.get("order_frequency_set", "")).strip()

        missing_demand = [ct for ct in existing_ctypes if not demand_link_map.get(ct)]
        if missing_demand:
            st.error(
                "Р”Р»СЏ СЃР»РµРґСѓСЋС‰РёС… С‚РёРїРѕРІ Р Рљ РЅРµ РІС‹Р±СЂР°РЅ РЅР°Р±РѕСЂ СЃРµР·РѕРЅРЅРѕСЃС‚Рё СЃРїСЂРѕСЃР°, РїРѕСЌС‚РѕРјСѓ РјРµРґРёР°РїР»Р°РЅ РґР°Р»СЊС€Рµ СЃС‡РёС‚Р°С‚СЊСЃСЏ РЅРµ Р±СѓРґРµС‚: "
                + ", ".join(missing_demand)
            )
            plan_calc_blockers.append("РЅРµ РІС‹Р±СЂР°РЅС‹ РЅР°Р±РѕСЂС‹ РєРѕСЌС„С„РёС†РёРµРЅС‚РѕРІ СЃРїСЂРѕСЃР°")

        missing_aov = [ct for ct in existing_ctypes if not aov_link_map.get(ct)]
        if metric_mode["needs_aov"] and missing_aov:
            st.error(
                "Р”Р»СЏ СЃР»РµРґСѓСЋС‰РёС… С‚РёРїРѕРІ Р Рљ РЅРµ РІС‹Р±СЂР°РЅ РЅР°Р±РѕСЂ СЃРµР·РѕРЅРЅРѕСЃС‚Рё AOV (СЃСЂРµРґРЅРёР№ С‡РµРє), "
                "РїРѕСЌС‚РѕРјСѓ РјРµРґРёР°РїР»Р°РЅ РґР°Р»СЊС€Рµ СЃС‡РёС‚Р°С‚СЊСЃСЏ РЅРµ Р±СѓРґРµС‚: "
                + ", ".join(missing_aov)
            )
            plan_calc_blockers.append("РЅРµ РІС‹Р±СЂР°РЅС‹ РЅР°Р±РѕСЂС‹ РєРѕСЌС„С„РёС†РёРµРЅС‚РѕРІ AOV")

    # ---------- 1.4. Р­Р»Р°СЃС‚РёС‡РЅРѕСЃС‚СЊ РјРµС‚СЂРёРє Рє СЃРµР·РѕРЅРЅРѕСЃС‚Рё СЃРїСЂРѕСЃР° ----------

    with st.expander("1.4. Р­Р»Р°СЃС‚РёС‡РЅРѕСЃС‚СЊ РјРµС‚СЂРёРє Рє СЃРµР·РѕРЅРЅРѕСЃС‚Рё СЃРїСЂРѕСЃР°", expanded=False):
        show_cr2_elasticity = bool((is_real_estate_preset and metric_mode["is_real_estate_full"]) or is_diy_preset)

        st.caption("РџР°РјСЏС‚РєР°: РјРµРЅСЊС€Рµ РґРµР»РёС‚РµР»СЊ = СЃРёР»СЊРЅРµРµ РІР»РёСЏРЅРёРµ.")
        st.caption(
            "Р—Р°РґР°Р№С‚Рµ РґРµР»РёС‚РµР»Рё РІР»РёСЏРЅРёСЏ СЃРµР·РѕРЅРЅРѕСЃС‚Рё СЃРїСЂРѕСЃР° РЅР° РјРµС‚СЂРёРєРё РґР»СЏ РєР°Р¶РґРѕРіРѕ С‚РёРїР° Р Рљ. "
            "Р§РµРј РјРµРЅСЊС€Рµ РґРµР»РёС‚РµР»СЊ, С‚РµРј СЃРёР»СЊРЅРµРµ РІР»РёСЏРЅРёРµ. РџСЂРѕС†РµРЅС‚РЅС‹Р№ СЂРµР·СѓР»СЊС‚Р°С‚ СЃРјРѕС‚СЂРёС‚Рµ РІ РїСЂРµРІСЊСЋ СЃРїСЂР°РІР°."
        )
        st.markdown(
            "**РљР°Рє СЃС‡РёС‚Р°СЋС‚СЃСЏ СЌС„С„РµРєС‚С‹ СЌР»Р°СЃС‚РёС‡РЅРѕСЃС‚Рё:**\n"
            "1. `CPC` РІ СЂРµР¶РёРјРµ СЂРѕСЃС‚Р° CPC: С„РѕСЂРјСѓР»Р° `(k-1)/div + 1`.\n"
            "2. `CTR` РІ СЂРµР¶РёРјРµ РїР°РґРµРЅРёСЏ CTR: С„РѕСЂРјСѓР»Р° `1 - (k-1)/div`.\n"
            "3. `CR` РІ СЂРµР¶РёРјРµ РїР°РґРµРЅРёСЏ CR: С„РѕСЂРјСѓР»Р° `1 - (k-1)/div`.\n"
            + ("4. `CR2` РІ РЅРµРґРІРёР¶РёРјРѕСЃС‚Рё СЃС‡РёС‚Р°РµС‚СЃСЏ РѕС‚РґРµР»СЊРЅРѕ С‚РѕР»СЊРєРѕ РґР»СЏ РїРѕР»РЅРѕР№ РІРѕСЂРѕРЅРєРё.\n" if show_cr2_elasticity else "")
            + "РџСЂРёРјРµСЂ: РґР»СЏ `k=1.10` Рё `CPC div=2` РїРѕР»СѓС‡РёС‚СЃСЏ `+5%` Рє CPC."
        )

        if "elasticity_df" not in st.session_state:
            st.session_state["elasticity_df"] = pd.DataFrame(
                {
                    "campaign_type": existing_ctypes,
                    "cpc_div": [1.0] * len(existing_ctypes),
                    "ctr_div": [2.0] * len(existing_ctypes),
                    "cr_div": [10.0] * len(existing_ctypes),
                    "cr2_div": [10.0] * len(existing_ctypes),
                }
            )

        elasticity_df = st.session_state["elasticity_df"]

        # РњРёРіСЂР°С†РёСЏ СЃС‚Р°СЂРѕРіРѕ С„РѕСЂРјР°С‚Р° РїСЂРѕС†РµРЅС‚РѕРІ/legacy РІ РґРµР»РёС‚РµР»Рё.
        if "cpc_div" not in elasticity_df.columns and "cpc_impact_pct_per_10" in elasticity_df.columns:
            src = pd.to_numeric(elasticity_df["cpc_impact_pct_per_10"], errors="coerce")
            elasticity_df["cpc_div"] = np.where(src > 0, 10.0 / src, 1.0)
        if "ctr_div" not in elasticity_df.columns and "ctr_impact_pct_per_10" in elasticity_df.columns:
            src = pd.to_numeric(elasticity_df["ctr_impact_pct_per_10"], errors="coerce").abs()
            elasticity_df["ctr_div"] = np.where(src > 0, 10.0 / src, 2.0)
        if "cr_div" not in elasticity_df.columns and "cr_impact_pct_per_10" in elasticity_df.columns:
            src = pd.to_numeric(elasticity_df["cr_impact_pct_per_10"], errors="coerce").abs()
            elasticity_df["cr_div"] = np.where(src > 0, 10.0 / src, 10.0)
        if "cpc_div" not in elasticity_df.columns and "cpc_up_pct_per_10" in elasticity_df.columns:
            src = pd.to_numeric(elasticity_df["cpc_up_pct_per_10"], errors="coerce")
            elasticity_df["cpc_div"] = np.where(src > 0, 10.0 / src, 1.0)
        if "ctr_div" not in elasticity_df.columns and "ctr_down_pct_per_10" in elasticity_df.columns:
            src = pd.to_numeric(elasticity_df["ctr_down_pct_per_10"], errors="coerce").abs()
            elasticity_df["ctr_div"] = np.where(src > 0, 10.0 / src, 2.0)
        if "cr_div" not in elasticity_df.columns and "cr_down_pct_per_10" in elasticity_df.columns:
            src = pd.to_numeric(elasticity_df["cr_down_pct_per_10"], errors="coerce").abs()
            elasticity_df["cr_div"] = np.where(src > 0, 10.0 / src, 10.0)
        if "cr2_div" not in elasticity_df.columns:
            elasticity_df["cr2_div"] = pd.to_numeric(elasticity_df.get("cr_div", 10.0), errors="coerce").fillna(10.0)
        for col_name, default_val in [("cpc_div", 1.0), ("ctr_div", 2.0), ("cr_div", 10.0), ("cr2_div", 10.0)]:
            if col_name in elasticity_df.columns:
                elasticity_df[col_name] = pd.to_numeric(elasticity_df[col_name], errors="coerce").fillna(default_val)

        elasticity_df = elasticity_df[
            elasticity_df["campaign_type"].isin(existing_ctypes)
        ]

        keep_cols = ["campaign_type", "preset", "cpc_div", "ctr_div", "cr_div", "cr2_div"]
        for c in keep_cols:
            if c not in elasticity_df.columns:
                if c == "campaign_type":
                    elasticity_df[c] = ""
                elif c == "preset":
                    elasticity_df[c] = "РљР°СЃС‚РѕРј"
                elif c == "cpc_div":
                    elasticity_df[c] = 1.0
                elif c == "ctr_div":
                    elasticity_df[c] = 2.0
                elif c == "cr_div":
                    elasticity_df[c] = 10.0
                elif c == "cr2_div":
                    elasticity_df[c] = 10.0
                else:
                    elasticity_df[c] = 0.0
        elasticity_df = elasticity_df[keep_cols]

        for ct in existing_ctypes:
            if ct not in elasticity_df["campaign_type"].tolist():
                elasticity_df = pd.concat(
                    [
                        elasticity_df,
                        pd.DataFrame(
                            [{
                                "campaign_type": ct,
                                "preset": "РЎСЂРµРґРЅРµРµ",
                                "cpc_div": 1.0,
                                "ctr_div": 2.0,
                                "cr_div": 10.0,
                                "cr2_div": 10.0,
                            }]
                        ),
                    ],
                    ignore_index=True,
                )

        def _style_impact_col(col: pd.Series) -> pd.Series:
            out = pd.Series([""] * len(col), index=col.index)
            vals = pd.to_numeric(col, errors="coerce")
            for idx_v, v in vals.items():
                if pd.isna(v):
                    continue
                if v > 0:
                    out.loc[idx_v] = "color: #00CDC5; font-weight: 700;"
                elif v < 0:
                    out.loc[idx_v] = "color: #FF6333; font-weight: 700;"
                else:
                    out.loc[idx_v] = "color: #D0D6DF;"
            return out

        def _coerce_div_for_preview(val: float, default_val: float) -> float:
            try:
                num = float(val)
                return num if num > 0 else default_val
            except Exception:
                return default_val

        def _row_divs_match_preset(row: pd.Series, preset_vals: dict[str, float], tol: float = 1e-9) -> bool:
            return (
                abs(_coerce_div_for_preview(row.get("cpc_div", 1.0), 1.0) - float(preset_vals["cpc_div"])) <= tol
                and abs(_coerce_div_for_preview(row.get("ctr_div", 2.0), 2.0) - float(preset_vals["ctr_div"])) <= tol
                and abs(_coerce_div_for_preview(row.get("cr_div", 10.0), 10.0) - float(preset_vals["cr_div"])) <= tol
                and (
                    not show_cr2_elasticity
                    or abs(_coerce_div_for_preview(row.get("cr2_div", 10.0), 10.0) - float(preset_vals["cr2_div"])) <= tol
                )
            )

        with st.form("elasticity_editor_form"):
            preset_map = {
                "РЎР»Р°Р±РѕРµ": {"cpc_div": 2.0, "ctr_div": 5.0, "cr_div": 15.0, "cr2_div": 15.0},
                "РЎСЂРµРґРЅРµРµ": {"cpc_div": 1.0, "ctr_div": 2.0, "cr_div": 10.0, "cr2_div": 10.0},
                "РЎРёР»СЊРЅРѕРµ": {"cpc_div": 0.5, "ctr_div": 1.0, "cr_div": 5.0, "cr2_div": 5.0},
            }
            st.caption("Р”Р»СЏ РєР°Р¶РґРѕРіРѕ С‚РёРїР° Р Рљ РјРѕР¶РЅРѕ РІС‹Р±СЂР°С‚СЊ РїСЂРµСЃРµС‚ РІ РєРѕР»РѕРЅРєРµ В«РџСЂРµСЃРµС‚В» РёР»Рё Р·Р°РґР°С‚СЊ Р·РЅР°С‡РµРЅРёСЏ РІСЂСѓС‡РЅСѓСЋ (РљР°СЃС‚РѕРј).")
            st.info("Р§С‚РѕР±С‹ РІС‹Р±СЂР°С‚СЊ РїСЂРµСЃРµС‚: РєР»РёРєРЅРёС‚Рµ СЏС‡РµР№РєСѓ РІ РєРѕР»РѕРЅРєРµ В«РџСЂРµСЃРµС‚ в–јВ» Рё РІС‹Р±РµСЂРёС‚Рµ РІР°СЂРёР°РЅС‚ РёР· СЃРїРёСЃРєР°.")
            st.caption("Р’ РєРѕР»РѕРЅРєР°С… CPC/CTR/CR СѓРєР°Р·С‹РІР°СЋС‚СЃСЏ РґРµР»РёС‚РµР»Рё РІР»РёСЏРЅРёСЏ. РџСЂРѕС†РµРЅС‚РЅС‹Р№ СЌС„С„РµРєС‚ РґР»СЏ РІС‹Р±СЂР°РЅРЅРѕРіРѕ РєРѕСЌС„С„РёС†РёРµРЅС‚Р° СЃРїСЂРѕСЃР° РїРѕРєР°Р·С‹РІР°РµС‚СЃСЏ РІ РїСЂРµРІСЊСЋ СЃРїСЂР°РІР°." + (" Р”Р»СЏ РїРѕР»РЅРѕР№ РІРѕСЂРѕРЅРєРё РЅРµРґРІРёР¶РёРјРѕСЃС‚Рё РѕС‚РґРµР»СЊРЅРѕ РЅР°СЃС‚СЂР°РёРІР°РµС‚СЃСЏ CR2." if show_cr2_elasticity else ""))
            st.markdown(
                "**РџСЂРµСЃРµС‚С‹ РІР»РёСЏРЅРёСЏ:**\n"
                "1. `РњСЏРіРєРёР№` - СЃР»Р°Р±РѕРµ РІР»РёСЏРЅРёРµ РЅР° РјРµС‚СЂРёРєРё (CPC СЂР°СЃС‚РµС‚ РјРµРЅСЊС€Рµ, CTR/CR РїР°РґР°СЋС‚ РјРµРЅСЊС€Рµ).\n"
                "2. `РЎСЂРµРґРЅРёР№` - СЃР±Р°Р»Р°РЅСЃРёСЂРѕРІР°РЅРЅС‹Р№ РІР°СЂРёР°РЅС‚ (РґРµС„РѕР»С‚ РґР»СЏ Р±РѕР»СЊС€РёРЅСЃС‚РІР° Р·Р°РґР°С‡).\n"
                "3. `РЎРёР»СЊРЅС‹Р№` - Р°РіСЂРµСЃСЃРёРІРЅРѕРµ РІР»РёСЏРЅРёРµ (РІС‹С€Рµ СЂРѕСЃС‚ CPC Рё СЃРёР»СЊРЅРµРµ РїР°РґРµРЅРёРµ CTR/CR).\n"
            )
            st.caption(
                "РџСЂРёРјРµСЂ СЌС„С„РµРєС‚Р° РґР»СЏ СЂРѕСЃС‚Р° СЃРїСЂРѕСЃР° РЅР° +10%:\n"
                "РњСЏРіРєРёР№: CPC +5.0%, CTR -2.0%, CR -0.67% | "
                "РЎСЂРµРґРЅРёР№: CPC +10.0%, CTR -5.0%, CR -1.0% | "
                "РЎРёР»СЊРЅС‹Р№: CPC +20.0%, CTR -10.0%, CR -2.0%."
            )
            st.caption(
                "РџСЂРёРјРµСЂ РїСЂРё +30% СЃРїСЂРѕСЃР° (РІ 3 СЂР°Р·Р° СЃРёР»СЊРЅРµРµ): "
                "РґР»СЏ РїСЂРµСЃРµС‚Р° В«РЎСЂРµРґРЅРµРµВ» РѕСЂРёРµРЅС‚РёСЂ: CPC +30.0%, CTR -15.0%, CR -3.0%."
            )
            st.caption(
                "РџСЂРёРјРµСЂ РґР»СЏ СЃРЅРёР¶РµРЅРёСЏ СЃРїСЂРѕСЃР° РґРѕ k=0.85 (-15%):\n"
                "РњСЏРіРєРёР№: CPC -7.5%, CTR +3.0%, CR +1.0% | "
                "РЎСЂРµРґРЅРёР№: CPC -15.0%, CTR +7.5%, CR +1.5% | "
                "РЎРёР»СЊРЅС‹Р№: CPC -30.0%, CTR +15.0%, CR +3.0%."
                "РЎРёР»СЊРЅС‹Р№: CPC -30.0%, CTR +15.0%, CR +3.0%."
            )
            top_left, top_right = st.columns([1.55, 1.05], vertical_alignment="bottom")
            with top_left:
                bulk_c1, bulk_c2, _bulk_sp = st.columns([1.2, 0.9, 1.2], vertical_alignment="bottom")
                with bulk_c1:
                    bulk_preset_name = st.selectbox(
                        "РњР°СЃСЃРѕРІС‹Р№ РїСЂРµСЃРµС‚",
                        options=[""] + list(preset_map.keys()),
                        key="elasticity_bulk_preset_select",
                    )
                with bulk_c2:
                    bulk_apply_clicked = st.form_submit_button("РџСЂРёРјРµРЅРёС‚СЊ РјР°СЃСЃРѕРІРѕ", type="primary")
            with top_right:
                preview_k_col, preview_btn_col, _preview_k_sp = st.columns([0.34, 0.20, 0.46], vertical_alignment="bottom")
                with preview_k_col:
                    preview_k_demand = st.number_input(
                        "РљРѕСЌС„С„РёС†РёРµРЅС‚ СЃРїСЂРѕСЃР° РґР»СЏ СЂР°СЃС‡РµС‚Р° РїСЂРµРІСЊСЋ",
                        min_value=0.10,
                        max_value=5.00,
                        value=float(st.session_state.get("elasticity_preview_k_demand", 1.10)),
                        step=0.01,
                        format="%.2f",
                        key="elasticity_preview_k_demand",
                        help="1.10 = +10% СЃРїСЂРѕСЃР°, 1.30 = +30%, 0.90 = -10%.",
                    )
                with preview_btn_col:
                    preview_recalc_clicked = st.form_submit_button("РћР±РЅРѕРІРёС‚СЊ")

            elasticity_row_h = max(28, int(ui_editor_row_height) - 8)
            compact_h = _table_height_for_rows(len(elasticity_df), min_height=140, max_height=280)

            edit_left, preview_right = st.columns([1.55, 1.05], vertical_alignment="top")
            with edit_left:
                st.caption("РќР°СЃС‚СЂРѕР№РєР° РїРѕ С‚РёРїР°Рј Р Рљ")
                elasticity_editor_key = f"elasticity_editor_{int(st.session_state.get('elasticity_editor_nonce', 0))}"
                elasticity_draft = st.data_editor(
                    elasticity_df,
                    num_rows="fixed",
                    use_container_width=True,
                    row_height=elasticity_row_h,
                    height=compact_h,
                    column_config={
                        "campaign_type": st.column_config.TextColumn("РўРёРї Р Рљ", disabled=True),
                        "preset": st.column_config.SelectboxColumn(
                            "РџСЂРµСЃРµС‚ в–ј (РєР»РёРєРЅРёС‚Рµ СЏС‡РµР№РєСѓ)",
                            options=["РљР°СЃС‚РѕРј", "РЎР»Р°Р±РѕРµ", "РЎСЂРµРґРЅРµРµ", "РЎРёР»СЊРЅРѕРµ"],
                            help="Р•СЃР»Рё Р·РЅР°С‡РµРЅРёСЏ РІ СЃС‚СЂРѕРєРµ СЃРѕРІРїР°РґР°СЋС‚ СЃ РїСЂРµСЃРµС‚РѕРј, РѕРЅ СЃРѕС…СЂР°РЅРёС‚СЃСЏ РєР°Рє РїСЂРµСЃРµС‚; РµСЃР»Рё РІС‹ РёР·РјРµРЅРёС‚Рµ С‡РёСЃР»Р° РІСЂСѓС‡РЅСѓСЋ, СЃС‚СЂРѕРєР° СЃРѕС…СЂР°РЅРёС‚СЃСЏ РєР°Рє В«РљР°СЃС‚РѕРјВ».",
                        ),
                        "cpc_div": st.column_config.NumberColumn(
                            "CPC",
                            format="%.2f",
                            help=mhelp("cpc_div"),
                        ),
                        "ctr_div": st.column_config.NumberColumn(
                            "CTR",
                            format="%.2f",
                            help=mhelp("ctr_div"),
                        ),
                        "cr_div": st.column_config.NumberColumn(
                            "CR",
                            format="%.2f",
                            help=mhelp("cr_div"),
                        ),
                        "cr2_div": st.column_config.NumberColumn(
                            "CR2",
                            format="%.2f",
                            help="Р”РµР»РёС‚РµР»СЊ РІР»РёСЏРЅРёСЏ СЃРїСЂРѕСЃР° РЅР° РІС‚РѕСЂСѓСЋ СЃС‚СѓРїРµРЅСЊ РєРѕРЅРІРµСЂСЃРёРё РІ Р¦Рћ.",
                        ),
                    },
                    key=elasticity_editor_key,
                )
                if not show_cr2_elasticity and "cr2_div" in elasticity_draft.columns:
                    elasticity_draft = elasticity_draft.drop(columns=["cr2_div"], errors="ignore")
            with preview_right:
                st.caption("Р РµР·СѓР»СЊС‚Р°С‚ РїРѕ С‚РµРєСѓС‰РёРј Р·РЅР°С‡РµРЅРёСЏРј РІ СЃС‚СЂРѕРєРµ")
                preview_effective = elasticity_draft[
                    ["campaign_type", "preset", "cpc_div", "ctr_div", "cr_div"] + (["cr2_div"] if show_cr2_elasticity else [])
                ].copy()
                saved_by_campaign = elasticity_df.set_index("campaign_type")
                for idx_row, row in preview_effective.iterrows():
                    campaign_type = str(row.get("campaign_type", ""))
                    p_name = str(row.get("preset", "РљР°СЃС‚РѕРј")).strip()
                    preview_effective.at[idx_row, "cpc_div"] = _coerce_div_for_preview(row.get("cpc_div", 1.0), 1.0)
                    preview_effective.at[idx_row, "ctr_div"] = _coerce_div_for_preview(row.get("ctr_div", 2.0), 2.0)
                    preview_effective.at[idx_row, "cr_div"] = _coerce_div_for_preview(row.get("cr_div", 10.0), 10.0)
                    if show_cr2_elasticity:
                        preview_effective.at[idx_row, "cr2_div"] = _coerce_div_for_preview(row.get("cr2_div", 10.0), 10.0)
                    if p_name not in preset_map or campaign_type not in saved_by_campaign.index:
                        continue
                    saved_row = saved_by_campaign.loc[campaign_type]
                    saved_preset = str(saved_row.get("preset", "РљР°СЃС‚РѕРј")).strip()
                    saved_cpc = _coerce_div_for_preview(saved_row.get("cpc_div", 1.0), 1.0)
                    saved_ctr = _coerce_div_for_preview(saved_row.get("ctr_div", 2.0), 2.0)
                    saved_cr = _coerce_div_for_preview(saved_row.get("cr_div", 10.0), 10.0)
                    raw_cpc = float(preview_effective.at[idx_row, "cpc_div"])
                    raw_ctr = float(preview_effective.at[idx_row, "ctr_div"])
                    raw_cr = float(preview_effective.at[idx_row, "cr_div"])
                    preset_changed_only = (
                        p_name != saved_preset
                        and abs(raw_cpc - saved_cpc) <= 1e-9
                        and abs(raw_ctr - saved_ctr) <= 1e-9
                        and abs(raw_cr - saved_cr) <= 1e-9
                        and (
                            not show_cr2_elasticity
                            or abs(
                                float(preview_effective.at[idx_row, "cr2_div"])
                                - _coerce_div_for_preview(saved_row.get("cr2_div", 10.0), 10.0)
                            ) <= 1e-9
                        )
                    )
                    if preset_changed_only:
                        p_vals = preset_map[p_name]
                        preview_effective.at[idx_row, "cpc_div"] = p_vals["cpc_div"]
                        preview_effective.at[idx_row, "ctr_div"] = p_vals["ctr_div"]
                        preview_effective.at[idx_row, "cr_div"] = p_vals["cr_div"]
                        if show_cr2_elasticity:
                            preview_effective.at[idx_row, "cr2_div"] = p_vals["cr2_div"]
                # РџРµСЂРµСЃС‡РµС‚ РїРѕРґ РІС‹Р±СЂР°РЅРЅС‹Р№ РєРѕСЌС„С„РёС†РёРµРЅС‚ СЃРїСЂРѕСЃР°:
                # Р·РЅР°С‡РµРЅРёСЏ РІ С‚Р°Р±Р»РёС†Рµ Р·Р°РґР°СЋС‚СЃСЏ РєР°Рє СЌС„С„РµРєС‚ РЅР° +10% СЃРїСЂРѕСЃР°, РјР°СЃС€С‚Р°Р±РёСЂСѓРµРј Р»РёРЅРµР№РЅРѕ.
                # scale = 1 РїСЂРё k=1.10; scale = 3 РїСЂРё k=1.30; scale = -1 РїСЂРё k=0.90.
                demand_delta = float(preview_k_demand) - 1.0
                cpc_div_vals = pd.to_numeric(preview_effective["cpc_div"], errors="coerce").replace(0, np.nan)
                ctr_div_vals = pd.to_numeric(preview_effective["ctr_div"], errors="coerce").replace(0, np.nan)
                cr_div_vals = pd.to_numeric(preview_effective["cr_div"], errors="coerce").replace(0, np.nan)
                cr2_div_vals = pd.to_numeric(preview_effective["cr2_div"], errors="coerce").replace(0, np.nan) if show_cr2_elasticity else None
                preview_effective["cpc_pct"] = (demand_delta / cpc_div_vals) * 100.0
                preview_effective["ctr_pct"] = -(demand_delta / ctr_div_vals) * 100.0
                preview_effective["cr_pct"] = -(demand_delta / cr_div_vals) * 100.0
                if show_cr2_elasticity:
                    preview_effective["cr2_pct"] = -(demand_delta / cr2_div_vals) * 100.0
                preview_cols = ["campaign_type", "cpc_pct", "ctr_pct", "cr_pct"] + (["cr2_pct"] if show_cr2_elasticity else [])
                preview_effective = preview_effective[preview_cols].rename(
                    columns={
                        "campaign_type": "РўРёРї Р Рљ",
                        "cpc_pct": "CPC, %",
                        "ctr_pct": "CTR, %",
                        "cr_pct": "CR, %",
                        "cr2_pct": "CR2, %",
                    }
                )
                st.dataframe(
                    preview_effective.style
                    .format({"CPC, %": "{:+.2f}%", "CTR, %": "{:+.2f}%", "CR, %": "{:+.2f}%", "CR2, %": "{:+.2f}%"})
                    .apply(_style_impact_col, axis=0, subset=["CPC, %", "CTR, %", "CR, %"] + (["CR2, %"] if show_cr2_elasticity else [])),
                    use_container_width=True,
                    hide_index=True,
                    height=compact_h,
                )
                st.caption("Р—РµР»РµРЅС‹Р№ = СЂРѕСЃС‚ РјРµС‚СЂРёРєРё, РєСЂР°СЃРЅС‹Р№ = СЃРЅРёР¶РµРЅРёРµ. Р—РЅР°С‡РµРЅРёСЏ РїРµСЂРµСЃС‡РёС‚Р°РЅС‹ РїРѕРґ РІС‹Р±СЂР°РЅРЅС‹Р№ РєРѕСЌС„С„РёС†РёРµРЅС‚ СЃРїСЂРѕСЃР°.")
            if _has_unsaved_changes(elasticity_draft, elasticity_df):
                st.caption("Р•СЃС‚СЊ РЅРµСЃРѕС…СЂР°РЅРµРЅРЅС‹Рµ РёР·РјРµРЅРµРЅРёСЏ СЌР»Р°СЃС‚РёС‡РЅРѕСЃС‚Рё.")
            btn_save_col, btn_reset_col, _btn_sp = st.columns([0.28, 0.34, 0.38])
            with btn_save_col:
                elasticity_saved = st.form_submit_button("РЎРѕС…СЂР°РЅРёС‚СЊ СЌР»Р°СЃС‚РёС‡РЅРѕСЃС‚СЊ", type="primary")
            with btn_reset_col:
                reset_recommended_clicked = st.form_submit_button("Р’РµСЂРЅСѓС‚СЊ СЂРµРєРѕРјРµРЅРґСѓРµРјС‹Рµ Р·РЅР°С‡РµРЅРёСЏ")

        if bulk_apply_clicked:
            if bulk_preset_name in preset_map:
                elasticity_draft = elasticity_draft.copy()
                elasticity_draft["preset"] = bulk_preset_name
                preset_vals = preset_map[bulk_preset_name]
                elasticity_draft["cpc_div"] = preset_vals["cpc_div"]
                elasticity_draft["ctr_div"] = preset_vals["ctr_div"]
                elasticity_draft["cr_div"] = preset_vals["cr_div"]
                if "cr2_div" in elasticity_draft.columns:
                    elasticity_draft["cr2_div"] = preset_vals["cr2_div"]
                st.session_state["elasticity_df"] = elasticity_draft.copy()
                elasticity_df = elasticity_draft.copy()
                st.session_state["elasticity_editor_nonce"] = int(st.session_state.get("elasticity_editor_nonce", 0)) + 1
                st.success(f"РњР°СЃСЃРѕРІРѕ РїСЂРёРјРµРЅРµРЅ РїСЂРµСЃРµС‚ В«{bulk_preset_name}В».")
                st.rerun()
            else:
                st.info("Р’С‹Р±РµСЂРёС‚Рµ РїСЂРµСЃРµС‚ РґР»СЏ РјР°СЃСЃРѕРІРѕРіРѕ РїСЂРёРјРµРЅРµРЅРёСЏ.")

        if elasticity_saved:
            elasticity_draft = elasticity_draft.copy()
            for idx_row, row in elasticity_draft.iterrows():
                preset_name = str(row.get("preset", "РљР°СЃС‚РѕРј")).strip()
                if preset_name in preset_map:
                    preset_vals = preset_map[preset_name]
                    if _row_divs_match_preset(row, preset_vals):
                        elasticity_draft.at[idx_row, "cpc_div"] = preset_vals["cpc_div"]
                        elasticity_draft.at[idx_row, "ctr_div"] = preset_vals["ctr_div"]
                        elasticity_draft.at[idx_row, "cr_div"] = preset_vals["cr_div"]
                        if "cr2_div" in elasticity_draft.columns:
                            elasticity_draft.at[idx_row, "cr2_div"] = preset_vals["cr2_div"]
                    else:
                        elasticity_draft.at[idx_row, "preset"] = "РљР°СЃС‚РѕРј"
            st.session_state["elasticity_df"] = elasticity_draft.copy()
            elasticity_df = elasticity_draft.copy()

        if reset_recommended_clicked:
            reset_df = elasticity_draft.copy()
            saved_by_campaign = elasticity_df.set_index("campaign_type") if not elasticity_df.empty else pd.DataFrame()
            for idx_row, row in reset_df.iterrows():
                campaign_type = str(row.get("campaign_type", ""))
                preset_name = str(row.get("preset", "РљР°СЃС‚РѕРј")).strip()
                if preset_name not in preset_map:
                    saved_preset = ""
                    if campaign_type and not saved_by_campaign.empty and campaign_type in saved_by_campaign.index:
                        saved_preset = str(saved_by_campaign.loc[campaign_type].get("preset", "")).strip()
                    preset_name = saved_preset if saved_preset in preset_map else "РЎСЂРµРґРЅРµРµ"
                preset_vals = preset_map[preset_name]
                reset_df.at[idx_row, "preset"] = preset_name
                reset_df.at[idx_row, "cpc_div"] = preset_vals["cpc_div"]
                reset_df.at[idx_row, "ctr_div"] = preset_vals["ctr_div"]
                reset_df.at[idx_row, "cr_div"] = preset_vals["cr_div"]
                if "cr2_div" in reset_df.columns:
                    reset_df.at[idx_row, "cr2_div"] = preset_vals["cr2_div"]
            st.session_state["elasticity_df"] = reset_df
            st.session_state["elasticity_editor_nonce"] = int(st.session_state.get("elasticity_editor_nonce", 0)) + 1
            st.rerun()

        def _safe_div(val: float, default_val: float) -> float:
            try:
                num = float(val)
                if pd.isna(num) or num <= 0:
                    return default_val
                return num
            except Exception:
                return default_val

        def _is_invalid_div(val: float) -> bool:
            try:
                num = float(val)
                return pd.isna(num) or num <= 0
            except Exception:
                return True

        had_invalid_values = False
        elasticity_map = {}
        for _, r in elasticity_df.iterrows():
            camp = str(r["campaign_type"])
            raw_cpc_div = r.get("cpc_div")
            raw_ctr_div = r.get("ctr_div")
            raw_cr_div = r.get("cr_div")
            raw_cr2_div = r.get("cr2_div")

            cpc_div = _safe_div(raw_cpc_div, 1.0)
            ctr_div = _safe_div(raw_ctr_div, 2.0)
            cr_div = _safe_div(raw_cr_div, 10.0)
            cr2_div = _safe_div(raw_cr2_div, 10.0)

            if (
                _is_invalid_div(raw_cpc_div)
                or _is_invalid_div(raw_ctr_div)
                or _is_invalid_div(raw_cr_div)
                or (show_cr2_elasticity and _is_invalid_div(raw_cr2_div))
            ):
                had_invalid_values = True

            elasticity_map[camp] = {
                "cpc_div": cpc_div,
                "ctr_div": ctr_div,
                "cr_div": cr_div,
                "cr2_div": cr2_div,
            }

        if had_invalid_values:
            st.warning(
                "РќРµРєРѕСЂСЂРµРєС‚РЅС‹Рµ Р·РЅР°С‡РµРЅРёСЏ РґРµР»РёС‚РµР»РµР№ СЌР»Р°СЃС‚РёС‡РЅРѕСЃС‚Рё (РїСѓСЃС‚С‹Рµ, РЅРµС‡РёСЃР»РѕРІС‹Рµ РёР»Рё <= 0) Р°РІС‚РѕРјР°С‚РёС‡РµСЃРєРё Р·Р°РјРµРЅРµРЅС‹: "
                + ("CPC=1.0, CTR=2.0, CR1=10.0, CR2=10.0." if show_cr2_elasticity else "CPC=1.0, CTR=2.0, CR=10.0.")
            )

    # ---------- 2. РЎСЂРµРґРЅРёР№ РјРµСЃСЏС† (Р±Р°Р·РѕРІС‹Рµ Р·РЅР°С‡РµРЅРёСЏ) ----------

    if plan_calc_blockers:
        st.info(
            "РџРѕРјРµСЃСЏС‡РЅС‹Р№ СЂР°СЃС‡С‘С‚ РїСЂРёРѕСЃС‚Р°РЅРѕРІР»РµРЅ. Р§С‚РѕР±С‹ РїРµСЂРµР№С‚Рё Рє Р±Р»РѕРєР°Рј 2вЂ“4, РёСЃРїСЂР°РІСЊС‚Рµ: "
            + "; ".join(dict.fromkeys(plan_calc_blockers))
            + "."
        )
        selected_month_nums = []

    ui_section_title("2. РЎСЂРµРґРЅРёР№ РјРµСЃСЏС† (Р±Р°Р·РѕРІС‹Рµ Р·РЅР°С‡РµРЅРёСЏ)")
    # Р‘РµСЂРµРј Р°РєС‚СѓР°Р»СЊРЅС‹Рµ РїСЂР°РІРёР»Р° РёР· session_state РїРµСЂРµРґ СЂР°СЃС‡РµС‚Р°РјРё,
    # С‡С‚РѕР±С‹ РёСЃРєР»СЋС‡РёС‚СЊ СЂР°СЃСЃРёРЅС…СЂРѕРЅ СЃ data_editor.
    ak_rules_df = st.session_state.get("ak_rules_df", ak_rules_df)

    base_rows = []
    for _, base_row in campaigns.iterrows():
        base = PlanInput(
            impressions=base_row["impressions_avg"],
            ctr=base_row["ctr_avg_percent"] / 100.0,
            cpc=base_row["cpc_avg"],
            cr=base_row["cr_avg_percent"] / 100.0,
            aov=base_row["aov_avg"],
            cr2=float(base_row.get("cr2_avg_percent", 0.0) or 0.0) / 100.0,
            reach=float(base_row.get("reach_avg", 0.0) or 0.0),
            preset_key=active_preset_key,
            funnel_mode=metric_mode["real_estate_funnel_mode"],
        )
        out = calculate_plan_month(base)
        out["campaign_type"] = base_row["campaign_type"]
        out["segment"] = str(base_row.get("segment", "B2C"))
        out["system"] = base_row["system"]
        out["format"] = base_row["format"]
        out["geo"] = str(base_row.get("geo", "") or "")
        if is_diy_preset:
            out["revenue"] = float(out.get("revenue", 0.0)) * (1.0 + VAT_RATE)
        if is_diy_preset:
            cap_avg = float(base_row.get("available_capacity_avg", 0.0) or 0.0)
            out["available_capacity"] = cap_avg
            out["client_count"] = float(base_row.get("client_count_avg", 0.0) or 0.0)
            out["absolute_new_clients"] = float(base_row.get("absolute_new_clients_avg", 0.0) or 0.0)
            out["returned_clients"] = float(base_row.get("returned_clients_avg", 0.0) or 0.0)
            out["new_clients"] = float(out.get("absolute_new_clients", 0.0)) + float(out.get("returned_clients", 0.0))
            out["order_frequency"] = float(base_row.get("order_frequency_avg", 0.0) or 0.0)
            out["new_clients_share_pct"] = float(base_row.get("new_clients_share_avg_percent", 0.0) or 0.0)
            out["sov_pct"] = (float(out.get("reach", 0.0)) / cap_avg * 100.0) if cap_avg > 0 else 0.0
            out["cac"] = (float(out.get("cost_with_vat_ak", 0.0)) / float(out.get("new_clients", 0.0))) if float(out.get("new_clients", 0.0)) > 0 else 0.0
            out["shipped_orders"] = float(out.get("target_leads", 0.0) or 0.0)
            out["shipped_revenue"] = float(out.get("shipped_orders", 0.0)) * float(out.get("aov", 0.0) or 0.0)
        base_rows.append(out)

    df_base = pd.DataFrame(base_rows)
    if not df_base.empty:
        base_ak_rate = resolve_ak_rate(float(df_base["cost"].sum()), ak_rules_df) if (use_ak_budget_metrics and ak_mode == "percent") else 0.0
        df_base = apply_budget_basis_metrics(
            df_base,
            use_vat_budget_metrics,
            use_ak=use_ak_budget_metrics,
            ak_mode=ak_mode,
            default_ak_rate=base_ak_rate,
            default_ak_fixed_wo_vat=float(ak_fixed_month_wo_vat),
            default_ak_fixed_rate=float(ak_fixed_percent) / 100.0,
        )
        df_base["ak_rate_pct"] = df_base["ak_rate"] * 100.0
        if is_diy_preset:
            df_base["cpa"] = np.where(
                pd.to_numeric(df_base.get("conversions", 0.0), errors="coerce").fillna(0.0) > 0,
                pd.to_numeric(df_base.get("cost_with_vat", 0.0), errors="coerce").fillna(0.0)
                / pd.to_numeric(df_base.get("conversions", 0.0), errors="coerce").fillna(0.0),
                0.0,
            )
            df_base["cac"] = np.where(
                pd.to_numeric(df_base.get("new_clients", 0.0), errors="coerce").fillna(0.0) > 0,
                pd.to_numeric(df_base.get("cost_with_vat_ak", 0.0), errors="coerce").fillna(0.0)
                / pd.to_numeric(df_base.get("new_clients", 0.0), errors="coerce").fillna(0.0),
                0.0,
            )
            df_base["shipped_cps"] = np.where(
                pd.to_numeric(df_base.get("shipped_orders", 0.0), errors="coerce").fillna(0.0) > 0,
                pd.to_numeric(df_base.get("cost_with_vat", 0.0), errors="coerce").fillna(0.0)
                / pd.to_numeric(df_base.get("shipped_orders", 0.0), errors="coerce").fillna(0.0),
                0.0,
            )
            df_base["shipped_aov"] = np.where(
                pd.to_numeric(df_base.get("shipped_orders", 0.0), errors="coerce").fillna(0.0) > 0,
                pd.to_numeric(df_base.get("shipped_revenue", 0.0), errors="coerce").fillna(0.0)
                / pd.to_numeric(df_base.get("shipped_orders", 0.0), errors="coerce").fillna(0.0),
                0.0,
            )
            shipped_budget_basis = pd.to_numeric(df_base.get("cost_with_vat_ak" if use_vat_budget_metrics and use_ak_budget_metrics else "cost_with_vat" if use_vat_budget_metrics else "total_budget_wo_vat_ak" if use_ak_budget_metrics else "cost", 0.0), errors="coerce").fillna(0.0)
            df_base["shipped_roas"] = np.where(
                shipped_budget_basis > 0,
                pd.to_numeric(df_base.get("shipped_revenue", 0.0), errors="coerce").fillna(0.0) / shipped_budget_basis,
                0.0,
            )
            df_base["shipped_drr_pct"] = np.where(
                pd.to_numeric(df_base.get("shipped_revenue", 0.0), errors="coerce").fillna(0.0) > 0,
                shipped_budget_basis / pd.to_numeric(df_base.get("shipped_revenue", 0.0), errors="coerce").fillna(0.0) * 100.0,
                0.0,
            )
            df_base = add_segment_budget_share(df_base)
            df_base = add_segment_value_share(df_base, value_col="conversions", out_col="order_share_segment_pct")
            df_base = add_segment_value_share(df_base, value_col="revenue", out_col="revenue_share_segment_pct")
            df_base = add_segment_value_share(df_base, value_col="shipped_orders", out_col="shipped_order_share_segment_pct")
            df_base = add_segment_value_share(df_base, value_col="shipped_revenue", out_col="shipped_revenue_share_segment_pct")

    if df_base.empty:
        st.info("РќРµС‚ РґР°РЅРЅС‹С… РґР»СЏ СЃСЂРµРґРЅРµРіРѕ РјРµСЃСЏС†Р°.")
    else:
        df_base_show = df_base.copy()
        if is_real_estate_preset:
            df_base_show = compute_real_estate_rates(df_base_show)
        df_base_show["ctr_pct"] = df_base_show["ctr"] * 100
        df_base_show["cr_pct"] = df_base_show["cr"] * 100
        df_base_show["drr_pct"] = df_base_show["drr"] * 100

        total_imp = df_base["impressions"].sum()
        total_reach = float(df_base["reach"].sum()) if "reach" in df_base.columns else 0.0
        total_frequency = (total_imp / total_reach) if total_reach > 0 else 0.0
        total_clicks = df_base["clicks"].sum()
        total_conv = df_base["conversions"].sum()
        total_leads = float(df_base["leads"].fillna(0.0).sum()) if "leads" in df_base.columns else 0.0
        total_target_leads = float(df_base["target_leads"].fillna(0.0).sum()) if "target_leads" in df_base.columns else float(total_conv)
        total_cost = df_base["cost"].sum()
        total_cost_with_vat = df_base["cost_with_vat"].sum()
        total_ak_wo_vat = df_base["ak_cost_wo_vat"].sum()
        total_cost_with_vat_ak = df_base["cost_with_vat_ak"].sum()
        total_rev = df_base["revenue"].sum()
        if use_ak_budget_metrics:
            total_budget_basis = total_cost_with_vat_ak if use_vat_budget_metrics else (total_cost + total_ak_wo_vat)
        else:
            total_budget_basis = total_cost_with_vat if use_vat_budget_metrics else total_cost
        if is_diy_preset:
            total_cpa = (total_cost_with_vat / total_conv) if total_conv > 0 else 0
        else:
            total_cpa = (total_budget_basis / total_conv) if total_conv > 0 else 0

        total_ctr = (total_clicks / total_imp * 100) if total_imp > 0 else 0
        total_cpc = (total_cost / total_clicks) if total_clicks > 0 else 0
        total_cr = (total_conv / total_clicks * 100) if total_clicks > 0 else 0
        total_cr1 = (total_leads / total_clicks * 100) if total_clicks > 0 else 0
        total_cr2 = (total_target_leads / total_leads * 100) if total_leads > 0 else 0
        total_cpm = (total_budget_basis / (total_imp / 1000)) if total_imp > 0 else 0
        total_cpl = (total_budget_basis / total_leads) if total_leads > 0 else 0
        total_cpql = (total_budget_basis / total_target_leads) if total_target_leads > 0 else 0
        total_roas = (total_rev / total_budget_basis) if total_budget_basis > 0 else 0
        total_drr = (total_budget_basis / total_rev * 100) if total_rev > 0 else 0
        st.session_state["mp_ref_base"] = {
            "impressions": float(total_imp),
            "reach": float(total_reach),
            "frequency": float(total_frequency),
            "clicks": float(total_clicks),
            "conversions": float(total_conv),
            "leads": float(total_leads),
            "target_leads": float(total_target_leads),
            "cost": float(total_cost),
            "cost_with_vat": float(total_cost_with_vat),
            "cost_with_vat_ak": float(total_cost_with_vat_ak),
            "revenue": float(total_rev),
            "ctr": float(total_ctr),
            "cpc": float(total_cpc),
            "cr": float(total_cr),
            "cr1": float(total_cr1),
            "cr2": float(total_cr2),
            "cpo": float(total_cpa),
            "cpl": float(total_cpl),
            "cpql": float(total_cpql),
            "roas": float(total_roas * 100.0),
            "drr": float(total_drr),
        }
        base_by_campaign = {}
        for _, r in df_base.iterrows():
            camp = str(r.get("campaign_type", "")).strip()
            if not camp:
                continue
            imp = float(r.get("impressions", 0.0))
            reach = float(r.get("reach", 0.0) or 0.0)
            frequency = (imp / reach) if reach > 0 else 0.0
            clicks = float(r.get("clicks", 0.0))
            conv = float(r.get("conversions", 0.0))
            cost = float(r.get("cost", 0.0))
            cost_with_vat = float(r.get("cost_with_vat", 0.0))
            cost_with_vat_ak = float(r.get("cost_with_vat_ak", 0.0))
            rev = float(r.get("revenue", 0.0))
            if use_ak_budget_metrics:
                budget_basis = cost_with_vat_ak if use_vat_budget_metrics else float(r.get("cost", 0.0) + r.get("ak_cost_wo_vat", 0.0))
            else:
                budget_basis = cost_with_vat if use_vat_budget_metrics else cost
            ctr = (clicks / imp * 100.0) if imp > 0 else 0.0
            cpc = (cost / clicks) if clicks > 0 else 0.0
            cr = (conv / clicks * 100.0) if clicks > 0 else 0.0
            leads = float(r.get("leads", 0.0) or 0.0)
            target_leads = float(r.get("target_leads", conv) or 0.0)
            cr1 = (leads / clicks * 100.0) if clicks > 0 else 0.0
            cr2 = (target_leads / leads * 100.0) if leads > 0 else 0.0
            cpo = (budget_basis / conv) if conv > 0 else 0.0
            cpl = (budget_basis / leads) if leads > 0 else 0.0
            cpql = (budget_basis / target_leads) if target_leads > 0 else 0.0
            roas = (rev / budget_basis * 100.0) if budget_basis > 0 else 0.0
            drr = (budget_basis / rev * 100.0) if rev > 0 else 0.0
            base_by_campaign[camp] = {
                "impressions": imp,
                "reach": reach,
                "frequency": frequency,
                "clicks": clicks,
                "conversions": conv,
                "leads": leads,
                "target_leads": target_leads,
                "cost": cost,
                "cost_with_vat": cost_with_vat,
                "cost_with_vat_ak": cost_with_vat_ak,
                "revenue": rev,
                "ctr": ctr,
                "cpc": cpc,
                "cr": cr,
                "cr1": cr1,
                "cr2": cr2,
                "cpo": cpo,
                "cpl": cpl,
                "cpql": cpql,
                "roas": roas,
                "drr": drr,
            }
        st.session_state["mp_ref_base_by_campaign"] = base_by_campaign

        total_row_raw = {
            "campaign_type": "РС‚РѕРіРѕ",
            "segment": "ALL",
            "system": "",
            "format": "",
            "impressions": total_imp,
            "ctr": total_ctr / 100,
            "cpc": total_cpc,
            "cr": total_cr / 100,
            "aov": None,
            "clicks": total_clicks,
            "conversions": total_conv,
            "leads": total_leads,
            "target_leads": total_target_leads,
            "cost": total_cost,
            "ak_cost_wo_vat": total_ak_wo_vat,
            "total_budget_wo_vat_ak": total_cost + total_ak_wo_vat,
            "cost_with_vat": total_cost_with_vat,
            "cost_with_vat_ak": total_cost_with_vat_ak,
            "vat_amount": total_cost_with_vat_ak - (total_cost + total_ak_wo_vat),
            "revenue": total_rev,
            "cpm": total_cpm,
            "cpa": total_cpa,
            "cpl": total_cpl,
            "cpql": total_cpql,
            "roas": total_roas,
            "drr": total_drr / 100,
            "ctr_pct": total_ctr,
            "cr_pct": total_cr,
            "cr1_pct": total_cr1,
            "cr2_pct": total_cr2,
            "drr_pct": total_drr,
        }
        if is_diy_preset:
            total_capacity = float(df_base["available_capacity"].sum()) if "available_capacity" in df_base.columns else 0.0
            total_row_raw["available_capacity"] = total_capacity
            total_row_raw["client_count"] = float(df_base["client_count"].sum()) if "client_count" in df_base.columns else 0.0
            total_row_raw["absolute_new_clients"] = float(df_base["absolute_new_clients"].sum()) if "absolute_new_clients" in df_base.columns else 0.0
            total_row_raw["returned_clients"] = float(df_base["returned_clients"].sum()) if "returned_clients" in df_base.columns else 0.0
            total_row_raw["new_clients"] = float(df_base["new_clients"].sum()) if "new_clients" in df_base.columns else 0.0
            total_row_raw["order_frequency"] = float(df_base["order_frequency"].mean()) if "order_frequency" in df_base.columns else 0.0
            total_row_raw["sov_pct"] = (total_reach / total_capacity * 100.0) if total_capacity > 0 else 0.0
            total_row_raw["new_clients_share_pct"] = float(df_base["new_clients_share_pct"].mean()) if "new_clients_share_pct" in df_base.columns else 0.0
            total_row_raw["cac"] = (total_cost_with_vat_ak / float(total_row_raw["new_clients"])) if float(total_row_raw["new_clients"]) > 0 else 0.0
            total_row_raw["shipped_orders"] = float(df_base["shipped_orders"].sum()) if "shipped_orders" in df_base.columns else 0.0
            total_row_raw["shipped_revenue"] = float(df_base["shipped_revenue"].sum()) if "shipped_revenue" in df_base.columns else 0.0
            total_row_raw["shipped_cps"] = (total_cost_with_vat / float(total_row_raw["shipped_orders"])) if float(total_row_raw["shipped_orders"]) > 0 else 0.0
            total_row_raw["shipped_aov"] = (float(total_row_raw["shipped_revenue"]) / float(total_row_raw["shipped_orders"])) if float(total_row_raw["shipped_orders"]) > 0 else 0.0
            total_row_raw["shipped_roas"] = (float(total_row_raw["shipped_revenue"]) / total_budget_basis) if total_budget_basis > 0 else 0.0
            total_row_raw["shipped_drr_pct"] = (total_budget_basis / float(total_row_raw["shipped_revenue"]) * 100.0) if float(total_row_raw["shipped_revenue"]) > 0 else 0.0
            total_row_raw["order_share_segment_pct"] = 100.0
            total_row_raw["revenue_share_segment_pct"] = 100.0
            total_row_raw["shipped_order_share_segment_pct"] = 100.0
            total_row_raw["shipped_revenue_share_segment_pct"] = 100.0
            total_row_raw["budget_share_segment_pct"] = 100.0
        segment_total_rows = []
        if show_segment_subtotals and "segment" in df_base.columns:
            for seg_name, seg_df in df_base.groupby("segment"):
                seg_imp = seg_df["impressions"].sum()
                seg_clicks = seg_df["clicks"].sum()
                seg_conv = seg_df["conversions"].sum()
                seg_leads = float(seg_df["leads"].fillna(0.0).sum()) if "leads" in seg_df.columns else 0.0
                seg_target_leads = float(seg_df["target_leads"].fillna(0.0).sum()) if "target_leads" in seg_df.columns else float(seg_conv)
                seg_cost = seg_df["cost"].sum()
                seg_cost_with_vat = seg_df["cost_with_vat"].sum()
                seg_cost_with_vat_ak = seg_df["cost_with_vat_ak"].sum()
                seg_ak_wo_vat = seg_df["ak_cost_wo_vat"].sum()
                seg_rev = seg_df["revenue"].sum()
                if use_ak_budget_metrics:
                    seg_budget_basis = seg_cost_with_vat_ak if use_vat_budget_metrics else (seg_cost + seg_ak_wo_vat)
                else:
                    seg_budget_basis = seg_cost_with_vat if use_vat_budget_metrics else seg_cost
                seg_ctr = (seg_clicks / seg_imp * 100) if seg_imp > 0 else 0
                seg_cpc = (seg_cost / seg_clicks) if seg_clicks > 0 else 0
                seg_cr = (seg_conv / seg_clicks * 100) if seg_clicks > 0 else 0
                seg_cr1 = (seg_leads / seg_clicks * 100) if seg_clicks > 0 else 0
                seg_cr2 = (seg_target_leads / seg_leads * 100) if seg_leads > 0 else 0
                seg_cpm = (seg_budget_basis / (seg_imp / 1000)) if seg_imp > 0 else 0
                seg_cpa = (seg_cost_with_vat / seg_conv) if (is_diy_preset and seg_conv > 0) else (seg_budget_basis / seg_conv) if seg_conv > 0 else 0
                seg_cpl = (seg_budget_basis / seg_leads) if seg_leads > 0 else 0
                seg_cpql = (seg_budget_basis / seg_target_leads) if seg_target_leads > 0 else 0
                seg_roas = (seg_rev / seg_budget_basis) if seg_budget_basis > 0 else 0
                seg_drr = (seg_budget_basis / seg_rev * 100) if seg_rev > 0 else 0
                segment_total_rows.append(
                    {
                        "campaign_type": f"РС‚РѕРіРѕ {seg_name}",
                        "segment": seg_name,
                        "system": "",
                        "format": "",
                        "impressions": seg_imp,
                        "reach": float(seg_df["reach"].sum()) if "reach" in seg_df.columns else 0.0,
                        "frequency": (seg_imp / float(seg_df["reach"].sum())) if ("reach" in seg_df.columns and float(seg_df["reach"].sum()) > 0) else 0.0,
                        "ctr": seg_ctr / 100,
                        "cpc": seg_cpc,
                        "cr": seg_cr / 100,
                        "aov": None,
                        "clicks": seg_clicks,
                        "conversions": seg_conv,
                        "leads": seg_leads,
                        "target_leads": seg_target_leads,
                        "cost": seg_cost,
                        "ak_cost_wo_vat": seg_ak_wo_vat,
                        "total_budget_wo_vat_ak": seg_cost + seg_ak_wo_vat,
                        "cost_with_vat": seg_cost_with_vat,
                        "cost_with_vat_ak": seg_cost_with_vat_ak,
                        "vat_amount": seg_cost_with_vat_ak - (seg_cost + seg_ak_wo_vat),
                        "revenue": seg_rev,
                        "cpm": seg_cpm,
                        "cpa": seg_cpa,
                        "cpl": seg_cpl,
                        "cpql": seg_cpql,
                        "roas": seg_roas,
                        "drr": seg_drr / 100,
                        "ctr_pct": seg_ctr,
                        "cr_pct": seg_cr,
                        "cr1_pct": seg_cr1,
                        "cr2_pct": seg_cr2,
                        "drr_pct": seg_drr,
                        "ak_rate_pct": (seg_ak_wo_vat / seg_cost * 100.0) if seg_cost > 0 else 0.0,
                    }
                )
                if is_diy_preset:
                    seg_capacity = float(seg_df["available_capacity"].sum()) if "available_capacity" in seg_df.columns else 0.0
                    segment_total_rows[-1]["available_capacity"] = seg_capacity
                    segment_total_rows[-1]["client_count"] = float(seg_df["client_count"].sum()) if "client_count" in seg_df.columns else 0.0
                    segment_total_rows[-1]["absolute_new_clients"] = float(seg_df["absolute_new_clients"].sum()) if "absolute_new_clients" in seg_df.columns else 0.0
                    segment_total_rows[-1]["returned_clients"] = float(seg_df["returned_clients"].sum()) if "returned_clients" in seg_df.columns else 0.0
                    segment_total_rows[-1]["new_clients"] = float(seg_df["new_clients"].sum()) if "new_clients" in seg_df.columns else 0.0
                    segment_total_rows[-1]["order_frequency"] = float(seg_df["order_frequency"].mean()) if "order_frequency" in seg_df.columns else 0.0
                    segment_total_rows[-1]["sov_pct"] = (float(segment_total_rows[-1].get("reach", 0.0)) / seg_capacity * 100.0) if seg_capacity > 0 else 0.0
                    segment_total_rows[-1]["new_clients_share_pct"] = float(seg_df["new_clients_share_pct"].mean()) if "new_clients_share_pct" in seg_df.columns else 0.0
                    segment_total_rows[-1]["cac"] = (seg_cost_with_vat_ak / float(segment_total_rows[-1]["new_clients"])) if float(segment_total_rows[-1]["new_clients"]) > 0 else 0.0
                    segment_total_rows[-1]["shipped_orders"] = float(seg_df["shipped_orders"].sum()) if "shipped_orders" in seg_df.columns else 0.0
                    segment_total_rows[-1]["shipped_revenue"] = float(seg_df["shipped_revenue"].sum()) if "shipped_revenue" in seg_df.columns else 0.0
                    segment_total_rows[-1]["shipped_cps"] = (seg_cost_with_vat / float(segment_total_rows[-1]["shipped_orders"])) if float(segment_total_rows[-1]["shipped_orders"]) > 0 else 0.0
                    segment_total_rows[-1]["shipped_aov"] = (float(segment_total_rows[-1]["shipped_revenue"]) / float(segment_total_rows[-1]["shipped_orders"])) if float(segment_total_rows[-1]["shipped_orders"]) > 0 else 0.0
                    segment_total_rows[-1]["shipped_roas"] = (float(segment_total_rows[-1]["shipped_revenue"]) / seg_budget_basis) if seg_budget_basis > 0 else 0.0
                    segment_total_rows[-1]["shipped_drr_pct"] = (seg_budget_basis / float(segment_total_rows[-1]["shipped_revenue"]) * 100.0) if float(segment_total_rows[-1]["shipped_revenue"]) > 0 else 0.0
                    segment_total_rows[-1]["order_share_segment_pct"] = 100.0
                    segment_total_rows[-1]["revenue_share_segment_pct"] = 100.0
                    segment_total_rows[-1]["shipped_order_share_segment_pct"] = 100.0
                    segment_total_rows[-1]["shipped_revenue_share_segment_pct"] = 100.0
                    segment_total_rows[-1]["budget_share_segment_pct"] = 100.0

        df_base_show = pd.concat(
            [df_base_show, pd.DataFrame(segment_total_rows + [total_row_raw])],
            ignore_index=True,
        )

        df_base_show["impressions"] = df_base_show["impressions"].map(
            lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
        )
        df_base_show["clicks"] = df_base_show["clicks"].map(
            lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
        )
        if "reach" in df_base_show.columns:
            df_base_show["reach"] = df_base_show["reach"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
            )
        if "frequency" in df_base_show.columns:
            df_base_show["frequency"] = df_base_show["frequency"].map(
                lambda x: "" if pd.isna(x) else f"{x:.2f}"
            )
        df_base_show["conversions"] = df_base_show["conversions"].map(
            lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
        )
        if "leads" in df_base_show.columns:
            df_base_show["leads"] = df_base_show["leads"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
            )
        if "target_leads" in df_base_show.columns:
            df_base_show["target_leads"] = df_base_show["target_leads"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
            )

        df_base_show["cost"] = df_base_show["cost"].map(
            lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
        )
        if "ak_cost_wo_vat" in df_base_show.columns:
            df_base_show["ak_cost_wo_vat"] = df_base_show["ak_cost_wo_vat"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
            )
        if "total_budget_wo_vat_ak" in df_base_show.columns:
            df_base_show["total_budget_wo_vat_ak"] = df_base_show["total_budget_wo_vat_ak"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
            )
        df_base_show["cost_with_vat"] = df_base_show["cost_with_vat"].map(
            lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
        )
        df_base_show["cost_with_vat_ak"] = df_base_show["cost_with_vat_ak"].map(
            lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
        )
        if "vat_amount" in df_base_show.columns:
            df_base_show["vat_amount"] = df_base_show["vat_amount"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
            )
        df_base_show["ak_rate_pct"] = df_base_show["ak_rate_pct"].map(
            lambda x: "" if pd.isna(x) else f"{x:.2f} %"
        )
        df_base_show["revenue"] = df_base_show["revenue"].map(
            lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
        )
        df_base_show["cpc"] = df_base_show["cpc"].map(
            lambda x: "" if pd.isna(x) else f"{x:.2f} в‚Ѕ".replace(",", " ")
        )
        df_base_show["cpm"] = df_base_show["cpm"].map(
            lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
        )
        df_base_show["cpa"] = df_base_show["cpa"].map(
            lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
        )
        df_base_show["aov"] = df_base_show["aov"].map(
            lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
        )
        if "cpl" in df_base_show.columns:
            df_base_show["cpl"] = df_base_show["cpl"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
            )
        if "cpql" in df_base_show.columns:
            df_base_show["cpql"] = df_base_show["cpql"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
            )

        df_base_show["ctr_pct"] = df_base_show["ctr_pct"].map(
            lambda x: "" if pd.isna(x) else f"{x:.2f} %"
        )
        df_base_show["cr_pct"] = df_base_show["cr_pct"].map(
            lambda x: "" if pd.isna(x) else f"{x:.2f} %"
        )
        if "cr1_pct" in df_base_show.columns:
            df_base_show["cr1_pct"] = df_base_show["cr1_pct"].map(
                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
            )
        if "cr2_pct" in df_base_show.columns:
            df_base_show["cr2_pct"] = df_base_show["cr2_pct"].map(
                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
            )
        df_base_show["drr_pct"] = df_base_show["drr_pct"].map(
            lambda x: "" if pd.isna(x) else f"{x:.2f} %"
        )
        df_base_show["ROAS"] = df_base_show["roas"].map(
            lambda x: "" if pd.isna(x) else f"{x * 100:.2f} %"
        )

        if is_real_estate_preset:
            base_show_cols = ["campaign_type", "system", "format", "geo"] + get_real_estate_table_cols(metric_mode)
        else:
            base_show_cols = [
                "campaign_type",
                "system",
                "format",
                "geo",
                "cost",
                "ak_cost_wo_vat",
                "total_budget_wo_vat_ak",
                "cost_with_vat",
                "cost_with_vat_ak",
                "vat_amount",
                "ak_rate_pct",
                "impressions",
                "reach",
                "frequency",
                "clicks",
                "cpc",
                "ctr_pct",
                "cr_pct",
                "conversions",
                "cpa",
                "aov",
                "revenue",
                "drr_pct",
                "ROAS",
            ]
        if is_diy_preset:
            base_show_cols.insert(3, "segment")
        if is_diy_preset:
            df_base_show["available_capacity"] = df_base_show["available_capacity"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
            )
            df_base_show["client_count"] = df_base_show["client_count"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
            )
            df_base_show["absolute_new_clients"] = df_base_show["absolute_new_clients"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
            )
            df_base_show["returned_clients"] = df_base_show["returned_clients"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
            )
            df_base_show["new_clients"] = df_base_show["new_clients"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
            )
            df_base_show["cac"] = df_base_show["cac"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
            )
            df_base_show["order_frequency"] = df_base_show["order_frequency"].map(
                lambda x: "" if pd.isna(x) else f"{x:.2f}"
            )
            df_base_show["sov_pct"] = df_base_show["sov_pct"].map(
                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
            )
            df_base_show["new_clients_share_pct"] = df_base_show["new_clients_share_pct"].map(
                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
            )
            df_base_show["order_share_segment_pct"] = df_base_show["order_share_segment_pct"].map(
                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
            )
            df_base_show["revenue_share_segment_pct"] = df_base_show["revenue_share_segment_pct"].map(
                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
            )
            df_base_show["budget_share_segment_pct"] = df_base_show["budget_share_segment_pct"].map(
                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
            )
            base_show_cols += ["available_capacity", "client_count", "absolute_new_clients", "returned_clients", "new_clients", "cac", "order_frequency", "shipped_orders", "shipped_cps", "shipped_aov", "shipped_revenue", "shipped_roas", "shipped_drr_pct", "shipped_order_share_segment_pct", "shipped_revenue_share_segment_pct", "sov_pct", "new_clients_share_pct", "order_share_segment_pct", "revenue_share_segment_pct", "budget_share_segment_pct"]
        # РџРѕСЂСЏРґРѕРє РјРµС‚СЂРёРє Р·Р°РґР°РЅ РїРѕРґ Р±РёР·РЅРµСЃ-Р»РѕРіРёРєСѓ РїСЂРѕРІРµСЂРєРё.
        df_base_show = safe_select_columns(df_base_show, base_show_cols, fill_value="")
        df_base_show = df_base_show.rename(columns=DISPLAY_COL_RENAME)
        if is_diy_preset:
            df_base_show = reorder_rows_with_segment_subtotals(
                df_base_show,
                DISPLAY_COL_RENAME["campaign_type"],
                DISPLAY_COL_RENAME["segment"],
            )
        # РџСЂРёРЅСѓРґРёС‚РµР»СЊРЅРѕ С„РёРєСЃРёСЂСѓРµРј Р·РЅР°С‡РµРЅРёСЏ РІ РїРѕСЃР»РµРґРЅРµР№ СЃС‚СЂРѕРєРµ РС‚РѕРіРѕ (РїРѕСЃР»Рµ РІСЃРµС… РїСЂРµРѕР±СЂР°Р·РѕРІР°РЅРёР№).
        if len(df_base_show) > 0:
            li = len(df_base_show) - 1
            df_base_show.at[li, DISPLAY_COL_RENAME["campaign_type"]] = "РС‚РѕРіРѕ"
            if DISPLAY_COL_RENAME.get("segment") in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["segment"]] = "ALL"
            df_base_show.at[li, DISPLAY_COL_RENAME["system"]] = ""
            df_base_show.at[li, DISPLAY_COL_RENAME["format"]] = ""
            if DISPLAY_COL_RENAME["geo"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["geo"]] = ""
            if DISPLAY_COL_RENAME["impressions"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["impressions"]] = f"{round(total_imp):,}".replace(",", " ")
            if DISPLAY_COL_RENAME["clicks"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["clicks"]] = f"{round(total_clicks):,}".replace(",", " ")
            if DISPLAY_COL_RENAME["reach"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["reach"]] = f"{round(total_reach):,}".replace(",", " ")
            if DISPLAY_COL_RENAME["frequency"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["frequency"]] = f"{total_frequency:.2f}"
            if DISPLAY_COL_RENAME["conversions"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["conversions"]] = f"{round(total_conv):,}".replace(",", " ")
            if DISPLAY_COL_RENAME["leads"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["leads"]] = f"{round(total_leads):,}".replace(",", " ")
            if DISPLAY_COL_RENAME["target_leads"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["target_leads"]] = f"{round(total_target_leads):,}".replace(",", " ")
            if DISPLAY_COL_RENAME["cost"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["cost"]] = f"{round(total_cost):,} в‚Ѕ".replace(",", " ")
            if DISPLAY_COL_RENAME["ak_cost_wo_vat"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["ak_cost_wo_vat"]] = f"{round(total_ak_wo_vat):,} в‚Ѕ".replace(",", " ")
            if DISPLAY_COL_RENAME["total_budget_wo_vat_ak"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["total_budget_wo_vat_ak"]] = f"{round(total_cost + total_ak_wo_vat):,} в‚Ѕ".replace(",", " ")
            if DISPLAY_COL_RENAME["cost_with_vat"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["cost_with_vat"]] = f"{round(total_cost_with_vat):,} в‚Ѕ".replace(",", " ")
            if DISPLAY_COL_RENAME["cost_with_vat_ak"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["cost_with_vat_ak"]] = f"{round(total_cost_with_vat_ak):,} в‚Ѕ".replace(",", " ")
            if DISPLAY_COL_RENAME["vat_amount"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["vat_amount"]] = f"{round(total_cost_with_vat_ak - (total_cost + total_ak_wo_vat)):,} в‚Ѕ".replace(",", " ")
            total_ak_rate_pct = (total_ak_wo_vat / total_cost * 100.0) if total_cost > 0 else 0.0
            if DISPLAY_COL_RENAME["ak_rate_pct"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["ak_rate_pct"]] = f"{total_ak_rate_pct:.2f} %"
            if DISPLAY_COL_RENAME["revenue"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["revenue"]] = f"{round(total_rev):,} в‚Ѕ".replace(",", " ")
            if DISPLAY_COL_RENAME["ctr"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["ctr"]] = f"{total_ctr:.2f} %"
            if DISPLAY_COL_RENAME["cpc"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["cpc"]] = f"{total_cpc:.2f} в‚Ѕ"
            if DISPLAY_COL_RENAME["cr"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["cr"]] = f"{total_cr:.2f} %"
            if DISPLAY_COL_RENAME["cr1_pct"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["cr1_pct"]] = f"{total_cr1:.2f} %"
            if DISPLAY_COL_RENAME["cr2_pct"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["cr2_pct"]] = f"{total_cr2:.2f} %"
            if DISPLAY_COL_RENAME["cpa"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["cpa"]] = f"{round(total_cpa):,} в‚Ѕ".replace(",", " ")
            if DISPLAY_COL_RENAME["cpl"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["cpl"]] = f"{round(total_cpl):,} в‚Ѕ".replace(",", " ")
            if DISPLAY_COL_RENAME["cpql"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["cpql"]] = f"{round(total_cpql):,} в‚Ѕ".replace(",", " ")
            if "ROAS" in df_base_show.columns:
                df_base_show.at[li, "ROAS"] = f"{total_roas * 100:.2f} %"
            if DISPLAY_COL_RENAME["drr"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["drr"]] = f"{total_drr:.2f} %"
            if DISPLAY_COL_RENAME["aov"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["aov"]] = ""
            if DISPLAY_COL_RENAME["available_capacity"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["available_capacity"]] = f"{round(total_row_raw.get('available_capacity', 0.0)):,}".replace(",", " ")
            if DISPLAY_COL_RENAME["client_count"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["client_count"]] = f"{round(total_row_raw.get('client_count', 0.0)):,}".replace(",", " ")
            if DISPLAY_COL_RENAME["absolute_new_clients"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["absolute_new_clients"]] = f"{round(total_row_raw.get('absolute_new_clients', 0.0)):,}".replace(",", " ")
            if DISPLAY_COL_RENAME["returned_clients"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["returned_clients"]] = f"{round(total_row_raw.get('returned_clients', 0.0)):,}".replace(",", " ")
            if DISPLAY_COL_RENAME["new_clients"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["new_clients"]] = f"{round(total_row_raw.get('new_clients', 0.0)):,}".replace(",", " ")
            if DISPLAY_COL_RENAME["cac"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["cac"]] = f"{round(total_row_raw.get('cac', 0.0)):,} в‚Ѕ".replace(",", " ")
            if DISPLAY_COL_RENAME["order_frequency"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["order_frequency"]] = f"{float(total_row_raw.get('order_frequency', 0.0)):.2f}"
            if DISPLAY_COL_RENAME["sov_pct"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["sov_pct"]] = f"{float(total_row_raw.get('sov_pct', 0.0)):.2f} %"
            if DISPLAY_COL_RENAME["new_clients_share_pct"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["new_clients_share_pct"]] = f"{float(total_row_raw.get('new_clients_share_pct', 0.0)):.2f} %"
            if DISPLAY_COL_RENAME["order_share_segment_pct"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["order_share_segment_pct"]] = "100.00 %"
            if DISPLAY_COL_RENAME["revenue_share_segment_pct"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["revenue_share_segment_pct"]] = "100.00 %"
            if DISPLAY_COL_RENAME["budget_share_segment_pct"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["budget_share_segment_pct"]] = "100.00 %"

        def highlight_total_row(row):
            style = [""] * len(row)
            camp_col = DISPLAY_COL_RENAME.get("campaign_type", "РќР°Р·РІР°РЅРёРµ РєР°РјРїР°РЅРёРё")
            camp_val = str(row.get(camp_col, ""))
            if camp_val == "РС‚РѕРіРѕ":
                style = [
                    f"background-color: #00CDC5; color: #081521; font-weight: 700; border-top: 2px solid {THEME_BORDER};"
                ] * len(row)
            elif camp_val.startswith("РС‚РѕРіРѕ "):
                style = [
                    f"background-color: #2C6E75; color: #DDEAF0; font-weight: 650; border-top: 1px solid {THEME_BORDER};"
                ] * len(row)
            return style

        styled_base = df_base_show.style.apply(highlight_total_row, axis=1)
        st.dataframe(styled_base, use_container_width=True)

    # ---------- 3. РљРѕСЌС„С„РёС†РёРµРЅС‚С‹ Рё СЂРµР·СѓР»СЊС‚Р°С‚С‹ РїРѕ РјРµСЃСЏС†Р°Рј ----------

    with st.expander("3. РљРѕСЌС„С„РёС†РёРµРЅС‚С‹ Рё СЂРµР·СѓР»СЊС‚Р°С‚С‹ РїРѕ РјРµСЃСЏС†Р°Рј", expanded=True):
        show_month_details = True

        coeff_values = {}
        for cs in coeff_sets:
            cs_id = get_coeff_set_id_value(cs)
            if not cs_id:
                continue
            cs_type = normalize_coeff_set_type(cs.get("type", "РЎРїСЂРѕСЃ (РїРѕ Р·Р°РїСЂРѕСЃР°Рј)"))
            df_res = cs.get("result")
            if df_res is None or df_res.empty:
                continue

            if cs_type == "РЎРїСЂРѕСЃ (РїРѕ Р·Р°РїСЂРѕСЃР°Рј)":
                col_coeff = "РљРѕСЌС„."
                metric_type = "demand"
            elif cs_type == "AOV (СЃСЂРµРґРЅРёР№ С‡РµРє)":
                col_coeff = "РљРѕСЌС„. AOV"
                metric_type = "aov"
            elif cs_type == "РњРµРґРёР№РЅС‹Рµ С…РІРѕСЃС‚С‹":
                col_coeff = "РљРѕСЌС„."
                metric_type = "media_tail"
            else:
                col_coeff = "РљРѕСЌС„."
                metric_type = "custom"

            if col_coeff not in df_res.columns or "РќРѕРјРµСЂ РјРµСЃСЏС†Р°" not in df_res.columns:
                continue

            for _, r in df_res.iterrows():
                try:
                    m = int(r["РќРѕРјРµСЂ РјРµСЃСЏС†Р°"])
                    k = float(r[col_coeff])
                except Exception:
                    continue
                if metric_type == "custom":
                    coeff_values[(cs_id, m, "demand")] = k
                    coeff_values[(cs_id, m, "aov")] = k
                else:
                    coeff_values[(cs_id, m, metric_type)] = k

        def get_k_demand(campaign_type: str, month_num: int) -> float:
            set_id = demand_link_map.get(campaign_type)
            if not set_id:
                return 1.0
            return coeff_values.get((set_id, month_num, "demand"), 1.0)

        def get_k_aov(campaign_type: str, month_num: int) -> float:
            set_id = aov_link_map.get(campaign_type)
            if not set_id:
                return 1.0
            return coeff_values.get((set_id, month_num, "aov"), 1.0)

        def get_k_media_tail(campaign_type: str, month_num: int) -> float:
            set_id = media_tail_link_map.get(campaign_type)
            if not set_id:
                return 1.0
            return coeff_values.get((set_id, month_num, "media_tail"), 1.0)

        def get_k_capacity(campaign_type: str, month_num: int) -> float:
            set_id = capacity_link_map.get(campaign_type)
            if not set_id:
                set_id = demand_link_map.get(campaign_type, "")
            if not set_id:
                return 1.0
            if (set_id, month_num, "demand") in coeff_values:
                return coeff_values[(set_id, month_num, "demand")]
            if (set_id, month_num, "media_tail") in coeff_values:
                return coeff_values[(set_id, month_num, "media_tail")]
            if (set_id, month_num, "aov") in coeff_values:
                return coeff_values[(set_id, month_num, "aov")]
            return 1.0

        def get_k_client_count(campaign_type: str, month_num: int) -> float:
            set_id = client_count_link_map.get(campaign_type)
            if not set_id:
                set_id = capacity_link_map.get(campaign_type, "")
            if not set_id:
                set_id = demand_link_map.get(campaign_type, "")
            if not set_id:
                return 1.0
            if (set_id, month_num, "demand") in coeff_values:
                return coeff_values[(set_id, month_num, "demand")]
            if (set_id, month_num, "media_tail") in coeff_values:
                return coeff_values[(set_id, month_num, "media_tail")]
            if (set_id, month_num, "aov") in coeff_values:
                return coeff_values[(set_id, month_num, "aov")]
            return 1.0

        def get_k_absolute_new_clients(campaign_type: str, month_num: int) -> float:
            set_id = absolute_new_clients_link_map.get(campaign_type)
            if not set_id:
                set_id = client_count_link_map.get(campaign_type, "")
            if not set_id:
                set_id = capacity_link_map.get(campaign_type, "")
            if not set_id:
                set_id = demand_link_map.get(campaign_type, "")
            if not set_id:
                return 1.0
            if (set_id, month_num, "demand") in coeff_values:
                return coeff_values[(set_id, month_num, "demand")]
            if (set_id, month_num, "media_tail") in coeff_values:
                return coeff_values[(set_id, month_num, "media_tail")]
            if (set_id, month_num, "aov") in coeff_values:
                return coeff_values[(set_id, month_num, "aov")]
            return 1.0

        def get_k_returned_clients(campaign_type: str, month_num: int) -> float:
            set_id = returned_clients_link_map.get(campaign_type)
            if not set_id:
                set_id = client_count_link_map.get(campaign_type, "")
            if not set_id:
                set_id = capacity_link_map.get(campaign_type, "")
            if not set_id:
                set_id = demand_link_map.get(campaign_type, "")
            if not set_id:
                return 1.0
            if (set_id, month_num, "demand") in coeff_values:
                return coeff_values[(set_id, month_num, "demand")]
            if (set_id, month_num, "media_tail") in coeff_values:
                return coeff_values[(set_id, month_num, "media_tail")]
            if (set_id, month_num, "aov") in coeff_values:
                return coeff_values[(set_id, month_num, "aov")]
            return 1.0

        def get_k_order_frequency(campaign_type: str, month_num: int) -> float:
            set_id = order_frequency_link_map.get(campaign_type)
            if not set_id:
                set_id = client_count_link_map.get(campaign_type, "")
            if not set_id:
                set_id = capacity_link_map.get(campaign_type, "")
            if not set_id:
                set_id = demand_link_map.get(campaign_type, "")
            if not set_id:
                return 1.0
            if (set_id, month_num, "demand") in coeff_values:
                return coeff_values[(set_id, month_num, "demand")]
            if (set_id, month_num, "media_tail") in coeff_values:
                return coeff_values[(set_id, month_num, "media_tail")]
            if (set_id, month_num, "aov") in coeff_values:
                return coeff_values[(set_id, month_num, "aov")]
            return 1.0

        def get_k_reach(campaign_type: str, month_num: int) -> float:
            set_id = demand_link_map.get(campaign_type, "")
            if not set_id:
                set_id = media_tail_link_map.get(campaign_type, "")
            if not set_id:
                return 1.0
            if (set_id, month_num, "demand") in coeff_values:
                return coeff_values[(set_id, month_num, "demand")]
            if (set_id, month_num, "media_tail") in coeff_values:
                return coeff_values[(set_id, month_num, "media_tail")]
            return 1.0

        all_months_results = []

        for month in selected_month_nums:
            month_name = month_names_full[month]

            coeff_rows = []
            base_capacity_map = {
                str(r["campaign_type"]): float(r.get("available_capacity_avg", 0.0) or 0.0)
                for _, r in campaigns.iterrows()
            }
            base_client_count_map = {
                str(r["campaign_type"]): float(r.get("client_count_avg", 0.0) or 0.0)
                for _, r in campaigns.iterrows()
            }
            base_absolute_new_clients_map = {
                str(r["campaign_type"]): float(r.get("absolute_new_clients_avg", 0.0) or 0.0)
                for _, r in campaigns.iterrows()
            }
            base_returned_clients_map = {
                str(r["campaign_type"]): float(r.get("returned_clients_avg", 0.0) or 0.0)
                for _, r in campaigns.iterrows()
            }
            base_order_frequency_map = {
                str(r["campaign_type"]): float(r.get("order_frequency_avg", 0.0) or 0.0)
                for _, r in campaigns.iterrows()
            }
            for _, base_row in campaigns.iterrows():
                camp_type = str(base_row["campaign_type"])

                k_demand = get_k_demand(camp_type, month)
                k_media_tail = get_k_media_tail(camp_type, month)
                el = elasticity_map.get(
                    camp_type,
                    {"cpc_div": 1.0, "ctr_div": 2.0, "cr_div": 10.0},
                )

                k_imp = k_demand * k_media_tail
                demand_delta = (k_demand - 1.0)
                k_cpc = 1.0 + demand_delta / el["cpc_div"]
                k_ctr = 1.0 - demand_delta / el["ctr_div"]
                k_cr = 1.0 - demand_delta / el["cr_div"]
                k_cr2 = 1.0 - demand_delta / el.get("cr2_div", el["cr_div"])
                k_ctr = max(0.0, k_ctr)
                k_cr = max(0.0, k_cr)
                k_cr2 = max(0.0, k_cr2)
                k_aov = get_k_aov(camp_type, month)
                k_capacity = get_k_capacity(camp_type, month)
                k_client_count = get_k_client_count(camp_type, month)
                k_absolute_new_clients = get_k_absolute_new_clients(camp_type, month)
                k_returned_clients = get_k_returned_clients(camp_type, month)
                k_order_frequency = get_k_order_frequency(camp_type, month)
                k_reach = get_k_reach(camp_type, month)

                coeff_row = {
                    "campaign_type": camp_type,
                    "k_imp": k_imp,
                    "k_reach": k_reach,
                    "k_capacity": k_capacity,
                    "available_capacity": float(base_row.get("available_capacity_avg", 0.0) or 0.0) * k_capacity,
                    "k_client_count": k_client_count,
                    "client_count": float(base_row.get("client_count_avg", 0.0) or 0.0) * k_client_count,
                    "k_absolute_new_clients": k_absolute_new_clients,
                    "absolute_new_clients": float(base_row.get("absolute_new_clients_avg", 0.0) or 0.0) * k_absolute_new_clients,
                    "k_returned_clients": k_returned_clients,
                    "returned_clients": float(base_row.get("returned_clients_avg", 0.0) or 0.0) * k_returned_clients,
                    "new_clients": (float(base_row.get("absolute_new_clients_avg", 0.0) or 0.0) * k_absolute_new_clients)
                    + (float(base_row.get("returned_clients_avg", 0.0) or 0.0) * k_returned_clients),
                    "k_order_frequency": k_order_frequency,
                    "order_frequency": float(base_row.get("order_frequency_avg", 0.0) or 0.0) * k_order_frequency,
                    "new_clients_share_pct": float(base_row.get("new_clients_share_avg_percent", 0.0) or 0.0),
                    "k_ctr": k_ctr,
                    "k_cpc": k_cpc,
                    "k_cr": k_cr,
                    "k_aov": k_aov,
                }
                if is_real_estate_preset and metric_mode["is_real_estate_full"]:
                    coeff_row["k_cr2"] = k_cr2
                coeff_rows.append(coeff_row)

            coeffs_default = pd.DataFrame(coeff_rows)
            if is_real_estate_preset:
                coeffs_default = coeffs_default.drop(columns=["k_aov"], errors="ignore")
            if not is_diy_preset:
                coeffs_default = coeffs_default.drop(
                    columns=["k_reach", "k_capacity", "available_capacity", "k_client_count", "client_count", "k_absolute_new_clients", "absolute_new_clients", "k_returned_clients", "returned_clients", "new_clients", "k_order_frequency", "order_frequency", "new_clients_share_pct"],
                    errors="ignore",
                )
            else:
                coeffs_default = coeffs_default[
                    [
                        "campaign_type",
                        "k_imp",
                        "k_reach",
                        "k_cpc",
                        "k_ctr",
                        "k_cr",
                        "k_aov",
                        "k_capacity",
                        "available_capacity",
                        "k_client_count",
                        "client_count",
                        "k_absolute_new_clients",
                        "absolute_new_clients",
                        "k_returned_clients",
                        "returned_clients",
                        "new_clients",
                        "k_order_frequency",
                        "order_frequency",
                        "new_clients_share_pct",
                    ]
                ]
            coeffs_month = coeffs_default.copy()
            coeffs_month["campaign_type"] = coeffs_month["campaign_type"].astype(str)

            rows = []
            # РµСЃР»Рё С…РѕС‚РёРј РїРѕРєР°Р·С‹РІР°С‚СЊ РґРµС‚Р°Р»Рё РїРѕ РјРµСЃСЏС†Р°Рј, СЃРѕР·РґР°С‘Рј СЂРµРґР°РєС‚РѕСЂ
            if show_month_details:
                st.markdown(f"#### РњРµСЃСЏС† {month}: {month_name}")
                col_left, col_right = st.columns([1.8, 1.0], vertical_alignment="top")
                month_row_height = max(36, int(ui_table_row_px))
                month_rows_count = max(len(coeffs_month), 1)
                month_table_height = max(140, min(520, 38 + month_rows_count * month_row_height + 6))
                with col_right:
                    edited_coeffs_month = st.data_editor(
                        coeffs_month,
                        num_rows="fixed",
                        use_container_width=True,
                        row_height=month_row_height,
                        height=month_table_height,
                        key=f"coeffs_month_{month}_v2",
                        column_config={
                            "campaign_type": st.column_config.TextColumn(
                                "РўРёРї Р Рљ / РќР°Р·РІР°РЅРёРµ", disabled=True, width="medium"
                            ),
                            "k_imp": st.column_config.NumberColumn(
                                "k_imp", format="%.2f", width="small",
                                help=mhelp("k_imp")
                            ),
                            "k_cpc": st.column_config.NumberColumn(
                                "k_cpc", format="%.2f", width="small",
                                help=mhelp("k_cpc")
                            ),
                            "k_reach": st.column_config.NumberColumn(
                                "k_reach", format="%.2f", width="small",
                                help=mhelp("k_reach")
                            ),
                            "k_ctr": st.column_config.NumberColumn(
                                "k_ctr", format="%.2f", width="small",
                                help=mhelp("k_ctr")
                            ),
                            "k_cr": st.column_config.NumberColumn(
                                "k_cr", format="%.2f", width="small",
                                help=mhelp("k_cr")
                            ),
                            "k_cr2": st.column_config.NumberColumn(
                                "k_cr2", format="%.2f", width="small",
                                help="РљРѕСЌС„С„РёС†РёРµРЅС‚ РІР»РёСЏРЅРёСЏ СЃРµР·РѕРЅРЅРѕСЃС‚Рё РЅР° CR2 РІ Р¦Рћ."
                            ),
                            "k_aov": st.column_config.NumberColumn(
                                "k_aov", format="%.2f", width="small",
                                help=mhelp("k_aov")
                            ),
                            "k_capacity": st.column_config.NumberColumn(
                                "k_capacity", format="%.2f", width="small",
                                disabled=not is_diy_preset,
                                help=mhelp("k_capacity"),
                            ),
                            "available_capacity": st.column_config.NumberColumn(
                                "Р”РѕСЃС‚СѓРїРЅР°СЏ РµРјРєРѕСЃС‚СЊ", format="%.0f", width="small",
                                disabled=True,
                            ),
                            "k_client_count": st.column_config.NumberColumn(
                                "k_client_count", format="%.2f", width="small",
                                disabled=not is_diy_preset,
                                help=mhelp("k_client_count"),
                            ),
                            "client_count": st.column_config.NumberColumn(
                                "РљРѕР»РёС‡РµСЃС‚РІРѕ РєР»РёРµРЅС‚РѕРІ", format="%.0f", width="small",
                                disabled=True,
                            ),
                            "k_absolute_new_clients": st.column_config.NumberColumn(
                                "k_absolute_new_clients", format="%.2f", width="small",
                                disabled=not is_diy_preset,
                                help=mhelp("k_absolute_new_clients"),
                            ),
                            "absolute_new_clients": st.column_config.NumberColumn(
                                "РђР±СЃРѕР»СЋС‚РЅРѕ РЅРѕРІС‹Рµ", format="%.0f", width="small",
                                disabled=True,
                            ),
                            "k_returned_clients": st.column_config.NumberColumn(
                                "k_returned_clients", format="%.2f", width="small",
                                disabled=not is_diy_preset,
                                help=mhelp("k_returned_clients"),
                            ),
                            "returned_clients": st.column_config.NumberColumn(
                                "Р’РµСЂРЅСѓРІС€РёРµСЃСЏ", format="%.0f", width="small",
                                disabled=True,
                            ),
                            "new_clients": st.column_config.NumberColumn(
                                "РќРѕРІС‹Рµ РєР»РёРµРЅС‚С‹", format="%.0f", width="small",
                                disabled=True,
                            ),
                            "k_order_frequency": st.column_config.NumberColumn(
                                "k_order_frequency", format="%.2f", width="small",
                                disabled=not is_diy_preset,
                                help=mhelp("k_order_frequency"),
                            ),
                            "order_frequency": st.column_config.NumberColumn(
                                "Р§Р°СЃС‚РѕС‚Р° Р·Р°РєР°Р·Р°", format="%.2f", width="small",
                                disabled=True,
                            ),
                            "new_clients_share_pct": st.column_config.NumberColumn(
                                "Р”РѕР»СЏ РЅРѕРІС‹С… РєР»РёРµРЅС‚РѕРІ, %", format="%.2f", width="small",
                                disabled=not is_diy_preset,
                            ),
                        },
                    )
                    edited_coeffs_month = edited_coeffs_month.copy()
                    if is_real_estate_preset and "k_aov" in edited_coeffs_month.columns:
                        edited_coeffs_month = edited_coeffs_month.drop(columns=["k_aov"], errors="ignore")
                    if not ((is_real_estate_preset and metric_mode["is_real_estate_full"]) or is_diy_preset) and "k_cr2" in edited_coeffs_month.columns:
                        edited_coeffs_month = edited_coeffs_month.drop(columns=["k_cr2"], errors="ignore")
                    edited_coeffs_month["campaign_type"] = edited_coeffs_month["campaign_type"].astype(str)
                    if is_diy_preset and "available_capacity" in edited_coeffs_month.columns:
                        edited_coeffs_month["available_capacity"] = edited_coeffs_month.apply(
                            lambda r: float(base_capacity_map.get(str(r.get("campaign_type", "")), 0.0))
                            * float(r.get("k_capacity", 1.0) or 1.0),
                            axis=1,
                        )
                    if is_diy_preset and "client_count" in edited_coeffs_month.columns:
                        edited_coeffs_month["client_count"] = edited_coeffs_month.apply(
                            lambda r: float(base_client_count_map.get(str(r.get("campaign_type", "")), 0.0))
                            * float(r.get("k_client_count", 1.0) or 1.0),
                            axis=1,
                        )
                    if is_diy_preset and "absolute_new_clients" in edited_coeffs_month.columns:
                        edited_coeffs_month["absolute_new_clients"] = edited_coeffs_month.apply(
                            lambda r: float(base_absolute_new_clients_map.get(str(r.get("campaign_type", "")), 0.0))
                            * float(r.get("k_absolute_new_clients", 1.0) or 1.0),
                            axis=1,
                        )
                    if is_diy_preset and "returned_clients" in edited_coeffs_month.columns:
                        edited_coeffs_month["returned_clients"] = edited_coeffs_month.apply(
                            lambda r: float(base_returned_clients_map.get(str(r.get("campaign_type", "")), 0.0))
                            * float(r.get("k_returned_clients", 1.0) or 1.0),
                            axis=1,
                        )
                    if is_diy_preset and "new_clients" in edited_coeffs_month.columns:
                        edited_coeffs_month["new_clients"] = (
                            pd.to_numeric(edited_coeffs_month.get("absolute_new_clients", 0.0), errors="coerce").fillna(0.0)
                            + pd.to_numeric(edited_coeffs_month.get("returned_clients", 0.0), errors="coerce").fillna(0.0)
                        )
                    if is_diy_preset and "order_frequency" in edited_coeffs_month.columns:
                        edited_coeffs_month["order_frequency"] = edited_coeffs_month.apply(
                            lambda r: float(base_order_frequency_map.get(str(r.get("campaign_type", "")), 0.0))
                            * float(r.get("k_order_frequency", 1.0) or 1.0),
                            axis=1,
                        )
                    coeffs_month = edited_coeffs_month
            else:
                # РµСЃР»Рё РїРѕРґСЂРѕР±РЅРѕСЃС‚Рё СЃРєСЂС‹С‚С‹ вЂ” Р±РµСЂС‘Рј coeffs_default РєР°Рє РµСЃС‚СЊ
                coeffs_month = coeffs_month.copy()
                col_left = st  # Р·Р°РіР»СѓС€РєР°, С‡С‚РѕР±С‹ РЅРёР¶Рµ РЅРµ РїР°РґР°Р»Рѕ

            for _, base_row in campaigns.iterrows():
                campaign_type = str(base_row["campaign_type"])

                k_row = coeffs_month[coeffs_month["campaign_type"] == campaign_type]
                if k_row.empty:
                    k_imp = k_reach = k_capacity = k_client_count = k_absolute_new_clients = k_returned_clients = k_order_frequency = k_ctr = k_cpc = k_cr = k_cr2 = k_aov = 1.0
                    available_capacity_month = float(base_row.get("available_capacity_avg", 0.0) or 0.0)
                    client_count_month = float(base_row.get("client_count_avg", 0.0) or 0.0)
                    absolute_new_clients_month = float(base_row.get("absolute_new_clients_avg", 0.0) or 0.0)
                    returned_clients_month = float(base_row.get("returned_clients_avg", 0.0) or 0.0)
                    order_frequency_month = float(base_row.get("order_frequency_avg", 0.0) or 0.0)
                    new_clients_share_month = 0.0
                else:
                    k_row = k_row.iloc[0]
                    k_imp = float(k_row["k_imp"])
                    k_reach = float(k_row.get("k_reach", k_imp) or k_imp)
                    k_capacity = float(k_row.get("k_capacity", 1.0) or 1.0)
                    k_client_count = float(k_row.get("k_client_count", 1.0) or 1.0)
                    k_absolute_new_clients = float(k_row.get("k_absolute_new_clients", 1.0) or 1.0)
                    k_returned_clients = float(k_row.get("k_returned_clients", 1.0) or 1.0)
                    k_order_frequency = float(k_row.get("k_order_frequency", 1.0) or 1.0)
                    k_ctr = float(k_row["k_ctr"])
                    k_cpc = float(k_row["k_cpc"])
                    k_cr = float(k_row["k_cr"])
                    k_cr2 = float(k_row.get("k_cr2", k_cr) or k_cr)
                    k_aov = float(k_row.get("k_aov", 1.0) or 1.0)
                    available_capacity_month = float(base_row.get("available_capacity_avg", 0.0) or 0.0) * k_capacity
                    client_count_month = float(base_row.get("client_count_avg", 0.0) or 0.0) * k_client_count
                    absolute_new_clients_month = float(base_row.get("absolute_new_clients_avg", 0.0) or 0.0) * k_absolute_new_clients
                    returned_clients_month = float(base_row.get("returned_clients_avg", 0.0) or 0.0) * k_returned_clients
                    order_frequency_month = float(base_row.get("order_frequency_avg", 0.0) or 0.0) * k_order_frequency
                    new_clients_share_month = float(k_row.get("new_clients_share_pct", 0.0) or 0.0)

                base = PlanInput(
                    impressions=base_row["impressions_avg"],
                    ctr=base_row["ctr_avg_percent"] / 100.0,
                    cpc=base_row["cpc_avg"],
                    cr=base_row["cr_avg_percent"] / 100.0,
                    aov=base_row["aov_avg"],
                    reach=float(base_row.get("reach_avg", 0.0) or 0.0),
                    cr2=float(base_row.get("cr2_avg_percent", 0.0) or 0.0) / 100.0,
                    preset_key=active_preset_key,
                    funnel_mode=metric_mode["real_estate_funnel_mode"],
                )

                month_inp = PlanInput(
                    impressions=base.impressions * k_imp,
                    ctr=base.ctr * k_ctr,
                    cpc=base.cpc * k_cpc,
                    cr=base.cr * k_cr,
                    aov=base.aov * k_aov,
                    reach=base.reach * k_reach,
                    cr2=base.cr2 * k_cr2 if ((is_real_estate_preset and metric_mode["is_real_estate_full"]) or is_diy_preset) else base.cr2,
                    preset_key=active_preset_key,
                    funnel_mode=metric_mode["real_estate_funnel_mode"],
                )

                out = calculate_plan_month(month_inp)
                out["month_num"] = month
                out["month_name"] = month_name
                out["campaign_type"] = campaign_type
                out["segment"] = str(base_row.get("segment", "B2C"))
                out["system"] = base_row["system"]
                out["format"] = base_row["format"]
                out["geo"] = str(base_row.get("geo", "") or "")
                out["k_demand_applied"] = float(k_demand)
                if is_diy_preset:
                    out["revenue"] = float(out.get("revenue", 0.0)) * (1.0 + VAT_RATE)
                if is_diy_preset:
                    out["available_capacity"] = available_capacity_month
                    out["client_count"] = client_count_month
                    out["absolute_new_clients"] = absolute_new_clients_month
                    out["returned_clients"] = returned_clients_month
                    out["new_clients"] = absolute_new_clients_month + returned_clients_month
                    out["order_frequency"] = order_frequency_month
                    out["sov_pct"] = (float(out.get("reach", 0.0)) / available_capacity_month * 100.0) if available_capacity_month > 0 else 0.0
                    out["new_clients_share_pct"] = new_clients_share_month
                    out["cac"] = (float(out.get("cost_with_vat_ak", 0.0)) / float(out.get("new_clients", 0.0))) if float(out.get("new_clients", 0.0)) > 0 else 0.0
                    out["shipped_orders"] = float(out.get("target_leads", 0.0) or 0.0)
                    out["shipped_revenue"] = float(out.get("shipped_orders", 0.0)) * float(out.get("aov", 0.0) or 0.0)
                rows.append(out)

            df_month = pd.DataFrame(rows)
            if not df_month.empty:
                month_total_wo_vat = float(df_month["cost"].sum())
                month_ak_rate = resolve_ak_rate(month_total_wo_vat, ak_rules_df) if (use_ak_budget_metrics and ak_mode == "percent") else 0.0
                df_month = apply_budget_basis_metrics(
                    df_month,
                    use_vat_budget_metrics,
                    use_ak=use_ak_budget_metrics,
                    ak_mode=ak_mode,
                    default_ak_rate=month_ak_rate,
                    default_ak_fixed_wo_vat=float(ak_fixed_month_wo_vat),
                    default_ak_fixed_rate=float(ak_fixed_percent) / 100.0,
                )
                df_month["ak_rate_pct"] = df_month["ak_rate"] * 100.0
                if is_diy_preset:
                    df_month["cpa"] = np.where(
                        pd.to_numeric(df_month.get("conversions", 0.0), errors="coerce").fillna(0.0) > 0,
                        pd.to_numeric(df_month.get("cost_with_vat", 0.0), errors="coerce").fillna(0.0)
                        / pd.to_numeric(df_month.get("conversions", 0.0), errors="coerce").fillna(0.0),
                        0.0,
                    )
                    df_month["shipped_cps"] = np.where(
                        pd.to_numeric(df_month.get("shipped_orders", 0.0), errors="coerce").fillna(0.0) > 0,
                        pd.to_numeric(df_month.get("cost_with_vat", 0.0), errors="coerce").fillna(0.0)
                        / pd.to_numeric(df_month.get("shipped_orders", 0.0), errors="coerce").fillna(0.0),
                        0.0,
                    )
                    df_month["shipped_aov"] = np.where(
                        pd.to_numeric(df_month.get("shipped_orders", 0.0), errors="coerce").fillna(0.0) > 0,
                        pd.to_numeric(df_month.get("shipped_revenue", 0.0), errors="coerce").fillna(0.0)
                        / pd.to_numeric(df_month.get("shipped_orders", 0.0), errors="coerce").fillna(0.0),
                        0.0,
                    )
                    shipped_budget_basis_month = pd.to_numeric(
                        df_month.get(
                            "cost_with_vat_ak" if use_vat_budget_metrics and use_ak_budget_metrics else "cost_with_vat" if use_vat_budget_metrics else "total_budget_wo_vat_ak" if use_ak_budget_metrics else "cost",
                            0.0,
                        ),
                        errors="coerce",
                    ).fillna(0.0)
                    df_month["shipped_roas"] = np.where(
                        shipped_budget_basis_month > 0,
                        pd.to_numeric(df_month.get("shipped_revenue", 0.0), errors="coerce").fillna(0.0) / shipped_budget_basis_month,
                        0.0,
                    )
                    df_month["shipped_drr_pct"] = np.where(
                        pd.to_numeric(df_month.get("shipped_revenue", 0.0), errors="coerce").fillna(0.0) > 0,
                        shipped_budget_basis_month / pd.to_numeric(df_month.get("shipped_revenue", 0.0), errors="coerce").fillna(0.0) * 100.0,
                        0.0,
                    )
                    df_month["cac"] = np.where(
                        pd.to_numeric(df_month.get("new_clients", 0.0), errors="coerce").fillna(0.0) > 0,
                        pd.to_numeric(df_month.get("cost_with_vat_ak", 0.0), errors="coerce").fillna(0.0)
                        / pd.to_numeric(df_month.get("new_clients", 0.0), errors="coerce").fillna(0.0),
                        0.0,
                    )
                    df_month = add_segment_budget_share(df_month)
                    df_month = add_segment_value_share(df_month, value_col="conversions", out_col="order_share_segment_pct")
                    df_month = add_segment_value_share(df_month, value_col="revenue", out_col="revenue_share_segment_pct")
                    df_month = add_segment_value_share(df_month, value_col="shipped_orders", out_col="shipped_order_share_segment_pct")
                    df_month = add_segment_value_share(df_month, value_col="shipped_revenue", out_col="shipped_revenue_share_segment_pct")
            all_months_results.append(df_month)

            if show_month_details:
                with col_left:
                    if not df_month.empty:
                        df_rows_show = df_month.copy()
                        if is_real_estate_preset:
                            df_rows_show = compute_real_estate_rates(df_rows_show)

                        df_rows_show["ctr_pct"] = df_rows_show["ctr"] * 100
                        df_rows_show["cr_pct"] = df_rows_show["cr"] * 100
                        df_rows_show["drr_pct"] = df_rows_show["drr"] * 100

                        total_imp = df_month["impressions"].sum()
                        total_reach = float(df_month["reach"].sum()) if "reach" in df_month.columns else 0.0
                        total_frequency = (total_imp / total_reach) if total_reach > 0 else 0.0
                        total_clicks = df_month["clicks"].sum()
                        total_conv = df_month["conversions"].sum()
                        total_leads = float(df_month["leads"].fillna(0.0).sum()) if "leads" in df_month.columns else 0.0
                        total_target_leads = float(df_month["target_leads"].fillna(0.0).sum()) if "target_leads" in df_month.columns else float(total_conv)
                        total_cost = df_month["cost"].sum()
                        total_cost_with_vat = df_month["cost_with_vat"].sum()
                        total_cost_with_vat_ak = df_month["cost_with_vat_ak"].sum()
                        total_ak_wo_vat = df_month["ak_cost_wo_vat"].sum()
                        month_ak_rate_effective = (total_ak_wo_vat / total_cost) if total_cost > 0 else 0.0
                        total_rev = df_month["revenue"].sum()
                        if use_ak_budget_metrics:
                            total_budget_basis = total_cost_with_vat_ak if use_vat_budget_metrics else (total_cost + total_ak_wo_vat)
                        else:
                            total_budget_basis = total_cost_with_vat if use_vat_budget_metrics else total_cost

                        total_ctr = (total_clicks / total_imp * 100) if total_imp > 0 else 0
                        total_cpc = (total_cost / total_clicks) if total_clicks > 0 else 0
                        total_cr = (total_conv / total_clicks * 100) if total_clicks > 0 else 0
                        total_cr1 = (total_leads / total_clicks * 100) if total_clicks > 0 else 0
                        total_cr2 = (total_target_leads / total_leads * 100) if total_leads > 0 else 0
                        total_cpm = (total_budget_basis / (total_imp / 1000)) if total_imp > 0 else 0
                        total_cpa = (total_cost_with_vat / total_conv) if (is_diy_preset and total_conv > 0) else (total_budget_basis / total_conv) if total_conv > 0 else 0
                        total_cpl = (total_budget_basis / total_leads) if total_leads > 0 else 0
                        total_cpql = (total_cost_with_vat / total_target_leads) if (is_diy_preset and total_target_leads > 0) else (total_budget_basis / total_target_leads) if total_target_leads > 0 else 0
                        total_roas = (total_rev / total_budget_basis) if total_budget_basis > 0 else 0
                        total_drr = (total_budget_basis / total_rev * 100) if total_rev > 0 else 0

                        total_row_month = {
                            "campaign_type": "РС‚РѕРіРѕ",
                            "segment": "ALL",
                            "system": "",
                            "format": "",
                            "geo": "",
                            "impressions": total_imp,
                            "reach": total_reach,
                            "frequency": total_frequency,
                            "ctr": total_ctr / 100,
                            "cpc": total_cpc,
                            "cr": total_cr / 100,
                            "aov": None,
                            "clicks": total_clicks,
                            "conversions": total_conv,
                            "leads": total_leads,
                            "target_leads": total_target_leads,
                            "cost": total_cost,
                            "ak_cost_wo_vat": total_ak_wo_vat,
                            "total_budget_wo_vat_ak": total_cost + total_ak_wo_vat,
                            "cost_with_vat": total_cost_with_vat,
                            "cost_with_vat_ak": total_cost_with_vat_ak,
                            "vat_amount": total_cost_with_vat_ak - (total_cost + total_ak_wo_vat),
                            "ak_rate_pct": month_ak_rate_effective * 100.0,
                            "revenue": total_rev,
                            "cpm": total_cpm,
                            "cpa": total_cpa,
                            "cpl": total_cpl,
                            "cpql": total_cpql,
                            "roas": total_roas,
                            "drr": total_drr / 100,
                            "ctr_pct": total_ctr,
                            "cr_pct": total_cr,
                            "cr1_pct": total_cr1,
                            "cr2_pct": total_cr2,
                            "drr_pct": total_drr,
                        }
                        if is_diy_preset:
                            total_capacity = float(df_month["available_capacity"].sum()) if "available_capacity" in df_month.columns else 0.0
                            total_row_month["available_capacity"] = total_capacity
                            total_row_month["client_count"] = float(df_month["client_count"].sum()) if "client_count" in df_month.columns else 0.0
                            total_row_month["absolute_new_clients"] = float(df_month["absolute_new_clients"].sum()) if "absolute_new_clients" in df_month.columns else 0.0
                            total_row_month["returned_clients"] = float(df_month["returned_clients"].sum()) if "returned_clients" in df_month.columns else 0.0
                            total_row_month["new_clients"] = float(df_month["new_clients"].sum()) if "new_clients" in df_month.columns else 0.0
                            total_row_month["order_frequency"] = float(df_month["order_frequency"].mean()) if "order_frequency" in df_month.columns else 0.0
                            total_row_month["sov_pct"] = (total_reach / total_capacity * 100.0) if total_capacity > 0 else 0.0
                            total_row_month["new_clients_share_pct"] = float(df_month["new_clients_share_pct"].mean()) if "new_clients_share_pct" in df_month.columns else 0.0
                            total_row_month["cac"] = (total_cost_with_vat_ak / float(total_row_month["new_clients"])) if float(total_row_month["new_clients"]) > 0 else 0.0
                            total_row_month["shipped_orders"] = float(df_month["shipped_orders"].sum()) if "shipped_orders" in df_month.columns else 0.0
                            total_row_month["shipped_revenue"] = float(df_month["shipped_revenue"].sum()) if "shipped_revenue" in df_month.columns else 0.0
                            total_row_month["shipped_cps"] = (total_cost_with_vat / float(total_row_month["shipped_orders"])) if float(total_row_month["shipped_orders"]) > 0 else 0.0
                            total_row_month["shipped_aov"] = (float(total_row_month["shipped_revenue"]) / float(total_row_month["shipped_orders"])) if float(total_row_month["shipped_orders"]) > 0 else 0.0
                            total_row_month["shipped_roas"] = (float(total_row_month["shipped_revenue"]) / total_budget_basis) if total_budget_basis > 0 else 0.0
                            total_row_month["shipped_drr_pct"] = (total_budget_basis / float(total_row_month["shipped_revenue"]) * 100.0) if float(total_row_month["shipped_revenue"]) > 0 else 0.0
                            total_row_month["order_share_segment_pct"] = 100.0
                            total_row_month["revenue_share_segment_pct"] = 100.0
                            total_row_month["shipped_order_share_segment_pct"] = 100.0
                            total_row_month["shipped_revenue_share_segment_pct"] = 100.0
                            total_row_month["budget_share_segment_pct"] = 100.0
                        segment_total_rows_month = []
                        if show_segment_subtotals and "segment" in df_month.columns:
                            for seg_name, seg_df in df_month.groupby("segment"):
                                seg_imp = seg_df["impressions"].sum()
                                seg_clicks = seg_df["clicks"].sum()
                                seg_conv = seg_df["conversions"].sum()
                                seg_leads = float(seg_df["leads"].fillna(0.0).sum()) if "leads" in seg_df.columns else 0.0
                                seg_target_leads = float(seg_df["target_leads"].fillna(0.0).sum()) if "target_leads" in seg_df.columns else float(seg_conv)
                                seg_cost = seg_df["cost"].sum()
                                seg_cost_with_vat = seg_df["cost_with_vat"].sum()
                                seg_cost_with_vat_ak = seg_df["cost_with_vat_ak"].sum()
                                seg_ak_wo_vat = seg_df["ak_cost_wo_vat"].sum()
                                seg_rev = seg_df["revenue"].sum()
                                if use_ak_budget_metrics:
                                    seg_budget_basis = seg_cost_with_vat_ak if use_vat_budget_metrics else (seg_cost + seg_ak_wo_vat)
                                else:
                                    seg_budget_basis = seg_cost_with_vat if use_vat_budget_metrics else seg_cost
                                seg_ctr = (seg_clicks / seg_imp * 100) if seg_imp > 0 else 0
                                seg_cpc = (seg_cost / seg_clicks) if seg_clicks > 0 else 0
                                seg_cr = (seg_conv / seg_clicks * 100) if seg_clicks > 0 else 0
                                seg_cr1 = (seg_leads / seg_clicks * 100) if seg_clicks > 0 else 0
                                seg_cr2 = (seg_target_leads / seg_leads * 100) if seg_leads > 0 else 0
                                seg_cpm = (seg_budget_basis / (seg_imp / 1000)) if seg_imp > 0 else 0
                                seg_cpa = (seg_cost_with_vat / seg_conv) if (is_diy_preset and seg_conv > 0) else (seg_budget_basis / seg_conv) if seg_conv > 0 else 0
                                seg_cpl = (seg_budget_basis / seg_leads) if seg_leads > 0 else 0
                                seg_cpql = (seg_budget_basis / seg_target_leads) if seg_target_leads > 0 else 0
                                seg_roas = (seg_rev / seg_budget_basis) if seg_budget_basis > 0 else 0
                                seg_drr = (seg_budget_basis / seg_rev * 100) if seg_rev > 0 else 0
                                seg_row = {
                                    "campaign_type": f"РС‚РѕРіРѕ {seg_name}",
                                    "segment": seg_name,
                                    "system": "",
                                    "format": "",
                                    "geo": "",
                                    "impressions": seg_imp,
                                    "reach": float(seg_df["reach"].sum()) if "reach" in seg_df.columns else 0.0,
                                    "frequency": (seg_imp / float(seg_df["reach"].sum())) if ("reach" in seg_df.columns and float(seg_df["reach"].sum()) > 0) else 0.0,
                                    "ctr": seg_ctr / 100,
                                    "cpc": seg_cpc,
                                    "cr": seg_cr / 100,
                                    "aov": None,
                                    "clicks": seg_clicks,
                                    "conversions": seg_conv,
                                    "leads": seg_leads,
                                    "target_leads": seg_target_leads,
                                    "cost": seg_cost,
                                    "ak_cost_wo_vat": seg_ak_wo_vat,
                                    "total_budget_wo_vat_ak": seg_cost + seg_ak_wo_vat,
                                    "cost_with_vat": seg_cost_with_vat,
                                    "cost_with_vat_ak": seg_cost_with_vat_ak,
                                    "vat_amount": seg_cost_with_vat_ak - (seg_cost + seg_ak_wo_vat),
                                    "ak_rate_pct": (seg_ak_wo_vat / seg_cost * 100.0) if seg_cost > 0 else 0.0,
                                    "revenue": seg_rev,
                                    "cpm": seg_cpm,
                                    "cpa": seg_cpa,
                                    "cpl": seg_cpl,
                                    "cpql": seg_cpql,
                                    "roas": seg_roas,
                                    "drr": seg_drr / 100,
                                    "ctr_pct": seg_ctr,
                                    "cr_pct": seg_cr,
                                    "cr1_pct": seg_cr1,
                                    "cr2_pct": seg_cr2,
                                    "drr_pct": seg_drr,
                                }
                                if is_diy_preset:
                                    seg_cap = float(seg_df["available_capacity"].sum()) if "available_capacity" in seg_df.columns else 0.0
                                    seg_row["available_capacity"] = seg_cap
                                    seg_row["client_count"] = float(seg_df["client_count"].sum()) if "client_count" in seg_df.columns else 0.0
                                    seg_row["absolute_new_clients"] = float(seg_df["absolute_new_clients"].sum()) if "absolute_new_clients" in seg_df.columns else 0.0
                                    seg_row["returned_clients"] = float(seg_df["returned_clients"].sum()) if "returned_clients" in seg_df.columns else 0.0
                                    seg_row["new_clients"] = float(seg_df["new_clients"].sum()) if "new_clients" in seg_df.columns else 0.0
                                    seg_row["order_frequency"] = float(seg_df["order_frequency"].mean()) if "order_frequency" in seg_df.columns else 0.0
                                    seg_row["sov_pct"] = (float(seg_row.get("reach", 0.0)) / seg_cap * 100.0) if seg_cap > 0 else 0.0
                                    seg_row["new_clients_share_pct"] = float(seg_df["new_clients_share_pct"].mean()) if "new_clients_share_pct" in seg_df.columns else 0.0
                                    seg_row["cac"] = (seg_cost_with_vat_ak / float(seg_row["new_clients"])) if float(seg_row["new_clients"]) > 0 else 0.0
                                    seg_row["shipped_orders"] = float(seg_df["shipped_orders"].sum()) if "shipped_orders" in seg_df.columns else 0.0
                                    seg_row["shipped_revenue"] = float(seg_df["shipped_revenue"].sum()) if "shipped_revenue" in seg_df.columns else 0.0
                                    seg_row["shipped_cps"] = (seg_cost_with_vat / float(seg_row["shipped_orders"])) if float(seg_row["shipped_orders"]) > 0 else 0.0
                                    seg_row["shipped_aov"] = (float(seg_row["shipped_revenue"]) / float(seg_row["shipped_orders"])) if float(seg_row["shipped_orders"]) > 0 else 0.0
                                    seg_row["shipped_roas"] = (float(seg_row["shipped_revenue"]) / seg_budget_basis) if seg_budget_basis > 0 else 0.0
                                    seg_row["shipped_drr_pct"] = (seg_budget_basis / float(seg_row["shipped_revenue"]) * 100.0) if float(seg_row["shipped_revenue"]) > 0 else 0.0
                                    seg_row["order_share_segment_pct"] = 100.0
                                    seg_row["revenue_share_segment_pct"] = 100.0
                                    seg_row["shipped_order_share_segment_pct"] = 100.0
                                    seg_row["shipped_revenue_share_segment_pct"] = 100.0
                                    seg_row["budget_share_segment_pct"] = 100.0
                                segment_total_rows_month.append(seg_row)

                        # РћСЃРЅРѕРІРЅР°СЏ С‚Р°Р±Р»РёС†Р°: С‚РѕР»СЊРєРѕ С‚РёРїС‹ Р Рљ (Р±РµР· TOTAL), С‡С‚РѕР±С‹ РїРѕСЃС‚СЂРѕС‡РЅРѕ
                        # СЃРѕРІРїР°РґР°С‚СЊ СЃ С‚Р°Р±Р»РёС†РµР№ РєРѕСЌС„С„РёС†РёРµРЅС‚РѕРІ СЃРїСЂР°РІР°.
                        df_rows_show["impressions"] = df_rows_show["impressions"].map(
                            lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
                        )
                        df_rows_show["clicks"] = df_rows_show["clicks"].map(
                            lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
                        )
                        if "reach" in df_rows_show.columns:
                            df_rows_show["reach"] = df_rows_show["reach"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
                            )
                        if "frequency" in df_rows_show.columns:
                            df_rows_show["frequency"] = df_rows_show["frequency"].map(
                                lambda x: "" if pd.isna(x) else f"{x:.2f}"
                            )
                        df_rows_show["conversions"] = df_rows_show["conversions"].map(
                            lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
                        )
                        if "leads" in df_rows_show.columns:
                            df_rows_show["leads"] = df_rows_show["leads"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
                            )
                        if "target_leads" in df_rows_show.columns:
                            df_rows_show["target_leads"] = df_rows_show["target_leads"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
                            )

                        df_rows_show["cost"] = df_rows_show["cost"].map(
                            lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
                        )
                        if "ak_cost_wo_vat" in df_rows_show.columns:
                            df_rows_show["ak_cost_wo_vat"] = df_rows_show["ak_cost_wo_vat"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
                            )
                        if "total_budget_wo_vat_ak" in df_rows_show.columns:
                            df_rows_show["total_budget_wo_vat_ak"] = df_rows_show["total_budget_wo_vat_ak"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
                            )
                        df_rows_show["cost_with_vat"] = df_rows_show["cost_with_vat"].map(
                            lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
                        )
                        df_rows_show["cost_with_vat_ak"] = df_rows_show["cost_with_vat_ak"].map(
                            lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
                        )
                        if "vat_amount" in df_rows_show.columns:
                            df_rows_show["vat_amount"] = df_rows_show["vat_amount"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
                            )
                        df_rows_show["ak_rate_pct"] = df_rows_show["ak_rate_pct"].map(
                            lambda x: "" if pd.isna(x) else f"{x:.2f} %"
                        )
                        df_rows_show["revenue"] = df_rows_show["revenue"].map(
                            lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
                        )
                        df_rows_show["cpc"] = df_rows_show["cpc"].map(
                            lambda x: "" if pd.isna(x) else f"{x:.2f} в‚Ѕ".replace(",", " ")
                        )
                        df_rows_show["cpm"] = df_rows_show["cpm"].map(
                            lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
                        )
                        df_rows_show["cpa"] = df_rows_show["cpa"].map(
                            lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
                        )
                        df_rows_show["aov"] = df_rows_show["aov"].map(
                            lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
                        )
                        if "cpl" in df_rows_show.columns:
                            df_rows_show["cpl"] = df_rows_show["cpl"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
                            )
                        if "cpql" in df_rows_show.columns:
                            df_rows_show["cpql"] = df_rows_show["cpql"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
                            )

                        df_rows_show["ctr_pct"] = df_rows_show["ctr_pct"].map(
                            lambda x: "" if pd.isna(x) else f"{x:.2f} %"
                        )
                        df_rows_show["cr_pct"] = df_rows_show["cr_pct"].map(
                            lambda x: "" if pd.isna(x) else f"{x:.2f} %"
                        )
                        if "cr1_pct" in df_rows_show.columns:
                            df_rows_show["cr1_pct"] = df_rows_show["cr1_pct"].map(
                                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
                            )
                        if "cr2_pct" in df_rows_show.columns:
                            df_rows_show["cr2_pct"] = df_rows_show["cr2_pct"].map(
                                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
                            )
                        df_rows_show["drr_pct"] = df_rows_show["drr_pct"].map(
                            lambda x: "" if pd.isna(x) else f"{x:.2f} %"
                        )
                        df_rows_show["ROAS"] = df_rows_show["roas"].map(
                            lambda x: "" if pd.isna(x) else f"{x * 100:.2f} %"
                        )
                        if is_diy_preset:
                            df_rows_show["available_capacity"] = df_rows_show["available_capacity"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
                            )
                            df_rows_show["client_count"] = df_rows_show["client_count"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
                            )
                            df_rows_show["absolute_new_clients"] = df_rows_show["absolute_new_clients"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
                            )
                            df_rows_show["returned_clients"] = df_rows_show["returned_clients"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
                            )
                            df_rows_show["new_clients"] = df_rows_show["new_clients"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
                            )
                            df_rows_show["cac"] = df_rows_show["cac"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,} в‚Ѕ".replace(",", " ")
                            )
                            df_rows_show["order_frequency"] = df_rows_show["order_frequency"].map(
                                lambda x: "" if pd.isna(x) else f"{x:.2f}"
                            )
                            df_rows_show["sov_pct"] = df_rows_show["sov_pct"].map(
                                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
                            )
                            df_rows_show["new_clients_share_pct"] = df_rows_show["new_clients_share_pct"].map(
                                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
                            )
                            df_rows_show["order_share_segment_pct"] = df_rows_show["order_share_segment_pct"].map(
                                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
                            )
                            df_rows_show["revenue_share_segment_pct"] = df_rows_show["revenue_share_segment_pct"].map(
                                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
                            )
                            df_rows_show["budget_share_segment_pct"] = df_rows_show["budget_share_segment_pct"].map(
                                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
                            )

                        if is_real_estate_preset:
                            month_show_cols = ["campaign_type", "system", "format", "geo"] + get_real_estate_table_cols(metric_mode)
                        else:
                            month_show_cols = [
                                "campaign_type",
                                "system",
                                "format",
                                "geo",
                                "cost",
                                "ak_cost_wo_vat",
                                "total_budget_wo_vat_ak",
                                "cost_with_vat",
                                "cost_with_vat_ak",
                                "vat_amount",
                                "ak_rate_pct",
                                "impressions",
                                "reach",
                                "frequency",
                                "clicks",
                                "cpc",
                                "ctr_pct",
                                "cr_pct",
                                "conversions",
                                "cpa",
                                "aov",
                                "revenue",
                                "drr_pct",
                                "ROAS",
                            ]
                        if is_diy_preset:
                            month_show_cols.insert(3, "segment")
                        if is_diy_preset:
                            month_show_cols += ["available_capacity", "client_count", "absolute_new_clients", "returned_clients", "new_clients", "cac", "order_frequency", "shipped_orders", "shipped_cps", "shipped_aov", "shipped_revenue", "shipped_roas", "shipped_drr_pct", "shipped_order_share_segment_pct", "shipped_revenue_share_segment_pct", "sov_pct", "new_clients_share_pct", "order_share_segment_pct", "revenue_share_segment_pct", "budget_share_segment_pct"]
                        df_rows_show = safe_select_columns(df_rows_show, month_show_cols, fill_value="")
                        df_rows_show = df_rows_show.rename(columns=DISPLAY_COL_RENAME)

                        # TOTAL РѕС‚РґРµР»СЊРЅРѕР№ СЃС‚СЂРѕРєРѕР№ РЅРёР¶Рµ РѕСЃРЅРѕРІРЅРѕР№ С‚Р°Р±Р»РёС†С‹.
                        total_month_show = pd.DataFrame(segment_total_rows_month + [total_row_month])
                        total_month_show["impressions"] = total_month_show["impressions"].map(lambda x: f"{round(x):,}".replace(",", " "))
                        if "reach" in total_month_show.columns:
                            total_month_show["reach"] = total_month_show["reach"].map(lambda x: f"{round(x):,}".replace(",", " "))
                        if "frequency" in total_month_show.columns:
                            total_month_show["frequency"] = total_month_show["frequency"].map(lambda x: f"{x:.2f}")
                        total_month_show["clicks"] = total_month_show["clicks"].map(lambda x: f"{round(x):,}".replace(",", " "))
                        total_month_show["conversions"] = total_month_show["conversions"].map(lambda x: f"{round(x):,}".replace(",", " "))
                        if "leads" in total_month_show.columns:
                            total_month_show["leads"] = total_month_show["leads"].map(lambda x: f"{round(x):,}".replace(",", " "))
                        if "target_leads" in total_month_show.columns:
                            total_month_show["target_leads"] = total_month_show["target_leads"].map(lambda x: f"{round(x):,}".replace(",", " "))
                        total_month_show["cost"] = total_month_show["cost"].map(lambda x: f"{round(x):,} в‚Ѕ".replace(",", " "))
                        if "ak_cost_wo_vat" in total_month_show.columns:
                            total_month_show["ak_cost_wo_vat"] = total_month_show["ak_cost_wo_vat"].map(lambda x: f"{round(x):,} в‚Ѕ".replace(",", " "))
                        if "total_budget_wo_vat_ak" in total_month_show.columns:
                            total_month_show["total_budget_wo_vat_ak"] = total_month_show["total_budget_wo_vat_ak"].map(lambda x: f"{round(x):,} в‚Ѕ".replace(",", " "))
                        total_month_show["cost_with_vat"] = total_month_show["cost_with_vat"].map(lambda x: f"{round(x):,} в‚Ѕ".replace(",", " "))
                        total_month_show["cost_with_vat_ak"] = total_month_show["cost_with_vat_ak"].map(lambda x: f"{round(x):,} в‚Ѕ".replace(",", " "))
                        if "vat_amount" in total_month_show.columns:
                            total_month_show["vat_amount"] = total_month_show["vat_amount"].map(lambda x: f"{round(x):,} в‚Ѕ".replace(",", " "))
                        total_month_show["ak_rate_pct"] = total_month_show["ak_rate_pct"].map(lambda x: f"{x:.2f} %")
                        total_month_show["revenue"] = total_month_show["revenue"].map(lambda x: f"{round(x):,} в‚Ѕ".replace(",", " "))
                        total_month_show["cpc"] = total_month_show["cpc"].map(lambda x: f"{x:.2f} в‚Ѕ")
                        total_month_show["cpm"] = total_month_show["cpm"].map(lambda x: f"{round(x):,} в‚Ѕ".replace(",", " "))
                        total_month_show["cpa"] = total_month_show["cpa"].map(lambda x: f"{round(x):,} в‚Ѕ".replace(",", " "))
                        if "cpl" in total_month_show.columns:
                            total_month_show["cpl"] = total_month_show["cpl"].map(lambda x: f"{round(x):,} в‚Ѕ".replace(",", " "))
                        if "cpql" in total_month_show.columns:
                            total_month_show["cpql"] = total_month_show["cpql"].map(lambda x: f"{round(x):,} в‚Ѕ".replace(",", " "))
                        total_month_show["aov"] = ""
                        if "client_count" in total_month_show.columns:
                            total_month_show["client_count"] = total_month_show["client_count"].map(lambda x: f"{round(x):,}".replace(",", " "))
                        if "absolute_new_clients" in total_month_show.columns:
                            total_month_show["absolute_new_clients"] = total_month_show["absolute_new_clients"].map(lambda x: f"{round(x):,}".replace(",", " "))
                        if "returned_clients" in total_month_show.columns:
                            total_month_show["returned_clients"] = total_month_show["returned_clients"].map(lambda x: f"{round(x):,}".replace(",", " "))
                        if "new_clients" in total_month_show.columns:
                            total_month_show["new_clients"] = total_month_show["new_clients"].map(lambda x: f"{round(x):,}".replace(",", " "))
                        if "cac" in total_month_show.columns:
                            total_month_show["cac"] = total_month_show["cac"].map(lambda x: f"{round(x):,} в‚Ѕ".replace(",", " "))
                        if "order_frequency" in total_month_show.columns:
                            total_month_show["order_frequency"] = total_month_show["order_frequency"].map(lambda x: f"{x:.2f}")
                        total_month_show["ctr_pct"] = total_month_show["ctr_pct"].map(lambda x: f"{x:.2f} %")
                        total_month_show["cr_pct"] = total_month_show["cr_pct"].map(lambda x: f"{x:.2f} %")
                        if "cr1_pct" in total_month_show.columns:
                            total_month_show["cr1_pct"] = total_month_show["cr1_pct"].map(lambda x: f"{x:.2f} %")
                        if "cr2_pct" in total_month_show.columns:
                            total_month_show["cr2_pct"] = total_month_show["cr2_pct"].map(lambda x: f"{x:.2f} %")
                        total_month_show["drr_pct"] = total_month_show["drr_pct"].map(lambda x: f"{x:.2f} %")
                        total_month_show["ROAS"] = total_month_show["roas"].map(lambda x: f"{x * 100:.2f} %")
                        if is_diy_preset:
                            total_month_show["available_capacity"] = total_month_show["available_capacity"].map(
                                lambda x: f"{round(x):,}".replace(",", " ")
                            )
                            total_month_show["sov_pct"] = total_month_show["sov_pct"].map(lambda x: f"{x:.2f} %")
                            total_month_show["new_clients_share_pct"] = total_month_show["new_clients_share_pct"].map(lambda x: f"{x:.2f} %")
                            total_month_show["order_share_segment_pct"] = total_month_show["order_share_segment_pct"].map(lambda x: f"{x:.2f} %")
                            total_month_show["revenue_share_segment_pct"] = total_month_show["revenue_share_segment_pct"].map(lambda x: f"{x:.2f} %")
                            total_month_show["budget_share_segment_pct"] = total_month_show["budget_share_segment_pct"].map(lambda x: f"{x:.2f} %")
                        if is_real_estate_preset:
                            total_month_cols = ["campaign_type", "system", "format", "geo"] + get_real_estate_table_cols(metric_mode)
                        else:
                            total_month_cols = [
                                "campaign_type", "system", "format", "geo",
                                "cost", "ak_cost_wo_vat", "total_budget_wo_vat_ak", "cost_with_vat", "cost_with_vat_ak", "vat_amount", "ak_rate_pct",
                                "impressions", "reach", "frequency", "clicks", "cpc", "ctr_pct", "cr_pct",
                                "conversions", "cpa", "aov", "revenue", "drr_pct", "ROAS",
                            ]
                        if is_diy_preset:
                            total_month_cols.insert(3, "segment")
                        if is_diy_preset:
                            total_month_cols += ["available_capacity", "client_count", "absolute_new_clients", "returned_clients", "new_clients", "cac", "order_frequency", "shipped_orders", "shipped_cps", "shipped_aov", "shipped_revenue", "shipped_roas", "shipped_drr_pct", "shipped_order_share_segment_pct", "shipped_revenue_share_segment_pct", "sov_pct", "new_clients_share_pct", "order_share_segment_pct", "revenue_share_segment_pct", "budget_share_segment_pct"]
                        total_month_show = safe_select_columns(total_month_show, total_month_cols, fill_value="").rename(columns=DISPLAY_COL_RENAME)

                        # РџСЂРѕСЃС‚РѕР№ РІР°СЂРёР°РЅС‚: РІРѕР·РІСЂР°С‰Р°РµРј TOTAL РѕР±СЂР°С‚РЅРѕ РІ РѕСЃРЅРѕРІРЅСѓСЋ С‚Р°Р±Р»РёС†Сѓ.
                        df_month_with_total_show = pd.concat([df_rows_show, total_month_show], ignore_index=True)
                        if is_diy_preset:
                            df_month_with_total_show = reorder_rows_with_segment_subtotals(
                                df_month_with_total_show,
                                DISPLAY_COL_RENAME["campaign_type"],
                                DISPLAY_COL_RENAME["segment"],
                            )

                        def _highlight_total_row_merged(row):
                            style = [""] * len(row)
                            camp_col = DISPLAY_COL_RENAME.get("campaign_type", "РќР°Р·РІР°РЅРёРµ РєР°РјРїР°РЅРёРё")
                            camp_val = str(row.get(camp_col, ""))
                            if camp_val == "РС‚РѕРіРѕ":
                                style = [
                                    f"background-color: #00CDC5; color: #081521; font-weight: 700; border-top: 2px solid {THEME_BORDER};"
                                ] * len(row)
                            elif camp_val.startswith("РС‚РѕРіРѕ "):
                                style = [
                                    f"background-color: #2C6E75; color: #DDEAF0; font-weight: 650; border-top: 1px solid {THEME_BORDER};"
                                ] * len(row)
                            return style

                        month_table_height_with_total = max(
                            140,
                            min(560, month_table_height + month_row_height + 4),
                        )
                        st.dataframe(
                            df_month_with_total_show.style.apply(_highlight_total_row_merged, axis=1),
                            use_container_width=True,
                            height=month_table_height_with_total,
                        )
                    else:
                        st.info("РќРµС‚ РґР°РЅРЅС‹С… РґР»СЏ СЌС‚РѕРіРѕ РјРµСЃСЏС†Р°.")

    # ---------- 4. РС‚РѕРіРё РїРѕ РІС‹Р±СЂР°РЅРЅС‹Рј РјРµСЃСЏС†Р°Рј (TOTAL) ----------

    if all_months_results:
        df_all = pd.concat(all_months_results, ignore_index=True)
    else:
        df_all = pd.DataFrame()

    if not df_all.empty:
        ref_imp = float(df_all["impressions"].sum())
        ref_reach = float(df_all["reach"].sum()) if "reach" in df_all.columns else 0.0
        ref_frequency = (ref_imp / ref_reach) if ref_reach > 0 else 0.0
        ref_clicks = float(df_all["clicks"].sum())
        ref_conv = float(df_all["conversions"].sum())
        ref_leads = float(df_all["leads"].fillna(0.0).sum()) if "leads" in df_all.columns else 0.0
        ref_target_leads = float(df_all["target_leads"].fillna(0.0).sum()) if "target_leads" in df_all.columns else float(ref_conv)
        ref_cost = float(df_all["cost"].sum())
        ref_cost_with_vat = float(df_all["cost_with_vat"].sum())
        ref_cost_with_vat_ak = float(df_all["cost_with_vat_ak"].sum())
        ref_ak_wo_vat = float(df_all["ak_cost_wo_vat"].sum())
        ref_rev = float(df_all["revenue"].sum())
        if use_ak_budget_metrics:
            ref_budget_basis = ref_cost_with_vat_ak if use_vat_budget_metrics else (ref_cost + ref_ak_wo_vat)
        else:
            ref_budget_basis = ref_cost_with_vat if use_vat_budget_metrics else ref_cost
        ref_ctr = (ref_clicks / ref_imp * 100.0) if ref_imp > 0 else 0.0
        ref_cpc = (ref_cost / ref_clicks) if ref_clicks > 0 else 0.0
        ref_cr = (ref_conv / ref_clicks * 100.0) if ref_clicks > 0 else 0.0
        ref_cpo = (ref_budget_basis / ref_conv) if ref_conv > 0 else 0.0
        ref_cpl = (ref_budget_basis / ref_leads) if ref_leads > 0 else 0.0
        ref_cpql = (ref_budget_basis / ref_target_leads) if ref_target_leads > 0 else 0.0
        ref_roas = (ref_rev / ref_budget_basis * 100.0) if ref_budget_basis > 0 else 0.0
        ref_drr = (ref_budget_basis / ref_rev * 100.0) if ref_rev > 0 else 0.0
        st.session_state["mp_ref_total"] = {
            "impressions": ref_imp,
            "reach": ref_reach,
            "frequency": ref_frequency,
            "clicks": ref_clicks,
            "conversions": ref_conv,
            "leads": ref_leads,
            "target_leads": ref_target_leads,
            "cost": ref_cost,
            "cost_with_vat": ref_cost_with_vat,
            "cost_with_vat_ak": ref_cost_with_vat_ak,
            "revenue": ref_rev,
            "ctr": ref_ctr,
            "cpc": ref_cpc,
            "cr": ref_cr,
            "cpo": ref_cpo,
            "cpl": ref_cpl,
            "cpql": ref_cpql,
            "roas": ref_roas,
            "drr": ref_drr,
        }
        total_by_campaign = {}
        agg_ct = df_all.groupby("campaign_type", as_index=False).agg(
            impressions=("impressions", "sum"),
            reach=("reach", "sum"),
            clicks=("clicks", "sum"),
            conversions=("conversions", "sum"),
            leads=("leads", "sum"),
            target_leads=("target_leads", "sum"),
            cost=("cost", "sum"),
            cost_with_vat=("cost_with_vat", "sum"),
            cost_with_vat_ak=("cost_with_vat_ak", "sum"),
            ak_cost_wo_vat=("ak_cost_wo_vat", "sum"),
            revenue=("revenue", "sum"),
        )
        for _, r in agg_ct.iterrows():
            camp = str(r.get("campaign_type", "")).strip()
            if not camp:
                continue
            imp = float(r.get("impressions", 0.0))
            reach = float(r.get("reach", 0.0) or 0.0)
            frequency = (imp / reach) if reach > 0 else 0.0
            clicks = float(r.get("clicks", 0.0))
            conv = float(r.get("conversions", 0.0))
            leads = float(r.get("leads", 0.0) or 0.0)
            target_leads = float(r.get("target_leads", conv) or 0.0)
            cost = float(r.get("cost", 0.0))
            cost_with_vat = float(r.get("cost_with_vat", 0.0))
            cost_with_vat_ak = float(r.get("cost_with_vat_ak", 0.0))
            ak_wo_vat = float(r.get("ak_cost_wo_vat", 0.0))
            rev = float(r.get("revenue", 0.0))
            if use_ak_budget_metrics:
                budget_basis = cost_with_vat_ak if use_vat_budget_metrics else (cost + ak_wo_vat)
            else:
                budget_basis = cost_with_vat if use_vat_budget_metrics else cost
            ctr = (clicks / imp * 100.0) if imp > 0 else 0.0
            cpc = (cost / clicks) if clicks > 0 else 0.0
            cr = (conv / clicks * 100.0) if clicks > 0 else 0.0
            cpo = (budget_basis / conv) if conv > 0 else 0.0
            cpl = (budget_basis / leads) if leads > 0 else 0.0
            cpql = (budget_basis / target_leads) if target_leads > 0 else 0.0
            roas = (rev / budget_basis * 100.0) if budget_basis > 0 else 0.0
            drr = (budget_basis / rev * 100.0) if rev > 0 else 0.0
            total_by_campaign[camp] = {
                "impressions": imp,
                "reach": reach,
                "frequency": frequency,
                "clicks": clicks,
                "conversions": conv,
                "leads": leads,
                "target_leads": target_leads,
                "cost": cost,
                "cost_with_vat": cost_with_vat,
                "cost_with_vat_ak": cost_with_vat_ak,
                "revenue": rev,
                "ctr": ctr,
                "cpc": cpc,
                "cr": cr,
                "cpo": cpo,
                "cpl": cpl,
                "cpql": cpql,
                "roas": roas,
                "drr": drr,
            }
        st.session_state["mp_ref_total_by_campaign"] = total_by_campaign
    else:
        st.session_state["mp_ref_total"] = None
        st.session_state["mp_ref_total_by_campaign"] = {}

    with st.sidebar:
        st.markdown("---")
        st.markdown("### Р‘С‹СЃС‚СЂР°СЏ СЃРІРµСЂРєР°")
        # Import compatibility: old state could store a legacy mode value.
        if st.session_state.get("mp_ref_mode") == "Р’РєР»СЋС‡РµРЅРѕ":
            st.session_state["mp_ref_mode"] = "РЎСЂРµРґРЅРёР№ РјРµСЃСЏС†"
        # If references are missing after import/rerun, rebuild them from base campaigns.
        if not st.session_state.get("mp_ref_base"):
            _bootstrap_reference_from_campaigns(st.session_state.get("campaigns_df"))
        if is_diy_preset:
            st.selectbox(
                "РЎРµРіРјРµРЅС‚ РІ СЂР°СЃС‡РµС‚Рµ",
                options=["Р’СЃРµ", "B2B", "B2C"],
                key="plan_segment_filter_sidebar",
                on_change=_sync_segment_from_sidebar,
            )
        ref_mode = st.selectbox(
            "Р§С‚Рѕ Р·Р°РєСЂРµРїРёС‚СЊ",
            options=["Р’С‹РєР»СЋС‡РµРЅРѕ", "РЎСЂРµРґРЅРёР№ РјРµСЃСЏС†", "TOTAL РІС‹Р±СЂР°РЅРЅС‹С… РјРµСЃСЏС†РµРІ"],
            key="mp_ref_mode",
        )
        all_ref_campaigns = sorted(
            set(st.session_state.get("mp_ref_base_by_campaign", {}).keys())
            | set(st.session_state.get("mp_ref_total_by_campaign", {}).keys())
        )
        valid_ref_campaigns = ["Р’СЃРµ С‚РёРїС‹ Р Рљ"] + all_ref_campaigns
        if st.session_state.get("mp_ref_campaign") not in valid_ref_campaigns:
            st.session_state["mp_ref_campaign"] = "Р’СЃРµ С‚РёРїС‹ Р Рљ"
        ref_campaign = st.selectbox(
            "РўРёРї Р Рљ",
            options=valid_ref_campaigns,
            key="mp_ref_campaign",
        )
        ref_src = None
        if ref_mode == "РЎСЂРµРґРЅРёР№ РјРµСЃСЏС†":
            if ref_campaign == "Р’СЃРµ С‚РёРїС‹ Р Рљ":
                ref_src = st.session_state.get("mp_ref_base")
            else:
                ref_src = st.session_state.get("mp_ref_base_by_campaign", {}).get(ref_campaign)
        elif ref_mode == "TOTAL РІС‹Р±СЂР°РЅРЅС‹С… РјРµСЃСЏС†РµРІ":
            if ref_campaign == "Р’СЃРµ С‚РёРїС‹ Р Рљ":
                ref_src = st.session_state.get("mp_ref_total")
            else:
                ref_src = st.session_state.get("mp_ref_total_by_campaign", {}).get(ref_campaign)

        # Fallback: right after import TOTAL may be unavailable until full plan calc.
        # Show base reference instead of empty block.
        if ref_src is None and ref_mode == "TOTAL РІС‹Р±СЂР°РЅРЅС‹С… РјРµСЃСЏС†РµРІ":
            if ref_campaign == "Р’СЃРµ С‚РёРїС‹ Р Рљ":
                ref_src = st.session_state.get("mp_ref_base")
            else:
                ref_src = st.session_state.get("mp_ref_base_by_campaign", {}).get(ref_campaign)

        if ref_mode != "Р’С‹РєР»СЋС‡РµРЅРѕ":
            if ref_src:
                st.dataframe(_build_ref_df(ref_src), use_container_width=True, hide_index=True)
            else:
                st.caption("РќРµС‚ РґР°РЅРЅС‹С… РґР»СЏ РІС‹Р±СЂР°РЅРЅРѕРіРѕ СЂРµР¶РёРјР° СЃРІРµСЂРєРё.")

    with st.expander("4. РС‚РѕРіРё РїРѕ РІС‹Р±СЂР°РЅРЅС‹Рј РјРµСЃСЏС†Р°Рј (TOTAL)", expanded=True):
        if df_all.empty:
            st.info("РќРµС‚ РґР°РЅРЅС‹С… РґР»СЏ РёС‚РѕРіРѕРІ. Р—Р°РїРѕР»РЅРёС‚Рµ РјРµРґРёР°РїР»Р°РЅ РїРѕ РјРµСЃСЏС†Р°Рј.")
        else:
            st.caption("РЎРІРѕРґРЅС‹Рµ РїРѕРєР°Р·Р°С‚РµР»Рё РїРѕ РІСЃРµРј РІС‹Р±СЂР°РЅРЅС‹Рј РјРµСЃСЏС†Р°Рј Рё С‚РёРїР°Рј Р Рљ.")

            agg = df_all.groupby(["month_num", "month_name"], as_index=False).agg(
                impressions=("impressions", "sum"),
                clicks=("clicks", "sum"),
                conversions=("conversions", "sum"),
                leads=("leads", "sum"),
                target_leads=("target_leads", "sum"),
                cost=("cost", "sum"),
                cost_with_vat=("cost_with_vat", "sum"),
                cost_with_vat_ak=("cost_with_vat_ak", "sum"),
                ak_cost_wo_vat=("ak_cost_wo_vat", "sum"),
                revenue=("revenue", "sum"),
            )
            if is_diy_preset:
                agg["reach"] = df_all.groupby(["month_num", "month_name"])["reach"].sum().values if "reach" in df_all.columns else 0.0
                agg["frequency"] = np.where(agg["reach"] > 0, agg["impressions"] / agg["reach"], 0.0)
                agg["available_capacity"] = df_all.groupby(["month_num", "month_name"])["available_capacity"].sum().values if "available_capacity" in df_all.columns else 0.0
                agg["client_count"] = df_all.groupby(["month_num", "month_name"])["client_count"].sum().values if "client_count" in df_all.columns else 0.0
                agg["absolute_new_clients"] = df_all.groupby(["month_num", "month_name"])["absolute_new_clients"].sum().values if "absolute_new_clients" in df_all.columns else 0.0
                agg["returned_clients"] = df_all.groupby(["month_num", "month_name"])["returned_clients"].sum().values if "returned_clients" in df_all.columns else 0.0
                agg["new_clients"] = df_all.groupby(["month_num", "month_name"])["new_clients"].sum().values if "new_clients" in df_all.columns else 0.0
                agg["order_frequency"] = df_all.groupby(["month_num", "month_name"])["order_frequency"].mean().values if "order_frequency" in df_all.columns else 0.0
                agg["sov_pct"] = np.where(agg["available_capacity"] > 0, agg["reach"] / agg["available_capacity"] * 100.0, 0.0)
                agg["new_clients_share_pct"] = df_all.groupby(["month_num", "month_name"])["new_clients_share_pct"].mean().values if "new_clients_share_pct" in df_all.columns else 0.0
                agg["order_share_segment_pct"] = 0.0
                agg["revenue_share_segment_pct"] = 0.0
            if use_ak_budget_metrics:
                budget_series = agg["cost_with_vat_ak"] if use_vat_budget_metrics else (agg["cost"] + agg["ak_cost_wo_vat"])
            else:
                budget_series = agg["cost_with_vat"] if use_vat_budget_metrics else agg["cost"]
            agg["total_budget_wo_vat_ak"] = agg["cost"] + agg["ak_cost_wo_vat"]
            agg["vat_amount"] = agg["cost_with_vat_ak"] - agg["total_budget_wo_vat_ak"]

            agg["ctr"] = np.where(
                agg["impressions"] > 0,
                agg["clicks"] / agg["impressions"] * 100,
                0.0,
            )
            agg["cr"] = np.where(
                agg["clicks"] > 0,
                agg["conversions"] / agg["clicks"] * 100,
                0.0,
            )
            agg["cr1_pct"] = np.where(
                agg["clicks"] > 0,
                agg["leads"] / agg["clicks"] * 100,
                0.0,
            )
            agg["cr2_pct"] = np.where(
                agg["leads"] > 0,
                agg["target_leads"] / agg["leads"] * 100,
                0.0,
            )
            agg["cpc"] = np.where(
                agg["clicks"] > 0,
                agg["cost"] / agg["clicks"],
                0.0,
            )
            agg["cpm"] = np.where(
                agg["impressions"] > 0,
                budget_series / (agg["impressions"] / 1000),
                0.0,
            )
            agg["cpa"] = np.where(
                agg["conversions"] > 0,
                budget_series / agg["conversions"],
                0.0,
            )
            agg["cpl"] = np.where(
                agg["leads"] > 0,
                budget_series / agg["leads"],
                0.0,
            )
            agg["cpql"] = np.where(
                agg["target_leads"] > 0,
                budget_series / agg["target_leads"],
                0.0,
            )
            agg["aov"] = np.where(
                agg["conversions"] > 0,
                agg["revenue"] / agg["conversions"],
                0.0,
            )
            agg["roas"] = np.where(
                budget_series > 0,
                agg["revenue"] / budget_series,
                0.0,
            )
            agg["drr"] = np.where(
                agg["revenue"] > 0,
                budget_series / agg["revenue"] * 100,
                0.0,
            )
            agg["ctr_pct"] = agg["ctr"]
            agg["cr_pct"] = agg["cr"]

            total_row = {
                "month_num": 999,
                "month_name": "TOTAL",
                "impressions": agg["impressions"].sum(),
                "reach": float(agg["reach"].sum()) if "reach" in agg.columns else 0.0,
                "clicks": agg["clicks"].sum(),
                "conversions": agg["conversions"].sum(),
                "leads": agg["leads"].sum(),
                "target_leads": agg["target_leads"].sum(),
                "cost": agg["cost"].sum(),
                "cost_with_vat": agg["cost_with_vat"].sum(),
                "cost_with_vat_ak": agg["cost_with_vat_ak"].sum(),
                "ak_cost_wo_vat": agg["ak_cost_wo_vat"].sum(),
                "total_budget_wo_vat_ak": agg["cost"].sum() + agg["ak_cost_wo_vat"].sum(),
                "vat_amount": agg["cost_with_vat_ak"].sum() - (agg["cost"].sum() + agg["ak_cost_wo_vat"].sum()),
                "revenue": agg["revenue"].sum(),
            }
            if use_ak_budget_metrics:
                total_budget_basis = total_row["cost_with_vat_ak"] if use_vat_budget_metrics else (total_row["cost"] + total_row["ak_cost_wo_vat"])
            else:
                total_budget_basis = total_row["cost_with_vat"] if use_vat_budget_metrics else total_row["cost"]
            if total_row["impressions"] > 0:
                total_row["ctr"] = total_row["clicks"] / total_row["impressions"] * 100
                total_row["cpm"] = total_budget_basis / (total_row["impressions"] / 1000)
            else:
                total_row["ctr"] = 0.0
                total_row["cpm"] = 0.0
            total_row["ctr_pct"] = total_row["ctr"]
            if total_row["clicks"] > 0:
                total_row["cr"] = total_row["conversions"] / total_row["clicks"] * 100
                total_row["cpc"] = total_row["cost"] / total_row["clicks"]
                total_row["cr1_pct"] = total_row["leads"] / total_row["clicks"] * 100
            else:
                total_row["cr"] = 0.0
                total_row["cpc"] = 0.0
                total_row["cr1_pct"] = 0.0
            total_row["cr_pct"] = total_row["cr"]
            if total_row["leads"] > 0:
                total_row["cr2_pct"] = total_row["target_leads"] / total_row["leads"] * 100
                total_row["cpl"] = total_budget_basis / total_row["leads"]
            else:
                total_row["cr2_pct"] = 0.0
                total_row["cpl"] = 0.0
            if total_row["conversions"] > 0:
                total_row["cpa"] = (total_row["cost_with_vat"] / total_row["conversions"]) if is_diy_preset else (total_budget_basis / total_row["conversions"])
                total_row["aov"] = total_row["revenue"] / total_row["conversions"]
            else:
                total_row["cpa"] = 0.0
                total_row["aov"] = 0.0
            if total_row["target_leads"] > 0:
                total_row["cpql"] = total_budget_basis / total_row["target_leads"]
            else:
                total_row["cpql"] = 0.0
            if total_budget_basis > 0:
                total_row["roas"] = total_row["revenue"] / total_budget_basis
            else:
                total_row["roas"] = 0.0
            if total_row["revenue"] > 0:
                total_row["drr"] = total_budget_basis / total_row["revenue"] * 100
            else:
                total_row["drr"] = 0.0
            if is_diy_preset:
                total_row["frequency"] = (total_row["impressions"] / total_row["reach"]) if total_row["reach"] > 0 else 0.0
                total_row["available_capacity"] = float(agg["available_capacity"].sum()) if "available_capacity" in agg.columns else 0.0
                total_row["client_count"] = float(agg["client_count"].sum()) if "client_count" in agg.columns else 0.0
                total_row["absolute_new_clients"] = float(agg["absolute_new_clients"].sum()) if "absolute_new_clients" in agg.columns else 0.0
                total_row["returned_clients"] = float(agg["returned_clients"].sum()) if "returned_clients" in agg.columns else 0.0
                total_row["new_clients"] = float(agg["new_clients"].sum()) if "new_clients" in agg.columns else 0.0
                total_row["order_frequency"] = float(agg["order_frequency"].mean()) if "order_frequency" in agg.columns else 0.0
                total_row["sov_pct"] = (total_row["reach"] / total_row["available_capacity"] * 100.0) if total_row["available_capacity"] > 0 else 0.0
                total_row["new_clients_share_pct"] = float(agg["new_clients_share_pct"].mean()) if "new_clients_share_pct" in agg.columns else 0.0
                total_row["cac"] = (total_row["cost_with_vat_ak"] / float(total_row["new_clients"])) if float(total_row["new_clients"]) > 0 else 0.0
                total_row["order_share_segment_pct"] = 100.0
                total_row["revenue_share_segment_pct"] = 100.0

            segment_total_rows = []
            if show_segment_subtotals and "segment" in df_all.columns:
                for seg_name, seg_df in df_all.groupby("segment"):
                    seg_cost = float(seg_df["cost"].sum())
                    seg_cost_with_vat = float(seg_df["cost_with_vat"].sum())
                    seg_cost_with_vat_ak = float(seg_df["cost_with_vat_ak"].sum())
                    seg_ak_wo_vat = float(seg_df["ak_cost_wo_vat"].sum())
                    if use_ak_budget_metrics:
                        seg_budget_basis = seg_cost_with_vat_ak if use_vat_budget_metrics else (seg_cost + seg_ak_wo_vat)
                    else:
                        seg_budget_basis = seg_cost_with_vat if use_vat_budget_metrics else seg_cost
                    seg_row = {
                        "month_num": 998,
                        "month_name": f"РС‚РѕРіРѕ {seg_name}",
                        "impressions": float(seg_df["impressions"].sum()),
                        "reach": float(seg_df["reach"].sum()) if "reach" in seg_df.columns else 0.0,
                        "clicks": float(seg_df["clicks"].sum()),
                        "conversions": float(seg_df["conversions"].sum()),
                        "leads": float(seg_df["leads"].fillna(0.0).sum()) if "leads" in seg_df.columns else 0.0,
                        "target_leads": float(seg_df["target_leads"].fillna(0.0).sum()) if "target_leads" in seg_df.columns else float(seg_df["conversions"].sum()),
                        "cost": seg_cost,
                        "cost_with_vat": seg_cost_with_vat,
                        "cost_with_vat_ak": seg_cost_with_vat_ak,
                        "ak_cost_wo_vat": seg_ak_wo_vat,
                        "total_budget_wo_vat_ak": seg_cost + seg_ak_wo_vat,
                        "vat_amount": seg_cost_with_vat_ak - (seg_cost + seg_ak_wo_vat),
                        "revenue": float(seg_df["revenue"].sum()),
                    }
                    seg_row["ctr"] = (seg_row["clicks"] / seg_row["impressions"] * 100.0) if seg_row["impressions"] > 0 else 0.0
                    seg_row["cr"] = (seg_row["conversions"] / seg_row["clicks"] * 100.0) if seg_row["clicks"] > 0 else 0.0
                    seg_row["ctr_pct"] = seg_row["ctr"]
                    seg_row["cr_pct"] = seg_row["cr"]
                    seg_row["cr1_pct"] = (seg_row["leads"] / seg_row["clicks"] * 100.0) if seg_row["clicks"] > 0 else 0.0
                    seg_row["cr2_pct"] = (seg_row["target_leads"] / seg_row["leads"] * 100.0) if seg_row["leads"] > 0 else 0.0
                    seg_row["cpc"] = (seg_row["cost"] / seg_row["clicks"]) if seg_row["clicks"] > 0 else 0.0
                    seg_row["cpm"] = (seg_budget_basis / (seg_row["impressions"] / 1000.0)) if seg_row["impressions"] > 0 else 0.0
                    seg_row["cpa"] = (seg_budget_basis / seg_row["conversions"]) if seg_row["conversions"] > 0 else 0.0
                    seg_row["cpl"] = (seg_budget_basis / seg_row["leads"]) if seg_row["leads"] > 0 else 0.0
                    seg_row["cpql"] = (seg_budget_basis / seg_row["target_leads"]) if seg_row["target_leads"] > 0 else 0.0
                    seg_row["aov"] = (seg_row["revenue"] / seg_row["conversions"]) if seg_row["conversions"] > 0 else 0.0
                    seg_row["roas"] = (seg_row["revenue"] / seg_budget_basis) if seg_budget_basis > 0 else 0.0
                    seg_row["drr"] = (seg_budget_basis / seg_row["revenue"] * 100.0) if seg_row["revenue"] > 0 else 0.0
                    if is_diy_preset:
                        seg_row["cpa"] = (seg_row["cost_with_vat"] / seg_row["conversions"]) if seg_row["conversions"] > 0 else 0.0
                        seg_row["frequency"] = (seg_row["impressions"] / seg_row["reach"]) if seg_row["reach"] > 0 else 0.0
                        seg_row["available_capacity"] = float(seg_df["available_capacity"].sum()) if "available_capacity" in seg_df.columns else 0.0
                        seg_row["client_count"] = float(seg_df["client_count"].sum()) if "client_count" in seg_df.columns else 0.0
                        seg_row["absolute_new_clients"] = float(seg_df["absolute_new_clients"].sum()) if "absolute_new_clients" in seg_df.columns else 0.0
                        seg_row["returned_clients"] = float(seg_df["returned_clients"].sum()) if "returned_clients" in seg_df.columns else 0.0
                        seg_row["new_clients"] = float(seg_df["new_clients"].sum()) if "new_clients" in seg_df.columns else 0.0
                        seg_row["order_frequency"] = float(seg_df["order_frequency"].mean()) if "order_frequency" in seg_df.columns else 0.0
                        seg_row["sov_pct"] = (seg_row["reach"] / seg_row["available_capacity"] * 100.0) if seg_row["available_capacity"] > 0 else 0.0
                        seg_row["new_clients_share_pct"] = float(seg_df["new_clients_share_pct"].mean()) if "new_clients_share_pct" in seg_df.columns else 0.0
                        seg_row["cac"] = (seg_row["cost_with_vat_ak"] / float(seg_row["new_clients"])) if float(seg_row["new_clients"]) > 0 else 0.0
                        seg_row["order_share_segment_pct"] = 100.0
                        seg_row["revenue_share_segment_pct"] = 100.0
                    segment_total_rows.append(seg_row)

            agg = pd.concat([agg, pd.DataFrame(segment_total_rows + [total_row])], ignore_index=True)

            if is_real_estate_preset:
                agg_cols = ["month_name"] + get_real_estate_table_cols(metric_mode)
            elif is_diy_preset:
                agg_cols = [
                    "month_name",
                    "impressions",
                    "reach",
                    "frequency",
                    "clicks",
                    "ctr",
                    "cpc",
                    "cost",
                    "ak_cost_wo_vat",
                    "total_budget_wo_vat_ak",
                    "cost_with_vat",
                    "cost_with_vat_ak",
                    "vat_amount",
                    "cr",
                    "conversions",
                    "aov",
                    "revenue",
                    "cpa",
                    "cpm",
                    "roas",
                    "drr",
                    "available_capacity",
                    "client_count",
                    "absolute_new_clients",
                    "returned_clients",
                    "new_clients",
                    "cac",
                    "order_frequency",
                    "shipped_orders",
                    "shipped_cps",
                    "shipped_aov",
                    "shipped_revenue",
                    "shipped_roas",
                    "shipped_drr_pct",
                    "shipped_order_share_segment_pct",
                    "shipped_revenue_share_segment_pct",
                    "sov_pct",
                    "new_clients_share_pct",
                    "order_share_segment_pct",
                    "revenue_share_segment_pct",
                ]
            else:
                agg_cols = [
                    "month_name",
                    "impressions",
                    "clicks",
                    "conversions",
                    "cost",
                    "cost_with_vat",
                    "cost_with_vat_ak",
                    "revenue",
                    "ctr",
                    "cpc",
                    "cr",
                    "cpm",
                    "cpa",
                    "roas",
                    "drr",
                ]
            agg_show = safe_select_columns(agg, agg_cols)
            agg_show = agg_show.rename(columns={"roas": "ROAS"})
            agg_show = agg_show.rename(columns=DISPLAY_COL_RENAME)

            if is_real_estate_preset:
                numeric_cols = [
                    "impressions", "clicks", "leads", "target_leads", "cost", "cost_with_vat", "cost_with_vat_ak",
                    "ctr", "cpc", "cr_pct", "cr1_pct", "cr2_pct", "cpm", "cpl", "cpql"
                ]
            else:
                numeric_cols = [
                    "impressions", "reach", "frequency", "clicks", "conversions", "cost", "ak_cost_wo_vat", "total_budget_wo_vat_ak", "cost_with_vat", "cost_with_vat_ak", "vat_amount", "revenue",
                    "ctr", "cpc", "cr", "aov", "cpm", "cpa", "ROAS", "drr", "available_capacity", "client_count", "absolute_new_clients", "returned_clients", "new_clients", "cac", "order_frequency", "shipped_orders", "shipped_cps", "shipped_aov", "shipped_revenue", "shipped_roas", "shipped_drr_pct", "shipped_order_share_segment_pct", "shipped_revenue_share_segment_pct", "sov_pct", "new_clients_share_pct", "order_share_segment_pct", "revenue_share_segment_pct"
                ]
            numeric_cols = [DISPLAY_COL_RENAME.get(c, c) for c in numeric_cols]
            numeric_cols = [c for c in numeric_cols if c in agg_show.columns]

            def _blend_hex(c1: str, c2: str, t: float) -> str:
                t = max(0.0, min(1.0, t))
                r1, g1, b1 = int(c1[1:3], 16), int(c1[3:5], 16), int(c1[5:7], 16)
                r2, g2, b2 = int(c2[1:3], 16), int(c2[3:5], 16), int(c2[5:7], 16)
                r = int(r1 + (r2 - r1) * t)
                g = int(g1 + (g2 - g1) * t)
                b = int(b1 + (b2 - b1) * t)
                return f"#{r:02x}{g:02x}{b:02x}"

            def _style_metric_col(col: pd.Series) -> pd.Series:
                styles = pd.Series([""] * len(col), index=col.index)
                mask = ~agg_show["РњРµСЃСЏС†"].astype(str).str.startswith("РС‚РѕРіРѕ")
                mask &= agg_show["РњРµСЃСЏС†"] != "TOTAL"
                vals = pd.to_numeric(col[mask], errors="coerce").dropna()
                if vals.empty:
                    return styles
                vmin, vmax = float(vals.min()), float(vals.max())
                span = (vmax - vmin) if vmax != vmin else 1.0
                base_low = "#173058"
                base_high = "#19B8B2"
                for idx in vals.index:
                    t = (float(col.loc[idx]) - vmin) / span
                    color = _blend_hex(base_low, base_high, t)
                    styles.loc[idx] = (
                        f"background-color: {color}; color: #EAF3EE; border: 1px solid {THEME_BORDER};"
                    )
                return styles

            def _highlight_total_row_total(row):
                style = [""] * len(row)
                month_val = str(row["РњРµСЃСЏС†"])
                if month_val == "TOTAL":
                    style = [
                        f"background-color: #00CDC5; color: #081521; font-weight: 700; border-top: 2px solid {THEME_BORDER};"
                    ] * len(row)
                elif month_val.startswith("РС‚РѕРіРѕ "):
                    style = [
                        f"background-color: #2C6E75; color: #DDEAF0; font-weight: 650; border-top: 1px solid {THEME_BORDER};"
                    ] * len(row)
                return style

            total_formatters = {
                DISPLAY_COL_RENAME["impressions"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["reach"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["frequency"]: lambda x: "" if pd.isna(x) else f"{x:.2f}",
                DISPLAY_COL_RENAME["clicks"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["conversions"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["leads"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["target_leads"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["cost"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["ak_cost_wo_vat"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["total_budget_wo_vat_ak"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["cost_with_vat"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["cost_with_vat_ak"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["vat_amount"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["revenue"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["aov"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["cpm"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["cpa"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["cpl"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["cpql"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["available_capacity"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["client_count"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["absolute_new_clients"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["returned_clients"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["new_clients"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["cac"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["order_frequency"]: lambda x: "" if pd.isna(x) else f"{x:.2f}",
                DISPLAY_COL_RENAME["shipped_orders"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["shipped_cps"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["shipped_aov"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["shipped_revenue"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["shipped_roas"]: lambda x: "" if pd.isna(x) else f"{x * 100:.2f} %",
                DISPLAY_COL_RENAME["shipped_drr_pct"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["shipped_order_share_segment_pct"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["shipped_revenue_share_segment_pct"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["cpc"]: lambda x: "" if pd.isna(x) else f"{x:.2f} \u20BD",
                DISPLAY_COL_RENAME["ctr"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["cr"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["cr1_pct"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["cr2_pct"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["sov_pct"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["new_clients_share_pct"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["order_share_segment_pct"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["revenue_share_segment_pct"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["drr"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                "ROAS": lambda x: "" if pd.isna(x) else f"{x * 100:.2f} %",
            }
            total_formatters = {k: v for k, v in total_formatters.items() if k in agg_show.columns}
            styled_total = (
                agg_show.style
                .format(total_formatters)
                .apply(_highlight_total_row_total, axis=1)
            )
            for c in numeric_cols:
                styled_total = styled_total.apply(_style_metric_col, axis=0, subset=[c])

            st.dataframe(styled_total, use_container_width=True)

    render_bottom_tab_switcher("РњРµРґРёР°РїР»Р°РЅ", "plan")
# ====================================================================
#                           РўРђР‘ В«Р”РђР“Р РђРњРњР«В»
# ====================================================================

with tab_charts:
    ui_section_title("РЎРІРѕРґРЅР°СЏ С‚Р°Р±Р»РёС†Р° Рё РґРёР°РіСЂР°РјРјС‹")
    use_vat_budget_metrics = st.session_state.get("use_vat_budget_metrics", True)
    use_ak_budget_metrics = st.session_state.get("use_ak_budget_metrics", False)
    st.markdown(
        """
        <div style="
            margin: 0.15rem 0 0.55rem 0;
            padding: 10px 12px;
            border-radius: 10px;
            border: 1px solid #2B7EE8;
            background: rgba(10, 116, 242, 0.12);
        ">
            <span style="font-weight: 800; color: #9EC5FF;">РќР°Р·РЅР°С‡РµРЅРёРµ Р±Р»РѕРєР°:</span>
            <span style="color: #EAF0FF; font-weight: 560;">
                Р’РєР»Р°РґРєР° РёСЃРїРѕР»СЊР·СѓРµС‚СЃСЏ РґР»СЏ С„РёРЅР°Р»СЊРЅРѕР№ РїСЂРѕРІРµСЂРєРё СЂР°СЃС‡РµС‚РѕРІ РјРµРґРёР°РїР»Р°РЅР°: С‡РµСЂРµР· С„РёР»СЊС‚СЂС‹ Рё РІРёР·СѓР°Р»РёР·Р°С†РёРё РјРѕР¶РЅРѕ РІР°Р»РёРґРёСЂРѕРІР°С‚СЊ РєРѕСЂСЂРµРєС‚РЅРѕСЃС‚СЊ С„РѕСЂРјСѓР», СЂР°СЃРїСЂРµРґРµР»РµРЅРёР№ Рё РёС‚РѕРіРѕРІС‹С… Р·РЅР°С‡РµРЅРёР№.
            </span>
        </div>
        """,
        unsafe_allow_html=True,
    )
    charts_intro_html = """
        <div class="tab-intro">
            <p>1) РЎРЅР°С‡Р°Р»Р° РІС‹Р±РµСЂРёС‚Рµ РїРµСЂРёРѕРґ (РјРµСЃСЏС†С‹), РєРѕС‚РѕСЂС‹Р№ С…РѕС‚РёС‚Рµ РїСЂРѕРІРµСЂРёС‚СЊ, РёР»Рё РѕСЃС‚Р°РІСЊС‚Рµ РІСЃРµ РЅР°СЃС‚СЂРѕР№РєРё Рё РїСЂРѕРІРµСЂСЊС‚Рµ РІРµСЃСЊ РїРµСЂРёРѕРґ СЂР°СЃС‡РµС‚Р° РјРµРґРёР°РїР»Р°РЅР° С†РµР»РёРєРѕРј.</p>
            <p>2) РџСЂРёРјРµРЅРёС‚Рµ С„РёР»СЊС‚СЂС‹ РїРѕ РЅСѓР¶РЅС‹Рј СЂР°Р·СЂРµР·Р°Рј (С‚РёРї РєР°РјРїР°РЅРёРё, СЃРёСЃС‚РµРјР°, С„РѕСЂРјР°С‚), С‡С‚РѕР±С‹ СЃСѓР·РёС‚СЊ РґР°РЅРЅС‹Рµ.</p>
            <p>3) РџСЂРѕРІРµСЂСЊС‚Рµ СЃРІРѕРґРЅСѓСЋ С‚Р°Р±Р»РёС†Сѓ РїРѕ РјРµСЃСЏС†Р°Рј: РѕРЅР° РїРѕРєР°Р·С‹РІР°РµС‚ РёС‚РѕРіРѕРІС‹Рµ РјРµС‚СЂРёРєРё Рё РїРѕРјРѕРіР°РµС‚ Р±С‹СЃС‚СЂРѕ СѓРІРёРґРµС‚СЊ РѕС‚РєР»РѕРЅРµРЅРёСЏ.</p>
            <p>4) РЎСЂР°РІРЅРёС‚Рµ РґРёРЅР°РјРёРєСѓ РєР»СЋС‡РµРІС‹С… РїРѕРєР°Р·Р°С‚РµР»РµР№ (РїРѕРєР°Р·С‹, РєР»РёРєРё, СЂР°СЃС…РѕРґ, РєРѕРЅРІРµСЂСЃРёРё, РІС‹СЂСѓС‡РєР°, ROAS/Р”Р Р ) РјРµР¶РґСѓ РјРµСЃСЏС†Р°РјРё.</p>
            <p>5) РџСЂРё РЅРµРѕР±С…РѕРґРёРјРѕСЃС‚Рё РІРєР»СЋС‡РёС‚Рµ РєСЂСѓРіРѕРІС‹Рµ РґРёР°РіСЂР°РјРјС‹, С‡С‚РѕР±С‹ РїСЂРѕРІРµСЂРёС‚СЊ СЃС‚СЂСѓРєС‚СѓСЂСѓ СЂР°СЃРїСЂРµРґРµР»РµРЅРёСЏ Р±СЋРґР¶РµС‚Р°, РєРѕРЅРІРµСЂСЃРёР№ Рё РІС‹СЂСѓС‡РєРё РїРѕ С‚РёРїР°Рј Р Рљ.</p>
            <p>6) Р•СЃР»Рё С†РёС„СЂС‹ РІС‹РіР»СЏРґСЏС‚ РЅРµР»РѕРіРёС‡РЅРѕ, РІРµСЂРЅРёС‚РµСЃСЊ РІРѕ РІРєР»Р°РґРєСѓ В«РњРµРґРёР°РїР»Р°РЅВ» Рё РІРЅРµСЃРёС‚Рµ С‚РѕС‡РµС‡РЅС‹Рµ РєРѕСЂСЂРµРєС‚РёСЂРѕРІРєРё РІ РЅСѓР¶РЅС‹Рµ РјРµСЃСЏС†С‹, Р·Р°С‚РµРј РІРµСЂРЅРёС‚РµСЃСЊ РІ В«Р”РёР°РіСЂР°РјРјС‹В» Рё РїСЂРѕРІРµРґРёС‚Рµ РїРѕРІС‚РѕСЂРЅСѓСЋ РїСЂРѕРІРµСЂРєСѓ.</p>
        </div>
    """
    if is_real_estate_preset:
        charts_intro_html = """
        <div class="tab-intro">
            <p>1) РЎРЅР°С‡Р°Р»Р° РІС‹Р±РµСЂРёС‚Рµ РїРµСЂРёРѕРґ (РјРµСЃСЏС†С‹), РєРѕС‚РѕСЂС‹Р№ С…РѕС‚РёС‚Рµ РїСЂРѕРІРµСЂРёС‚СЊ, РёР»Рё РѕСЃС‚Р°РІСЊС‚Рµ РІСЃРµ РЅР°СЃС‚СЂРѕР№РєРё Рё РїСЂРѕРІРµСЂСЊС‚Рµ РІРµСЃСЊ РїРµСЂРёРѕРґ СЂР°СЃС‡РµС‚Р° РјРµРґРёР°РїР»Р°РЅР° С†РµР»РёРєРѕРј.</p>
            <p>2) РџСЂРёРјРµРЅРёС‚Рµ С„РёР»СЊС‚СЂС‹ РїРѕ РЅСѓР¶РЅС‹Рј СЂР°Р·СЂРµР·Р°Рј (С‚РёРї РєР°РјРїР°РЅРёРё, СЃРёСЃС‚РµРјР°, С„РѕСЂРјР°С‚), С‡С‚РѕР±С‹ СЃСѓР·РёС‚СЊ РґР°РЅРЅС‹Рµ.</p>
            <p>3) РџСЂРѕРІРµСЂСЊС‚Рµ СЃРІРѕРґРЅСѓСЋ С‚Р°Р±Р»РёС†Сѓ РїРѕ РјРµСЃСЏС†Р°Рј: РѕРЅР° РїРѕРєР°Р·С‹РІР°РµС‚ РёС‚РѕРіРѕРІС‹Рµ РјРµС‚СЂРёРєРё Рё РїРѕРјРѕРіР°РµС‚ Р±С‹СЃС‚СЂРѕ СѓРІРёРґРµС‚СЊ РѕС‚РєР»РѕРЅРµРЅРёСЏ.</p>
            <p>4) РЎСЂР°РІРЅРёС‚Рµ РґРёРЅР°РјРёРєСѓ РєР»СЋС‡РµРІС‹С… РїРѕРєР°Р·Р°С‚РµР»РµР№ (РїРѕРєР°Р·С‹, РєР»РёРєРё, Р±СЋРґР¶РµС‚, Р»РёРґС‹, Р¦Рћ, CPL/CPQL) РјРµР¶РґСѓ РјРµСЃСЏС†Р°РјРё.</p>
            <p>5) РџСЂРё РЅРµРѕР±С…РѕРґРёРјРѕСЃС‚Рё РІРєР»СЋС‡РёС‚Рµ РєСЂСѓРіРѕРІС‹Рµ РґРёР°РіСЂР°РјРјС‹, С‡С‚РѕР±С‹ РїСЂРѕРІРµСЂРёС‚СЊ СЃС‚СЂСѓРєС‚СѓСЂСѓ СЂР°СЃРїСЂРµРґРµР»РµРЅРёСЏ Р±СЋРґР¶РµС‚Р°, Р»РёРґРѕРІ Рё Р¦Рћ РїРѕ С‚РёРїР°Рј Р Рљ.</p>
            <p>6) Р•СЃР»Рё С†РёС„СЂС‹ РІС‹РіР»СЏРґСЏС‚ РЅРµР»РѕРіРёС‡РЅРѕ, РІРµСЂРЅРёС‚РµСЃСЊ РІРѕ РІРєР»Р°РґРєСѓ В«РњРµРґРёР°РїР»Р°РЅВ» Рё РІРЅРµСЃРёС‚Рµ С‚РѕС‡РµС‡РЅС‹Рµ РєРѕСЂСЂРµРєС‚РёСЂРѕРІРєРё РІ РЅСѓР¶РЅС‹Рµ РјРµСЃСЏС†С‹, Р·Р°С‚РµРј РІРµСЂРЅРёС‚РµСЃСЊ РІ В«Р”РёР°РіСЂР°РјРјС‹В» Рё РїСЂРѕРІРµРґРёС‚Рµ РїРѕРІС‚РѕСЂРЅСѓСЋ РїСЂРѕРІРµСЂРєСѓ.</p>
        </div>
        """
    st.markdown(charts_intro_html, unsafe_allow_html=True)

    if "df_all" not in locals() or df_all.empty:
        st.info("РЎРЅР°С‡Р°Р»Р° Р·Р°РїРѕР»РЅРёС‚Рµ РґР°РЅРЅС‹Рµ РЅР° РІРєР»Р°РґРєРµ В«РњРµРґРёР°РїР»Р°РЅВ» Рё СЂР°СЃСЃС‡РёС‚Р°Р№С‚Рµ РјРµСЃСЏС†С‹.")
    else:
        st.caption("Р’С‹Р±РµСЂРёС‚Рµ РјРµСЃСЏС†С‹ Рё С‚РёРїС‹ Р Рљ РґР»СЏ СЃРІРѕРґРЅРѕР№ С‚Р°Р±Р»РёС†С‹ Рё РєСЂСѓРіРѕРІС‹С… РґРёР°РіСЂР°РјРј.")
        st.caption(
            "Р РµР¶РёРј Р±СЋРґР¶РµС‚РЅС‹С… РјРµС‚СЂРёРє: "
            + (
                "СЃ РќР”РЎ 22% Рё РђРљ"
                if use_ak_budget_metrics and use_vat_budget_metrics
                else "СЃ РќР”РЎ 22%"
                if use_vat_budget_metrics
                else "Р±РµР· РќР”РЎ"
            )
        )

        month_names_full = {
            1: "РЇРЅРІР°СЂСЊ",
            2: "Р¤РµРІСЂР°Р»СЊ",
            3: "РњР°СЂС‚",
            4: "РђРїСЂРµР»СЊ",
            5: "РњР°Р№",
            6: "РСЋРЅСЊ",
            7: "РСЋР»СЊ",
            8: "РђРІРіСѓСЃС‚",
            9: "РЎРµРЅС‚СЏР±СЂСЊ",
            10: "РћРєС‚СЏР±СЂСЊ",
            11: "РќРѕСЏР±СЂСЊ",
            12: "Р”РµРєР°Р±СЂСЊ",
        }

        chart_campaign_palette = [
            "#0066E0", "#00CDC5", "#9747FF", "#FF6333",
            "#3D8EF0", "#42DDD6", "#B07BFF", "#FF8A66",
            "#2D5BFF", "#17B890", "#C77DFF", "#FF9F1C",
        ]

        def _build_campaign_color_map(campaign_names: list[str]) -> dict[str, str]:
            ordered_names = []
            for name in campaign_names:
                name_str = str(name).strip()
                if name_str and name_str not in ordered_names:
                    ordered_names.append(name_str)
            return {
                name: chart_campaign_palette[idx % len(chart_campaign_palette)]
                for idx, name in enumerate(ordered_names)
            }

        available_months = sorted(df_all["month_num"].unique())
        month_options = [f"{m}. {month_names_full[m]}" for m in available_months]

        selected_labels = st.multiselect(
            "РњРµСЃСЏС†С‹ РІ СЃРІРѕРґРЅРѕР№ С‚Р°Р±Р»РёС†Рµ Рё РґРёР°РіСЂР°РјРјР°С…:",
            options=month_options,
            default=month_options,
            key="charts_months_multiselect",
        )

        if not selected_labels:
            st.info("Р’С‹Р±РµСЂРёС‚Рµ С…РѕС‚СЏ Р±С‹ РѕРґРёРЅ РјРµСЃСЏС†.")
        else:
            selected_month_nums_charts = [int(label.split(".")[0]) for label in selected_labels]

            all_campaign_types = df_all["campaign_type"].unique().tolist()
            campaign_color_map = _build_campaign_color_map(all_campaign_types)
            selected_campaign_types = st.multiselect(
                "РўРёРїС‹ Р Рљ РґР»СЏ РІРєР»СЋС‡РµРЅРёСЏ:",
                options=all_campaign_types,
                default=all_campaign_types,
                key="charts_campaign_types_multiselect",
            )

            if not selected_campaign_types:
                st.info("Р’С‹Р±РµСЂРёС‚Рµ С…РѕС‚СЏ Р±С‹ РѕРґРёРЅ С‚РёРї Р Рљ.")
            else:
                selected_segments_charts = None
                if is_diy_preset and "segment" in df_all.columns:
                    all_segments = sorted(df_all["segment"].dropna().astype(str).unique().tolist())
                    selected_segments_charts = st.multiselect(
                        "РЎРµРіРјРµРЅС‚С‹ РґР»СЏ РІРєР»СЋС‡РµРЅРёСЏ:",
                        options=all_segments,
                        default=all_segments,
                        key="charts_segments_multiselect",
                    )
                    if not selected_segments_charts:
                        st.info("Р’С‹Р±РµСЂРёС‚Рµ С…РѕС‚СЏ Р±С‹ РѕРґРёРЅ СЃРµРіРјРµРЅС‚.")
                if is_diy_preset and selected_segments_charts == []:
                    df_sel = pd.DataFrame()
                else:
                    mask = df_all["month_num"].isin(selected_month_nums_charts) & df_all[
                        "campaign_type"
                    ].isin(selected_campaign_types)
                    if selected_segments_charts is not None:
                        mask &= df_all["segment"].isin(selected_segments_charts)
                    df_sel = df_all[mask]

                if df_sel.empty:
                    st.info("Р”Р»СЏ РІС‹Р±СЂР°РЅРЅС‹С… РјРµСЃСЏС†РµРІ Рё С‚РёРїРѕРІ Р Рљ РЅРµС‚ РґР°РЅРЅС‹С….")
                else:
                    # ---------- РЎРІРѕРґРЅР°СЏ С‚Р°Р±Р»РёС†Р° РїРѕ РјРµСЃСЏС†Р°Рј ----------
                    agg_m = df_sel.groupby(["month_num", "month_name"], as_index=False).agg(
                        impressions=("impressions", "sum"),
                        clicks=("clicks", "sum"),
                        conversions=("conversions", "sum"),
                        leads=("leads", "sum"),
                        target_leads=("target_leads", "sum"),
                        cost=("cost", "sum"),
                        cost_with_vat=("cost_with_vat", "sum"),
                        cost_with_vat_ak=("cost_with_vat_ak", "sum"),
                        ak_cost_wo_vat=("ak_cost_wo_vat", "sum"),
                        revenue=("revenue", "sum"),
                    )
                    if use_ak_budget_metrics:
                        budget_series = agg_m["cost_with_vat_ak"] if use_vat_budget_metrics else (agg_m["cost"] + agg_m["ak_cost_wo_vat"])
                    else:
                        budget_series = agg_m["cost_with_vat"] if use_vat_budget_metrics else agg_m["cost"]

                    agg_m["ctr"] = np.where(
                        agg_m["impressions"] > 0,
                        agg_m["clicks"] / agg_m["impressions"] * 100,
                        0.0,
                    )
                    agg_m["cr"] = np.where(
                        agg_m["clicks"] > 0,
                        agg_m["conversions"] / agg_m["clicks"] * 100,
                        0.0,
                    )
                    agg_m["cr1_pct"] = np.where(
                        agg_m["clicks"] > 0,
                        agg_m["leads"] / agg_m["clicks"] * 100,
                        0.0,
                    )
                    agg_m["cr2_pct"] = np.where(
                        agg_m["leads"] > 0,
                        agg_m["target_leads"] / agg_m["leads"] * 100,
                        0.0,
                    )
                    agg_m["cpm"] = np.where(
                        agg_m["impressions"] > 0,
                        budget_series / (agg_m["impressions"] / 1000),
                        0.0,
                    )
                    agg_m["cpa"] = np.where(
                        agg_m["conversions"] > 0,
                        budget_series / agg_m["conversions"],
                        0.0,
                    )
                    agg_m["cpl"] = np.where(
                        agg_m["leads"] > 0,
                        budget_series / agg_m["leads"],
                        0.0,
                    )
                    agg_m["cpql"] = np.where(
                        agg_m["target_leads"] > 0,
                        budget_series / agg_m["target_leads"],
                        0.0,
                    )
                    agg_m["aov"] = np.where(
                        agg_m["conversions"] > 0,
                        agg_m["revenue"] / agg_m["conversions"],
                        0.0,
                    )
                    agg_m["cpc"] = np.where(
                        agg_m["clicks"] > 0,
                        agg_m["cost"] / agg_m["clicks"],
                        0.0,
                    )
                    agg_m["roas"] = np.where(
                        budget_series > 0,
                        agg_m["revenue"] / budget_series,
                        0.0,
                    )
                    agg_m["drr"] = np.where(
                        agg_m["revenue"] > 0,
                        budget_series / agg_m["revenue"] * 100,
                        0.0,
                    )

                    selected_month_nums_charts = sorted(selected_month_nums_charts)
                    month_headers = [month_names_full[m] for m in selected_month_nums_charts]

                    # РљРѕСЌС„С„РёС†РёРµРЅС‚С‹, РёСЃРїРѕР»СЊР·РѕРІР°РЅРЅС‹Рµ РІ СЂР°СЃС‡РµС‚Р°С… (РІ СЃСЂРµРґРЅРµРј РїРѕ РІС‹Р±СЂР°РЅРЅС‹Рј С‚РёРїР°Рј Р Рљ)
                    if "campaigns" in locals() and not campaigns.empty:
                        base_for_coeffs = campaigns[
                            campaigns["campaign_type"].isin(selected_campaign_types)
                        ][
                            [
                                "campaign_type",
                                "impressions_avg",
                                "ctr_avg_percent",
                                "cpc_avg",
                                "cr_avg_percent",
                                "cr2_avg_percent",
                                "aov_avg",
                            ]
                        ].copy()
                        if not base_for_coeffs.empty:
                            coeff_raw = df_sel.merge(base_for_coeffs, on="campaign_type", how="left")
                            coeff_raw["base_ctr"] = coeff_raw["ctr_avg_percent"] / 100.0
                            coeff_raw["base_cr"] = coeff_raw["cr_avg_percent"] / 100.0
                            coeff_raw["base_cr2"] = coeff_raw["cr2_avg_percent"] / 100.0

                            coeff_raw["k_imp"] = np.where(
                                coeff_raw["impressions_avg"] > 0,
                                coeff_raw["impressions"] / coeff_raw["impressions_avg"],
                                np.nan,
                            )
                            coeff_raw["k_ctr"] = np.where(
                                coeff_raw["base_ctr"] > 0,
                                coeff_raw["ctr"] / coeff_raw["base_ctr"],
                                np.nan,
                            )
                            coeff_raw["k_cpc"] = np.where(
                                coeff_raw["cpc_avg"] > 0,
                                coeff_raw["cpc"] / coeff_raw["cpc_avg"],
                                np.nan,
                            )
                            coeff_raw["k_cr"] = np.where(
                                coeff_raw["base_cr"] > 0,
                                coeff_raw["cr"] / coeff_raw["base_cr"],
                                np.nan,
                            )
                            coeff_raw["k_cr2"] = np.where(
                                coeff_raw["base_cr2"] > 0,
                                coeff_raw.get("cr2", np.nan) / coeff_raw["base_cr2"],
                                np.nan,
                            )
                            coeff_raw["k_aov"] = np.where(
                                coeff_raw["aov_avg"] > 0,
                                coeff_raw["aov"] / coeff_raw["aov_avg"],
                                np.nan,
                            )

                            agg_kwargs = {
                                "k_imp": ("k_imp", "mean"),
                                "k_ctr": ("k_ctr", "mean"),
                                "k_cpc": ("k_cpc", "mean"),
                                "k_cr": ("k_cr", "mean"),
                            }
                            if is_real_estate_preset and metric_mode["is_real_estate_full"]:
                                agg_kwargs["k_cr2"] = ("k_cr2", "mean")
                            if not is_real_estate_preset:
                                agg_kwargs["k_aov"] = ("k_aov", "mean")
                            coeff_month = coeff_raw.groupby("month_num", as_index=False).agg(**agg_kwargs)
                            coeff_month_map = coeff_month.set_index("month_num").to_dict("index")

                            # РћС‚РґРµР»СЊРЅРѕ РїРѕРєР°Р·С‹РІР°РµРј РєРѕСЌС„С„РёС†РёРµРЅС‚ "РјРµРґРёР№РЅС‹С… С…РІРѕСЃС‚РѕРІ" РєР°Рє Р·РЅР°С‡РµРЅРёРµ РёР· РЅР°Р±РѕСЂР°
                            # (Р±РµР· СѓСЃСЂРµРґРЅРµРЅРёСЏ РїРѕ С‚РёРїР°Рј Р Рљ)
                            media_tail_month_map = {}
                            media_tail_values_by_set = {}
                            coeff_sets_for_tail = st.session_state.get("coeff_sets", [])
                            for cs_tail in coeff_sets_for_tail:
                                if normalize_coeff_set_type(cs_tail.get("type")) != "РњРµРґРёР№РЅС‹Рµ С…РІРѕСЃС‚С‹":
                                    continue
                                cs_tail_name = str(cs_tail.get("name", "")).strip()
                                df_tail = cs_tail.get("result")
                                if not cs_tail_name or df_tail is None or getattr(df_tail, "empty", True):
                                    continue
                                if "РќРѕРјРµСЂ РјРµСЃСЏС†Р°" not in df_tail.columns or "РљРѕСЌС„." not in df_tail.columns:
                                    continue
                                mm = {}
                                for _, rtail in df_tail.iterrows():
                                    try:
                                        mm[int(rtail["РќРѕРјРµСЂ РјРµСЃСЏС†Р°"])] = float(rtail["РљРѕСЌС„."])
                                    except Exception:
                                        continue
                                media_tail_values_by_set[cs_tail_name] = mm

                            coeff_links_for_tail = st.session_state.get("coeff_sets_links_new", pd.DataFrame())
                            coeff_links_for_tail = coeff_links_for_tail[
                                coeff_links_for_tail["campaign_type"].isin(selected_campaign_types)
                            ] if not coeff_links_for_tail.empty else pd.DataFrame()

                            for m in selected_month_nums_charts:
                                set_names = []
                                if not coeff_links_for_tail.empty and "media_tail_set" in coeff_links_for_tail.columns:
                                    for _, rr in coeff_links_for_tail.iterrows():
                                        set_name = str(rr.get("media_tail_set", "")).strip()
                                        if set_name:
                                            set_names.append(set_name)

                                uniq_set_names = sorted(set(set_names))
                                if len(uniq_set_names) == 1:
                                    media_tail_month_map[m] = float(
                                        media_tail_values_by_set.get(uniq_set_names[0], {}).get(m, 1.0)
                                    )
                                elif len(uniq_set_names) == 0:
                                    media_tail_month_map[m] = 1.0
                                else:
                                    media_tail_month_map[m] = None

                            coeff_rows = []
                            coeff_labels = [
                                ("k_imp", "РљРѕСЌС„. РїРѕРєР°Р·С‹"),
                                ("k_ctr", "РљРѕСЌС„. CTR"),
                                ("k_cpc", "РљРѕСЌС„. CPC"),
                                ("k_cr", "РљРѕСЌС„. CR"),
                            ]
                            if is_real_estate_preset and metric_mode["is_real_estate_full"]:
                                coeff_labels.append(("k_cr2", "РљРѕСЌС„. CR2"))
                            if not is_real_estate_preset:
                                coeff_labels.append(("k_aov", "РљРѕСЌС„. AOV"))
                            for key_name, label in coeff_labels:
                                row = {"РњРµС‚СЂРёРєР°": label}
                                for m in selected_month_nums_charts:
                                    val = coeff_month_map.get(m, {}).get(key_name, np.nan)
                                    row[month_names_full[m]] = "" if pd.isna(val) else f"{val:.2f}"
                                coeff_rows.append(row)

                            row_media_tail = {"РњРµС‚СЂРёРєР°": "РљРѕСЌС„. РјРµРґРёР№РЅС‹Рµ С…РІРѕСЃС‚С‹"}
                            for m in selected_month_nums_charts:
                                v = media_tail_month_map.get(m, 1.0)
                                row_media_tail[month_names_full[m]] = "РЅРµСЃРє." if v is None else f"{v:.2f}"
                            coeff_rows.append(row_media_tail)

                            coeff_table = pd.DataFrame(coeff_rows).set_index("РњРµС‚СЂРёРєР°")
                            with st.expander("РљРѕСЌС„С„РёС†РёРµРЅС‚С‹, РёСЃРїРѕР»СЊР·РѕРІР°РЅРЅС‹Рµ РІ СЂР°СЃС‡РµС‚Рµ РїРѕ РјРµСЃСЏС†Р°Рј", expanded=False):
                                st.dataframe(coeff_table, use_container_width=True)
                            st.caption("РљРѕСЌС„. РјРµРґРёР№РЅС‹Рµ С…РІРѕСЃС‚С‹: 1.00 = С…РІРѕСЃС‚С‹ РЅРµ РїСЂРёРјРµРЅСЏСЋС‚СЃСЏ.")

                    # РўСЂР°РЅСЃРїРѕРЅРёСЂРѕРІР°РЅРЅР°СЏ С‚Р°Р±Р»РёС†Р°: РјРµСЃСЏС†С‹ РІ РєРѕР»РѕРЅРєР°С…, РјРµС‚СЂРёРєРё РІ СЃС‚СЂРѕРєР°С…
                    agg_by_month = agg_m.set_index("month_num").to_dict("index")

                    total_impressions = agg_m["impressions"].sum()
                    total_clicks = agg_m["clicks"].sum()
                    total_conversions = agg_m["conversions"].sum()
                    total_cost = agg_m["cost"].sum()
                    total_cost_with_vat = agg_m["cost_with_vat"].sum()
                    total_cost_with_vat_ak = agg_m["cost_with_vat_ak"].sum()
                    total_ak_wo_vat = agg_m["ak_cost_wo_vat"].sum()
                    total_revenue = agg_m["revenue"].sum()
                    if use_ak_budget_metrics:
                        total_budget_basis = total_cost_with_vat_ak if use_vat_budget_metrics else (total_cost + total_ak_wo_vat)
                    else:
                        total_budget_basis = total_cost_with_vat if use_vat_budget_metrics else total_cost
                    total_metrics = {
                        "impressions": total_impressions,
                        "clicks": total_clicks,
                        "ctr": (total_clicks / total_impressions * 100) if total_impressions > 0 else 0.0,
                        "cpc": (total_cost / total_clicks) if total_clicks > 0 else 0.0,
                        "cpm": (total_budget_basis / (total_impressions / 1000.0)) if total_impressions > 0 else 0.0,
                        "cost": total_cost,
                        "cost_with_vat": total_cost_with_vat,
                        "cost_with_vat_ak": total_cost_with_vat_ak,
                        "cr": (total_conversions / total_clicks * 100) if total_clicks > 0 else 0.0,
                        "cr1_pct": (float(agg_m["leads"].sum()) / total_clicks * 100) if total_clicks > 0 else 0.0,
                        "cr2_pct": (float(agg_m["target_leads"].sum()) / float(agg_m["leads"].sum()) * 100) if float(agg_m["leads"].sum()) > 0 else 0.0,
                        "leads": float(agg_m["leads"].sum()),
                        "target_leads": float(agg_m["target_leads"].sum()),
                        "cpl": (total_budget_basis / float(agg_m["leads"].sum())) if float(agg_m["leads"].sum()) > 0 else 0.0,
                        "cpql": (total_budget_basis / float(agg_m["target_leads"].sum())) if float(agg_m["target_leads"].sum()) > 0 else 0.0,
                        "aov": (total_revenue / total_conversions) if total_conversions > 0 else 0.0,
                        "conversions": total_conversions,
                        "cpa": (total_budget_basis / total_conversions) if total_conversions > 0 else 0.0,
                        "revenue": total_revenue,
                        "roas": (total_revenue / total_budget_basis) if total_budget_basis > 0 else 0.0,
                        "drr": (total_budget_basis / total_revenue * 100) if total_revenue > 0 else 0.0,
                    }

                    def _fmt_int(v):
                        return f"{round(v):,}".replace(",", " ")

                    def _fmt_rub(v):
                        return f"{round(v):,} в‚Ѕ".replace(",", " ")

                    def _fmt_rub2(v):
                        return f"{v:.2f} в‚Ѕ"

                    def _fmt_pct(v):
                        return f"{v:.2f} %"

                    def _fmt_mult(v):
                        return f"{v * 100:.2f} %"

                    if is_real_estate_preset:
                        metric_specs = get_real_estate_display_metric_specs(metric_mode)
                    else:
                        metric_specs = [
                            ("РџРѕРєР°Р·С‹", "impressions"),
                            ("РљР»РёРєРё", "clicks"),
                            ("CTR", "ctr"),
                            ("CPC", "cpc"),
                            ("Р Р°СЃС…РѕРґ", "cost"),
                            ("Р‘СЋРґР¶РµС‚ СЃ РќР”РЎ", "cost_with_vat"),
                            ("Р‘СЋРґР¶РµС‚ СЃ РќР”РЎ Рё РђРљ", "cost_with_vat_ak"),
                            ("CR", "cr"),
                            ("AOV", "aov"),
                            ("РљРѕРЅРІРµСЂСЃРёРё", "conversions"),
                            ("CPO", "cpa"),
                            ("Р’С‹СЂСѓС‡РєР°", "revenue"),
                            ("ROAS", "roas"),
                            ("Р”Р Р ", "drr"),
                        ]

                    rows = []
                    metric_series_map = {}
                    for metric_label, metric_key in metric_specs:
                        row = {"РњРµС‚СЂРёРєР°": metric_label}
                        month_value_map = {}
                        for m in selected_month_nums_charts:
                            month_col = month_names_full[m]
                            val = agg_by_month.get(m, {}).get(metric_key, 0.0)
                            month_value_map[month_col] = float(val)
                            if metric_key in ["impressions", "clicks", "conversions", "leads", "target_leads"]:
                                row[month_col] = _fmt_int(val)
                            elif metric_key in ["cost", "cost_with_vat", "cost_with_vat_ak", "revenue", "cpm", "cpa", "aov", "cpl", "cpql"]:
                                row[month_col] = _fmt_rub(val)
                            elif metric_key == "cpc":
                                row[month_col] = _fmt_rub2(val)
                            elif metric_key in ["ctr", "cr", "drr", "cr1_pct", "cr2_pct"]:
                                row[month_col] = _fmt_pct(val)
                            elif metric_key == "roas":
                                row[month_col] = _fmt_mult(val)
                            else:
                                row[month_col] = val

                        total_val = total_metrics.get(metric_key, 0.0)
                        if metric_key in ["impressions", "clicks", "conversions", "leads", "target_leads"]:
                            row["TOTAL"] = _fmt_int(total_val)
                        elif metric_key in ["cost", "cost_with_vat", "cost_with_vat_ak", "revenue", "cpm", "cpa", "aov", "cpl", "cpql"]:
                            row["TOTAL"] = _fmt_rub(total_val)
                        elif metric_key == "cpc":
                            row["TOTAL"] = _fmt_rub2(total_val)
                        elif metric_key in ["ctr", "cr", "drr", "cr1_pct", "cr2_pct"]:
                            row["TOTAL"] = _fmt_pct(total_val)
                        elif metric_key == "roas":
                            row["TOTAL"] = _fmt_mult(total_val)
                        else:
                            row["TOTAL"] = _fmt_int(total_val)
                        rows.append(row)
                        metric_series_map[metric_label] = {
                            "metric_key": metric_key,
                            "months": month_value_map,
                            "total": float(total_val),
                        }

                    summary_matrix = pd.DataFrame(rows).set_index("РњРµС‚СЂРёРєР°")
                    ui_section_title("РЎРІРѕРґРЅР°СЏ С‚Р°Р±Р»РёС†Р° (РјРµСЃСЏС†С‹ РІ РєРѕР»РѕРЅРєР°С…)")
                    heat_cols = month_headers

                    def _parse_num_for_heat(v):
                        if pd.isna(v):
                            return None
                        if isinstance(v, (int, float, np.integer, np.floating)):
                            return float(v)
                        s = str(v).strip()
                        if not s:
                            return None
                        for token in ["в‚Ѕ", "%", "Г—"]:
                            s = s.replace(token, "")
                        s = s.replace(" ", "").replace(",", ".")
                        try:
                            return float(s)
                        except Exception:
                            return None

                    def _blend_hex(c1: str, c2: str, t: float) -> str:
                        t = max(0.0, min(1.0, t))
                        r1, g1, b1 = int(c1[1:3], 16), int(c1[3:5], 16), int(c1[5:7], 16)
                        r2, g2, b2 = int(c2[1:3], 16), int(c2[3:5], 16), int(c2[5:7], 16)
                        r = int(r1 + (r2 - r1) * t)
                        g = int(g1 + (g2 - g1) * t)
                        b = int(b1 + (b2 - b1) * t)
                        return f"#{r:02x}{g:02x}{b:02x}"

                    base_low = "#173058"
                    base_high = "#19B8B2"

                    def _row_heat_style(row):
                        styles = [""] * len(row)
                        vals = [_parse_num_for_heat(row.get(col)) for col in heat_cols]
                        nums = [v for v in vals if v is not None]
                        if not nums:
                            return styles

                        vmin, vmax = min(nums), max(nums)
                        span = (vmax - vmin) if vmax != vmin else 1.0

                        for i, col in enumerate(row.index):
                            if col not in heat_cols:
                                continue
                            value = _parse_num_for_heat(row[col])
                            if value is None:
                                continue
                            t = (value - vmin) / span
                            color = _blend_hex(base_low, base_high, t)
                            styles[i] = (
                                f"background-color: {color}; color: #EAF3EE; "
                                f"border: 1px solid {THEME_BORDER};"
                            )
                        return styles

                    st.caption(
                        "РћС‚РјРµС‡Р°Р№С‚Рµ С‡РµРєР±РѕРєСЃС‹ СЃР»РµРІР° Сѓ РЅСѓР¶РЅС‹С… СЃС‚СЂРѕРє: РґР»СЏ РѕРґРЅРѕР№ РјРµС‚СЂРёРєРё СЃС‚СЂРѕРёС‚СЃСЏ РіСЂР°С„РёРє С„Р°РєС‚РёС‡РµСЃРєРёС… Р·РЅР°С‡РµРЅРёР№, "
                        "РґР»СЏ РЅРµСЃРєРѕР»СЊРєРёС… РјРµС‚СЂРёРє вЂ” СЃСЂР°РІРЅРёС‚РµР»СЊРЅР°СЏ РґРёРЅР°РјРёРєР° РІ РѕС‚РЅРѕСЃРёС‚РµР»СЊРЅРѕРј РІРёРґРµ."
                    )

                    selection_state_key = "charts_summary_selected_metrics"
                    prev_selected_metrics = st.session_state.get(selection_state_key, [])
                    styled_summary = summary_matrix.style.apply(_row_heat_style, axis=1)
                    summary_selector_df = pd.DataFrame(
                        {
                            "Р“СЂР°С„РёРє": summary_matrix.index.isin(prev_selected_metrics),
                        }
                    )
                    selector_col, table_col = st.columns([0.07, 0.93], vertical_alignment="top")
                    with selector_col:
                        summary_editor = st.data_editor(
                            summary_selector_df,
                            use_container_width=True,
                            hide_index=True,
                            column_config={
                                "Р“СЂР°С„РёРє": st.column_config.CheckboxColumn(
                                    "Р“СЂР°С„РёРє",
                                    help="РћС‚РјРµС‚СЊС‚Рµ СЃС‚СЂРѕРєСѓ, С‡С‚РѕР±С‹ СЃСЂР°Р·Сѓ РїРѕСЃС‚СЂРѕРёС‚СЊ РіСЂР°С„РёРє РЅРёР¶Рµ.",
                                    width="small",
                                ),
                            },
                            key="charts_summary_selector_editor",
                        )
                    with table_col:
                        st.dataframe(styled_summary, use_container_width=True)

                    selected_metrics_for_chart = summary_editor.loc[
                        summary_editor["Р“СЂР°С„РёРє"].fillna(False),
                    ].index.tolist()
                    selected_metrics_for_chart = [
                        str(summary_matrix.index[idx])
                        for idx in selected_metrics_for_chart
                        if 0 <= idx < len(summary_matrix.index)
                    ]
                    st.session_state[selection_state_key] = selected_metrics_for_chart

                    if selected_metrics_for_chart:
                        chart_rows = []
                        if len(selected_metrics_for_chart) == 1:
                            metric_label = selected_metrics_for_chart[0]
                            metric_payload = metric_series_map.get(metric_label, {})
                            month_map = metric_payload.get("months", {})
                            for month_col in month_headers:
                                chart_rows.append(
                                    {
                                        "РњРµСЃСЏС†": month_col,
                                        "РњРµС‚СЂРёРєР°": metric_label,
                                        "Р—РЅР°С‡РµРЅРёРµ": float(month_map.get(month_col, 0.0)),
                                    }
                                )
                            chart_title = f"Р”РёРЅР°РјРёРєР° РјРµС‚СЂРёРєРё: {metric_label}"
                            chart_y_title = metric_label
                            chart_note = None
                        else:
                            for metric_label in selected_metrics_for_chart:
                                metric_payload = metric_series_map.get(metric_label, {})
                                month_map = metric_payload.get("months", {})
                                base_value = None
                                for month_col in month_headers:
                                    value = float(month_map.get(month_col, 0.0))
                                    if base_value is None and value != 0:
                                        base_value = value
                                if base_value is None:
                                    base_value = 0.0
                                for month_col in month_headers:
                                    value = float(month_map.get(month_col, 0.0))
                                    relative_value = (value / base_value * 100.0) if base_value not in (None, 0.0) else 0.0
                                    chart_rows.append(
                                        {
                                            "РњРµСЃСЏС†": month_col,
                                            "РњРµС‚СЂРёРєР°": metric_label,
                                            "Р—РЅР°С‡РµРЅРёРµ": relative_value,
                                        }
                                    )
                            chart_title = "РЎСЂР°РІРЅРµРЅРёРµ РґРёРЅР°РјРёРєРё РІС‹Р±СЂР°РЅРЅС‹С… РјРµС‚СЂРёРє"
                            chart_y_title = "РРЅРґРµРєСЃ, %"
                            chart_note = (
                                "Р”Р»СЏ РЅРµСЃРєРѕР»СЊРєРёС… СЃС‚СЂРѕРє РіСЂР°С„РёРє СЃС‚СЂРѕРёС‚СЃСЏ РІ РѕС‚РЅРѕСЃРёС‚РµР»СЊРЅРѕРј РІРёРґРµ: "
                                "РїРµСЂРІР°СЏ РЅРµРЅСѓР»РµРІР°СЏ С‚РѕС‡РєР° РєР°Р¶РґРѕР№ РјРµС‚СЂРёРєРё РїСЂРёРЅРёРјР°РµС‚СЃСЏ Р·Р° 100%."
                            )

                        chart_df = pd.DataFrame(chart_rows)
                        fig_selected_metrics = px.line(
                            chart_df,
                            x="РњРµСЃСЏС†",
                            y="Р—РЅР°С‡РµРЅРёРµ",
                            color="РњРµС‚СЂРёРєР°",
                            markers=True,
                            title=chart_title,
                        )
                        fig_selected_metrics.update_layout(
                            height=420,
                            margin=dict(l=12, r=12, t=64, b=12),
                            paper_bgcolor="rgba(0,0,0,0)",
                            plot_bgcolor="rgba(0,0,0,0)",
                            font=dict(color=THEME_PLOT_TEXT),
                            xaxis_title="РњРµСЃСЏС†",
                            yaxis_title=chart_y_title,
                            legend_title_text="РњРµС‚СЂРёРєР°",
                        )
                        fig_selected_metrics.update_traces(line=dict(width=3))
                        st.plotly_chart(fig_selected_metrics, use_container_width=True)
                        if chart_note:
                            st.caption(chart_note)
                    else:
                        st.caption("РћС‚РјРµС‚СЊС‚Рµ С‡РµРєР±РѕРєСЃ СЃР»РµРІР° Сѓ РЅСѓР¶РЅРѕР№ СЃС‚СЂРѕРєРё, С‡С‚РѕР±С‹ РїРѕСЃС‚СЂРѕРёС‚СЊ РіСЂР°С„РёРє РїРѕ РјРµС‚СЂРёРєРµ.")

                    # ---------- РљСЂСѓРіРѕРІС‹Рµ РґРёР°РіСЂР°РјРјС‹ РїРѕ С‚РёРїР°Рј Р Рљ ----------

                    ui_section_title("Р Р°СЃРїСЂРµРґРµР»РµРЅРёРµ РїРѕ С‚РёРїР°Рј Р Рљ")

                    agg_ct = df_sel.groupby("campaign_type", as_index=False).agg(
                        cost=("cost", "sum"),
                        conversions=("conversions", "sum"),
                        leads=("leads", "sum"),
                        target_leads=("target_leads", "sum"),
                        revenue=("revenue", "sum"),
                    )
                    show_pies = st.checkbox(
                        "РџРѕРєР°Р·С‹РІР°С‚СЊ РєСЂСѓРіРѕРІС‹Рµ РґРёР°РіСЂР°РјРјС‹",
                        value=True,
                        key="charts_show_pies",
                    )

                    if show_pies:
                        pie_color_map = {
                            ct_name: campaign_color_map.get(ct_name, chart_campaign_palette[0])
                            for ct_name in agg_ct["campaign_type"].astype(str).tolist()
                        }
                        st.caption("РћР±С‰Р°СЏ Р»РµРіРµРЅРґР°")
                        legend_cols_count = min(4, max(1, len(agg_ct)))
                        legend_cols = st.columns(legend_cols_count)
                        for i, ct_name in enumerate(agg_ct["campaign_type"].tolist()):
                            c = pie_color_map.get(str(ct_name), chart_campaign_palette[0])
                            with legend_cols[i % legend_cols_count]:
                                st.markdown(
                                    f"<span style='color:{c}; font-weight:700;'>в—Џ</span> "
                                    f"<span style='color:{THEME_LEGEND_TEXT};'>{ct_name}</span>",
                                    unsafe_allow_html=True,
                                )

                        col_pie1, col_pie2, col_pie3 = st.columns(3)
                        if is_real_estate_preset:
                            pie_specs = [
                                ("cost", "Р”РѕР»СЏ Р±СЋРґР¶РµС‚Р° РїРѕ С‚РёРїР°Рј Р Рљ"),
                                ("leads" if metric_mode["is_real_estate_full"] else "target_leads", "Р”РѕР»СЏ Р»РёРґРѕРІ РїРѕ С‚РёРїР°Рј Р Рљ" if metric_mode["is_real_estate_full"] else "Р”РѕР»СЏ Р¦Рћ РїРѕ С‚РёРїР°Рј Р Рљ"),
                                ("target_leads" if metric_mode["is_real_estate_full"] else "clicks", "Р”РѕР»СЏ Р¦Рћ РїРѕ С‚РёРїР°Рј Р Рљ" if metric_mode["is_real_estate_full"] else "Р”РѕР»СЏ РєР»РёРєРѕРІ РїРѕ С‚РёРїР°Рј Р Рљ"),
                            ]
                            if "clicks" not in agg_ct.columns:
                                agg_ct = agg_ct.merge(df_sel.groupby("campaign_type", as_index=False).agg(clicks=("clicks", "sum")), on="campaign_type", how="left")
                        else:
                            pie_specs = [
                                ("cost", "Р”РѕР»СЏ Р±СЋРґР¶РµС‚Р° РїРѕ С‚РёРїР°Рј Р Рљ"),
                                ("conversions", "Р”РѕР»СЏ РєРѕРЅРІРµСЂСЃРёР№ РїРѕ С‚РёРїР°Рј Р Рљ"),
                                ("revenue", "Р”РѕР»СЏ РґРѕС…РѕРґР° РїРѕ С‚РёРїР°Рј Р Рљ"),
                            ]
                        with col_pie1:
                            fig_cost = px.pie(
                                agg_ct,
                                names="campaign_type",
                                values=pie_specs[0][0],
                                title=pie_specs[0][1],
                                color="campaign_type",
                                color_discrete_map=pie_color_map,
                            )
                            fig_cost.update_traces(textposition="inside", textinfo="percent+label")
                            fig_cost.update_layout(
                                height=520,
                                margin=dict(l=12, r=12, t=64, b=12),
                                showlegend=False,
                                paper_bgcolor="rgba(0,0,0,0)",
                                plot_bgcolor="rgba(0,0,0,0)",
                                font=dict(color=THEME_PLOT_TEXT),
                            )
                            st.plotly_chart(fig_cost, use_container_width=True)

                        with col_pie2:
                            fig_conv = px.pie(
                                agg_ct,
                                names="campaign_type",
                                values=pie_specs[1][0],
                                title=pie_specs[1][1],
                                color="campaign_type",
                                color_discrete_map=pie_color_map,
                            )
                            fig_conv.update_traces(textposition="inside", textinfo="percent+label")
                            fig_conv.update_layout(
                                height=520,
                                margin=dict(l=12, r=12, t=64, b=12),
                                showlegend=False,
                                paper_bgcolor="rgba(0,0,0,0)",
                                plot_bgcolor="rgba(0,0,0,0)",
                                font=dict(color=THEME_PLOT_TEXT),
                            )
                            st.plotly_chart(fig_conv, use_container_width=True)

                        with col_pie3:
                            fig_rev = px.pie(
                                agg_ct,
                                names="campaign_type",
                                values=pie_specs[2][0],
                                title=pie_specs[2][1],
                                color="campaign_type",
                                color_discrete_map=pie_color_map,
                            )
                            fig_rev.update_traces(textposition="inside", textinfo="percent+label")
                            fig_rev.update_layout(
                                height=520,
                                margin=dict(l=12, r=12, t=64, b=12),
                                showlegend=False,
                                paper_bgcolor="rgba(0,0,0,0)",
                                plot_bgcolor="rgba(0,0,0,0)",
                                font=dict(color=THEME_PLOT_TEXT),
                            )
                            st.plotly_chart(fig_rev, use_container_width=True)

                        if is_diy_preset and "segment" in df_sel.columns:
                            seg_work = df_sel.copy()
                            seg_work["segment_norm"] = (
                                seg_work["segment"].astype(str).str.upper().str.strip()
                            )
                            seg_work = seg_work[seg_work["segment_norm"].isin(["B2B", "B2C"])]

                            if not seg_work.empty:
                                if use_ak_budget_metrics:
                                    if use_vat_budget_metrics:
                                        seg_work["budget_basis"] = seg_work["cost_with_vat_ak"]
                                    else:
                                        seg_work["budget_basis"] = seg_work["cost"] + seg_work["ak_cost_wo_vat"]
                                else:
                                    seg_work["budget_basis"] = (
                                        seg_work["cost_with_vat"] if use_vat_budget_metrics else seg_work["cost"]
                                    )

                                agg_seg = (
                                    seg_work.groupby("segment_norm", as_index=False)
                                    .agg(
                                        budget_basis=("budget_basis", "sum"),
                                        revenue=("revenue", "sum"),
                                        orders=("conversions", "sum"),
                                    )
                                )
                                agg_seg["segment_norm"] = pd.Categorical(
                                    agg_seg["segment_norm"], categories=["B2C", "B2B"], ordered=True
                                )
                                agg_seg = agg_seg.sort_values("segment_norm")

                                ui_section_title("DIY: Р Р°СЃРїСЂРµРґРµР»РµРЅРёРµ РјРµР¶РґСѓ B2B Рё B2C")
                                seg_palette = ["#3D8EF0", "#00CDC5"]
                                seg_pie1, seg_pie2, seg_pie3 = st.columns(3)

                                with seg_pie1:
                                    fig_seg_budget = px.pie(
                                        agg_seg,
                                        names="segment_norm",
                                        values="budget_basis",
                                        title="Р”РѕР»СЏ Р±СЋРґР¶РµС‚Р° РјРµР¶РґСѓ B2B Рё B2C",
                                        color_discrete_sequence=seg_palette,
                                    )
                                    fig_seg_budget.update_traces(textposition="inside", textinfo="percent+label")
                                    fig_seg_budget.update_layout(
                                        height=520,
                                        margin=dict(l=12, r=12, t=64, b=12),
                                        showlegend=False,
                                        paper_bgcolor="rgba(0,0,0,0)",
                                        plot_bgcolor="rgba(0,0,0,0)",
                                        font=dict(color=THEME_PLOT_TEXT),
                                    )
                                    st.plotly_chart(fig_seg_budget, use_container_width=True)

                                with seg_pie2:
                                    fig_seg_rev = px.pie(
                                        agg_seg,
                                        names="segment_norm",
                                        values="revenue",
                                        title="Р”РѕР»СЏ РІС‹СЂСѓС‡РєРё РјРµР¶РґСѓ B2B Рё B2C",
                                        color_discrete_sequence=seg_palette,
                                    )
                                    fig_seg_rev.update_traces(textposition="inside", textinfo="percent+label")
                                    fig_seg_rev.update_layout(
                                        height=520,
                                        margin=dict(l=12, r=12, t=64, b=12),
                                        showlegend=False,
                                        paper_bgcolor="rgba(0,0,0,0)",
                                        plot_bgcolor="rgba(0,0,0,0)",
                                        font=dict(color=THEME_PLOT_TEXT),
                                    )
                                    st.plotly_chart(fig_seg_rev, use_container_width=True)

                                with seg_pie3:
                                    fig_seg_orders = px.pie(
                                        agg_seg,
                                        names="segment_norm",
                                        values="orders",
                                        title="Р”РѕР»СЏ Р·Р°РєР°Р·РѕРІ РјРµР¶РґСѓ B2B Рё B2C",
                                        color_discrete_sequence=seg_palette,
                                    )
                                    fig_seg_orders.update_traces(textposition="inside", textinfo="percent+label")
                                    fig_seg_orders.update_layout(
                                        height=520,
                                        margin=dict(l=12, r=12, t=64, b=12),
                                        showlegend=False,
                                        paper_bgcolor="rgba(0,0,0,0)",
                                        plot_bgcolor="rgba(0,0,0,0)",
                                        font=dict(color=THEME_PLOT_TEXT),
                                    )
                                    st.plotly_chart(fig_seg_orders, use_container_width=True)

    render_bottom_tab_switcher("Р”РёР°РіСЂР°РјРјС‹", "charts")

# ====================================================================
#                           РўРђР‘ В«EXPORTВ»
# ====================================================================

with tab_export:
    ui_section_title("Р­РєСЃРїРѕСЂС‚")
    st.markdown(
        """
        <div class="tab-intro">
            <p>Р’РєР»Р°РґРєР° РїРѕРґРіРѕС‚РѕРІРєРё Рё Р·Р°РїСѓСЃРєР° СЌРєСЃРїРѕСЂС‚Р° РјРµРґРёР°РїР»Р°РЅР° РІ Excel.</p>
            <p>Р”РѕСЃС‚СѓРїРЅС‹: РІС‹РіСЂСѓР·РєР° РґРµС‚Р°Р»СЊРЅС‹С… РґР°РЅРЅС‹С…, СЃРІРѕРґРєРё РїРѕ РјРµСЃСЏС†Р°Рј Рё СЌРєСЃРїРѕСЂС‚ РІ С€Р°Р±Р»РѕРЅ.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown("### Р­РєСЃРїРѕСЂС‚ РїСЂРѕРµРєС‚Р°")
    st.caption("РЎРѕС…СЂР°РЅРёС‚Рµ С‚РµРєСѓС‰РµРµ СЃРѕСЃС‚РѕСЏРЅРёРµ РїСЂРѕРµРєС‚Р° РІ JSON, С‡С‚РѕР±С‹ РїРµСЂРµРґР°С‚СЊ РµРіРѕ РґСЂСѓРіРѕРјСѓ РїРѕР»СЊР·РѕРІР°С‚РµР»СЋ.")

    project_payload = export_project_state_payload()
    project_json = json.dumps(project_payload, ensure_ascii=False, indent=2).encode("utf-8")
    project_ts = dt.datetime.now().strftime("%Y%m%d_%H%M")
    st.download_button(
        "РЎРєР°С‡Р°С‚СЊ РїСЂРѕРµРєС‚ (JSON)",
        data=project_json,
        file_name=f"mediaplan_project_{project_ts}.json",
        mime="application/json",
        key="download_project_json",
    )

    st.markdown("---")
    if "df_all" not in locals() or df_all.empty:
        st.info("РќРµС‚ РґР°РЅРЅС‹С… РґР»СЏ СЌРєСЃРїРѕСЂС‚Р°. РЎРЅР°С‡Р°Р»Р° СЂР°СЃСЃС‡РёС‚Р°Р№С‚Рµ РјРµРґРёР°РїР»Р°РЅ РЅР° РІРєР»Р°РґРєРµ В«РњРµРґРёР°РїР»Р°РЅВ».")
    else:
        month_names_full = {
            1: "РЇРЅРІР°СЂСЊ",
            2: "Р¤РµРІСЂР°Р»СЊ",
            3: "РњР°СЂС‚",
            4: "РђРїСЂРµР»СЊ",
            5: "РњР°Р№",
            6: "РСЋРЅСЊ",
            7: "РСЋР»СЊ",
            8: "РђРІРіСѓСЃС‚",
            9: "РЎРµРЅС‚СЏР±СЂСЊ",
            10: "РћРєС‚СЏР±СЂСЊ",
            11: "РќРѕСЏР±СЂСЊ",
            12: "Р”РµРєР°Р±СЂСЊ",
        }

        export_months_available = sorted(df_all["month_num"].dropna().astype(int).unique().tolist())
        export_month_options = [f"{m}. {month_names_full.get(m, str(m))}" for m in export_months_available]
        export_month_labels = st.multiselect(
            "РњРµСЃСЏС†С‹ РґР»СЏ СЌРєСЃРїРѕСЂС‚Р°",
            options=export_month_options,
            default=export_month_options,
            key="export_months_multiselect",
        )

        export_ctypes_all = sorted(df_all["campaign_type"].dropna().astype(str).unique().tolist())
        export_ctypes_selected = st.multiselect(
            "РўРёРїС‹ Р Рљ РґР»СЏ СЌРєСЃРїРѕСЂС‚Р°",
            options=export_ctypes_all,
            default=export_ctypes_all,
            key="export_campaign_types_multiselect",
        )

        if not export_month_labels:
            st.info("Р’С‹Р±РµСЂРёС‚Рµ С…РѕС‚СЏ Р±С‹ РѕРґРёРЅ РјРµСЃСЏС†.")
        elif not export_ctypes_selected:
            st.info("Р’С‹Р±РµСЂРёС‚Рµ С…РѕС‚СЏ Р±С‹ РѕРґРёРЅ С‚РёРї Р Рљ.")
        else:
            export_month_nums = [int(lbl.split(".")[0]) for lbl in export_month_labels]
            export_mask = df_all["month_num"].isin(export_month_nums) & df_all["campaign_type"].isin(export_ctypes_selected)
            df_export = df_all.loc[export_mask].copy()

            if df_export.empty:
                st.info("РќРµС‚ РґР°РЅРЅС‹С… РґР»СЏ РІС‹Р±СЂР°РЅРЅС‹С… С„РёР»СЊС‚СЂРѕРІ СЌРєСЃРїРѕСЂС‚Р°.")
            else:
                st.caption(f"РЎС‚СЂРѕРє Рє СЌРєСЃРїРѕСЂС‚Сѓ: {len(df_export)}")

                st.markdown("#### Р­РєСЃРїРѕСЂС‚ РґР°РЅРЅС‹С…")
                st.caption("ZIP-Р°СЂС…РёРІ СЃ С‚Р°Р±Р»РёС†Р°РјРё fact/dim РІ CSV (UTF-8). Р”Р»СЏ РёРјРїРѕСЂС‚Р° РІ Power BI РёСЃРїРѕР»СЊР·СѓР№С‚Рµ Get data -> Text/CSV.")

                bi_export_df = df_export.copy()
                bi_use_vat = bool(st.session_state.get("use_vat_budget_metrics", True))
                bi_use_ak = bool(st.session_state.get("use_ak_budget_metrics", False))

                if bi_use_ak:
                    if bi_use_vat:
                        bi_export_df["budget_basis"] = bi_export_df.get("cost_with_vat_ak", 0.0)
                    else:
                        bi_export_df["budget_basis"] = (
                            pd.to_numeric(bi_export_df.get("cost", 0.0), errors="coerce").fillna(0.0)
                            + pd.to_numeric(bi_export_df.get("ak_cost_wo_vat", 0.0), errors="coerce").fillna(0.0)
                        )
                else:
                    bi_export_df["budget_basis"] = (
                        bi_export_df.get("cost_with_vat", 0.0) if bi_use_vat else bi_export_df.get("cost", 0.0)
                    )

                if "segment" not in bi_export_df.columns:
                    bi_export_df["segment"] = "ALL"
                bi_export_df["segment"] = bi_export_df["segment"].fillna("ALL").astype(str)
                if "geo" not in bi_export_df.columns:
                    bi_export_df["geo"] = ""
                bi_export_df["geo"] = bi_export_df["geo"].fillna("").astype(str)
                bi_export_df["orders"] = pd.to_numeric(
                    bi_export_df.get("conversions", 0.0), errors="coerce"
                ).fillna(0.0)

                dim_month = (
                    bi_export_df[["month_num", "month_name"]]
                    .dropna(subset=["month_num"])
                    .drop_duplicates()
                    .sort_values("month_num")
                    .rename(columns={"month_num": "month_key"})
                    .reset_index(drop=True)
                )
                if not dim_month.empty:
                    dim_month["month_key"] = dim_month["month_key"].astype(int)

                dim_campaign = (
                    bi_export_df[["campaign_type", "system", "format", "geo", "segment"]]
                    .fillna("")
                    .drop_duplicates()
                    .reset_index(drop=True)
                )
                dim_campaign.insert(0, "campaign_key", np.arange(1, len(dim_campaign) + 1, dtype=int))

                fact = bi_export_df.merge(
                    dim_campaign,
                    on=["campaign_type", "system", "format", "geo", "segment"],
                    how="left",
                ).copy()
                fact["month_key"] = pd.to_numeric(fact.get("month_num"), errors="coerce").fillna(0).astype(int)

                fact_cols = [
                    "month_key", "campaign_key",
                    "month_num", "month_name", "campaign_type", "segment", "system", "format", "geo",
                    "impressions", "reach", "frequency", "clicks", "orders", "conversions", "leads", "target_leads",
                    "cost", "ak_cost_wo_vat", "total_budget_wo_vat_ak", "cost_with_vat", "cost_with_vat_ak", "vat_amount", "budget_basis",
                    "ctr", "cpc", "cr", "cr1", "cr2", "cpm", "cpa", "cpl", "cpql", "aov", "revenue", "roas", "drr",
                    "available_capacity", "client_count", "absolute_new_clients", "returned_clients", "new_clients", "cac", "order_frequency", "sov_pct", "new_clients_share_pct", "order_share_segment_pct", "revenue_share_segment_pct",
                    "ak_rate", "ak_rate_pct",
                ]
                for c in fact_cols:
                    if c not in fact.columns:
                        fact[c] = np.nan
                fact = fact[fact_cols].copy()

                dim_segment = (
                    bi_export_df[["segment"]]
                    .fillna("ALL")
                    .drop_duplicates()
                    .sort_values("segment")
                    .reset_index(drop=True)
                )
                dim_segment.insert(0, "segment_key", np.arange(1, len(dim_segment) + 1, dtype=int))

                readme_text = (
                    "media_planner BI export package\n"
                    "Import in Power BI Desktop: Get data -> Text/CSV.\n"
                    "Files:\n"
                    "- fact_mediaplan.csv (main fact table)\n"
                    "- dim_month.csv (join by month_key)\n"
                    "- dim_campaign.csv (join by campaign_key)\n"
                    "- dim_segment.csv (segment dictionary)\n"
                    "Notes:\n"
                    f"- budget_basis uses current VAT/AK mode (use_vat={bi_use_vat}, use_ak={bi_use_ak}).\n"
                    "- orders duplicates conversions for BI naming convenience.\n"
                )
                if is_real_estate_preset:
                    readme_text += (
                        "- real_estate preset: target_leads = ??, cpql uses current budget_basis.\n"
                        "- full funnel mode also includes leads and cpl.\n"
                    )
                bi_zip_buf = io.BytesIO()
                with zipfile.ZipFile(bi_zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                    zf.writestr("fact_mediaplan.csv", _to_csv_bytes(fact))
                    zf.writestr("dim_month.csv", _to_csv_bytes(dim_month))
                    zf.writestr("dim_campaign.csv", _to_csv_bytes(dim_campaign))
                    zf.writestr("dim_segment.csv", _to_csv_bytes(dim_segment))
                    zf.writestr("README.txt", readme_text.encode("utf-8"))
                bi_zip_buf.seek(0)

                bi_ts = dt.datetime.now().strftime("%Y%m%d_%H%M")
                st.download_button(
                    "РЎРєР°С‡Р°С‚СЊ BI-РїР°РєРµС‚ (Power BI CSV ZIP)",
                    data=bi_zip_buf.getvalue(),
                    file_name=f"mediaplan_bi_export_{bi_ts}.zip",
                    mime="application/zip",
                    key="download_export_bi_zip",
                )

                # ---------- 1) Р”РµС‚Р°Р»СЊРЅС‹Р№ Excel ----------
                timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M")
                detail_buf = io.BytesIO()
                detail_engine = "xlsxwriter" if HAS_XLSXWRITER else "openpyxl"
                with pd.ExcelWriter(detail_buf, engine=detail_engine) as writer:
                    detail_cols = [
                        "month_num", "month_name", "campaign_type", "system", "format", "geo",
                        "impressions", "reach", "frequency", "clicks", "ctr", "cpc", "cost", "ak_cost_wo_vat", "total_budget_wo_vat_ak", "cost_with_vat",
                        "cost_with_vat_ak", "vat_amount", "ak_rate", "ak_rate_pct",
                        "cr", "cr1", "cr2", "conversions", "leads", "target_leads", "aov", "cpm", "cpa", "cpl", "cpql", "revenue", "roas", "drr",
                        "available_capacity", "client_count", "absolute_new_clients", "returned_clients", "new_clients", "cac", "order_frequency", "sov_pct", "new_clients_share_pct", "order_share_segment_pct", "revenue_share_segment_pct", "budget_share_segment_pct",
                    ]
                    for c in detail_cols:
                        if c not in df_export.columns:
                            df_export[c] = np.nan
                    df_detail = safe_select_columns(df_export, detail_cols).rename(
                        columns={
                            "month_num": "РњРµСЃСЏС† в„–",
                            "month_name": "РњРµСЃСЏС†",
                            "campaign_type": "РўРёРї Р Рљ",
                            "system": "РЎРёСЃС‚РµРјР°",
                            "format": "Р¤РѕСЂРјР°С‚",
                            "geo": "Р“Р•Рћ",
                            "impressions": "РџРѕРєР°Р·С‹",
                            "reach": "РћС…РІР°С‚",
                            "frequency": "Р§Р°СЃС‚РѕС‚Р°",
                            "clicks": "РљР»РёРєРё",
                            "ctr": "CTR (РґРѕР»СЏ)",
                            "cpc": "CPC, в‚Ѕ",
                            "cost": "Р‘СЋРґР¶РµС‚ Р±РµР· РќР”РЎ, в‚Ѕ",
                            "cost_with_vat": "Р‘СЋРґР¶РµС‚ СЃ РќР”РЎ, в‚Ѕ",
                            "ak_rate": "РђРљ (РґРѕР»СЏ)",
                            "ak_rate_pct": "РђРљ, %",
                            "ak_cost_wo_vat": "РђРљ Р±РµР· РќР”РЎ, в‚Ѕ",
                            "total_budget_wo_vat_ak": "РўРѕС‚Р°Р» Р±СЋРґР¶РµС‚ Р±РµР· РќР”РЎ СЃ СѓС‡РµС‚РѕРј РђРљ, в‚Ѕ",
                            "cost_with_vat_ak": "Р‘СЋРґР¶РµС‚ СЃ РќР”РЎ Рё РђРљ, в‚Ѕ",
                            "vat_amount": "РќР”РЎ, в‚Ѕ",
                            "cr": "CR (РґРѕР»СЏ)",
                            "cr1": "CR1 РІ Р›РёРґ (РґРѕР»СЏ)",
                            "cr2": "CR2 РІ Р¦Рћ (РґРѕР»СЏ)",
                            "conversions": "РљРѕРЅРІРµСЂСЃРёРё",
                            "leads": "Р›РёРґС‹",
                            "target_leads": "Р¦Рћ",
                            "aov": "AOV, в‚Ѕ",
                            "cpm": "CPM, в‚Ѕ",
                            "cpa": "CPO, в‚Ѕ",
                            "cpl": "CPL, в‚Ѕ",
                            "cpql": "CPQL, в‚Ѕ",
                            "revenue": "Р’С‹СЂСѓС‡РєР°, в‚Ѕ",
                            "roas": "ROAS (РґРѕР»СЏ)",
                            "drr": "Р”Р Р  (РґРѕР»СЏ)",
                            "available_capacity": "Р”РѕСЃС‚СѓРїРЅР°СЏ РµРјРєРѕСЃС‚СЊ",
                            "client_count": "РљРѕР»РёС‡РµСЃС‚РІРѕ РєР»РёРµРЅС‚РѕРІ",
                            "absolute_new_clients": "РљРѕР»-РІРѕ Р°Р±СЃРѕР»СЋС‚РЅРѕ РЅРѕРІС‹С… РєР»РёРµРЅС‚РѕРІ",
                            "returned_clients": "РљРѕР»-РІРѕ РІРµСЂРЅСѓРІС€РёС…СЃСЏ РєР»РёРµРЅС‚РѕРІ",
                            "new_clients": "РљРѕР»-РІРѕ РЅРѕРІС‹С… РєР»РёРµРЅС‚РѕРІ",
                            "cac": "CAC, в‚Ѕ",
                            "order_frequency": "Р§Р°СЃС‚РѕС‚Р° Р·Р°РєР°Р·Р°",
                            "sov_pct": "SOV, %",
                            "new_clients_share_pct": "Р”РѕР»СЏ РЅРѕРІС‹С… РєР»РёРµРЅС‚РѕРІ, %",
                            "order_share_segment_pct": "Р”РѕР»СЏ Р·Р°РєР°Р·РѕРІ, %",
                            "revenue_share_segment_pct": "Р”РѕР»СЏ РґРѕС…РѕРґР°, %",
                            "budget_share_segment_pct": "Р”РѕР»СЏ СЂРµРєР»Р°РјРЅРѕРіРѕ Р±СЋРґР¶РµС‚Р°, %",
                        }
                    )
                    df_detail.to_excel(writer, sheet_name="Р”РµС‚Р°Р»СЊРЅРѕ", index=False)

                    agg_month = df_export.groupby(["month_num", "month_name"], as_index=False).agg(
                        impressions=("impressions", "sum"),
                        clicks=("clicks", "sum"),
                        conversions=("conversions", "sum"),
                        leads=("leads", "sum"),
                        target_leads=("target_leads", "sum"),
                        cost=("cost", "sum"),
                        cost_with_vat=("cost_with_vat", "sum"),
                        ak_cost_wo_vat=("ak_cost_wo_vat", "sum"),
                        cost_with_vat_ak=("cost_with_vat_ak", "sum"),
                        revenue=("revenue", "sum"),
                    ).sort_values("month_num")
                    agg_month["ctr_pct"] = np.where(agg_month["impressions"] > 0, agg_month["clicks"] / agg_month["impressions"] * 100.0, 0.0)
                    agg_month["cpc"] = np.where(agg_month["clicks"] > 0, agg_month["cost"] / agg_month["clicks"], 0.0)
                    agg_month["cr_pct"] = np.where(agg_month["clicks"] > 0, agg_month["conversions"] / agg_month["clicks"] * 100.0, 0.0)
                    agg_month["cr1_pct"] = np.where(agg_month["clicks"] > 0, agg_month["leads"] / agg_month["clicks"] * 100.0, 0.0)
                    agg_month["cr2_pct"] = np.where(agg_month["leads"] > 0, agg_month["target_leads"] / agg_month["leads"] * 100.0, 0.0)
                    export_use_vat = bool(st.session_state.get("use_vat_budget_metrics", True))
                    export_use_ak = bool(st.session_state.get("use_ak_budget_metrics", False))
                    if export_use_ak:
                        budget_series = agg_month["cost_with_vat_ak"] if export_use_vat else (agg_month["cost"] + agg_month["ak_cost_wo_vat"])
                    else:
                        budget_series = agg_month["cost_with_vat"] if export_use_vat else agg_month["cost"]
                    agg_month["cpm"] = np.where(agg_month["impressions"] > 0, budget_series / (agg_month["impressions"] / 1000.0), 0.0)
                    agg_month["cpo"] = np.where(agg_month["conversions"] > 0, budget_series / agg_month["conversions"], 0.0)
                    agg_month["cpl"] = np.where(agg_month["leads"] > 0, budget_series / agg_month["leads"], 0.0)
                    agg_month["cpql"] = np.where(agg_month["target_leads"] > 0, budget_series / agg_month["target_leads"], 0.0)
                    agg_month["roas_pct"] = np.where(budget_series > 0, agg_month["revenue"] / budget_series * 100.0, 0.0)
                    agg_month["drr_pct"] = np.where(agg_month["revenue"] > 0, budget_series / agg_month["revenue"] * 100.0, 0.0)
                    agg_month = agg_month.rename(
                        columns={
                            "month_num": "РњРµСЃСЏС† в„–",
                            "month_name": "РњРµСЃСЏС†",
                            "impressions": "РџРѕРєР°Р·С‹",
                            "clicks": "РљР»РёРєРё",
                            "conversions": "РљРѕРЅРІРµСЂСЃРёРё",
                            "leads": "Р›РёРґС‹",
                            "target_leads": "Р¦Рћ",
                            "cost": "Р‘СЋРґР¶РµС‚ Р±РµР· РќР”РЎ, в‚Ѕ",
                            "cost_with_vat": "Р‘СЋРґР¶РµС‚ СЃ РќР”РЎ, в‚Ѕ",
                            "ak_cost_wo_vat": "РђРљ Р±РµР· РќР”РЎ, в‚Ѕ",
                            "cost_with_vat_ak": "Р‘СЋРґР¶РµС‚ СЃ РќР”РЎ Рё РђРљ, в‚Ѕ",
                            "revenue": "Р’С‹СЂСѓС‡РєР°, в‚Ѕ",
                            "ctr_pct": "CTR, %",
                            "cpc": "CPC, в‚Ѕ",
                            "cr_pct": "CR, %",
                            "cr1_pct": "CR1 РІ Р›РёРґ, %",
                            "cr2_pct": "CR2 РІ Р¦Рћ, %",
                            "cpm": "CPM, в‚Ѕ",
                            "cpo": "CPO, в‚Ѕ",
                            "cpl": "CPL, в‚Ѕ",
                            "cpql": "CPQL, в‚Ѕ",
                            "roas_pct": "ROAS, %",
                            "drr_pct": "Р”Р Р , %",
                        }
                    )
                    agg_month.to_excel(writer, sheet_name="РЎРІРѕРґРєР°_РїРѕ_РјРµСЃСЏС†Р°Рј", index=False)

                detail_buf.seek(0)
                st.download_button(
                    "РЎРєР°С‡Р°С‚СЊ Excel (РґРµС‚Р°Р»СЊРЅРѕ + СЃРІРѕРґРєР°)",
                    data=detail_buf.getvalue(),
                    file_name=f"mediaplan_export_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_export_detail_xlsx",
                    type="primary",
                )

                # ---------- 2) Р­РєСЃРїРѕСЂС‚ РІ С€Р°Р±Р»РѕРЅ ----------
                template_campaigns = campaigns[campaigns["campaign_type"].isin(export_ctypes_selected)].copy()
                template_month_nums = sorted(export_month_nums)
                if len(template_month_nums) > 12:
                    st.warning("Р”Р»СЏ СЌРєСЃРїРѕСЂС‚Р° РІ С€Р°Р±Р»РѕРЅ РёСЃРїРѕР»СЊР·СѓРµС‚СЃСЏ РјР°РєСЃРёРјСѓРј 12 РјРµСЃСЏС†РµРІ.")
                    template_month_nums = template_month_nums[:12]
                if len(template_campaigns) > 10:
                    st.warning("Р”Р»СЏ СЌРєСЃРїРѕСЂС‚Р° РІ С€Р°Р±Р»РѕРЅ РёСЃРїРѕР»СЊР·СѓРµС‚СЃСЏ РјР°РєСЃРёРјСѓРј 10 С‚РёРїРѕРІ Р Рљ (РѕРіСЂР°РЅРёС‡РµРЅРёРµ С€Р°Р±Р»РѕРЅР°).")
                    template_campaigns = template_campaigns.head(10)

                if not template_campaigns.empty and template_month_nums:
                    try:
                        export_preset_key = str(st.session_state.get("metric_preset_key", "ecom")).strip().lower()
                        if export_preset_key not in METRIC_PRESETS:
                            export_preset_key = "ecom"
                        allow_ecom_template = export_preset_key == "ecom"
                        allow_diy_template = export_preset_key == "diy"
                        real_estate_funnel_mode_export = str(st.session_state.get("real_estate_funnel_mode", "simple"))
                        allow_real_estate_simple_template = export_preset_key == "real_estate" and real_estate_funnel_mode_export == "simple"
                        allow_real_estate_full_template = export_preset_key == "real_estate" and real_estate_funnel_mode_export == "full"
                        resolved_tpl_ecom = resolve_template_path("ecom")
                        if resolved_tpl_ecom:
                            tpl_buf_ecom = build_excel_from_template(
                                df_all=df_export,
                                campaigns=template_campaigns,
                                selected_month_nums=template_month_nums,
                                template_kind="ecom",
                            )
                            st.download_button(
                                "РЎРєР°С‡Р°С‚СЊ Excel РїРѕ С€Р°Р±Р»РѕРЅСѓ E-com",
                                data=tpl_buf_ecom.getvalue(),
                                file_name=f"mediaplan_template_ecom_{timestamp}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_export_template_xlsx_ecom",
                                disabled=not allow_ecom_template,
                            )
                        else:
                            st.info("РЁР°Р±Р»РѕРЅ E-com РЅРµ РЅР°Р№РґРµРЅ. РћР¶РёРґР°РµРјС‹Рµ РїСѓС‚Рё: " + ", ".join(TEMPLATE_PATHS_ECOM))

                        resolved_tpl_diy = resolve_template_path("diy")
                        if resolved_tpl_diy:
                            tpl_buf_diy = build_excel_from_template(
                                df_all=df_export,
                                campaigns=template_campaigns,
                                selected_month_nums=template_month_nums,
                                template_kind="diy",
                            )
                            st.download_button(
                                "РЎРєР°С‡Р°С‚СЊ Excel РїРѕ С€Р°Р±Р»РѕРЅСѓ DIY",
                                data=tpl_buf_diy.getvalue(),
                                file_name=f"mediaplan_template_diy_{timestamp}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_export_template_xlsx_diy",
                                disabled=not allow_diy_template,
                            )
                        else:
                            st.info("РЁР°Р±Р»РѕРЅ DIY РЅРµ РЅР°Р№РґРµРЅ. Р”РѕР±Р°РІСЊС‚Рµ С„Р°Р№Р» РІ РѕРґРёРЅ РёР· РїСѓС‚РµР№: " + ", ".join(TEMPLATE_PATHS_DIY))

                        resolved_tpl_real_estate_simple = resolve_template_path("real_estate_simple")
                        if resolved_tpl_real_estate_simple:
                            tpl_buf_real_estate_simple = build_excel_from_template(
                                df_all=df_export,
                                campaigns=template_campaigns,
                                selected_month_nums=template_month_nums,
                                template_kind="real_estate_simple",
                            )
                            st.download_button(
                                "РЎРєР°С‡Р°С‚СЊ Excel РїРѕ С€Р°Р±Р»РѕРЅСѓ РќРµРґРІРёР¶РёРјРѕСЃС‚СЊ (СѓРїСЂРѕС‰РµРЅРЅР°СЏ РІРѕСЂРѕРЅРєР°)",
                                data=tpl_buf_real_estate_simple.getvalue(),
                                file_name=f"mediaplan_template_real_estate_simple_{timestamp}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_export_template_xlsx_real_estate_simple",
                                disabled=not allow_real_estate_simple_template,
                            )
                        elif export_preset_key == "real_estate" and real_estate_funnel_mode_export == "simple":
                            st.info("РЁР°Р±Р»РѕРЅ РќРµРґРІРёР¶РёРјРѕСЃС‚СЊ (СѓРїСЂРѕС‰РµРЅРЅР°СЏ РІРѕСЂРѕРЅРєР°) РЅРµ РЅР°Р№РґРµРЅ. Р”РѕР±Р°РІСЊС‚Рµ С„Р°Р№Р» РІ РѕРґРёРЅ РёР· РїСѓС‚РµР№: " + ", ".join(TEMPLATE_PATHS_REAL_ESTATE_SIMPLE))

                        resolved_tpl_real_estate_full = resolve_template_path("real_estate_full")
                        if resolved_tpl_real_estate_full:
                            tpl_buf_real_estate_full = build_excel_from_template(
                                df_all=df_export,
                                campaigns=template_campaigns,
                                selected_month_nums=template_month_nums,
                                template_kind="real_estate_full",
                            )
                            st.download_button(
                                "РЎРєР°С‡Р°С‚СЊ Excel РїРѕ С€Р°Р±Р»РѕРЅСѓ РќРµРґРІРёР¶РёРјРѕСЃС‚СЊ (РїРѕР»РЅР°СЏ РІРѕСЂРѕРЅРєР°)",
                                data=tpl_buf_real_estate_full.getvalue(),
                                file_name=f"mediaplan_template_real_estate_full_{timestamp}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_export_template_xlsx_real_estate_full",
                                disabled=not allow_real_estate_full_template,
                            )
                        elif export_preset_key == "real_estate" and real_estate_funnel_mode_export == "full":
                            st.info("РЁР°Р±Р»РѕРЅ РќРµРґРІРёР¶РёРјРѕСЃС‚СЊ (РїРѕР»РЅР°СЏ РІРѕСЂРѕРЅРєР°) РЅРµ РЅР°Р№РґРµРЅ. Р”РѕР±Р°РІСЊС‚Рµ С„Р°Р№Р» РІ РѕРґРёРЅ РёР· РїСѓС‚РµР№: " + ", ".join(TEMPLATE_PATHS_REAL_ESTATE_FULL))

                    except FileNotFoundError:
                        st.error("Р¤Р°Р№Р» С€Р°Р±Р»РѕРЅР° РґР»СЏ РІС‹Р±СЂР°РЅРЅРѕРіРѕ С‚РёРїР° РЅРµ РЅР°Р№РґРµРЅ.")
                    except Exception as e:
                        st.error(f"РћС€РёР±РєР° СЌРєСЃРїРѕСЂС‚Р° РІ С€Р°Р±Р»РѕРЅ: {e}")
    st.markdown("---")
    st.markdown("### РРјРїРѕСЂС‚ РїСЂРѕРµРєС‚Р°")
    st.caption("Р—Р°РіСЂСѓР·РёС‚Рµ JSON-С„Р°Р№Р» РїСЂРѕРµРєС‚Р°, С‡С‚РѕР±С‹ РІРѕСЃСЃС‚Р°РЅРѕРІРёС‚СЊ СЂР°СЃС‡РµС‚С‹ Рё РЅР°СЃС‚СЂРѕР№РєРё РІ СЌС‚РѕРј РїСЂРёР»РѕР¶РµРЅРёРё.")
    import_result = st.session_state.pop("_pending_project_import_result", None)
    if isinstance(import_result, dict):
        if import_result.get("ok"):
            st.success(str(import_result.get("msg", "Р”Р°РЅРЅС‹Рµ РїСЂРѕРµРєС‚Р° РёРјРїРѕСЂС‚РёСЂРѕРІР°РЅС‹.")))
        else:
            st.error(str(import_result.get("msg", "РќРµ СѓРґР°Р»РѕСЃСЊ РёРјРїРѕСЂС‚РёСЂРѕРІР°С‚СЊ РїСЂРѕРµРєС‚.")))

    uploaded_project = st.file_uploader(
        "РРјРїРѕСЂС‚ РїСЂРѕРµРєС‚Р° (JSON)",
        type=["json"],
        key="upload_project_json",
        help="РџРѕСЃР»Рµ РёРјРїРѕСЂС‚Р° СЃС‚СЂР°РЅРёС†Р° РїРµСЂРµР·Р°РіСЂСѓР·РёС‚СЃСЏ Рё РІРѕСЃСЃС‚Р°РЅРѕРІРёС‚ РЅР°Р±РѕСЂС‹ РєРѕСЌС„С„РёС†РёРµРЅС‚РѕРІ, СЃРІСЏР·РєРё Рё РЅР°СЃС‚СЂРѕР№РєРё.",
    )
    queue_project_import_from_upload(uploaded_project, "export_tab")

    render_bottom_tab_switcher("Export/Import", "export")


