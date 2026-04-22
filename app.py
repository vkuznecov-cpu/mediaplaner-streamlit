import streamlit as st
import numpy as np
import pandas as pd
import io
import json
import os
import hashlib
import base64
import zipfile
import re
from dataclasses import dataclass
from io import BytesIO
from openpyxl import load_workbook
import datetime as dt
import plotly.express as px
import streamlit.components.v1 as components

try:
    import xlsxwriter  # noqa: F401
    HAS_XLSXWRITER = True
except Exception:
    HAS_XLSXWRITER = False

try:
    from statsmodels.tsa.holtwinters import ExponentialSmoothing
    HAS_STATSMODELS = True
except Exception:
    ExponentialSmoothing = None
    HAS_STATSMODELS = False


# ---------- ВНЕШНЙ ВД (CSS) ----------

st.set_page_config(page_title="Медиапланер 12 месяцев (кампании)", layout="wide")

THEME_BORDER = "#1D2A44"
THEME_PLOT_TEXT = "#EAF0FF"
THEME_LEGEND_TEXT = "#D4DDF2"
THEME_CARD_BG = "#111A2E"
VAT_RATE = 0.22
USE_EXCEL_ROUNDDOWN = True

METRIC_PRESETS = {
    "ecom": {
        "label": "E-com",
        "description": "Базовый пресет для e-commerce: показы, CTR, CPC, CR, AOV и производные метрики.",
    },
    "diy": {
        "label": "DIY",
        "description": "Пресет для DIY: продажи (от кликов), CR в продажу, SOV и доля новых клиентов.",
    },
    "real_estate": {
        "label": "Недвижимость",
        "description": "Пресет для недвижимости: лиды, целевые обращения, CPL/CPQL без выручки и AOV.",
    },
}

REAL_ESTATE_FUNNEL_OPTIONS = {
    "simple": "Упрощённая воронка",
    "full": "Полная воронка",
}

MONTH_NAMES_FULL = {
    1: "Январь",
    2: "Февраль",
    3: "Март",
    4: "Апрель",
    5: "Май",
    6: "Июнь",
    7: "Июль",
    8: "Август",
    9: "Сентябрь",
    10: "Октябрь",
    11: "Ноябрь",
    12: "Декабрь",
}

st.markdown(
    """
    <style>
    .main {
        background-color: #0B1220;
        color: #EAF0FF;
    }
    .stApp {
        background-color: #0B1220;
        color: #EAF0FF;
    }
    [data-testid="block-container"] {
        padding-top: 1.05rem;
        padding-bottom: 0.9rem;
    }
    section[data-testid="stSidebar"] {
        background-color: #111A2E;
        border-right: 1px solid #1D2A44;
    }
    h1, h2, h3, h4 {
        font-weight: 700;
        letter-spacing: 0.02em;
        color: #EAF0FF;
    }
    p, li, label {
        color: #D4DDF2;
    }
    .stDataFrame {
        border-radius: 12px;
        border: 1px solid #1D2A44;
    }
    .stDownloadButton button {
        background: #0066E0;
        color: #FFFFFF;
        font-weight: 700;
        border-radius: 999px;
        border: 1px solid #2B7EE8;
    }
    .stDownloadButton button:hover {
        background: #0A74F2;
        border-color: #0A74F2;
    }
    .stButton > button, .stFormSubmitButton > button {
        border-radius: 10px;
        border: 1px solid #27406F;
        background: #111A2E;
        color: #EAF0FF;
        font-weight: 600;
    }
    .stButton > button:hover, .stFormSubmitButton > button:hover {
        border-color: #0066E0;
        background: #14203A;
    }
    button[kind="primary"],
    button[kind="primaryFormSubmit"],
    .stButton > button[kind="primary"],
    .stFormSubmitButton > button[kind="primary"],
    .stFormSubmitButton > button[kind="primaryFormSubmit"],
    button[data-testid="baseButton-primary"],
    button[data-testid="stBaseButton-primary"] {
        background: linear-gradient(180deg, #0A74F2 0%, #0066E0 100%) !important;
        border: 1px solid #3D8EF0 !important;
        color: #FFFFFF !important;
        box-shadow: 0 0 0 1px rgba(61, 142, 240, 0.25), 0 6px 16px rgba(0, 102, 224, 0.35);
    }
    button[kind="primary"]:hover,
    button[kind="primaryFormSubmit"]:hover,
    .stButton > button[kind="primary"]:hover,
    .stFormSubmitButton > button[kind="primary"]:hover,
    .stFormSubmitButton > button[kind="primaryFormSubmit"]:hover,
    button[data-testid="baseButton-primary"]:hover,
    button[data-testid="stBaseButton-primary"]:hover {
        background: linear-gradient(180deg, #1C82F6 0%, #0A74F2 100%) !important;
        border-color: #74AEF6 !important;
        color: #FFFFFF !important;
    }
    [data-baseweb="tag"] {
        background-color: #1A2A47 !important;
        border: 1px solid #2C4D82 !important;
        color: #9EC5FF !important;
    }
    button[data-baseweb="tab"] {
        color: #9FB0D1 !important;
    }
    button[data-baseweb="tab"][aria-selected="true"] {
        color: #0066E0 !important;
    }
    [data-baseweb="tab-highlight"] {
        background-color: #0066E0 !important;
    }
    .bottom-tab-nav {
        display: flex;
        gap: 1rem;
        align-items: end;
        border-bottom: 1px solid #1D2A44;
        padding-bottom: 6px;
        margin-top: 8px;
    }
    .bottom-tab-btn {
        background: transparent;
        border: 0;
        border-bottom: 2px solid transparent;
        color: #9FB0D1;
        font-weight: 650;
        font-size: 1rem;
        padding: 0 0 8px 0;
        cursor: pointer;
    }
    .bottom-tab-btn.is-active {
        color: #EAF0FF;
        border-bottom-color: #0066E0;
    }
    .ui-section-title {
        margin: 0.22rem 0 0.42rem 0;
        font-size: 1.35rem;
        font-weight: 700;
        letter-spacing: 0.01em;
        line-height: 1.2;
    }
    .tab-intro {
        margin: 0.2rem 0 0.7rem 0;
        padding: 10px 12px;
        border: 1px solid #1D2A44;
        border-radius: 10px;
        background: rgba(17, 26, 46, 0.55);
        animation: introIn 220ms ease both;
    }
    .tab-intro p {
        margin: 0.12rem 0;
        color: #D4DDF2;
        line-height: 1.45;
    }
    [data-testid="stExpander"] {
        margin-bottom: 0.3rem;
    }
    [data-testid="stExpander"] summary p {
        font-weight: 650;
        letter-spacing: 0.01em;
    }
    [data-testid="stExpander"] div[data-testid="stForm"] {
        border: 0 !important;
        padding: 0 !important;
        background: transparent !important;
        box-shadow: none !important;
    }
    /* MOTION START */
    .stButton > button,
    .stFormSubmitButton > button,
    .stDownloadButton > button {
        transition: transform 140ms ease, box-shadow 180ms ease, background-color 180ms ease, border-color 180ms ease;
    }
    .stButton > button:hover,
    .stFormSubmitButton > button:hover,
    .stDownloadButton > button:hover {
        transform: translateY(-1px);
        box-shadow: 0 6px 14px rgba(0, 0, 0, 0.22);
    }
    .stButton > button:active,
    .stFormSubmitButton > button:active,
    .stDownloadButton > button:active {
        transform: translateY(0);
        box-shadow: 0 2px 6px rgba(0, 0, 0, 0.16);
    }
    [data-testid="stExpander"] {
        transition: border-color 180ms ease, box-shadow 180ms ease, transform 160ms ease;
    }
    [data-testid="stExpander"]:hover {
        border-color: #2C4D82;
        box-shadow: 0 6px 16px rgba(0, 0, 0, 0.18);
        transform: translateY(-1px);
    }
    [data-testid="stDataFrame"] {
        transition: border-color 180ms ease, box-shadow 180ms ease;
    }
    [data-testid="stDataFrame"]:hover {
        border-color: #2C4D82;
        box-shadow: 0 8px 20px rgba(0, 0, 0, 0.2);
    }
    [data-baseweb="tab"] {
        transition: color 160ms ease;
    }
    [data-baseweb="tab-highlight"] {
        transition: transform 180ms ease, width 180ms ease, left 180ms ease;
    }
    .ui-section-title {
        animation: fadeInUp 260ms ease both;
    }
    @keyframes fadeInUp {
        from { opacity: 0; transform: translateY(4px); }
        to { opacity: 1; transform: translateY(0); }
    }
    @keyframes introIn {
        from { opacity: 0; transform: translateY(6px); }
        to { opacity: 1; transform: translateY(0); }
    }
    /* MOTION END */
    </style>
    """,
    unsafe_allow_html=True,
)

# ---------- МОДЕЛЬ ФУНКЦИЙ РАСЧЁТА ----------

@dataclass
class PlanInput:
    impressions: float  # Показы
    ctr: float          # CTR как доля (0.01 = 1%)
    cpc: float          # ₽
    cr: float = 0.0     # CR как доля (0.02 = 2%) / CR1 для недвижимости
    aov: float = 0.0    # ₽
    cr2: float = 0.0    # CR2 как доля (для полной воронки недвижимости)
    reach: float = 0.0  # Охват (актуально для DIY)
    preset_key: str = "ecom"
    funnel_mode: str = "simple"


def calculate_plan_month(inp: PlanInput) -> dict:
    raw_clicks = inp.impressions * inp.ctr
    clicks = np.floor(raw_clicks) if USE_EXCEL_ROUNDDOWN else raw_clicks
    # В шаблоне Excel стоимость считается от S*CTR (до округления кликов).
    cost = raw_clicks * inp.cpc
    leads = np.nan
    target_leads = np.nan
    cr1 = np.nan
    cr2 = np.nan

    if inp.preset_key == "real_estate":
        if inp.funnel_mode == "full":
            leads = np.floor(clicks * inp.cr) if USE_EXCEL_ROUNDDOWN else (clicks * inp.cr)
            target_leads = np.floor(leads * inp.cr2) if USE_EXCEL_ROUNDDOWN else (leads * inp.cr2)
            conv = target_leads
            revenue = 0.0
            cr1 = inp.cr
            cr2 = inp.cr2
            cr_total = (target_leads / clicks) if clicks > 0 else 0.0
        else:
            target_leads = np.floor(clicks * inp.cr) if USE_EXCEL_ROUNDDOWN else (clicks * inp.cr)
            conv = target_leads
            revenue = 0.0
            cr_total = inp.cr
        cpm = cost / (inp.impressions / 1000) if inp.impressions > 0 else 0
        cpa = cost / conv if conv > 0 else 0
        roas = 0.0
        drr = 0.0
    elif inp.preset_key == "diy":
        conv = np.floor(clicks * inp.cr) if USE_EXCEL_ROUNDDOWN else (clicks * inp.cr)
        target_leads = np.floor(conv * inp.cr2) if USE_EXCEL_ROUNDDOWN else (conv * inp.cr2)
        revenue = conv * inp.aov
        cpm = cost / (inp.impressions / 1000) if inp.impressions > 0 else 0
        cpa = cost / conv if conv > 0 else 0
        roas = revenue / cost if cost > 0 else 0
        drr = cost / revenue if revenue > 0 else 0
        cr_total = inp.cr
        cr2 = inp.cr2
    else:
        conv = np.floor(clicks * inp.cr) if USE_EXCEL_ROUNDDOWN else (clicks * inp.cr)
        revenue = conv * inp.aov
        cpm = cost / (inp.impressions / 1000) if inp.impressions > 0 else 0
        cpa = cost / conv if conv > 0 else 0
        roas = revenue / cost if cost > 0 else 0
        drr = cost / revenue if revenue > 0 else 0
        cr_total = inp.cr

    reach = max(0.0, float(inp.reach or 0.0))
    frequency = (inp.impressions / reach) if reach > 0 else 0.0

    return {
        "impressions": inp.impressions,
        "reach": reach,
        "frequency": frequency,
        "ctr": inp.ctr,
        "cpc": inp.cpc,
        "cr": cr_total,
        "cr1": cr1,
        "cr2": cr2,
        "aov": inp.aov,
        "clicks": clicks,
        "conversions": conv,
        "leads": leads,
        "target_leads": target_leads if not pd.isna(target_leads) else conv,
        "cost": cost,
        "revenue": revenue,
        "cpm": cpm,
        "cpa": cpa,
        "roas": roas,
        "drr": drr,
    }


def calc_month_for_all_campaigns(base_campaigns: pd.DataFrame,
                                 coeffs_month: pd.DataFrame,
                                 month_num: int,
                                 month_name: str) -> pd.DataFrame:
    rows = []
    for _, base_row in base_campaigns.iterrows():
        campaign_type = str(base_row["campaign_type"])
        k_row = coeffs_month[coeffs_month["campaign_type"] == campaign_type]
        if k_row.empty:
            k_imp = k_ctr = k_cpc = k_cr = k_aov = 1.0
        else:
            k_row = k_row.iloc[0]
            k_imp = k_row["k_imp"]
            k_ctr = k_row["k_ctr"]
            k_cpc = k_row["k_cpc"]
            k_cr = k_row["k_cr"]
            k_aov = k_row["k_aov"]

        base = PlanInput(
            impressions=base_row["impressions_avg"],
            ctr=base_row["ctr_avg_percent"] / 100.0,
            cpc=base_row["cpc_avg"],
            cr=base_row["cr_avg_percent"] / 100.0,
            aov=base_row["aov_avg"],
        )

        month_inp = PlanInput(
            impressions=base.impressions * k_imp,
            ctr=base.ctr * k_ctr,
            cpc=base.cpc * k_cpc,
            cr=base.cr * k_cr,
            aov=base.aov * k_aov,
        )

        out = calculate_plan_month(month_inp)
        out["month_num"] = month_num
        out["month_name"] = month_name
        out["campaign_type"] = campaign_type
        out["system"] = base_row["system"]
        out["format"] = base_row["format"]

        rows.append(out)

    return pd.DataFrame(rows)


def normalize_coeff_set_type(raw_type: str) -> str:
    """
    Normalize legacy/custom spelling variants of coefficient set types.
    This keeps older session_state values compatible after UI text changes.
    """
    val = str(raw_type or "").strip()
    compact = val.lower().replace(" ", "")
    if compact in {"спрос(позапросам)", "спроспозапросам", "demand"}:
        return "Спрос (по запросам)"
    if compact in {"aov(среднийчек)", "aovсреднийчек", "aov"}:
        return "AOV (средний чек)"
    if compact in {"кастомныйнабор", "custom", "customset"}:
        return "Кастомный набор"
    if compact in {"медийныехвосты", "медийныйхвост", "mediatails", "media_tail", "media tails"}:
        return "Медийные хвосты"
    return val


def get_coeff_set_id_value(coeff_set: dict) -> str:
    try:
        return str(int(coeff_set.get("id", 0) or 0))
    except Exception:
        return ""


def get_coeff_set_label(coeff_set: dict) -> str:
    set_id = get_coeff_set_id_value(coeff_set)
    set_name = str(coeff_set.get("name", "")).strip()
    if not set_id:
        return set_name
    return f"[{set_id}] {set_name}" if set_name else f"[{set_id}]"


def normalize_coeff_link_value(raw_value, coeff_sets: list[dict], allowed_ids: set[str] | None = None) -> str:
    raw_text = str(raw_value or "").strip()
    if raw_text == "":
        return ""

    coeff_sets_by_id = {
        get_coeff_set_id_value(cs): cs
        for cs in coeff_sets
        if isinstance(cs, dict) and get_coeff_set_id_value(cs)
    }

    match = re.match(r"^\[(\d+)\]", raw_text)
    parsed_id = match.group(1) if match else raw_text if raw_text.isdigit() else ""
    if parsed_id and parsed_id in coeff_sets_by_id:
        if allowed_ids is None or parsed_id in allowed_ids:
            return parsed_id

    for set_id, coeff_set in coeff_sets_by_id.items():
        if allowed_ids is not None and set_id not in allowed_ids:
            continue
        set_name = str(coeff_set.get("name", "")).strip()
        if raw_text == set_name or raw_text == get_coeff_set_label(coeff_set):
            return set_id

    return ""


def _norm_campaign_identity_value(value) -> str:
    return str(value or "").strip()


def build_campaign_identity_key(
    campaign_type: str = "",
    geo: str = "",
    segment: str = "",
) -> str:
    return "||".join(
        [
            _norm_campaign_identity_value(campaign_type),
            _norm_campaign_identity_value(geo),
            _norm_campaign_identity_value(segment).upper(),
        ]
    )


def get_campaign_identity_key_from_row(row) -> str:
    return build_campaign_identity_key(
        campaign_type=row.get("campaign_type", ""),
        geo=row.get("geo", ""),
        segment=row.get("segment", ""),
    )


def get_campaign_identity_label(row) -> str:
    campaign_type = _norm_campaign_identity_value(row.get("campaign_type", ""))
    geo = _norm_campaign_identity_value(row.get("geo", ""))
    segment = _norm_campaign_identity_value(row.get("segment", "")).upper()
    parts = [p for p in [geo, segment] if p]
    return f"{campaign_type} ({' | '.join(parts)})" if parts else campaign_type


def ui_section_title(text: str) -> None:
    st.markdown(f"<div class='ui-section-title'>{text}</div>", unsafe_allow_html=True)


def parse_float_loose(value, default: float = 0.0) -> float:
    """Parse numeric values robustly (supports spaces, commas and percent sign)."""
    try:
        if value is None:
            return float(default)
        if isinstance(value, str):
            cleaned = value.strip().replace(" ", "").replace("%", "").replace(",", ".")
            if cleaned == "":
                return float(default)
            return float(cleaned)
        if pd.isna(value):
            return float(default)
        return float(value)
    except Exception:
        return float(default)


def image_file_to_data_uri(image_path: str) -> str | None:
    if not image_path or not os.path.exists(image_path):
        return None
    ext = os.path.splitext(image_path)[1].lower()
    mime = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".webp": "image/webp",
        ".svg": "image/svg+xml",
    }.get(ext)
    if not mime:
        return None
    try:
        raw = open(image_path, "rb").read()
    except Exception:
        return None
    encoded = base64.b64encode(raw).decode("ascii")
    return f"data:{mime};base64,{encoded}"


def _df_to_payload(df: pd.DataFrame | None) -> dict:
    if df is None or not isinstance(df, pd.DataFrame):
        return {"columns": [], "rows": []}
    safe_df = df.replace({np.nan: None})
    return {
        "columns": list(safe_df.columns),
        "rows": safe_df.to_dict(orient="records"),
    }


def _df_from_payload(payload: dict | None) -> pd.DataFrame:
    if not payload or not isinstance(payload, dict):
        return pd.DataFrame()
    cols = payload.get("columns", [])
    rows = payload.get("rows", [])
    if not isinstance(cols, list) or not isinstance(rows, list):
        return pd.DataFrame()
    return pd.DataFrame(rows, columns=cols if cols else None)


def _to_csv_bytes(df: pd.DataFrame) -> bytes:
    if df is None or not isinstance(df, pd.DataFrame):
        return b""
    return df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


def build_planning_periods(start_month: int, start_year: int, horizon_months: int = 12) -> list[dict]:
    periods = []
    start_month = int(start_month)
    start_year = int(start_year)
    for slot in range(1, horizon_months + 1):
        month_offset = slot - 1
        month_num = ((start_month - 1 + month_offset) % 12) + 1
        year = start_year + ((start_month - 1 + month_offset) // 12)
        month_name_short = MONTH_NAMES_FULL[month_num]
        month_name = f"{month_name_short} {year}"
        periods.append(
            {
                "planning_slot": slot,
                "month_num": month_num,
                "month_year": year,
                "month_name": month_name,
                "month_name_short": month_name_short,
                "period_label": f"{slot}. {month_name}",
            }
        )
    return periods


def sync_multiselect_state(key: str, valid_options: list[str], default_to_all: bool = True) -> list[str]:
    current_values = [str(v) for v in st.session_state.get(key, [])]
    valid_values = [value for value in current_values if value in valid_options]
    if valid_values:
        st.session_state[key] = valid_values
    elif default_to_all:
        st.session_state[key] = valid_options.copy()
    else:
        st.session_state[key] = []
    return list(st.session_state[key])


def calc_new_clients_share_pct(new_clients: float, client_count: float) -> float:
    new_clients = float(new_clients or 0.0)
    client_count = float(client_count or 0.0)
    return (new_clients / client_count * 100.0) if client_count > 0 else 0.0


def safe_select_columns(df: pd.DataFrame, columns: list[str], fill_value=np.nan) -> pd.DataFrame:
    work = df.copy()
    for col in columns:
        if col not in work.columns:
            work[col] = fill_value
    return work[columns].copy()


def get_real_estate_funnel_mode() -> str:
    mode = str(st.session_state.get("real_estate_funnel_mode", "simple")).strip().lower()
    return mode if mode in REAL_ESTATE_FUNNEL_OPTIONS else "simple"


def get_metric_mode(preset_key: str | None = None, funnel_mode: str | None = None) -> dict:
    preset = str(preset_key or st.session_state.get("metric_preset_key", "ecom")).strip().lower()
    if preset not in METRIC_PRESETS:
        preset = "ecom"
    re_mode = str(funnel_mode or get_real_estate_funnel_mode()).strip().lower()
    if re_mode not in REAL_ESTATE_FUNNEL_OPTIONS:
        re_mode = "simple"
    return {
        "preset_key": preset,
        "is_diy": preset == "diy",
        "is_real_estate": preset == "real_estate",
        "is_real_estate_simple": preset == "real_estate" and re_mode == "simple",
        "is_real_estate_full": preset == "real_estate" and re_mode == "full",
        "real_estate_funnel_mode": re_mode,
        "needs_aov": preset in {"ecom", "diy"},
        "needs_capacity": preset == "diy",
    }


def get_campaign_required_cols(metric_mode: dict) -> list[str]:
    required = ["impressions_avg", "ctr_avg_percent", "cpc_avg", "cr_avg_percent"]
    if metric_mode.get("needs_aov"):
        required.append("aov_avg")
    if metric_mode.get("is_diy"):
        required.append("shipped_aov_avg")
        required.append("cr2_avg_percent")
    elif metric_mode.get("is_real_estate_full"):
        required.append("cr2_avg_percent")
    return required


def get_real_estate_display_metric_specs(metric_mode: dict) -> list[tuple[str, str]]:
    common = [
        ("Показы", "impressions"),
        ("Клики", "clicks"),
        ("CTR", "ctr"),
        ("CPC", "cpc"),
        ("Бюджет", "cost"),
        ("Бюджет с НДС", "cost_with_vat"),
        ("Бюджет с НДС и АК", "cost_with_vat_ak"),
    ]
    if metric_mode.get("is_real_estate_full"):
        return common + [
            ("CR1 в Лид", "cr1_pct"),
            ("Лиды", "leads"),
            ("CPL", "cpl"),
            ("CR2 в ЦО", "cr2_pct"),
            ("ЦО", "target_leads"),
            ("CPQL", "cpql"),
        ]
    return common + [
        ("CR в ЦО", "cr_pct"),
        ("ЦО", "target_leads"),
        ("CPQL", "cpql"),
    ]


def get_real_estate_table_cols(metric_mode: dict) -> list[str]:
    if metric_mode.get("is_real_estate_full"):
        return [
            "impressions", "clicks", "ctr_pct", "cpc", "cost", "cost_with_vat", "cost_with_vat_ak",
            "cr1_pct", "leads", "cpl", "cr2_pct", "target_leads", "cpql",
        ]
    return [
        "impressions", "clicks", "ctr_pct", "cpc", "cost", "cost_with_vat", "cost_with_vat_ak",
        "cr_pct", "target_leads", "cpql",
    ]


def get_diy_plan_table_cols() -> list[str]:
    return [
        "cost",
        "ak_cost_wo_vat",
        "total_budget_wo_vat_ak",
        "cost_with_vat",
        "cost_with_vat_ak",
        "vat_amount",
        "ak_rate_pct",
        "impressions",
        "clicks",
        "cpc",
        "ctr_pct",
        "cr_pct",
        "conversions",
        "cpa",
        "aov",
        "revenue",
        "drr_pct",
        "ROAS",
        "available_capacity",
        "client_count",
        "absolute_new_clients",
        "returned_clients",
        "new_clients",
        "cac",
        "order_frequency",
        "shipped_orders",
        "shipped_cps",
        "shipped_aov",
        "shipped_revenue",
        "shipped_roas",
        "shipped_drr_pct",
        "shipped_order_share_segment_pct",
        "shipped_revenue_share_segment_pct",
        "sov_pct",
        "new_clients_share_pct",
        "order_share_segment_pct",
        "revenue_share_segment_pct",
        "budget_share_segment_pct",
    ]


def get_diy_display_metric_specs() -> list[tuple[str, str]]:
    return [
        ("Бюджет", "cost"),
        ("АК без НДС", "ak_cost_wo_vat"),
        ("Тотал бюджет без НДС с учетом АК", "total_budget_wo_vat_ak"),
        ("Бюджет с НДС", "cost_with_vat"),
        ("Бюджет с НДС и АК", "cost_with_vat_ak"),
        ("НДС, ₽", "vat_amount"),
        ("АК, %", "ak_rate_pct"),
        ("Показы", "impressions"),
        ("Клики", "clicks"),
        ("CPC", "cpc"),
        ("CTR", "ctr"),
        ("CR", "cr"),
        ("Конверсии", "conversions"),
        ("CPO", "cpa"),
        ("Средний чек по оформл. доходу, с НДС", "aov"),
        ("Выручка", "revenue"),
        ("ROAS", "roas"),
        ("ДРР", "drr"),
        ("Доступная емкость", "available_capacity"),
        ("Количество клиентов", "client_count"),
        ("Кол-во абсолютно новых клиентов", "absolute_new_clients"),
        ("Кол-во вернувшихся клиентов", "returned_clients"),
        ("Кол-во новых клиентов", "new_clients"),
        ("CAC", "cac"),
        ("Частота заказа", "order_frequency"),
        ("Кол-во отгруженных заказов", "shipped_orders"),
        ("Стоимость отгр. заказа, ₽ с НДС", "shipped_cps"),
        ("Средний чек по отгр. доходу, с НДС", "shipped_aov"),
        ("Выручка (отгруженный доход), ₽ с НДС", "shipped_revenue"),
        ("ROAS отгр.", "shipped_roas"),
        ("ДРР отгр., %", "shipped_drr_pct"),
        ("SOV, %", "sov_pct"),
        ("Доля новых клиентов, %", "new_clients_share_pct"),
    ]


def compute_real_estate_rates(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "target_leads" not in out.columns and "conversions" in out.columns:
        out["target_leads"] = out["conversions"]
    if "leads" not in out.columns:
        out["leads"] = np.nan
    out["cr1_pct"] = np.where(out["clicks"] > 0, out["leads"] / out["clicks"] * 100.0, 0.0)
    out["cr2_pct"] = np.where(out["leads"] > 0, out["target_leads"] / out["leads"] * 100.0, 0.0)
    return out


def _bootstrap_reference_from_campaigns(campaigns_df: pd.DataFrame | None) -> None:
    """
    Prepare quick-reference sidebar data right after import.
    This keeps "Быстрая сверка" usable even before full recalculation blocks run.
    """
    if campaigns_df is None or not isinstance(campaigns_df, pd.DataFrame) or campaigns_df.empty:
        return

    metric_mode = get_metric_mode()
    required_cols = {"campaign_type", *get_campaign_required_cols(metric_mode)}
    if not required_cols.issubset(set(campaigns_df.columns)):
        return

    work = campaigns_df.copy()
    work["campaign_type"] = work["campaign_type"].astype(str).str.strip()
    work = work[work["campaign_type"] != ""]
    if work.empty:
        return

    use_vat = bool(st.session_state.get("use_vat_budget_metrics", True))
    base_by_campaign = {}
    total_imp = total_clicks = total_conv = total_cost = total_rev = 0.0

    for _, row in work.iterrows():
        inp = PlanInput(
            impressions=max(0.0, parse_float_loose(row.get("impressions_avg"), 0.0)),
            ctr=max(0.0, parse_float_loose(row.get("ctr_avg_percent"), 0.0) / 100.0),
            cpc=max(0.0, parse_float_loose(row.get("cpc_avg"), 0.0)),
            cr=max(0.0, parse_float_loose(row.get("cr_avg_percent"), 0.0) / 100.0),
            aov=max(0.0, parse_float_loose(row.get("aov_avg"), 0.0)),
            cr2=max(0.0, parse_float_loose(row.get("cr2_avg_percent"), 0.0) / 100.0),
            preset_key=metric_mode["preset_key"],
            funnel_mode=metric_mode["real_estate_funnel_mode"],
        )
        out = calculate_plan_month(inp)
        if metric_mode.get("is_diy"):
            shipped_orders = float(out.get("target_leads", 0.0) or 0.0)
            shipped_aov = max(0.0, parse_float_loose(row.get("shipped_aov_avg"), 0.0))
            out["shipped_orders"] = shipped_orders
            out["shipped_revenue"] = shipped_orders * shipped_aov
            out["shipped_aov"] = shipped_aov
        camp = str(row.get("campaign_type", "")).strip()
        if not camp:
            continue

        imp = float(out.get("impressions", 0.0))
        clicks = float(out.get("clicks", 0.0))
        conv = float(out.get("conversions", 0.0))
        leads = float(out.get("leads", 0.0) or 0.0)
        target_leads = float(out.get("target_leads", conv) or 0.0)
        cost = float(out.get("cost", 0.0))
        rev = float(out.get("revenue", 0.0))
        cost_with_vat = cost * (1.0 + VAT_RATE)
        budget_basis = cost_with_vat if use_vat else cost
        cpo = (budget_basis / conv) if conv > 0 else 0.0
        cpl = (budget_basis / leads) if leads > 0 else 0.0
        cpql = (budget_basis / target_leads) if target_leads > 0 else 0.0
        roas = (rev / budget_basis) if budget_basis > 0 else 0.0
        drr = (budget_basis / rev * 100.0) if rev > 0 else 0.0

        base_by_campaign[camp] = {
            "impressions": imp,
            "clicks": clicks,
            "conversions": conv,
            "leads": leads,
            "target_leads": target_leads,
            "cost": cost,
            "cost_with_vat": cost_with_vat,
            "cost_with_vat_ak": cost_with_vat,
            "revenue": rev,
            "ctr": (clicks / imp * 100.0) if imp > 0 else 0.0,
            "cpc": (cost / clicks) if clicks > 0 else 0.0,
            "cr": (conv / clicks * 100.0) if clicks > 0 else 0.0,
            "cpo": cpo,
            "cpl": cpl,
            "cpql": cpql,
            "roas": roas,
            "drr": drr,
        }

        total_imp += imp
        total_clicks += clicks
        total_conv += conv
        total_cost += cost
        total_rev += rev

    if not base_by_campaign:
        return

    total_cost_with_vat = total_cost * (1.0 + VAT_RATE)
    total_budget_basis = total_cost_with_vat if use_vat else total_cost
    st.session_state["mp_ref_base_by_campaign"] = base_by_campaign
    st.session_state["mp_ref_base"] = {
        "impressions": float(total_imp),
        "clicks": float(total_clicks),
        "conversions": float(total_conv),
        "cost": float(total_cost),
        "cost_with_vat": float(total_cost_with_vat),
        "cost_with_vat_ak": float(total_cost_with_vat),
        "revenue": float(total_rev),
        "ctr": float((total_clicks / total_imp * 100.0) if total_imp > 0 else 0.0),
        "cpc": float((total_cost / total_clicks) if total_clicks > 0 else 0.0),
        "cr": float((total_conv / total_clicks * 100.0) if total_clicks > 0 else 0.0),
        "cpo": float((total_budget_basis / total_conv) if total_conv > 0 else 0.0),
        "roas": float((total_rev / total_budget_basis) if total_budget_basis > 0 else 0.0),
        "drr": float((total_budget_basis / total_rev * 100.0) if total_rev > 0 else 0.0),
    }


def export_project_state_payload() -> dict:
    coeff_sets_src = st.session_state.get("coeff_sets", [])
    coeff_sets_payload = []
    for cs in coeff_sets_src:
        if not isinstance(cs, dict):
            continue
        coeff_sets_payload.append(
            {
                "id": int(cs.get("id", 0) or 0),
                "name": str(cs.get("name", "")),
                "type": str(cs.get("type", "")),
                "start_month": int(cs.get("start_month", 1) or 1),
                "start_year": int(cs.get("start_year", dt.date.today().year) or dt.date.today().year),
                "period_months": int(cs.get("period_months", 24) or 24),
                "queries": [str(q) for q in cs.get("queries", []) if str(q).strip()],
                "df_data": _df_to_payload(cs.get("df_data")),
                "result": _df_to_payload(cs.get("result")),
            }
        )

    payload = {
        "schema_version": 1,
        "exported_at": dt.datetime.now().isoformat(timespec="seconds"),
        "app": "media_planner",
        "state": {
            "campaigns_df": _df_to_payload(st.session_state.get("campaigns_df")),
            "coeff_sets": coeff_sets_payload,
            "coeff_sets_links_new": _df_to_payload(st.session_state.get("coeff_sets_links_new")),
            "elasticity_df": _df_to_payload(st.session_state.get("elasticity_df")),
            "ak_rules_df": _df_to_payload(st.session_state.get("ak_rules_df")),
            "use_vat_budget_metrics": bool(st.session_state.get("use_vat_budget_metrics", True)),
            "use_ak_budget_metrics": bool(st.session_state.get("use_ak_budget_metrics", False)),
            "ak_mode": str(st.session_state.get("ak_mode", "percent")),
            "ak_fixed_month_wo_vat": float(st.session_state.get("ak_fixed_month_wo_vat", 200000.0) or 0.0),
            "ak_fixed_percent": float(st.session_state.get("ak_fixed_percent", 2.0) or 0.0),
            "planning_months_multiselect": list(st.session_state.get("planning_months_multiselect", [])),
            "planning_start_month": int(st.session_state.get("planning_start_month", 1) or 1),
            "planning_start_year": int(st.session_state.get("planning_start_year", dt.date.today().year) or dt.date.today().year),
            "metric_preset_key": str(st.session_state.get("metric_preset_key", "ecom")),
            "real_estate_funnel_mode": str(st.session_state.get("real_estate_funnel_mode", "simple")),
        },
    }
    return payload


def import_project_state_payload(payload: dict) -> tuple[bool, str]:
    if not isinstance(payload, dict):
        return False, "Неверный формат файла импорта."
    if payload.get("schema_version") != 1:
        return False, "Неподдерживаемая версия файла импорта."

    state = payload.get("state")
    if not isinstance(state, dict):
        return False, "В файле нет блока state."

    st.session_state["campaigns_df"] = _df_from_payload(state.get("campaigns_df"))
    st.session_state["coeff_sets_links_new"] = _df_from_payload(state.get("coeff_sets_links_new"))
    st.session_state["elasticity_df"] = _df_from_payload(state.get("elasticity_df"))
    st.session_state["ak_rules_df"] = _df_from_payload(state.get("ak_rules_df"))

    imported_sets = []
    for idx, cs in enumerate(state.get("coeff_sets", []), start=1):
        if not isinstance(cs, dict):
            continue
        imported_sets.append(
            {
                "id": int(cs.get("id", idx) or idx),
                "name": str(cs.get("name", f"Набор {idx}")),
                "type": str(cs.get("type", "Спрос (по запросам)")),
                "start_month": int(cs.get("start_month", 1) or 1),
                "start_year": int(cs.get("start_year", dt.date.today().year) or dt.date.today().year),
                "period_months": int(cs.get("period_months", 24) or 24),
                "queries": [str(q) for q in cs.get("queries", []) if str(q).strip()],
                "df_data": _df_from_payload(cs.get("df_data")),
                "result": _df_from_payload(cs.get("result")),
            }
        )
    st.session_state["coeff_sets"] = imported_sets

    st.session_state["use_vat_budget_metrics"] = bool(state.get("use_vat_budget_metrics", True))
    st.session_state["use_ak_budget_metrics"] = bool(state.get("use_ak_budget_metrics", False))
    st.session_state["ak_mode"] = str(state.get("ak_mode", "percent"))
    st.session_state["ak_fixed_month_wo_vat"] = float(state.get("ak_fixed_month_wo_vat", 200000.0) or 0.0)
    st.session_state["ak_fixed_percent"] = float(state.get("ak_fixed_percent", 2.0) or 0.0)
    st.session_state["planning_months_multiselect"] = list(state.get("planning_months_multiselect", []))
    st.session_state["planning_start_month"] = int(state.get("planning_start_month", 1) or 1)
    st.session_state["planning_start_year"] = int(state.get("planning_start_year", dt.date.today().year) or dt.date.today().year)
    preset_key = str(state.get("metric_preset_key", "ecom"))
    st.session_state["metric_preset_key"] = preset_key if preset_key in METRIC_PRESETS else "ecom"
    st.session_state["real_estate_funnel_mode"] = str(state.get("real_estate_funnel_mode", "simple"))
    _bootstrap_reference_from_campaigns(st.session_state.get("campaigns_df"))

    return True, "Данные проекта импортированы."


if "_pending_project_import_payload" in st.session_state:
    _pending_payload = st.session_state.pop("_pending_project_import_payload")
    _ok, _msg = import_project_state_payload(_pending_payload)
    st.session_state["_pending_project_import_result"] = {"ok": bool(_ok), "msg": str(_msg)}


def queue_project_import_from_upload(uploaded_file, source_key: str) -> None:
    """
    Queue project import from a Streamlit UploadedFile only once per file content.
    Prevents rerun loops when uploader keeps the selected file between reruns.
    """
    sig_key = f"_project_upload_sig_{source_key}"
    if uploaded_file is None:
        st.session_state.pop(sig_key, None)
        return

    raw_bytes = uploaded_file.getvalue()
    payload_sig = hashlib.sha1(raw_bytes).hexdigest()
    if st.session_state.get(sig_key) == payload_sig:
        return

    st.session_state[sig_key] = payload_sig
    try:
        imported_payload = json.loads(raw_bytes.decode("utf-8"))
        st.session_state["_pending_project_import_payload"] = imported_payload
        st.rerun()
    except Exception as e:
        st.error(f"Не удалось прочитать файл проекта: {e}")


def make_ak_rules_signature(ak_rules_df: pd.DataFrame | None) -> tuple:
    if ak_rules_df is None or not isinstance(ak_rules_df, pd.DataFrame) or ak_rules_df.empty:
        return tuple()
    work = ak_rules_df.copy()
    if "min_budget_wo_vat" not in work.columns:
        work["min_budget_wo_vat"] = 0.0
    if "ak_percent" not in work.columns:
        work["ak_percent"] = 0.0
    work["min_budget_wo_vat"] = work["min_budget_wo_vat"].map(lambda v: round(parse_float_loose(v, 0.0), 4))
    work["ak_percent"] = work["ak_percent"].map(lambda v: round(parse_float_loose(v, 0.0), 4))
    work = work[["min_budget_wo_vat", "ak_percent"]].sort_values("min_budget_wo_vat", kind="stable")
    return tuple((float(r["min_budget_wo_vat"]), float(r["ak_percent"])) for _, r in work.iterrows())


def make_vat_ak_settings_signature(
    use_vat_budget_metrics: bool,
    use_ak_budget_metrics: bool,
    ak_mode: str,
    ak_fixed_month_wo_vat: float,
    ak_fixed_percent: float,
    ak_rules_df: pd.DataFrame | None,
) -> tuple:
    return (
        bool(use_vat_budget_metrics),
        bool(use_ak_budget_metrics),
        str(ak_mode),
        round(float(ak_fixed_month_wo_vat), 4) if str(ak_mode) == "fixed" else None,
        round(float(ak_fixed_percent), 4) if str(ak_mode) == "fixed_percent" else None,
        make_ak_rules_signature(ak_rules_df) if str(ak_mode) == "percent" else tuple(),
    )


def normalize_ak_rules_df(ak_rules_df: pd.DataFrame | None) -> pd.DataFrame:
    if ak_rules_df is None or not isinstance(ak_rules_df, pd.DataFrame):
        return pd.DataFrame([{"min_budget_wo_vat": 0.0, "ak_percent": 0.0}])
    work = ak_rules_df.copy()
    for col in ["min_budget_wo_vat", "ak_percent"]:
        if col not in work.columns:
            work[col] = 0.0
    work["min_budget_wo_vat"] = work["min_budget_wo_vat"].map(lambda v: parse_float_loose(v, 0.0))
    work["ak_percent"] = work["ak_percent"].map(lambda v: parse_float_loose(v, 0.0))
    work = work[["min_budget_wo_vat", "ak_percent"]].sort_values("min_budget_wo_vat", kind="stable").reset_index(drop=True)
    if work.empty:
        work = pd.DataFrame([{"min_budget_wo_vat": 0.0, "ak_percent": 0.0}])
    return work


def get_vat_ak_ui_config(metric_mode: dict, is_real_estate_preset: bool) -> dict:
    ak_mode_options = [
        ("percent", "Процент по шкале"),
        ("fixed", "Фиксированная сумма в месяц"),
        ("fixed_percent", "Фиксированный процент"),
    ]
    warning_text = (
        "<b>Учитывать НДС 22% в расчетах</b>: настройка влияет на CPC, CPL и CPQL."
        if is_real_estate_preset and metric_mode["is_real_estate_full"]
        else "<b>Учитывать НДС 22% в расчетах</b>: настройка влияет на CPC и CPQL."
        if is_real_estate_preset
        else "<b>Учитывать НДС 22% в расчетах</b>: настройка влияет на CPC, CPO, ROAS и ДРР."
    )
    mode_label_by_value = {value: label for value, label in ak_mode_options}
    mode_value_by_label = {label: value for value, label in ak_mode_options}
    return {
        "warning_text": warning_text,
        "ak_mode_labels": [label for _, label in ak_mode_options],
        "mode_label_by_value": mode_label_by_value,
        "mode_value_by_label": mode_value_by_label,
    }


def get_vat_ak_apply_state(
    use_vat_budget_metrics: bool,
    use_ak_budget_metrics: bool,
    ak_mode: str,
    ak_fixed_month_wo_vat: float,
    ak_fixed_percent: float,
    ak_rules_df: pd.DataFrame | None,
    last_applied_sig: tuple | None,
) -> dict:
    current_sig = make_vat_ak_settings_signature(
        use_vat_budget_metrics=use_vat_budget_metrics,
        use_ak_budget_metrics=use_ak_budget_metrics,
        ak_mode=ak_mode,
        ak_fixed_month_wo_vat=ak_fixed_month_wo_vat,
        ak_fixed_percent=ak_fixed_percent,
        ak_rules_df=ak_rules_df,
    )
    return {
        "current_sig": current_sig,
        "last_applied_sig": last_applied_sig,
        "dirty": current_sig != last_applied_sig,
    }


def resolve_ak_rate(total_cost_wo_vat: float, ak_rules_df: pd.DataFrame) -> float:
    if ak_rules_df is None or ak_rules_df.empty:
        return 0.0
    rules = ak_rules_df.copy()
    rules["min_budget_wo_vat"] = rules["min_budget_wo_vat"].map(lambda v: parse_float_loose(v, 0.0))
    rules["ak_percent"] = rules["ak_percent"].map(lambda v: parse_float_loose(v, 0.0))
    rules = rules.sort_values("min_budget_wo_vat")
    matched = rules[rules["min_budget_wo_vat"] <= float(total_cost_wo_vat)]
    if matched.empty:
        return 0.0
    return max(0.0, float(matched.iloc[-1]["ak_percent"]) / 100.0)


def apply_budget_basis_metrics(
    df: pd.DataFrame,
    use_vat: bool,
    use_ak: bool = False,
    ak_mode: str = "percent",
    ak_rate_by_month: dict | None = None,
    default_ak_rate: float = 0.0,
    ak_fixed_by_month: dict | None = None,
    default_ak_fixed_wo_vat: float = 0.0,
    default_ak_fixed_rate: float = 0.0,
    vat_rate: float = VAT_RATE,
) -> pd.DataFrame:
    out = df.copy()
    out["cost_with_vat"] = out["cost"].astype(float) * (1.0 + vat_rate)
    out["ak_rate"] = 0.0
    out["ak_fixed_month_wo_vat"] = 0.0
    out["ak_cost_wo_vat"] = 0.0

    if use_ak:
        if ak_mode == "fixed":
            if "month_num" in out.columns:
                if ak_fixed_by_month:
                    out["ak_fixed_month_wo_vat"] = out["month_num"].map(
                        lambda m: float(ak_fixed_by_month.get(int(m), default_ak_fixed_wo_vat))
                    )
                else:
                    out["ak_fixed_month_wo_vat"] = float(default_ak_fixed_wo_vat)
                month_sum = out.groupby("month_num")["cost"].transform("sum")
                share = np.where(month_sum > 0, out["cost"] / month_sum, 0.0)
                out["ak_cost_wo_vat"] = out["ak_fixed_month_wo_vat"] * share
            else:
                total_cost = float(out["cost"].sum())
                fixed_total = float(default_ak_fixed_wo_vat)
                out["ak_fixed_month_wo_vat"] = fixed_total
                share = np.where(total_cost > 0, out["cost"] / total_cost, 0.0)
                out["ak_cost_wo_vat"] = fixed_total * share
            out["ak_rate"] = np.where(out["cost"] > 0, out["ak_cost_wo_vat"] / out["cost"], 0.0)
        elif ak_mode == "fixed_percent":
            out["ak_rate"] = float(default_ak_fixed_rate)
            out["ak_rate"] = pd.to_numeric(out["ak_rate"], errors="coerce").fillna(float(default_ak_fixed_rate))
            out["ak_cost_wo_vat"] = out["cost"] * out["ak_rate"]
        else:
            if "month_num" in out.columns and ak_rate_by_month:
                out["ak_rate"] = out["month_num"].map(
                    lambda m: float(ak_rate_by_month.get(int(m), default_ak_rate))
                )
            else:
                out["ak_rate"] = float(default_ak_rate)
            out["ak_rate"] = pd.to_numeric(out["ak_rate"], errors="coerce").fillna(float(default_ak_rate))
            out["ak_cost_wo_vat"] = out["cost"] * out["ak_rate"]

    out["ak_cost_with_vat"] = out["ak_cost_wo_vat"] * (1.0 + vat_rate)
    out["total_budget_wo_vat_ak"] = out["cost"] + out["ak_cost_wo_vat"]
    out["cost_with_vat_ak"] = (out["cost"] + out["ak_cost_wo_vat"]) * (1.0 + vat_rate)
    out["vat_amount"] = out["cost_with_vat_ak"] - out["total_budget_wo_vat_ak"]
    if use_ak:
        budget_basis = out["cost_with_vat_ak"] if use_vat else (out["cost"] + out["ak_cost_wo_vat"])
    else:
        budget_basis = out["cost_with_vat"] if use_vat else out["cost"]
    out["cpm"] = np.where(
        out["impressions"] > 0,
        budget_basis / (out["impressions"] / 1000.0),
        0.0,
    )
    out["cpa"] = np.where(
        out["conversions"] > 0,
        budget_basis / out["conversions"],
        0.0,
    )
    if "leads" in out.columns:
        out["cpl"] = np.where(out["leads"] > 0, budget_basis / out["leads"], 0.0)
    if "target_leads" in out.columns:
        out["cpql"] = np.where(out["target_leads"] > 0, budget_basis / out["target_leads"], 0.0)
    out["roas"] = np.where(
        budget_basis > 0,
        out["revenue"] / budget_basis,
        0.0,
    )
    out["drr"] = np.where(
        out["revenue"] > 0,
        budget_basis / out["revenue"],
        0.0,
    )
    out["budget_basis"] = budget_basis
    return out


def render_bottom_tab_switcher(current_tab: str, key_suffix: str) -> None:
    st.markdown("---")
    nav_id = f"bottom-tab-nav-{key_suffix}"
    st.markdown(
        f"""
        <div id="{nav_id}" class="bottom-tab-nav">
            <button type="button" class="bottom-tab-btn" data-tab-idx="0">Коэффициенты</button>
            <button type="button" class="bottom-tab-btn" data-tab-idx="1">Настройки расчета</button>
            <button type="button" class="bottom-tab-btn" data-tab-idx="2">Медиаплан</button>
            <button type="button" class="bottom-tab-btn" data-tab-idx="3">Диаграммы</button>
            <button type="button" class="bottom-tab-btn" data-tab-idx="4">Export/Import</button>
        </div>
        """,
        unsafe_allow_html=True,
    )
    components.html(
        f"""
        <script>
        (function() {{
            const tablists = Array.from(window.parent.document.querySelectorAll('[role="tablist"]'));
            if (!tablists.length) return;
            const topmostTablist = tablists
                .map(tl => ({{ tl, top: tl.getBoundingClientRect().top }}))
                .sort((a, b) => a.top - b.top)[0]?.tl;
            const nav = window.parent.document.getElementById("{nav_id}");
            if (!topmostTablist || !nav) return;

            const topTabs = topmostTablist.querySelectorAll('button[role="tab"]');
            const navBtns = nav.querySelectorAll('.bottom-tab-btn');
            if (!topTabs || topTabs.length < 5 || !navBtns.length) return;

            const syncActive = () => {{
                let activeIdx = 0;
                for (let i = 0; i < topTabs.length; i++) {{
                    if (topTabs[i].getAttribute('aria-selected') === 'true') {{
                        activeIdx = i;
                        break;
                    }}
                }}
                navBtns.forEach((btn, i) => {{
                    if (i === activeIdx) btn.classList.add('is-active');
                    else btn.classList.remove('is-active');
                }});
            }};

            navBtns.forEach((btn) => {{
                btn.onclick = () => {{
                    const idx = Number(btn.getAttribute('data-tab-idx') || 0);
                    if (topTabs[idx]) {{
                        topTabs[idx].click();
                        setTimeout(syncActive, 30);
                    }}
                }};
            }});

            syncActive();
            setTimeout(syncActive, 40);
            const obs = new MutationObserver(() => syncActive());
            obs.observe(topmostTablist, {{
                subtree: true,
                attributes: true,
                attributeFilter: ['aria-selected', 'class']
            }});
        }})();
        </script>
        """,
        height=0,
        width=0,
    )


DISPLAY_COL_RENAME = {
    "segment": "Сегмент",
    "campaign_type": "Название кампании",
    "system": "Рекламная система",
    "format": "Формат/таргетинги",
    "geo": "ГЕО",
    "month_name": "Месяц",
    "impressions": "Показы",
    "reach": "Охват",
    "frequency": "Частота",
    "ctr_pct": "CTR",
    "ctr": "CTR",
    "cpc": "CPC",
    "cr_pct": "CR",
    "cr": "CR",
    "cr1_pct": "CR1 в Лид",
    "cr2_pct": "CR2 в ЦО",
    "aov": "AOV",
    "clicks": "Клики",
    "conversions": "Конверсии",
    "leads": "Лиды",
    "target_leads": "ЦО",
    "cost": "Бюджет",
    "cost_with_vat": "Бюджет с НДС",
    "cost_with_vat_ak": "Бюджет с НДС и АК",
    "ak_cost_wo_vat": "АК без НДС",
    "total_budget_wo_vat_ak": "Тотал бюджет без НДС с учетом АК",
    "vat_amount": "НДС, ₽",
    "ak_rate_pct": "АК, %",
    "revenue": "Доход",
    "cpm": "CPM",
    "cpa": "CPO",
    "cpl": "CPL",
    "cpql": "CPQL",
    "drr_pct": "ДРР",
    "drr": "ДРР",
    "ROAS": "ROAS",
    "sov_pct": "SOV, %",
    "available_capacity": "Доступная емкость",
    "client_count": "Количество клиентов",
    "absolute_new_clients": "Кол-во абсолютно новых клиентов",
    "returned_clients": "Кол-во вернувшихся клиентов",
    "new_clients": "Кол-во новых клиентов",
    "cac": "CAC",
    "order_frequency": "Частота заказа",
    "new_clients_share_pct": "Доля новых клиентов, %",
    "budget_share_segment_pct": "Доля рекламного бюджета, %",
    "order_share_segment_pct": "Доля заказов, %",
    "revenue_share_segment_pct": "Доля дохода, %",
    "shipped_orders": "Кол-во отгруженных заказов",
    "shipped_cps": "Стоимость отгр. заказа, ₽ с НДС",
    "shipped_order_share_segment_pct": "Доля отгр. заказов, %",
    "shipped_revenue": "Выручка (отгруженный доход), ₽ с НДС",
    "shipped_roas": "ROAS отгр.",
    "shipped_drr_pct": "ДРР отгр., %",
    "shipped_aov": "Средний чек по отгр. доходу, с НДС",
    "shipped_revenue_share_segment_pct": "Доля выручки, %",
}

METRIC_HELP = {
    "impressions": "Показы: количество рекламных показов.",
    "reach": "Охват: число уникальных пользователей, увидевших рекламу.",
    "frequency": "Частота = Показы / Охват.",
    "clicks": "Клики: Показы × CTR.",
    "ctr": "CTR, % = Клики / Показы × 100%.",
    "cpc": "CPC = Бюджет / Клики.",
    "cr": "CR, % = Конверсии / Клики × 100%.",
    "aov": "AOV = Доход / Конверсии.",
    "conversions": "Конверсии = Клики × CR.",
    "cost": "Бюджет = Клики × CPC.",
    "cost_with_vat": "Бюджет с НДС = Бюджет × 1.22 (при включенном учете НДС).",
    "cost_with_vat_ak": "Бюджет с НДС и АК = (Бюджет + АК без НДС) × 1.22.",
    "total_budget_wo_vat_ak": "Тотал бюджет без НДС с учетом АК = Бюджет + АК без НДС.",
    "vat_amount": "НДС, ₽ = Тотал бюджет с НДС и АК - Тотал бюджет без НДС и АК.",
    "revenue": "Доход = Конверсии × AOV. Если средний чек вводится с НДС, доход также считается с НДС без дополнительного накидывания.",
    "client_count": "Количество клиентов: базовое значение по типу РК, к которому применяется месячный коэффициент.",
    "absolute_new_clients": "Абсолютно новые клиенты: базовое значение по типу РК, к которому применяется месячный коэффициент.",
    "returned_clients": "Вернувшиеся клиенты: базовое значение по типу РК, к которому применяется месячный коэффициент.",
    "new_clients": "Новые клиенты = Абсолютно новые клиенты + Вернувшиеся клиенты.",
    "cac": "CAC = Бюджет с НДС и АК / Кол-во новых клиентов.",
    "order_frequency": "Частота заказа: базовое значение по типу РК, к которому применяется месячный коэффициент.",
    "shipped_orders": "Отгруженные заказы = Оформленные заказы × CR2.",
    "shipped_cps": "Стоимость отгр. заказа, ₽ с НДС = Бюджет с НДС / Кол-во отгруженных заказов.",
    "shipped_revenue": "Отгруженный доход = Кол-во отгруженных заказов × Средний чек. Если средний чек вводится с НДС, отгруженный доход также считается с НДС без дополнительного накидывания.",
    "shipped_roas": "ROAS отгр. = Отгруженный доход / Бюджетная база.",
    "shipped_drr_pct": "ДРР отгр., % = Бюджетная база / Отгруженный доход × 100%.",
    "shipped_aov": "Средний чек по отгр. доходу, с НДС = Отгруженный доход / Кол-во отгруженных заказов.",
    "cpm": "CPM = Бюджетная база / (Показы / 1000). База зависит от режима НДС.",
    "cpo": "CPO = Бюджетная база / Конверсии. База зависит от режима НДС.",
    "roas": "ROAS = Доход / Бюджетная база. База зависит от режима НДС.",
    "drr": "ДРР, % = Бюджетная база / Доход × 100%. База зависит от режима НДС.",
    "k_imp": "k_imp: коэффициент показов. Применяется к Показы.",
    "k_reach": "k_reach: коэффициент охвата. Применяется к Охват.",
    "k_ctr": "k_ctr: коэффициент CTR. Применяется к CTR.",
    "k_cpc": "k_cpc: коэффициент CPC. Применяется к CPC.",
    "k_cr": "k_cr: коэффициент CR. Применяется к CR.",
    "k_aov": "k_aov: коэффициент AOV. Применяется к AOV.",
    "k_client_count": "k_client_count: коэффициент количества клиентов. Применяется к метрике Количество клиентов.",
    "k_absolute_new_clients": "k_absolute_new_clients: коэффициент для абсолютно новых клиентов.",
    "k_returned_clients": "k_returned_clients: коэффициент для вернувшихся клиентов.",
    "k_order_frequency": "k_order_frequency: коэффициент частоты заказа.",
    "cpc_div": "Делитель влияния на CPC: k_cpc = 1 + (k_demand - 1) / div.",
    "ctr_div": "Делитель влияния на CTR: k_ctr = 1 - (k_demand - 1) / div.",
    "cr_div": "Делитель влияния на CR: k_cr = 1 - (k_demand - 1) / div.",
}


def mhelp(key: str, fallback: str = "") -> str:
    return METRIC_HELP.get(key, fallback)


def reorder_rows_with_segment_subtotals(
    df_show: pd.DataFrame,
    campaign_col: str,
    segment_col: str,
) -> pd.DataFrame:
    """Order rows as: campaigns by segment -> segment subtotal -> ... -> grand total."""
    if df_show is None or df_show.empty:
        return df_show
    if campaign_col not in df_show.columns or segment_col not in df_show.columns:
        return df_show

    work = df_show.copy()
    campaign_vals = work[campaign_col].astype(str)
    is_total_any = campaign_vals.str.startswith("Итого")
    is_total_grand = campaign_vals.eq("Итого")
    is_total_segment = is_total_any & ~is_total_grand
    is_campaign = ~is_total_any

    seg_order = ["B2C", "B2B"]
    segments_present = work.loc[is_campaign, segment_col].astype(str).unique().tolist()
    for seg in segments_present:
        if seg not in seg_order:
            seg_order.append(seg)

    ordered_parts = []
    used_idx: set[int] = set()
    for seg in seg_order:
        seg_campaigns = work[is_campaign & (work[segment_col].astype(str) == seg)]
        if not seg_campaigns.empty:
            ordered_parts.append(seg_campaigns)
            used_idx.update(seg_campaigns.index.tolist())
            seg_total = work[is_total_segment & (campaign_vals == f"Итого {seg}")]
            if not seg_total.empty:
                ordered_parts.append(seg_total)
                used_idx.update(seg_total.index.tolist())

    # Fallback for any rows that didn't get included.
    if ordered_parts:
        ordered = pd.concat(ordered_parts, ignore_index=True)
        campaign_rows_left = work[is_campaign & ~work.index.isin(list(used_idx))]
        seg_totals_left = work[is_total_segment & ~work.index.isin(list(used_idx))]
        if not campaign_rows_left.empty:
            ordered = pd.concat([ordered, campaign_rows_left], ignore_index=True)
            used_idx.update(campaign_rows_left.index.tolist())
        if not seg_totals_left.empty:
            ordered = pd.concat([ordered, seg_totals_left], ignore_index=True)
            used_idx.update(seg_totals_left.index.tolist())
    else:
        ordered = work[is_campaign].copy()
        used_idx.update(ordered.index.tolist())

    grand_total = work[is_total_grand & ~work.index.isin(list(used_idx))]
    if not grand_total.empty:
        ordered = pd.concat([ordered, grand_total], ignore_index=True)
    return ordered


def add_segment_budget_share(
    df: pd.DataFrame | None,
    segment_col: str = "segment",
    budget_col: str = "cost_with_vat_ak",
    out_col: str = "budget_share_segment_pct",
) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df
    out = df.copy()
    if segment_col not in out.columns or budget_col not in out.columns:
        out[out_col] = 0.0
        return out
    seg_total = out.groupby(segment_col)[budget_col].transform("sum")
    out[out_col] = np.where(seg_total > 0, out[budget_col] / seg_total * 100.0, 0.0)
    return out


def add_segment_value_share(
    df: pd.DataFrame | None,
    segment_col: str = "segment",
    value_col: str = "conversions",
    out_col: str = "order_share_segment_pct",
) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df
    out = df.copy()
    if segment_col not in out.columns or value_col not in out.columns:
        out[out_col] = 0.0
        return out
    values = pd.to_numeric(out[value_col], errors="coerce").fillna(0.0)
    seg_total = values.groupby(out[segment_col]).transform("sum")
    out[out_col] = np.where(seg_total > 0, values / seg_total * 100.0, 0.0)
    return out


def forecast_ets_like(
    series: list[float],
    periods: int = 12,
    season_length: int = 12,
    alpha: float = 0.35,
    beta: float = 0.10,
    gamma: float = 0.25,
) -> list[float]:
    """
    ETS forecast with statsmodels when available; otherwise fallback to a
    lightweight additive Holt-Winters implementation.
    """
    vals = [float(x) for x in series if pd.notna(x)]
    n = len(vals)
    if n == 0:
        return [0.0] * periods

    # Primary path: statsmodels ETS (closest to Excel FORECAST.ETS behavior).
    if HAS_STATSMODELS and n >= season_length * 2:
        try:
            # Keep smoothing params inside valid open interval for stable fitting.
            a = min(max(float(alpha), 1e-4), 0.9999)
            b = min(max(float(beta), 1e-4), 0.9999)
            g = min(max(float(gamma), 1e-4), 0.9999)

            model = ExponentialSmoothing(
                vals,
                trend="add",
                seasonal="add",
                seasonal_periods=season_length,
                initialization_method="estimated",
            )
            # Use user-selected alpha/beta/gamma (instead of auto-optimized fit).
            fitted = model.fit(
                optimized=False,
                smoothing_level=a,
                smoothing_trend=b,
                smoothing_seasonal=g,
            )
            fc = fitted.forecast(periods)
            return [max(0.0, float(v)) for v in fc]
        except Exception:
            # fallback below
            pass

    if n < season_length * 2:
        # For short history, repeat last available seasonal pattern.
        tail = vals[-season_length:] if n >= season_length else vals
        rep = []
        while len(rep) < periods:
            rep.extend(tail)
        return [max(0.0, float(v)) for v in rep[:periods]]

    n_seasons = n // season_length
    season_avgs = []
    for j in range(n_seasons):
        chunk = vals[j * season_length:(j + 1) * season_length]
        season_avgs.append(float(np.mean(chunk)))

    seasonals = [0.0] * season_length
    for i in range(season_length):
        acc = 0.0
        cnt = 0
        for j in range(n_seasons):
            idx = j * season_length + i
            if idx < n:
                acc += vals[idx] - season_avgs[j]
                cnt += 1
        seasonals[i] = acc / max(cnt, 1)

    level = vals[0] - seasonals[0]
    trend = (
        ((vals[season_length] - seasonals[0]) - (vals[0] - seasonals[0])) / season_length
        if n > season_length else 0.0
    )

    for t in range(n):
        s = seasonals[t % season_length]
        prev_level = level
        level = alpha * (vals[t] - s) + (1 - alpha) * (level + trend)
        trend = beta * (level - prev_level) + (1 - beta) * trend
        seasonals[t % season_length] = gamma * (vals[t] - level) + (1 - gamma) * s

    fc = []
    for m in range(1, periods + 1):
        idx = (n + m - 1) % season_length
        yhat = level + m * trend + seasonals[idx]
        fc.append(max(0.0, float(yhat)))
    return fc


# ---------- ЭКСПОРТ В ШАБЛОН EXCEL ----------

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATHS_ECOM = [
    os.path.join(BASE_DIR, "templates", "Шаблоны МП.xlsx"),
    os.path.join(BASE_DIR, "Шаблоны МП.xlsx"),
    os.path.join(BASE_DIR, "Shablony-MP.xlsx"),
]
TEMPLATE_PATHS_DIY = [
    os.path.join(BASE_DIR, "templates", "Шаблоны МП DIY_актуальный.xlsx"),
    os.path.join(BASE_DIR, "Шаблоны МП DIY_актуальный.xlsx"),
    os.path.join(BASE_DIR, "templates", "Шаблоны МП DIY.xlsx"),
    os.path.join(BASE_DIR, "Шаблоны МП DIY.xlsx"),
    os.path.join(BASE_DIR, "templates", "Shablony-MP-DIY.xlsx"),
]
TEMPLATE_PATHS_REAL_ESTATE_SIMPLE = [
    os.path.join(BASE_DIR, "templates", "Шаблоны МП Недвижка упр. воронка.xlsx"),
    os.path.join(BASE_DIR, "Шаблоны МП Недвижка упр. воронка.xlsx"),
]
TEMPLATE_PATHS_REAL_ESTATE_FULL = [
    os.path.join(BASE_DIR, "templates", "Шаблоны МП Недвижка полн. воронка.xlsx"),
    os.path.join(BASE_DIR, "Шаблоны МП Недвижка полн. воронка.xlsx"),
]
TEMPLATE_PATHS_BY_KIND = {
    "ecom": TEMPLATE_PATHS_ECOM,
    "diy": TEMPLATE_PATHS_DIY,
    "real_estate_simple": TEMPLATE_PATHS_REAL_ESTATE_SIMPLE,
    "real_estate_full": TEMPLATE_PATHS_REAL_ESTATE_FULL,
}


def build_excel_from_template(df_all: pd.DataFrame,
                              campaigns: pd.DataFrame,
                              selected_periods: list[dict],
                              template_kind: str = "ecom",
                              compact_empty_rows: bool = True) -> BytesIO:

    template_paths = TEMPLATE_PATHS_BY_KIND.get(template_kind, TEMPLATE_PATHS_ECOM)
    existing_templates = [p for p in template_paths if os.path.exists(p)]
    template_path = (
        max(existing_templates, key=os.path.getmtime)
        if existing_templates
        else template_paths[0]
    )

    wb = load_workbook(template_path)

    periods_candidates = [s for s in wb.sheetnames if str(s).endswith("_Periods")]
    total_candidates = [s for s in wb.sheetnames if str(s).endswith("_Total")]

    if periods_candidates:
        ws = wb[periods_candidates[0]]
    elif len(wb.worksheets) >= 2:
        ws = wb.worksheets[1]
    else:
        ws = wb.worksheets[0]

    if total_candidates:
        ws_total = wb[total_candidates[0]]
    else:
        ws_total = wb.worksheets[0]

    DEFAULT_START_ROW_JAN = 3
    DEFAULT_BLOCK_STEP = 13
    DEFAULT_ROWS_PER_MONTH = 10
    DEFAULT_START_ROW_TOTAL = 3

    COL_SYSTEM = "B"
    COL_FORMAT = "C"
    COL_TARGETING = "D"
    is_real_estate_template = template_kind in {"real_estate_simple", "real_estate_full"}
    is_real_estate_full_template = template_kind == "real_estate_full"
    is_diy_template = template_kind == "diy"

    if is_diy_template:
        COL_PERIOD = "G"
        COL_MODEL = "H"
        COL_CPC = "I"
        COL_GEO = "E"
        COL_DEMAND_COEFF = None
        COL_OTHER = "M"
        COL_TOTAL_GEO = "E"
        COL_AK = "O"
        COL_IMPRESSIONS = "V"
        COL_CTR = "AH"
        COL_CR = "AL"
        COL_CR2 = "AW"
        COL_AOV = "AS"
        COL_SHIPPED_AOV = "BD"
        COL_AVAILABLE_CAPACITY = "AE"
        COL_FREQUENCY = None
        COL_CLIENT_COUNT = "Y"
        COL_ABSOLUTE_NEW_CLIENTS = "Z"
        COL_RETURNED_CLIENTS = "AA"
        COL_ORDER_FREQUENCY = "AC"
        COL_NEW_CLIENTS_SHARE = None
        COL_REACH = None
    else:
        COL_PERIOD = "F"
        COL_MODEL = "G"
        COL_CPC = "H"
        COL_GEO = "K" if is_real_estate_template else None
        COL_DEMAND_COEFF = "L" if is_real_estate_template else None
        COL_OTHER = "M" if is_real_estate_template else None
        COL_TOTAL_GEO = "X" if is_real_estate_template else None
        COL_SHIPPED_AOV = None
        COL_CLIENT_COUNT = None
        COL_ABSOLUTE_NEW_CLIENTS = None
        COL_RETURNED_CLIENTS = None
        COL_ORDER_FREQUENCY = None
    if is_real_estate_template:
        COL_AK = "Q"
        COL_IMPRESSIONS = "V"
        COL_CTR = "AD"
        COL_CR = "AH"
        COL_CR2 = "AL" if is_real_estate_full_template else None
        COL_AOV = None
        COL_NEW_CLIENTS_SHARE = None
        COL_AVAILABLE_CAPACITY = None
        COL_REACH = None
        COL_FREQUENCY = None
    elif not is_diy_template:
        COL_AK = "N"
        COL_IMPRESSIONS = "S"
        COL_CTR = "AA"
        COL_CR = "AE"
        COL_CR2 = None
        COL_AOV = "AN"
        COL_NEW_CLIENTS_SHARE = None
        COL_AVAILABLE_CAPACITY = None
        COL_REACH = None
        COL_FREQUENCY = None

    def _safe_text(v) -> str:
        if v is None:
            return ""
        return str(v).strip().lower()

    def _norm_segment(v) -> str:
        t = _safe_text(v).replace(" ", "")
        if "b2b" in t:
            return "B2B"
        if "b2c" in t:
            return "B2C"
        return ""

    def _is_header_row(sheet, r: int) -> bool:
        b = sheet[f"{COL_SYSTEM}{r}"].value
        c = sheet[f"{COL_FORMAT}{r}"].value
        d = sheet[f"{COL_TARGETING}{r}"].value
        f = sheet[f"{COL_PERIOD}{r}"].value
        s = sheet[f"{COL_IMPRESSIONS}{r}"].value
        if not all(v is not None and str(v).strip() != "" for v in (b, c, d, f, s)):
            return False
        # Exclude subtotal/total rows where template keeps dashes in C/D.
        if str(c).strip() == "-" and str(d).strip() == "-":
            return False
        return True

    def _detect_periods_layout() -> tuple[int, int, int]:
        header_rows = []
        for r in range(1, min(ws.max_row, 500) + 1):
            if _is_header_row(ws, r):
                header_rows.append(r)
                if len(header_rows) >= 3:
                    break

        if not header_rows:
            return DEFAULT_START_ROW_JAN, DEFAULT_BLOCK_STEP, DEFAULT_ROWS_PER_MONTH

        header_row = header_rows[0]
        next_header_row = header_rows[1] if len(header_rows) > 1 else None
        start_row = header_row + 1

        if next_header_row:
            rows_per_month = max(1, next_header_row - start_row - 2)
            block_step = next_header_row - header_row
        else:
            search_end = min(ws.max_row, start_row + 50)
            total_row = None
            for r in range(start_row, search_end + 1):
                c = _safe_text(ws[f"{COL_FORMAT}{r}"].value)
                d = _safe_text(ws[f"{COL_TARGETING}{r}"].value)
                b = _safe_text(ws[f"{COL_SYSTEM}{r}"].value)
                if b and c == "-" and d == "-":
                    total_row = r
                    break
            rows_per_month = max(1, total_row - start_row) if total_row else DEFAULT_ROWS_PER_MONTH
            block_step = max(DEFAULT_BLOCK_STEP, rows_per_month + 2)

        return start_row, block_step, rows_per_month

    def _detect_total_layout(default_rows_per_month: int) -> tuple[int, int]:
        header_row = None
        for r in range(1, min(ws_total.max_row, 200) + 1):
            if _is_header_row(ws_total, r):
                header_row = r
                break

        if header_row is None:
            return DEFAULT_START_ROW_TOTAL, default_rows_per_month

        start_row = header_row + 1
        total_like_row = None
        for r in range(start_row, min(ws_total.max_row, start_row + 100) + 1):
            c = _safe_text(ws_total[f"{COL_FORMAT}{r}"].value)
            d = _safe_text(ws_total[f"{COL_TARGETING}{r}"].value)
            b = _safe_text(ws_total[f"{COL_SYSTEM}{r}"].value)
            if b and c == "-" and d == "-":
                total_like_row = r
                break

        rows_per_month = max(1, total_like_row - start_row) if total_like_row else default_rows_per_month
        return start_row, rows_per_month

    START_ROW_JAN, BLOCK_STEP, ROWS_PER_MONTH = _detect_periods_layout()
    START_ROW_TOTAL, ROWS_PER_MONTH_TOTAL = _detect_total_layout(ROWS_PER_MONTH)
    period_rows_for_hide: list[int] = []

    def _write_period_row(row_excel: int, camp_row: pd.Series | None, row_data: pd.Series | None, period_str: str):
        ws.row_dimensions[row_excel].hidden = False
        if camp_row is None or row_data is None:
            ws[f"{COL_SYSTEM}{row_excel}"] = None
            ws[f"{COL_FORMAT}{row_excel}"] = None
            ws[f"{COL_TARGETING}{row_excel}"] = None
            ws[f"{COL_PERIOD}{row_excel}"] = None
            ws[f"{COL_MODEL}{row_excel}"] = None
            if COL_GEO:
                ws[f"{COL_GEO}{row_excel}"] = None
            if COL_DEMAND_COEFF:
                ws[f"{COL_DEMAND_COEFF}{row_excel}"] = None
            if COL_OTHER:
                ws[f"{COL_OTHER}{row_excel}"] = None
            ws[f"{COL_IMPRESSIONS}{row_excel}"] = None
            ws[f"{COL_CTR}{row_excel}"] = None
            ws[f"{COL_CPC}{row_excel}"] = None
            ws[f"{COL_AK}{row_excel}"] = None
            ws[f"{COL_CR}{row_excel}"] = None
            if COL_CR2:
                ws[f"{COL_CR2}{row_excel}"] = None
            if COL_AOV:
                ws[f"{COL_AOV}{row_excel}"] = None
            if COL_NEW_CLIENTS_SHARE:
                ws[f"{COL_NEW_CLIENTS_SHARE}{row_excel}"] = None
            if COL_AVAILABLE_CAPACITY:
                ws[f"{COL_AVAILABLE_CAPACITY}{row_excel}"] = None
            if COL_REACH:
                ws[f"{COL_REACH}{row_excel}"] = None
            if COL_FREQUENCY:
                ws[f"{COL_FREQUENCY}{row_excel}"] = None
            if COL_CLIENT_COUNT:
                ws[f"{COL_CLIENT_COUNT}{row_excel}"] = None
            if COL_ABSOLUTE_NEW_CLIENTS:
                ws[f"{COL_ABSOLUTE_NEW_CLIENTS}{row_excel}"] = None
            if COL_RETURNED_CLIENTS:
                ws[f"{COL_RETURNED_CLIENTS}{row_excel}"] = None
            if COL_ORDER_FREQUENCY:
                ws[f"{COL_ORDER_FREQUENCY}{row_excel}"] = None
            if COL_SHIPPED_AOV:
                ws[f"{COL_SHIPPED_AOV}{row_excel}"] = None
            return

        impressions = float(row_data["impressions"])
        ctr = float(row_data["ctr"])
        cpc = float(row_data["cpc"])
        cr = float(row_data["cr1"] if is_real_estate_full_template else row_data["cr"])
        cr2 = float(row_data.get("cr2", 0.0) or 0.0)
        aov = float(row_data.get("aov", 0.0) or 0.0)
        shipped_aov = float(row_data.get("shipped_aov", 0.0) or 0.0)
        ak_rate = float(row_data.get("ak_rate", 0.0))
        if pd.isna(ak_rate):
            ak_rate = 0.0
        if pd.isna(cr2):
            cr2 = 0.0
        new_clients_share_pct = float(row_data.get("new_clients_share_pct", 0.0))
        if pd.isna(new_clients_share_pct):
            new_clients_share_pct = 0.0
        # App stores this metric in percents (e.g. 25), while Excel percent cells
        # typically expect a fraction (0.25) to display 25%.
        if new_clients_share_pct > 1:
            new_clients_share_pct = new_clients_share_pct / 100.0
        available_capacity = float(row_data.get("available_capacity", 0.0))
        if pd.isna(available_capacity):
            available_capacity = 0.0
        reach = float(row_data.get("reach", 0.0) or 0.0)
        if pd.isna(reach):
            reach = 0.0
        frequency = float(row_data.get("frequency", 0.0) or 0.0)
        if pd.isna(frequency):
            frequency = (impressions / reach) if reach > 0 else 0.0
        client_count = float(row_data.get("client_count", 0.0) or 0.0)
        absolute_new_clients = float(row_data.get("absolute_new_clients", 0.0) or 0.0)
        returned_clients = float(row_data.get("returned_clients", 0.0) or 0.0)
        order_frequency_raw = row_data.get("order_frequency", 0.0)
        order_frequency = None if pd.isna(order_frequency_raw) else float(order_frequency_raw or 0.0)
        demand_coeff = float(row_data.get("k_demand_applied", 1.0) or 1.0)
        if pd.isna(demand_coeff):
            demand_coeff = 1.0

        ws[f"{COL_SYSTEM}{row_excel}"] = camp_row.get("system", "")
        ws[f"{COL_FORMAT}{row_excel}"] = camp_row.get("campaign_type", "")
        ws[f"{COL_TARGETING}{row_excel}"] = camp_row.get("format", "")
        ws[f"{COL_PERIOD}{row_excel}"] = period_str
        ws[f"{COL_MODEL}{row_excel}"] = "CPC"
        if COL_GEO:
            ws[f"{COL_GEO}{row_excel}"] = camp_row.get("geo", "")
        if COL_DEMAND_COEFF:
            ws[f"{COL_DEMAND_COEFF}{row_excel}"] = demand_coeff
        if COL_OTHER:
            ws[f"{COL_OTHER}{row_excel}"] = None
        ws[f"{COL_IMPRESSIONS}{row_excel}"] = impressions
        ws[f"{COL_CTR}{row_excel}"] = ctr
        ws[f"{COL_CPC}{row_excel}"] = (cpc / 1.22) if is_diy_template else cpc
        ws[f"{COL_AK}{row_excel}"] = ak_rate
        ws[f"{COL_CR}{row_excel}"] = cr
        if COL_CR2:
            ws[f"{COL_CR2}{row_excel}"] = cr2
        if COL_AOV:
            ws[f"{COL_AOV}{row_excel}"] = aov
        if COL_NEW_CLIENTS_SHARE:
            ws[f"{COL_NEW_CLIENTS_SHARE}{row_excel}"] = new_clients_share_pct
        if COL_AVAILABLE_CAPACITY:
            ws[f"{COL_AVAILABLE_CAPACITY}{row_excel}"] = available_capacity
        if COL_REACH:
            ws[f"{COL_REACH}{row_excel}"] = reach
        if COL_FREQUENCY:
            ws[f"{COL_FREQUENCY}{row_excel}"] = frequency
        if COL_CLIENT_COUNT:
            ws[f"{COL_CLIENT_COUNT}{row_excel}"] = client_count
        if COL_ABSOLUTE_NEW_CLIENTS:
            ws[f"{COL_ABSOLUTE_NEW_CLIENTS}{row_excel}"] = absolute_new_clients
        if COL_RETURNED_CLIENTS:
            ws[f"{COL_RETURNED_CLIENTS}{row_excel}"] = returned_clients
        if COL_ORDER_FREQUENCY:
            ws[f"{COL_ORDER_FREQUENCY}{row_excel}"] = order_frequency
        if COL_SHIPPED_AOV:
            ws[f"{COL_SHIPPED_AOV}{row_excel}"] = shipped_aov

    def _write_total_row(row_excel: int, camp_row: pd.Series | None):
        ws_total.row_dimensions[row_excel].hidden = False
        if camp_row is None:
            ws_total[f"{COL_SYSTEM}{row_excel}"] = None
            ws_total[f"{COL_FORMAT}{row_excel}"] = None
            ws_total[f"{COL_TARGETING}{row_excel}"] = None
            if COL_TOTAL_GEO:
                ws_total[f"{COL_TOTAL_GEO}{row_excel}"] = None
            return
        ws_total[f"{COL_SYSTEM}{row_excel}"] = camp_row.get("system", "")
        ws_total[f"{COL_FORMAT}{row_excel}"] = camp_row.get("campaign_type", "")
        ws_total[f"{COL_TARGETING}{row_excel}"] = camp_row.get("format", "")
        if COL_TOTAL_GEO:
            ws_total[f"{COL_TOTAL_GEO}{row_excel}"] = camp_row.get("geo", "")

    def _collect_diy_period_rows(block_start_row: int) -> tuple[list[int], list[int]] | None:
        scan_from = block_start_row
        scan_to = min(ws.max_row, block_start_row + BLOCK_STEP + 8)
        row_b2c_total = None
        row_b2b_total = None
        for r in range(scan_from, scan_to + 1):
            t = _safe_text(ws[f"{COL_SYSTEM}{r}"].value)
            c = _safe_text(ws[f"{COL_FORMAT}{r}"].value)
            d = _safe_text(ws[f"{COL_TARGETING}{r}"].value)
            if c != "-" or d != "-":
                continue
            if "b2c" in t and row_b2c_total is None:
                row_b2c_total = r
            elif "b2b" in t and row_b2b_total is None:
                row_b2b_total = r

        if row_b2c_total is None or row_b2b_total is None or row_b2b_total <= row_b2c_total:
            return None

        b2c_rows = list(range(block_start_row, row_b2c_total))
        b2b_rows = list(range(row_b2c_total + 1, row_b2b_total))
        return b2c_rows, b2b_rows

    def _collect_diy_total_rows() -> tuple[list[int], list[int]] | None:
        scan_from = START_ROW_TOTAL
        scan_to = min(ws_total.max_row, START_ROW_TOTAL + ROWS_PER_MONTH_TOTAL + 40)
        row_b2c_total = None
        row_b2b_total = None
        for r in range(scan_from, scan_to + 1):
            t = _safe_text(ws_total[f"{COL_SYSTEM}{r}"].value)
            c = _safe_text(ws_total[f"{COL_FORMAT}{r}"].value)
            d = _safe_text(ws_total[f"{COL_TARGETING}{r}"].value)
            if c != "-" or d != "-":
                continue
            if "b2c" in t and row_b2c_total is None:
                row_b2c_total = r
            elif "b2b" in t and row_b2b_total is None:
                row_b2b_total = r

        if row_b2c_total is None or row_b2b_total is None or row_b2b_total <= row_b2c_total:
            return None

        b2c_rows = list(range(START_ROW_TOTAL, row_b2c_total))
        b2b_rows = list(range(row_b2c_total + 1, row_b2b_total))
        return b2c_rows, b2b_rows

    for block_index, period in enumerate(selected_periods):
        month_num = int(period.get("month_num", 1) or 1)
        year = int(period.get("month_year", dt.date.today().year) or dt.date.today().year)
        planning_slot = int(period.get("planning_slot", block_index + 1) or (block_index + 1))
        block_start_row = START_ROW_JAN + block_index * BLOCK_STEP
        start = dt.date(year, month_num, 1)
        end = dt.date(year, 12, 31) if month_num == 12 else dt.date(year, month_num + 1, 1) - dt.timedelta(days=1)
        period_str = f"{start.strftime('%d.%m.%Y')} - {end.strftime('%d.%m.%Y')}"
        if "planning_slot" in df_all.columns:
            df_month = df_all[df_all["planning_slot"] == planning_slot]
        else:
            df_month = df_all[df_all["month_num"] == month_num]

        diy_period_rows = _collect_diy_period_rows(block_start_row) if (template_kind == "diy" and "segment" in campaigns.columns) else None

        if diy_period_rows is not None:
            b2c_rows, b2b_rows = diy_period_rows
            period_rows_for_hide.extend(b2c_rows)
            period_rows_for_hide.extend(b2b_rows)
            camps = campaigns.copy()
            camps["_seg"] = camps["segment"].map(_norm_segment)
            camps_b2c = camps[camps["_seg"] == "B2C"]
            camps_b2b = camps[camps["_seg"] == "B2B"]

            for row_excel, (_, camp) in zip(b2c_rows, camps_b2c.iterrows()):
                row_data = df_month[df_month["campaign_type"] == camp["campaign_type"]]
                _write_period_row(row_excel, camp, (None if row_data.empty else row_data.iloc[0]), period_str)
            for row_excel in b2c_rows[len(camps_b2c):]:
                _write_period_row(row_excel, None, None, period_str)

            for row_excel, (_, camp) in zip(b2b_rows, camps_b2b.iterrows()):
                row_data = df_month[df_month["campaign_type"] == camp["campaign_type"]]
                _write_period_row(row_excel, camp, (None if row_data.empty else row_data.iloc[0]), period_str)
            for row_excel in b2b_rows[len(camps_b2b):]:
                _write_period_row(row_excel, None, None, period_str)
        else:
            period_rows_for_hide.extend(block_start_row + i for i in range(ROWS_PER_MONTH))
            for i, (_, camp) in enumerate(campaigns.iterrows()):
                if i >= ROWS_PER_MONTH:
                    break
                row_excel = block_start_row + i
                row_data = df_month[df_month["campaign_type"] == camp["campaign_type"]]
                _write_period_row(row_excel, camp, (None if row_data.empty else row_data.iloc[0]), period_str)
            for i in range(min(len(campaigns), ROWS_PER_MONTH), ROWS_PER_MONTH):
                _write_period_row(block_start_row + i, None, None, period_str)

    if compact_empty_rows:
        rows_to_hide_periods = []
        for row_excel in sorted(set(period_rows_for_hide)):
            is_empty_main = (
                ws[f"{COL_SYSTEM}{row_excel}"].value in (None, "")
                and ws[f"{COL_FORMAT}{row_excel}"].value in (None, "")
                and ws[f"{COL_TARGETING}{row_excel}"].value in (None, "")
            )
            is_empty_metrics = (
                ws[f"{COL_IMPRESSIONS}{row_excel}"].value in (None, "")
                and ws[f"{COL_CTR}{row_excel}"].value in (None, "")
                and ws[f"{COL_CPC}{row_excel}"].value in (None, "")
                and ws[f"{COL_AK}{row_excel}"].value in (None, "")
                and ws[f"{COL_CR}{row_excel}"].value in (None, "")
                and (not COL_AOV or ws[f"{COL_AOV}{row_excel}"].value in (None, ""))
                and (not COL_NEW_CLIENTS_SHARE or ws[f"{COL_NEW_CLIENTS_SHARE}{row_excel}"].value in (None, ""))
                and (not COL_AVAILABLE_CAPACITY or ws[f"{COL_AVAILABLE_CAPACITY}{row_excel}"].value in (None, ""))
                and (not COL_REACH or ws[f"{COL_REACH}{row_excel}"].value in (None, ""))
                and (not COL_FREQUENCY or ws[f"{COL_FREQUENCY}{row_excel}"].value in (None, ""))
                and (not COL_CLIENT_COUNT or ws[f"{COL_CLIENT_COUNT}{row_excel}"].value in (None, ""))
                and (not COL_ABSOLUTE_NEW_CLIENTS or ws[f"{COL_ABSOLUTE_NEW_CLIENTS}{row_excel}"].value in (None, ""))
                and (not COL_RETURNED_CLIENTS or ws[f"{COL_RETURNED_CLIENTS}{row_excel}"].value in (None, ""))
                and (not COL_ORDER_FREQUENCY or ws[f"{COL_ORDER_FREQUENCY}{row_excel}"].value in (None, ""))
                and (not COL_SHIPPED_AOV or ws[f"{COL_SHIPPED_AOV}{row_excel}"].value in (None, ""))
            )
            if is_empty_main and is_empty_metrics:
                rows_to_hide_periods.append(row_excel)
        for row_idx in sorted(set(rows_to_hide_periods)):
            ws.row_dimensions[row_idx].hidden = True

    diy_total_rows = _collect_diy_total_rows() if (template_kind == "diy" and "segment" in campaigns.columns) else None

    if diy_total_rows is not None:
        b2c_rows_t, b2b_rows_t = diy_total_rows
        total_rows_for_hide = list(b2c_rows_t) + list(b2b_rows_t)
        camps = campaigns.copy()
        camps["_seg"] = camps["segment"].map(_norm_segment)
        camps_b2c = camps[camps["_seg"] == "B2C"]
        camps_b2b = camps[camps["_seg"] == "B2B"]

        for row_excel, (_, camp) in zip(b2c_rows_t, camps_b2c.iterrows()):
            _write_total_row(row_excel, camp)
        for row_excel in b2c_rows_t[len(camps_b2c):]:
            _write_total_row(row_excel, None)

        for row_excel, (_, camp) in zip(b2b_rows_t, camps_b2b.iterrows()):
            _write_total_row(row_excel, camp)
        for row_excel in b2b_rows_t[len(camps_b2b):]:
            _write_total_row(row_excel, None)
    else:
        total_rows_for_hide = list(range(START_ROW_TOTAL, START_ROW_TOTAL + ROWS_PER_MONTH_TOTAL))
        for i in range(ROWS_PER_MONTH_TOTAL):
            row_excel = START_ROW_TOTAL + i
            if i < len(campaigns):
                _write_total_row(row_excel, campaigns.iloc[i])
            else:
                _write_total_row(row_excel, None)

    if compact_empty_rows:
        rows_to_hide_total = []
        for row_excel in total_rows_for_hide:
            if (
                ws_total[f"{COL_SYSTEM}{row_excel}"].value in (None, "")
                and ws_total[f"{COL_FORMAT}{row_excel}"].value in (None, "")
                and ws_total[f"{COL_TARGETING}{row_excel}"].value in (None, "")
            ):
                rows_to_hide_total.append(row_excel)
        for row_idx in sorted(set(rows_to_hide_total)):
            ws_total.row_dimensions[row_idx].hidden = True

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
def resolve_template_path(template_kind: str = "ecom") -> str | None:
    template_paths = TEMPLATE_PATHS_BY_KIND.get(template_kind, TEMPLATE_PATHS_ECOM)
    existing_templates = [p for p in template_paths if os.path.exists(p)]
    if not existing_templates:
        return None
    return max(existing_templates, key=os.path.getmtime)


def build_template_export_payload(
    df_export: pd.DataFrame,
    campaigns_export: pd.DataFrame,
    collapse_geo_to_rf: bool = False,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    df_template = df_export.copy()
    campaigns_template = campaigns_export.copy()

    if not collapse_geo_to_rf:
        return df_template, campaigns_template

    if "geo" not in df_template.columns:
        df_template["geo"] = ""
    if "geo" not in campaigns_template.columns:
        campaigns_template["geo"] = ""

    group_cols = ["planning_slot", "month_num", "month_year", "month_name", "campaign_type", "system", "format"]
    if "segment" in df_template.columns:
        group_cols.append("segment")

    sum_candidates = [
        "impressions", "reach", "clicks", "conversions", "leads", "target_leads",
        "cost", "ak_cost_wo_vat", "total_budget_wo_vat_ak", "cost_with_vat", "cost_with_vat_ak", "vat_amount",
        "revenue", "available_capacity", "client_count", "absolute_new_clients", "returned_clients", "new_clients",
        "shipped_orders", "shipped_revenue",
    ]
    agg_map = {col: "sum" for col in sum_candidates if col in df_template.columns}
    df_template = df_template.groupby(group_cols, as_index=False, sort=False).agg(agg_map)
    df_template["geo"] = "РФ"

    def _safe_ratio(num_col: str, den_col: str, pct: bool = False) -> pd.Series:
        if num_col not in df_template.columns or den_col not in df_template.columns:
            return pd.Series(0.0, index=df_template.index, dtype=float)
        num = pd.to_numeric(df_template[num_col], errors="coerce").fillna(0.0)
        den = pd.to_numeric(df_template[den_col], errors="coerce").fillna(0.0)
        out = np.where(den > 0, num / den, 0.0)
        if pct:
            out = out * 100.0
        return pd.Series(out, index=df_template.index, dtype=float)

    df_template["frequency"] = _safe_ratio("impressions", "reach", pct=False)
    df_template["ctr"] = _safe_ratio("clicks", "impressions", pct=False)
    df_template["cpc"] = _safe_ratio("cost", "clicks", pct=False)
    if "conversions" in df_template.columns:
        df_template["cr"] = _safe_ratio("conversions", "clicks", pct=False)
    elif "target_leads" in df_template.columns:
        df_template["cr"] = _safe_ratio("target_leads", "clicks", pct=False)
    else:
        df_template["cr"] = 0.0
    if "shipped_orders" in df_template.columns:
        df_template["cr2"] = _safe_ratio("shipped_orders", "conversions", pct=False)
    elif "target_leads" in df_template.columns and "leads" in df_template.columns:
        df_template["cr2"] = _safe_ratio("target_leads", "leads", pct=False)
    else:
        df_template["cr2"] = 0.0
    df_template["aov"] = _safe_ratio("revenue", "conversions", pct=False)
    df_template["shipped_aov"] = _safe_ratio("shipped_revenue", "shipped_orders", pct=False)
    df_template["ak_rate"] = _safe_ratio("ak_cost_wo_vat", "cost", pct=False)
    df_template["ak_rate_pct"] = df_template["ak_rate"] * 100.0
    df_template["new_clients_share_pct"] = _safe_ratio("new_clients", "client_count", pct=True)
    df_template["sov_pct"] = _safe_ratio("reach", "available_capacity", pct=True)
    # For collapsed RF export this metric is intentionally left empty:
    # the specialist can set it manually in the template.
    df_template["order_frequency"] = np.nan
    df_template["cac"] = _safe_ratio("cost_with_vat_ak", "new_clients", pct=False)
    df_template["cpm"] = np.where(
        pd.to_numeric(df_template.get("impressions", 0.0), errors="coerce").fillna(0.0) > 0,
        pd.to_numeric(df_template.get("cost_with_vat_ak", 0.0), errors="coerce").fillna(0.0)
        / (pd.to_numeric(df_template.get("impressions", 0.0), errors="coerce").fillna(0.0) / 1000.0),
        0.0,
    )
    df_template["roas"] = _safe_ratio("revenue", "cost_with_vat_ak", pct=False)
    df_template["drr"] = _safe_ratio("cost_with_vat_ak", "revenue", pct=False)
    df_template["shipped_roas"] = _safe_ratio("shipped_revenue", "cost_with_vat_ak", pct=False)
    df_template["shipped_drr_pct"] = _safe_ratio("cost_with_vat_ak", "shipped_revenue", pct=True)
    df_template["shipped_cps"] = _safe_ratio("cost_with_vat_ak", "shipped_orders", pct=False)

    campaign_key_cols = ["campaign_type", "system", "format"]
    if "segment" in campaigns_template.columns:
        campaign_key_cols.append("segment")
    campaigns_template["geo"] = "РФ"
    campaigns_template = (
        campaigns_template
        .drop_duplicates(subset=campaign_key_cols, keep="first")
        .reset_index(drop=True)
    )

    return df_template, campaigns_template


# ---------- ЗАГОЛОВОК  ТАБЫ ----------

coeff_header_uri = image_file_to_data_uri(os.path.join(BASE_DIR, "assets", "coeff_header_mascot.png"))
plan_header_uri = image_file_to_data_uri(os.path.join(BASE_DIR, "assets", "plan_header_mascot.png"))
charts_header_uri = image_file_to_data_uri(os.path.join(BASE_DIR, "assets", "charts_header_mascot.png"))
export_header_uri = image_file_to_data_uri(os.path.join(BASE_DIR, "assets", "export_header_mascot.png"))
top_header_uri = coeff_header_uri or plan_header_uri or charts_header_uri or export_header_uri
if top_header_uri:
    top_header_html = """
        <style>
        .app-top-hero {
            position: relative;
            border: 1px solid #1D2A44;
            border-radius: 14px;
            overflow: hidden;
            margin: 0.15rem 0 0.85rem 0;
            min-height: 250px;
            background: #0B1220;
        }
        .app-top-hero-bg {
            position: absolute;
            inset: 0;
            background-image: url("__TOP_HEADER_URI__");
            background-size: contain;
            background-repeat: no-repeat;
            background-position: calc(100% - 20px) center;
            opacity: 0.0;
            transform: scale(1.0);
            filter: none;
            transition: opacity 220ms ease;
        }
        .app-top-hero-fade {
            position: absolute;
            inset: 0;
            background: linear-gradient(90deg, rgba(8, 16, 30, 0.90) 0%, rgba(8, 16, 30, 0.48) 42%, rgba(8, 16, 30, 0.02) 72%, rgba(8, 16, 30, 0.00) 100%);
        }
        .app-top-hero-content {
            position: relative;
            z-index: 2;
            padding: 14px 16px 14px 16px;
        }
        .app-top-hero-content .intro-card {
            margin: 0.2rem 0 0.2rem 0;
            padding: 12px 14px;
            width: fit-content;
            max-width: calc(100% - 360px);
            border-radius: 12px;
            border: 1px solid #3D8EF0;
            border-left: 4px solid #00CDC5;
            background: linear-gradient(180deg, rgba(0, 102, 224, 0.20) 0%, rgba(17, 26, 46, 0.84) 100%);
            box-shadow: 0 8px 18px rgba(0, 0, 0, 0.20);
        }
        @media (max-width: 1150px) {
            .app-top-hero-content .intro-card {
                width: 100%;
            }
            .app-top-hero-bg {
                background-position: calc(100% - 12px) center;
            }
        }
        .app-top-hero-content .intro-card p {
            margin: 0 0 6px 0;
            color: #EAF0FF;
            line-height: 1.38;
            font-size: clamp(0.93rem, 0.92vw, 1.06rem);
            overflow-wrap: anywhere;
        }
        .app-top-hero-content .intro-card p.one-line {
            white-space: normal;
        }
        .app-top-hero-content .intro-card p:last-child {
            margin-bottom: 0;
        }
        @media (max-width: 1150px) {
            .app-top-hero-content .intro-card p.one-line {
                white-space: normal;
            }
        }
        </style>
        <div class="app-top-hero">
            <div id="app-top-hero-bg" class="app-top-hero-bg"></div>
            <div class="app-top-hero-fade"></div>
            <div class="app-top-hero-content">
                <h1 style="font-weight: 700; letter-spacing: 0.02em; margin: 0 0 6px 0; max-width: 70%;">
                    Медиапланер <span style="color: #00CDC5; font-size: 1.25em;">E-promo</span>
                </h1>
                <div class="intro-card">
                    <p><span style="font-weight: 800; color: #9EC5FF;">Что это:</span>
                    инструмент для расчета медиаплана на выбранный период (от 1 до 12 месяцев) по типам рекламных кампаний.</p>
                    <p class="one-line"><span style="font-weight: 800; color: #9EC5FF;">Зачем нужен:</span>
                    чтобы упростить алгоритм расчета медиаплана, сократить время на первичный расчет и ускорить внесение последующих правок.</p>
                </div>
            </div>
        </div>
        """
    st.markdown(
        top_header_html.replace("__TOP_HEADER_URI__", top_header_uri),
        unsafe_allow_html=True,
    )
else:
    st.markdown(
        """
        <h1 style="font-weight: 700; letter-spacing: 0.02em;">
            Медиапланер <span style="color: #00CDC5; font-size: 1.25em;">E-promo</span>
        </h1>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        """
        <div style="
            margin: 0.2rem 0 0.8rem 0;
            padding: 12px 14px;
            border-radius: 12px;
            border: 1px solid #3D8EF0;
            border-left: 4px solid #00CDC5;
            background: linear-gradient(180deg, rgba(0, 102, 224, 0.18) 0%, rgba(17, 26, 46, 0.82) 100%);
            box-shadow: 0 8px 18px rgba(0, 0, 0, 0.20);
        ">
            <p style="margin: 0 0 6px 0; color: #EAF0FF; line-height: 1.45;">
                <span style="font-weight: 800; color: #9EC5FF;">Как это работает:</span>
                Добавьте один или несколько коэффициентов по выбранным месяцам (от 1 до 12 месяцев) на основе доступных сценариев.
            </p>
            <p style="margin: 0; color: #EAF0FF; line-height: 1.45;">
                <span style="font-weight: 800; color: #9EC5FF;">Зачем нужно:</span>
                Чтобы гибко учитывать сезонные колебания, изменения спроса по выбранным месяцам и влияние медийных хвостов.
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )

ui_compact_tables = True
ui_editor_row_height = 28
ui_table_row_px = 29
ui_table_font_px = "12px"

st.markdown(
    f"""
    <style>
    [data-testid="stDataFrame"] [role="gridcell"],
    [data-testid="stDataFrame"] [role="columnheader"] {{
        font-size: {ui_table_font_px};
        line-height: 1.0 !important;
        padding-top: 2px !important;
        padding-bottom: 2px !important;
        min-height: {ui_table_row_px}px !important;
        height: {ui_table_row_px}px !important;
        max-height: {ui_table_row_px}px !important;
        box-sizing: border-box !important;
    }}
    </style>
    """,
    unsafe_allow_html=True,
)

tab_coeffs, tab_setup, tab_plan, tab_charts, tab_export = st.tabs(
    ["Коэффициенты", "Настройки расчета", "Медиаплан", "Диаграммы", "Export/Import"]
)
if top_header_uri:
    header_tab_bg_map_json = json.dumps(
        {
            0: coeff_header_uri or "",
            1: plan_header_uri or "",
            2: plan_header_uri or "",
            3: charts_header_uri or "",
            4: export_header_uri or "",
        },
        ensure_ascii=False,
    )
    components.html(
        """
        <script>
        (function() {
            const TAB_BG_MAP = __TAB_BG_MAP__;
            const VISIBLE_OPACITY = "1.00";
            const COEFF_TAB_INDEX = 0;
            const SETUP_TAB_INDEX = 1;
            const PLAN_TAB_INDEX = 2;
            const CHARTS_TAB_INDEX = 3;
            const EXPORT_TAB_INDEX = 4;
            const BASE_BG_POSITION = "calc(100% - 20px) center";
            const COEFF_BG_POSITION = "calc(100% + 95px) calc(50% + 77px)";
            const PLAN_BG_POSITION = "calc(100% + 120px) center";
            const CHARTS_BG_SIZE = "47% auto";
            const CHARTS_BG_POSITION = "calc(100% + 120px) calc(50% + 30px)";
            const EXPORT_BG_SIZE = "24.8% auto";
            const EXPORT_BG_POSITION = "calc(100% - 20px) calc(50% + 30px)";

            const getTopTablist = (doc) => {
                const tablists = Array.from(doc.querySelectorAll('[role="tablist"]'));
                if (!tablists.length) return null;
                return tablists
                    .map(tl => ({ tl, top: tl.getBoundingClientRect().top }))
                    .sort((a, b) => a.top - b.top)[0]?.tl || null;
            };

            const applyHeaderImage = () => {
                const doc = window.parent.document;
                const bg = doc.getElementById("app-top-hero-bg");
                if (!bg) return false;
                const topTablist = getTopTablist(doc);
                if (!topTablist) return false;
                const tabs = topTablist.querySelectorAll('button[role="tab"]');
                if (!tabs || !tabs.length) return false;

                let activeIdx = 0;
                for (let i = 0; i < tabs.length; i++) {
                    if (tabs[i].getAttribute("aria-selected") === "true") {
                        activeIdx = i;
                        break;
                    }
                }
                const key = String(activeIdx);
                const imgUrl = TAB_BG_MAP[key] || "";
                bg.style.backgroundImage = imgUrl ? `url(${imgUrl})` : "";
                if (activeIdx === COEFF_TAB_INDEX) {
                    bg.style.backgroundSize = "52% auto";
                    bg.style.backgroundPosition = COEFF_BG_POSITION;
                } else if (activeIdx === SETUP_TAB_INDEX) {
                    bg.style.backgroundSize = "58% auto";
                    bg.style.backgroundPosition = PLAN_BG_POSITION;
                } else if (activeIdx === PLAN_TAB_INDEX) {
                    bg.style.backgroundSize = "58% auto";
                    bg.style.backgroundPosition = PLAN_BG_POSITION;
                } else if (activeIdx === CHARTS_TAB_INDEX) {
                    bg.style.backgroundSize = CHARTS_BG_SIZE;
                    bg.style.backgroundPosition = CHARTS_BG_POSITION;
                } else if (activeIdx === EXPORT_TAB_INDEX) {
                    bg.style.backgroundSize = EXPORT_BG_SIZE;
                    bg.style.backgroundPosition = EXPORT_BG_POSITION;
                } else {
                    bg.style.backgroundSize = "58% auto";
                    bg.style.backgroundPosition = BASE_BG_POSITION;
                }
                bg.style.opacity = imgUrl ? VISIBLE_OPACITY : "0.0";
                return true;
            };

            let attempts = 0;
            const init = () => {
                applyHeaderImage();
                attempts += 1;
                if (attempts < 30) setTimeout(init, 120);
            };
            init();

            const topTablist = getTopTablist(window.parent.document);
            if (topTablist) {
                const obs = new MutationObserver(() => applyHeaderImage());
                obs.observe(topTablist, {
                    subtree: true,
                    attributes: true,
                    attributeFilter: ["aria-selected", "class"],
                });
            }
        })();
        </script>
        """.replace("__TAB_BG_MAP__", header_tab_bg_map_json),
        height=0,
        width=0,
    )

# Быстрый импорт доступен всегда (даже до первого расчета на вкладке "Медиаплан").
with st.sidebar:
    st.markdown("### Быстрый импорт проекта")
    st.caption("Загрузите JSON проекта сразу после запуска приложения.")
    uploaded_project_quick = st.file_uploader(
        "Импорт проекта (JSON)",
        type=["json"],
        key="upload_project_json_quick",
    )
    queue_project_import_from_upload(uploaded_project_quick, "quick_sidebar")
# ====================================================================
#                        ТАБ «КОЭФФИЦИЕНТЫ»
# ====================================================================

with tab_coeffs:
    st.markdown(
        """
        <div class="tab-intro">
            <p>1) Создайте набор коэффициентов и выберите тип: <b>Спрос</b>, <b>AOV</b>, <b>Кастомный набор</b> или <b>Медийные хвосты</b>.</p>
            <p>2) Укажите стартовый месяц и год, затем заполните значения по периодам в таблице по данным из <a href="https://wordstat.yandex.ru/" target="_blank">Вордстат</a>.</p>
            <p>3) Нажмите кнопку применения/расчета, чтобы сохранить набор и получить итоговые коэффициенты.</p>
            <p>4) Повторите для всех наборов, которые планируете использовать в «Медиаплане».</p>
            <p>5) Проверьте, что у каждого набора заполнены данные, иначе он не будет влиять на расчет.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    ui_section_title("Коэффициенты сезонности")
    st.caption(
        "Создавайте один или несколько наборов коэффициентов сезонности: "
        "по спросу (запросы) или по AOV (средний чек). Для каждого набора считайте индексы по месяцам."
    )
    st.markdown(
        """
        <div style="
            margin: 8px 0 14px 0;
            display: inline-block;
            width: fit-content;
            max-width: 100%;
            white-space: nowrap;
            padding: 10px 12px;
            border-radius: 10px;
            border: 1px solid #FF8A66;
            background: rgba(255, 99, 51, 0.14);
            color: #FFD9CC;
            font-weight: 600;
        ">
            Важно: для корректного расчета медиаплана сначала обязательно рассчитайте минимум два набора:
            «Сезонность» (спрос) и «AOV» (средний чек).
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.caption(
        "Источники данных: для «Спрос» — Wordstat; для «AOV» — фактический средний чек из аналитики; "
        "для «Медийных хвостов» — обратитесь к специалистам по медийной рекламе для расчета влияния."
    )

    # ---------------- ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ----------------

    def generate_months_list(start_month: int, start_year: int, count: int = 24):
        month_names = [
            "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
            "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
        ]
        months = []
        current_month = start_month
        current_year = start_year
        for _ in range(count):
            months.append(
                {
                    "period": f"{month_names[current_month - 1]} {current_year}",
                    "month_num": current_month,
                    "year": current_year,
                }
            )
            current_month += 1
            if current_month > 12:
                current_month = 1
                current_year += 1
        return months

    def calculate_seasonality_coeffs_demand(df_data: pd.DataFrame, query_cols: list[str]) -> pd.DataFrame:
        """
        Спрос (по запросам), логика как в Excel-файле:
        1) Для каждого календарного месяца суммируем по каждому запросу (по всем годам).
        2) Внутри месяца считаем веса запросов как доли (сумма по запросу / сумма по всем запросам месяца).
        3) Сред. взв. месяца = сумма(сумма_по_запросу_месяца * вес_запроса_в_месяце).
        4) Среднее по 12 месяцам = СРЗНАЧ(Сред. взв. по месяцам).
        5) Коэф. месяца = Сред. взв. месяца / Среднее по 12 месяцам.
        """
        month_names_map = {
            1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
            5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
            9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
        }

        df = df_data.copy()

        avg_weighted_by_month: dict[int, float] = {}

        for m in range(1, 13):
            df_m = df[df["month_num"] == m]
            if df_m.empty:
                avg_weighted_by_month[m] = 0.0
                continue

            sums_per_query = df_m[query_cols].sum(axis=0)
            total_sum_month = float(sums_per_query.sum())
            if total_sum_month <= 0:
                avg_weighted_by_month[m] = 0.0
                continue

            weights = sums_per_query / total_sum_month
            avg_weighted = float((sums_per_query * weights).sum())
            avg_weighted_by_month[m] = avg_weighted

        values = list(avg_weighted_by_month.values())
        avg_all_months = float(np.mean(values)) if values else 1.0
        if avg_all_months == 0:
            avg_all_months = 1.0

        rows = []
        for m in range(1, 13):
            avg_w = avg_weighted_by_month.get(m, 0.0)
            coeff = avg_w / avg_all_months if avg_all_months > 0 else 1.0
            rows.append(
                {
                    "Номер месяца": m,
                    "Месяц": month_names_map[m],
                    "Сред. взв.": round(avg_w, 2),
                    "Коэф.": round(coeff, 2),
                }
            )

        return pd.DataFrame(rows)

    def calculate_seasonality_coeffs_aov(df_data: pd.DataFrame) -> pd.DataFrame:
        """
        AOV (средний чек) по месяцам:
        1) Для каждого календарного месяца считаем средний AOV по строкам (по годам).
        2) Считаем средний AOV по 12 месяцам.
        3) Коэф. AOV месяца = Сред. AOV месяца / средний AOV по 12 месяцам.
        """
        month_names_map = {
            1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
            5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
            9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
        }

        df = df_data.copy()

        # средний AOV по каждому календарному месяцу
        if "AOV" not in df.columns:
            df["AOV"] = 0.0
        avg_aov_by_month = df.groupby("month_num")["AOV"].mean()

        avg_all = float(avg_aov_by_month.mean()) if not avg_aov_by_month.empty else 1.0
        if avg_all == 0:
            avg_all = 1.0

        rows = []
        for m in range(1, 13):
            avg_m = float(avg_aov_by_month.get(m, 0.0))
            coeff = avg_m / avg_all if avg_all > 0 else 1.0
            rows.append(
                {
                    "Номер месяца": m,
                    "Месяц": month_names_map[m],
                    "Сред. AOV": round(avg_m, 2),
                    "Коэф. AOV": round(coeff, 2),
                }
            )

        return pd.DataFrame(rows)

    # ---------------- ИНИЦИАЛИЗАЦИЯ СОСТОЯНИЯ ----------------

    if "coeff_sets" not in st.session_state:
        st.session_state["coeff_sets"] = []
    if "coeff_active_set_id" not in st.session_state:
        st.session_state["coeff_active_set_id"] = None

    def _set_active_coeff_set(set_id: int) -> None:
        st.session_state["coeff_active_set_id"] = int(set_id)


    # ---------------- ДОБАВЛЕНИЕ НОВОГО НАБОРА ----------------

    col_add_set, col_help = st.columns([2, 3])
    with col_add_set:
        if st.button("Добавить новый набор коэффициентов", key="add_coeff_set"):
            new_id = len(st.session_state["coeff_sets"]) + 1
            st.session_state["coeff_sets"].append(
                {
                    "id": new_id,
                    "name": f"Набор {new_id}",
                    "type": "Спрос (по запросам)",  # или "AOV (средний чек)"
                    "start_month": 1,
                    "start_year": 2024,
                    "period_months": 24,
                    "queries": [],
                    "df_data": None,
                    "result": None,
                }
            )
            _set_active_coeff_set(new_id)
            st.rerun()
    with col_help:
        st.empty()

    # ---------------- СПИСОК НАБОРОВ ----------------

    if not st.session_state["coeff_sets"]:
        st.session_state["coeff_active_set_id"] = None
        st.info("Пока нет ни одного набора коэффициентов. Нажмите «Добавить новый набор коэффициентов».")
    else:
        existing_coeff_set_ids = [int(cs["id"]) for cs in st.session_state["coeff_sets"]]
        if st.session_state.get("coeff_active_set_id") not in existing_coeff_set_ids:
            st.session_state["coeff_active_set_id"] = existing_coeff_set_ids[-1]
        for idx, cs in enumerate(st.session_state["coeff_sets"]):
            set_id = cs["id"]
            with st.expander(f"Набор {set_id}: {cs['name']}", expanded=st.session_state.get("coeff_active_set_id") == set_id):

                # ---- 1-2. Название и тип набора (компактный layout) ----
                meta_c1, meta_c2, _meta_spacer = st.columns([2.2, 2.2, 1.6], vertical_alignment="bottom")
                with meta_c1:
                    cs["name"] = st.text_input(
                        "Название набора",
                        value=cs["name"],
                        key=f"cs_name_{set_id}",
                        on_change=_set_active_coeff_set,
                        args=(set_id,),
                    )
                with meta_c2:
                    cs["type"] = normalize_coeff_set_type(cs.get("type"))
                    cs["type"] = st.selectbox(
                        "Тип набора:",
                        options=["Спрос (по запросам)", "AOV (средний чек)", "Кастомный набор", "Медийные хвосты"],
                        help=(
                            "• Спрос (по запросам)\n"
                            "  Автоматические коэффициенты на основе Wordstat.\n\n"
                            "• AOV (средний чек)\n"
                            "  Автоматические коэффициенты среднего чека.\n\n"
                            "• Кастомный набор\n"
                            "  Ручной ввод коэффициентов на 12 месяцев.\n\n"
                            "• Медийные хвосты\n"
                            "  Коэффициенты показов в месяцах, накладываются поверх коэффициента спроса."
                        ),
                        index=(
                            0 if cs.get("type") == "Спрос (по запросам)"
                            else 1 if cs.get("type") == "AOV (средний чек)"
                            else 2 if cs.get("type") == "Кастомный набор"
                            else 3
                        ),
                        key=f"cs_type_{set_id}",
                        on_change=_set_active_coeff_set,
                        args=(set_id,),
                    )

                # ---- 3. Период 24 месяца (компактный layout) ----
                col_m, col_y, _period_spacer = st.columns([1.8, 1.8, 2.4], vertical_alignment="bottom")
                with col_m:
                    cs["start_month"] = st.selectbox(
                        "Начальный месяц периода:",
                        options=list(range(1, 13)),
                        format_func=lambda x: [
                            "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
                            "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
                        ][x - 1],
                        index=cs["start_month"] - 1,
                        key=f"cs_month_{set_id}",
                        on_change=_set_active_coeff_set,
                        args=(set_id,),
                    )
                with col_y:
                    cs["start_year"] = st.number_input(
                        "Начальный год периода:",
                        min_value=2020,
                        max_value=2030,
                        value=cs["start_year"],
                        step=1,
                        key=f"cs_year_{set_id}",
                        on_change=_set_active_coeff_set,
                        args=(set_id,),
                    )
                if "period_months" not in cs:
                    cs["period_months"] = 24
                cs["period_months"] = int(cs.get("period_months", 24) or 24)
                period_len_col, _period_len_spacer = st.columns([1.8, 4.2], vertical_alignment="bottom")
                with period_len_col:
                    cs["period_months"] = int(
                        st.number_input(
                            "Длительность периода, мес.",
                            min_value=24,
                            max_value=240,
                            value=int(cs["period_months"]),
                            step=12,
                            key=f"cs_period_months_{set_id}",
                            on_change=_set_active_coeff_set,
                            args=(set_id,),
                            help="По умолчанию 24 месяца. Можно увеличить период до 36, 60, 84 месяцев и больше.",
                        )
                    )
                period_years = cs["period_months"] / 12.0
                st.caption(
                    f"Текущий период: {int(cs['period_months'])} мес. ({period_years:.0f} г.)"
                    if float(period_years).is_integer()
                    else f"Текущий период: {int(cs['period_months'])} мес. ({period_years:.1f} г.)"
                )

                # ---- 4. Интерфейс в зависимости от типа набора ----

                if cs["type"] == "Спрос (по запросам)":
                    # ===== РЕЖИМ СПРОС (ПО ЗАПРОСАМ) =====
                    st.markdown("**Поисковые запросы:**")

                    if not cs.get("queries"):
                        cs["queries"] = ["Запрос 1"]

                    to_remove = []
                    for q_idx, q in enumerate(cs["queries"]):
                        query_row, _query_spacer = st.columns([1.6, 4.4])
                        with query_row:
                            col_q, col_del = st.columns([5, 1])
                            with col_q:
                                new_q = st.text_input(
                                    f"Запрос {q_idx + 1}:",
                                    value=q,
                                    key=f"cs_q_{set_id}_{q_idx}",
                                    label_visibility="collapsed",
                                    on_change=_set_active_coeff_set,
                                    args=(set_id,),
                                )
                                cs["queries"][q_idx] = new_q
                            with col_del:
                                if st.button("🗑️", key=f"cs_del_{set_id}_{q_idx}"):
                                    _set_active_coeff_set(set_id)
                                    to_remove.append(q_idx)

                    for r in sorted(to_remove, reverse=True):
                        cs["queries"].pop(r)
                        st.rerun()

                    if st.button("➕ Добавить запрос", key=f"cs_add_q_{set_id}"):
                        _set_active_coeff_set(set_id)
                        cs["queries"].append("")
                        st.rerun()

                    # Генерация таблицы на выбранный период
                    if st.button("Применить настройки набора", key=f"cs_apply_{set_id}", type="primary"):
                        _set_active_coeff_set(set_id)
                        queries_clean = [q.strip() for q in cs["queries"] if q.strip()]
                        if not queries_clean:
                            st.error("Добавьте хотя бы один поисковый запрос.")
                        else:
                            period_months = int(cs.get("period_months", 24) or 24)
                            months_period = generate_months_list(cs["start_month"], cs["start_year"], period_months)
                            df_cs = pd.DataFrame(months_period)
                            for q in queries_clean:
                                df_cs[q] = 0
                            cs["df_data"] = df_cs
                            st.success("Настроен период и запросы, заполните таблицу ниже.")
                            st.rerun()

                    # Редактор данных и расчёт
                    if cs.get("df_data") is not None:
                        demand_left, demand_right = st.columns([1.65, 1.0], vertical_alignment="top")
                        with demand_left:
                            st.markdown("**Заполните данные (можно копировать из Excel через Ctrl+C → Ctrl+V):**")
                            df_edit = st.data_editor(
                                cs["df_data"],
                                width="stretch",
                                row_height=ui_editor_row_height,
                                num_rows="fixed",
                                column_config={
                                    "period": st.column_config.TextColumn("Период", disabled=True),
                                    "month_num": st.column_config.NumberColumn("Месяц №", disabled=True),
                                    "year": st.column_config.NumberColumn("Год", disabled=True),
                                },
                                key=f"cs_editor_{set_id}",
                            )
                            cs["df_data"] = df_edit

                            if st.button("Рассчитать коэффициенты для этого набора", key=f"cs_calc_{set_id}", type="primary"):
                                _set_active_coeff_set(set_id)
                                queries_clean = [q.strip() for q in cs["queries"] if q.strip()]
                                if not queries_clean:
                                    st.error("Добавьте хотя бы один поисковый запрос.")
                                else:
                                    df_coeffs = calculate_seasonality_coeffs_demand(df_edit, queries_clean)
                                    cs["result"] = df_coeffs
                                    st.success("Коэффициенты рассчитаны!")
                        with demand_right:
                            st.markdown("**Итоговые коэффициенты спроса**")
                            if cs.get("result") is not None:
                                st.dataframe(cs["result"], width="stretch", height=420)
                            else:
                                st.info("После расчёта здесь появится итоговая таблица со средневзвешенными значениями и коэффициентами по месяцам.")

                elif cs["type"] == "AOV (средний чек)":
                    # ===== РЕЖИМ AOV (СРЕДНИЙ ЧЕК) =====
                    st.markdown("**Данные по среднему чеку (AOV):**")
                    st.caption(
                        f"Для каждого месяца выбранного периода укажите средний чек (AOV). Сейчас: {int(cs.get('period_months', 24) or 24)} мес."
                    )

                    # Генерация таблицы на выбранный период (только AOV)
                    if st.button("Применить настройки набора (AOV)", key=f"cs_apply_aov_{set_id}", type="primary"):
                        _set_active_coeff_set(set_id)
                        period_months = int(cs.get("period_months", 24) or 24)
                        months_period = generate_months_list(cs["start_month"], cs["start_year"], period_months)
                        df_cs = pd.DataFrame(months_period)
                        df_cs["AOV"] = 0.0
                        cs["df_data"] = df_cs
                        st.success("Настроен период. Заполните AOV по месяцам ниже.")
                        st.rerun()

                    # Редактор данных и расчёт для AOV
                    if cs.get("df_data") is not None:
                        aov_left, aov_right = st.columns([1.65, 1.0], vertical_alignment="top")
                        with aov_left:
                            st.markdown("**Заполните AOV (средний чек) по месяцам:**")
                            df_edit = st.data_editor(
                                cs["df_data"],
                                width="stretch",
                                row_height=ui_editor_row_height,
                                num_rows="fixed",
                                column_config={
                                    "period": st.column_config.TextColumn("Период", disabled=True),
                                    "month_num": st.column_config.NumberColumn("Месяц №", disabled=True),
                                    "year": st.column_config.NumberColumn("Год", disabled=True),
                                    "AOV": st.column_config.NumberColumn(
                                        "AOV (средний чек)", format="%.2f", help=mhelp("aov")
                                    ),
                                },
                                key=f"cs_editor_aov_{set_id}",
                            )
                            cs["df_data"] = df_edit

                        if st.button("Рассчитать коэффициенты AOV для этого набора", key=f"cs_calc_aov_{set_id}", type="primary"):
                            _set_active_coeff_set(set_id)
                            df_coeffs = calculate_seasonality_coeffs_aov(df_edit)
                            cs["result"] = df_coeffs
                            st.success("Коэффициенты AOV рассчитаны!")
                        with aov_right:
                            st.markdown("**Итоговые коэффициенты AOV**")
                            if cs.get("result") is not None:
                                st.dataframe(cs["result"], width="stretch", height=420)
                            else:
                                st.info("После расчета здесь появится итоговая таблица коэффициентов.")
                elif cs["type"] == "Кастомный набор":
                    # ===== РЕЖИМ КАСТОМНЫЙ НАБОР =====
                    st.markdown("**Кастомные коэффициенты на 12 месяцев:**")
                    st.caption(
                        "Заполните коэффициенты вручную. Этот набор можно использовать и для спроса, и для AOV."
                    )

                    month_names_map = {
                        1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
                        5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
                        9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
                    }

                    if (
                        cs.get("df_data") is None
                        or "month_num" not in cs["df_data"].columns
                        or "Коэф." not in cs["df_data"].columns
                    ):
                        cs["df_data"] = pd.DataFrame(
                            {
                                "month_num": list(range(1, 13)),
                                "Месяц": [month_names_map[m] for m in range(1, 13)],
                                "Коэф.": [1.0] * 12,
                            }
                        )
                    elif "Коэф. AOV" in cs["df_data"].columns:
                        custom_df = cs["df_data"].copy()
                        custom_df = custom_df.drop(columns=["Коэф. AOV"], errors="ignore")
                        cs["df_data"] = custom_df

                    custom_left, custom_right = st.columns([1.65, 1.0], vertical_alignment="top")
                    with custom_left:
                        df_edit = st.data_editor(
                            cs["df_data"],
                            width="stretch",
                            row_height=ui_editor_row_height,
                            num_rows="fixed",
                            column_config={
                                "month_num": st.column_config.NumberColumn("Месяц №", disabled=True),
                                "Месяц": st.column_config.TextColumn("Месяц", disabled=True),
                                "Коэф.": st.column_config.NumberColumn(
                                    "Коэффициент", format="%.2f",
                                    help="Коэффициент месяца: 1.00 = без изменений, 1.20 = +20%, 0.80 = -20%."
                                ),
                            },
                            key=f"cs_editor_custom_{set_id}",
                        )
                        cs["df_data"] = df_edit
                    with custom_right:
                        st.markdown("**Действия**")
                        if st.button("Применить кастомный набор", key=f"cs_apply_custom_{set_id}", type="primary"):
                            _set_active_coeff_set(set_id)
                            df_coeffs = df_edit.copy()
                            if "Коэф." not in df_coeffs.columns:
                                df_coeffs["Коэф."] = 1.0
                            df_coeffs["Номер месяца"] = df_coeffs["month_num"].astype(int)
                            df_coeffs = df_coeffs[["Номер месяца", "Месяц", "Коэф."]]
                            cs["result"] = df_coeffs
                            st.success("Кастомные коэффициенты сохранены!")
                        else:
                            st.caption("Вносите значения слева и применяйте набор кнопкой выше.")
                else:
                    # ===== РЕЖИМ МЕДИЙНЫЕ ХВОСТЫ =====
                    st.markdown("**Медийные хвосты (коэффициенты показов на 12 месяцев):**")
                    st.caption(
                        "Этот набор накладывается на спрос и влияет только на показы "
                        "(k_imp = k_demand × k_media_tail)."
                    )

                    month_names_map = {
                        1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
                        5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
                        9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
                    }

                    if (
                        cs.get("df_data") is None
                        or "month_num" not in cs["df_data"].columns
                        or "Коэф." not in cs["df_data"].columns
                    ):
                        cs["df_data"] = pd.DataFrame(
                            {
                                "month_num": list(range(1, 13)),
                                "Месяц": [month_names_map[m] for m in range(1, 13)],
                                "Коэф.": [1.0] * 12,
                            }
                        )

                    media_left, media_right = st.columns([1.65, 1.0], vertical_alignment="top")
                    with media_left:
                        df_edit = st.data_editor(
                            cs["df_data"],
                            width="stretch",
                            row_height=ui_editor_row_height,
                            num_rows="fixed",
                            column_config={
                                "month_num": st.column_config.NumberColumn("Месяц №", disabled=True),
                                "Месяц": st.column_config.TextColumn("Месяц", disabled=True),
                                "Коэф.": st.column_config.NumberColumn(
                                    "Коэффициент хвостов", format="%.2f",
                                    help="Медийный хвост для показов: k_imp = k_demand × k_media_tail."
                                ),
                            },
                            key=f"cs_editor_media_tail_{set_id}",
                        )
                        cs["df_data"] = df_edit
                    with media_right:
                        st.markdown("**Действия**")
                        if st.button("Применить набор медийных хвостов", key=f"cs_apply_media_tail_{set_id}", type="primary"):
                            _set_active_coeff_set(set_id)
                            df_coeffs = df_edit.copy()
                            if "Коэф." not in df_coeffs.columns:
                                df_coeffs["Коэф."] = 1.0
                            df_coeffs["Номер месяца"] = df_coeffs["month_num"].astype(int)
                            df_coeffs = df_coeffs[["Номер месяца", "Месяц", "Коэф."]]
                            cs["result"] = df_coeffs
                            st.success("Набор медийных хвостов сохранен!")
                        else:
                            st.caption("Набор влияет только на показы и накладывается поверх спроса.")
                    cs_type_norm = normalize_coeff_set_type(cs.get("type"))
                    if cs_type_norm == "Спрос (по запросам)":
                        st.markdown("### Итоговые коэффициенты")
                        st.dataframe(cs["result"], width="stretch")

                    buf = io.BytesIO()
                    sheet_name = (cs["name"] or "Коэффициенты")[:31]
                    excel_engine = "xlsxwriter" if HAS_XLSXWRITER else "openpyxl"
                    with pd.ExcelWriter(buf, engine=excel_engine) as writer:
                        cs["result"].to_excel(writer, sheet_name=sheet_name, index=False)
                    buf.seek(0)

                    st.download_button(
                        label="📥 Скачать коэффициенты этого набора",
                        data=buf,
                        file_name=f"coeffs_set_{set_id}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"cs_dl_{set_id}",
                    )
    render_bottom_tab_switcher("Коэффициенты", "coeffs")
# ====================================================================
#                      ТАБ «НАСТРОЙКИ РАСЧЕТА»
# ====================================================================

with tab_setup:
    st.markdown(
        """
        <div class="tab-intro">
            <p>1) Выберите месяцы, для которых нужно рассчитать медиаплан.</p>
            <p>2) Выберите пресет метрик <b>E-com</b>, <b>DIY</b> или <b>Недвижимость</b>. Пресет влияет на состав метрик и логику отображения отдельных показателей в расчете.</p>
            <p>3) Заполните данные по типам рекламных кампаний и базовым метрикам для расчета среднего месяца. Значения в блок «Средний месяц» вносятся <b><span style="color:#9EC5FF;">без НДС</span></b>. Помните: сезонность среднего месяца равна 1, и от нее рассчитываются все выбранные месяцы через коэффициенты.</p>
            <p>4) При необходимости настройте учет <b>НДС</b> и <b>АК</b>: можно включить НДС, задать фиксированную АК на месяц или использовать шкалу АК от TOTAL бюджета месяца без НДС.</p>
            <p>5) Назначьте для каждого типа РК наборы коэффициентов: <b>Спрос</b>, <b>AOV</b>, <b>Кастомный набор</b> и при необходимости <b>Медийные хвосты</b>. Для пресета <b>Недвижимость</b> набор <b>AOV</b> не используется.</p>
            <p>6) Настройте эластичность метрик к сезонности спроса: можно выбрать пресет настроек или задать значения вручную для <b>CPC</b>, <b>CTR</b> и <b>CR</b>.</p>
            <p>7) Проверьте базовый «Средний месяц»: он служит опорой для дальнейшей помесячной работы.</p>
            <p>8) После проверки базы переходите во вкладку «Медиаплан» для работы по месяцам и сверки TOTAL.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    if "use_vat_budget_metrics" not in st.session_state:
        st.session_state["use_vat_budget_metrics"] = True
    if "use_ak_budget_metrics" not in st.session_state:
        st.session_state["use_ak_budget_metrics"] = False
    if "ak_mode" not in st.session_state:
        st.session_state["ak_mode"] = "percent"
    if "ak_fixed_month_wo_vat" not in st.session_state:
        st.session_state["ak_fixed_month_wo_vat"] = 200000.0
    if "ak_fixed_percent" not in st.session_state:
        st.session_state["ak_fixed_percent"] = 2.0
    if "ak_rules_editor_nonce" not in st.session_state:
        st.session_state["ak_rules_editor_nonce"] = 0
    if "elasticity_editor_nonce" not in st.session_state:
        st.session_state["elasticity_editor_nonce"] = 0
    default_ak_rules = pd.DataFrame(
        [
            {"min_budget_wo_vat": 0.0, "ak_percent": 0.0},
            {"min_budget_wo_vat": 1500000.0, "ak_percent": 8.0},
            {"min_budget_wo_vat": 3000000.0, "ak_percent": 4.0},
            {"min_budget_wo_vat": 6000000.0, "ak_percent": 2.0},
            {"min_budget_wo_vat": 10000000.0, "ak_percent": 0.0},
        ]
    )
    if "ak_rules_df" not in st.session_state:
        st.session_state["ak_rules_df"] = default_ak_rules.copy()
    use_vat_budget_metrics = bool(st.session_state.get("use_vat_budget_metrics", True))
    use_ak_budget_metrics = bool(st.session_state.get("use_ak_budget_metrics", False))
    ak_mode = str(st.session_state.get("ak_mode", "fixed"))
    ak_fixed_month_wo_vat = float(st.session_state.get("ak_fixed_month_wo_vat", 200000.0))
    ak_fixed_percent = float(st.session_state.get("ak_fixed_percent", 2.0))
    ak_rules_df = st.session_state.get("ak_rules_df", pd.DataFrame())
    ak_rules_invalid = (
        not isinstance(ak_rules_df, pd.DataFrame)
        or ak_rules_df.empty
        or "min_budget_wo_vat" not in ak_rules_df.columns
        or "ak_percent" not in ak_rules_df.columns
    )
    if not ak_rules_invalid:
        _min_col = pd.to_numeric(ak_rules_df["min_budget_wo_vat"], errors="coerce")
        _pct_col = pd.to_numeric(ak_rules_df["ak_percent"], errors="coerce")
        ak_rules_invalid = _min_col.isna().all() or _pct_col.isna().all()
    if ak_rules_invalid:
        st.session_state["ak_rules_df"] = default_ak_rules.copy()
        ak_rules_df = st.session_state["ak_rules_df"]
        st.session_state["ak_rules_editor_nonce"] = int(st.session_state.get("ak_rules_editor_nonce", 0)) + 1

    month_names_full = MONTH_NAMES_FULL.copy()

    def _fmt_ref_int(v: float) -> str:
        return f"{int(round(v)):,}".replace(",", " ")

    def _fmt_ref_rub(v: float) -> str:
        return f"{int(round(v)):,} ₽".replace(",", " ")

    def _fmt_ref_rub2(v: float) -> str:
        return f"{float(v):.2f} ₽"

    def _fmt_ref_pct(v: float) -> str:
        return f"{float(v):.2f}%"

    def _fmt_ref_roas(v: float) -> str:
        return f"{float(v):.2f}"

    def _build_ref_df(ref: dict) -> pd.DataFrame:
        if not ref:
            return pd.DataFrame(columns=["Показатель", "Значение"])
        if is_real_estate_preset:
            rows = [
                {"Показатель": "Показы", "Значение": _fmt_ref_int(ref.get("impressions", 0.0))},
                {"Показатель": "Клики", "Значение": _fmt_ref_int(ref.get("clicks", 0.0))},
                {"Показатель": "Бюджет", "Значение": _fmt_ref_rub(ref.get("cost", 0.0))},
                {"Показатель": "Бюджет с НДС", "Значение": _fmt_ref_rub(ref.get("cost_with_vat", 0.0))},
                {"Показатель": "Бюджет с НДС и АК", "Значение": _fmt_ref_rub(ref.get("cost_with_vat_ak", 0.0))},
                {"Показатель": "CTR", "Значение": _fmt_ref_pct(ref.get("ctr", 0.0))},
                {"Показатель": "CPC", "Значение": _fmt_ref_rub2(ref.get("cpc", 0.0))},
            ]
            if metric_mode["is_real_estate_full"]:
                rows += [
                    {"Показатель": "CR1 в Лид", "Значение": _fmt_ref_pct(ref.get("cr1", 0.0))},
                    {"Показатель": "Лиды", "Значение": _fmt_ref_int(ref.get("leads", 0.0))},
                    {"Показатель": "CPL", "Значение": _fmt_ref_rub(ref.get("cpl", 0.0))},
                    {"Показатель": "CR2 в ЦО", "Значение": _fmt_ref_pct(ref.get("cr2", 0.0))},
                    {"Показатель": "ЦО", "Значение": _fmt_ref_int(ref.get("target_leads", ref.get("conversions", 0.0)))},
                    {"Показатель": "CPQL", "Значение": _fmt_ref_rub(ref.get("cpql", 0.0))},
                ]
            else:
                rows += [
                    {"Показатель": "CR в ЦО", "Значение": _fmt_ref_pct(ref.get("cr", 0.0))},
                    {"Показатель": "ЦО", "Значение": _fmt_ref_int(ref.get("target_leads", ref.get("conversions", 0.0)))},
                    {"Показатель": "CPQL", "Значение": _fmt_ref_rub(ref.get("cpql", ref.get("cpo", 0.0)))},
                ]
        else:
            rows = [
                {"Показатель": "Показы", "Значение": _fmt_ref_int(ref.get("impressions", 0.0))},
                {"Показатель": "Клики", "Значение": _fmt_ref_int(ref.get("clicks", 0.0))},
                {"Показатель": "Конверсии", "Значение": _fmt_ref_int(ref.get("conversions", 0.0))},
                {"Показатель": "Бюджет", "Значение": _fmt_ref_rub(ref.get("cost", 0.0))},
                {"Показатель": "Бюджет с НДС", "Значение": _fmt_ref_rub(ref.get("cost_with_vat", 0.0))},
                {"Показатель": "Бюджет с НДС и АК", "Значение": _fmt_ref_rub(ref.get("cost_with_vat_ak", 0.0))},
                {"Показатель": "Доход", "Значение": _fmt_ref_rub(ref.get("revenue", 0.0))},
                {"Показатель": "CTR", "Значение": _fmt_ref_pct(ref.get("ctr", 0.0))},
                {"Показатель": "CPC", "Значение": _fmt_ref_rub2(ref.get("cpc", 0.0))},
                {"Показатель": "CR", "Значение": _fmt_ref_pct(ref.get("cr", 0.0))},
                {"Показатель": "CPO", "Значение": _fmt_ref_rub(ref.get("cpo", 0.0))},
                {"Показатель": "ROAS", "Значение": _fmt_ref_roas(ref.get("roas", 0.0))},
                {"Показатель": "ДРР", "Значение": _fmt_ref_pct(ref.get("drr", 0.0))},
            ]
        return pd.DataFrame(rows)

    # ---------- 0. Горизонт планирования ----------

    ui_section_title("0. Горизонт планирования")
    st.caption(
        "Выбранные месяцы напрямую влияют на помесячный расчет, итоги в блоке TOTAL "
        "и данные на вкладке «Диаграммы»."
    )

    if "planning_start_month" not in st.session_state:
        st.session_state["planning_start_month"] = 1
    if "planning_start_year" not in st.session_state:
        st.session_state["planning_start_year"] = dt.date.today().year
    period_col_month, period_col_year, _period_sp = st.columns([1.2, 1.0, 2.8], vertical_alignment="bottom")
    with period_col_month:
        planning_start_month = int(
            st.selectbox(
                "Стартовый месяц периода:",
                options=list(month_names_full.keys()),
                format_func=lambda m: month_names_full.get(int(m), str(m)),
                key="planning_start_month",
            )
        )
    with period_col_year:
        planning_start_year = int(
            st.number_input(
                "Стартовый год:",
                min_value=2000,
                max_value=2100,
                step=1,
                key="planning_start_year",
            )
        )
    planning_periods = build_planning_periods(planning_start_month, planning_start_year, horizon_months=12)
    all_month_labels = [p["period_label"] for p in planning_periods]
    sync_multiselect_state("planning_months_multiselect", all_month_labels, default_to_all=True)

    selected_month_labels = st.multiselect(
        "Выберите месяцы, для которых считать медиаплан:",
        options=all_month_labels,
        key="planning_months_multiselect",
    )

    plan_calc_blockers = []
    if not selected_month_labels:
        st.warning("Выберите хотя бы один месяц для расчёта медиаплана.")
        plan_calc_blockers.append("не выбраны месяцы планирования")

    planning_periods_by_label = {p["period_label"]: p for p in planning_periods}
    selected_periods = [planning_periods_by_label[label] for label in selected_month_labels if label in planning_periods_by_label]
    selected_periods_by_slot = {int(p["planning_slot"]): p for p in selected_periods}
    selected_month_nums = [int(p["month_num"]) for p in selected_periods]
    selected_planning_slots = [int(p["planning_slot"]) for p in selected_periods]
    months_count = len(selected_periods)

    ui_section_title("0.1 Пресет метрик")
    st.caption(
        "После выбора горизонта планирования выберите пресет метрик. "
        "Доступны пресеты: E-com, DIY и Недвижимость."
    )
    preset_keys = list(METRIC_PRESETS.keys())
    if "metric_preset_key" not in st.session_state or st.session_state["metric_preset_key"] not in preset_keys:
        st.session_state["metric_preset_key"] = "ecom"
    preset_col, _preset_spacer = st.columns([1, 3])
    with preset_col:
        active_preset_key = st.selectbox(
            "Выберите пресет метрик:",
            options=preset_keys,
            index=preset_keys.index(st.session_state["metric_preset_key"]),
            format_func=lambda k: METRIC_PRESETS[k]["label"],
            key="metric_preset_key",
        )
    st.caption(f"Активный пресет: {METRIC_PRESETS[active_preset_key]['label']}")
    if "real_estate_funnel_mode" not in st.session_state:
        st.session_state["real_estate_funnel_mode"] = "simple"
    metric_mode = get_metric_mode(active_preset_key, st.session_state.get("real_estate_funnel_mode", "simple"))
    is_diy_preset = metric_mode["is_diy"]
    is_real_estate_preset = metric_mode["is_real_estate"]
    if is_real_estate_preset:
        re_col, _re_spacer = st.columns([1.2, 2.8])
        with re_col:
            selected_re_mode = st.selectbox(
                "Режим воронки:",
                options=list(REAL_ESTATE_FUNNEL_OPTIONS.keys()),
                index=list(REAL_ESTATE_FUNNEL_OPTIONS.keys()).index(metric_mode["real_estate_funnel_mode"]),
                format_func=lambda k: REAL_ESTATE_FUNNEL_OPTIONS[k],
                key="real_estate_funnel_mode",
            )
        metric_mode = get_metric_mode(active_preset_key, selected_re_mode)
        DISPLAY_COL_RENAME["conversions"] = "ЦО"
        DISPLAY_COL_RENAME["target_leads"] = "ЦО"
        DISPLAY_COL_RENAME["revenue"] = "Выручка"
        if metric_mode["is_real_estate_full"]:
            DISPLAY_COL_RENAME["cr"] = "CR в ЦО"
            DISPLAY_COL_RENAME["cr_pct"] = "CR в ЦО"
            DISPLAY_COL_RENAME["cpa"] = "CPQL"
        else:
            DISPLAY_COL_RENAME["cr"] = "CR в ЦО"
            DISPLAY_COL_RENAME["cr_pct"] = "CR в ЦО"
            DISPLAY_COL_RENAME["cpa"] = "CPQL"
    elif is_diy_preset:
        DISPLAY_COL_RENAME["conversions"] = "Кол-во оформленных заказов"
        DISPLAY_COL_RENAME["cr"] = "CR1"
        DISPLAY_COL_RENAME["cr_pct"] = "CR1"
        DISPLAY_COL_RENAME["revenue"] = "Оформленный доход, ₽ с НДС"
        DISPLAY_COL_RENAME["aov"] = "Средний чек по оформл. доходу, с НДС"
        DISPLAY_COL_RENAME["cpa"] = "Стоимость оформл. заказа, ₽ с НДС"
        DISPLAY_COL_RENAME["ROAS"] = "ROAS оформл."
        DISPLAY_COL_RENAME["drr"] = "ДРР от оформл. дохода, %"
        DISPLAY_COL_RENAME["drr_pct"] = "ДРР от оформл. дохода, %"
        DISPLAY_COL_RENAME["cr2_pct"] = "CR2"
        DISPLAY_COL_RENAME["target_leads"] = "Кол-во отгруженных заказов"
        DISPLAY_COL_RENAME["cpql"] = "Стоимость отгр. заказа, ₽ с НДС"
    else:
        DISPLAY_COL_RENAME["conversions"] = "Конверсии"
        DISPLAY_COL_RENAME["cr"] = "CR"
        DISPLAY_COL_RENAME["cr_pct"] = "CR"
        DISPLAY_COL_RENAME["revenue"] = "Доход"
        DISPLAY_COL_RENAME["cpa"] = "CPO"

    def _df_for_compare(df: pd.DataFrame) -> pd.DataFrame:
        if df is None:
            return pd.DataFrame()
        tmp = df.copy().reset_index(drop=True)
        tmp = tmp.reindex(sorted(tmp.columns), axis=1)
        return tmp.fillna("")

    def _has_unsaved_changes(draft_df: pd.DataFrame, saved_df: pd.DataFrame) -> bool:
        return not _df_for_compare(draft_df).equals(_df_for_compare(saved_df))

    def _table_height_for_rows(rows_count: int, min_height: int = 120, max_height: int = 560) -> int:
        header_px = 34 if ui_compact_tables else 38
        row_px = ui_table_row_px
        h = header_px + max(rows_count, 1) * row_px + 6
        return max(min_height, min(h, max_height))

    # ---------- 1. Типы РК (средний месяц) ----------

    ui_section_title("1. Типы рекламных кампаний в среднем месяце")
    with st.expander("Типы рекламных кампаний в среднем месяце", expanded=True):

        default_campaigns = pd.DataFrame(
            [
                ["Поиск бренд", "B2C", "Яндекс", "Поиск", "",        500_000, 5.0, 15.0, 5.0, 50.0, 5000.0, 5000.0, 220_000, 0.0, 0.0, 0.0, 0.0, 0.0],
                ["РСЯ баннеры", "B2C", "Яндекс", "РСЯ баннеры", "", 1_000_000, 1.0, 10.0, 2.0, 40.0, 3000.0, 3000.0, 450_000, 0.0, 0.0, 0.0, 0.0, 0.0],
                ["Видео YouTube", "B2B", "YouTube", "Видео", "",     300_000, 0.7, 20.0, 1.5, 35.0, 4000.0, 4000.0, 180_000, 0.0, 0.0, 0.0, 0.0, 0.0],
            ],
            columns=[
                "campaign_type", "segment", "system", "format", "geo",
                "impressions_avg", "ctr_avg_percent", "cpc_avg", "cr_avg_percent", "cr2_avg_percent", "aov_avg", "shipped_aov_avg", "reach_avg",
                "available_capacity_avg", "client_count_avg", "absolute_new_clients_avg", "returned_clients_avg", "order_frequency_avg",
            ],
        )

        if "campaigns_df" not in st.session_state:
            st.session_state["campaigns_df"] = default_campaigns.copy()
        if "segment" not in st.session_state["campaigns_df"].columns:
            st.session_state["campaigns_df"]["segment"] = "B2C"
        if "geo" not in st.session_state["campaigns_df"].columns:
            st.session_state["campaigns_df"]["geo"] = ""
        if "available_capacity_avg" not in st.session_state["campaigns_df"].columns:
            st.session_state["campaigns_df"]["available_capacity_avg"] = 0.0
        if "client_count_avg" not in st.session_state["campaigns_df"].columns:
            st.session_state["campaigns_df"]["client_count_avg"] = 0.0
        if "absolute_new_clients_avg" not in st.session_state["campaigns_df"].columns:
            st.session_state["campaigns_df"]["absolute_new_clients_avg"] = 0.0
        if "returned_clients_avg" not in st.session_state["campaigns_df"].columns:
            st.session_state["campaigns_df"]["returned_clients_avg"] = 0.0
        if "order_frequency_avg" not in st.session_state["campaigns_df"].columns:
            st.session_state["campaigns_df"]["order_frequency_avg"] = 0.0
        if "shipped_aov_avg" not in st.session_state["campaigns_df"].columns:
            st.session_state["campaigns_df"]["shipped_aov_avg"] = 0.0
        if "reach_avg" not in st.session_state["campaigns_df"].columns:
            st.session_state["campaigns_df"]["reach_avg"] = 0.0
        if "cr2_avg_percent" not in st.session_state["campaigns_df"].columns:
            st.session_state["campaigns_df"]["cr2_avg_percent"] = 0.0

        campaign_cols_order = [
            "campaign_type",
            "system",
            "format",
            "geo",
            "segment",
            "impressions_avg",
            "ctr_avg_percent",
            "cpc_avg",
            "cr_avg_percent",
            "cr2_avg_percent",
            "aov_avg",
            "shipped_aov_avg",
            "reach_avg",
            "available_capacity_avg",
            "client_count_avg",
            "absolute_new_clients_avg",
            "returned_clients_avg",
            "order_frequency_avg",
        ]
        st.session_state["campaigns_df"] = st.session_state["campaigns_df"].reindex(
            columns=[c for c in campaign_cols_order if c in st.session_state["campaigns_df"].columns]
        )

        campaigns_column_config = {
            "campaign_type": st.column_config.TextColumn("Название кампании"),
            "system": st.column_config.TextColumn("Рекламная система"),
            "format": st.column_config.TextColumn("Формат/таргетинги"),
            "geo": st.column_config.TextColumn("ГЕО"),
            "impressions_avg": st.column_config.NumberColumn(
                "Показы (средний месяц)", format="%.0f", help=mhelp("impressions")
            ),
            "ctr_avg_percent": st.column_config.NumberColumn(
                "CTR, % (средний месяц)", format="%.2f", help=mhelp("ctr")
            ),
            "cpc_avg": st.column_config.NumberColumn(
                "CPC, ₽ (средний месяц)", format="%.2f", help=mhelp("cpc")
            ),
            "cr_avg_percent": st.column_config.NumberColumn("CR1, % (средний месяц)" if is_diy_preset else "CR в ЦО, % (средний месяц)" if metric_mode["is_real_estate_simple"] else "CR1 в Лид, % (средний месяц)" if metric_mode["is_real_estate_full"] else "CR, % (средний месяц)", format="%.2f", help=mhelp("cr")),
            "cr2_avg_percent": st.column_config.NumberColumn("CR2, % (средний месяц)", format="%.2f", help="Отгруженные заказы = оформленные заказы × CR2."),
            "reach_avg": st.column_config.NumberColumn(
                "Охват (средний месяц)", format="%.0f", help=mhelp("reach")
            ),
        }
        campaigns_editor_columns = [
            "campaign_type",
            "system",
            "format",
            "impressions_avg",
            "ctr_avg_percent",
            "cpc_avg",
            "cr_avg_percent",
        ]
        if is_real_estate_preset or is_diy_preset:
            campaigns_editor_columns.insert(3, "geo")
        if metric_mode["is_real_estate_full"] or is_diy_preset:
            campaigns_editor_columns.append("cr2_avg_percent")
        elif metric_mode["needs_aov"]:
            campaigns_editor_columns.append("aov_avg")
        if is_diy_preset:
            campaigns_column_config["aov_avg"] = st.column_config.NumberColumn(
                "Средний чек по оформл. доходу, с НДС (база)", format="%.2f", help=mhelp("aov")
            )
            campaigns_column_config["shipped_aov_avg"] = st.column_config.NumberColumn(
                "Средний чек по отгр. доходу, с НДС (база)", format="%.2f", help=mhelp("shipped_aov")
            )
            campaigns_editor_columns.append("aov_avg")
            campaigns_editor_columns.append("shipped_aov_avg")
        if is_diy_preset:
            campaigns_column_config["segment"] = st.column_config.SelectboxColumn("Сегмент", options=["B2B", "B2C"])
            campaigns_editor_columns.insert(4, "segment")
            campaigns_column_config["available_capacity_avg"] = st.column_config.NumberColumn(
                "Доступная емкость (база)", format="%.0f"
            )
            campaigns_column_config["client_count_avg"] = st.column_config.NumberColumn(
                "Количество клиентов (база)", format="%.0f", help=mhelp("client_count")
            )
            campaigns_column_config["absolute_new_clients_avg"] = st.column_config.NumberColumn(
                "Кол-во абсолютно новых клиентов (база)", format="%.0f", help=mhelp("absolute_new_clients")
            )
            campaigns_column_config["returned_clients_avg"] = st.column_config.NumberColumn(
                "Кол-во вернувшихся клиентов (база)", format="%.0f", help=mhelp("returned_clients")
            )
            campaigns_column_config["order_frequency_avg"] = st.column_config.NumberColumn(
                "Частота заказа (база)", format="%.2f", help=mhelp("order_frequency")
            )
            campaigns_editor_columns += ["available_capacity_avg", "client_count_avg", "absolute_new_clients_avg", "returned_clients_avg", "order_frequency_avg"]

        with st.form("campaigns_editor_form"):
            campaigns_draft = st.data_editor(
                st.session_state["campaigns_df"],
                num_rows="dynamic",
                width="stretch",
                row_height=ui_editor_row_height,
                column_order=campaigns_editor_columns,
                column_config=campaigns_column_config,
                key="campaigns_editor",
            )
            if _has_unsaved_changes(campaigns_draft, st.session_state["campaigns_df"]):
                st.caption("Есть несохраненные изменения в таблице РК.")
            camp_btn_c1, camp_btn_c2 = st.columns([1.2, 2.8])
            with camp_btn_c1:
                campaigns_saved = st.form_submit_button("Сохранить таблицу РК", type="primary")
            with camp_btn_c2:
                campaign_row_added = st.form_submit_button("➕ Добавить тип РК")

        if campaigns_saved:
            st.session_state["campaigns_df"] = campaigns_draft.copy()

        if campaign_row_added:
            new_row = {
                "campaign_type": "",
                "segment": "B2C",
                "system": "",
                "format": "",
                "geo": "",
                "impressions_avg": 0,
                "ctr_avg_percent": 0.0,
                "cpc_avg": 0.0,
                "cr_avg_percent": 0.0,
                "cr2_avg_percent": 0.0,
                "aov_avg": 0.0,
                "shipped_aov_avg": 0.0,
                "reach_avg": 0.0,
                "available_capacity_avg": 0.0,
                "client_count_avg": 0.0,
                "absolute_new_clients_avg": 0.0,
                "returned_clients_avg": 0.0,
                "order_frequency_avg": 0.0,
            }
            st.session_state["campaigns_df"] = pd.concat(
                [campaigns_draft, pd.DataFrame([new_row])],
                ignore_index=True,
            )
            st.rerun()

        campaigns = st.session_state["campaigns_df"].copy()
        campaigns = campaigns.dropna(subset=["campaign_type"])
        campaigns = campaigns[campaigns["campaign_type"].astype(str).str.strip() != ""]

        if campaigns.empty:
            st.warning("Добавьте хотя бы один тип РК.")
            plan_calc_blockers.append("не заполнены типы РК")

        required_cols = get_campaign_required_cols(metric_mode)

        def row_has_missing(row):
            for col in required_cols:
                val = row.get(col)
                if val is None:
                    return True
                if isinstance(val, str) and val.strip() == "":
                    return True
                try:
                    if pd.isna(val):
                        return True
                except Exception:
                    pass
            return False

        missing_mask = campaigns.apply(row_has_missing, axis=1)

        if missing_mask.any():
            required_labels = ["Показы", "CTR", "CPC"]
            if metric_mode["is_real_estate_full"]:
                required_labels += ["CR1 в Лид", "CR2 в ЦО"]
            elif metric_mode["is_real_estate_simple"]:
                required_labels += ["CR в ЦО"]
            elif is_diy_preset:
                required_labels += ["CR1", "CR2", "Средний чек по оформл. доходу", "Средний чек по отгр. доходу"]
            else:
                required_labels += ["CR", "AOV"]
            st.error(
                "Заполните, пожалуйста, все поля в среднем месяце ("
                + ", ".join(required_labels)
                + ") "
                "для каждого типа РК. Проверь строки, где есть пустые ячейки."
            )
            plan_calc_blockers.append("не заполнены обязательные поля в таблице РК")

        segment_filter_options = ["Все", "B2B", "B2C"]
        if "plan_segment_filter" not in st.session_state:
            st.session_state["plan_segment_filter"] = "Все"
        if "plan_segment_filter_top" not in st.session_state:
            st.session_state["plan_segment_filter_top"] = st.session_state["plan_segment_filter"]
        if "plan_segment_filter_sidebar" not in st.session_state:
            st.session_state["plan_segment_filter_sidebar"] = st.session_state["plan_segment_filter"]

        def _sync_segment_from_top():
            val = str(st.session_state.get("plan_segment_filter_top", "Все"))
            if val not in segment_filter_options:
                val = "Все"
            st.session_state["plan_segment_filter"] = val
            st.session_state["plan_segment_filter_sidebar"] = val

        def _sync_segment_from_sidebar():
            val = str(st.session_state.get("plan_segment_filter_sidebar", "Все"))
            if val not in segment_filter_options:
                val = "Все"
            st.session_state["plan_segment_filter"] = val
            st.session_state["plan_segment_filter_top"] = val

        if is_diy_preset:
            st.radio(
                "Показывать кампании сегмента:",
                options=segment_filter_options,
                horizontal=True,
                key="plan_segment_filter_top",
                on_change=_sync_segment_from_top,
            )
            segment_filter = str(st.session_state.get("plan_segment_filter", "Все"))
            if segment_filter != "Все":
                campaigns = campaigns[campaigns["segment"].astype(str).str.upper() == segment_filter]
                if campaigns.empty:
                    st.warning(f"Для сегмента {segment_filter} нет кампаний.")
                    plan_calc_blockers.append(f"для сегмента {segment_filter} нет кампаний для расчёта")
            visible_segments = sorted(campaigns["segment"].astype(str).str.upper().unique().tolist())
            show_segment_subtotals = len(visible_segments) > 1
        else:
            st.session_state["plan_segment_filter"] = "Все"
            st.session_state["plan_segment_filter_top"] = "Все"
            st.session_state["plan_segment_filter_sidebar"] = "Все"
            show_segment_subtotals = False

        existing_ctypes = campaigns["campaign_type"].astype(str).unique().tolist()
        campaign_identity_rows = campaigns.copy()
        for identity_col, default_val in [("system", ""), ("format", ""), ("geo", ""), ("segment", "ALL")]:
            if identity_col not in campaign_identity_rows.columns:
                campaign_identity_rows[identity_col] = default_val
        campaign_identity_rows["campaign_type"] = campaign_identity_rows["campaign_type"].astype(str).str.strip()
        campaign_identity_rows["system"] = campaign_identity_rows["system"].astype(str).str.strip()
        campaign_identity_rows["format"] = campaign_identity_rows["format"].astype(str).str.strip()
        campaign_identity_rows["geo"] = campaign_identity_rows["geo"].astype(str).str.strip()
        campaign_identity_rows["segment"] = campaign_identity_rows["segment"].astype(str).str.strip().str.upper()
        campaign_identity_rows["campaign_key"] = campaign_identity_rows.apply(get_campaign_identity_key_from_row, axis=1)
        campaign_identity_rows = (
            campaign_identity_rows[
                ["campaign_key", "campaign_type", "system", "format", "geo", "segment"]
            ]
            .drop_duplicates(subset=["campaign_key"], keep="first")
            .reset_index(drop=True)
        )

    with st.expander("1.2 НДС и АК", expanded=True):
        vat_warning_text = (
            "<b>Учитывать НДС 22% в расчетах</b>: настройка влияет на CPC, CPL и CPQL."
            if is_real_estate_preset and metric_mode["is_real_estate_full"]
            else "<b>Учитывать НДС 22% в расчетах</b>: настройка влияет на CPC и CPQL."
            if is_real_estate_preset
            else "<b>Учитывать НДС 22% в расчетах</b>: настройка влияет на CPC, CPO, ROAS и ДРР."
        )
        st.markdown(
            f"""
            <div style="
                margin: 10px 0 8px 0;
                display: inline-block;
                width: fit-content;
                max-width: 100%;
                padding: 10px 12px;
                border-radius: 10px;
                border: 1px solid #FF8A66;
                background: rgba(255, 99, 51, 0.14);
                color: #FFD9CC;
                font-weight: 600;
            ">
                {vat_warning_text}
            </div>
            """,
            unsafe_allow_html=True,
        )
        use_vat_budget_metrics = st.checkbox(
                "Учитывать НДС 22% в расчетах",
            key="use_vat_budget_metrics",
        )
        use_ak_budget_metrics = st.checkbox(
            "Учитывать АК в расчетах",
            key="use_ak_budget_metrics",
        )
        vat_ak_ui_config = get_vat_ak_ui_config(metric_mode, is_real_estate_preset)
        vat_warning_text = vat_ak_ui_config["warning_text"]
        ak_mode_labels = vat_ak_ui_config["ak_mode_labels"]
        mode_label_by_value = vat_ak_ui_config["mode_label_by_value"]
        mode_value_by_label = vat_ak_ui_config["mode_value_by_label"]
        ak_mode_label = st.selectbox(
            "Способ расчета АК",
            options=ak_mode_labels,
            index=ak_mode_labels.index(mode_label_by_value.get(ak_mode, ak_mode_labels[0])),
            key="ak_mode_select",
        )
        ak_mode = mode_value_by_label.get(ak_mode_label, "percent")
        st.session_state["ak_mode"] = ak_mode
        if ak_mode == "fixed":
            ak_fixed_month_wo_vat = st.number_input(
                "АК в месяц (без НДС), ₽",
                min_value=0.0,
                step=10000.0,
                format="%.0f",
                key="ak_fixed_month_wo_vat",
            )
            st.caption("Фиксированная АК месяца распределяется между типами РК пропорционально их бюджету без НДС.")
        elif ak_mode == "fixed_percent":
            ak_fixed_percent = st.number_input(
                "АК, % для всех типов РК",
                min_value=0.0,
                step=0.5,
                format="%.2f",
                key="ak_fixed_percent",
            )
            st.caption("Один и тот же процент АК применяется ко всем типам РК и месяцам в расчёте.")
        if ak_mode == "percent":
            with st.expander("Настройка АК по шкале (TOTAL бюджета месяца без НДС)", expanded=use_ak_budget_metrics):
                st.caption(
                    "Для режима «Процент по шкале»: АК выбирается по наибольшему подходящему порогу TOTAL бюджета месяца без НДС."
                )
                ak_editor_key = f"ak_rules_editor_{int(st.session_state.get('ak_rules_editor_nonce', 0))}"
                scale_col, _spacer_col = st.columns([1, 2], vertical_alignment="top")
                with scale_col:
                    ak_rules_draft = st.data_editor(
                        st.session_state["ak_rules_df"],
                        num_rows="dynamic",
                        width="stretch",
                        row_height=ui_editor_row_height,
                        column_config={
                            "min_budget_wo_vat": st.column_config.NumberColumn(
                                "Бюджет от (без НДС), ₽",
                                min_value=0.0,
                                step=10000.0,
                                format="%.0f ₽",
                            ),
                            "ak_percent": st.column_config.NumberColumn(
                                "АК, %",
                                min_value=0.0,
                                step=0.5,
                                format="%.2f%%",
                            ),
                        },
                        key=ak_editor_key,
                    )
                if isinstance(ak_rules_draft, pd.DataFrame):
                    ak_rules_clean = normalize_ak_rules_df(ak_rules_draft)
                    st.session_state["ak_rules_df"] = ak_rules_clean
                    ak_rules_df = ak_rules_clean
        vat_ak_apply_state = get_vat_ak_apply_state(
            use_vat_budget_metrics=use_vat_budget_metrics,
            use_ak_budget_metrics=use_ak_budget_metrics,
            ak_mode=ak_mode,
            ak_fixed_month_wo_vat=ak_fixed_month_wo_vat,
            ak_fixed_percent=ak_fixed_percent,
            ak_rules_df=st.session_state.get("ak_rules_df"),
            last_applied_sig=st.session_state.get("vat_ak_last_applied_sig"),
        )
        current_vat_ak_sig = vat_ak_apply_state["current_sig"]
        last_applied_vat_ak_sig = vat_ak_apply_state["last_applied_sig"]
        vat_ak_dirty = vat_ak_apply_state["dirty"]
        apply_vat_ak_clicked = st.button(
            "Применить настройки НДС и АК",
            type="primary" if vat_ak_dirty else "secondary",
            key="apply_vat_ak_settings_btn",
        )
        if apply_vat_ak_clicked:
            st.session_state["vat_ak_last_applied_sig"] = current_vat_ak_sig
            if vat_ak_dirty:
                st.success("Настройки НДС и АК применены.")
            else:
                st.info("Изменений нет: текущие настройки уже применены.")
        st.caption(
            "CPC всегда считается от бюджета без НДС. При выключенном режиме НДС расчеты возвращаются к старой логике."
        )

    # ---------- 1.3. Наборы коэффициентов для типов РК ----------

    with st.expander("1.3. Наборы коэффициентов для типов РК", expanded=True):
        st.markdown(
            """
            <div style="
                margin: 8px 0 14px 0;
                display: inline-block;
                width: fit-content;
                max-width: 100%;
                white-space: nowrap;
                padding: 10px 12px;
                border-radius: 10px;
                border: 1px solid #FF8A66;
                background: rgba(255, 99, 51, 0.14);
                color: #FFD9CC;
                font-weight: 600;
            ">
                Важно: без назначенных наборов расчет не пойдет.
            </div>
            """,
            unsafe_allow_html=True,
        )

        coeff_sets = st.session_state.get("coeff_sets", [])
        coeff_sets_by_id = {
            get_coeff_set_id_value(cs): cs
            for cs in coeff_sets
            if isinstance(cs, dict) and get_coeff_set_id_value(cs)
        }

        demand_set_ids = [
            get_coeff_set_id_value(cs) for cs in coeff_sets
            if normalize_coeff_set_type(cs.get("type")) in ["Спрос (по запросам)", "Кастомный набор"]
            and get_coeff_set_id_value(cs)
        ]
        aov_set_ids = [
            get_coeff_set_id_value(cs) for cs in coeff_sets
            if normalize_coeff_set_type(cs.get("type")) in ["AOV (средний чек)", "Кастомный набор"]
            and get_coeff_set_id_value(cs)
        ]
        media_tail_set_ids = [
            get_coeff_set_id_value(cs) for cs in coeff_sets
            if normalize_coeff_set_type(cs.get("type")) == "Медийные хвосты"
            and get_coeff_set_id_value(cs)
        ]
        capacity_set_ids = [get_coeff_set_id_value(cs) for cs in coeff_sets if get_coeff_set_id_value(cs)]
        valid_demand_set_ids = set(demand_set_ids)
        valid_aov_set_ids = set(aov_set_ids)
        valid_media_tail_set_ids = set(media_tail_set_ids)
        valid_capacity_set_ids = set(capacity_set_ids)
        demand_set_labels = [get_coeff_set_label(coeff_sets_by_id[set_id]) for set_id in demand_set_ids if set_id in coeff_sets_by_id]
        aov_set_labels = [get_coeff_set_label(coeff_sets_by_id[set_id]) for set_id in aov_set_ids if set_id in coeff_sets_by_id]
        media_tail_set_labels = [get_coeff_set_label(coeff_sets_by_id[set_id]) for set_id in media_tail_set_ids if set_id in coeff_sets_by_id]
        capacity_set_labels = [get_coeff_set_label(coeff_sets_by_id[set_id]) for set_id in capacity_set_ids if set_id in coeff_sets_by_id]

        if "coeff_sets_links_new" not in st.session_state:
            st.session_state["coeff_sets_links_new"] = pd.DataFrame(
                columns=["campaign_key", "campaign_type", "system", "format", "geo", "segment", "demand_set", "aov_set", "media_tail_set", "capacity_set", "client_count_set", "absolute_new_clients_set", "returned_clients_set", "order_frequency_set"]
            )

        coeff_links_prev = st.session_state["coeff_sets_links_new"].copy()
        prev_demand_by_key = {}
        prev_demand_by_campaign = {}
        prev_aov_by_key = {}
        prev_aov_by_campaign = {}
        prev_media_tail_by_key = {}
        prev_media_tail_by_campaign = {}
        prev_capacity_by_key = {}
        prev_capacity_by_campaign = {}
        prev_client_count_by_key = {}
        prev_client_count_by_campaign = {}
        prev_absolute_new_clients_by_key = {}
        prev_absolute_new_clients_by_campaign = {}
        prev_returned_clients_by_key = {}
        prev_returned_clients_by_campaign = {}
        prev_order_frequency_by_key = {}
        prev_order_frequency_by_campaign = {}
        for _, r in coeff_links_prev.iterrows():
            ct = str(r.get("campaign_type", "")).strip()
            row_key = str(r.get("campaign_key", "")).strip() or build_campaign_identity_key(
                campaign_type=ct,
                geo=r.get("geo", ""),
                segment=r.get("segment", ""),
            )
            if not ct:
                continue
            demand_val = normalize_coeff_link_value(r.get("demand_set", ""), coeff_sets, valid_demand_set_ids)
            aov_val = normalize_coeff_link_value(r.get("aov_set", ""), coeff_sets, valid_aov_set_ids)
            media_tail_val = normalize_coeff_link_value(r.get("media_tail_set", ""), coeff_sets, valid_media_tail_set_ids)
            capacity_val = normalize_coeff_link_value(r.get("capacity_set", ""), coeff_sets, valid_capacity_set_ids)
            client_count_val = normalize_coeff_link_value(r.get("client_count_set", ""), coeff_sets, valid_capacity_set_ids)
            absolute_new_clients_val = normalize_coeff_link_value(r.get("absolute_new_clients_set", ""), coeff_sets, valid_capacity_set_ids)
            returned_clients_val = normalize_coeff_link_value(r.get("returned_clients_set", ""), coeff_sets, valid_capacity_set_ids)
            order_frequency_val = normalize_coeff_link_value(r.get("order_frequency_set", ""), coeff_sets, valid_capacity_set_ids)
            if row_key and row_key not in prev_demand_by_key:
                prev_demand_by_key[row_key] = demand_val
                prev_aov_by_key[row_key] = aov_val
                prev_media_tail_by_key[row_key] = media_tail_val
                prev_capacity_by_key[row_key] = capacity_val
                prev_client_count_by_key[row_key] = client_count_val
                prev_absolute_new_clients_by_key[row_key] = absolute_new_clients_val
                prev_returned_clients_by_key[row_key] = returned_clients_val
                prev_order_frequency_by_key[row_key] = order_frequency_val
            if ct not in prev_demand_by_campaign:
                prev_demand_by_campaign[ct] = demand_val
                prev_aov_by_campaign[ct] = aov_val
                prev_media_tail_by_campaign[ct] = media_tail_val
                prev_capacity_by_campaign[ct] = capacity_val
                prev_client_count_by_campaign[ct] = client_count_val
                prev_absolute_new_clients_by_campaign[ct] = absolute_new_clients_val
                prev_returned_clients_by_campaign[ct] = returned_clients_val
                prev_order_frequency_by_campaign[ct] = order_frequency_val

        coeff_links_new = pd.DataFrame(
            {
                "campaign_key": campaign_identity_rows["campaign_key"],
                "campaign_type": campaign_identity_rows["campaign_type"],
                "geo": campaign_identity_rows["geo"],
                "segment": campaign_identity_rows["segment"],
                "demand_set": [
                    prev_demand_by_key.get(row_key, prev_demand_by_campaign.get(ct, ""))
                    for row_key, ct in zip(campaign_identity_rows["campaign_key"], campaign_identity_rows["campaign_type"])
                ],
                "aov_set": [
                    prev_aov_by_key.get(row_key, prev_aov_by_campaign.get(ct, ""))
                    for row_key, ct in zip(campaign_identity_rows["campaign_key"], campaign_identity_rows["campaign_type"])
                ],
                "media_tail_set": [
                    prev_media_tail_by_key.get(row_key, prev_media_tail_by_campaign.get(ct, ""))
                    for row_key, ct in zip(campaign_identity_rows["campaign_key"], campaign_identity_rows["campaign_type"])
                ],
                "capacity_set": [
                    prev_capacity_by_key.get(row_key, prev_capacity_by_campaign.get(ct, prev_demand_by_campaign.get(ct, "")))
                    for row_key, ct in zip(campaign_identity_rows["campaign_key"], campaign_identity_rows["campaign_type"])
                ],
                "client_count_set": [
                    prev_client_count_by_key.get(row_key, prev_client_count_by_campaign.get(ct, prev_capacity_by_campaign.get(ct, prev_demand_by_campaign.get(ct, ""))))
                    for row_key, ct in zip(campaign_identity_rows["campaign_key"], campaign_identity_rows["campaign_type"])
                ],
                "absolute_new_clients_set": [
                    prev_absolute_new_clients_by_key.get(row_key, prev_absolute_new_clients_by_campaign.get(ct, prev_client_count_by_campaign.get(ct, prev_capacity_by_campaign.get(ct, prev_demand_by_campaign.get(ct, "")))))
                    for row_key, ct in zip(campaign_identity_rows["campaign_key"], campaign_identity_rows["campaign_type"])
                ],
                "returned_clients_set": [
                    prev_returned_clients_by_key.get(row_key, prev_returned_clients_by_campaign.get(ct, prev_client_count_by_campaign.get(ct, prev_capacity_by_campaign.get(ct, prev_demand_by_campaign.get(ct, "")))))
                    for row_key, ct in zip(campaign_identity_rows["campaign_key"], campaign_identity_rows["campaign_type"])
                ],
                "order_frequency_set": [
                    prev_order_frequency_by_key.get(row_key, prev_order_frequency_by_campaign.get(ct, prev_client_count_by_campaign.get(ct, prev_capacity_by_campaign.get(ct, prev_demand_by_campaign.get(ct, "")))))
                    for row_key, ct in zip(campaign_identity_rows["campaign_key"], campaign_identity_rows["campaign_type"])
                ],
            }
        )
        if not is_diy_preset:
            coeff_links_new = coeff_links_new.drop(columns=["capacity_set", "client_count_set", "absolute_new_clients_set", "returned_clients_set", "order_frequency_set"], errors="ignore")
        if not metric_mode["needs_aov"]:
            coeff_links_new = coeff_links_new.drop(columns=["aov_set"], errors="ignore")

        dangling_links_detected = False
        if "demand_set" in coeff_links_new.columns:
            invalid_mask = coeff_links_new["demand_set"].astype(str).str.strip().ne("") & ~coeff_links_new["demand_set"].isin(valid_demand_set_ids)
            if invalid_mask.any():
                coeff_links_new.loc[invalid_mask, "demand_set"] = ""
                dangling_links_detected = True
        if "aov_set" in coeff_links_new.columns:
            invalid_mask = coeff_links_new["aov_set"].astype(str).str.strip().ne("") & ~coeff_links_new["aov_set"].isin(valid_aov_set_ids)
            if invalid_mask.any():
                coeff_links_new.loc[invalid_mask, "aov_set"] = ""
                dangling_links_detected = True
        if "media_tail_set" in coeff_links_new.columns:
            invalid_mask = coeff_links_new["media_tail_set"].astype(str).str.strip().ne("") & ~coeff_links_new["media_tail_set"].isin(valid_media_tail_set_ids)
            if invalid_mask.any():
                coeff_links_new.loc[invalid_mask, "media_tail_set"] = ""
                dangling_links_detected = True
        if "capacity_set" in coeff_links_new.columns:
            invalid_mask = coeff_links_new["capacity_set"].astype(str).str.strip().ne("") & ~coeff_links_new["capacity_set"].isin(valid_capacity_set_ids)
            if invalid_mask.any():
                coeff_links_new.loc[invalid_mask, "capacity_set"] = ""
                dangling_links_detected = True
        if "client_count_set" in coeff_links_new.columns:
            invalid_mask = coeff_links_new["client_count_set"].astype(str).str.strip().ne("") & ~coeff_links_new["client_count_set"].isin(valid_capacity_set_ids)
            if invalid_mask.any():
                coeff_links_new.loc[invalid_mask, "client_count_set"] = ""
                dangling_links_detected = True
        if "absolute_new_clients_set" in coeff_links_new.columns:
            invalid_mask = coeff_links_new["absolute_new_clients_set"].astype(str).str.strip().ne("") & ~coeff_links_new["absolute_new_clients_set"].isin(valid_capacity_set_ids)
            if invalid_mask.any():
                coeff_links_new.loc[invalid_mask, "absolute_new_clients_set"] = ""
                dangling_links_detected = True
        if "returned_clients_set" in coeff_links_new.columns:
            invalid_mask = coeff_links_new["returned_clients_set"].astype(str).str.strip().ne("") & ~coeff_links_new["returned_clients_set"].isin(valid_capacity_set_ids)
            if invalid_mask.any():
                coeff_links_new.loc[invalid_mask, "returned_clients_set"] = ""
                dangling_links_detected = True
        if "order_frequency_set" in coeff_links_new.columns:
            invalid_mask = coeff_links_new["order_frequency_set"].astype(str).str.strip().ne("") & ~coeff_links_new["order_frequency_set"].isin(valid_capacity_set_ids)
            if invalid_mask.any():
                coeff_links_new.loc[invalid_mask, "order_frequency_set"] = ""
                dangling_links_detected = True
        if dangling_links_detected:
            st.session_state["coeff_sets_links_new"] = coeff_links_new.copy()

        coeff_links_editor = coeff_links_new.copy()
        coeff_link_allowed_ids = {
            "demand_set": valid_demand_set_ids,
            "aov_set": valid_aov_set_ids,
            "media_tail_set": valid_media_tail_set_ids,
            "capacity_set": valid_capacity_set_ids,
            "client_count_set": valid_capacity_set_ids,
            "absolute_new_clients_set": valid_capacity_set_ids,
            "returned_clients_set": valid_capacity_set_ids,
            "order_frequency_set": valid_capacity_set_ids,
        }
        for col_name, allowed_ids in coeff_link_allowed_ids.items():
            if col_name not in coeff_links_editor.columns:
                continue
            coeff_links_editor[col_name] = coeff_links_editor[col_name].map(
                lambda raw_value: get_coeff_set_label(coeff_sets_by_id[raw_value])
                if str(raw_value).strip() in allowed_ids and str(raw_value).strip() in coeff_sets_by_id
                else ""
            )

        if metric_mode["needs_aov"]:
            st.markdown(
                "Для каждой **строки РК** выберите набор сезонности **спроса** и набор сезонности **AOV (средний чек)**."
            )
        else:
            st.markdown(
                "Для каждой **строки РК** выберите набор сезонности **спроса**. Для пресета «Недвижимость» набор AOV не используется."
            )

        with st.form("coeff_sets_links_form"):
            bulk_cols = [2, 2, 1.2] if metric_mode["needs_aov"] else [2, 1.2]
            bulk_columns = st.columns(bulk_cols, vertical_alignment="bottom")
            bulk_c1 = bulk_columns[0]
            bulk_c2 = bulk_columns[1] if metric_mode["needs_aov"] else None
            bulk_c3 = bulk_columns[-1]
            with bulk_c1:
                bulk_demand_set = st.selectbox(
                    "Набор спроса для всех РК",
                    options=[""] + demand_set_labels,
                    help="Быстро назначает один набор сезонности спроса всем типам РК в таблице ниже.",
                    key="bulk_demand_set_for_all",
                )
            bulk_aov_set = ""
            if metric_mode["needs_aov"] and bulk_c2 is not None:
                with bulk_c2:
                    bulk_aov_set = st.selectbox(
                        "Набор AOV для всех РК",
                        options=[""] + aov_set_labels,
                        help="Быстро назначает один набор сезонности AOV всем типам РК в таблице ниже.",
                        key="bulk_aov_set_for_all",
                    )
            with bulk_c3:
                apply_all_clicked = st.form_submit_button("Применить ко всем", type="primary")

            links_column_config = {
                "campaign_type": st.column_config.TextColumn(
                    "Тип РК",
                    disabled=True,
                ),
                "geo": st.column_config.TextColumn("ГЕО", disabled=True),
                "segment": st.column_config.TextColumn("Сегмент", disabled=True),
                "demand_set": st.column_config.SelectboxColumn(
                    "Набор сезонности спроса",
                    options=demand_set_labels,
                    help="Влияет на показы (k_imp), а также через эластичность на CTR/CPC/CR.",
                ),
                "media_tail_set": st.column_config.SelectboxColumn(
                    "Набор медийных хвостов (показы)",
                    options=[""] + media_tail_set_labels,
                    help="Множитель показов поверх спроса: k_imp = k_demand * k_media_tail.",
                ),
            }
            if metric_mode["needs_aov"]:
                links_column_config["aov_set"] = st.column_config.SelectboxColumn(
                    "Набор сезонности AOV",
                    options=aov_set_labels,
                    help="Влияет на средний чек: k_aov.",
                )
            if is_diy_preset:
                links_column_config["capacity_set"] = st.column_config.SelectboxColumn(
                    "Набор для емкости",
                    options=capacity_set_labels,
                    help="Коэффициенты этого набора применяются к доступной емкости (k_capacity).",
                )
                links_column_config["client_count_set"] = st.column_config.SelectboxColumn(
                    "Набор для клиентов",
                    options=capacity_set_labels,
                    help="Коэффициенты этого набора применяются к количеству клиентов (k_client_count).",
                )
                links_column_config["absolute_new_clients_set"] = st.column_config.SelectboxColumn(
                    "Набор для АНК",
                    options=capacity_set_labels,
                    help="Коэффициенты этого набора применяются к абсолютно новым клиентам.",
                )
                links_column_config["returned_clients_set"] = st.column_config.SelectboxColumn(
                    "Набор для вернувшихся",
                    options=capacity_set_labels,
                    help="Коэффициенты этого набора применяются к вернувшимся клиентам.",
                )
                links_column_config["order_frequency_set"] = st.column_config.SelectboxColumn(
                    "Набор для частоты заказа",
                    options=capacity_set_labels,
                    help="Коэффициенты этого набора применяются к частоте заказа.",
                )

            coeff_links_draft = st.data_editor(
                coeff_links_editor.set_index("campaign_key"),
                num_rows="fixed",
                width="stretch",
                row_height=ui_editor_row_height,
                column_config=links_column_config,
                key="coeff_sets_links_new_editor",
                hide_index=True,
            )
            coeff_links_draft = coeff_links_draft.reset_index(names="campaign_key")
            if _has_unsaved_changes(coeff_links_draft, coeff_links_new):
                st.caption("Есть несохраненные изменения в выборе наборов.")
            links_saved = st.form_submit_button("Сохранить выбор наборов", type="primary")

        if apply_all_clicked or links_saved:
            coeff_links_new = coeff_links_draft.copy()
            for col_name, allowed_ids in coeff_link_allowed_ids.items():
                if col_name in coeff_links_new.columns:
                    coeff_links_new[col_name] = coeff_links_new[col_name].map(
                        lambda raw_value: normalize_coeff_link_value(raw_value, coeff_sets, allowed_ids)
                    )
            if apply_all_clicked:
                bulk_demand_set_id = normalize_coeff_link_value(bulk_demand_set, coeff_sets, valid_demand_set_ids)
                bulk_aov_set_id = normalize_coeff_link_value(bulk_aov_set, coeff_sets, valid_aov_set_ids)
                if bulk_demand_set_id:
                    coeff_links_new["demand_set"] = bulk_demand_set_id
                if bulk_aov_set_id and "aov_set" in coeff_links_new.columns:
                    coeff_links_new["aov_set"] = bulk_aov_set_id
            st.session_state["coeff_sets_links_new"] = coeff_links_new
            if apply_all_clicked:
                # После массового применения перерисовываем форму,
                # чтобы data_editor отобразил новые значения в строках.
                st.session_state.pop("coeff_sets_links_new_editor", None)
                st.rerun()

        aov_link_map = {}
        demand_link_map = {}
        media_tail_link_map = {}
        capacity_link_map = {}
        client_count_link_map = {}
        absolute_new_clients_link_map = {}
        returned_clients_link_map = {}
        order_frequency_link_map = {}

        for _, r in coeff_links_new.iterrows():
            campaign_key = str(r.get("campaign_key", "")).strip()
            if not campaign_key:
                continue
            demand_link_map[campaign_key] = str(r.get("demand_set", "")).strip()
            aov_link_map[campaign_key] = str(r.get("aov_set", "")).strip()
            media_tail_link_map[campaign_key] = str(r.get("media_tail_set", "")).strip()
            capacity_link_map[campaign_key] = str(r.get("capacity_set", "")).strip()
            client_count_link_map[campaign_key] = str(r.get("client_count_set", "")).strip()
            absolute_new_clients_link_map[campaign_key] = str(r.get("absolute_new_clients_set", "")).strip()
            returned_clients_link_map[campaign_key] = str(r.get("returned_clients_set", "")).strip()
            order_frequency_link_map[campaign_key] = str(r.get("order_frequency_set", "")).strip()

        missing_demand = [
            get_campaign_identity_label(r)
            for _, r in campaign_identity_rows.iterrows()
            if not demand_link_map.get(str(r.get("campaign_key", "")).strip())
        ]
        if missing_demand:
            st.error(
                "Для следующих строк РК не выбран набор сезонности спроса, поэтому медиаплан дальше считаться не будет: "
                + ", ".join(missing_demand)
            )
            plan_calc_blockers.append("не выбраны наборы коэффициентов спроса")

        missing_aov = [
            get_campaign_identity_label(r)
            for _, r in campaign_identity_rows.iterrows()
            if not aov_link_map.get(str(r.get("campaign_key", "")).strip())
        ]
        if metric_mode["needs_aov"] and missing_aov:
            st.error(
                "Для следующих строк РК не выбран набор сезонности AOV (средний чек), "
                "поэтому медиаплан дальше считаться не будет: "
                + ", ".join(missing_aov)
            )
            plan_calc_blockers.append("не выбраны наборы коэффициентов AOV")

    # ---------- 1.4. Эластичность метрик к сезонности спроса ----------

    with st.expander("1.4. Эластичность метрик к сезонности спроса", expanded=False):
        show_cr2_elasticity = bool((is_real_estate_preset and metric_mode["is_real_estate_full"]) or is_diy_preset)

        st.caption("Памятка: меньше делитель = сильнее влияние.")
        st.caption(
            "Задайте делители влияния сезонности спроса на метрики для каждой строки РК. "
            "Чем меньше делитель, тем сильнее влияние. Процентный результат смотрите в превью справа."
        )
        st.markdown(
            "**Как считаются эффекты эластичности:**\n"
            "1. `CPC` в режиме роста CPC: формула `(k-1)/div + 1`.\n"
            "2. `CTR` в режиме падения CTR: формула `1 - (k-1)/div`.\n"
            "3. `CR` в режиме падения CR: формула `1 - (k-1)/div`.\n"
            + ("4. `CR2` в недвижимости считается отдельно только для полной воронки.\n" if show_cr2_elasticity else "")
            + "Пример: для `k=1.10` и `CPC div=2` получится `+5%` к CPC."
        )

        if "elasticity_df" not in st.session_state:
            st.session_state["elasticity_df"] = pd.DataFrame(
                {
                "campaign_key": campaign_identity_rows["campaign_key"],
                "campaign_type": campaign_identity_rows["campaign_type"],
                "geo": campaign_identity_rows["geo"],
                "segment": campaign_identity_rows["segment"],
                    "cpc_div": [1.0] * len(campaign_identity_rows),
                    "ctr_div": [2.0] * len(campaign_identity_rows),
                    "cr_div": [10.0] * len(campaign_identity_rows),
                    "cr2_div": [10.0] * len(campaign_identity_rows),
                }
            )

        elasticity_df = st.session_state["elasticity_df"]

        # Миграция старого формата процентов/legacy в делители.
        if "cpc_div" not in elasticity_df.columns and "cpc_impact_pct_per_10" in elasticity_df.columns:
            src = pd.to_numeric(elasticity_df["cpc_impact_pct_per_10"], errors="coerce")
            elasticity_df["cpc_div"] = np.where(src > 0, 10.0 / src, 1.0)
        if "ctr_div" not in elasticity_df.columns and "ctr_impact_pct_per_10" in elasticity_df.columns:
            src = pd.to_numeric(elasticity_df["ctr_impact_pct_per_10"], errors="coerce").abs()
            elasticity_df["ctr_div"] = np.where(src > 0, 10.0 / src, 2.0)
        if "cr_div" not in elasticity_df.columns and "cr_impact_pct_per_10" in elasticity_df.columns:
            src = pd.to_numeric(elasticity_df["cr_impact_pct_per_10"], errors="coerce").abs()
            elasticity_df["cr_div"] = np.where(src > 0, 10.0 / src, 10.0)
        if "cpc_div" not in elasticity_df.columns and "cpc_up_pct_per_10" in elasticity_df.columns:
            src = pd.to_numeric(elasticity_df["cpc_up_pct_per_10"], errors="coerce")
            elasticity_df["cpc_div"] = np.where(src > 0, 10.0 / src, 1.0)
        if "ctr_div" not in elasticity_df.columns and "ctr_down_pct_per_10" in elasticity_df.columns:
            src = pd.to_numeric(elasticity_df["ctr_down_pct_per_10"], errors="coerce").abs()
            elasticity_df["ctr_div"] = np.where(src > 0, 10.0 / src, 2.0)
        if "cr_div" not in elasticity_df.columns and "cr_down_pct_per_10" in elasticity_df.columns:
            src = pd.to_numeric(elasticity_df["cr_down_pct_per_10"], errors="coerce").abs()
            elasticity_df["cr_div"] = np.where(src > 0, 10.0 / src, 10.0)
        if "cr2_div" not in elasticity_df.columns:
            elasticity_df["cr2_div"] = pd.to_numeric(elasticity_df.get("cr_div", 10.0), errors="coerce").fillna(10.0)
        for col_name, default_val in [("cpc_div", 1.0), ("ctr_div", 2.0), ("cr_div", 10.0), ("cr2_div", 10.0)]:
            if col_name in elasticity_df.columns:
                elasticity_df[col_name] = pd.to_numeric(elasticity_df[col_name], errors="coerce").fillna(default_val)

        if "campaign_key" not in elasticity_df.columns:
            elasticity_df["campaign_key"] = elasticity_df.apply(
                lambda row: build_campaign_identity_key(
                    campaign_type=row.get("campaign_type", ""),
                    geo=row.get("geo", ""),
                    segment=row.get("segment", ""),
                ),
                axis=1,
            )

        elasticity_df = elasticity_df[
            elasticity_df["campaign_key"].astype(str).isin(campaign_identity_rows["campaign_key"].astype(str))
            | elasticity_df["campaign_type"].astype(str).isin(existing_ctypes)
        ]

        keep_cols = ["campaign_key", "campaign_type", "geo", "segment", "preset", "cpc_div", "ctr_div", "cr_div", "cr2_div"]
        for c in keep_cols:
            if c not in elasticity_df.columns:
                if c == "campaign_key":
                    elasticity_df[c] = ""
                elif c == "campaign_type":
                    elasticity_df[c] = ""
                elif c == "geo":
                    elasticity_df[c] = ""
                elif c == "segment":
                    elasticity_df[c] = "ALL"
                elif c == "preset":
                    elasticity_df[c] = "Кастом"
                elif c == "cpc_div":
                    elasticity_df[c] = 1.0
                elif c == "ctr_div":
                    elasticity_df[c] = 2.0
                elif c == "cr_div":
                    elasticity_df[c] = 10.0
                elif c == "cr2_div":
                    elasticity_df[c] = 10.0
                else:
                    elasticity_df[c] = 0.0
        elasticity_df = elasticity_df[keep_cols]

        existing_elasticity_keys = set(elasticity_df["campaign_key"].astype(str).tolist())
        legacy_by_campaign_type = (
            elasticity_df.drop_duplicates(subset=["campaign_type"], keep="first")
            .set_index("campaign_type")
            if not elasticity_df.empty
            else pd.DataFrame()
        )
        for _, campaign_row in campaign_identity_rows.iterrows():
            row_key = str(campaign_row.get("campaign_key", "")).strip()
            campaign_type = str(campaign_row.get("campaign_type", "")).strip()
            if row_key in existing_elasticity_keys:
                continue
            legacy_row = None
            if campaign_type and not legacy_by_campaign_type.empty and campaign_type in legacy_by_campaign_type.index:
                legacy_row = legacy_by_campaign_type.loc[campaign_type]
                if isinstance(legacy_row, pd.DataFrame):
                    legacy_row = legacy_row.iloc[0]
            preset_value = str(legacy_row.get("preset", "Среднее")).strip() if legacy_row is not None else "Среднее"
            cpc_div_value = float(legacy_row.get("cpc_div", 1.0)) if legacy_row is not None else 1.0
            ctr_div_value = float(legacy_row.get("ctr_div", 2.0)) if legacy_row is not None else 2.0
            cr_div_value = float(legacy_row.get("cr_div", 10.0)) if legacy_row is not None else 10.0
            cr2_div_value = float(legacy_row.get("cr2_div", 10.0)) if legacy_row is not None else 10.0
            elasticity_df = pd.concat(
                [
                    elasticity_df,
                    pd.DataFrame(
                        [{
                            "campaign_key": row_key,
                            "campaign_type": campaign_type,
                            "geo": str(campaign_row.get("geo", "")),
                            "segment": str(campaign_row.get("segment", "ALL")),
                            "preset": preset_value if preset_value else "Среднее",
                            "cpc_div": cpc_div_value,
                            "ctr_div": ctr_div_value,
                            "cr_div": cr_div_value,
                            "cr2_div": cr2_div_value,
                        }]
                    ),
                ],
                ignore_index=True,
            )

        def _style_impact_col(col: pd.Series) -> pd.Series:
            out = pd.Series([""] * len(col), index=col.index)
            vals = pd.to_numeric(col, errors="coerce")
            for idx_v, v in vals.items():
                if pd.isna(v):
                    continue
                if v > 0:
                    out.loc[idx_v] = "color: #00CDC5; font-weight: 700;"
                elif v < 0:
                    out.loc[idx_v] = "color: #FF6333; font-weight: 700;"
                else:
                    out.loc[idx_v] = "color: #D0D6DF;"
            return out

        def _coerce_div_for_preview(val: float, default_val: float) -> float:
            try:
                num = float(val)
                return num if num > 0 else default_val
            except Exception:
                return default_val

        def _row_divs_match_preset(row: pd.Series, preset_vals: dict[str, float], tol: float = 1e-9) -> bool:
            return (
                abs(_coerce_div_for_preview(row.get("cpc_div", 1.0), 1.0) - float(preset_vals["cpc_div"])) <= tol
                and abs(_coerce_div_for_preview(row.get("ctr_div", 2.0), 2.0) - float(preset_vals["ctr_div"])) <= tol
                and abs(_coerce_div_for_preview(row.get("cr_div", 10.0), 10.0) - float(preset_vals["cr_div"])) <= tol
                and (
                    not show_cr2_elasticity
                    or abs(_coerce_div_for_preview(row.get("cr2_div", 10.0), 10.0) - float(preset_vals["cr2_div"])) <= tol
                )
            )

        with st.form("elasticity_editor_form"):
            preset_map = {
                "Слабое": {"cpc_div": 2.0, "ctr_div": 5.0, "cr_div": 15.0, "cr2_div": 15.0},
                "Среднее": {"cpc_div": 1.0, "ctr_div": 2.0, "cr_div": 10.0, "cr2_div": 10.0},
                "Сильное": {"cpc_div": 0.5, "ctr_div": 1.0, "cr_div": 5.0, "cr2_div": 5.0},
            }
            st.caption("Для каждой строки РК можно выбрать пресет в колонке «Пресет» или задать значения вручную (Кастом).")
            st.info("Чтобы выбрать пресет: кликните ячейку в колонке «Пресет ▼» и выберите вариант из списка.")
            st.caption("В колонках CPC/CTR/CR указываются делители влияния. Процентный эффект для выбранного коэффициента спроса показывается в превью справа." + (" Для полной воронки недвижимости отдельно настраивается CR2." if show_cr2_elasticity else ""))
            st.markdown(
                "**Пресеты влияния:**\n"
                "1. `Мягкий` - слабое влияние на метрики (CPC растет меньше, CTR/CR падают меньше).\n"
                "2. `Средний` - сбалансированный вариант (дефолт для большинства задач).\n"
                "3. `Сильный` - агрессивное влияние (выше рост CPC и сильнее падение CTR/CR).\n"
            )
            st.caption(
                "Пример эффекта для роста спроса на +10%:\n"
                "Мягкий: CPC +5.0%, CTR -2.0%, CR -0.67% | "
                "Средний: CPC +10.0%, CTR -5.0%, CR -1.0% | "
                "Сильный: CPC +20.0%, CTR -10.0%, CR -2.0%."
            )
            st.caption(
                "Пример при +30% спроса (в 3 раза сильнее): "
                "для пресета «Среднее» ориентир: CPC +30.0%, CTR -15.0%, CR -3.0%."
            )
            st.caption(
                "Пример для снижения спроса до k=0.85 (-15%):\n"
                "Мягкий: CPC -7.5%, CTR +3.0%, CR +1.0% | "
                "Средний: CPC -15.0%, CTR +7.5%, CR +1.5% | "
                "Сильный: CPC -30.0%, CTR +15.0%, CR +3.0%."
                "Сильный: CPC -30.0%, CTR +15.0%, CR +3.0%."
            )
            top_left, top_right = st.columns([1.55, 1.05], vertical_alignment="bottom")
            with top_left:
                bulk_c1, bulk_c2, _bulk_sp = st.columns([1.2, 0.9, 1.2], vertical_alignment="bottom")
                with bulk_c1:
                    bulk_preset_name = st.selectbox(
                        "Массовый пресет",
                        options=[""] + list(preset_map.keys()),
                        key="elasticity_bulk_preset_select",
                    )
                with bulk_c2:
                    bulk_apply_clicked = st.form_submit_button("Применить массово", type="primary")
            with top_right:
                preview_k_col, preview_btn_col, _preview_k_sp = st.columns([0.34, 0.20, 0.46], vertical_alignment="bottom")
                with preview_k_col:
                    preview_k_demand = st.number_input(
                        "Коэффициент спроса для расчета превью",
                        min_value=0.10,
                        max_value=5.00,
                        value=float(st.session_state.get("elasticity_preview_k_demand", 1.10)),
                        step=0.01,
                        format="%.2f",
                        key="elasticity_preview_k_demand",
                        help="1.10 = +10% спроса, 1.30 = +30%, 0.90 = -10%.",
                    )
                with preview_btn_col:
                    preview_recalc_clicked = st.form_submit_button("Обновить")

            elasticity_row_h = max(28, int(ui_editor_row_height) - 8)
            compact_h = _table_height_for_rows(len(elasticity_df), min_height=140, max_height=280)

            edit_left, preview_right = st.columns([1.55, 1.05], vertical_alignment="top")
            with edit_left:
                st.caption("Настройка по строкам РК")
                elasticity_editor_key = f"elasticity_editor_{int(st.session_state.get('elasticity_editor_nonce', 0))}"
                elasticity_draft = st.data_editor(
                    elasticity_df.set_index("campaign_key"),
                    num_rows="fixed",
                    width="stretch",
                    row_height=elasticity_row_h,
                    height=compact_h,
                    column_config={
                        "campaign_type": st.column_config.TextColumn("Тип РК", disabled=True),
                        "geo": st.column_config.TextColumn("ГЕО", disabled=True),
                        "segment": st.column_config.TextColumn("Сегмент", disabled=True),
                        "preset": st.column_config.SelectboxColumn(
                            "Пресет ▼ (кликните ячейку)",
                            options=["Кастом", "Слабое", "Среднее", "Сильное"],
                            help="Если значения в строке совпадают с пресетом, он сохранится как пресет; если вы измените числа вручную, строка сохранится как «Кастом».",
                        ),
                        "cpc_div": st.column_config.NumberColumn(
                            "CPC",
                            format="%.2f",
                            help=mhelp("cpc_div"),
                        ),
                        "ctr_div": st.column_config.NumberColumn(
                            "CTR",
                            format="%.2f",
                            help=mhelp("ctr_div"),
                        ),
                        "cr_div": st.column_config.NumberColumn(
                            "CR",
                            format="%.2f",
                            help=mhelp("cr_div"),
                        ),
                        "cr2_div": st.column_config.NumberColumn(
                            "CR2",
                            format="%.2f",
                            help="Делитель влияния спроса на вторую ступень конверсии в ЦО.",
                        ),
                    },
                    key=elasticity_editor_key,
                    hide_index=True,
                )
                elasticity_draft = elasticity_draft.reset_index(names="campaign_key")
                if not show_cr2_elasticity and "cr2_div" in elasticity_draft.columns:
                    elasticity_draft = elasticity_draft.drop(columns=["cr2_div"], errors="ignore")
            with preview_right:
                st.caption("Результат по текущим значениям в строке")
                preview_effective = elasticity_draft[
                    ["campaign_key", "campaign_type", "geo", "segment", "preset", "cpc_div", "ctr_div", "cr_div"] + (["cr2_div"] if show_cr2_elasticity else [])
                ].copy()
                saved_by_campaign = elasticity_df.set_index("campaign_key")
                for idx_row, row in preview_effective.iterrows():
                    campaign_key = str(row.get("campaign_key", ""))
                    p_name = str(row.get("preset", "Кастом")).strip()
                    preview_effective.at[idx_row, "cpc_div"] = _coerce_div_for_preview(row.get("cpc_div", 1.0), 1.0)
                    preview_effective.at[idx_row, "ctr_div"] = _coerce_div_for_preview(row.get("ctr_div", 2.0), 2.0)
                    preview_effective.at[idx_row, "cr_div"] = _coerce_div_for_preview(row.get("cr_div", 10.0), 10.0)
                    if show_cr2_elasticity:
                        preview_effective.at[idx_row, "cr2_div"] = _coerce_div_for_preview(row.get("cr2_div", 10.0), 10.0)
                    if p_name not in preset_map or campaign_key not in saved_by_campaign.index:
                        continue
                    saved_row = saved_by_campaign.loc[campaign_key]
                    saved_preset = str(saved_row.get("preset", "Кастом")).strip()
                    saved_cpc = _coerce_div_for_preview(saved_row.get("cpc_div", 1.0), 1.0)
                    saved_ctr = _coerce_div_for_preview(saved_row.get("ctr_div", 2.0), 2.0)
                    saved_cr = _coerce_div_for_preview(saved_row.get("cr_div", 10.0), 10.0)
                    raw_cpc = float(preview_effective.at[idx_row, "cpc_div"])
                    raw_ctr = float(preview_effective.at[idx_row, "ctr_div"])
                    raw_cr = float(preview_effective.at[idx_row, "cr_div"])
                    preset_changed_only = (
                        p_name != saved_preset
                        and abs(raw_cpc - saved_cpc) <= 1e-9
                        and abs(raw_ctr - saved_ctr) <= 1e-9
                        and abs(raw_cr - saved_cr) <= 1e-9
                        and (
                            not show_cr2_elasticity
                            or abs(
                                float(preview_effective.at[idx_row, "cr2_div"])
                                - _coerce_div_for_preview(saved_row.get("cr2_div", 10.0), 10.0)
                            ) <= 1e-9
                        )
                    )
                    if preset_changed_only:
                        p_vals = preset_map[p_name]
                        preview_effective.at[idx_row, "cpc_div"] = p_vals["cpc_div"]
                        preview_effective.at[idx_row, "ctr_div"] = p_vals["ctr_div"]
                        preview_effective.at[idx_row, "cr_div"] = p_vals["cr_div"]
                        if show_cr2_elasticity:
                            preview_effective.at[idx_row, "cr2_div"] = p_vals["cr2_div"]
                # Пересчет под выбранный коэффициент спроса:
                # значения в таблице задаются как эффект на +10% спроса, масштабируем линейно.
                # scale = 1 при k=1.10; scale = 3 при k=1.30; scale = -1 при k=0.90.
                demand_delta = float(preview_k_demand) - 1.0
                cpc_div_vals = pd.to_numeric(preview_effective["cpc_div"], errors="coerce").replace(0, np.nan)
                ctr_div_vals = pd.to_numeric(preview_effective["ctr_div"], errors="coerce").replace(0, np.nan)
                cr_div_vals = pd.to_numeric(preview_effective["cr_div"], errors="coerce").replace(0, np.nan)
                cr2_div_vals = pd.to_numeric(preview_effective["cr2_div"], errors="coerce").replace(0, np.nan) if show_cr2_elasticity else None
                preview_effective["cpc_pct"] = (demand_delta / cpc_div_vals) * 100.0
                preview_effective["ctr_pct"] = -(demand_delta / ctr_div_vals) * 100.0
                preview_effective["cr_pct"] = -(demand_delta / cr_div_vals) * 100.0
                if show_cr2_elasticity:
                    preview_effective["cr2_pct"] = -(demand_delta / cr2_div_vals) * 100.0
                preview_cols = ["campaign_type", "geo", "segment", "cpc_pct", "ctr_pct", "cr_pct"] + (["cr2_pct"] if show_cr2_elasticity else [])
                preview_effective = preview_effective[preview_cols].rename(
                    columns={
                        "campaign_type": "Тип РК",
                        "geo": "ГЕО",
                        "segment": "Сегмент",
                        "cpc_pct": "CPC, %",
                        "ctr_pct": "CTR, %",
                        "cr_pct": "CR, %",
                        "cr2_pct": "CR2, %",
                    }
                )
                st.dataframe(
                    preview_effective.style
                    .format({"CPC, %": "{:+.2f}%", "CTR, %": "{:+.2f}%", "CR, %": "{:+.2f}%", "CR2, %": "{:+.2f}%"})
                    .apply(_style_impact_col, axis=0, subset=["CPC, %", "CTR, %", "CR, %"] + (["CR2, %"] if show_cr2_elasticity else [])),
                    width="stretch",
                    hide_index=True,
                    height=compact_h,
                )
                st.caption("Зеленый = рост метрики, красный = снижение. Значения пересчитаны под выбранный коэффициент спроса.")
            if _has_unsaved_changes(elasticity_draft, elasticity_df):
                st.caption("Есть несохраненные изменения эластичности.")
            btn_save_col, btn_reset_col, _btn_sp = st.columns([0.28, 0.34, 0.38])
            with btn_save_col:
                elasticity_saved = st.form_submit_button("Сохранить эластичность", type="primary")
            with btn_reset_col:
                reset_recommended_clicked = st.form_submit_button("Вернуть рекомендуемые значения")

        if bulk_apply_clicked:
            if bulk_preset_name in preset_map:
                elasticity_draft = elasticity_draft.copy()
                elasticity_draft["preset"] = bulk_preset_name
                preset_vals = preset_map[bulk_preset_name]
                elasticity_draft["cpc_div"] = preset_vals["cpc_div"]
                elasticity_draft["ctr_div"] = preset_vals["ctr_div"]
                elasticity_draft["cr_div"] = preset_vals["cr_div"]
                if "cr2_div" in elasticity_draft.columns:
                    elasticity_draft["cr2_div"] = preset_vals["cr2_div"]
                st.session_state["elasticity_df"] = elasticity_draft.copy()
                elasticity_df = elasticity_draft.copy()
                st.session_state["elasticity_editor_nonce"] = int(st.session_state.get("elasticity_editor_nonce", 0)) + 1
                st.success(f"Массово применен пресет «{bulk_preset_name}».")
                st.rerun()
            else:
                st.info("Выберите пресет для массового применения.")

        if elasticity_saved:
            elasticity_draft = elasticity_draft.copy()
            for idx_row, row in elasticity_draft.iterrows():
                preset_name = str(row.get("preset", "Кастом")).strip()
                if preset_name in preset_map:
                    preset_vals = preset_map[preset_name]
                    if _row_divs_match_preset(row, preset_vals):
                        elasticity_draft.at[idx_row, "cpc_div"] = preset_vals["cpc_div"]
                        elasticity_draft.at[idx_row, "ctr_div"] = preset_vals["ctr_div"]
                        elasticity_draft.at[idx_row, "cr_div"] = preset_vals["cr_div"]
                        if "cr2_div" in elasticity_draft.columns:
                            elasticity_draft.at[idx_row, "cr2_div"] = preset_vals["cr2_div"]
                    else:
                        elasticity_draft.at[idx_row, "preset"] = "Кастом"
            st.session_state["elasticity_df"] = elasticity_draft.copy()
            elasticity_df = elasticity_draft.copy()

        if reset_recommended_clicked:
            reset_df = elasticity_draft.copy()
            saved_by_campaign = elasticity_df.set_index("campaign_key") if not elasticity_df.empty else pd.DataFrame()
            for idx_row, row in reset_df.iterrows():
                campaign_key = str(row.get("campaign_key", ""))
                preset_name = str(row.get("preset", "Кастом")).strip()
                if preset_name not in preset_map:
                    saved_preset = ""
                    if campaign_key and not saved_by_campaign.empty and campaign_key in saved_by_campaign.index:
                        saved_preset = str(saved_by_campaign.loc[campaign_key].get("preset", "")).strip()
                    preset_name = saved_preset if saved_preset in preset_map else "Среднее"
                preset_vals = preset_map[preset_name]
                reset_df.at[idx_row, "preset"] = preset_name
                reset_df.at[idx_row, "cpc_div"] = preset_vals["cpc_div"]
                reset_df.at[idx_row, "ctr_div"] = preset_vals["ctr_div"]
                reset_df.at[idx_row, "cr_div"] = preset_vals["cr_div"]
                if "cr2_div" in reset_df.columns:
                    reset_df.at[idx_row, "cr2_div"] = preset_vals["cr2_div"]
            st.session_state["elasticity_df"] = reset_df
            st.session_state["elasticity_editor_nonce"] = int(st.session_state.get("elasticity_editor_nonce", 0)) + 1
            st.rerun()

        def _safe_div(val: float, default_val: float) -> float:
            try:
                num = float(val)
                if pd.isna(num) or num <= 0:
                    return default_val
                return num
            except Exception:
                return default_val

        def _is_invalid_div(val: float) -> bool:
            try:
                num = float(val)
                return pd.isna(num) or num <= 0
            except Exception:
                return True

        had_invalid_values = False
        elasticity_map = {}
        for _, r in elasticity_df.iterrows():
            camp = str(r.get("campaign_key", "")).strip()
            if not camp:
                camp = build_campaign_identity_key(
                    campaign_type=r.get("campaign_type", ""),
                    geo=r.get("geo", ""),
                    segment=r.get("segment", ""),
                )
            raw_cpc_div = r.get("cpc_div")
            raw_ctr_div = r.get("ctr_div")
            raw_cr_div = r.get("cr_div")
            raw_cr2_div = r.get("cr2_div")

            cpc_div = _safe_div(raw_cpc_div, 1.0)
            ctr_div = _safe_div(raw_ctr_div, 2.0)
            cr_div = _safe_div(raw_cr_div, 10.0)
            cr2_div = _safe_div(raw_cr2_div, 10.0)

            if (
                _is_invalid_div(raw_cpc_div)
                or _is_invalid_div(raw_ctr_div)
                or _is_invalid_div(raw_cr_div)
                or (show_cr2_elasticity and _is_invalid_div(raw_cr2_div))
            ):
                had_invalid_values = True

            elasticity_map[camp] = {
                "cpc_div": cpc_div,
                "ctr_div": ctr_div,
                "cr_div": cr_div,
                "cr2_div": cr2_div,
            }

        if had_invalid_values:
            st.warning(
                "Некорректные значения делителей эластичности (пустые, нечисловые или <= 0) автоматически заменены: "
                + ("CPC=1.0, CTR=2.0, CR1=10.0, CR2=10.0." if show_cr2_elasticity else "CPC=1.0, CTR=2.0, CR=10.0.")
            )

with tab_setup:
    st.caption("Ниже показан полный базовый «Средний месяц», который используется как опора для дальнейшего помесячного расчета.")

    # ---------- 2. Средний месяц (базовые значения) ----------

    if plan_calc_blockers:
        st.info(
            "Помесячный расчёт приостановлен. Чтобы перейти к блокам 2–4, исправьте: "
            + "; ".join(dict.fromkeys(plan_calc_blockers))
            + "."
        )
        selected_month_nums = []
        selected_planning_slots = []
        selected_periods = []
        selected_periods_by_slot = {}

    ui_section_title("2. Средний месяц (базовые значения)")
    # Берем актуальные правила из session_state перед расчетами,
    # чтобы исключить рассинхрон с data_editor.
    ak_rules_df = st.session_state.get("ak_rules_df", ak_rules_df)

    base_rows = []
    for _, base_row in campaigns.iterrows():
        base = PlanInput(
            impressions=base_row["impressions_avg"],
            ctr=base_row["ctr_avg_percent"] / 100.0,
            cpc=base_row["cpc_avg"],
            cr=base_row["cr_avg_percent"] / 100.0,
            aov=base_row["aov_avg"],
            cr2=float(base_row.get("cr2_avg_percent", 0.0) or 0.0) / 100.0,
            reach=float(base_row.get("reach_avg", 0.0) or 0.0),
            preset_key=active_preset_key,
            funnel_mode=metric_mode["real_estate_funnel_mode"],
        )
        out = calculate_plan_month(base)
        out["campaign_type"] = base_row["campaign_type"]
        out["segment"] = str(base_row.get("segment", "B2C"))
        out["system"] = base_row["system"]
        out["format"] = base_row["format"]
        out["geo"] = str(base_row.get("geo", "") or "")
        if is_diy_preset:
            cap_avg = float(base_row.get("available_capacity_avg", 0.0) or 0.0)
            shipped_aov_avg = float(base_row.get("shipped_aov_avg", 0.0) or 0.0)
            out["available_capacity"] = cap_avg
            out["client_count"] = float(base_row.get("client_count_avg", 0.0) or 0.0)
            out["absolute_new_clients"] = float(base_row.get("absolute_new_clients_avg", 0.0) or 0.0)
            out["returned_clients"] = float(base_row.get("returned_clients_avg", 0.0) or 0.0)
            out["new_clients"] = float(out.get("absolute_new_clients", 0.0)) + float(out.get("returned_clients", 0.0))
            out["order_frequency"] = float(base_row.get("order_frequency_avg", 0.0) or 0.0)
            out["new_clients_share_pct"] = calc_new_clients_share_pct(out["new_clients"], out["client_count"])
            out["sov_pct"] = (float(out.get("reach", 0.0)) / cap_avg * 100.0) if cap_avg > 0 else 0.0
            out["cac"] = (float(out.get("cost_with_vat_ak", 0.0)) / float(out.get("new_clients", 0.0))) if float(out.get("new_clients", 0.0)) > 0 else 0.0
            out["shipped_orders"] = float(out.get("target_leads", 0.0) or 0.0)
            out["shipped_revenue"] = float(out.get("shipped_orders", 0.0)) * shipped_aov_avg
            out["shipped_aov"] = shipped_aov_avg
        base_rows.append(out)

    df_base = pd.DataFrame(base_rows)
    if not df_base.empty:
        base_ak_rate = resolve_ak_rate(float(df_base["cost"].sum()), ak_rules_df) if (use_ak_budget_metrics and ak_mode == "percent") else 0.0
        df_base = apply_budget_basis_metrics(
            df_base,
            use_vat_budget_metrics,
            use_ak=use_ak_budget_metrics,
            ak_mode=ak_mode,
            default_ak_rate=base_ak_rate,
            default_ak_fixed_wo_vat=float(ak_fixed_month_wo_vat),
            default_ak_fixed_rate=float(ak_fixed_percent) / 100.0,
        )
        df_base["ak_rate_pct"] = df_base["ak_rate"] * 100.0
        if is_diy_preset:
            df_base["cpa"] = np.where(
                pd.to_numeric(df_base.get("conversions", 0.0), errors="coerce").fillna(0.0) > 0,
                pd.to_numeric(df_base.get("cost_with_vat", 0.0), errors="coerce").fillna(0.0)
                / pd.to_numeric(df_base.get("conversions", 0.0), errors="coerce").fillna(0.0),
                0.0,
            )
            df_base["cac"] = np.where(
                pd.to_numeric(df_base.get("new_clients", 0.0), errors="coerce").fillna(0.0) > 0,
                pd.to_numeric(df_base.get("cost_with_vat_ak", 0.0), errors="coerce").fillna(0.0)
                / pd.to_numeric(df_base.get("new_clients", 0.0), errors="coerce").fillna(0.0),
                0.0,
            )
            df_base["shipped_cps"] = np.where(
                pd.to_numeric(df_base.get("shipped_orders", 0.0), errors="coerce").fillna(0.0) > 0,
                pd.to_numeric(df_base.get("cost_with_vat", 0.0), errors="coerce").fillna(0.0)
                / pd.to_numeric(df_base.get("shipped_orders", 0.0), errors="coerce").fillna(0.0),
                0.0,
            )
            df_base["shipped_aov"] = np.where(
                pd.to_numeric(df_base.get("shipped_orders", 0.0), errors="coerce").fillna(0.0) > 0,
                pd.to_numeric(df_base.get("shipped_revenue", 0.0), errors="coerce").fillna(0.0)
                / pd.to_numeric(df_base.get("shipped_orders", 0.0), errors="coerce").fillna(0.0),
                0.0,
            )
            shipped_budget_basis = pd.to_numeric(df_base.get("cost_with_vat_ak" if use_vat_budget_metrics and use_ak_budget_metrics else "cost_with_vat" if use_vat_budget_metrics else "total_budget_wo_vat_ak" if use_ak_budget_metrics else "cost", 0.0), errors="coerce").fillna(0.0)
            df_base["shipped_roas"] = np.where(
                shipped_budget_basis > 0,
                pd.to_numeric(df_base.get("shipped_revenue", 0.0), errors="coerce").fillna(0.0) / shipped_budget_basis,
                0.0,
            )
            df_base["shipped_drr_pct"] = np.where(
                pd.to_numeric(df_base.get("shipped_revenue", 0.0), errors="coerce").fillna(0.0) > 0,
                shipped_budget_basis / pd.to_numeric(df_base.get("shipped_revenue", 0.0), errors="coerce").fillna(0.0) * 100.0,
                0.0,
            )
            df_base = add_segment_budget_share(df_base)
            df_base = add_segment_value_share(df_base, value_col="conversions", out_col="order_share_segment_pct")
            df_base = add_segment_value_share(df_base, value_col="revenue", out_col="revenue_share_segment_pct")
            df_base = add_segment_value_share(df_base, value_col="shipped_orders", out_col="shipped_order_share_segment_pct")
            df_base = add_segment_value_share(df_base, value_col="shipped_revenue", out_col="shipped_revenue_share_segment_pct")

    if df_base.empty:
        st.info("Нет данных для среднего месяца.")
    else:
        df_base_show = df_base.copy()
        if is_real_estate_preset:
            df_base_show = compute_real_estate_rates(df_base_show)
        df_base_show["ctr_pct"] = df_base_show["ctr"] * 100
        df_base_show["cr_pct"] = df_base_show["cr"] * 100
        df_base_show["drr_pct"] = df_base_show["drr"] * 100

        total_imp = df_base["impressions"].sum()
        total_reach = float(df_base["reach"].sum()) if "reach" in df_base.columns else 0.0
        total_frequency = (total_imp / total_reach) if total_reach > 0 else 0.0
        total_clicks = df_base["clicks"].sum()
        total_conv = df_base["conversions"].sum()
        total_leads = float(df_base["leads"].fillna(0.0).sum()) if "leads" in df_base.columns else 0.0
        total_target_leads = float(df_base["target_leads"].fillna(0.0).sum()) if "target_leads" in df_base.columns else float(total_conv)
        total_cost = df_base["cost"].sum()
        total_cost_with_vat = df_base["cost_with_vat"].sum()
        total_ak_wo_vat = df_base["ak_cost_wo_vat"].sum()
        total_cost_with_vat_ak = df_base["cost_with_vat_ak"].sum()
        total_rev = df_base["revenue"].sum()
        if use_ak_budget_metrics:
            total_budget_basis = total_cost_with_vat_ak if use_vat_budget_metrics else (total_cost + total_ak_wo_vat)
        else:
            total_budget_basis = total_cost_with_vat if use_vat_budget_metrics else total_cost
        if is_diy_preset:
            total_cpa = (total_cost_with_vat / total_conv) if total_conv > 0 else 0
        else:
            total_cpa = (total_budget_basis / total_conv) if total_conv > 0 else 0

        total_ctr = (total_clicks / total_imp * 100) if total_imp > 0 else 0
        total_cpc = (total_cost / total_clicks) if total_clicks > 0 else 0
        total_cr = (total_conv / total_clicks * 100) if total_clicks > 0 else 0
        total_cr1 = (total_leads / total_clicks * 100) if total_clicks > 0 else 0
        total_cr2 = (total_target_leads / total_leads * 100) if total_leads > 0 else 0
        total_cpm = (total_budget_basis / (total_imp / 1000)) if total_imp > 0 else 0
        total_cpl = (total_budget_basis / total_leads) if total_leads > 0 else 0
        total_cpql = (total_budget_basis / total_target_leads) if total_target_leads > 0 else 0
        total_roas = (total_rev / total_budget_basis) if total_budget_basis > 0 else 0
        total_drr = (total_budget_basis / total_rev * 100) if total_rev > 0 else 0
        st.session_state["mp_ref_base"] = {
            "impressions": float(total_imp),
            "reach": float(total_reach),
            "frequency": float(total_frequency),
            "clicks": float(total_clicks),
            "conversions": float(total_conv),
            "leads": float(total_leads),
            "target_leads": float(total_target_leads),
            "cost": float(total_cost),
            "cost_with_vat": float(total_cost_with_vat),
            "cost_with_vat_ak": float(total_cost_with_vat_ak),
            "revenue": float(total_rev),
            "ctr": float(total_ctr),
            "cpc": float(total_cpc),
            "cr": float(total_cr),
            "cr1": float(total_cr1),
            "cr2": float(total_cr2),
            "cpo": float(total_cpa),
            "cpl": float(total_cpl),
            "cpql": float(total_cpql),
            "roas": float(total_roas),
            "drr": float(total_drr),
        }
        base_by_campaign = {}
        for _, r in df_base.iterrows():
            camp = str(r.get("campaign_type", "")).strip()
            if not camp:
                continue
            imp = float(r.get("impressions", 0.0))
            reach = float(r.get("reach", 0.0) or 0.0)
            frequency = (imp / reach) if reach > 0 else 0.0
            clicks = float(r.get("clicks", 0.0))
            conv = float(r.get("conversions", 0.0))
            cost = float(r.get("cost", 0.0))
            cost_with_vat = float(r.get("cost_with_vat", 0.0))
            cost_with_vat_ak = float(r.get("cost_with_vat_ak", 0.0))
            rev = float(r.get("revenue", 0.0))
            if use_ak_budget_metrics:
                budget_basis = cost_with_vat_ak if use_vat_budget_metrics else float(r.get("cost", 0.0) + r.get("ak_cost_wo_vat", 0.0))
            else:
                budget_basis = cost_with_vat if use_vat_budget_metrics else cost
            ctr = (clicks / imp * 100.0) if imp > 0 else 0.0
            cpc = (cost / clicks) if clicks > 0 else 0.0
            cr = (conv / clicks * 100.0) if clicks > 0 else 0.0
            leads = float(r.get("leads", 0.0) or 0.0)
            target_leads = float(r.get("target_leads", conv) or 0.0)
            cr1 = (leads / clicks * 100.0) if clicks > 0 else 0.0
            cr2 = (target_leads / leads * 100.0) if leads > 0 else 0.0
            cpo = (budget_basis / conv) if conv > 0 else 0.0
            cpl = (budget_basis / leads) if leads > 0 else 0.0
            cpql = (budget_basis / target_leads) if target_leads > 0 else 0.0
            roas = (rev / budget_basis) if budget_basis > 0 else 0.0
            drr = (budget_basis / rev * 100.0) if rev > 0 else 0.0
            base_by_campaign[camp] = {
                "impressions": imp,
                "reach": reach,
                "frequency": frequency,
                "clicks": clicks,
                "conversions": conv,
                "leads": leads,
                "target_leads": target_leads,
                "cost": cost,
                "cost_with_vat": cost_with_vat,
                "cost_with_vat_ak": cost_with_vat_ak,
                "revenue": rev,
                "ctr": ctr,
                "cpc": cpc,
                "cr": cr,
                "cr1": cr1,
                "cr2": cr2,
                "cpo": cpo,
                "cpl": cpl,
                "cpql": cpql,
                "roas": roas,
                "drr": drr,
            }
        st.session_state["mp_ref_base_by_campaign"] = base_by_campaign

        total_row_raw = {
            "campaign_type": "Итого",
            "segment": "ALL",
            "system": "",
            "format": "",
            "impressions": total_imp,
            "ctr": total_ctr / 100,
            "cpc": total_cpc,
            "cr": total_cr / 100,
            "aov": None,
            "clicks": total_clicks,
            "conversions": total_conv,
            "leads": total_leads,
            "target_leads": total_target_leads,
            "cost": total_cost,
            "ak_cost_wo_vat": total_ak_wo_vat,
            "total_budget_wo_vat_ak": total_cost + total_ak_wo_vat,
            "cost_with_vat": total_cost_with_vat,
            "cost_with_vat_ak": total_cost_with_vat_ak,
            "vat_amount": total_cost_with_vat_ak - (total_cost + total_ak_wo_vat),
            "revenue": total_rev,
            "cpm": total_cpm,
            "cpa": total_cpa,
            "cpl": total_cpl,
            "cpql": total_cpql,
            "roas": total_roas,
            "drr": total_drr / 100,
            "ctr_pct": total_ctr,
            "cr_pct": total_cr,
            "cr1_pct": total_cr1,
            "cr2_pct": total_cr2,
            "drr_pct": total_drr,
        }
        if is_diy_preset:
            total_capacity = float(df_base["available_capacity"].sum()) if "available_capacity" in df_base.columns else 0.0
            total_row_raw["available_capacity"] = total_capacity
            total_row_raw["client_count"] = float(df_base["client_count"].sum()) if "client_count" in df_base.columns else 0.0
            total_row_raw["absolute_new_clients"] = float(df_base["absolute_new_clients"].sum()) if "absolute_new_clients" in df_base.columns else 0.0
            total_row_raw["returned_clients"] = float(df_base["returned_clients"].sum()) if "returned_clients" in df_base.columns else 0.0
            total_row_raw["new_clients"] = float(df_base["new_clients"].sum()) if "new_clients" in df_base.columns else 0.0
            total_row_raw["order_frequency"] = float(df_base["order_frequency"].mean()) if "order_frequency" in df_base.columns else 0.0
            total_row_raw["sov_pct"] = (total_reach / total_capacity * 100.0) if total_capacity > 0 else 0.0
            total_row_raw["new_clients_share_pct"] = calc_new_clients_share_pct(total_row_raw["new_clients"], total_row_raw["client_count"])
            total_row_raw["cac"] = (total_cost_with_vat_ak / float(total_row_raw["new_clients"])) if float(total_row_raw["new_clients"]) > 0 else 0.0
            total_row_raw["shipped_orders"] = float(df_base["shipped_orders"].sum()) if "shipped_orders" in df_base.columns else 0.0
            total_row_raw["shipped_revenue"] = float(df_base["shipped_revenue"].sum()) if "shipped_revenue" in df_base.columns else 0.0
            total_row_raw["shipped_cps"] = (total_cost_with_vat / float(total_row_raw["shipped_orders"])) if float(total_row_raw["shipped_orders"]) > 0 else 0.0
            total_row_raw["shipped_aov"] = (float(total_row_raw["shipped_revenue"]) / float(total_row_raw["shipped_orders"])) if float(total_row_raw["shipped_orders"]) > 0 else 0.0
            total_row_raw["shipped_roas"] = (float(total_row_raw["shipped_revenue"]) / total_budget_basis) if total_budget_basis > 0 else 0.0
            total_row_raw["shipped_drr_pct"] = (total_budget_basis / float(total_row_raw["shipped_revenue"]) * 100.0) if float(total_row_raw["shipped_revenue"]) > 0 else 0.0
            total_row_raw["order_share_segment_pct"] = 100.0
            total_row_raw["revenue_share_segment_pct"] = 100.0
            total_row_raw["shipped_order_share_segment_pct"] = 100.0
            total_row_raw["shipped_revenue_share_segment_pct"] = 100.0
            total_row_raw["budget_share_segment_pct"] = 100.0
        segment_total_rows = []
        if show_segment_subtotals and "segment" in df_base.columns:
            for seg_name, seg_df in df_base.groupby("segment"):
                seg_imp = seg_df["impressions"].sum()
                seg_clicks = seg_df["clicks"].sum()
                seg_conv = seg_df["conversions"].sum()
                seg_leads = float(seg_df["leads"].fillna(0.0).sum()) if "leads" in seg_df.columns else 0.0
                seg_target_leads = float(seg_df["target_leads"].fillna(0.0).sum()) if "target_leads" in seg_df.columns else float(seg_conv)
                seg_cost = seg_df["cost"].sum()
                seg_cost_with_vat = seg_df["cost_with_vat"].sum()
                seg_cost_with_vat_ak = seg_df["cost_with_vat_ak"].sum()
                seg_ak_wo_vat = seg_df["ak_cost_wo_vat"].sum()
                seg_rev = seg_df["revenue"].sum()
                if use_ak_budget_metrics:
                    seg_budget_basis = seg_cost_with_vat_ak if use_vat_budget_metrics else (seg_cost + seg_ak_wo_vat)
                else:
                    seg_budget_basis = seg_cost_with_vat if use_vat_budget_metrics else seg_cost
                seg_ctr = (seg_clicks / seg_imp * 100) if seg_imp > 0 else 0
                seg_cpc = (seg_cost / seg_clicks) if seg_clicks > 0 else 0
                seg_cr = (seg_conv / seg_clicks * 100) if seg_clicks > 0 else 0
                seg_cr1 = (seg_leads / seg_clicks * 100) if seg_clicks > 0 else 0
                seg_cr2 = (seg_target_leads / seg_leads * 100) if seg_leads > 0 else 0
                seg_cpm = (seg_budget_basis / (seg_imp / 1000)) if seg_imp > 0 else 0
                seg_cpa = (seg_cost_with_vat / seg_conv) if (is_diy_preset and seg_conv > 0) else (seg_budget_basis / seg_conv) if seg_conv > 0 else 0
                seg_cpl = (seg_budget_basis / seg_leads) if seg_leads > 0 else 0
                seg_cpql = (seg_budget_basis / seg_target_leads) if seg_target_leads > 0 else 0
                seg_roas = (seg_rev / seg_budget_basis) if seg_budget_basis > 0 else 0
                seg_drr = (seg_budget_basis / seg_rev * 100) if seg_rev > 0 else 0
                segment_total_rows.append(
                    {
                        "campaign_type": f"Итого {seg_name}",
                        "segment": seg_name,
                        "system": "",
                        "format": "",
                        "impressions": seg_imp,
                        "reach": float(seg_df["reach"].sum()) if "reach" in seg_df.columns else 0.0,
                        "frequency": (seg_imp / float(seg_df["reach"].sum())) if ("reach" in seg_df.columns and float(seg_df["reach"].sum()) > 0) else 0.0,
                        "ctr": seg_ctr / 100,
                        "cpc": seg_cpc,
                        "cr": seg_cr / 100,
                        "aov": None,
                        "clicks": seg_clicks,
                        "conversions": seg_conv,
                        "leads": seg_leads,
                        "target_leads": seg_target_leads,
                        "cost": seg_cost,
                        "ak_cost_wo_vat": seg_ak_wo_vat,
                        "total_budget_wo_vat_ak": seg_cost + seg_ak_wo_vat,
                        "cost_with_vat": seg_cost_with_vat,
                        "cost_with_vat_ak": seg_cost_with_vat_ak,
                        "vat_amount": seg_cost_with_vat_ak - (seg_cost + seg_ak_wo_vat),
                        "revenue": seg_rev,
                        "cpm": seg_cpm,
                        "cpa": seg_cpa,
                        "cpl": seg_cpl,
                        "cpql": seg_cpql,
                        "roas": seg_roas,
                        "drr": seg_drr / 100,
                        "ctr_pct": seg_ctr,
                        "cr_pct": seg_cr,
                        "cr1_pct": seg_cr1,
                        "cr2_pct": seg_cr2,
                        "drr_pct": seg_drr,
                        "ak_rate_pct": (seg_ak_wo_vat / seg_cost * 100.0) if seg_cost > 0 else 0.0,
                    }
                )
                if is_diy_preset:
                    seg_capacity = float(seg_df["available_capacity"].sum()) if "available_capacity" in seg_df.columns else 0.0
                    segment_total_rows[-1]["available_capacity"] = seg_capacity
                    segment_total_rows[-1]["client_count"] = float(seg_df["client_count"].sum()) if "client_count" in seg_df.columns else 0.0
                    segment_total_rows[-1]["absolute_new_clients"] = float(seg_df["absolute_new_clients"].sum()) if "absolute_new_clients" in seg_df.columns else 0.0
                    segment_total_rows[-1]["returned_clients"] = float(seg_df["returned_clients"].sum()) if "returned_clients" in seg_df.columns else 0.0
                    segment_total_rows[-1]["new_clients"] = float(seg_df["new_clients"].sum()) if "new_clients" in seg_df.columns else 0.0
                    segment_total_rows[-1]["order_frequency"] = float(seg_df["order_frequency"].mean()) if "order_frequency" in seg_df.columns else 0.0
                    segment_total_rows[-1]["sov_pct"] = (float(segment_total_rows[-1].get("reach", 0.0)) / seg_capacity * 100.0) if seg_capacity > 0 else 0.0
                    segment_total_rows[-1]["new_clients_share_pct"] = calc_new_clients_share_pct(
                        segment_total_rows[-1]["new_clients"],
                        segment_total_rows[-1]["client_count"],
                    )
                    segment_total_rows[-1]["cac"] = (seg_cost_with_vat_ak / float(segment_total_rows[-1]["new_clients"])) if float(segment_total_rows[-1]["new_clients"]) > 0 else 0.0
                    segment_total_rows[-1]["shipped_orders"] = float(seg_df["shipped_orders"].sum()) if "shipped_orders" in seg_df.columns else 0.0
                    segment_total_rows[-1]["shipped_revenue"] = float(seg_df["shipped_revenue"].sum()) if "shipped_revenue" in seg_df.columns else 0.0
                    segment_total_rows[-1]["shipped_cps"] = (seg_cost_with_vat / float(segment_total_rows[-1]["shipped_orders"])) if float(segment_total_rows[-1]["shipped_orders"]) > 0 else 0.0
                    segment_total_rows[-1]["shipped_aov"] = (float(segment_total_rows[-1]["shipped_revenue"]) / float(segment_total_rows[-1]["shipped_orders"])) if float(segment_total_rows[-1]["shipped_orders"]) > 0 else 0.0
                    segment_total_rows[-1]["shipped_roas"] = (float(segment_total_rows[-1]["shipped_revenue"]) / seg_budget_basis) if seg_budget_basis > 0 else 0.0
                    segment_total_rows[-1]["shipped_drr_pct"] = (seg_budget_basis / float(segment_total_rows[-1]["shipped_revenue"]) * 100.0) if float(segment_total_rows[-1]["shipped_revenue"]) > 0 else 0.0
                    segment_total_rows[-1]["order_share_segment_pct"] = 100.0
                    segment_total_rows[-1]["revenue_share_segment_pct"] = 100.0
                    segment_total_rows[-1]["shipped_order_share_segment_pct"] = 100.0
                    segment_total_rows[-1]["shipped_revenue_share_segment_pct"] = 100.0
                    segment_total_rows[-1]["budget_share_segment_pct"] = 100.0

        extra_base_rows = pd.DataFrame.from_records(
            segment_total_rows + [total_row_raw],
            columns=df_base_show.columns,
        )
        df_base_show = pd.DataFrame.from_records(
            df_base_show.to_dict("records") + extra_base_rows.to_dict("records"),
            columns=df_base_show.columns,
        )

        df_base_show["impressions"] = df_base_show["impressions"].map(
            lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
        )
        df_base_show["clicks"] = df_base_show["clicks"].map(
            lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
        )
        if "reach" in df_base_show.columns:
            df_base_show["reach"] = df_base_show["reach"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
            )
        if "frequency" in df_base_show.columns:
            df_base_show["frequency"] = df_base_show["frequency"].map(
                lambda x: "" if pd.isna(x) else f"{x:.2f}"
            )
        df_base_show["conversions"] = df_base_show["conversions"].map(
            lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
        )
        if "leads" in df_base_show.columns:
            df_base_show["leads"] = df_base_show["leads"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
            )
        if "target_leads" in df_base_show.columns:
            df_base_show["target_leads"] = df_base_show["target_leads"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
            )

        df_base_show["cost"] = df_base_show["cost"].map(
            lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
        )
        if "ak_cost_wo_vat" in df_base_show.columns:
            df_base_show["ak_cost_wo_vat"] = df_base_show["ak_cost_wo_vat"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
            )
        if "total_budget_wo_vat_ak" in df_base_show.columns:
            df_base_show["total_budget_wo_vat_ak"] = df_base_show["total_budget_wo_vat_ak"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
            )
        df_base_show["cost_with_vat"] = df_base_show["cost_with_vat"].map(
            lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
        )
        df_base_show["cost_with_vat_ak"] = df_base_show["cost_with_vat_ak"].map(
            lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
        )
        if "vat_amount" in df_base_show.columns:
            df_base_show["vat_amount"] = df_base_show["vat_amount"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
            )
        df_base_show["ak_rate_pct"] = df_base_show["ak_rate_pct"].map(
            lambda x: "" if pd.isna(x) else f"{x:.2f} %"
        )
        df_base_show["revenue"] = df_base_show["revenue"].map(
            lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
        )
        df_base_show["cpc"] = df_base_show["cpc"].map(
            lambda x: "" if pd.isna(x) else f"{x:.2f} ₽".replace(",", " ")
        )
        df_base_show["cpm"] = df_base_show["cpm"].map(
            lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
        )
        df_base_show["cpa"] = df_base_show["cpa"].map(
            lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
        )
        df_base_show["aov"] = df_base_show["aov"].map(
            lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
        )
        if "cpl" in df_base_show.columns:
            df_base_show["cpl"] = df_base_show["cpl"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
            )
        if "cpql" in df_base_show.columns:
            df_base_show["cpql"] = df_base_show["cpql"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
            )

        df_base_show["ctr_pct"] = df_base_show["ctr_pct"].map(
            lambda x: "" if pd.isna(x) else f"{x:.2f} %"
        )
        df_base_show["cr_pct"] = df_base_show["cr_pct"].map(
            lambda x: "" if pd.isna(x) else f"{x:.2f} %"
        )
        if "cr1_pct" in df_base_show.columns:
            df_base_show["cr1_pct"] = df_base_show["cr1_pct"].map(
                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
            )
        if "cr2_pct" in df_base_show.columns:
            df_base_show["cr2_pct"] = df_base_show["cr2_pct"].map(
                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
            )
        df_base_show["drr_pct"] = df_base_show["drr_pct"].map(
            lambda x: "" if pd.isna(x) else f"{x:.2f} %"
        )
        df_base_show["ROAS"] = df_base_show["roas"].map(
            lambda x: "" if pd.isna(x) else f"{x:.2f}"
        )

        if is_real_estate_preset:
            base_show_cols = ["campaign_type", "system", "format", "geo"] + get_real_estate_table_cols(metric_mode)
        else:
            base_show_cols = [
                "campaign_type",
                "system",
                "format",
                "geo",
                "cost",
                "ak_cost_wo_vat",
                "total_budget_wo_vat_ak",
                "cost_with_vat",
                "cost_with_vat_ak",
                "vat_amount",
                "ak_rate_pct",
                "impressions",
                "clicks",
                "cpc",
                "ctr_pct",
                "cr_pct",
                "conversions",
                "cpa",
                "aov",
                "revenue",
                "drr_pct",
                "ROAS",
            ]
        if is_diy_preset:
            base_show_cols.insert(3, "segment")
        if is_diy_preset:
            df_base_show["available_capacity"] = df_base_show["available_capacity"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
            )
            df_base_show["client_count"] = df_base_show["client_count"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
            )
            df_base_show["absolute_new_clients"] = df_base_show["absolute_new_clients"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
            )
            df_base_show["returned_clients"] = df_base_show["returned_clients"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
            )
            df_base_show["new_clients"] = df_base_show["new_clients"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
            )
            df_base_show["cac"] = df_base_show["cac"].map(
                lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
            )
            df_base_show["order_frequency"] = df_base_show["order_frequency"].map(
                lambda x: "" if pd.isna(x) else f"{x:.2f}"
            )
            df_base_show["sov_pct"] = df_base_show["sov_pct"].map(
                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
            )
            df_base_show["new_clients_share_pct"] = df_base_show["new_clients_share_pct"].map(
                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
            )
            df_base_show["order_share_segment_pct"] = df_base_show["order_share_segment_pct"].map(
                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
            )
            df_base_show["revenue_share_segment_pct"] = df_base_show["revenue_share_segment_pct"].map(
                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
            )
            df_base_show["budget_share_segment_pct"] = df_base_show["budget_share_segment_pct"].map(
                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
            )
            base_show_cols += ["available_capacity", "client_count", "absolute_new_clients", "returned_clients", "new_clients", "cac", "order_frequency", "shipped_orders", "shipped_cps", "shipped_aov", "shipped_revenue", "shipped_roas", "shipped_drr_pct", "shipped_order_share_segment_pct", "shipped_revenue_share_segment_pct", "sov_pct", "new_clients_share_pct", "order_share_segment_pct", "revenue_share_segment_pct", "budget_share_segment_pct"]
        # Порядок метрик задан под бизнес-логику проверки.
        df_base_show = safe_select_columns(df_base_show, base_show_cols, fill_value="")
        df_base_show = df_base_show.rename(columns=DISPLAY_COL_RENAME)
        if is_diy_preset:
            df_base_show = reorder_rows_with_segment_subtotals(
                df_base_show,
                DISPLAY_COL_RENAME["campaign_type"],
                DISPLAY_COL_RENAME["segment"],
            )
        # Дальше в итоговую строку записываются уже форматированные строки,
        # поэтому заранее переводим display-таблицу в object.
        df_base_show = df_base_show.astype(object)
        # Принудительно фиксируем значения в последней строке Итого (после всех преобразований).
        if len(df_base_show) > 0:
            li = len(df_base_show) - 1
            df_base_show.at[li, DISPLAY_COL_RENAME["campaign_type"]] = "Итого"
            if DISPLAY_COL_RENAME.get("segment") in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["segment"]] = "ALL"
            df_base_show.at[li, DISPLAY_COL_RENAME["system"]] = ""
            df_base_show.at[li, DISPLAY_COL_RENAME["format"]] = ""
            if DISPLAY_COL_RENAME["geo"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["geo"]] = ""
            if DISPLAY_COL_RENAME["impressions"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["impressions"]] = f"{round(total_imp):,}".replace(",", " ")
            if DISPLAY_COL_RENAME["clicks"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["clicks"]] = f"{round(total_clicks):,}".replace(",", " ")
            if DISPLAY_COL_RENAME["reach"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["reach"]] = f"{round(total_reach):,}".replace(",", " ")
            if DISPLAY_COL_RENAME["frequency"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["frequency"]] = f"{total_frequency:.2f}"
            if DISPLAY_COL_RENAME["conversions"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["conversions"]] = f"{round(total_conv):,}".replace(",", " ")
            if DISPLAY_COL_RENAME["leads"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["leads"]] = f"{round(total_leads):,}".replace(",", " ")
            if DISPLAY_COL_RENAME["target_leads"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["target_leads"]] = f"{round(total_target_leads):,}".replace(",", " ")
            if DISPLAY_COL_RENAME["cost"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["cost"]] = f"{round(total_cost):,} ₽".replace(",", " ")
            if DISPLAY_COL_RENAME["ak_cost_wo_vat"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["ak_cost_wo_vat"]] = f"{round(total_ak_wo_vat):,} ₽".replace(",", " ")
            if DISPLAY_COL_RENAME["total_budget_wo_vat_ak"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["total_budget_wo_vat_ak"]] = f"{round(total_cost + total_ak_wo_vat):,} ₽".replace(",", " ")
            if DISPLAY_COL_RENAME["cost_with_vat"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["cost_with_vat"]] = f"{round(total_cost_with_vat):,} ₽".replace(",", " ")
            if DISPLAY_COL_RENAME["cost_with_vat_ak"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["cost_with_vat_ak"]] = f"{round(total_cost_with_vat_ak):,} ₽".replace(",", " ")
            if DISPLAY_COL_RENAME["vat_amount"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["vat_amount"]] = f"{round(total_cost_with_vat_ak - (total_cost + total_ak_wo_vat)):,} ₽".replace(",", " ")
            total_ak_rate_pct = (total_ak_wo_vat / total_cost * 100.0) if total_cost > 0 else 0.0
            if DISPLAY_COL_RENAME["ak_rate_pct"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["ak_rate_pct"]] = f"{total_ak_rate_pct:.2f} %"
            if DISPLAY_COL_RENAME["revenue"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["revenue"]] = f"{round(total_rev):,} ₽".replace(",", " ")
            if DISPLAY_COL_RENAME["ctr"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["ctr"]] = f"{total_ctr:.2f} %"
            if DISPLAY_COL_RENAME["cpc"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["cpc"]] = f"{total_cpc:.2f} ₽"
            if DISPLAY_COL_RENAME["cr"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["cr"]] = f"{total_cr:.2f} %"
            if DISPLAY_COL_RENAME["cr1_pct"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["cr1_pct"]] = f"{total_cr1:.2f} %"
            if DISPLAY_COL_RENAME["cr2_pct"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["cr2_pct"]] = f"{total_cr2:.2f} %"
            if DISPLAY_COL_RENAME["cpa"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["cpa"]] = f"{round(total_cpa):,} ₽".replace(",", " ")
            if DISPLAY_COL_RENAME["cpl"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["cpl"]] = f"{round(total_cpl):,} ₽".replace(",", " ")
            if DISPLAY_COL_RENAME["cpql"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["cpql"]] = f"{round(total_cpql):,} ₽".replace(",", " ")
            if "ROAS" in df_base_show.columns:
                df_base_show.at[li, "ROAS"] = f"{total_roas:.2f}"
            if DISPLAY_COL_RENAME["drr"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["drr"]] = f"{total_drr:.2f} %"
            if DISPLAY_COL_RENAME["aov"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["aov"]] = ""
            if DISPLAY_COL_RENAME["available_capacity"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["available_capacity"]] = f"{round(total_row_raw.get('available_capacity', 0.0)):,}".replace(",", " ")
            if DISPLAY_COL_RENAME["client_count"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["client_count"]] = f"{round(total_row_raw.get('client_count', 0.0)):,}".replace(",", " ")
            if DISPLAY_COL_RENAME["absolute_new_clients"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["absolute_new_clients"]] = f"{round(total_row_raw.get('absolute_new_clients', 0.0)):,}".replace(",", " ")
            if DISPLAY_COL_RENAME["returned_clients"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["returned_clients"]] = f"{round(total_row_raw.get('returned_clients', 0.0)):,}".replace(",", " ")
            if DISPLAY_COL_RENAME["new_clients"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["new_clients"]] = f"{round(total_row_raw.get('new_clients', 0.0)):,}".replace(",", " ")
            if DISPLAY_COL_RENAME["cac"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["cac"]] = f"{round(total_row_raw.get('cac', 0.0)):,} ₽".replace(",", " ")
            if DISPLAY_COL_RENAME["order_frequency"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["order_frequency"]] = f"{float(total_row_raw.get('order_frequency', 0.0)):.2f}"
            if DISPLAY_COL_RENAME["sov_pct"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["sov_pct"]] = f"{float(total_row_raw.get('sov_pct', 0.0)):.2f} %"
            if DISPLAY_COL_RENAME["new_clients_share_pct"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["new_clients_share_pct"]] = f"{float(total_row_raw.get('new_clients_share_pct', 0.0)):.2f} %"
            if DISPLAY_COL_RENAME["order_share_segment_pct"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["order_share_segment_pct"]] = "100.00 %"
            if DISPLAY_COL_RENAME["revenue_share_segment_pct"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["revenue_share_segment_pct"]] = "100.00 %"
            if DISPLAY_COL_RENAME["budget_share_segment_pct"] in df_base_show.columns:
                df_base_show.at[li, DISPLAY_COL_RENAME["budget_share_segment_pct"]] = "100.00 %"

        # Таблица уже полностью подготовлена для отображения,
        # поэтому приводим значения к строкам и убираем NaN, чтобы Streamlit
        # не ловил проблемы сериализации при смешении чисел и форматированных строк.
        df_base_show = df_base_show.where(pd.notna(df_base_show), "")
        df_base_show = df_base_show.astype(str)

        def highlight_total_row(row):
            style = [""] * len(row)
            camp_col = DISPLAY_COL_RENAME.get("campaign_type", "Название кампании")
            camp_val = str(row.get(camp_col, ""))
            if camp_val == "Итого":
                style = [
                    f"background-color: #00CDC5; color: #081521; font-weight: 700; border-top: 2px solid {THEME_BORDER};"
                ] * len(row)
            elif camp_val.startswith("Итого "):
                style = [
                    f"background-color: #2C6E75; color: #DDEAF0; font-weight: 650; border-top: 1px solid {THEME_BORDER};"
                ] * len(row)
            return style

        styled_base = df_base_show.style.apply(highlight_total_row, axis=1)
        st.dataframe(styled_base, width="stretch")

    render_bottom_tab_switcher("Настройки расчета", "setup")

with tab_plan:
    st.markdown(
        """
        <div class="tab-intro">
            <p>1) Используйте полный блок «Средний месяц» как опорную базу для всех помесячных расчетов.</p>
            <p>2) Ниже работайте с коэффициентами и результатами по каждому выбранному месяцу.</p>
            <p>3) Коэффициенты в помесячной таблице можно редактировать прямо в ячейках и сразу видеть эффект на результат.</p>
            <p>4) После правок сверяйте TOTAL и переходите в «Диаграммы» для финальной проверки.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    ui_section_title("2. Средний месяц (базовые значения)")
    if df_base.empty:
        st.info("Нет данных для среднего месяца.")
    else:
        styled_base_plan = df_base_show.style.apply(highlight_total_row, axis=1)
        st.dataframe(styled_base_plan, width="stretch")

    # ---------- 3. Коэффициенты и результаты по месяцам ----------

    with st.expander("3. Коэффициенты и результаты по месяцам", expanded=True):
        show_month_details = True

        coeff_values = {}
        for cs in coeff_sets:
            cs_id = get_coeff_set_id_value(cs)
            if not cs_id:
                continue
            cs_type = normalize_coeff_set_type(cs.get("type", "Спрос (по запросам)"))
            df_res = cs.get("result")
            if df_res is None or df_res.empty:
                continue

            if cs_type == "Спрос (по запросам)":
                col_coeff = "Коэф."
                metric_type = "demand"
            elif cs_type == "AOV (средний чек)":
                col_coeff = "Коэф. AOV"
                metric_type = "aov"
            elif cs_type == "Медийные хвосты":
                col_coeff = "Коэф."
                metric_type = "media_tail"
            else:
                col_coeff = "Коэф."
                metric_type = "custom"

            if col_coeff not in df_res.columns or "Номер месяца" not in df_res.columns:
                continue

            for _, r in df_res.iterrows():
                try:
                    m = int(r["Номер месяца"])
                    k = float(r[col_coeff])
                except Exception:
                    continue
                if metric_type == "custom":
                    coeff_values[(cs_id, m, "demand")] = k
                    coeff_values[(cs_id, m, "aov")] = k
                else:
                    coeff_values[(cs_id, m, metric_type)] = k

        def get_k_demand(campaign_key: str, month_num: int) -> float:
            set_id = demand_link_map.get(campaign_key)
            if not set_id:
                return 1.0
            return coeff_values.get((set_id, month_num, "demand"), 1.0)

        def get_k_aov(campaign_key: str, month_num: int) -> float:
            set_id = aov_link_map.get(campaign_key)
            if not set_id:
                return 1.0
            return coeff_values.get((set_id, month_num, "aov"), 1.0)

        def get_k_media_tail(campaign_key: str, month_num: int) -> float:
            set_id = media_tail_link_map.get(campaign_key)
            if not set_id:
                return 1.0
            return coeff_values.get((set_id, month_num, "media_tail"), 1.0)

        def get_k_capacity(campaign_key: str, month_num: int) -> float:
            set_id = capacity_link_map.get(campaign_key)
            if not set_id:
                set_id = demand_link_map.get(campaign_key, "")
            if not set_id:
                return 1.0
            if (set_id, month_num, "demand") in coeff_values:
                return coeff_values[(set_id, month_num, "demand")]
            if (set_id, month_num, "media_tail") in coeff_values:
                return coeff_values[(set_id, month_num, "media_tail")]
            if (set_id, month_num, "aov") in coeff_values:
                return coeff_values[(set_id, month_num, "aov")]
            return 1.0

        def get_k_client_count(campaign_key: str, month_num: int) -> float:
            set_id = client_count_link_map.get(campaign_key)
            if not set_id:
                set_id = capacity_link_map.get(campaign_key, "")
            if not set_id:
                set_id = demand_link_map.get(campaign_key, "")
            if not set_id:
                return 1.0
            if (set_id, month_num, "demand") in coeff_values:
                return coeff_values[(set_id, month_num, "demand")]
            if (set_id, month_num, "media_tail") in coeff_values:
                return coeff_values[(set_id, month_num, "media_tail")]
            if (set_id, month_num, "aov") in coeff_values:
                return coeff_values[(set_id, month_num, "aov")]
            return 1.0

        def get_k_absolute_new_clients(campaign_key: str, month_num: int) -> float:
            set_id = absolute_new_clients_link_map.get(campaign_key)
            if not set_id:
                set_id = client_count_link_map.get(campaign_key, "")
            if not set_id:
                set_id = capacity_link_map.get(campaign_key, "")
            if not set_id:
                set_id = demand_link_map.get(campaign_key, "")
            if not set_id:
                return 1.0
            if (set_id, month_num, "demand") in coeff_values:
                return coeff_values[(set_id, month_num, "demand")]
            if (set_id, month_num, "media_tail") in coeff_values:
                return coeff_values[(set_id, month_num, "media_tail")]
            if (set_id, month_num, "aov") in coeff_values:
                return coeff_values[(set_id, month_num, "aov")]
            return 1.0

        def get_k_returned_clients(campaign_key: str, month_num: int) -> float:
            set_id = returned_clients_link_map.get(campaign_key)
            if not set_id:
                set_id = client_count_link_map.get(campaign_key, "")
            if not set_id:
                set_id = capacity_link_map.get(campaign_key, "")
            if not set_id:
                set_id = demand_link_map.get(campaign_key, "")
            if not set_id:
                return 1.0
            if (set_id, month_num, "demand") in coeff_values:
                return coeff_values[(set_id, month_num, "demand")]
            if (set_id, month_num, "media_tail") in coeff_values:
                return coeff_values[(set_id, month_num, "media_tail")]
            if (set_id, month_num, "aov") in coeff_values:
                return coeff_values[(set_id, month_num, "aov")]
            return 1.0

        def get_k_order_frequency(campaign_key: str, month_num: int) -> float:
            set_id = order_frequency_link_map.get(campaign_key)
            if not set_id:
                set_id = client_count_link_map.get(campaign_key, "")
            if not set_id:
                set_id = capacity_link_map.get(campaign_key, "")
            if not set_id:
                set_id = demand_link_map.get(campaign_key, "")
            if not set_id:
                return 1.0
            if (set_id, month_num, "demand") in coeff_values:
                return coeff_values[(set_id, month_num, "demand")]
            if (set_id, month_num, "media_tail") in coeff_values:
                return coeff_values[(set_id, month_num, "media_tail")]
            if (set_id, month_num, "aov") in coeff_values:
                return coeff_values[(set_id, month_num, "aov")]
            return 1.0

        def get_k_reach(campaign_key: str, month_num: int) -> float:
            set_id = demand_link_map.get(campaign_key, "")
            if not set_id:
                set_id = media_tail_link_map.get(campaign_key, "")
            if not set_id:
                return 1.0
            if (set_id, month_num, "demand") in coeff_values:
                return coeff_values[(set_id, month_num, "demand")]
            if (set_id, month_num, "media_tail") in coeff_values:
                return coeff_values[(set_id, month_num, "media_tail")]
            return 1.0

        all_months_results = []

        for period in selected_periods:
            planning_slot = int(period["planning_slot"])
            month = int(period["month_num"])
            month_year = int(period["month_year"])
            month_name = str(period["month_name"])

            coeff_rows = []
            base_capacity_map = {
                str(r["campaign_type"]): float(r.get("available_capacity_avg", 0.0) or 0.0)
                for _, r in campaigns.iterrows()
            }
            base_client_count_map = {
                str(r["campaign_type"]): float(r.get("client_count_avg", 0.0) or 0.0)
                for _, r in campaigns.iterrows()
            }
            base_absolute_new_clients_map = {
                str(r["campaign_type"]): float(r.get("absolute_new_clients_avg", 0.0) or 0.0)
                for _, r in campaigns.iterrows()
            }
            base_returned_clients_map = {
                str(r["campaign_type"]): float(r.get("returned_clients_avg", 0.0) or 0.0)
                for _, r in campaigns.iterrows()
            }
            base_order_frequency_map = {
                str(r["campaign_type"]): float(r.get("order_frequency_avg", 0.0) or 0.0)
                for _, r in campaigns.iterrows()
            }
            for _, base_row in campaigns.iterrows():
                camp_type = str(base_row["campaign_type"])
                campaign_key = get_campaign_identity_key_from_row(base_row)

                k_demand = get_k_demand(campaign_key, planning_slot)
                k_media_tail = get_k_media_tail(campaign_key, planning_slot)
                el = elasticity_map.get(
                    campaign_key,
                    {"cpc_div": 1.0, "ctr_div": 2.0, "cr_div": 10.0},
                )

                k_imp = k_demand * k_media_tail
                demand_delta = (k_demand - 1.0)
                k_cpc = 1.0 + demand_delta / el["cpc_div"]
                k_ctr = 1.0 - demand_delta / el["ctr_div"]
                k_cr = 1.0 - demand_delta / el["cr_div"]
                k_cr2 = 1.0 - demand_delta / el.get("cr2_div", el["cr_div"])
                k_ctr = max(0.0, k_ctr)
                k_cr = max(0.0, k_cr)
                k_cr2 = max(0.0, k_cr2)
                k_aov = get_k_aov(campaign_key, planning_slot)
                k_capacity = get_k_capacity(campaign_key, planning_slot)
                k_client_count = get_k_client_count(campaign_key, planning_slot)
                k_absolute_new_clients = get_k_absolute_new_clients(campaign_key, planning_slot)
                k_returned_clients = get_k_returned_clients(campaign_key, planning_slot)
                k_order_frequency = get_k_order_frequency(campaign_key, planning_slot)
                k_reach = get_k_reach(campaign_key, planning_slot)

                coeff_row = {
                    "campaign_key": campaign_key,
                    "campaign_type": camp_type,
                    "geo": str(base_row.get("geo", "")),
                    "segment": str(base_row.get("segment", "ALL")).upper(),
                    "k_imp": k_imp,
                    "k_reach": k_reach,
                    "k_capacity": k_capacity,
                    "available_capacity": float(base_row.get("available_capacity_avg", 0.0) or 0.0) * k_capacity,
                    "k_client_count": k_client_count,
                    "client_count": float(base_row.get("client_count_avg", 0.0) or 0.0) * k_client_count,
                    "k_absolute_new_clients": k_absolute_new_clients,
                    "absolute_new_clients": float(base_row.get("absolute_new_clients_avg", 0.0) or 0.0) * k_absolute_new_clients,
                    "k_returned_clients": k_returned_clients,
                    "returned_clients": float(base_row.get("returned_clients_avg", 0.0) or 0.0) * k_returned_clients,
                    "new_clients": (float(base_row.get("absolute_new_clients_avg", 0.0) or 0.0) * k_absolute_new_clients)
                    + (float(base_row.get("returned_clients_avg", 0.0) or 0.0) * k_returned_clients),
                    "k_order_frequency": k_order_frequency,
                    "order_frequency": float(base_row.get("order_frequency_avg", 0.0) or 0.0) * k_order_frequency,
                    "k_ctr": k_ctr,
                    "k_cpc": k_cpc,
                    "k_cr": k_cr,
                    "k_aov": k_aov,
                }
                if is_real_estate_preset and metric_mode["is_real_estate_full"]:
                    coeff_row["k_cr2"] = k_cr2
                coeff_rows.append(coeff_row)

            coeffs_default = pd.DataFrame(coeff_rows)
            if is_real_estate_preset:
                coeffs_default = coeffs_default.drop(columns=["k_aov"], errors="ignore")
            if not is_diy_preset:
                coeffs_default = coeffs_default.drop(
                    columns=["k_reach", "k_capacity", "available_capacity", "k_client_count", "client_count", "k_absolute_new_clients", "absolute_new_clients", "k_returned_clients", "returned_clients", "new_clients", "k_order_frequency", "order_frequency"],
                    errors="ignore",
                )
            else:
                coeffs_default = coeffs_default[
                    [
                        "campaign_key",
                        "campaign_type",
                        "geo",
                        "segment",
                        "k_imp",
                        "k_cpc",
                        "k_ctr",
                        "k_cr",
                        "k_aov",
                        "k_capacity",
                        "k_client_count",
                        "k_absolute_new_clients",
                        "k_returned_clients",
                        "k_order_frequency",
                    ]
                ]
            coeffs_month = coeffs_default.copy()
            coeffs_month["campaign_type"] = coeffs_month["campaign_type"].astype(str)
            if "campaign_key" in coeffs_month.columns:
                coeffs_month["campaign_key"] = coeffs_month["campaign_key"].astype(str)

            rows = []
            # если хотим показывать детали по месяцам, создаём редактор
            if show_month_details:
                st.markdown(f"#### Месяц {planning_slot}: {month_name}")
                col_left, col_right = st.columns([1.8, 1.0], vertical_alignment="top")
                month_row_height = max(36, int(ui_table_row_px))
                month_rows_count = max(len(coeffs_month), 1)
                month_table_height = max(140, min(520, 38 + month_rows_count * month_row_height + 6))
                with col_right:
                    edited_coeffs_month = st.data_editor(
                        coeffs_month.set_index("campaign_key") if "campaign_key" in coeffs_month.columns else coeffs_month,
                        num_rows="fixed",
                        width="stretch",
                        row_height=month_row_height,
                        height=month_table_height,
                        key=f"coeffs_month_{planning_slot}_v2",
                        hide_index=True,
                        column_config={
                            "campaign_type": st.column_config.TextColumn(
                                "Тип РК / Название", disabled=True, width="medium"
                            ),
                            "geo": st.column_config.TextColumn("ГЕО", disabled=True, width="small"),
                            "segment": st.column_config.TextColumn("Сегмент", disabled=True, width="small"),
                            "k_imp": st.column_config.NumberColumn(
                                "k_imp", format="%.2f", width="small",
                                help=mhelp("k_imp")
                            ),
                            "k_cpc": st.column_config.NumberColumn(
                                "k_cpc", format="%.2f", width="small",
                                help=mhelp("k_cpc")
                            ),
                            "k_reach": st.column_config.NumberColumn(
                                "k_reach", format="%.2f", width="small",
                                help=mhelp("k_reach")
                            ),
                            "k_ctr": st.column_config.NumberColumn(
                                "k_ctr", format="%.2f", width="small",
                                help=mhelp("k_ctr")
                            ),
                            "k_cr": st.column_config.NumberColumn(
                                "k_cr", format="%.2f", width="small",
                                help=mhelp("k_cr")
                            ),
                            "k_cr2": st.column_config.NumberColumn(
                                "k_cr2", format="%.2f", width="small",
                                help="Коэффициент влияния сезонности на CR2 в ЦО."
                            ),
                            "k_aov": st.column_config.NumberColumn(
                                "k_aov", format="%.2f", width="small",
                                help=mhelp("k_aov")
                            ),
                            "k_capacity": st.column_config.NumberColumn(
                                "k_capacity", format="%.2f", width="small",
                                disabled=not is_diy_preset,
                                help=mhelp("k_capacity"),
                            ),
                            "k_client_count": st.column_config.NumberColumn(
                                "k_client_count", format="%.2f", width="small",
                                disabled=not is_diy_preset,
                                help=mhelp("k_client_count"),
                            ),
                            "k_absolute_new_clients": st.column_config.NumberColumn(
                                "k_absolute_new_clients", format="%.2f", width="small",
                                disabled=not is_diy_preset,
                                help=mhelp("k_absolute_new_clients"),
                            ),
                            "k_returned_clients": st.column_config.NumberColumn(
                                "k_returned_clients", format="%.2f", width="small",
                                disabled=not is_diy_preset,
                                help=mhelp("k_returned_clients"),
                            ),
                            "k_order_frequency": st.column_config.NumberColumn(
                                "k_order_frequency", format="%.2f", width="small",
                                disabled=not is_diy_preset,
                                help=mhelp("k_order_frequency"),
                            ),
                        },
                    )
                    if "campaign_key" in coeffs_month.columns:
                        edited_coeffs_month = edited_coeffs_month.reset_index(names="campaign_key")
                    edited_coeffs_month = edited_coeffs_month.copy()
                    if is_real_estate_preset and "k_aov" in edited_coeffs_month.columns:
                        edited_coeffs_month = edited_coeffs_month.drop(columns=["k_aov"], errors="ignore")
                    if not ((is_real_estate_preset and metric_mode["is_real_estate_full"]) or is_diy_preset) and "k_cr2" in edited_coeffs_month.columns:
                        edited_coeffs_month = edited_coeffs_month.drop(columns=["k_cr2"], errors="ignore")
                    edited_coeffs_month["campaign_type"] = edited_coeffs_month["campaign_type"].astype(str)
                    if "campaign_key" in edited_coeffs_month.columns:
                        edited_coeffs_month["campaign_key"] = edited_coeffs_month["campaign_key"].astype(str)
                    coeffs_month = edited_coeffs_month
            else:
                # если подробности скрыты — берём coeffs_default как есть
                coeffs_month = coeffs_month.copy()
                col_left = st  # заглушка, чтобы ниже не падало

            for _, base_row in campaigns.iterrows():
                campaign_type = str(base_row["campaign_type"])
                campaign_key = get_campaign_identity_key_from_row(base_row)
                if "campaign_key" in coeffs_month.columns:
                    k_row = coeffs_month[coeffs_month["campaign_key"] == campaign_key]
                else:
                    k_row = coeffs_month[coeffs_month["campaign_type"] == campaign_type]
                if k_row.empty:
                    k_imp = k_reach = k_capacity = k_client_count = k_absolute_new_clients = k_returned_clients = k_order_frequency = k_ctr = k_cpc = k_cr = k_cr2 = k_aov = 1.0
                    available_capacity_month = float(base_row.get("available_capacity_avg", 0.0) or 0.0)
                    client_count_month = float(base_row.get("client_count_avg", 0.0) or 0.0)
                    absolute_new_clients_month = float(base_row.get("absolute_new_clients_avg", 0.0) or 0.0)
                    returned_clients_month = float(base_row.get("returned_clients_avg", 0.0) or 0.0)
                    order_frequency_month = float(base_row.get("order_frequency_avg", 0.0) or 0.0)
                    shipped_aov_month = float(base_row.get("shipped_aov_avg", 0.0) or 0.0)
                else:
                    k_row = k_row.iloc[0]
                    k_imp = float(k_row["k_imp"])
                    k_reach = float(k_row.get("k_reach", k_imp) or k_imp)
                    k_capacity = float(k_row.get("k_capacity", 1.0) or 1.0)
                    k_client_count = float(k_row.get("k_client_count", 1.0) or 1.0)
                    k_absolute_new_clients = float(k_row.get("k_absolute_new_clients", 1.0) or 1.0)
                    k_returned_clients = float(k_row.get("k_returned_clients", 1.0) or 1.0)
                    k_order_frequency = float(k_row.get("k_order_frequency", 1.0) or 1.0)
                    k_ctr = float(k_row["k_ctr"])
                    k_cpc = float(k_row["k_cpc"])
                    k_cr = float(k_row["k_cr"])
                    k_cr2 = float(k_row.get("k_cr2", k_cr) or k_cr)
                    k_aov = float(k_row.get("k_aov", 1.0) or 1.0)
                    available_capacity_month = float(base_row.get("available_capacity_avg", 0.0) or 0.0) * k_capacity
                    client_count_month = float(base_row.get("client_count_avg", 0.0) or 0.0) * k_client_count
                    absolute_new_clients_month = float(base_row.get("absolute_new_clients_avg", 0.0) or 0.0) * k_absolute_new_clients
                    returned_clients_month = float(base_row.get("returned_clients_avg", 0.0) or 0.0) * k_returned_clients
                    order_frequency_month = float(base_row.get("order_frequency_avg", 0.0) or 0.0) * k_order_frequency
                    shipped_aov_month = float(base_row.get("shipped_aov_avg", 0.0) or 0.0) * k_aov

                base = PlanInput(
                    impressions=base_row["impressions_avg"],
                    ctr=base_row["ctr_avg_percent"] / 100.0,
                    cpc=base_row["cpc_avg"],
                    cr=base_row["cr_avg_percent"] / 100.0,
                    aov=base_row["aov_avg"],
                    reach=float(base_row.get("reach_avg", 0.0) or 0.0),
                    cr2=float(base_row.get("cr2_avg_percent", 0.0) or 0.0) / 100.0,
                    preset_key=active_preset_key,
                    funnel_mode=metric_mode["real_estate_funnel_mode"],
                )

                month_inp = PlanInput(
                    impressions=base.impressions * k_imp,
                    ctr=base.ctr * k_ctr,
                    cpc=base.cpc * k_cpc,
                    cr=base.cr * k_cr,
                    aov=base.aov * k_aov,
                    reach=base.reach * k_reach,
                    cr2=base.cr2 * k_cr2 if ((is_real_estate_preset and metric_mode["is_real_estate_full"]) or is_diy_preset) else base.cr2,
                    preset_key=active_preset_key,
                    funnel_mode=metric_mode["real_estate_funnel_mode"],
                )

                out = calculate_plan_month(month_inp)
                out["planning_slot"] = planning_slot
                out["month_num"] = month
                out["month_year"] = month_year
                out["month_name"] = month_name
                out["campaign_type"] = campaign_type
                out["segment"] = str(base_row.get("segment", "B2C"))
                out["system"] = base_row["system"]
                out["format"] = base_row["format"]
                out["geo"] = str(base_row.get("geo", "") or "")
                out["k_demand_applied"] = float(k_demand)
                if is_diy_preset:
                    out["available_capacity"] = available_capacity_month
                    out["client_count"] = client_count_month
                    out["absolute_new_clients"] = absolute_new_clients_month
                    out["returned_clients"] = returned_clients_month
                    out["new_clients"] = absolute_new_clients_month + returned_clients_month
                    out["order_frequency"] = order_frequency_month
                    out["sov_pct"] = (float(out.get("reach", 0.0)) / available_capacity_month * 100.0) if available_capacity_month > 0 else 0.0
                    out["new_clients_share_pct"] = calc_new_clients_share_pct(out["new_clients"], out["client_count"])
                    out["cac"] = (float(out.get("cost_with_vat_ak", 0.0)) / float(out.get("new_clients", 0.0))) if float(out.get("new_clients", 0.0)) > 0 else 0.0
                    out["shipped_orders"] = float(out.get("target_leads", 0.0) or 0.0)
                    out["shipped_revenue"] = float(out.get("shipped_orders", 0.0)) * shipped_aov_month
                    out["shipped_aov"] = shipped_aov_month
                rows.append(out)

            df_month = pd.DataFrame(rows)
            if not df_month.empty:
                month_total_wo_vat = float(df_month["cost"].sum())
                month_ak_rate = resolve_ak_rate(month_total_wo_vat, ak_rules_df) if (use_ak_budget_metrics and ak_mode == "percent") else 0.0
                df_month = apply_budget_basis_metrics(
                    df_month,
                    use_vat_budget_metrics,
                    use_ak=use_ak_budget_metrics,
                    ak_mode=ak_mode,
                    default_ak_rate=month_ak_rate,
                    default_ak_fixed_wo_vat=float(ak_fixed_month_wo_vat),
                    default_ak_fixed_rate=float(ak_fixed_percent) / 100.0,
                )
                df_month["ak_rate_pct"] = df_month["ak_rate"] * 100.0
                if is_diy_preset:
                    df_month["cpa"] = np.where(
                        pd.to_numeric(df_month.get("conversions", 0.0), errors="coerce").fillna(0.0) > 0,
                        pd.to_numeric(df_month.get("cost_with_vat", 0.0), errors="coerce").fillna(0.0)
                        / pd.to_numeric(df_month.get("conversions", 0.0), errors="coerce").fillna(0.0),
                        0.0,
                    )
                    df_month["shipped_cps"] = np.where(
                        pd.to_numeric(df_month.get("shipped_orders", 0.0), errors="coerce").fillna(0.0) > 0,
                        pd.to_numeric(df_month.get("cost_with_vat", 0.0), errors="coerce").fillna(0.0)
                        / pd.to_numeric(df_month.get("shipped_orders", 0.0), errors="coerce").fillna(0.0),
                        0.0,
                    )
                    df_month["shipped_aov"] = np.where(
                        pd.to_numeric(df_month.get("shipped_orders", 0.0), errors="coerce").fillna(0.0) > 0,
                        pd.to_numeric(df_month.get("shipped_revenue", 0.0), errors="coerce").fillna(0.0)
                        / pd.to_numeric(df_month.get("shipped_orders", 0.0), errors="coerce").fillna(0.0),
                        0.0,
                    )
                    shipped_budget_basis_month = pd.to_numeric(
                        df_month.get(
                            "cost_with_vat_ak" if use_vat_budget_metrics and use_ak_budget_metrics else "cost_with_vat" if use_vat_budget_metrics else "total_budget_wo_vat_ak" if use_ak_budget_metrics else "cost",
                            0.0,
                        ),
                        errors="coerce",
                    ).fillna(0.0)
                    df_month["shipped_roas"] = np.where(
                        shipped_budget_basis_month > 0,
                        pd.to_numeric(df_month.get("shipped_revenue", 0.0), errors="coerce").fillna(0.0) / shipped_budget_basis_month,
                        0.0,
                    )
                    df_month["shipped_drr_pct"] = np.where(
                        pd.to_numeric(df_month.get("shipped_revenue", 0.0), errors="coerce").fillna(0.0) > 0,
                        shipped_budget_basis_month / pd.to_numeric(df_month.get("shipped_revenue", 0.0), errors="coerce").fillna(0.0) * 100.0,
                        0.0,
                    )
                    df_month["cac"] = np.where(
                        pd.to_numeric(df_month.get("new_clients", 0.0), errors="coerce").fillna(0.0) > 0,
                        pd.to_numeric(df_month.get("cost_with_vat_ak", 0.0), errors="coerce").fillna(0.0)
                        / pd.to_numeric(df_month.get("new_clients", 0.0), errors="coerce").fillna(0.0),
                        0.0,
                    )
                    df_month = add_segment_budget_share(df_month)
                    df_month = add_segment_value_share(df_month, value_col="conversions", out_col="order_share_segment_pct")
                    df_month = add_segment_value_share(df_month, value_col="revenue", out_col="revenue_share_segment_pct")
                    df_month = add_segment_value_share(df_month, value_col="shipped_orders", out_col="shipped_order_share_segment_pct")
                    df_month = add_segment_value_share(df_month, value_col="shipped_revenue", out_col="shipped_revenue_share_segment_pct")
            all_months_results.append(df_month)

            if show_month_details:
                with col_left:
                    if not df_month.empty:
                        df_rows_show = df_month.copy()
                        if is_real_estate_preset:
                            df_rows_show = compute_real_estate_rates(df_rows_show)

                        df_rows_show["ctr_pct"] = df_rows_show["ctr"] * 100
                        df_rows_show["cr_pct"] = df_rows_show["cr"] * 100
                        df_rows_show["drr_pct"] = df_rows_show["drr"] * 100

                        total_imp = df_month["impressions"].sum()
                        total_reach = float(df_month["reach"].sum()) if "reach" in df_month.columns else 0.0
                        total_frequency = (total_imp / total_reach) if total_reach > 0 else 0.0
                        total_clicks = df_month["clicks"].sum()
                        total_conv = df_month["conversions"].sum()
                        total_leads = float(df_month["leads"].fillna(0.0).sum()) if "leads" in df_month.columns else 0.0
                        total_target_leads = float(df_month["target_leads"].fillna(0.0).sum()) if "target_leads" in df_month.columns else float(total_conv)
                        total_cost = df_month["cost"].sum()
                        total_cost_with_vat = df_month["cost_with_vat"].sum()
                        total_cost_with_vat_ak = df_month["cost_with_vat_ak"].sum()
                        total_ak_wo_vat = df_month["ak_cost_wo_vat"].sum()
                        month_ak_rate_effective = (total_ak_wo_vat / total_cost) if total_cost > 0 else 0.0
                        total_rev = df_month["revenue"].sum()
                        if use_ak_budget_metrics:
                            total_budget_basis = total_cost_with_vat_ak if use_vat_budget_metrics else (total_cost + total_ak_wo_vat)
                        else:
                            total_budget_basis = total_cost_with_vat if use_vat_budget_metrics else total_cost

                        total_ctr = (total_clicks / total_imp * 100) if total_imp > 0 else 0
                        total_cpc = (total_cost / total_clicks) if total_clicks > 0 else 0
                        total_cr = (total_conv / total_clicks * 100) if total_clicks > 0 else 0
                        total_cr1 = (total_leads / total_clicks * 100) if total_clicks > 0 else 0
                        total_cr2 = (total_target_leads / total_leads * 100) if total_leads > 0 else 0
                        total_cpm = (total_budget_basis / (total_imp / 1000)) if total_imp > 0 else 0
                        total_cpa = (total_cost_with_vat / total_conv) if (is_diy_preset and total_conv > 0) else (total_budget_basis / total_conv) if total_conv > 0 else 0
                        total_cpl = (total_budget_basis / total_leads) if total_leads > 0 else 0
                        total_cpql = (total_cost_with_vat / total_target_leads) if (is_diy_preset and total_target_leads > 0) else (total_budget_basis / total_target_leads) if total_target_leads > 0 else 0
                        total_roas = (total_rev / total_budget_basis) if total_budget_basis > 0 else 0
                        total_drr = (total_budget_basis / total_rev * 100) if total_rev > 0 else 0

                        total_row_month = {
                            "campaign_type": "Итого",
                            "segment": "ALL",
                            "system": "",
                            "format": "",
                            "geo": "",
                            "impressions": total_imp,
                            "reach": total_reach,
                            "frequency": total_frequency,
                            "ctr": total_ctr / 100,
                            "cpc": total_cpc,
                            "cr": total_cr / 100,
                            "aov": None,
                            "clicks": total_clicks,
                            "conversions": total_conv,
                            "leads": total_leads,
                            "target_leads": total_target_leads,
                            "cost": total_cost,
                            "ak_cost_wo_vat": total_ak_wo_vat,
                            "total_budget_wo_vat_ak": total_cost + total_ak_wo_vat,
                            "cost_with_vat": total_cost_with_vat,
                            "cost_with_vat_ak": total_cost_with_vat_ak,
                            "vat_amount": total_cost_with_vat_ak - (total_cost + total_ak_wo_vat),
                            "ak_rate_pct": month_ak_rate_effective * 100.0,
                            "revenue": total_rev,
                            "cpm": total_cpm,
                            "cpa": total_cpa,
                            "cpl": total_cpl,
                            "cpql": total_cpql,
                            "roas": total_roas,
                            "drr": total_drr / 100,
                            "ctr_pct": total_ctr,
                            "cr_pct": total_cr,
                            "cr1_pct": total_cr1,
                            "cr2_pct": total_cr2,
                            "drr_pct": total_drr,
                        }
                        if is_diy_preset:
                            total_capacity = float(df_month["available_capacity"].sum()) if "available_capacity" in df_month.columns else 0.0
                            total_row_month["available_capacity"] = total_capacity
                            total_row_month["client_count"] = float(df_month["client_count"].sum()) if "client_count" in df_month.columns else 0.0
                            total_row_month["absolute_new_clients"] = float(df_month["absolute_new_clients"].sum()) if "absolute_new_clients" in df_month.columns else 0.0
                            total_row_month["returned_clients"] = float(df_month["returned_clients"].sum()) if "returned_clients" in df_month.columns else 0.0
                            total_row_month["new_clients"] = float(df_month["new_clients"].sum()) if "new_clients" in df_month.columns else 0.0
                            total_row_month["order_frequency"] = float(df_month["order_frequency"].mean()) if "order_frequency" in df_month.columns else 0.0
                            total_row_month["sov_pct"] = (total_reach / total_capacity * 100.0) if total_capacity > 0 else 0.0
                            total_row_month["new_clients_share_pct"] = calc_new_clients_share_pct(
                                total_row_month["new_clients"],
                                total_row_month["client_count"],
                            )
                            total_row_month["cac"] = (total_cost_with_vat_ak / float(total_row_month["new_clients"])) if float(total_row_month["new_clients"]) > 0 else 0.0
                            total_row_month["shipped_orders"] = float(df_month["shipped_orders"].sum()) if "shipped_orders" in df_month.columns else 0.0
                            total_row_month["shipped_revenue"] = float(df_month["shipped_revenue"].sum()) if "shipped_revenue" in df_month.columns else 0.0
                            total_row_month["shipped_cps"] = (total_cost_with_vat / float(total_row_month["shipped_orders"])) if float(total_row_month["shipped_orders"]) > 0 else 0.0
                            total_row_month["shipped_aov"] = (float(total_row_month["shipped_revenue"]) / float(total_row_month["shipped_orders"])) if float(total_row_month["shipped_orders"]) > 0 else 0.0
                            total_row_month["shipped_roas"] = (float(total_row_month["shipped_revenue"]) / total_budget_basis) if total_budget_basis > 0 else 0.0
                            total_row_month["shipped_drr_pct"] = (total_budget_basis / float(total_row_month["shipped_revenue"]) * 100.0) if float(total_row_month["shipped_revenue"]) > 0 else 0.0
                            total_row_month["order_share_segment_pct"] = 100.0
                            total_row_month["revenue_share_segment_pct"] = 100.0
                            total_row_month["shipped_order_share_segment_pct"] = 100.0
                            total_row_month["shipped_revenue_share_segment_pct"] = 100.0
                            total_row_month["budget_share_segment_pct"] = 100.0
                        segment_total_rows_month = []
                        if show_segment_subtotals and "segment" in df_month.columns:
                            for seg_name, seg_df in df_month.groupby("segment"):
                                seg_imp = seg_df["impressions"].sum()
                                seg_clicks = seg_df["clicks"].sum()
                                seg_conv = seg_df["conversions"].sum()
                                seg_leads = float(seg_df["leads"].fillna(0.0).sum()) if "leads" in seg_df.columns else 0.0
                                seg_target_leads = float(seg_df["target_leads"].fillna(0.0).sum()) if "target_leads" in seg_df.columns else float(seg_conv)
                                seg_cost = seg_df["cost"].sum()
                                seg_cost_with_vat = seg_df["cost_with_vat"].sum()
                                seg_cost_with_vat_ak = seg_df["cost_with_vat_ak"].sum()
                                seg_ak_wo_vat = seg_df["ak_cost_wo_vat"].sum()
                                seg_rev = seg_df["revenue"].sum()
                                if use_ak_budget_metrics:
                                    seg_budget_basis = seg_cost_with_vat_ak if use_vat_budget_metrics else (seg_cost + seg_ak_wo_vat)
                                else:
                                    seg_budget_basis = seg_cost_with_vat if use_vat_budget_metrics else seg_cost
                                seg_ctr = (seg_clicks / seg_imp * 100) if seg_imp > 0 else 0
                                seg_cpc = (seg_cost / seg_clicks) if seg_clicks > 0 else 0
                                seg_cr = (seg_conv / seg_clicks * 100) if seg_clicks > 0 else 0
                                seg_cr1 = (seg_leads / seg_clicks * 100) if seg_clicks > 0 else 0
                                seg_cr2 = (seg_target_leads / seg_leads * 100) if seg_leads > 0 else 0
                                seg_cpm = (seg_budget_basis / (seg_imp / 1000)) if seg_imp > 0 else 0
                                seg_cpa = (seg_cost_with_vat / seg_conv) if (is_diy_preset and seg_conv > 0) else (seg_budget_basis / seg_conv) if seg_conv > 0 else 0
                                seg_cpl = (seg_budget_basis / seg_leads) if seg_leads > 0 else 0
                                seg_cpql = (seg_budget_basis / seg_target_leads) if seg_target_leads > 0 else 0
                                seg_roas = (seg_rev / seg_budget_basis) if seg_budget_basis > 0 else 0
                                seg_drr = (seg_budget_basis / seg_rev * 100) if seg_rev > 0 else 0
                                seg_row = {
                                    "campaign_type": f"Итого {seg_name}",
                                    "segment": seg_name,
                                    "system": "",
                                    "format": "",
                                    "geo": "",
                                    "impressions": seg_imp,
                                    "reach": float(seg_df["reach"].sum()) if "reach" in seg_df.columns else 0.0,
                                    "frequency": (seg_imp / float(seg_df["reach"].sum())) if ("reach" in seg_df.columns and float(seg_df["reach"].sum()) > 0) else 0.0,
                                    "ctr": seg_ctr / 100,
                                    "cpc": seg_cpc,
                                    "cr": seg_cr / 100,
                                    "aov": None,
                                    "clicks": seg_clicks,
                                    "conversions": seg_conv,
                                    "leads": seg_leads,
                                    "target_leads": seg_target_leads,
                                    "cost": seg_cost,
                                    "ak_cost_wo_vat": seg_ak_wo_vat,
                                    "total_budget_wo_vat_ak": seg_cost + seg_ak_wo_vat,
                                    "cost_with_vat": seg_cost_with_vat,
                                    "cost_with_vat_ak": seg_cost_with_vat_ak,
                                    "vat_amount": seg_cost_with_vat_ak - (seg_cost + seg_ak_wo_vat),
                                    "ak_rate_pct": (seg_ak_wo_vat / seg_cost * 100.0) if seg_cost > 0 else 0.0,
                                    "revenue": seg_rev,
                                    "cpm": seg_cpm,
                                    "cpa": seg_cpa,
                                    "cpl": seg_cpl,
                                    "cpql": seg_cpql,
                                    "roas": seg_roas,
                                    "drr": seg_drr / 100,
                                    "ctr_pct": seg_ctr,
                                    "cr_pct": seg_cr,
                                    "cr1_pct": seg_cr1,
                                    "cr2_pct": seg_cr2,
                                    "drr_pct": seg_drr,
                                }
                                if is_diy_preset:
                                    seg_cap = float(seg_df["available_capacity"].sum()) if "available_capacity" in seg_df.columns else 0.0
                                    seg_row["available_capacity"] = seg_cap
                                    seg_row["client_count"] = float(seg_df["client_count"].sum()) if "client_count" in seg_df.columns else 0.0
                                    seg_row["absolute_new_clients"] = float(seg_df["absolute_new_clients"].sum()) if "absolute_new_clients" in seg_df.columns else 0.0
                                    seg_row["returned_clients"] = float(seg_df["returned_clients"].sum()) if "returned_clients" in seg_df.columns else 0.0
                                    seg_row["new_clients"] = float(seg_df["new_clients"].sum()) if "new_clients" in seg_df.columns else 0.0
                                    seg_row["order_frequency"] = float(seg_df["order_frequency"].mean()) if "order_frequency" in seg_df.columns else 0.0
                                    seg_row["sov_pct"] = (float(seg_row.get("reach", 0.0)) / seg_cap * 100.0) if seg_cap > 0 else 0.0
                                    seg_row["new_clients_share_pct"] = calc_new_clients_share_pct(
                                        seg_row["new_clients"],
                                        seg_row["client_count"],
                                    )
                                    seg_row["cac"] = (seg_cost_with_vat_ak / float(seg_row["new_clients"])) if float(seg_row["new_clients"]) > 0 else 0.0
                                    seg_row["shipped_orders"] = float(seg_df["shipped_orders"].sum()) if "shipped_orders" in seg_df.columns else 0.0
                                    seg_row["shipped_revenue"] = float(seg_df["shipped_revenue"].sum()) if "shipped_revenue" in seg_df.columns else 0.0
                                    seg_row["shipped_cps"] = (seg_cost_with_vat / float(seg_row["shipped_orders"])) if float(seg_row["shipped_orders"]) > 0 else 0.0
                                    seg_row["shipped_aov"] = (float(seg_row["shipped_revenue"]) / float(seg_row["shipped_orders"])) if float(seg_row["shipped_orders"]) > 0 else 0.0
                                    seg_row["shipped_roas"] = (float(seg_row["shipped_revenue"]) / seg_budget_basis) if seg_budget_basis > 0 else 0.0
                                    seg_row["shipped_drr_pct"] = (seg_budget_basis / float(seg_row["shipped_revenue"]) * 100.0) if float(seg_row["shipped_revenue"]) > 0 else 0.0
                                    seg_row["order_share_segment_pct"] = 100.0
                                    seg_row["revenue_share_segment_pct"] = 100.0
                                    seg_row["shipped_order_share_segment_pct"] = 100.0
                                    seg_row["shipped_revenue_share_segment_pct"] = 100.0
                                    seg_row["budget_share_segment_pct"] = 100.0
                                segment_total_rows_month.append(seg_row)

                        # Основная таблица: только типы РК (без TOTAL), чтобы построчно
                        # совпадать с таблицей коэффициентов справа.
                        df_rows_show["impressions"] = df_rows_show["impressions"].map(
                            lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
                        )
                        df_rows_show["clicks"] = df_rows_show["clicks"].map(
                            lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
                        )
                        if "reach" in df_rows_show.columns:
                            df_rows_show["reach"] = df_rows_show["reach"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
                            )
                        if "frequency" in df_rows_show.columns:
                            df_rows_show["frequency"] = df_rows_show["frequency"].map(
                                lambda x: "" if pd.isna(x) else f"{x:.2f}"
                            )
                        df_rows_show["conversions"] = df_rows_show["conversions"].map(
                            lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
                        )
                        if "leads" in df_rows_show.columns:
                            df_rows_show["leads"] = df_rows_show["leads"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
                            )
                        if "target_leads" in df_rows_show.columns:
                            df_rows_show["target_leads"] = df_rows_show["target_leads"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
                            )

                        df_rows_show["cost"] = df_rows_show["cost"].map(
                            lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
                        )
                        if "ak_cost_wo_vat" in df_rows_show.columns:
                            df_rows_show["ak_cost_wo_vat"] = df_rows_show["ak_cost_wo_vat"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
                            )
                        if "total_budget_wo_vat_ak" in df_rows_show.columns:
                            df_rows_show["total_budget_wo_vat_ak"] = df_rows_show["total_budget_wo_vat_ak"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
                            )
                        df_rows_show["cost_with_vat"] = df_rows_show["cost_with_vat"].map(
                            lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
                        )
                        df_rows_show["cost_with_vat_ak"] = df_rows_show["cost_with_vat_ak"].map(
                            lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
                        )
                        if "vat_amount" in df_rows_show.columns:
                            df_rows_show["vat_amount"] = df_rows_show["vat_amount"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
                            )
                        df_rows_show["ak_rate_pct"] = df_rows_show["ak_rate_pct"].map(
                            lambda x: "" if pd.isna(x) else f"{x:.2f} %"
                        )
                        df_rows_show["revenue"] = df_rows_show["revenue"].map(
                            lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
                        )
                        df_rows_show["cpc"] = df_rows_show["cpc"].map(
                            lambda x: "" if pd.isna(x) else f"{x:.2f} ₽".replace(",", " ")
                        )
                        df_rows_show["cpm"] = df_rows_show["cpm"].map(
                            lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
                        )
                        df_rows_show["cpa"] = df_rows_show["cpa"].map(
                            lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
                        )
                        df_rows_show["aov"] = df_rows_show["aov"].map(
                            lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
                        )
                        if "cpl" in df_rows_show.columns:
                            df_rows_show["cpl"] = df_rows_show["cpl"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
                            )
                        if "cpql" in df_rows_show.columns:
                            df_rows_show["cpql"] = df_rows_show["cpql"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
                            )

                        df_rows_show["ctr_pct"] = df_rows_show["ctr_pct"].map(
                            lambda x: "" if pd.isna(x) else f"{x:.2f} %"
                        )
                        df_rows_show["cr_pct"] = df_rows_show["cr_pct"].map(
                            lambda x: "" if pd.isna(x) else f"{x:.2f} %"
                        )
                        if "cr1_pct" in df_rows_show.columns:
                            df_rows_show["cr1_pct"] = df_rows_show["cr1_pct"].map(
                                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
                            )
                        if "cr2_pct" in df_rows_show.columns:
                            df_rows_show["cr2_pct"] = df_rows_show["cr2_pct"].map(
                                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
                            )
                        df_rows_show["drr_pct"] = df_rows_show["drr_pct"].map(
                            lambda x: "" if pd.isna(x) else f"{x:.2f} %"
                        )
                        df_rows_show["ROAS"] = df_rows_show["roas"].map(
                            lambda x: "" if pd.isna(x) else f"{x:.2f}"
                        )
                        if is_diy_preset:
                            df_rows_show["available_capacity"] = df_rows_show["available_capacity"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
                            )
                            df_rows_show["client_count"] = df_rows_show["client_count"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
                            )
                            df_rows_show["absolute_new_clients"] = df_rows_show["absolute_new_clients"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
                            )
                            df_rows_show["returned_clients"] = df_rows_show["returned_clients"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
                            )
                            df_rows_show["new_clients"] = df_rows_show["new_clients"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " ")
                            )
                            df_rows_show["cac"] = df_rows_show["cac"].map(
                                lambda x: "" if pd.isna(x) else f"{round(x):,} ₽".replace(",", " ")
                            )
                            df_rows_show["order_frequency"] = df_rows_show["order_frequency"].map(
                                lambda x: "" if pd.isna(x) else f"{x:.2f}"
                            )
                            df_rows_show["sov_pct"] = df_rows_show["sov_pct"].map(
                                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
                            )
                            df_rows_show["new_clients_share_pct"] = df_rows_show["new_clients_share_pct"].map(
                                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
                            )
                            df_rows_show["order_share_segment_pct"] = df_rows_show["order_share_segment_pct"].map(
                                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
                            )
                            df_rows_show["revenue_share_segment_pct"] = df_rows_show["revenue_share_segment_pct"].map(
                                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
                            )
                            df_rows_show["budget_share_segment_pct"] = df_rows_show["budget_share_segment_pct"].map(
                                lambda x: "" if pd.isna(x) else f"{x:.2f} %"
                            )

                        if is_real_estate_preset:
                            month_show_cols = ["campaign_type", "system", "format", "geo"] + get_real_estate_table_cols(metric_mode)
                        elif is_diy_preset:
                            month_show_cols = ["campaign_type", "system", "format", "geo"] + get_diy_plan_table_cols()
                        else:
                            month_show_cols = [
                                "campaign_type",
                                "system",
                                "format",
                                "geo",
                                "cost",
                                "ak_cost_wo_vat",
                                "total_budget_wo_vat_ak",
                                "cost_with_vat",
                                "cost_with_vat_ak",
                                "vat_amount",
                                "ak_rate_pct",
                                "impressions",
                                "clicks",
                                "cpc",
                                "ctr_pct",
                                "cr_pct",
                                "conversions",
                                "cpa",
                                "aov",
                                "revenue",
                                "drr_pct",
                                "ROAS",
                            ]
                        if is_diy_preset:
                            month_show_cols.insert(3, "segment")
                        df_rows_show = safe_select_columns(df_rows_show, month_show_cols, fill_value="")
                        df_rows_show = df_rows_show.rename(columns=DISPLAY_COL_RENAME)

                        # TOTAL отдельной строкой ниже основной таблицы.
                        total_month_show = pd.DataFrame(segment_total_rows_month + [total_row_month])
                        total_month_show["impressions"] = total_month_show["impressions"].map(lambda x: f"{round(x):,}".replace(",", " "))
                        if "reach" in total_month_show.columns:
                            total_month_show["reach"] = total_month_show["reach"].map(lambda x: f"{round(x):,}".replace(",", " "))
                        if "frequency" in total_month_show.columns:
                            total_month_show["frequency"] = total_month_show["frequency"].map(lambda x: f"{x:.2f}")
                        total_month_show["clicks"] = total_month_show["clicks"].map(lambda x: f"{round(x):,}".replace(",", " "))
                        total_month_show["conversions"] = total_month_show["conversions"].map(lambda x: f"{round(x):,}".replace(",", " "))
                        if "leads" in total_month_show.columns:
                            total_month_show["leads"] = total_month_show["leads"].map(lambda x: f"{round(x):,}".replace(",", " "))
                        if "target_leads" in total_month_show.columns:
                            total_month_show["target_leads"] = total_month_show["target_leads"].map(lambda x: f"{round(x):,}".replace(",", " "))
                        total_month_show["cost"] = total_month_show["cost"].map(lambda x: f"{round(x):,} ₽".replace(",", " "))
                        if "ak_cost_wo_vat" in total_month_show.columns:
                            total_month_show["ak_cost_wo_vat"] = total_month_show["ak_cost_wo_vat"].map(lambda x: f"{round(x):,} ₽".replace(",", " "))
                        if "total_budget_wo_vat_ak" in total_month_show.columns:
                            total_month_show["total_budget_wo_vat_ak"] = total_month_show["total_budget_wo_vat_ak"].map(lambda x: f"{round(x):,} ₽".replace(",", " "))
                        total_month_show["cost_with_vat"] = total_month_show["cost_with_vat"].map(lambda x: f"{round(x):,} ₽".replace(",", " "))
                        total_month_show["cost_with_vat_ak"] = total_month_show["cost_with_vat_ak"].map(lambda x: f"{round(x):,} ₽".replace(",", " "))
                        if "vat_amount" in total_month_show.columns:
                            total_month_show["vat_amount"] = total_month_show["vat_amount"].map(lambda x: f"{round(x):,} ₽".replace(",", " "))
                        total_month_show["ak_rate_pct"] = total_month_show["ak_rate_pct"].map(lambda x: f"{x:.2f} %")
                        total_month_show["revenue"] = total_month_show["revenue"].map(lambda x: f"{round(x):,} ₽".replace(",", " "))
                        total_month_show["cpc"] = total_month_show["cpc"].map(lambda x: f"{x:.2f} ₽")
                        total_month_show["cpm"] = total_month_show["cpm"].map(lambda x: f"{round(x):,} ₽".replace(",", " "))
                        total_month_show["cpa"] = total_month_show["cpa"].map(lambda x: f"{round(x):,} ₽".replace(",", " "))
                        if "cpl" in total_month_show.columns:
                            total_month_show["cpl"] = total_month_show["cpl"].map(lambda x: f"{round(x):,} ₽".replace(",", " "))
                        if "cpql" in total_month_show.columns:
                            total_month_show["cpql"] = total_month_show["cpql"].map(lambda x: f"{round(x):,} ₽".replace(",", " "))
                        total_month_show["aov"] = ""
                        if "client_count" in total_month_show.columns:
                            total_month_show["client_count"] = total_month_show["client_count"].map(lambda x: f"{round(x):,}".replace(",", " "))
                        if "absolute_new_clients" in total_month_show.columns:
                            total_month_show["absolute_new_clients"] = total_month_show["absolute_new_clients"].map(lambda x: f"{round(x):,}".replace(",", " "))
                        if "returned_clients" in total_month_show.columns:
                            total_month_show["returned_clients"] = total_month_show["returned_clients"].map(lambda x: f"{round(x):,}".replace(",", " "))
                        if "new_clients" in total_month_show.columns:
                            total_month_show["new_clients"] = total_month_show["new_clients"].map(lambda x: f"{round(x):,}".replace(",", " "))
                        if "cac" in total_month_show.columns:
                            total_month_show["cac"] = total_month_show["cac"].map(lambda x: f"{round(x):,} ₽".replace(",", " "))
                        if "order_frequency" in total_month_show.columns:
                            total_month_show["order_frequency"] = total_month_show["order_frequency"].map(lambda x: f"{x:.2f}")
                        total_month_show["ctr_pct"] = total_month_show["ctr_pct"].map(lambda x: f"{x:.2f} %")
                        total_month_show["cr_pct"] = total_month_show["cr_pct"].map(lambda x: f"{x:.2f} %")
                        if "cr1_pct" in total_month_show.columns:
                            total_month_show["cr1_pct"] = total_month_show["cr1_pct"].map(lambda x: f"{x:.2f} %")
                        if "cr2_pct" in total_month_show.columns:
                            total_month_show["cr2_pct"] = total_month_show["cr2_pct"].map(lambda x: f"{x:.2f} %")
                        total_month_show["drr_pct"] = total_month_show["drr_pct"].map(lambda x: f"{x:.2f} %")
                        total_month_show["ROAS"] = total_month_show["roas"].map(lambda x: f"{x:.2f}")
                        if is_diy_preset:
                            total_month_show["available_capacity"] = total_month_show["available_capacity"].map(
                                lambda x: f"{round(x):,}".replace(",", " ")
                            )
                            total_month_show["sov_pct"] = total_month_show["sov_pct"].map(lambda x: f"{x:.2f} %")
                            total_month_show["new_clients_share_pct"] = total_month_show["new_clients_share_pct"].map(lambda x: f"{x:.2f} %")
                            total_month_show["order_share_segment_pct"] = total_month_show["order_share_segment_pct"].map(lambda x: f"{x:.2f} %")
                            total_month_show["revenue_share_segment_pct"] = total_month_show["revenue_share_segment_pct"].map(lambda x: f"{x:.2f} %")
                            total_month_show["budget_share_segment_pct"] = total_month_show["budget_share_segment_pct"].map(lambda x: f"{x:.2f} %")
                        if is_real_estate_preset:
                            total_month_cols = ["campaign_type", "system", "format", "geo"] + get_real_estate_table_cols(metric_mode)
                        elif is_diy_preset:
                            total_month_cols = ["campaign_type", "system", "format", "geo"] + get_diy_plan_table_cols()
                        else:
                            total_month_cols = [
                                "campaign_type", "system", "format", "geo",
                                "cost", "ak_cost_wo_vat", "total_budget_wo_vat_ak", "cost_with_vat", "cost_with_vat_ak", "vat_amount", "ak_rate_pct",
                                "impressions", "clicks", "cpc", "ctr_pct", "cr_pct",
                                "conversions", "cpa", "aov", "revenue", "drr_pct", "ROAS",
                            ]
                        if is_diy_preset:
                            total_month_cols.insert(3, "segment")
                        total_month_show = safe_select_columns(total_month_show, total_month_cols, fill_value="").rename(columns=DISPLAY_COL_RENAME)

                        # Простой вариант: возвращаем TOTAL обратно в основную таблицу.
                        df_month_with_total_show = pd.concat([df_rows_show, total_month_show], ignore_index=True)
                        if is_diy_preset:
                            df_month_with_total_show = reorder_rows_with_segment_subtotals(
                                df_month_with_total_show,
                                DISPLAY_COL_RENAME["campaign_type"],
                                DISPLAY_COL_RENAME["segment"],
                            )

                        # Это display-таблица: приводим всё к строкам, чтобы не смешивать
                        # float и форматированные строковые значения в одних и тех же колонках.
                        df_month_with_total_show = df_month_with_total_show.where(pd.notna(df_month_with_total_show), "")
                        df_month_with_total_show = df_month_with_total_show.astype(str)

                        def _highlight_total_row_merged(row):
                            style = [""] * len(row)
                            camp_col = DISPLAY_COL_RENAME.get("campaign_type", "Название кампании")
                            camp_val = str(row.get(camp_col, ""))
                            if camp_val == "Итого":
                                style = [
                                    f"background-color: #00CDC5; color: #081521; font-weight: 700; border-top: 2px solid {THEME_BORDER};"
                                ] * len(row)
                            elif camp_val.startswith("Итого "):
                                style = [
                                    f"background-color: #2C6E75; color: #DDEAF0; font-weight: 650; border-top: 1px solid {THEME_BORDER};"
                                ] * len(row)
                            return style

                        month_table_height_with_total = max(
                            140,
                            min(560, month_table_height + month_row_height + 4),
                        )
                        st.dataframe(
                            df_month_with_total_show.style.apply(_highlight_total_row_merged, axis=1),
                            width="stretch",
                            height=month_table_height_with_total,
                        )
                    else:
                        st.info("Нет данных для этого месяца.")

    # ---------- 4. Итоги по выбранным месяцам (TOTAL) ----------

    if all_months_results:
        df_all = pd.concat(all_months_results, ignore_index=True)
    else:
        df_all = pd.DataFrame()

    if not df_all.empty:
        ref_imp = float(df_all["impressions"].sum())
        ref_reach = float(df_all["reach"].sum()) if "reach" in df_all.columns else 0.0
        ref_frequency = (ref_imp / ref_reach) if ref_reach > 0 else 0.0
        ref_clicks = float(df_all["clicks"].sum())
        ref_conv = float(df_all["conversions"].sum())
        ref_leads = float(df_all["leads"].fillna(0.0).sum()) if "leads" in df_all.columns else 0.0
        ref_target_leads = float(df_all["target_leads"].fillna(0.0).sum()) if "target_leads" in df_all.columns else float(ref_conv)
        ref_cost = float(df_all["cost"].sum())
        ref_cost_with_vat = float(df_all["cost_with_vat"].sum())
        ref_cost_with_vat_ak = float(df_all["cost_with_vat_ak"].sum())
        ref_ak_wo_vat = float(df_all["ak_cost_wo_vat"].sum())
        ref_rev = float(df_all["revenue"].sum())
        if use_ak_budget_metrics:
            ref_budget_basis = ref_cost_with_vat_ak if use_vat_budget_metrics else (ref_cost + ref_ak_wo_vat)
        else:
            ref_budget_basis = ref_cost_with_vat if use_vat_budget_metrics else ref_cost
        ref_ctr = (ref_clicks / ref_imp * 100.0) if ref_imp > 0 else 0.0
        ref_cpc = (ref_cost / ref_clicks) if ref_clicks > 0 else 0.0
        ref_cr = (ref_conv / ref_clicks * 100.0) if ref_clicks > 0 else 0.0
        ref_cpo = (ref_budget_basis / ref_conv) if ref_conv > 0 else 0.0
        ref_cpl = (ref_budget_basis / ref_leads) if ref_leads > 0 else 0.0
        ref_cpql = (ref_budget_basis / ref_target_leads) if ref_target_leads > 0 else 0.0
        ref_roas = (ref_rev / ref_budget_basis) if ref_budget_basis > 0 else 0.0
        ref_drr = (ref_budget_basis / ref_rev * 100.0) if ref_rev > 0 else 0.0
        st.session_state["mp_ref_total"] = {
            "impressions": ref_imp,
            "reach": ref_reach,
            "frequency": ref_frequency,
            "clicks": ref_clicks,
            "conversions": ref_conv,
            "leads": ref_leads,
            "target_leads": ref_target_leads,
            "cost": ref_cost,
            "cost_with_vat": ref_cost_with_vat,
            "cost_with_vat_ak": ref_cost_with_vat_ak,
            "revenue": ref_rev,
            "ctr": ref_ctr,
            "cpc": ref_cpc,
            "cr": ref_cr,
            "cpo": ref_cpo,
            "cpl": ref_cpl,
            "cpql": ref_cpql,
            "roas": ref_roas,
            "drr": ref_drr,
        }
        total_by_campaign = {}
        agg_ct = df_all.groupby("campaign_type", as_index=False).agg(
            impressions=("impressions", "sum"),
            reach=("reach", "sum"),
            clicks=("clicks", "sum"),
            conversions=("conversions", "sum"),
            leads=("leads", "sum"),
            target_leads=("target_leads", "sum"),
            cost=("cost", "sum"),
            cost_with_vat=("cost_with_vat", "sum"),
            cost_with_vat_ak=("cost_with_vat_ak", "sum"),
            ak_cost_wo_vat=("ak_cost_wo_vat", "sum"),
            revenue=("revenue", "sum"),
        )
        for _, r in agg_ct.iterrows():
            camp = str(r.get("campaign_type", "")).strip()
            if not camp:
                continue
            imp = float(r.get("impressions", 0.0))
            reach = float(r.get("reach", 0.0) or 0.0)
            frequency = (imp / reach) if reach > 0 else 0.0
            clicks = float(r.get("clicks", 0.0))
            conv = float(r.get("conversions", 0.0))
            leads = float(r.get("leads", 0.0) or 0.0)
            target_leads = float(r.get("target_leads", conv) or 0.0)
            cost = float(r.get("cost", 0.0))
            cost_with_vat = float(r.get("cost_with_vat", 0.0))
            cost_with_vat_ak = float(r.get("cost_with_vat_ak", 0.0))
            ak_wo_vat = float(r.get("ak_cost_wo_vat", 0.0))
            rev = float(r.get("revenue", 0.0))
            if use_ak_budget_metrics:
                budget_basis = cost_with_vat_ak if use_vat_budget_metrics else (cost + ak_wo_vat)
            else:
                budget_basis = cost_with_vat if use_vat_budget_metrics else cost
            ctr = (clicks / imp * 100.0) if imp > 0 else 0.0
            cpc = (cost / clicks) if clicks > 0 else 0.0
            cr = (conv / clicks * 100.0) if clicks > 0 else 0.0
            cpo = (budget_basis / conv) if conv > 0 else 0.0
            cpl = (budget_basis / leads) if leads > 0 else 0.0
            cpql = (budget_basis / target_leads) if target_leads > 0 else 0.0
            roas = (rev / budget_basis) if budget_basis > 0 else 0.0
            drr = (budget_basis / rev * 100.0) if rev > 0 else 0.0
            total_by_campaign[camp] = {
                "impressions": imp,
                "reach": reach,
                "frequency": frequency,
                "clicks": clicks,
                "conversions": conv,
                "leads": leads,
                "target_leads": target_leads,
                "cost": cost,
                "cost_with_vat": cost_with_vat,
                "cost_with_vat_ak": cost_with_vat_ak,
                "revenue": rev,
                "ctr": ctr,
                "cpc": cpc,
                "cr": cr,
                "cpo": cpo,
                "cpl": cpl,
                "cpql": cpql,
                "roas": roas,
                "drr": drr,
            }
        st.session_state["mp_ref_total_by_campaign"] = total_by_campaign
    else:
        st.session_state["mp_ref_total"] = None
        st.session_state["mp_ref_total_by_campaign"] = {}

    with st.sidebar:
        st.markdown("---")
        st.markdown("### Быстрая сверка")
        # Import compatibility: old state could store a legacy mode value.
        if st.session_state.get("mp_ref_mode") == "Включено":
            st.session_state["mp_ref_mode"] = "Средний месяц"
        # If references are missing after import/rerun, rebuild them from base campaigns.
        if not st.session_state.get("mp_ref_base"):
            _bootstrap_reference_from_campaigns(st.session_state.get("campaigns_df"))
        if is_diy_preset:
            st.selectbox(
                "Сегмент в расчете",
                options=["Все", "B2B", "B2C"],
                key="plan_segment_filter_sidebar",
                on_change=_sync_segment_from_sidebar,
            )
        ref_mode = st.selectbox(
            "Что закрепить",
            options=["Выключено", "Средний месяц", "TOTAL выбранных месяцев"],
            key="mp_ref_mode",
        )
        all_ref_campaigns = sorted(
            set(st.session_state.get("mp_ref_base_by_campaign", {}).keys())
            | set(st.session_state.get("mp_ref_total_by_campaign", {}).keys())
        )
        valid_ref_campaigns = ["Все типы РК"] + all_ref_campaigns
        if st.session_state.get("mp_ref_campaign") not in valid_ref_campaigns:
            st.session_state["mp_ref_campaign"] = "Все типы РК"
        ref_campaign = st.selectbox(
            "Тип РК",
            options=valid_ref_campaigns,
            key="mp_ref_campaign",
        )
        ref_src = None
        if ref_mode == "Средний месяц":
            if ref_campaign == "Все типы РК":
                ref_src = st.session_state.get("mp_ref_base")
            else:
                ref_src = st.session_state.get("mp_ref_base_by_campaign", {}).get(ref_campaign)
        elif ref_mode == "TOTAL выбранных месяцев":
            if ref_campaign == "Все типы РК":
                ref_src = st.session_state.get("mp_ref_total")
            else:
                ref_src = st.session_state.get("mp_ref_total_by_campaign", {}).get(ref_campaign)

        # Fallback: right after import TOTAL may be unavailable until full plan calc.
        # Show base reference instead of empty block.
        if ref_src is None and ref_mode == "TOTAL выбранных месяцев":
            if ref_campaign == "Все типы РК":
                ref_src = st.session_state.get("mp_ref_base")
            else:
                ref_src = st.session_state.get("mp_ref_base_by_campaign", {}).get(ref_campaign)

        if ref_mode != "Выключено":
            if ref_src:
                st.dataframe(_build_ref_df(ref_src), width="stretch", hide_index=True)
            else:
                st.caption("Нет данных для выбранного режима сверки.")

    with st.expander("4. Итоги по выбранным месяцам (TOTAL)", expanded=True):
        if df_all.empty:
            st.info("Нет данных для итогов. Заполните медиаплан по месяцам.")
        else:
            st.caption("Сводные показатели по всем выбранным месяцам и типам РК.")

            agg = df_all.groupby(["planning_slot", "month_num", "month_year", "month_name"], as_index=False).agg(
                impressions=("impressions", "sum"),
                clicks=("clicks", "sum"),
                conversions=("conversions", "sum"),
                leads=("leads", "sum"),
                target_leads=("target_leads", "sum"),
                cost=("cost", "sum"),
                cost_with_vat=("cost_with_vat", "sum"),
                cost_with_vat_ak=("cost_with_vat_ak", "sum"),
                ak_cost_wo_vat=("ak_cost_wo_vat", "sum"),
                revenue=("revenue", "sum"),
            )
            if is_diy_preset:
                agg["reach"] = df_all.groupby(["planning_slot", "month_num", "month_year", "month_name"])["reach"].sum().values if "reach" in df_all.columns else 0.0
                agg["frequency"] = np.where(agg["reach"] > 0, agg["impressions"] / agg["reach"], 0.0)
                agg["available_capacity"] = df_all.groupby(["planning_slot", "month_num", "month_year", "month_name"])["available_capacity"].sum().values if "available_capacity" in df_all.columns else 0.0
                agg["client_count"] = df_all.groupby(["planning_slot", "month_num", "month_year", "month_name"])["client_count"].sum().values if "client_count" in df_all.columns else 0.0
                agg["absolute_new_clients"] = df_all.groupby(["planning_slot", "month_num", "month_year", "month_name"])["absolute_new_clients"].sum().values if "absolute_new_clients" in df_all.columns else 0.0
                agg["returned_clients"] = df_all.groupby(["planning_slot", "month_num", "month_year", "month_name"])["returned_clients"].sum().values if "returned_clients" in df_all.columns else 0.0
                agg["new_clients"] = df_all.groupby(["planning_slot", "month_num", "month_year", "month_name"])["new_clients"].sum().values if "new_clients" in df_all.columns else 0.0
                agg["order_frequency"] = df_all.groupby(["planning_slot", "month_num", "month_year", "month_name"])["order_frequency"].mean().values if "order_frequency" in df_all.columns else 0.0
                agg["sov_pct"] = np.where(agg["available_capacity"] > 0, agg["reach"] / agg["available_capacity"] * 100.0, 0.0)
                agg["new_clients_share_pct"] = np.where(
                    agg["client_count"] > 0,
                    agg["new_clients"] / agg["client_count"] * 100.0,
                    0.0,
                )
                agg["order_share_segment_pct"] = 0.0
                agg["revenue_share_segment_pct"] = 0.0
            if use_ak_budget_metrics:
                budget_series = agg["cost_with_vat_ak"] if use_vat_budget_metrics else (agg["cost"] + agg["ak_cost_wo_vat"])
            else:
                budget_series = agg["cost_with_vat"] if use_vat_budget_metrics else agg["cost"]
            agg["total_budget_wo_vat_ak"] = agg["cost"] + agg["ak_cost_wo_vat"]
            agg["vat_amount"] = agg["cost_with_vat_ak"] - agg["total_budget_wo_vat_ak"]

            agg["ctr"] = np.where(
                agg["impressions"] > 0,
                agg["clicks"] / agg["impressions"] * 100,
                0.0,
            )
            agg["cr"] = np.where(
                agg["clicks"] > 0,
                agg["conversions"] / agg["clicks"] * 100,
                0.0,
            )
            agg["cr1_pct"] = np.where(
                agg["clicks"] > 0,
                agg["leads"] / agg["clicks"] * 100,
                0.0,
            )
            agg["cr2_pct"] = np.where(
                agg["leads"] > 0,
                agg["target_leads"] / agg["leads"] * 100,
                0.0,
            )
            agg["cpc"] = np.where(
                agg["clicks"] > 0,
                agg["cost"] / agg["clicks"],
                0.0,
            )
            agg["cpm"] = np.where(
                agg["impressions"] > 0,
                budget_series / (agg["impressions"] / 1000),
                0.0,
            )
            agg["cpa"] = np.where(
                agg["conversions"] > 0,
                budget_series / agg["conversions"],
                0.0,
            )
            agg["cpl"] = np.where(
                agg["leads"] > 0,
                budget_series / agg["leads"],
                0.0,
            )
            agg["cpql"] = np.where(
                agg["target_leads"] > 0,
                budget_series / agg["target_leads"],
                0.0,
            )
            agg["aov"] = np.where(
                agg["conversions"] > 0,
                agg["revenue"] / agg["conversions"],
                0.0,
            )
            agg["roas"] = np.where(
                budget_series > 0,
                agg["revenue"] / budget_series,
                0.0,
            )
            agg["drr"] = np.where(
                agg["revenue"] > 0,
                budget_series / agg["revenue"] * 100,
                0.0,
            )
            agg["ctr_pct"] = agg["ctr"]
            agg["cr_pct"] = agg["cr"]

            total_row = {
                "planning_slot": 999,
                "month_num": 999,
                "month_year": 999,
                "month_name": "TOTAL",
                "impressions": agg["impressions"].sum(),
                "reach": float(agg["reach"].sum()) if "reach" in agg.columns else 0.0,
                "clicks": agg["clicks"].sum(),
                "conversions": agg["conversions"].sum(),
                "leads": agg["leads"].sum(),
                "target_leads": agg["target_leads"].sum(),
                "cost": agg["cost"].sum(),
                "cost_with_vat": agg["cost_with_vat"].sum(),
                "cost_with_vat_ak": agg["cost_with_vat_ak"].sum(),
                "ak_cost_wo_vat": agg["ak_cost_wo_vat"].sum(),
                "total_budget_wo_vat_ak": agg["cost"].sum() + agg["ak_cost_wo_vat"].sum(),
                "vat_amount": agg["cost_with_vat_ak"].sum() - (agg["cost"].sum() + agg["ak_cost_wo_vat"].sum()),
                "revenue": agg["revenue"].sum(),
            }
            if use_ak_budget_metrics:
                total_budget_basis = total_row["cost_with_vat_ak"] if use_vat_budget_metrics else (total_row["cost"] + total_row["ak_cost_wo_vat"])
            else:
                total_budget_basis = total_row["cost_with_vat"] if use_vat_budget_metrics else total_row["cost"]
            if total_row["impressions"] > 0:
                total_row["ctr"] = total_row["clicks"] / total_row["impressions"] * 100
                total_row["cpm"] = total_budget_basis / (total_row["impressions"] / 1000)
            else:
                total_row["ctr"] = 0.0
                total_row["cpm"] = 0.0
            total_row["ctr_pct"] = total_row["ctr"]
            if total_row["clicks"] > 0:
                total_row["cr"] = total_row["conversions"] / total_row["clicks"] * 100
                total_row["cpc"] = total_row["cost"] / total_row["clicks"]
                total_row["cr1_pct"] = total_row["leads"] / total_row["clicks"] * 100
            else:
                total_row["cr"] = 0.0
                total_row["cpc"] = 0.0
                total_row["cr1_pct"] = 0.0
            total_row["cr_pct"] = total_row["cr"]
            if total_row["leads"] > 0:
                total_row["cr2_pct"] = total_row["target_leads"] / total_row["leads"] * 100
                total_row["cpl"] = total_budget_basis / total_row["leads"]
            else:
                total_row["cr2_pct"] = 0.0
                total_row["cpl"] = 0.0
            if total_row["conversions"] > 0:
                total_row["cpa"] = (total_row["cost_with_vat"] / total_row["conversions"]) if is_diy_preset else (total_budget_basis / total_row["conversions"])
                total_row["aov"] = total_row["revenue"] / total_row["conversions"]
            else:
                total_row["cpa"] = 0.0
                total_row["aov"] = 0.0
            if total_row["target_leads"] > 0:
                total_row["cpql"] = total_budget_basis / total_row["target_leads"]
            else:
                total_row["cpql"] = 0.0
            if total_budget_basis > 0:
                total_row["roas"] = total_row["revenue"] / total_budget_basis
            else:
                total_row["roas"] = 0.0
            if total_row["revenue"] > 0:
                total_row["drr"] = total_budget_basis / total_row["revenue"] * 100
            else:
                total_row["drr"] = 0.0
            if is_diy_preset:
                total_row["frequency"] = (total_row["impressions"] / total_row["reach"]) if total_row["reach"] > 0 else 0.0
                total_row["available_capacity"] = float(agg["available_capacity"].sum()) if "available_capacity" in agg.columns else 0.0
                total_row["client_count"] = float(agg["client_count"].sum()) if "client_count" in agg.columns else 0.0
                total_row["absolute_new_clients"] = float(agg["absolute_new_clients"].sum()) if "absolute_new_clients" in agg.columns else 0.0
                total_row["returned_clients"] = float(agg["returned_clients"].sum()) if "returned_clients" in agg.columns else 0.0
                total_row["new_clients"] = float(agg["new_clients"].sum()) if "new_clients" in agg.columns else 0.0
                total_row["order_frequency"] = float(agg["order_frequency"].mean()) if "order_frequency" in agg.columns else 0.0
                total_row["sov_pct"] = (total_row["reach"] / total_row["available_capacity"] * 100.0) if total_row["available_capacity"] > 0 else 0.0
                total_row["new_clients_share_pct"] = calc_new_clients_share_pct(total_row["new_clients"], total_row["client_count"])
                total_row["cac"] = (total_row["cost_with_vat_ak"] / float(total_row["new_clients"])) if float(total_row["new_clients"]) > 0 else 0.0
                total_row["order_share_segment_pct"] = 100.0
                total_row["revenue_share_segment_pct"] = 100.0

            segment_total_rows = []
            if show_segment_subtotals and "segment" in df_all.columns:
                for seg_name, seg_df in df_all.groupby("segment"):
                    seg_cost = float(seg_df["cost"].sum())
                    seg_cost_with_vat = float(seg_df["cost_with_vat"].sum())
                    seg_cost_with_vat_ak = float(seg_df["cost_with_vat_ak"].sum())
                    seg_ak_wo_vat = float(seg_df["ak_cost_wo_vat"].sum())
                    if use_ak_budget_metrics:
                        seg_budget_basis = seg_cost_with_vat_ak if use_vat_budget_metrics else (seg_cost + seg_ak_wo_vat)
                    else:
                        seg_budget_basis = seg_cost_with_vat if use_vat_budget_metrics else seg_cost
                    seg_row = {
                        "planning_slot": 998,
                        "month_num": 998,
                        "month_year": 998,
                        "month_name": f"Итого {seg_name}",
                        "impressions": float(seg_df["impressions"].sum()),
                        "reach": float(seg_df["reach"].sum()) if "reach" in seg_df.columns else 0.0,
                        "clicks": float(seg_df["clicks"].sum()),
                        "conversions": float(seg_df["conversions"].sum()),
                        "leads": float(seg_df["leads"].fillna(0.0).sum()) if "leads" in seg_df.columns else 0.0,
                        "target_leads": float(seg_df["target_leads"].fillna(0.0).sum()) if "target_leads" in seg_df.columns else float(seg_df["conversions"].sum()),
                        "cost": seg_cost,
                        "cost_with_vat": seg_cost_with_vat,
                        "cost_with_vat_ak": seg_cost_with_vat_ak,
                        "ak_cost_wo_vat": seg_ak_wo_vat,
                        "total_budget_wo_vat_ak": seg_cost + seg_ak_wo_vat,
                        "vat_amount": seg_cost_with_vat_ak - (seg_cost + seg_ak_wo_vat),
                        "revenue": float(seg_df["revenue"].sum()),
                    }
                    seg_row["ctr"] = (seg_row["clicks"] / seg_row["impressions"] * 100.0) if seg_row["impressions"] > 0 else 0.0
                    seg_row["cr"] = (seg_row["conversions"] / seg_row["clicks"] * 100.0) if seg_row["clicks"] > 0 else 0.0
                    seg_row["ctr_pct"] = seg_row["ctr"]
                    seg_row["cr_pct"] = seg_row["cr"]
                    seg_row["cr1_pct"] = (seg_row["leads"] / seg_row["clicks"] * 100.0) if seg_row["clicks"] > 0 else 0.0
                    seg_row["cr2_pct"] = (seg_row["target_leads"] / seg_row["leads"] * 100.0) if seg_row["leads"] > 0 else 0.0
                    seg_row["cpc"] = (seg_row["cost"] / seg_row["clicks"]) if seg_row["clicks"] > 0 else 0.0
                    seg_row["cpm"] = (seg_budget_basis / (seg_row["impressions"] / 1000.0)) if seg_row["impressions"] > 0 else 0.0
                    seg_row["cpa"] = (seg_budget_basis / seg_row["conversions"]) if seg_row["conversions"] > 0 else 0.0
                    seg_row["cpl"] = (seg_budget_basis / seg_row["leads"]) if seg_row["leads"] > 0 else 0.0
                    seg_row["cpql"] = (seg_budget_basis / seg_row["target_leads"]) if seg_row["target_leads"] > 0 else 0.0
                    seg_row["aov"] = (seg_row["revenue"] / seg_row["conversions"]) if seg_row["conversions"] > 0 else 0.0
                    seg_row["roas"] = (seg_row["revenue"] / seg_budget_basis) if seg_budget_basis > 0 else 0.0
                    seg_row["drr"] = (seg_budget_basis / seg_row["revenue"] * 100.0) if seg_row["revenue"] > 0 else 0.0
                    if is_diy_preset:
                        seg_row["cpa"] = (seg_row["cost_with_vat"] / seg_row["conversions"]) if seg_row["conversions"] > 0 else 0.0
                        seg_row["frequency"] = (seg_row["impressions"] / seg_row["reach"]) if seg_row["reach"] > 0 else 0.0
                        seg_row["available_capacity"] = float(seg_df["available_capacity"].sum()) if "available_capacity" in seg_df.columns else 0.0
                        seg_row["client_count"] = float(seg_df["client_count"].sum()) if "client_count" in seg_df.columns else 0.0
                        seg_row["absolute_new_clients"] = float(seg_df["absolute_new_clients"].sum()) if "absolute_new_clients" in seg_df.columns else 0.0
                        seg_row["returned_clients"] = float(seg_df["returned_clients"].sum()) if "returned_clients" in seg_df.columns else 0.0
                        seg_row["new_clients"] = float(seg_df["new_clients"].sum()) if "new_clients" in seg_df.columns else 0.0
                        seg_row["order_frequency"] = float(seg_df["order_frequency"].mean()) if "order_frequency" in seg_df.columns else 0.0
                        seg_row["sov_pct"] = (seg_row["reach"] / seg_row["available_capacity"] * 100.0) if seg_row["available_capacity"] > 0 else 0.0
                        seg_row["new_clients_share_pct"] = calc_new_clients_share_pct(seg_row["new_clients"], seg_row["client_count"])
                        seg_row["cac"] = (seg_row["cost_with_vat_ak"] / float(seg_row["new_clients"])) if float(seg_row["new_clients"]) > 0 else 0.0
                        seg_row["order_share_segment_pct"] = 100.0
                        seg_row["revenue_share_segment_pct"] = 100.0
                    segment_total_rows.append(seg_row)

            extra_agg_rows = pd.DataFrame.from_records(
                segment_total_rows + [total_row],
                columns=agg.columns,
            )
            agg = pd.DataFrame.from_records(
                agg.to_dict("records") + extra_agg_rows.to_dict("records"),
                columns=agg.columns,
            )

            if is_real_estate_preset:
                agg_cols = ["month_name"] + get_real_estate_table_cols(metric_mode)
            elif is_diy_preset:
                agg_cols = ["month_name"] + [
                    "impressions",
                    "clicks",
                    "ctr",
                    "cpc",
                    "cost",
                    "ak_cost_wo_vat",
                    "total_budget_wo_vat_ak",
                    "cost_with_vat",
                    "cost_with_vat_ak",
                    "vat_amount",
                    "ak_rate_pct",
                    "cr",
                    "conversions",
                    "cpa",
                    "aov",
                    "revenue",
                    "roas",
                    "drr",
                    "available_capacity",
                    "client_count",
                    "absolute_new_clients",
                    "returned_clients",
                    "new_clients",
                    "cac",
                    "order_frequency",
                    "shipped_orders",
                    "shipped_cps",
                    "shipped_aov",
                    "shipped_revenue",
                    "shipped_roas",
                    "shipped_drr_pct",
                    "shipped_order_share_segment_pct",
                    "shipped_revenue_share_segment_pct",
                    "sov_pct",
                    "new_clients_share_pct",
                    "order_share_segment_pct",
                    "revenue_share_segment_pct",
                    "budget_share_segment_pct",
                ]
            else:
                agg_cols = [
                    "month_name",
                    "impressions",
                    "clicks",
                    "conversions",
                    "cost",
                    "cost_with_vat",
                    "cost_with_vat_ak",
                    "revenue",
                    "ctr",
                    "cpc",
                    "cr",
                    "cpa",
                    "roas",
                    "drr",
                ]
            agg_show = safe_select_columns(agg, agg_cols)
            agg_show = agg_show.rename(columns={"roas": "ROAS"})
            agg_show = agg_show.rename(columns=DISPLAY_COL_RENAME)

            if is_real_estate_preset:
                numeric_cols = [
                    "impressions", "clicks", "leads", "target_leads", "cost", "cost_with_vat", "cost_with_vat_ak",
                    "ctr", "cpc", "cr_pct", "cr1_pct", "cr2_pct", "cpl", "cpql"
                ]
            else:
                numeric_cols = [
                    "impressions", "clicks", "conversions", "cost", "ak_cost_wo_vat", "total_budget_wo_vat_ak", "cost_with_vat", "cost_with_vat_ak", "vat_amount", "ak_rate_pct", "revenue",
                    "ctr", "cpc", "cr", "aov", "cpa", "ROAS", "drr", "available_capacity", "client_count", "absolute_new_clients", "returned_clients", "new_clients", "cac", "order_frequency", "shipped_orders", "shipped_cps", "shipped_aov", "shipped_revenue", "shipped_roas", "shipped_drr_pct", "shipped_order_share_segment_pct", "shipped_revenue_share_segment_pct", "sov_pct", "new_clients_share_pct", "order_share_segment_pct", "revenue_share_segment_pct", "budget_share_segment_pct"
                ]
            numeric_cols = [DISPLAY_COL_RENAME.get(c, c) for c in numeric_cols]
            numeric_cols = [c for c in numeric_cols if c in agg_show.columns]

            def _blend_hex(c1: str, c2: str, t: float) -> str:
                t = max(0.0, min(1.0, t))
                r1, g1, b1 = int(c1[1:3], 16), int(c1[3:5], 16), int(c1[5:7], 16)
                r2, g2, b2 = int(c2[1:3], 16), int(c2[3:5], 16), int(c2[5:7], 16)
                r = int(r1 + (r2 - r1) * t)
                g = int(g1 + (g2 - g1) * t)
                b = int(b1 + (b2 - b1) * t)
                return f"#{r:02x}{g:02x}{b:02x}"

            def _style_metric_col(col: pd.Series) -> pd.Series:
                styles = pd.Series([""] * len(col), index=col.index)
                mask = ~agg_show["Месяц"].astype(str).str.startswith("Итого")
                mask &= agg_show["Месяц"] != "TOTAL"
                vals = pd.to_numeric(col[mask], errors="coerce").dropna()
                if vals.empty:
                    return styles
                vmin, vmax = float(vals.min()), float(vals.max())
                span = (vmax - vmin) if vmax != vmin else 1.0
                base_low = "#173058"
                base_high = "#19B8B2"
                for idx in vals.index:
                    t = (float(col.loc[idx]) - vmin) / span
                    color = _blend_hex(base_low, base_high, t)
                    styles.loc[idx] = (
                        f"background-color: {color}; color: #EAF3EE; border: 1px solid {THEME_BORDER};"
                    )
                return styles

            def _highlight_total_row_total(row):
                style = [""] * len(row)
                month_val = str(row["Месяц"])
                if month_val == "TOTAL":
                    style = [
                        f"background-color: #00CDC5; color: #081521; font-weight: 700; border-top: 2px solid {THEME_BORDER};"
                    ] * len(row)
                elif month_val.startswith("Итого "):
                    style = [
                        f"background-color: #2C6E75; color: #DDEAF0; font-weight: 650; border-top: 1px solid {THEME_BORDER};"
                    ] * len(row)
                return style

            total_formatters = {
                DISPLAY_COL_RENAME["impressions"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["reach"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["frequency"]: lambda x: "" if pd.isna(x) else f"{x:.2f}",
                DISPLAY_COL_RENAME["clicks"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["conversions"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["leads"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["target_leads"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["cost"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["ak_cost_wo_vat"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["total_budget_wo_vat_ak"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["cost_with_vat"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["cost_with_vat_ak"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["vat_amount"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["ak_rate_pct"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["revenue"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["aov"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["cpm"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["cpa"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["cpl"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["cpql"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["available_capacity"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["client_count"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["absolute_new_clients"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["returned_clients"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["new_clients"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["cac"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["order_frequency"]: lambda x: "" if pd.isna(x) else f"{x:.2f}",
                DISPLAY_COL_RENAME["shipped_orders"]: lambda x: "" if pd.isna(x) else f"{round(x):,}".replace(",", " "),
                DISPLAY_COL_RENAME["shipped_cps"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["shipped_aov"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["shipped_revenue"]: lambda x: "" if pd.isna(x) else f"{round(x):,} \u20BD".replace(",", " "),
                DISPLAY_COL_RENAME["shipped_roas"]: lambda x: "" if pd.isna(x) else f"{x:.2f}",
                DISPLAY_COL_RENAME["shipped_drr_pct"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["shipped_order_share_segment_pct"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["shipped_revenue_share_segment_pct"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["cpc"]: lambda x: "" if pd.isna(x) else f"{x:.2f} \u20BD",
                DISPLAY_COL_RENAME["ctr"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["cr"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["cr1_pct"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["cr2_pct"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["sov_pct"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["new_clients_share_pct"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["order_share_segment_pct"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["revenue_share_segment_pct"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["budget_share_segment_pct"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                DISPLAY_COL_RENAME["drr"]: lambda x: "" if pd.isna(x) else f"{x:.2f} %",
                "ROAS": lambda x: "" if pd.isna(x) else f"{x:.2f}",
            }
            total_formatters = {k: v for k, v in total_formatters.items() if k in agg_show.columns}
            styled_total = (
                agg_show.style
                .format(total_formatters)
                .apply(_highlight_total_row_total, axis=1)
            )
            for c in numeric_cols:
                styled_total = styled_total.apply(_style_metric_col, axis=0, subset=[c])

            st.dataframe(styled_total, width="stretch")

    render_bottom_tab_switcher("Медиаплан", "plan")
# ====================================================================
#                           ТАБ «ДИАГРАММЫ»
# ====================================================================

with tab_charts:
    ui_section_title("Сводная таблица и диаграммы")
    use_vat_budget_metrics = st.session_state.get("use_vat_budget_metrics", True)
    use_ak_budget_metrics = st.session_state.get("use_ak_budget_metrics", False)
    st.markdown(
        """
        <div style="
            margin: 0.15rem 0 0.55rem 0;
            padding: 10px 12px;
            border-radius: 10px;
            border: 1px solid #2B7EE8;
            background: rgba(10, 116, 242, 0.12);
        ">
            <span style="font-weight: 800; color: #9EC5FF;">Назначение блока:</span>
            <span style="color: #EAF0FF; font-weight: 560;">
                Вкладка используется для финальной проверки расчетов медиаплана: через фильтры и визуализации можно валидировать корректность формул, распределений и итоговых значений.
            </span>
        </div>
        """,
        unsafe_allow_html=True,
    )
    charts_intro_html = """
        <div class="tab-intro">
            <p>1) Сначала выберите период (месяцы), который хотите проверить, или оставьте все настройки и проверьте весь период расчета медиаплана целиком.</p>
            <p>2) Примените фильтры по нужным разрезам (тип кампании, система, формат), чтобы сузить данные.</p>
            <p>3) Проверьте сводную таблицу по месяцам: она показывает итоговые метрики и помогает быстро увидеть отклонения.</p>
            <p>4) Сравните динамику ключевых показателей (показы, клики, расход, конверсии, выручка, ROAS/ДРР) между месяцами.</p>
            <p>5) При необходимости включите круговые диаграммы, чтобы проверить структуру распределения бюджета, конверсий и выручки по типам РК.</p>
            <p>6) Если цифры выглядят нелогично, вернитесь во вкладку «Медиаплан» и внесите точечные корректировки в нужные месяцы, затем вернитесь в «Диаграммы» и проведите повторную проверку.</p>
        </div>
    """
    if is_real_estate_preset:
        charts_intro_html = """
        <div class="tab-intro">
            <p>1) Сначала выберите период (месяцы), который хотите проверить, или оставьте все настройки и проверьте весь период расчета медиаплана целиком.</p>
            <p>2) Примените фильтры по нужным разрезам (тип кампании, система, формат), чтобы сузить данные.</p>
            <p>3) Проверьте сводную таблицу по месяцам: она показывает итоговые метрики и помогает быстро увидеть отклонения.</p>
            <p>4) Сравните динамику ключевых показателей (показы, клики, бюджет, лиды, ЦО, CPL/CPQL) между месяцами.</p>
            <p>5) При необходимости включите круговые диаграммы, чтобы проверить структуру распределения бюджета, лидов и ЦО по типам РК.</p>
            <p>6) Если цифры выглядят нелогично, вернитесь во вкладку «Медиаплан» и внесите точечные корректировки в нужные месяцы, затем вернитесь в «Диаграммы» и проведите повторную проверку.</p>
        </div>
        """
    st.markdown(charts_intro_html, unsafe_allow_html=True)

    if "df_all" not in locals() or df_all.empty:
        st.info("Сначала заполните данные на вкладке «Медиаплан» и рассчитайте месяцы.")
    else:
        st.caption("Выберите месяцы и типы РК для сводной таблицы и круговых диаграмм.")
        st.caption(
            "Режим бюджетных метрик: "
            + (
                "с НДС 22% и АК"
                if use_ak_budget_metrics and use_vat_budget_metrics
                else "с НДС 22%"
                if use_vat_budget_metrics
                else "без НДС"
            )
        )

        chart_campaign_palette = [
            "#0066E0", "#00CDC5", "#9747FF", "#FF6333",
            "#3D8EF0", "#42DDD6", "#B07BFF", "#FF8A66",
            "#2D5BFF", "#17B890", "#C77DFF", "#FF9F1C",
        ]

        def _build_campaign_color_map(campaign_names: list[str]) -> dict[str, str]:
            ordered_names = []
            for name in campaign_names:
                name_str = str(name).strip()
                if name_str and name_str not in ordered_names:
                    ordered_names.append(name_str)
            return {
                name: chart_campaign_palette[idx % len(chart_campaign_palette)]
                for idx, name in enumerate(ordered_names)
            }

        available_months_df = (
            df_all[["planning_slot", "month_name"]]
            .dropna(subset=["planning_slot"])
            .drop_duplicates()
            .sort_values("planning_slot")
        )
        month_options = [
            f"{int(r['planning_slot'])}. {str(r['month_name'])}"
            for _, r in available_months_df.iterrows()
        ]
        sync_multiselect_state("charts_months_multiselect", month_options, default_to_all=True)

        selected_labels = st.multiselect(
            "Месяцы в сводной таблице и диаграммах:",
            options=month_options,
            default=month_options,
            key="charts_months_multiselect",
        )

        if not selected_labels:
            st.info("Выберите хотя бы один месяц.")
        else:
            selected_month_slots_charts = [int(label.split(".")[0]) for label in selected_labels]

            all_campaign_types = df_all["campaign_type"].unique().tolist()
            campaign_color_map = _build_campaign_color_map(all_campaign_types)
            selected_campaign_types = st.multiselect(
                "Типы РК для включения:",
                options=all_campaign_types,
                default=all_campaign_types,
                key="charts_campaign_types_multiselect",
            )

            if not selected_campaign_types:
                st.info("Выберите хотя бы один тип РК.")
            else:
                selected_segments_charts = None
                if is_diy_preset and "segment" in df_all.columns:
                    all_segments = sorted(df_all["segment"].dropna().astype(str).unique().tolist())
                    selected_segments_charts = st.multiselect(
                        "Сегменты для включения:",
                        options=all_segments,
                        default=all_segments,
                        key="charts_segments_multiselect",
                    )
                    if not selected_segments_charts:
                        st.info("Выберите хотя бы один сегмент.")
                if is_diy_preset and selected_segments_charts == []:
                    df_sel = pd.DataFrame()
                else:
                    mask = df_all["planning_slot"].isin(selected_month_slots_charts) & df_all[
                        "campaign_type"
                    ].isin(selected_campaign_types)
                    if selected_segments_charts is not None:
                        mask &= df_all["segment"].isin(selected_segments_charts)
                    df_sel = df_all[mask]

                if df_sel.empty:
                    st.info("Для выбранных месяцев и типов РК нет данных.")
                else:
                    # ---------- Сводная таблица по месяцам ----------
                    agg_m = df_sel.groupby(["planning_slot", "month_num", "month_year", "month_name"], as_index=False).agg(
                        impressions=("impressions", "sum"),
                        clicks=("clicks", "sum"),
                        conversions=("conversions", "sum"),
                        leads=("leads", "sum"),
                        target_leads=("target_leads", "sum"),
                        cost=("cost", "sum"),
                        cost_with_vat=("cost_with_vat", "sum"),
                        cost_with_vat_ak=("cost_with_vat_ak", "sum"),
                        ak_cost_wo_vat=("ak_cost_wo_vat", "sum"),
                        revenue=("revenue", "sum"),
                    )
                    if is_diy_preset:
                        agg_m["reach"] = df_sel.groupby(["planning_slot", "month_num", "month_year", "month_name"])["reach"].sum().values if "reach" in df_sel.columns else 0.0
                        agg_m["frequency"] = np.where(agg_m["reach"] > 0, agg_m["impressions"] / agg_m["reach"], 0.0)
                        agg_m["available_capacity"] = df_sel.groupby(["planning_slot", "month_num", "month_year", "month_name"])["available_capacity"].sum().values if "available_capacity" in df_sel.columns else 0.0
                        agg_m["client_count"] = df_sel.groupby(["planning_slot", "month_num", "month_year", "month_name"])["client_count"].sum().values if "client_count" in df_sel.columns else 0.0
                        agg_m["absolute_new_clients"] = df_sel.groupby(["planning_slot", "month_num", "month_year", "month_name"])["absolute_new_clients"].sum().values if "absolute_new_clients" in df_sel.columns else 0.0
                        agg_m["returned_clients"] = df_sel.groupby(["planning_slot", "month_num", "month_year", "month_name"])["returned_clients"].sum().values if "returned_clients" in df_sel.columns else 0.0
                        agg_m["new_clients"] = df_sel.groupby(["planning_slot", "month_num", "month_year", "month_name"])["new_clients"].sum().values if "new_clients" in df_sel.columns else 0.0
                        agg_m["order_frequency"] = df_sel.groupby(["planning_slot", "month_num", "month_year", "month_name"])["order_frequency"].mean().values if "order_frequency" in df_sel.columns else 0.0
                        agg_m["shipped_orders"] = df_sel.groupby(["planning_slot", "month_num", "month_year", "month_name"])["shipped_orders"].sum().values if "shipped_orders" in df_sel.columns else 0.0
                        agg_m["shipped_revenue"] = df_sel.groupby(["planning_slot", "month_num", "month_year", "month_name"])["shipped_revenue"].sum().values if "shipped_revenue" in df_sel.columns else 0.0
                        agg_m["sov_pct"] = np.where(agg_m["available_capacity"] > 0, agg_m["reach"] / agg_m["available_capacity"] * 100.0, 0.0)
                        agg_m["new_clients_share_pct"] = np.where(agg_m["client_count"] > 0, agg_m["new_clients"] / agg_m["client_count"] * 100.0, 0.0)
                    if use_ak_budget_metrics:
                        budget_series = agg_m["cost_with_vat_ak"] if use_vat_budget_metrics else (agg_m["cost"] + agg_m["ak_cost_wo_vat"])
                    else:
                        budget_series = agg_m["cost_with_vat"] if use_vat_budget_metrics else agg_m["cost"]
                    agg_m["total_budget_wo_vat_ak"] = agg_m["cost"] + agg_m["ak_cost_wo_vat"]
                    agg_m["vat_amount"] = agg_m["cost_with_vat_ak"] - agg_m["total_budget_wo_vat_ak"]
                    agg_m["ak_rate_pct"] = np.where(agg_m["cost"] > 0, agg_m["ak_cost_wo_vat"] / agg_m["cost"] * 100.0, 0.0)

                    agg_m["ctr"] = np.where(
                        agg_m["impressions"] > 0,
                        agg_m["clicks"] / agg_m["impressions"] * 100,
                        0.0,
                    )
                    agg_m["cr"] = np.where(
                        agg_m["clicks"] > 0,
                        agg_m["conversions"] / agg_m["clicks"] * 100,
                        0.0,
                    )
                    agg_m["cr1_pct"] = np.where(
                        agg_m["clicks"] > 0,
                        agg_m["leads"] / agg_m["clicks"] * 100,
                        0.0,
                    )
                    agg_m["cr2_pct"] = np.where(
                        agg_m["leads"] > 0,
                        agg_m["target_leads"] / agg_m["leads"] * 100,
                        0.0,
                    )
                    agg_m["cpm"] = np.where(
                        agg_m["impressions"] > 0,
                        budget_series / (agg_m["impressions"] / 1000),
                        0.0,
                    )
                    agg_m["cpa"] = np.where(
                        agg_m["conversions"] > 0,
                        budget_series / agg_m["conversions"],
                        0.0,
                    )
                    agg_m["cpl"] = np.where(
                        agg_m["leads"] > 0,
                        budget_series / agg_m["leads"],
                        0.0,
                    )
                    agg_m["cpql"] = np.where(
                        agg_m["target_leads"] > 0,
                        budget_series / agg_m["target_leads"],
                        0.0,
                    )
                    agg_m["aov"] = np.where(
                        agg_m["conversions"] > 0,
                        agg_m["revenue"] / agg_m["conversions"],
                        0.0,
                    )
                    agg_m["cpc"] = np.where(
                        agg_m["clicks"] > 0,
                        agg_m["cost"] / agg_m["clicks"],
                        0.0,
                    )
                    agg_m["roas"] = np.where(
                        budget_series > 0,
                        agg_m["revenue"] / budget_series,
                        0.0,
                    )
                    agg_m["drr"] = np.where(
                        agg_m["revenue"] > 0,
                        budget_series / agg_m["revenue"] * 100,
                        0.0,
                    )
                    if is_diy_preset:
                        agg_m["cpa"] = np.where(
                            agg_m["conversions"] > 0,
                            agg_m["cost_with_vat"] / agg_m["conversions"],
                            0.0,
                        )
                        agg_m["shipped_cps"] = np.where(
                            agg_m["shipped_orders"] > 0,
                            agg_m["cost_with_vat"] / agg_m["shipped_orders"],
                            0.0,
                        )
                        agg_m["shipped_aov"] = np.where(
                            agg_m["shipped_orders"] > 0,
                            agg_m["shipped_revenue"] / agg_m["shipped_orders"],
                            0.0,
                        )
                        agg_m["shipped_roas"] = np.where(
                            budget_series > 0,
                            agg_m["shipped_revenue"] / budget_series,
                            0.0,
                        )
                        agg_m["shipped_drr_pct"] = np.where(
                            agg_m["shipped_revenue"] > 0,
                            budget_series / agg_m["shipped_revenue"] * 100.0,
                            0.0,
                        )
                        agg_m["order_share_segment_pct"] = 100.0
                        agg_m["revenue_share_segment_pct"] = 100.0
                        agg_m["shipped_order_share_segment_pct"] = 100.0
                        agg_m["shipped_revenue_share_segment_pct"] = 100.0
                        agg_m["budget_share_segment_pct"] = 100.0
                        agg_m["cac"] = np.where(
                            agg_m["new_clients"] > 0,
                            agg_m["cost_with_vat_ak"] / agg_m["new_clients"],
                            0.0,
                        )

                    selected_month_slots_charts = sorted(selected_month_slots_charts)
                    month_name_by_slot = {
                        int(r["planning_slot"]): str(r["month_name"])
                        for _, r in agg_m[["planning_slot", "month_name"]].drop_duplicates().iterrows()
                    }
                    month_headers = [month_name_by_slot.get(slot, str(slot)) for slot in selected_month_slots_charts]

                    # Коэффициенты, использованные в расчетах (в среднем по выбранным типам РК)
                    if "campaigns" in locals() and not campaigns.empty:
                        base_for_coeffs = campaigns[
                            campaigns["campaign_type"].isin(selected_campaign_types)
                        ][
                            [
                                "campaign_type",
                                "impressions_avg",
                                "ctr_avg_percent",
                                "cpc_avg",
                                "cr_avg_percent",
                                "cr2_avg_percent",
                                "aov_avg",
                            ]
                        ].copy()
                        if not base_for_coeffs.empty:
                            coeff_raw = df_sel.merge(base_for_coeffs, on="campaign_type", how="left")
                            coeff_raw["base_ctr"] = coeff_raw["ctr_avg_percent"] / 100.0
                            coeff_raw["base_cr"] = coeff_raw["cr_avg_percent"] / 100.0
                            coeff_raw["base_cr2"] = coeff_raw["cr2_avg_percent"] / 100.0

                            coeff_raw["k_imp"] = np.where(
                                coeff_raw["impressions_avg"] > 0,
                                coeff_raw["impressions"] / coeff_raw["impressions_avg"],
                                np.nan,
                            )
                            coeff_raw["k_ctr"] = np.where(
                                coeff_raw["base_ctr"] > 0,
                                coeff_raw["ctr"] / coeff_raw["base_ctr"],
                                np.nan,
                            )
                            coeff_raw["k_cpc"] = np.where(
                                coeff_raw["cpc_avg"] > 0,
                                coeff_raw["cpc"] / coeff_raw["cpc_avg"],
                                np.nan,
                            )
                            coeff_raw["k_cr"] = np.where(
                                coeff_raw["base_cr"] > 0,
                                coeff_raw["cr"] / coeff_raw["base_cr"],
                                np.nan,
                            )
                            coeff_raw["k_cr2"] = np.where(
                                coeff_raw["base_cr2"] > 0,
                                coeff_raw.get("cr2", np.nan) / coeff_raw["base_cr2"],
                                np.nan,
                            )
                            coeff_raw["k_aov"] = np.where(
                                coeff_raw["aov_avg"] > 0,
                                coeff_raw["aov"] / coeff_raw["aov_avg"],
                                np.nan,
                            )

                            agg_kwargs = {
                                "k_imp": ("k_imp", "mean"),
                                "k_ctr": ("k_ctr", "mean"),
                                "k_cpc": ("k_cpc", "mean"),
                                "k_cr": ("k_cr", "mean"),
                            }
                            if is_real_estate_preset and metric_mode["is_real_estate_full"]:
                                agg_kwargs["k_cr2"] = ("k_cr2", "mean")
                            if not is_real_estate_preset:
                                agg_kwargs["k_aov"] = ("k_aov", "mean")
                            coeff_month = coeff_raw.groupby("planning_slot", as_index=False).agg(**agg_kwargs)
                            coeff_month_map = coeff_month.set_index("planning_slot").to_dict("index")

                            # Отдельно показываем коэффициент "медийных хвостов" как значение из набора
                            # (без усреднения по типам РК)
                            media_tail_month_map = {}
                            media_tail_values_by_set = {}
                            coeff_sets_for_tail = st.session_state.get("coeff_sets", [])
                            for cs_tail in coeff_sets_for_tail:
                                if normalize_coeff_set_type(cs_tail.get("type")) != "Медийные хвосты":
                                    continue
                                cs_tail_name = str(cs_tail.get("name", "")).strip()
                                df_tail = cs_tail.get("result")
                                if not cs_tail_name or df_tail is None or getattr(df_tail, "empty", True):
                                    continue
                                if "Номер месяца" not in df_tail.columns or "Коэф." not in df_tail.columns:
                                    continue
                                mm = {}
                                for _, rtail in df_tail.iterrows():
                                    try:
                                        mm[int(rtail["Номер месяца"])] = float(rtail["Коэф."])
                                    except Exception:
                                        continue
                                media_tail_values_by_set[cs_tail_name] = mm

                            coeff_links_for_tail = st.session_state.get("coeff_sets_links_new", pd.DataFrame())
                            coeff_links_for_tail = coeff_links_for_tail[
                                coeff_links_for_tail["campaign_type"].isin(selected_campaign_types)
                            ] if not coeff_links_for_tail.empty else pd.DataFrame()

                            for m in selected_month_slots_charts:
                                set_names = []
                                if not coeff_links_for_tail.empty and "media_tail_set" in coeff_links_for_tail.columns:
                                    for _, rr in coeff_links_for_tail.iterrows():
                                        set_name = str(rr.get("media_tail_set", "")).strip()
                                        if set_name:
                                            set_names.append(set_name)

                                uniq_set_names = sorted(set(set_names))
                                if len(uniq_set_names) == 1:
                                        media_tail_month_map[m] = float(
                                            media_tail_values_by_set.get(uniq_set_names[0], {}).get(m, 1.0)
                                        )
                                elif len(uniq_set_names) == 0:
                                    media_tail_month_map[m] = 1.0
                                else:
                                    media_tail_month_map[m] = None

                            coeff_rows = []
                            coeff_labels = [
                                ("k_imp", "Коэф. показы"),
                                ("k_ctr", "Коэф. CTR"),
                                ("k_cpc", "Коэф. CPC"),
                                ("k_cr", "Коэф. CR"),
                            ]
                            if is_real_estate_preset and metric_mode["is_real_estate_full"]:
                                coeff_labels.append(("k_cr2", "Коэф. CR2"))
                            if not is_real_estate_preset:
                                coeff_labels.append(("k_aov", "Коэф. AOV"))
                            for key_name, label in coeff_labels:
                                row = {"Метрика": label}
                                for m in selected_month_slots_charts:
                                    val = coeff_month_map.get(m, {}).get(key_name, np.nan)
                                    row[month_name_by_slot.get(m, str(m))] = "" if pd.isna(val) else f"{val:.2f}"
                                coeff_rows.append(row)

                            row_media_tail = {"Метрика": "Коэф. медийные хвосты"}
                            for m in selected_month_slots_charts:
                                v = media_tail_month_map.get(m, 1.0)
                                row_media_tail[month_name_by_slot.get(m, str(m))] = "неск." if v is None else f"{v:.2f}"
                            coeff_rows.append(row_media_tail)

                            coeff_table = pd.DataFrame(coeff_rows).set_index("Метрика")
                            with st.expander("Коэффициенты, использованные в расчете по месяцам", expanded=False):
                                st.dataframe(coeff_table, width="stretch")
                            st.caption("Коэф. медийные хвосты: 1.00 = хвосты не применяются.")

                    # Транспонированная таблица: месяцы в колонках, метрики в строках
                    agg_by_month = agg_m.set_index("planning_slot").to_dict("index")

                    total_impressions = agg_m["impressions"].sum()
                    total_clicks = agg_m["clicks"].sum()
                    total_conversions = agg_m["conversions"].sum()
                    total_cost = agg_m["cost"].sum()
                    total_cost_with_vat = agg_m["cost_with_vat"].sum()
                    total_cost_with_vat_ak = agg_m["cost_with_vat_ak"].sum()
                    total_ak_wo_vat = agg_m["ak_cost_wo_vat"].sum()
                    total_revenue = agg_m["revenue"].sum()
                    if use_ak_budget_metrics:
                        total_budget_basis = total_cost_with_vat_ak if use_vat_budget_metrics else (total_cost + total_ak_wo_vat)
                    else:
                        total_budget_basis = total_cost_with_vat if use_vat_budget_metrics else total_cost
                    total_metrics = {
                        "impressions": total_impressions,
                        "clicks": total_clicks,
                        "ctr": (total_clicks / total_impressions * 100) if total_impressions > 0 else 0.0,
                        "cpc": (total_cost / total_clicks) if total_clicks > 0 else 0.0,
                        "cpm": (total_budget_basis / (total_impressions / 1000.0)) if total_impressions > 0 else 0.0,
                        "cost": total_cost,
                        "cost_with_vat": total_cost_with_vat,
                        "cost_with_vat_ak": total_cost_with_vat_ak,
                        "cr": (total_conversions / total_clicks * 100) if total_clicks > 0 else 0.0,
                        "cr1_pct": (float(agg_m["leads"].sum()) / total_clicks * 100) if total_clicks > 0 else 0.0,
                        "cr2_pct": (float(agg_m["target_leads"].sum()) / float(agg_m["leads"].sum()) * 100) if float(agg_m["leads"].sum()) > 0 else 0.0,
                        "leads": float(agg_m["leads"].sum()),
                        "target_leads": float(agg_m["target_leads"].sum()),
                        "cpl": (total_budget_basis / float(agg_m["leads"].sum())) if float(agg_m["leads"].sum()) > 0 else 0.0,
                        "cpql": (total_budget_basis / float(agg_m["target_leads"].sum())) if float(agg_m["target_leads"].sum()) > 0 else 0.0,
                        "aov": (total_revenue / total_conversions) if total_conversions > 0 else 0.0,
                        "conversions": total_conversions,
                        "cpa": (total_budget_basis / total_conversions) if total_conversions > 0 else 0.0,
                        "revenue": total_revenue,
                        "roas": (total_revenue / total_budget_basis) if total_budget_basis > 0 else 0.0,
                        "drr": (total_budget_basis / total_revenue * 100) if total_revenue > 0 else 0.0,
                    }
                    if is_diy_preset:
                        total_reach = float(agg_m["reach"].sum()) if "reach" in agg_m.columns else 0.0
                        total_available_capacity = float(agg_m["available_capacity"].sum()) if "available_capacity" in agg_m.columns else 0.0
                        total_client_count = float(agg_m["client_count"].sum()) if "client_count" in agg_m.columns else 0.0
                        total_new_clients = float(agg_m["new_clients"].sum()) if "new_clients" in agg_m.columns else 0.0
                        total_shipped_orders = float(agg_m["shipped_orders"].sum()) if "shipped_orders" in agg_m.columns else 0.0
                        total_shipped_revenue = float(agg_m["shipped_revenue"].sum()) if "shipped_revenue" in agg_m.columns else 0.0
                        total_metrics.update(
                            {
                                "total_budget_wo_vat_ak": total_cost + total_ak_wo_vat,
                                "vat_amount": total_cost_with_vat_ak - (total_cost + total_ak_wo_vat),
                                "ak_rate_pct": (total_ak_wo_vat / total_cost * 100.0) if total_cost > 0 else 0.0,
                                "reach": total_reach,
                                "frequency": (total_impressions / total_reach) if total_reach > 0 else 0.0,
                                "available_capacity": total_available_capacity,
                                "client_count": total_client_count,
                                "absolute_new_clients": float(agg_m["absolute_new_clients"].sum()) if "absolute_new_clients" in agg_m.columns else 0.0,
                                "returned_clients": float(agg_m["returned_clients"].sum()) if "returned_clients" in agg_m.columns else 0.0,
                                "new_clients": total_new_clients,
                                "cac": (total_cost_with_vat_ak / total_new_clients) if total_new_clients > 0 else 0.0,
                                "order_frequency": float(agg_m["order_frequency"].mean()) if "order_frequency" in agg_m.columns else 0.0,
                                "shipped_orders": total_shipped_orders,
                                "shipped_cps": (total_cost_with_vat / total_shipped_orders) if total_shipped_orders > 0 else 0.0,
                                "shipped_aov": (total_shipped_revenue / total_shipped_orders) if total_shipped_orders > 0 else 0.0,
                                "shipped_revenue": total_shipped_revenue,
                                "shipped_roas": (total_shipped_revenue / total_budget_basis) if total_budget_basis > 0 else 0.0,
                                "shipped_drr_pct": (total_budget_basis / total_shipped_revenue * 100.0) if total_shipped_revenue > 0 else 0.0,
                                "shipped_order_share_segment_pct": 100.0,
                                "shipped_revenue_share_segment_pct": 100.0,
                                "sov_pct": (total_reach / total_available_capacity * 100.0) if total_available_capacity > 0 else 0.0,
                                "new_clients_share_pct": calc_new_clients_share_pct(total_new_clients, total_client_count),
                                "order_share_segment_pct": 100.0,
                                "revenue_share_segment_pct": 100.0,
                                "budget_share_segment_pct": 100.0,
                            }
                        )

                    def _fmt_int(v):
                        return f"{round(v):,}".replace(",", " ")

                    def _fmt_rub(v):
                        return f"{round(v):,} ₽".replace(",", " ")

                    def _fmt_rub2(v):
                        return f"{v:.2f} ₽"

                    def _fmt_pct(v):
                        return f"{v:.2f} %"

                    def _fmt_mult(v):
                        return f"{v:.2f}"

                    if is_real_estate_preset:
                        metric_specs = get_real_estate_display_metric_specs(metric_mode)
                    elif is_diy_preset:
                        metric_specs = get_diy_display_metric_specs()
                    else:
                        metric_specs = [
                            ("Показы", "impressions"),
                            ("Клики", "clicks"),
                            ("CTR", "ctr"),
                            ("CPC", "cpc"),
                            ("Расход", "cost"),
                            ("Бюджет с НДС", "cost_with_vat"),
                            ("Бюджет с НДС и АК", "cost_with_vat_ak"),
                            ("CR", "cr"),
                            ("AOV", "aov"),
                            ("Конверсии", "conversions"),
                            ("CPO", "cpa"),
                            ("Выручка", "revenue"),
                            ("ROAS", "roas"),
                            ("ДРР", "drr"),
                        ]

                    rows = []
                    metric_series_map = {}
                    for metric_label, metric_key in metric_specs:
                        row = {"Метрика": metric_label}
                        month_value_map = {}
                        for m in selected_month_slots_charts:
                            month_col = month_name_by_slot.get(m, str(m))
                            val = agg_by_month.get(m, {}).get(metric_key, 0.0)
                            month_value_map[month_col] = float(val)
                            if metric_key in ["impressions", "clicks", "conversions", "leads", "target_leads", "reach", "available_capacity", "client_count", "absolute_new_clients", "returned_clients", "new_clients", "shipped_orders"]:
                                row[month_col] = _fmt_int(val)
                            elif metric_key in ["cost", "ak_cost_wo_vat", "total_budget_wo_vat_ak", "cost_with_vat", "cost_with_vat_ak", "vat_amount", "revenue", "cpm", "cpa", "aov", "cpl", "cpql", "cac", "shipped_cps", "shipped_aov", "shipped_revenue"]:
                                row[month_col] = _fmt_rub(val)
                            elif metric_key == "cpc":
                                row[month_col] = _fmt_rub2(val)
                            elif metric_key in ["ctr", "cr", "drr", "cr1_pct", "cr2_pct", "ak_rate_pct", "shipped_drr_pct", "shipped_order_share_segment_pct", "shipped_revenue_share_segment_pct", "sov_pct", "new_clients_share_pct", "order_share_segment_pct", "revenue_share_segment_pct", "budget_share_segment_pct"]:
                                row[month_col] = _fmt_pct(val)
                            elif metric_key in ["roas", "shipped_roas"]:
                                row[month_col] = _fmt_mult(val)
                            elif metric_key == "order_frequency":
                                row[month_col] = f"{float(val):.2f}"
                            else:
                                row[month_col] = val

                        total_val = total_metrics.get(metric_key, 0.0)
                        if metric_key in ["impressions", "clicks", "conversions", "leads", "target_leads", "reach", "available_capacity", "client_count", "absolute_new_clients", "returned_clients", "new_clients", "shipped_orders"]:
                            row["TOTAL"] = _fmt_int(total_val)
                        elif metric_key in ["cost", "ak_cost_wo_vat", "total_budget_wo_vat_ak", "cost_with_vat", "cost_with_vat_ak", "vat_amount", "revenue", "cpm", "cpa", "aov", "cpl", "cpql", "cac", "shipped_cps", "shipped_aov", "shipped_revenue"]:
                            row["TOTAL"] = _fmt_rub(total_val)
                        elif metric_key == "cpc":
                            row["TOTAL"] = _fmt_rub2(total_val)
                        elif metric_key in ["ctr", "cr", "drr", "cr1_pct", "cr2_pct", "ak_rate_pct", "shipped_drr_pct", "shipped_order_share_segment_pct", "shipped_revenue_share_segment_pct", "sov_pct", "new_clients_share_pct", "order_share_segment_pct", "revenue_share_segment_pct", "budget_share_segment_pct"]:
                            row["TOTAL"] = _fmt_pct(total_val)
                        elif metric_key in ["roas", "shipped_roas"]:
                            row["TOTAL"] = _fmt_mult(total_val)
                        elif metric_key == "order_frequency":
                            row["TOTAL"] = f"{float(total_val):.2f}"
                        else:
                            row["TOTAL"] = _fmt_int(total_val)
                        rows.append(row)
                        metric_series_map[metric_label] = {
                            "metric_key": metric_key,
                            "months": month_value_map,
                            "total": float(total_val),
                        }

                    summary_matrix = pd.DataFrame(rows).set_index("Метрика")
                    ui_section_title("Сводная таблица (месяцы в колонках)")
                    heat_cols = month_headers

                    def _parse_num_for_heat(v):
                        if pd.isna(v):
                            return None
                        if isinstance(v, (int, float, np.integer, np.floating)):
                            return float(v)
                        s = str(v).strip()
                        if not s:
                            return None
                        for token in ["₽", "%", "×"]:
                            s = s.replace(token, "")
                        s = s.replace(" ", "").replace(",", ".")
                        try:
                            return float(s)
                        except Exception:
                            return None

                    def _blend_hex(c1: str, c2: str, t: float) -> str:
                        t = max(0.0, min(1.0, t))
                        r1, g1, b1 = int(c1[1:3], 16), int(c1[3:5], 16), int(c1[5:7], 16)
                        r2, g2, b2 = int(c2[1:3], 16), int(c2[3:5], 16), int(c2[5:7], 16)
                        r = int(r1 + (r2 - r1) * t)
                        g = int(g1 + (g2 - g1) * t)
                        b = int(b1 + (b2 - b1) * t)
                        return f"#{r:02x}{g:02x}{b:02x}"

                    base_low = "#173058"
                    base_high = "#19B8B2"

                    def _row_heat_style(row):
                        styles = [""] * len(row)
                        vals = [_parse_num_for_heat(row.get(col)) for col in heat_cols]
                        nums = [v for v in vals if v is not None]
                        if not nums:
                            return styles

                        vmin, vmax = min(nums), max(nums)
                        span = (vmax - vmin) if vmax != vmin else 1.0

                        for i, col in enumerate(row.index):
                            if col not in heat_cols:
                                continue
                            value = _parse_num_for_heat(row[col])
                            if value is None:
                                continue
                            t = (value - vmin) / span
                            color = _blend_hex(base_low, base_high, t)
                            styles[i] = (
                                f"background-color: {color}; color: #EAF3EE; "
                                f"border: 1px solid {THEME_BORDER};"
                            )
                        return styles

                    st.caption(
                        "Отмечайте чекбоксы слева у нужных строк: для одной метрики строится график фактических значений, "
                        "для нескольких метрик — сравнительная динамика в относительном виде."
                    )

                    selection_state_key = "charts_summary_selected_metrics"
                    prev_selected_metrics = st.session_state.get(selection_state_key, [])
                    styled_summary = summary_matrix.style.apply(_row_heat_style, axis=1)
                    summary_selector_df = pd.DataFrame(
                        {
                            "График": summary_matrix.index.isin(prev_selected_metrics),
                        }
                    )
                    selector_col, table_col = st.columns([0.07, 0.93], vertical_alignment="top")
                    with selector_col:
                        summary_editor = st.data_editor(
                            summary_selector_df,
                            width="stretch",
                            hide_index=True,
                            column_config={
                                "График": st.column_config.CheckboxColumn(
                                    "График",
                                    help="Отметьте строку, чтобы сразу построить график ниже.",
                                    width="small",
                                ),
                            },
                            key="charts_summary_selector_editor",
                        )
                    with table_col:
                        st.dataframe(styled_summary, width="stretch")

                    selected_metrics_for_chart = summary_editor.loc[
                        summary_editor["График"].fillna(False),
                    ].index.tolist()
                    selected_metrics_for_chart = [
                        str(summary_matrix.index[idx])
                        for idx in selected_metrics_for_chart
                        if 0 <= idx < len(summary_matrix.index)
                    ]
                    st.session_state[selection_state_key] = selected_metrics_for_chart

                    if selected_metrics_for_chart:
                        chart_rows = []
                        if len(selected_metrics_for_chart) == 1:
                            metric_label = selected_metrics_for_chart[0]
                            metric_payload = metric_series_map.get(metric_label, {})
                            month_map = metric_payload.get("months", {})
                            for month_col in month_headers:
                                chart_rows.append(
                                    {
                                        "Месяц": month_col,
                                        "Метрика": metric_label,
                                        "Значение": float(month_map.get(month_col, 0.0)),
                                    }
                                )
                            chart_title = f"Динамика метрики: {metric_label}"
                            chart_y_title = metric_label
                            chart_note = None
                        else:
                            for metric_label in selected_metrics_for_chart:
                                metric_payload = metric_series_map.get(metric_label, {})
                                month_map = metric_payload.get("months", {})
                                base_value = None
                                for month_col in month_headers:
                                    value = float(month_map.get(month_col, 0.0))
                                    if base_value is None and value != 0:
                                        base_value = value
                                if base_value is None:
                                    base_value = 0.0
                                for month_col in month_headers:
                                    value = float(month_map.get(month_col, 0.0))
                                    relative_value = (value / base_value * 100.0) if base_value not in (None, 0.0) else 0.0
                                    chart_rows.append(
                                        {
                                            "Месяц": month_col,
                                            "Метрика": metric_label,
                                            "Значение": relative_value,
                                        }
                                    )
                            chart_title = "Сравнение динамики выбранных метрик"
                            chart_y_title = "Индекс, %"
                            chart_note = (
                                "Для нескольких строк график строится в относительном виде: "
                                "первая ненулевая точка каждой метрики принимается за 100%."
                            )

                        chart_df = pd.DataFrame(chart_rows)
                        fig_selected_metrics = px.line(
                            chart_df,
                            x="Месяц",
                            y="Значение",
                            color="Метрика",
                            markers=True,
                            title=chart_title,
                        )
                        fig_selected_metrics.update_layout(
                            height=420,
                            margin=dict(l=12, r=12, t=64, b=12),
                            paper_bgcolor="rgba(0,0,0,0)",
                            plot_bgcolor="rgba(0,0,0,0)",
                            font=dict(color=THEME_PLOT_TEXT),
                            xaxis_title="Месяц",
                            yaxis_title=chart_y_title,
                            legend_title_text="Метрика",
                        )
                        fig_selected_metrics.update_traces(line=dict(width=3))
                        st.plotly_chart(fig_selected_metrics, width="stretch")
                        if chart_note:
                            st.caption(chart_note)
                    else:
                        st.caption("Отметьте чекбокс слева у нужной строки, чтобы построить график по метрике.")

                    # ---------- Круговые диаграммы по типам РК ----------

                    ui_section_title("Распределение по типам РК")

                    agg_ct = df_sel.groupby("campaign_type", as_index=False).agg(
                        cost=("cost", "sum"),
                        conversions=("conversions", "sum"),
                        leads=("leads", "sum"),
                        target_leads=("target_leads", "sum"),
                        revenue=("revenue", "sum"),
                    )
                    show_pies = st.checkbox(
                        "Показывать круговые диаграммы",
                        value=True,
                        key="charts_show_pies",
                    )

                    if show_pies:
                        pie_color_map = {
                            ct_name: campaign_color_map.get(ct_name, chart_campaign_palette[0])
                            for ct_name in agg_ct["campaign_type"].astype(str).tolist()
                        }
                        st.caption("Общая легенда")
                        legend_cols_count = min(4, max(1, len(agg_ct)))
                        legend_cols = st.columns(legend_cols_count)
                        for i, ct_name in enumerate(agg_ct["campaign_type"].tolist()):
                            c = pie_color_map.get(str(ct_name), chart_campaign_palette[0])
                            with legend_cols[i % legend_cols_count]:
                                st.markdown(
                                    f"<span style='color:{c}; font-weight:700;'>●</span> "
                                    f"<span style='color:{THEME_LEGEND_TEXT};'>{ct_name}</span>",
                                    unsafe_allow_html=True,
                                )

                        col_pie1, col_pie2, col_pie3 = st.columns(3)
                        if is_real_estate_preset:
                            pie_specs = [
                                ("cost", "Доля бюджета по типам РК"),
                                ("leads" if metric_mode["is_real_estate_full"] else "target_leads", "Доля лидов по типам РК" if metric_mode["is_real_estate_full"] else "Доля ЦО по типам РК"),
                                ("target_leads" if metric_mode["is_real_estate_full"] else "clicks", "Доля ЦО по типам РК" if metric_mode["is_real_estate_full"] else "Доля кликов по типам РК"),
                            ]
                            if "clicks" not in agg_ct.columns:
                                agg_ct = agg_ct.merge(df_sel.groupby("campaign_type", as_index=False).agg(clicks=("clicks", "sum")), on="campaign_type", how="left")
                        else:
                            pie_specs = [
                                ("cost", "Доля бюджета по типам РК"),
                                ("conversions", "Доля конверсий по типам РК"),
                                ("revenue", "Доля дохода по типам РК"),
                            ]
                        with col_pie1:
                            fig_cost = px.pie(
                                agg_ct,
                                names="campaign_type",
                                values=pie_specs[0][0],
                                title=pie_specs[0][1],
                                color="campaign_type",
                                color_discrete_map=pie_color_map,
                            )
                            fig_cost.update_traces(textposition="inside", textinfo="percent+label")
                            fig_cost.update_layout(
                                height=520,
                                margin=dict(l=12, r=12, t=64, b=12),
                                showlegend=False,
                                paper_bgcolor="rgba(0,0,0,0)",
                                plot_bgcolor="rgba(0,0,0,0)",
                                font=dict(color=THEME_PLOT_TEXT),
                            )
                            st.plotly_chart(fig_cost, width="stretch")

                        with col_pie2:
                            fig_conv = px.pie(
                                agg_ct,
                                names="campaign_type",
                                values=pie_specs[1][0],
                                title=pie_specs[1][1],
                                color="campaign_type",
                                color_discrete_map=pie_color_map,
                            )
                            fig_conv.update_traces(textposition="inside", textinfo="percent+label")
                            fig_conv.update_layout(
                                height=520,
                                margin=dict(l=12, r=12, t=64, b=12),
                                showlegend=False,
                                paper_bgcolor="rgba(0,0,0,0)",
                                plot_bgcolor="rgba(0,0,0,0)",
                                font=dict(color=THEME_PLOT_TEXT),
                            )
                            st.plotly_chart(fig_conv, width="stretch")

                        with col_pie3:
                            fig_rev = px.pie(
                                agg_ct,
                                names="campaign_type",
                                values=pie_specs[2][0],
                                title=pie_specs[2][1],
                                color="campaign_type",
                                color_discrete_map=pie_color_map,
                            )
                            fig_rev.update_traces(textposition="inside", textinfo="percent+label")
                            fig_rev.update_layout(
                                height=520,
                                margin=dict(l=12, r=12, t=64, b=12),
                                showlegend=False,
                                paper_bgcolor="rgba(0,0,0,0)",
                                plot_bgcolor="rgba(0,0,0,0)",
                                font=dict(color=THEME_PLOT_TEXT),
                            )
                            st.plotly_chart(fig_rev, width="stretch")

                        if is_diy_preset and "segment" in df_sel.columns:
                            seg_work = df_sel.copy()
                            seg_work["segment_norm"] = (
                                seg_work["segment"].astype(str).str.upper().str.strip()
                            )
                            seg_work = seg_work[seg_work["segment_norm"].isin(["B2B", "B2C"])]

                            if not seg_work.empty:
                                if use_ak_budget_metrics:
                                    if use_vat_budget_metrics:
                                        seg_work["budget_basis"] = seg_work["cost_with_vat_ak"]
                                    else:
                                        seg_work["budget_basis"] = seg_work["cost"] + seg_work["ak_cost_wo_vat"]
                                else:
                                    seg_work["budget_basis"] = (
                                        seg_work["cost_with_vat"] if use_vat_budget_metrics else seg_work["cost"]
                                    )

                                agg_seg = (
                                    seg_work.groupby("segment_norm", as_index=False)
                                    .agg(
                                        budget_basis=("budget_basis", "sum"),
                                        revenue=("revenue", "sum"),
                                        orders=("conversions", "sum"),
                                    )
                                )
                                agg_seg["segment_norm"] = pd.Categorical(
                                    agg_seg["segment_norm"], categories=["B2C", "B2B"], ordered=True
                                )
                                agg_seg = agg_seg.sort_values("segment_norm")

                                ui_section_title("DIY: Распределение между B2B и B2C")
                                seg_palette = ["#3D8EF0", "#00CDC5"]
                                seg_pie1, seg_pie2, seg_pie3 = st.columns(3)

                                with seg_pie1:
                                    fig_seg_budget = px.pie(
                                        agg_seg,
                                        names="segment_norm",
                                        values="budget_basis",
                                        title="Доля бюджета между B2B и B2C",
                                        color_discrete_sequence=seg_palette,
                                    )
                                    fig_seg_budget.update_traces(textposition="inside", textinfo="percent+label")
                                    fig_seg_budget.update_layout(
                                        height=520,
                                        margin=dict(l=12, r=12, t=64, b=12),
                                        showlegend=False,
                                        paper_bgcolor="rgba(0,0,0,0)",
                                        plot_bgcolor="rgba(0,0,0,0)",
                                        font=dict(color=THEME_PLOT_TEXT),
                                    )
                                    st.plotly_chart(fig_seg_budget, width="stretch")

                                with seg_pie2:
                                    fig_seg_rev = px.pie(
                                        agg_seg,
                                        names="segment_norm",
                                        values="revenue",
                                        title="Доля выручки между B2B и B2C",
                                        color_discrete_sequence=seg_palette,
                                    )
                                    fig_seg_rev.update_traces(textposition="inside", textinfo="percent+label")
                                    fig_seg_rev.update_layout(
                                        height=520,
                                        margin=dict(l=12, r=12, t=64, b=12),
                                        showlegend=False,
                                        paper_bgcolor="rgba(0,0,0,0)",
                                        plot_bgcolor="rgba(0,0,0,0)",
                                        font=dict(color=THEME_PLOT_TEXT),
                                    )
                                    st.plotly_chart(fig_seg_rev, width="stretch")

                                with seg_pie3:
                                    fig_seg_orders = px.pie(
                                        agg_seg,
                                        names="segment_norm",
                                        values="orders",
                                        title="Доля заказов между B2B и B2C",
                                        color_discrete_sequence=seg_palette,
                                    )
                                    fig_seg_orders.update_traces(textposition="inside", textinfo="percent+label")
                                    fig_seg_orders.update_layout(
                                        height=520,
                                        margin=dict(l=12, r=12, t=64, b=12),
                                        showlegend=False,
                                        paper_bgcolor="rgba(0,0,0,0)",
                                        plot_bgcolor="rgba(0,0,0,0)",
                                        font=dict(color=THEME_PLOT_TEXT),
                                    )
                                    st.plotly_chart(fig_seg_orders, width="stretch")

    render_bottom_tab_switcher("Диаграммы", "charts")

# ====================================================================
#                           ТАБ «EXPORT»
# ====================================================================

with tab_export:
    ui_section_title("Экспорт")
    st.markdown(
        """
        <div class="tab-intro">
            <p>Вкладка подготовки и запуска экспорта медиаплана в Excel.</p>
            <p>Доступны: выгрузка детальных данных, сводки по месяцам и экспорт в шаблон.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown("### Экспорт проекта")
    st.caption("Сохраните текущее состояние проекта в JSON, чтобы передать его другому пользователю.")

    project_payload = export_project_state_payload()
    project_json = json.dumps(project_payload, ensure_ascii=False, indent=2).encode("utf-8")
    project_ts = dt.datetime.now().strftime("%Y%m%d_%H%M")
    st.download_button(
        "Скачать проект (JSON)",
        data=project_json,
        file_name=f"mediaplan_project_{project_ts}.json",
        mime="application/json",
        key="download_project_json",
    )

    st.markdown("---")
    if "df_all" not in locals() or df_all.empty:
        st.info("Нет данных для экспорта. Сначала рассчитайте медиаплан на вкладке «Медиаплан».")
    else:
        export_periods_df = (
            df_all[["planning_slot", "month_name"]]
            .dropna(subset=["planning_slot"])
            .drop_duplicates()
            .sort_values("planning_slot")
        )
        export_month_options = [
            f"{int(r['planning_slot'])}. {str(r['month_name'])}"
            for _, r in export_periods_df.iterrows()
        ]
        sync_multiselect_state("export_months_multiselect", export_month_options, default_to_all=True)
        export_month_labels = st.multiselect(
            "Месяцы для экспорта",
            options=export_month_options,
            default=export_month_options,
            key="export_months_multiselect",
        )

        export_ctypes_all = sorted(df_all["campaign_type"].dropna().astype(str).unique().tolist())
        export_ctypes_selected = st.multiselect(
            "Типы РК для экспорта",
            options=export_ctypes_all,
            default=export_ctypes_all,
            key="export_campaign_types_multiselect",
        )

        if not export_month_labels:
            st.info("Выберите хотя бы один месяц.")
        elif not export_ctypes_selected:
            st.info("Выберите хотя бы один тип РК.")
        else:
            export_month_slots = [int(lbl.split(".")[0]) for lbl in export_month_labels]
            export_mask = df_all["planning_slot"].isin(export_month_slots) & df_all["campaign_type"].isin(export_ctypes_selected)
            df_export = df_all.loc[export_mask].copy()

            if df_export.empty:
                st.info("Нет данных для выбранных фильтров экспорта.")
            else:
                st.caption(f"Строк к экспорту: {len(df_export)}")

                st.markdown("#### Экспорт данных")
                st.caption("ZIP-архив с таблицами fact/dim в CSV (UTF-8). Для импорта в Power BI используйте Get data -> Text/CSV.")

                bi_export_df = df_export.copy()
                bi_use_vat = bool(st.session_state.get("use_vat_budget_metrics", True))
                bi_use_ak = bool(st.session_state.get("use_ak_budget_metrics", False))

                if bi_use_ak:
                    if bi_use_vat:
                        bi_export_df["budget_basis"] = bi_export_df.get("cost_with_vat_ak", 0.0)
                    else:
                        bi_export_df["budget_basis"] = (
                            pd.to_numeric(bi_export_df.get("cost", 0.0), errors="coerce").fillna(0.0)
                            + pd.to_numeric(bi_export_df.get("ak_cost_wo_vat", 0.0), errors="coerce").fillna(0.0)
                        )
                else:
                    bi_export_df["budget_basis"] = (
                        bi_export_df.get("cost_with_vat", 0.0) if bi_use_vat else bi_export_df.get("cost", 0.0)
                    )

                if "segment" not in bi_export_df.columns:
                    bi_export_df["segment"] = "ALL"
                bi_export_df["segment"] = bi_export_df["segment"].fillna("ALL").astype(str)
                if "geo" not in bi_export_df.columns:
                    bi_export_df["geo"] = ""
                bi_export_df["geo"] = bi_export_df["geo"].fillna("").astype(str)
                bi_export_df["orders"] = pd.to_numeric(
                    bi_export_df.get("conversions", 0.0), errors="coerce"
                ).fillna(0.0)

                dim_month = (
                    bi_export_df[["planning_slot", "month_num", "month_year", "month_name"]]
                    .dropna(subset=["planning_slot"])
                    .drop_duplicates()
                    .sort_values("planning_slot")
                    .rename(columns={"planning_slot": "month_key"})
                    .reset_index(drop=True)
                )
                if not dim_month.empty:
                    dim_month["month_key"] = dim_month["month_key"].astype(int)

                dim_campaign = (
                    bi_export_df[["campaign_type", "system", "format", "geo", "segment"]]
                    .fillna("")
                    .drop_duplicates()
                    .reset_index(drop=True)
                )
                dim_campaign.insert(0, "campaign_key", np.arange(1, len(dim_campaign) + 1, dtype=int))

                fact = bi_export_df.merge(
                    dim_campaign,
                    on=["campaign_type", "system", "format", "geo", "segment"],
                    how="left",
                ).copy()
                fact["month_key"] = pd.to_numeric(fact.get("planning_slot"), errors="coerce").fillna(0).astype(int)

                fact_cols = [
                    "month_key", "campaign_key",
                    "planning_slot", "month_num", "month_year", "month_name", "campaign_type", "segment", "system", "format", "geo",
                    "impressions", "clicks", "orders", "conversions", "leads", "target_leads",
                    "cost", "ak_cost_wo_vat", "total_budget_wo_vat_ak", "cost_with_vat", "cost_with_vat_ak", "vat_amount", "budget_basis",
                    "ctr", "cpc", "cr", "cr1", "cr2", "cpa", "cpl", "cpql", "aov", "revenue", "roas", "drr",
                    "available_capacity", "client_count", "absolute_new_clients", "returned_clients", "new_clients", "cac", "order_frequency", "sov_pct", "new_clients_share_pct", "order_share_segment_pct", "revenue_share_segment_pct",
                    "ak_rate", "ak_rate_pct",
                ]
                for c in fact_cols:
                    if c not in fact.columns:
                        fact[c] = np.nan
                fact = fact[fact_cols].copy()

                dim_segment = (
                    bi_export_df[["segment"]]
                    .fillna("ALL")
                    .drop_duplicates()
                    .sort_values("segment")
                    .reset_index(drop=True)
                )
                dim_segment.insert(0, "segment_key", np.arange(1, len(dim_segment) + 1, dtype=int))

                readme_text = (
                    "media_planner BI export package\n"
                    "Import in Power BI Desktop: Get data -> Text/CSV.\n"
                    "Files:\n"
                    "- fact_mediaplan.csv (main fact table)\n"
                    "- dim_month.csv (join by month_key)\n"
                    "- dim_campaign.csv (join by campaign_key)\n"
                    "- dim_segment.csv (segment dictionary)\n"
                    "Notes:\n"
                    f"- budget_basis uses current VAT/AK mode (use_vat={bi_use_vat}, use_ak={bi_use_ak}).\n"
                    "- orders duplicates conversions for BI naming convenience.\n"
                )
                if is_real_estate_preset:
                    readme_text += (
                        "- real_estate preset: target_leads = ??, cpql uses current budget_basis.\n"
                        "- full funnel mode also includes leads and cpl.\n"
                    )
                bi_zip_buf = io.BytesIO()
                with zipfile.ZipFile(bi_zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                    zf.writestr("fact_mediaplan.csv", _to_csv_bytes(fact))
                    zf.writestr("dim_month.csv", _to_csv_bytes(dim_month))
                    zf.writestr("dim_campaign.csv", _to_csv_bytes(dim_campaign))
                    zf.writestr("dim_segment.csv", _to_csv_bytes(dim_segment))
                    zf.writestr("README.txt", readme_text.encode("utf-8"))
                bi_zip_buf.seek(0)

                bi_ts = dt.datetime.now().strftime("%Y%m%d_%H%M")
                st.download_button(
                    "Скачать BI-пакет (Power BI CSV ZIP)",
                    data=bi_zip_buf.getvalue(),
                    file_name=f"mediaplan_bi_export_{bi_ts}.zip",
                    mime="application/zip",
                    key="download_export_bi_zip",
                )

                # ---------- 1) Детальный Excel ----------
                timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M")
                detail_buf = io.BytesIO()
                detail_engine = "xlsxwriter" if HAS_XLSXWRITER else "openpyxl"
                with pd.ExcelWriter(detail_buf, engine=detail_engine) as writer:
                    detail_cols = [
                        "planning_slot", "month_num", "month_year", "month_name", "campaign_type", "system", "format", "geo",
                        "impressions", "clicks", "ctr", "cpc", "cost", "ak_cost_wo_vat", "total_budget_wo_vat_ak", "cost_with_vat",
                        "cost_with_vat_ak", "vat_amount", "ak_rate", "ak_rate_pct",
                        "cr", "cr1", "cr2", "conversions", "leads", "target_leads", "aov", "cpa", "cpl", "cpql", "revenue", "roas", "drr",
                        "available_capacity", "client_count", "absolute_new_clients", "returned_clients", "new_clients", "cac", "order_frequency", "sov_pct", "new_clients_share_pct", "order_share_segment_pct", "revenue_share_segment_pct", "budget_share_segment_pct",
                    ]
                    for c in detail_cols:
                        if c not in df_export.columns:
                            df_export[c] = np.nan
                    df_detail = safe_select_columns(df_export, detail_cols).rename(
                        columns={
                            "planning_slot": "Месяц плана №",
                            "month_num": "Календарный месяц №",
                            "month_year": "Год",
                            "month_name": "Месяц",
                            "campaign_type": "Тип РК",
                            "system": "Система",
                            "format": "Формат",
                            "geo": "ГЕО",
                            "impressions": "Показы",
                            "clicks": "Клики",
                            "ctr": "CTR (доля)",
                            "cpc": "CPC, ₽",
                            "cost": "Бюджет без НДС, ₽",
                            "cost_with_vat": "Бюджет с НДС, ₽",
                            "ak_rate": "АК (доля)",
                            "ak_rate_pct": "АК, %",
                            "ak_cost_wo_vat": "АК без НДС, ₽",
                            "total_budget_wo_vat_ak": "Тотал бюджет без НДС с учетом АК, ₽",
                            "cost_with_vat_ak": "Бюджет с НДС и АК, ₽",
                            "vat_amount": "НДС, ₽",
                            "cr": "CR (доля)",
                            "cr1": "CR1 в Лид (доля)",
                            "cr2": "CR2 в ЦО (доля)",
                            "conversions": "Конверсии",
                            "leads": "Лиды",
                            "target_leads": "ЦО",
                            "aov": "AOV, ₽",
                            "cpa": "CPO, ₽",
                            "cpl": "CPL, ₽",
                            "cpql": "CPQL, ₽",
                            "revenue": "Выручка, ₽",
                            "roas": "ROAS",
                            "drr": "ДРР (доля)",
                            "available_capacity": "Доступная емкость",
                            "client_count": "Количество клиентов",
                            "absolute_new_clients": "Кол-во абсолютно новых клиентов",
                            "returned_clients": "Кол-во вернувшихся клиентов",
                            "new_clients": "Кол-во новых клиентов",
                            "cac": "CAC, ₽",
                            "order_frequency": "Частота заказа",
                            "sov_pct": "SOV, %",
                            "new_clients_share_pct": "Доля новых клиентов, %",
                            "order_share_segment_pct": "Доля заказов, %",
                            "revenue_share_segment_pct": "Доля дохода, %",
                            "budget_share_segment_pct": "Доля рекламного бюджета, %",
                        }
                    )
                    df_detail.to_excel(writer, sheet_name="Детально", index=False)

                    agg_month = df_export.groupby(["planning_slot", "month_num", "month_year", "month_name"], as_index=False).agg(
                        impressions=("impressions", "sum"),
                        clicks=("clicks", "sum"),
                        conversions=("conversions", "sum"),
                        leads=("leads", "sum"),
                        target_leads=("target_leads", "sum"),
                        cost=("cost", "sum"),
                        cost_with_vat=("cost_with_vat", "sum"),
                        ak_cost_wo_vat=("ak_cost_wo_vat", "sum"),
                        cost_with_vat_ak=("cost_with_vat_ak", "sum"),
                        revenue=("revenue", "sum"),
                    ).sort_values("planning_slot")
                    agg_month["ctr_pct"] = np.where(agg_month["impressions"] > 0, agg_month["clicks"] / agg_month["impressions"] * 100.0, 0.0)
                    agg_month["cpc"] = np.where(agg_month["clicks"] > 0, agg_month["cost"] / agg_month["clicks"], 0.0)
                    agg_month["cr_pct"] = np.where(agg_month["clicks"] > 0, agg_month["conversions"] / agg_month["clicks"] * 100.0, 0.0)
                    agg_month["cr1_pct"] = np.where(agg_month["clicks"] > 0, agg_month["leads"] / agg_month["clicks"] * 100.0, 0.0)
                    agg_month["cr2_pct"] = np.where(agg_month["leads"] > 0, agg_month["target_leads"] / agg_month["leads"] * 100.0, 0.0)
                    export_use_vat = bool(st.session_state.get("use_vat_budget_metrics", True))
                    export_use_ak = bool(st.session_state.get("use_ak_budget_metrics", False))
                    if export_use_ak:
                        budget_series = agg_month["cost_with_vat_ak"] if export_use_vat else (agg_month["cost"] + agg_month["ak_cost_wo_vat"])
                    else:
                        budget_series = agg_month["cost_with_vat"] if export_use_vat else agg_month["cost"]
                    agg_month["cpo"] = np.where(agg_month["conversions"] > 0, budget_series / agg_month["conversions"], 0.0)
                    agg_month["cpl"] = np.where(agg_month["leads"] > 0, budget_series / agg_month["leads"], 0.0)
                    agg_month["cpql"] = np.where(agg_month["target_leads"] > 0, budget_series / agg_month["target_leads"], 0.0)
                    agg_month["roas"] = np.where(budget_series > 0, agg_month["revenue"] / budget_series, 0.0)
                    agg_month["drr_pct"] = np.where(agg_month["revenue"] > 0, budget_series / agg_month["revenue"] * 100.0, 0.0)
                    agg_month = agg_month.rename(
                        columns={
                            "planning_slot": "Месяц плана №",
                            "month_num": "Календарный месяц №",
                            "month_year": "Год",
                            "month_name": "Месяц",
                            "impressions": "Показы",
                            "clicks": "Клики",
                            "conversions": "Конверсии",
                            "leads": "Лиды",
                            "target_leads": "ЦО",
                            "cost": "Бюджет без НДС, ₽",
                            "cost_with_vat": "Бюджет с НДС, ₽",
                            "ak_cost_wo_vat": "АК без НДС, ₽",
                            "cost_with_vat_ak": "Бюджет с НДС и АК, ₽",
                            "revenue": "Выручка, ₽",
                            "ctr_pct": "CTR, %",
                            "cpc": "CPC, ₽",
                            "cr_pct": "CR, %",
                            "cr1_pct": "CR1 в Лид, %",
                            "cr2_pct": "CR2 в ЦО, %",
                            "cpo": "CPO, ₽",
                            "cpl": "CPL, ₽",
                            "cpql": "CPQL, ₽",
                            "roas": "ROAS",
                            "drr_pct": "ДРР, %",
                        }
                    )
                    agg_month.to_excel(writer, sheet_name="Сводка_по_месяцам", index=False)

                detail_buf.seek(0)
                st.download_button(
                    "Скачать Excel (детально + сводка)",
                    data=detail_buf.getvalue(),
                    file_name=f"mediaplan_export_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_export_detail_xlsx",
                    type="primary",
                )

                # ---------- 2) Экспорт в шаблон ----------
                collapse_geo_template_export = st.checkbox(
                    "Схлопнуть ГЕО в РФ для шаблонного экспорта",
                    value=False,
                    key="collapse_geo_template_export",
                    help="Объединяет строки одинаковых типов РК из разных ГЕО и выгружает их одной строкой с ГЕО = РФ.",
                )

                template_campaigns_raw = campaigns[campaigns["campaign_type"].isin(export_ctypes_selected)].copy()
                template_periods = (
                    df_export[["planning_slot", "month_num", "month_year", "month_name"]]
                    .drop_duplicates()
                    .sort_values("planning_slot")
                    .to_dict("records")
                )
                if len(template_periods) > 12:
                    st.warning("Для экспорта в шаблон используется максимум 12 месяцев.")
                    template_periods = template_periods[:12]

                template_df_export, template_campaigns = build_template_export_payload(
                    df_export=df_export,
                    campaigns_export=template_campaigns_raw,
                    collapse_geo_to_rf=collapse_geo_template_export,
                )

                if collapse_geo_template_export:
                    st.caption(
                        f"Шаблонный экспорт будет собран без разбивки по ГЕО: {len(df_export)} строк -> {len(template_df_export)} строк."
                    )

                if not template_campaigns.empty and template_periods:
                    try:
                        export_preset_key = str(st.session_state.get("metric_preset_key", "ecom")).strip().lower()
                        if export_preset_key not in METRIC_PRESETS:
                            export_preset_key = "ecom"
                        allow_ecom_template = export_preset_key == "ecom"
                        allow_diy_template = export_preset_key == "diy"
                        real_estate_funnel_mode_export = str(st.session_state.get("real_estate_funnel_mode", "simple"))
                        allow_real_estate_simple_template = export_preset_key == "real_estate" and real_estate_funnel_mode_export == "simple"
                        allow_real_estate_full_template = export_preset_key == "real_estate" and real_estate_funnel_mode_export == "full"
                        resolved_tpl_ecom = resolve_template_path("ecom")
                        if resolved_tpl_ecom:
                            tpl_buf_ecom = build_excel_from_template(
                                df_all=template_df_export,
                                campaigns=template_campaigns,
                                selected_periods=template_periods,
                                template_kind="ecom",
                            )
                            st.download_button(
                                "Скачать Excel по шаблону E-com",
                                data=tpl_buf_ecom.getvalue(),
                                file_name=f"mediaplan_template_ecom_{timestamp}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_export_template_xlsx_ecom",
                                disabled=not allow_ecom_template,
                            )
                        else:
                            st.info("Шаблон E-com не найден. Ожидаемые пути: " + ", ".join(TEMPLATE_PATHS_ECOM))

                        resolved_tpl_diy = resolve_template_path("diy")
                        if resolved_tpl_diy:
                            tpl_buf_diy = build_excel_from_template(
                                df_all=template_df_export,
                                campaigns=template_campaigns,
                                selected_periods=template_periods,
                                template_kind="diy",
                            )
                            st.download_button(
                                "Скачать Excel по шаблону DIY",
                                data=tpl_buf_diy.getvalue(),
                                file_name=f"mediaplan_template_diy_{timestamp}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_export_template_xlsx_diy",
                                disabled=not allow_diy_template,
                            )
                        else:
                            st.info("Шаблон DIY не найден. Добавьте файл в один из путей: " + ", ".join(TEMPLATE_PATHS_DIY))

                        resolved_tpl_real_estate_simple = resolve_template_path("real_estate_simple")
                        if resolved_tpl_real_estate_simple:
                            tpl_buf_real_estate_simple = build_excel_from_template(
                                df_all=template_df_export,
                                campaigns=template_campaigns,
                                selected_periods=template_periods,
                                template_kind="real_estate_simple",
                            )
                            st.download_button(
                                "Скачать Excel по шаблону Недвижимость (упрощенная воронка)",
                                data=tpl_buf_real_estate_simple.getvalue(),
                                file_name=f"mediaplan_template_real_estate_simple_{timestamp}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_export_template_xlsx_real_estate_simple",
                                disabled=not allow_real_estate_simple_template,
                            )
                        elif export_preset_key == "real_estate" and real_estate_funnel_mode_export == "simple":
                            st.info("Шаблон Недвижимость (упрощенная воронка) не найден. Добавьте файл в один из путей: " + ", ".join(TEMPLATE_PATHS_REAL_ESTATE_SIMPLE))

                        resolved_tpl_real_estate_full = resolve_template_path("real_estate_full")
                        if resolved_tpl_real_estate_full:
                            tpl_buf_real_estate_full = build_excel_from_template(
                                df_all=template_df_export,
                                campaigns=template_campaigns,
                                selected_periods=template_periods,
                                template_kind="real_estate_full",
                            )
                            st.download_button(
                                "Скачать Excel по шаблону Недвижимость (полная воронка)",
                                data=tpl_buf_real_estate_full.getvalue(),
                                file_name=f"mediaplan_template_real_estate_full_{timestamp}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_export_template_xlsx_real_estate_full",
                                disabled=not allow_real_estate_full_template,
                            )
                        elif export_preset_key == "real_estate" and real_estate_funnel_mode_export == "full":
                            st.info("Шаблон Недвижимость (полная воронка) не найден. Добавьте файл в один из путей: " + ", ".join(TEMPLATE_PATHS_REAL_ESTATE_FULL))

                    except FileNotFoundError:
                        st.error("Файл шаблона для выбранного типа не найден.")
                    except Exception as e:
                        st.error(f"Ошибка экспорта в шаблон: {e}")
    st.markdown("---")
    st.markdown("### Импорт проекта")
    st.caption("Загрузите JSON-файл проекта, чтобы восстановить расчеты и настройки в этом приложении.")
    import_result = st.session_state.pop("_pending_project_import_result", None)
    if isinstance(import_result, dict):
        if import_result.get("ok"):
            st.success(str(import_result.get("msg", "Данные проекта импортированы.")))
        else:
            st.error(str(import_result.get("msg", "Не удалось импортировать проект.")))

    uploaded_project = st.file_uploader(
        "Импорт проекта (JSON)",
        type=["json"],
        key="upload_project_json",
        help="После импорта страница перезагрузится и восстановит наборы коэффициентов, связки и настройки.",
    )
    queue_project_import_from_upload(uploaded_project, "export_tab")

    render_bottom_tab_switcher("Export/Import", "export")


